"""Genera el Listado de Compras en PDF y Excel — puro, sin tocar la base ni la red.

Mismo criterio que core/exportar_precios.py (banda de encabezado repetida en
cada página, una sección por página con su título embebido en la tabla vía
repeatRows para que se repita solo si se corta), pero con su propia
estructura de columnas: acá no hay "grupo" ni resaltado de "nuevo precio",
en cambio se agrupa primero por fecha y, dentro de cada fecha, por
proveedor.

Recibe los datos ya armados (ver "filas" en cada función) y devuelve los
bytes del archivo — no sabe nada de la base, eso lo resuelve quien llama
(app/main.py). Los archivos son temporarios: esto solo arma bytes en
memoria, nunca los guarda en ningún lado.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"

SUFIJOS_UNIDAD_COMPRA = {"kilo": "k", "unidad": "u", "cubeta": "c"}


def _formatear_numero(valor) -> str:
    """Muestra un número sin decimales de sobra (16 en vez de 16.0), igual que el resto de la app."""
    if valor is None:
        return "—"
    return f"{float(valor):.2f}".rstrip("0").rstrip(".")


def _formatear_moneda(valor) -> str:
    """"$45.000": símbolo $, "." cada tres cifras, redondeado al peso entero. Sin importe -> "Sin precio"."""
    if valor is None:
        return "Sin precio"
    entero = round(float(valor))
    negativo = entero < 0
    texto = f"{abs(entero):,}".replace(",", ".")
    return f"${'-' if negativo else ''}{texto}"


def _texto_cantidad(fila: dict) -> str:
    """"40 cajones × 20k": mismo criterio de abreviatura (k/u/c) que ya usa Últimas Compras."""
    cantidad = _formatear_numero(fila.get("cantidad_cajones"))
    contenido = _formatear_numero(fila.get("contenido_por_cajon"))
    sufijo = SUFIJOS_UNIDAD_COMPRA.get(fila.get("unidad_compra"), "")
    return f"{cantidad} cajones × {contenido}{sufijo}"


def _agrupar_por_fecha_y_proveedor(filas: list[dict]) -> list[tuple]:
    """Agrupa las filas por fecha (más reciente primero) y, dentro de cada fecha, por proveedor
    (por código de puesto). Devuelve [(fecha, [(proveedor_nombre, proveedor_codigo, filas), ...]), ...].
    """
    por_fecha: dict = {}
    for fila in filas:
        por_fecha.setdefault(fila["fecha_operacion"], []).append(fila)

    resultado = []
    for fecha in sorted(por_fecha.keys(), reverse=True):
        por_proveedor: dict = {}
        for fila in por_fecha[fecha]:
            clave = (fila["proveedor_codigo_puesto"], fila["proveedor_nombre"])
            por_proveedor.setdefault(clave, []).append(fila)

        proveedores = [
            (nombre, codigo, por_proveedor[(codigo, nombre)])
            for codigo, nombre in sorted(por_proveedor.keys())
        ]
        resultado.append((fecha, proveedores))
    return resultado


OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm  # topMargin: deja aire entre el filete y el cuerpo


def _dibujar_encabezado(canvas, documento, subtitulo: str):
    """Título "Listado de Compras" negro a la izquierda, subtítulo con el rango de fechas, y un filete
    verde fino separador — directo en el canvas, en TODAS las páginas (mismo criterio que
    core/exportar_precios.py: un flowable normal solo aparece una vez, esto tiene que repetirse).
    """
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Listado de Compras")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 10)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def generar_pdf_listado_compras(fecha_desde: date, fecha_hasta: date, filas: list[dict]) -> bytes:
    """Arma el PDF del Listado de Compras entre dos fechas.

    filas: las mismas que devuelve app.db.buscar_compras (fecha_operacion, proveedor_nombre,
    proveedor_codigo_puesto, articulo_nombre, cantidad_cajones, contenido_por_cajon, unidad_compra,
    importe, ...).

    Cada FECHA arranca en su propia página; dentro de una fecha, cada proveedor es su propia tabla
    (título "fecha — proveedor (código)" + encabezado de columnas embebidos vía repeatRows=2, para que
    se repitan solos si ese proveedor tiene tantas compras que la tabla se corta sola entre páginas).
    """
    buffer = BytesIO()
    subtitulo = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _dibujar_encabezado_pagina(canvas, documento):
        _dibujar_encabezado(canvas, documento, subtitulo)

    estilo_titulo_tabla = ParagraphStyle(
        "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=9.5, textColor=VERDE_ENCABEZADO
    )
    estilo_articulo = ParagraphStyle("articulo", fontName="Helvetica", fontSize=9.5, textColor=colors.black)
    estilo_cantidad = ParagraphStyle("cantidad", fontName="Helvetica", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)
    estilo_importe = ParagraphStyle("importe", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.black)

    elementos = []
    grupos_por_fecha = _agrupar_por_fecha_y_proveedor(filas)

    for indice_fecha, (fecha_operacion, proveedores) in enumerate(grupos_por_fecha):
        if indice_fecha > 0:
            elementos.append(PageBreak())

        for indice_proveedor, (proveedor_nombre, proveedor_codigo, filas_proveedor) in enumerate(proveedores):
            if indice_proveedor > 0:
                elementos.append(Spacer(1, 10))

            titulo = f"{fecha_operacion.strftime('%d/%m/%Y')} — {proveedor_nombre} ({proveedor_codigo})"
            datos_tabla = [
                [Paragraph(titulo, estilo_titulo_tabla), "", ""],
                [
                    Paragraph("Artículo", estilo_encabezado_tabla),
                    Paragraph("Cantidad", estilo_encabezado_tabla),
                    Paragraph("Importe", estilo_encabezado_tabla),
                ],
            ]
            estilos_filas = []
            for indice_dato, fila in enumerate(filas_proveedor):
                indice_tabla = indice_dato + 2
                datos_tabla.append(
                    [
                        Paragraph(fila["articulo_nombre"], estilo_articulo),
                        Paragraph(_texto_cantidad(fila), estilo_cantidad),
                        Paragraph(_formatear_moneda(fila.get("importe")), estilo_importe),
                    ]
                )
                if indice_dato % 2 == 1:
                    estilos_filas.append(("BACKGROUND", (0, indice_tabla), (-1, indice_tabla), GRIS_FILA_ALTERNADA))

            tabla = Table(
                datos_tabla,
                colWidths=[ancho_util * 0.5, ancho_util * 0.28, ancho_util * 0.22],
                repeatRows=2,
            )
            tabla.setStyle(
                TableStyle(
                    [
                        ("SPAN", (0, 0), (-1, 0)),
                        ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                        ("LINEBELOW", (0, 2), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                        ("TOPPADDING", (0, 0), (-1, 0), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 1), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        *estilos_filas,
                    ]
                )
            )
            elementos.append(tabla)

    documento.build(elementos, onFirstPage=_dibujar_encabezado_pagina, onLaterPages=_dibujar_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_listado_compras(fecha_desde: date, fecha_hasta: date, filas: list[dict]) -> bytes:
    """Arma el Excel del Listado de Compras entre dos fechas, mismas secciones y columnas que el PDF."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Listado de Compras"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_fecha = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_proveedor = Font(bold=True, size=10.5)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Listado de Compras")
    for columna in range(1, 4):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    subtitulo = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    hoja.cell(row=fila_actual, column=1, value=subtitulo).font = fuente_normal
    fila_actual += 2

    for fecha_operacion, proveedores in _agrupar_por_fecha_y_proveedor(filas):
        hoja.cell(row=fila_actual, column=1, value=fecha_operacion.strftime("%d/%m/%Y")).font = fuente_fecha
        fila_actual += 1

        for proveedor_nombre, proveedor_codigo, filas_proveedor in proveedores:
            hoja.cell(
                row=fila_actual, column=1, value=f"{proveedor_nombre} ({proveedor_codigo})"
            ).font = fuente_proveedor
            fila_actual += 1

            for columna, encabezado in enumerate(("Artículo", "Cantidad", "Importe"), start=1):
                celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
                celda.font = fuente_encabezado_tabla
                celda.fill = relleno_verde_claro
            fila_actual += 1

            for fila in filas_proveedor:
                hoja.cell(row=fila_actual, column=1, value=fila["articulo_nombre"])
                hoja.cell(row=fila_actual, column=2, value=_texto_cantidad(fila))
                importe = fila.get("importe")
                celda_importe = hoja.cell(row=fila_actual, column=3, value=float(importe) if importe is not None else "Sin precio")
                if importe is not None:
                    celda_importe.number_format = '"$"#,##0'
                fila_actual += 1

            fila_actual += 1  # renglón en blanco entre proveedores

        fila_actual += 1  # renglón en blanco entre fechas

    hoja.column_dimensions[get_column_letter(1)].width = 34
    hoja.column_dimensions[get_column_letter(2)].width = 22
    hoja.column_dimensions[get_column_letter(3)].width = 14

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
