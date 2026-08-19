"""Genera el Excel de Disponibles — puro, sin tocar la base ni la red.

Recibe los datos ya armados (ver "filas" en generar_excel_disponibles) y
devuelve los bytes del archivo, calcado del formato real que se manda hoy
por mail (hoja "Stock Actualizado"). No sabe nada de clientes ni de la
base — eso lo resuelve app/main.py.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

FILA_ENCABEZADO = 6
FILA_PRIMER_DATO = FILA_ENCABEZADO + 1

AZUL_ENCABEZADO_HEX = "1F4E79"  # banda azul oscuro del encabezado, letra blanca encima

_BORDE_FINO = Side(style="thin", color="000000")
_BORDE_CELDA = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_FINO, bottom=_BORDE_FINO)


def _texto_fecha(fecha_desde: date, fecha_hasta: date) -> str:
    """"Fecha: 14/08/2026" si es un solo día, o "Fecha: 14/08/2026 al 15/08/2026" si el rango es de dos."""
    texto_desde = fecha_desde.strftime("%d/%m/%Y")
    if fecha_hasta == fecha_desde:
        return f"Fecha: {texto_desde}"
    return f"Fecha: {texto_desde} al {fecha_hasta.strftime('%d/%m/%Y')}"


def generar_excel_disponibles(fecha_desde: date, fecha_hasta: date, filas: list[dict], nombre_empresa: str) -> bytes:
    """Arma el Excel de Disponibles, mismo layout que el archivo real que se manda hoy a mano:

    A1 (combinada A1:C1): título "{EMPRESA} - Disponibilidad de Stock" (el nombre viene como
    parámetro y se escribe en mayúsculas — mismo código para varias empresas, este módulo no lee
    variables de entorno, eso lo resuelve app/main.py). A2 (combinada A2:C2): "Fecha: ...".
    Filas 3-5 vacías. Fila 6: encabezado Código/Producto/Stock. Datos desde la fila 7, en el
    orden en que vienen "filas" — ya no se reordena acá.

    Estética: encabezado con banda azul oscuro y letra blanca, y bordes finos
    en toda la tabla (encabezado y datos) — el layout de filas/columnas no
    cambia, solo el aspecto.

    filas: [{"codigo": str | None, "nombre": str, "cantidad": float}, ...]. codigo vacío deja la
    celda vacía (ej. Frutilla). cantidad se escribe siempre, incluso si es 0.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Stock Actualizado"

    hoja.merge_cells("A1:C1")
    hoja["A1"] = f"{nombre_empresa.upper()} - Disponibilidad de Stock"
    hoja["A1"].font = Font(bold=True, size=14)

    hoja.merge_cells("A2:C2")
    hoja["A2"] = _texto_fecha(fecha_desde, fecha_hasta)
    hoja["A2"].font = Font(bold=True)

    relleno_encabezado = PatternFill(start_color=AZUL_ENCABEZADO_HEX, end_color=AZUL_ENCABEZADO_HEX, fill_type="solid")
    for columna, encabezado in enumerate(("Código", "Producto", "Stock"), start=1):
        celda = hoja.cell(row=FILA_ENCABEZADO, column=columna, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno_encabezado
        celda.border = _BORDE_CELDA

    fila_actual = FILA_PRIMER_DATO
    for fila in filas:
        hoja.cell(row=fila_actual, column=1, value=fila.get("codigo") or None)
        hoja.cell(row=fila_actual, column=2, value=fila["nombre"])
        hoja.cell(row=fila_actual, column=3, value=float(fila["cantidad"]))
        for columna in (1, 2, 3):
            hoja.cell(row=fila_actual, column=columna).border = _BORDE_CELDA
        fila_actual += 1

    hoja.column_dimensions["A"].width = 12
    hoja.column_dimensions["B"].width = 34
    hoja.column_dimensions["C"].width = 12

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
