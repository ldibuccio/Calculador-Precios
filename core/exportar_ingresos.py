"""Genera Ingresos a Depósito en PDF y Excel — puro, sin tocar la base.

Mismo criterio que core/exportar_retiros.py: banda de encabezado repetida
en cada página, una tabla por proveedor con su subtotal (así se factura),
y el total general al final. Las cantidades son SIEMPRE las reales que
pesó/contó Depósito; el rechazo parcial se muestra aparte para explicar
por qué el número no coincide con lo comprado, y las compras sin precio
van marcadas (falta completarlas antes de facturar).

grupos/totales: los mismos que arma _grupos_ingresos_deposito en
app/main.py — esto solo arma bytes en memoria, nunca guarda nada.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
ROJO_MARCA = colors.Color(0.6, 0.11, 0.11)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"
ROJO_MARCA_HEX = "991B1B"

SUFIJOS_UNIDAD_COMPRA = {"kilo": "k", "unidad": "u", "cubeta": "c"}

OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm


def _formatear_numero(valor) -> str:
    if valor is None:
        return "—"
    return f"{float(valor):.2f}".rstrip("0").rstrip(".")


def _formatear_moneda(valor) -> str:
    if valor is None:
        return "—"
    entero = round(float(valor))
    negativo = entero < 0
    texto = f"{abs(entero):,}".replace(",", ".")
    return f"${'-' if negativo else ''}{texto}"


def _texto_recepcion(fila: dict) -> str:
    return fila["procesada_el"].strftime("%d/%m %H:%M") if fila["procesada_el"] else "—"


def _texto_guia(fila: dict) -> str:
    if fila["guia_id"] is None:
        return "—"
    return f"{fila['guia_id']}.{fila['guia_punto']}"


def _texto_cantidad_real(fila: dict) -> str:
    """"8 × 20k" con las columnas REALES; una rechazada total/no ingresada no tiene nada contado ("—")."""
    if fila["cantidad_cajones_real"] is None:
        return "—"
    sufijo = SUFIJOS_UNIDAD_COMPRA.get(fila.get("unidad_compra"), "")
    return f"{_formatear_numero(fila['cantidad_cajones_real'])} × {_formatear_numero(fila['contenido_por_cajon_real'])}{sufijo}"


def _texto_sena(fila: dict) -> str:
    """La seña por cajón y, entre paréntesis, lo señado por los cajones reales ("$500 ($4.000)")."""
    if fila.get("sena") is None:
        return "—"
    texto = _formatear_moneda(fila["sena"])
    if fila.get("total_sena") is not None:
        texto += f" ({_formatear_moneda(fila['total_sena'])})"
    return texto


def _texto_estado(fila: dict) -> str:
    estado = fila["estado_etiqueta"]
    if fila["cantidad_cajones_rechazada"] is not None:
        estado += f" ({_formatear_numero(fila['cantidad_cajones_rechazada'])} rech.)"
    return estado


def _armar_subtitulo(fecha_desde: date, fecha_hasta: date, filtros_texto: list[str]) -> str:
    subtitulo = f"Recepciones del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    if filtros_texto:
        subtitulo += " — " + ", ".join(filtros_texto)
    return subtitulo


def _dibujar_encabezado(canvas, documento, subtitulo: str):
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Ingresos a Depósito")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 10)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def generar_pdf_ingresos_deposito(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], grupos: list[dict], totales: dict
) -> bytes:
    """Arma el PDF de Ingresos a Depósito: una tabla por proveedor con subtotal + total general al final."""
    buffer = BytesIO()
    subtitulo = _armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto)
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado(canvas, doc, subtitulo)

    estilo_titulo_tabla = ParagraphStyle(
        "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=8.5, textColor=VERDE_ENCABEZADO
    )
    estilo_dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=8.5, textColor=colors.black)
    estilo_dato_gris = ParagraphStyle("dato_gris", fontName="Helvetica", fontSize=8, textColor=GRIS_TEXTO_AYUDA)
    estilo_numero = ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.black)
    estilo_marca = ParagraphStyle("marca", fontName="Helvetica-Bold", fontSize=8, textColor=ROJO_MARCA)
    estilo_subtotal = ParagraphStyle("subtotal", fontName="Helvetica-Bold", fontSize=9, textColor=colors.black)
    estilo_total = ParagraphStyle("total", fontName="Helvetica-Bold", fontSize=13, textColor=colors.black)
    estilo_aviso = ParagraphStyle("aviso", fontName="Helvetica-Bold", fontSize=9.5, textColor=ROJO_MARCA)
    estilo_vacio = ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []
    if not grupos:
        elementos.append(Paragraph("No se encontraron ingresos con estos filtros.", estilo_vacio))

    encabezados = ("Recepción", "Guía", "Artículo", "Cantidad real", "Precio", "Seña", "Total", "Estado")
    anchos = [
        ancho_util * 0.11, ancho_util * 0.07, ancho_util * 0.18, ancho_util * 0.14,
        ancho_util * 0.11, ancho_util * 0.13, ancho_util * 0.12, ancho_util * 0.14,
    ]
    for indice_grupo, grupo in enumerate(grupos):
        if indice_grupo > 0:
            elementos.append(Spacer(1, 14))

        titulo = f"{grupo['proveedor_nombre']} ({grupo['proveedor_codigo_puesto']})"
        datos_tabla = [
            [Paragraph(titulo, estilo_titulo_tabla)] + [""] * (len(encabezados) - 1),
            [Paragraph(encabezado, estilo_encabezado_tabla) for encabezado in encabezados],
        ]
        estilos_filas = []
        for indice, fila in enumerate(grupo["filas"]):
            precio = "<font color='#991B1B'><b>SIN PRECIO</b></font>" if fila["sin_precio"] else _formatear_moneda(fila["importe"])
            datos_tabla.append(
                [
                    Paragraph(_texto_recepcion(fila), estilo_dato_gris),
                    Paragraph(_texto_guia(fila), estilo_dato_gris),
                    Paragraph(fila["articulo_nombre"], estilo_dato),
                    Paragraph(_texto_cantidad_real(fila), estilo_dato),
                    Paragraph(precio, estilo_dato),
                    Paragraph(_texto_sena(fila), estilo_dato),
                    Paragraph(_formatear_moneda(fila["total"]), estilo_numero),
                    Paragraph(_texto_estado(fila), estilo_marca if fila["estado_etiqueta"] != "Recepcionada" else estilo_dato_gris),
                ]
            )
            if indice % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice + 2), (-1, indice + 2), GRIS_FILA_ALTERNADA))

        indice_subtotal = len(datos_tabla)
        datos_tabla.append(
            [
                Paragraph("Subtotal", estilo_subtotal), "", "", "", "", "",
                Paragraph(_formatear_moneda(grupo["subtotal"]), estilo_subtotal), "",
            ]
        )

        tabla = Table(datos_tabla, colWidths=anchos, repeatRows=2)
        tabla.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 2), (-1, -2), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("LINEABOVE", (0, indice_subtotal), (-1, indice_subtotal), 1, VERDE_ENCABEZADO),
                    ("TOPPADDING", (0, 0), (-1, 0), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

    if grupos:
        elementos.append(Spacer(1, 16))
        elementos.append(Paragraph(f"Total general: {_formatear_moneda(totales['total_general'])}", estilo_total))
        if totales["cantidad_sin_precio"] > 0:
            elementos.append(Spacer(1, 5))
            elementos.append(
                Paragraph(
                    f"Ojo: {totales['cantidad_sin_precio']} compra{'s' if totales['cantidad_sin_precio'] != 1 else ''} "
                    "sin precio cargado (marcadas SIN PRECIO) — no suman al total, completalas antes de facturar.",
                    estilo_aviso,
                )
            )

    documento.build(elementos, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_ingresos_deposito(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], grupos: list[dict], totales: dict
) -> bytes:
    """Arma el Excel de Ingresos a Depósito: secciones por proveedor con subtotal + total general al final."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Ingresos a Depósito"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_proveedor = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_marca = Font(bold=True, color=ROJO_MARCA_HEX, size=9)
    fuente_subtotal = Font(bold=True)
    fuente_total = Font(bold=True, size=13)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Ingresos a Depósito")
    for columna in range(1, 13):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(row=fila_actual, column=1, value=_armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto)).font = fuente_normal
    fila_actual += 2

    if not grupos:
        hoja.cell(row=fila_actual, column=1, value="No se encontraron ingresos con estos filtros.").font = fuente_normal

    encabezados = (
        "Fecha", "Hora", "Guía", "Artículo", "Cajones reales", "Contenido real",
        "Rechazados", "Precio por bulto", "Seña por bulto", "Total seña", "Total", "Estado",
    )
    for grupo in grupos:
        hoja.cell(
            row=fila_actual, column=1, value=f"{grupo['proveedor_nombre']} ({grupo['proveedor_codigo_puesto']})"
        ).font = fuente_proveedor
        fila_actual += 1

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for fila in grupo["filas"]:
            hoja.cell(row=fila_actual, column=1, value=fila["procesada_el"].strftime("%d/%m/%Y") if fila["procesada_el"] else "—")
            hoja.cell(row=fila_actual, column=2, value=fila["procesada_el"].strftime("%H:%M") if fila["procesada_el"] else "—")
            hoja.cell(row=fila_actual, column=3, value=_texto_guia(fila))
            hoja.cell(row=fila_actual, column=4, value=fila["articulo_nombre"])
            if fila["cantidad_cajones_real"] is not None:
                hoja.cell(row=fila_actual, column=5, value=float(fila["cantidad_cajones_real"]))
            if fila["contenido_por_cajon_real"] is not None:
                sufijo = SUFIJOS_UNIDAD_COMPRA.get(fila.get("unidad_compra"), "")
                hoja.cell(row=fila_actual, column=6, value=f"{_formatear_numero(fila['contenido_por_cajon_real'])}{sufijo}")
            if fila["cantidad_cajones_rechazada"] is not None:
                hoja.cell(row=fila_actual, column=7, value=float(fila["cantidad_cajones_rechazada"]))
            if fila["sin_precio"]:
                hoja.cell(row=fila_actual, column=8, value="SIN PRECIO").font = fuente_marca
            elif fila["importe"] is not None:
                celda_precio = hoja.cell(row=fila_actual, column=8, value=float(fila["importe"]))
                celda_precio.number_format = '"$"#,##0'
            if fila.get("sena") is not None:
                celda_sena = hoja.cell(row=fila_actual, column=9, value=float(fila["sena"]))
                celda_sena.number_format = '"$"#,##0'
            if fila.get("total_sena") is not None:
                celda_total_sena = hoja.cell(row=fila_actual, column=10, value=float(fila["total_sena"]))
                celda_total_sena.number_format = '"$"#,##0'
            if fila["total"] is not None:
                celda_total = hoja.cell(row=fila_actual, column=11, value=float(fila["total"]))
                celda_total.number_format = '"$"#,##0'
            celda_estado = hoja.cell(row=fila_actual, column=12, value=_texto_estado(fila))
            if fila["estado_etiqueta"] != "Recepcionada":
                celda_estado.font = fuente_marca
            fila_actual += 1

        hoja.cell(row=fila_actual, column=1, value="Subtotal").font = fuente_subtotal
        celda_subtotal = hoja.cell(row=fila_actual, column=11, value=float(grupo["subtotal"]))
        celda_subtotal.font = fuente_subtotal
        celda_subtotal.number_format = '"$"#,##0'
        fila_actual += 2

    if grupos:
        hoja.cell(row=fila_actual, column=1, value="Total general").font = fuente_total
        celda_total_general = hoja.cell(row=fila_actual, column=11, value=float(totales["total_general"]))
        celda_total_general.font = fuente_total
        celda_total_general.number_format = '"$"#,##0'
        fila_actual += 1
        if totales["cantidad_sin_precio"] > 0:
            hoja.cell(
                row=fila_actual, column=1,
                value=f"Ojo: {totales['cantidad_sin_precio']} compras sin precio cargado — no suman al total, "
                "completalas antes de facturar.",
            ).font = fuente_marca

    for columna, ancho in enumerate((12, 8, 9, 22, 14, 14, 12, 15, 14, 12, 13, 24), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
