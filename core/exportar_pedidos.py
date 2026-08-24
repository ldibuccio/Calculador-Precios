"""Genera Buscar Pedidos en PDF y Excel — puro, sin tocar la base.

Mismo criterio visual que los otros exports (core/exportar_ingresos.py):
banda de encabezado por página, una tabla por FECHA de pedido con su
subtotal, y el total general al final. Los kilos son SIEMPRE los
kilos_enviados que grabó el depósito al armar (lo que se factura): un
renglón sin kilaje dice "sin kilaje" — jamás se calcula el de la ficha
acá. Los anulados se muestran marcados y no suman.

grupos/totales: los que arma _grupos_buscar_pedidos en app/main.py.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
ROJO_MARCA = colors.Color(0.6, 0.11, 0.11)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"
ROJO_MARCA_HEX = "991B1B"

OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm


def _formatear_numero(valor) -> str:
    if valor is None:
        return "—"
    return f"{float(valor):.2f}".rstrip("0").rstrip(".")


def _texto_kilos(fila: dict) -> str:
    if fila["anulado"]:
        return "—"
    if fila["kilos"] is None:
        return "SIN KILAJE"
    return f"{_formatear_numero(fila['kilos'])} kg"


def _texto_estado(fila: dict) -> str:
    if fila["anulado"]:
        return "Anulado"
    if not fila["armado"]:
        return "Sin armar"
    return "Armado"


def _armar_subtitulo(fecha_desde: date, fecha_hasta: date, nombre_cliente: str) -> str:
    return (
        f"Cliente {nombre_cliente} — pedidos del {fecha_desde.strftime('%d/%m/%Y')} al "
        f"{fecha_hasta.strftime('%d/%m/%Y')}. Los kilos son los ENVIADOS por el depósito "
        "(lo que se factura), nunca los de la ficha."
    )


def _dibujar_encabezado(canvas, documento, subtitulo: str):
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Pedidos")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 9)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def generar_pdf_pedidos(
    fecha_desde: date, fecha_hasta: date, nombre_cliente: str, grupos: list[dict], totales: dict
) -> bytes:
    """Arma el PDF de Buscar Pedidos: una tabla por fecha con subtotal + total general al final."""
    buffer = BytesIO()
    subtitulo = _armar_subtitulo(fecha_desde, fecha_hasta, nombre_cliente)
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado(canvas, doc, subtitulo)

    estilo_titulo_tabla = ParagraphStyle(
        "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=8.5, textColor=VERDE_ENCABEZADO
    )
    estilo_dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=8.5, textColor=colors.black)
    estilo_numero = ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.black)
    estilo_marca = ParagraphStyle("marca", fontName="Helvetica-Bold", fontSize=8, textColor=ROJO_MARCA)
    estilo_gris = ParagraphStyle("gris", fontName="Helvetica", fontSize=8, textColor=GRIS_TEXTO_AYUDA)
    estilo_subtotal = ParagraphStyle("subtotal", fontName="Helvetica-Bold", fontSize=9, textColor=colors.black)
    estilo_total = ParagraphStyle("total", fontName="Helvetica-Bold", fontSize=13, textColor=colors.black)
    estilo_aviso = ParagraphStyle("aviso", fontName="Helvetica-Bold", fontSize=9.5, textColor=ROJO_MARCA)
    estilo_vacio = ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []
    if not grupos:
        elementos.append(Paragraph("No se encontraron pedidos con estos filtros.", estilo_vacio))

    encabezados = ("Artículo", "Sucursal", "Bultos", "Kilos enviados", "Estado")
    anchos = [ancho_util * 0.34, ancho_util * 0.14, ancho_util * 0.14, ancho_util * 0.2, ancho_util * 0.18]
    for indice_grupo, grupo in enumerate(grupos):
        if indice_grupo > 0:
            elementos.append(Spacer(1, 14))

        datos_tabla = [
            [Paragraph(f"Pedido del {grupo['fecha_mostrar']}", estilo_titulo_tabla)] + [""] * (len(encabezados) - 1),
            [Paragraph(encabezado, estilo_encabezado_tabla) for encabezado in encabezados],
        ]
        estilos_filas = []
        for indice, fila in enumerate(grupo["filas"]):
            estilo_fila = estilo_gris if fila["anulado"] else estilo_dato
            kilos_texto = _texto_kilos(fila)
            estilo_kilos = estilo_marca if kilos_texto == "SIN KILAJE" else estilo_numero
            datos_tabla.append(
                [
                    Paragraph(fila["articulo_nombre"], estilo_fila),
                    Paragraph(fila["sucursal"] or "—", estilo_fila),
                    Paragraph(_formatear_numero(fila["bultos"]), estilo_fila),
                    Paragraph(kilos_texto, estilo_gris if fila["anulado"] else estilo_kilos),
                    Paragraph(_texto_estado(fila), estilo_marca if fila["anulado"] or not fila["armado"] else estilo_gris),
                ]
            )
            if indice % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice + 2), (-1, indice + 2), GRIS_FILA_ALTERNADA))

        indice_subtotal = len(datos_tabla)
        subtotal_kilos = f"{_formatear_numero(grupo['kilos'])} kg"
        if grupo["sin_kilaje"]:
            subtotal_kilos += f" ({grupo['sin_kilaje']} sin kilaje)"
        datos_tabla.append(
            [
                Paragraph("Subtotal", estilo_subtotal),
                "",
                Paragraph(_formatear_numero(grupo["bultos"]), estilo_subtotal),
                Paragraph(subtotal_kilos, estilo_subtotal),
                "",
            ]
        )

        tabla = Table(datos_tabla, colWidths=anchos, repeatRows=2)
        tabla.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 2), (-1, -2), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("LINEABOVE", (0, indice_subtotal), (-1, indice_subtotal), 1, VERDE_ENCABEZADO),
                    ("TOPPADDING", (0, 0), (-1, 0), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

    if grupos:
        elementos.append(Spacer(1, 16))
        elementos.append(
            Paragraph(
                f"Total: {_formatear_numero(totales['bultos'])} bultos — {_formatear_numero(totales['kilos'])} kg enviados",
                estilo_total,
            )
        )
        if totales["sin_kilaje"] or totales["anulados"]:
            partes = []
            if totales["sin_kilaje"]:
                partes.append(f"{totales['sin_kilaje']} {'renglón' if totales['sin_kilaje'] == 1 else 'renglones'} sin kilaje (no suman kilos)")
            if totales["anulados"]:
                partes.append(f"{totales['anulados']} anulado{'s' if totales['anulados'] != 1 else ''} (no suman)")
            elementos.append(Spacer(1, 5))
            elementos.append(Paragraph("Ojo: " + " — ".join(partes) + ".", estilo_aviso))

    documento.build(elementos, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_pedidos(
    fecha_desde: date, fecha_hasta: date, nombre_cliente: str, grupos: list[dict], totales: dict
) -> bytes:
    """Arma el Excel de Buscar Pedidos: secciones por fecha con subtotal + total general al final."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Pedidos"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_fecha = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_marca = Font(bold=True, color=ROJO_MARCA_HEX, size=9)
    fuente_subtotal = Font(bold=True)
    fuente_total = Font(bold=True, size=13)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Pedidos")
    for columna in range(1, 7):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(row=fila_actual, column=1, value=_armar_subtitulo(fecha_desde, fecha_hasta, nombre_cliente)).font = fuente_normal
    fila_actual += 2

    if not grupos:
        hoja.cell(row=fila_actual, column=1, value="No se encontraron pedidos con estos filtros.").font = fuente_normal

    encabezados = ("Fecha", "Artículo", "Sucursal", "Bultos", "Kilos enviados", "Estado")
    for grupo in grupos:
        hoja.cell(row=fila_actual, column=1, value=f"Pedido del {grupo['fecha_mostrar']}").font = fuente_fecha
        fila_actual += 1

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for fila in grupo["filas"]:
            hoja.cell(row=fila_actual, column=1, value=grupo["fecha_mostrar"])
            hoja.cell(row=fila_actual, column=2, value=fila["articulo_nombre"])
            hoja.cell(row=fila_actual, column=3, value=fila["sucursal"] or "—")
            hoja.cell(row=fila_actual, column=4, value=float(fila["bultos"]))
            if fila["anulado"]:
                hoja.cell(row=fila_actual, column=5, value="—")
            elif fila["kilos"] is not None:
                hoja.cell(row=fila_actual, column=5, value=round(float(fila["kilos"]), 2))
            else:
                hoja.cell(row=fila_actual, column=5, value="SIN KILAJE").font = fuente_marca
            celda_estado = hoja.cell(row=fila_actual, column=6, value=_texto_estado(fila))
            if fila["anulado"] or not fila["armado"]:
                celda_estado.font = fuente_marca
            fila_actual += 1

        hoja.cell(row=fila_actual, column=1, value="Subtotal").font = fuente_subtotal
        celda = hoja.cell(row=fila_actual, column=4, value=float(grupo["bultos"]))
        celda.font = fuente_subtotal
        celda = hoja.cell(row=fila_actual, column=5, value=round(float(grupo["kilos"]), 2))
        celda.font = fuente_subtotal
        if grupo["sin_kilaje"]:
            hoja.cell(row=fila_actual, column=6, value=f"{grupo['sin_kilaje']} sin kilaje").font = fuente_marca
        fila_actual += 2

    if grupos:
        hoja.cell(row=fila_actual, column=1, value="Total").font = fuente_total
        celda = hoja.cell(row=fila_actual, column=4, value=float(totales["bultos"]))
        celda.font = fuente_total
        celda = hoja.cell(row=fila_actual, column=5, value=round(float(totales["kilos"]), 2))
        celda.font = fuente_total
        fila_actual += 1
        if totales["sin_kilaje"] or totales["anulados"]:
            hoja.cell(
                row=fila_actual, column=1,
                value=f"Ojo: {totales['sin_kilaje']} sin kilaje (no suman kilos) — {totales['anulados']} anulados (no suman).",
            ).font = fuente_marca

    for columna, ancho in enumerate((12, 26, 10, 9, 14, 12), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
