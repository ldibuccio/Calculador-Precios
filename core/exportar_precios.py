"""Genera la Lista de Precios en PDF y Excel — puro, sin tocar la base ni la red.

Recibe los datos ya armados (ver "filas" en cada función) y devuelve los
bytes del archivo. No sabe nada de clientes, fichas ni de la base — eso lo
resuelve quien llama (app/main.py). Así se puede reusar tal cual desde
/precios/consultar y, más adelante, desde "Guardar y generar listado" en
la carga de precios, sin duplicar nada.

Los archivos son temporarios: esto solo arma bytes en memoria, nunca los
guarda en ningún lado.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
ROJO_PRECIO_NUEVO = colors.Color(0.80, 0.10, 0.10)
ROSA_FILA_NUEVA = colors.Color(0.98, 0.91, 0.91)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
VERDE_ENCABEZADO_HEX = "2E8C57"  # equivalente aproximado de (0.18, 0.55, 0.34) en hex, para Excel
ROJO_PRECIO_NUEVO_HEX = "CC1A1A"
ROJO_FONDO_CLARO_HEX = "FCE4E4"

# Orden fijo de las secciones. Lo que no tiene ninguno de los grupos
# válidos cae en "Sin clasificar", al final, para que se note en vez de
# perderse.
ORDEN_GRUPOS = [
    ("fruta", "FRUTA"),
    ("hortaliza", "HORTALIZA"),
    ("hoja", "HOJA"),
    ("pesada", "PESADA"),
    (None, "SIN CLASIFICAR"),
]

TEXTO_UNIDAD = {"kilo": "por kilo", "unidad": "por unidad", "cubeta": "por cubeta"}


def _texto_unidad(unidad) -> str:
    return TEXTO_UNIDAD.get(unidad, "—")


def _formatear_moneda(valor) -> str:
    """"$ 45.000": símbolo $ con espacio, "." cada tres cifras, redondeado al peso entero."""
    entero = round(float(valor))
    negativo = entero < 0
    texto = f"{abs(entero):,}".replace(",", ".")
    return f"$ {'-' if negativo else ''}{texto}"


def _agrupar_y_ordenar_filas(filas: list[dict]) -> dict:
    """Agrupa las filas por grupo (fruta/hortaliza/hoja/pesada/sin clasificar) y ordena cada sección por nombre."""
    grupos: dict = {clave: [] for clave, _ in ORDEN_GRUPOS}
    for fila in filas:
        clave = fila.get("grupo") if fila.get("grupo") in grupos else None
        grupos[clave].append(fila)
    for lista_filas in grupos.values():
        lista_filas.sort(key=lambda f: f["articulo_nombre"])
    return grupos


LEYENDA_PRECIO_NUEVO = "Nuevo precio indica los productos cuyo precio fue actualizado."
PIE_PAGINA = "* Todos los precios están expresados en pesos y no incluyen IVA."


# Posiciones (desde el borde superior de la página) de cada línea del
# encabezado: título negro a la izquierda, Cliente, Vigencia, el filete
# verde fino, y la leyenda con el puntito rojo — sin banda de color, como
# en el formato de referencia.
OFFSET_TITULO = 15 * mm
OFFSET_CLIENTE = 22 * mm
OFFSET_VIGENCIA = 27.5 * mm
OFFSET_LINEA = 32 * mm
OFFSET_LEYENDA = 38 * mm
ALTURA_ENCABEZADO = 46 * mm  # topMargin: deja aire entre la leyenda y el cuerpo


def _dibujar_encabezado(canvas, documento, cliente_nombre: str, fecha_texto: str, nombre_empresa: str):
    """Título, Cliente, Vigencia, filete verde y leyenda — directo en el canvas, en TODAS las páginas.

    Va en el canvas (no como Paragraph en el flujo normal) para que se
    repita en cada página, incluso si una sección se corta y sigue en la
    siguiente — un flowable normal solo aparece una vez, en el lugar donde
    cae dentro del flujo.

    El título lleva el nombre de la empresa ("Lista de Precios — Frutamax"):
    el mismo cliente puede recibir listas de más de una empresa, sin el
    nombre no puede distinguirlas.
    """
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, f"Lista de Precios — {nombre_empresa}")

    canvas.setFont("Helvetica", 10)
    canvas.drawString(x, alto_pagina - OFFSET_CLIENTE, f"Cliente: {cliente_nombre}")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.drawString(x, alto_pagina - OFFSET_VIGENCIA, f"Vigencia: {fecha_texto} · Precios + IVA")

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    # El puntito va en un drawString aparte (con el espacio incluido en el
    # mismo texto) para que quede realmente rojo y separado de la leyenda
    # gris que sigue, sin depender de cómo un lector de PDF una texto
    # dibujado en llamadas distintas.
    y_leyenda = alto_pagina - OFFSET_LEYENDA
    canvas.setFont("Helvetica-Oblique", 9)
    canvas.setFillColor(ROJO_PRECIO_NUEVO)
    prefijo_bullet = "• "
    canvas.drawString(x, y_leyenda, prefijo_bullet)
    ancho_bullet = canvas.stringWidth(prefijo_bullet, "Helvetica-Oblique", 9)
    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.drawString(x + ancho_bullet, y_leyenda, LEYENDA_PRECIO_NUEVO)

    canvas.restoreState()


def generar_pdf_lista_precios(
    cliente_nombre: str, fecha: date, filas: list[dict], nombre_empresa: str
) -> bytes:
    """Arma el PDF de la Lista de Precios de un cliente a una fecha, con el formato ya definido.

    filas: [{"articulo_nombre", "grupo", "precio", "unidad", "es_nuevo"}, ...]. es_nuevo es "empezó
    a regir en la fecha de esta lista" y se resalta siempre, sea hoy o una fecha pasada: consultar
    para atrás es justamente para ver qué cambió ESE día.

    Cada grupo (Fruta/Hortaliza/Pesada/Sin clasificar) arranca en su propia página, y lleva su título
    de sección DENTRO de la tabla (repeatRows=2, junto con el encabezado de columnas) para que, si una
    sección es tan larga que igual se corta sola entre dos páginas, el título se repita solo.
    """
    buffer = BytesIO()
    fecha_texto = fecha.strftime("%d/%m/%Y")
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _dibujar_encabezado_pagina(canvas, documento):
        _dibujar_encabezado(canvas, documento, cliente_nombre, fecha_texto, nombre_empresa)

    estilo_ayuda = ParagraphStyle(
        "ayuda", fontName="Helvetica-Oblique", fontSize=9, textColor=GRIS_TEXTO_AYUDA, spaceBefore=16
    )
    estilo_seccion_en_tabla = ParagraphStyle(
        "seccion_en_tabla", fontName="Helvetica-Bold", fontSize=13, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=9.5, textColor=VERDE_ENCABEZADO
    )
    estilo_producto = ParagraphStyle("producto", fontName="Helvetica", fontSize=9.5, textColor=colors.black)
    estilo_precio = ParagraphStyle("precio", fontName="Helvetica-Bold", fontSize=9.5, textColor=VERDE_ENCABEZADO)
    estilo_precio_nuevo = ParagraphStyle(
        "precio_nuevo", fontName="Helvetica-Bold", fontSize=9.5, textColor=ROJO_PRECIO_NUEVO
    )
    estilo_badge_nuevo = ParagraphStyle("badge_nuevo", fontName="Helvetica-Bold", fontSize=8, textColor=ROJO_PRECIO_NUEVO)
    estilo_unidad = ParagraphStyle("unidad", fontName="Helvetica", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []

    grupos = _agrupar_y_ordenar_filas(filas)
    hay_grupo_previo = False
    for clave, titulo in ORDEN_GRUPOS:
        filas_grupo = grupos[clave]
        if not filas_grupo:
            continue

        if hay_grupo_previo:
            elementos.append(PageBreak())
        hay_grupo_previo = True

        datos_tabla = [
            [Paragraph(titulo, estilo_seccion_en_tabla), "", "", ""],
            [
                Paragraph("Producto", estilo_encabezado_tabla),
                Paragraph("Precio", estilo_encabezado_tabla),
                Paragraph("Unidad", estilo_encabezado_tabla),
                "",
            ],
        ]
        # Comandos de TableStyle propios de cada fila de datos (resaltado
        # rosa + franja roja izquierda en las nuevas, gris muy suave
        # alternado en el resto) — se arman acá porque dependen del índice
        # real de cada fila dentro de la tabla (2 = primera fila de datos,
        # después del título y el encabezado de columnas).
        estilos_filas = []
        for indice_dato, fila in enumerate(filas_grupo):
            es_nueva = bool(fila.get("es_nuevo"))
            indice_tabla = indice_dato + 2

            precio_texto = _formatear_moneda(fila["precio"])
            celda_precio = Paragraph(precio_texto, estilo_precio_nuevo if es_nueva else estilo_precio)
            celda_badge = Paragraph("• Nuevo precio", estilo_badge_nuevo) if es_nueva else ""

            datos_tabla.append(
                [
                    Paragraph(fila["articulo_nombre"], estilo_producto),
                    celda_precio,
                    Paragraph(_texto_unidad(fila.get("unidad")), estilo_unidad),
                    celda_badge,
                ]
            )

            if es_nueva:
                estilos_filas.append(("BACKGROUND", (0, indice_tabla), (-1, indice_tabla), ROSA_FILA_NUEVA))
                estilos_filas.append(("LINEBEFORE", (0, indice_tabla), (0, indice_tabla), 3, ROJO_PRECIO_NUEVO))
            elif indice_dato % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice_tabla), (-1, indice_tabla), GRIS_FILA_ALTERNADA))

        tabla = Table(
            datos_tabla,
            colWidths=[ancho_util * 0.40, ancho_util * 0.16, ancho_util * 0.16, ancho_util * 0.28],
            repeatRows=2,
        )
        tabla.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 2), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("TOPPADDING", (0, 0), (-1, 0), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("TOPPADDING", (0, 1), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

    elementos.append(Paragraph(PIE_PAGINA, estilo_ayuda))

    documento.build(elementos, onFirstPage=_dibujar_encabezado_pagina, onLaterPages=_dibujar_encabezado_pagina)
    return buffer.getvalue()


# Formato del Excel de la Lista de Precios (pedido explícito, 21/08/2026):
# el diseño que el dueño ya usaba a mano — fecha arriba, tres columnas
# (Producto | Precio Anterior | Precio Desde HOY), formato contable con
# centavos, cambios resaltados en naranja con el precio en rojo, y el pie
# "PRECIOS POR KG - SIN IVA". Cambiarlo solo si él lo pide.
NARANJA_CAMBIO_HEX = "FFC000"
ROJO_CAMBIO_HEX = "C00000"
FORMATO_CONTABLE = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"_);_(@_)'


def generar_excel_lista_precios(
    cliente_nombre: str, fecha: date, filas: list[dict], es_hoy: bool, nombre_empresa: str
) -> bytes:
    """Arma el Excel de la Lista de Precios con el formato de planilla que el cliente ya conoce.

    Una sola tabla alfabética (sin secciones por grupo — eso queda para el
    PDF): fecha arriba, columnas Producto | Precio Anterior | Precio Desde
    HOY, y el pie "PRECIOS POR KG - SIN IVA". El precio anterior figura
    SIEMPRE: si el artículo nunca cambió de precio, se repite el vigente —
    así el cliente ve todas las filas iguales salvo las resaltadas.

    El resaltado (fondo naranja + precio en rojo) marca las filas cuyo
    precio EMPEZÓ A REGIR en la fecha de esta lista (fila["es_nuevo"]), no
    las que difieren del anterior: un precio que cambió hace una semana no
    es una novedad de hoy y resaltarlo llenaba la lista de falsos avisos.
    Para una fecha pasada, el encabezado dice "Precio al dd/mm" en vez de
    "Precio Desde HOY" — pero el resaltado va igual, que es para lo que se
    consulta para atrás.

    Los artículos que NO se venden por kilo llevan la unidad pegada al
    nombre ("Mango (por unidad)"): el pie dice "POR KG" y sin esa
    aclaración el precio se leería mal.

    cliente_nombre/nombre_empresa no aparecen en la hoja (la planilla
    original no los lleva): viajan en el nombre del archivo, que arma
    quien llama.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Lista de Precios"

    borde_fino = Side(style="thin", color="000000")
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    centrado = Alignment(horizontal="center", vertical="center")
    fuente_encabezado = Font(bold=True, size=11)
    relleno_naranja = PatternFill(start_color=NARANJA_CAMBIO_HEX, end_color=NARANJA_CAMBIO_HEX, fill_type="solid")
    fuente_roja = Font(bold=True, color=ROJO_CAMBIO_HEX)

    # Fila 1: la fecha, sola, centrada sobre las tres columnas (21/8/2026,
    # sin ceros a la izquierda — igual que la planilla original).
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    celda_fecha = hoja.cell(row=1, column=1, value=f"{fecha.day}/{fecha.month}/{fecha.year}")
    celda_fecha.alignment = centrado
    celda_fecha.font = fuente_encabezado
    for columna in range(1, 4):
        hoja.cell(row=1, column=columna).border = borde

    encabezado_precio = "Precio Desde HOY" if es_hoy else f"Precio al {fecha.day}/{fecha.month}"
    for columna, encabezado in enumerate(("Producto", "Precio Anterior", encabezado_precio), start=1):
        celda = hoja.cell(row=2, column=columna, value=encabezado)
        celda.font = fuente_encabezado
        celda.border = borde

    fila_actual = 3
    for fila in sorted(filas, key=lambda f: f["articulo_nombre"].lower()):
        nombre = fila["articulo_nombre"]
        unidad = fila.get("unidad")
        if unidad and unidad != "kilo":
            nombre = f"{nombre} ({_texto_unidad(unidad)})"

        precio = float(fila["precio"])
        # El anterior SIEMPRE figura: sin un precio previo cargado, se
        # repite el vigente (fila sin cambio, no se resalta).
        precio_anterior = float(fila["precio_anterior"]) if fila.get("precio_anterior") is not None else precio
        cambio = bool(fila.get("es_nuevo"))

        celda_nombre = hoja.cell(row=fila_actual, column=1, value=nombre)
        celda_nombre.border = borde

        celda_anterior = hoja.cell(row=fila_actual, column=2, value=precio_anterior)
        celda_anterior.number_format = FORMATO_CONTABLE
        celda_anterior.border = borde

        celda_precio = hoja.cell(row=fila_actual, column=3, value=precio)
        celda_precio.number_format = FORMATO_CONTABLE
        celda_precio.border = borde
        if cambio:
            celda_precio.fill = relleno_naranja
            celda_precio.font = fuente_roja

        fila_actual += 1

    hoja.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=3)
    celda_pie = hoja.cell(row=fila_actual, column=1, value="PRECIOS POR KG - SIN IVA")
    celda_pie.alignment = centrado
    celda_pie.font = fuente_encabezado
    for columna in range(1, 4):
        hoja.cell(row=fila_actual, column=columna).border = borde

    hoja.column_dimensions[get_column_letter(1)].width = 24
    hoja.column_dimensions[get_column_letter(2)].width = 17
    hoja.column_dimensions[get_column_letter(3)].width = 17

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
