"""Genera el Excel del Remanente del depósito — puro, sin tocar la base ni la red.

Mismo molde que exportar_disponibles: recibe las porciones ya armadas y
devuelve los bytes. No sabe de dónde salen los números — eso lo resuelve
app/main.py.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

# Los grupos y su orden salen del MISMO lugar que los usa el resto del sistema.
# Escribir acá otra lista sería una segunda versión de "cuáles son los grupos",
# y el día que se agregue uno este archivo lo mandaría al final sin que nadie
# se entere.
from core.rentabilidad import ETIQUETA_SIN_GRUPO, ETIQUETAS_GRUPO, ORDEN_GRUPOS

# Las cajas armadas van APARTE y al final: en el piso son una pila propia, no
# están mezcladas con la fruta suelta, y el que cuenta las cuenta en otro
# momento. No es un grupo de artículo — por eso no está en ORDEN_GRUPOS.
SECCION_PROCESADAS = "Cajas Procesadas"

FILA_ENCABEZADO = 4
FILA_PRIMER_DATO = FILA_ENCABEZADO + 1

AZUL_ENCABEZADO_HEX = "1F4E79"  # el mismo de Disponibles

_BORDE_FINO = Side(style="thin", color="000000")
_BORDE_CELDA = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_FINO, bottom=_BORDE_FINO)
# El total se despega de los datos con una raya gruesa arriba: impreso, es lo
# que hace que se lea como cierre y no como un renglón más de la lista.
_BORDE_GRUESO = Side(style="medium", color="000000")
_BORDE_TOTAL = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_GRUESO, bottom=_BORDE_FINO)

GRIS_SECCION_HEX = "D9E2F3"

# El subtotal se pinta IGUAL que el título de su sección, y es el punto: los dos
# son chrome, no mercadería. Impreso, la sección queda encerrada entre dos
# franjas grises —encabezado arriba, cierre abajo— y ningún renglón con número
# adentro se confunde con ellas. Sin el relleno, "Subtotal Hortaliza · 132" es
# una fila con nombre y número, o sea exactamente lo que parece un producto.
_BORDE_SUBTOTAL = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_FINO, bottom=_BORDE_FINO)


def _secciones(porciones: list[dict]) -> list[tuple]:
    """[(título, [porciones]), ...] en el orden en que se recorre el depósito.

    Primero los grupos de artículo, en el orden fijo del sistema, y al final
    las cajas procesadas. Una sección sin nada no sale: una hoja impresa con
    "HOJA" y ningún renglón abajo hace dudar de si falta algo o no hay.
    """
    procesadas = [p for p in porciones if p.get("procesada")]
    sueltas = [p for p in porciones if not p.get("procesada")]

    secciones = []
    for grupo in ORDEN_GRUPOS:
        del_grupo = [p for p in sueltas if p.get("grupo") == grupo]
        if del_grupo:
            secciones.append((ETIQUETAS_GRUPO.get(grupo, ETIQUETA_SIN_GRUPO), del_grupo))
    # Un grupo que el sistema todavía no conoce no puede desaparecer del
    # archivo: cae en "Sin grupo" con el resto, que es donde se va a notar.
    conocidos = set(ORDEN_GRUPOS)
    huerfanas = [p for p in sueltas if p.get("grupo") not in conocidos]
    if huerfanas:
        secciones.append((ETIQUETA_SIN_GRUPO, huerfanas))
    if procesadas:
        secciones.append((SECCION_PROCESADAS, procesadas))
    return secciones


def generar_excel_remanente(fecha: date, porciones: list[dict]) -> bytes:
    """El remanente en una hoja: una fila por porción más el total al pie, en el orden en que viene.

    porciones: [{"nombre", "bultos", "grupo", "procesada"}, ...] — ya
    ordenadas por app/main.py. Acá no se REORDENA nada: solo se reparten en
    secciones respetando el orden que traen, así que dentro de cada grupo
    quedan igual que en la pantalla. Si el Excel las reordenara, serían dos
    criterios y se irían separando.

    LAS SECCIONES son los grupos del artículo (Fruta, Hortaliza, Hoja,
    Pesada, Sin grupo) más una propia al final, "Cajas Procesadas", con las
    cajas armadas a una ficha. Es para caminar el depósito: las cajas armadas
    son una pila aparte y se cuentan en otro momento. La segunda NO va ahí
    —son bultos sueltos de calidad menor, no cajas— y se queda con su
    artículo.

    SIN PLATA, y no es un olvido: el costo por bulto vive en el LOTE, no en
    el artículo —los mismos 30 bultos pueden venir de tres compras a tres
    precios—, así que poner un número de plata acá obligaría a elegir una
    valuación que ninguna pantalla calcula hoy. Y para lo que sirve este
    archivo no hace falta: se cuentan cajones, no pesos.

    La columna "Contado" va VACÍA a propósito: es para escribir a mano
    contra el conteo físico. Si saliera precargada con lo del sistema, el
    que cuenta transcribe en vez de contar y se pierde el control cruzado
    — el mismo criterio de la pantalla de Stock Físico.

    Cada sección CIERRA con su subtotal y al pie va el total general. No hay
    total por ARTÍCULO, que es otra cosa y es justo la suma que el dueño pidió
    no mostrar ("80 cajones sin procesar + 40 cajas armadas no son 120
    bultos"): el subtotal es de la sección, o sea de una zona del depósito.

    Para qué sirve cada uno, que no es lo mismo: el del pie dice si el archivo
    impreso está completo (si no se cortó una hoja); el de la sección acota
    DÓNDE BUSCAR cuando algo no cierra, sin tener que rehacer la suma entera.

    El número va COMO VALOR, no como fórmula: lo que importa es lo que
    quedó impreso en el papel, y una fórmula no se imprime distinto pero
    sí puede cambiar si alguien toca una celda antes de imprimir.

    La celda "Contado" del total también va vacía, por lo mismo que las de
    arriba: si el que cuenta ve un total del sistema al pie, tiene contra
    qué cuadrar sin haber contado.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Remanente"

    hoja.merge_cells("A1:C1")
    hoja["A1"] = "Remanente del depósito"
    hoja["A1"].font = Font(bold=True, size=14)

    hoja.merge_cells("A2:C2")
    hoja["A2"] = f"Al {fecha.strftime('%d/%m/%Y')}"
    hoja["A2"].font = Font(bold=True)

    relleno = PatternFill(start_color=AZUL_ENCABEZADO_HEX, end_color=AZUL_ENCABEZADO_HEX, fill_type="solid")
    for columna, encabezado in enumerate(("Producto", "Sistema", "Contado"), start=1):
        celda = hoja.cell(row=FILA_ENCABEZADO, column=columna, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
        celda.border = _BORDE_CELDA

    relleno_seccion = PatternFill(start_color=GRIS_SECCION_HEX, end_color=GRIS_SECCION_HEX,
                                  fill_type="solid")
    fila_actual = FILA_PRIMER_DATO
    for titulo, del_grupo in _secciones(porciones):
        for columna in (1, 2, 3):
            celda = hoja.cell(row=fila_actual, column=columna,
                              value=titulo if columna == 1 else None)
            celda.font = Font(bold=True)
            celda.fill = relleno_seccion
            celda.border = _BORDE_CELDA
        fila_actual += 1
        for porcion in del_grupo:
            hoja.cell(row=fila_actual, column=1, value=porcion["nombre"])
            hoja.cell(row=fila_actual, column=2, value=float(porcion["bultos"]))
            # La tercera queda vacía: se llena a mano.
            for columna in (1, 2, 3):
                hoja.cell(row=fila_actual, column=columna).border = _BORDE_CELDA
            fila_actual += 1

        # El cierre de la sección. Mismas tres reglas que el total del pie:
        # el número va como VALOR (lo que importa es lo que quedó impreso), la
        # celda "Contado" va VACÍA (si el que cuenta ve un subtotal del sistema
        # ya tiene contra qué cuadrar sin haber contado), y se pinta como
        # chrome para que no se lea como un renglón de mercadería.
        cuantas = len(del_grupo)
        etiqueta = f"Subtotal {titulo} — {cuantas} {'renglón' if cuantas == 1 else 'renglones'}"
        subtotal = round(sum(float(p["bultos"]) for p in del_grupo), 2)
        for columna, valor in ((1, etiqueta), (2, subtotal), (3, None)):
            celda = hoja.cell(row=fila_actual, column=columna, value=valor)
            celda.font = Font(bold=True)
            celda.fill = relleno_seccion
            celda.border = _BORDE_SUBTOTAL
        fila_actual += 1

    # El total cuenta las PORCIONES, no las filas escritas: los títulos de
    # sección no son renglones que alguien tenga que ir a contar.
    renglones = len(porciones)
    total = round(sum(float(p["bultos"]) for p in porciones), 2)
    etiqueta = f"TOTAL — {renglones} {'renglón' if renglones == 1 else 'renglones'}"
    hoja.cell(row=fila_actual, column=1, value=etiqueta)
    hoja.cell(row=fila_actual, column=2, value=total)
    for columna in (1, 2, 3):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.font = Font(bold=True)
        celda.border = _BORDE_TOTAL

    hoja.column_dimensions["A"].width = 34
    hoja.column_dimensions["B"].width = 12
    hoja.column_dimensions["C"].width = 12

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
