"""Genera Rentabilidad de Pedidos en PDF y Excel — puro, sin tocar la base.

Mismo criterio visual que los otros exports (core/exportar_ingresos.py):
banda de encabezado repetida por página, una tabla por grupo de artículos
con su subtotal, el total general al final y — SIEMPRE que exista — la
sección de "no calculables" con su motivo y su peso en bultos: lo que no
se pudo calcular no suma como cero ni desaparece del papel.

grupos/totales/no_calculables: los que arma core.rentabilidad — esto solo
arma bytes en memoria, nunca guarda nada.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
ROJO_MARCA = colors.Color(0.6, 0.11, 0.11)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"
ROJO_MARCA_HEX = "991B1B"

OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm

SUFIJOS_UNIDAD_VENTA = {"kilo": "k", "unidad": "u", "cubeta": "c"}


def _formatear_numero(valor) -> str:
    if valor is None:
        return "—"
    return f"{float(valor):.2f}".rstrip("0").rstrip(".")


def _formatear_moneda(valor) -> str:
    if valor is None:
        return "—"
    entero = round(float(valor))
    negativo = entero < 0
    texto = f"{abs(entero):,}".replace(",", ".")
    return f"${'-' if negativo else ''}{texto}"


def _formatear_pct(valor) -> str:
    if valor is None:
        return "—"
    return f"{float(valor):.1f}%"


def _armar_subtitulo(fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], cantidad_fechas: int) -> str:
    subtitulo = (
        f"Pedidos del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} "
        f"({cantidad_fechas} día{'s' if cantidad_fechas != 1 else ''} con pedido) — "
        "bultos = lo PEDIDO, sin ajustar por armado (estimación, no facturación)"
    )
    if filtros_texto:
        subtitulo += " — " + ", ".join(filtros_texto)
    return subtitulo


def _dibujar_encabezado(canvas, documento, subtitulo: str):
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Rentabilidad de Pedidos")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 8.5)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def generar_pdf_rentabilidad(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], resultado: dict
) -> bytes:
    """Arma el PDF de Rentabilidad: una tabla por grupo con subtotal, total general y los no calculables al final."""
    buffer = BytesIO()
    subtitulo = _armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto, len(resultado["fechas_incluidas"]))
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado(canvas, doc, subtitulo)

    estilo_titulo_tabla = ParagraphStyle(
        "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=8.5, textColor=VERDE_ENCABEZADO
    )
    estilo_dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=8.5, textColor=colors.black)
    estilo_numero = ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.black)
    estilo_negativo = ParagraphStyle("negativo", fontName="Helvetica-Bold", fontSize=8.5, textColor=ROJO_MARCA)
    estilo_subtotal = ParagraphStyle("subtotal", fontName="Helvetica-Bold", fontSize=9, textColor=colors.black)
    estilo_total = ParagraphStyle("total", fontName="Helvetica-Bold", fontSize=13, textColor=colors.black)
    estilo_aviso = ParagraphStyle("aviso", fontName="Helvetica-Bold", fontSize=9.5, textColor=ROJO_MARCA)
    estilo_vacio = ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []
    grupos = resultado["grupos"]
    totales = resultado["totales"]
    no_calculables = resultado["no_calculables"]

    if not grupos and not no_calculables:
        elementos.append(Paragraph("No se encontraron pedidos con estos filtros.", estilo_vacio))

    encabezados = ("Artículo", "Bultos", "Precio", "Costo", "Venta", "Costo total", "Renta $", "Renta %")
    anchos = [
        ancho_util * 0.24, ancho_util * 0.08, ancho_util * 0.11, ancho_util * 0.11,
        ancho_util * 0.13, ancho_util * 0.13, ancho_util * 0.12, ancho_util * 0.08,
    ]
    for indice_grupo, grupo in enumerate(grupos):
        if indice_grupo > 0:
            elementos.append(Spacer(1, 14))

        datos_tabla = [
            [Paragraph(grupo["etiqueta"], estilo_titulo_tabla)] + [""] * (len(encabezados) - 1),
            [Paragraph(encabezado, estilo_encabezado_tabla) for encabezado in encabezados],
        ]
        estilos_filas = []
        for indice, fila in enumerate(grupo["filas"]):
            sufijo = SUFIJOS_UNIDAD_VENTA.get(fila["unidad_venta"], "")
            estilo_renta = estilo_negativo if fila["renta_pesos"] < 0 else estilo_numero
            datos_tabla.append(
                [
                    Paragraph(fila["articulo_nombre"], estilo_dato),
                    Paragraph(_formatear_numero(fila["bultos"]), estilo_dato),
                    Paragraph(f"{_formatear_moneda(fila['precio_promedio'])}/{sufijo}" if sufijo else _formatear_moneda(fila["precio_promedio"]), estilo_dato),
                    Paragraph(f"{_formatear_moneda(fila['costo_promedio'])}/{sufijo}" if sufijo else _formatear_moneda(fila["costo_promedio"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["venta"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["costo_total"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["renta_pesos"]), estilo_renta),
                    Paragraph(_formatear_pct(fila["renta_pct"]), estilo_renta),
                ]
            )
            if indice % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice + 2), (-1, indice + 2), GRIS_FILA_ALTERNADA))

        subtotal = grupo["subtotal"]
        indice_subtotal = len(datos_tabla)
        datos_tabla.append(
            [
                Paragraph("Subtotal", estilo_subtotal),
                Paragraph(_formatear_numero(subtotal["bultos"]), estilo_subtotal),
                "", "",
                Paragraph(_formatear_moneda(subtotal["venta"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["costo_total"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["renta_pesos"]), estilo_subtotal),
                Paragraph(_formatear_pct(subtotal["renta_pct"]), estilo_subtotal),
            ]
        )

        tabla = Table(datos_tabla, colWidths=anchos, repeatRows=2)
        tabla.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 2), (-1, -2), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("LINEABOVE", (0, indice_subtotal), (-1, indice_subtotal), 1, VERDE_ENCABEZADO),
                    ("TOPPADDING", (0, 0), (-1, 0), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

    if grupos:
        elementos.append(Spacer(1, 16))
        elementos.append(
            Paragraph(
                f"Total: venta {_formatear_moneda(totales['venta'])} — costo {_formatear_moneda(totales['costo_total'])} — "
                f"renta {_formatear_moneda(totales['renta_pesos'])} ({_formatear_pct(totales['renta_pct'])})",
                estilo_total,
            )
        )

    if no_calculables:
        elementos.append(Spacer(1, 14))
        elementos.append(
            Paragraph(
                f"Quedaron AFUERA del cálculo {totales['no_calculables_casos']} artículo{'s' if totales['no_calculables_casos'] != 1 else ''} "
                f"({_formatear_numero(totales['no_calculables_bultos'])} bultos) — no sumaron como cero:",
                estilo_aviso,
            )
        )
        elementos.append(Spacer(1, 6))
        datos_nc = [
            [
                Paragraph("Artículo", estilo_encabezado_tabla),
                Paragraph("Bultos", estilo_encabezado_tabla),
                Paragraph("Días", estilo_encabezado_tabla),
                Paragraph("Motivo", estilo_encabezado_tabla),
            ]
        ]
        for entrada in no_calculables:
            datos_nc.append(
                [
                    Paragraph(entrada["articulo_nombre"], estilo_dato),
                    Paragraph(_formatear_numero(entrada["bultos"]), estilo_dato),
                    Paragraph(str(entrada["dias"]), estilo_dato),
                    Paragraph(entrada["motivo_etiqueta"], estilo_dato),
                ]
            )
        tabla_nc = Table(
            datos_nc,
            colWidths=[ancho_util * 0.26, ancho_util * 0.1, ancho_util * 0.08, ancho_util * 0.56],
            repeatRows=1,
        )
        tabla_nc.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elementos.append(tabla_nc)

    documento.build(elementos, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_rentabilidad(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], resultado: dict
) -> bytes:
    """Arma el Excel de Rentabilidad: secciones por grupo con subtotal, total general y los no calculables al final."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Rentabilidad de Pedidos"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_grupo = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_marca = Font(bold=True, color=ROJO_MARCA_HEX, size=9)
    fuente_subtotal = Font(bold=True)
    fuente_total = Font(bold=True, size=13)

    grupos = resultado["grupos"]
    totales = resultado["totales"]
    no_calculables = resultado["no_calculables"]

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Rentabilidad de Pedidos")
    for columna in range(1, 11):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(
        row=fila_actual, column=1,
        value=_armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto, len(resultado["fechas_incluidas"])),
    ).font = fuente_normal
    fila_actual += 2

    if not grupos and not no_calculables:
        hoja.cell(row=fila_actual, column=1, value="No se encontraron pedidos con estos filtros.").font = fuente_normal

    encabezados = (
        "Artículo", "Bultos", "Unidades", "Unidad", "Precio", "Costo",
        "Venta", "Costo total", "Renta $", "Renta %",
    )
    for grupo in grupos:
        hoja.cell(row=fila_actual, column=1, value=grupo["etiqueta"]).font = fuente_grupo
        fila_actual += 1

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for fila in grupo["filas"]:
            hoja.cell(row=fila_actual, column=1, value=fila["articulo_nombre"])
            hoja.cell(row=fila_actual, column=2, value=float(fila["bultos"]))
            hoja.cell(row=fila_actual, column=3, value=round(float(fila["unidades"]), 2))
            hoja.cell(row=fila_actual, column=4, value=fila["unidad_venta"] or "—")
            for columna, valor in ((5, fila["precio_promedio"]), (6, fila["costo_promedio"]),
                                   (7, fila["venta"]), (8, fila["costo_total"]), (9, fila["renta_pesos"])):
                if valor is not None:
                    celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
                    celda.number_format = '"$"#,##0'
            if fila["renta_pct"] is not None:
                celda = hoja.cell(row=fila_actual, column=10, value=round(float(fila["renta_pct"]) / 100, 4))
                celda.number_format = "0.0%"
            if fila["renta_pesos"] < 0:
                hoja.cell(row=fila_actual, column=9).font = fuente_marca
            fila_actual += 1

        subtotal = grupo["subtotal"]
        hoja.cell(row=fila_actual, column=1, value="Subtotal").font = fuente_subtotal
        celda = hoja.cell(row=fila_actual, column=2, value=float(subtotal["bultos"]))
        celda.font = fuente_subtotal
        for columna, valor in ((7, subtotal["venta"]), (8, subtotal["costo_total"]), (9, subtotal["renta_pesos"])):
            celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
            celda.font = fuente_subtotal
            celda.number_format = '"$"#,##0'
        if subtotal["renta_pct"] is not None:
            celda = hoja.cell(row=fila_actual, column=10, value=round(float(subtotal["renta_pct"]) / 100, 4))
            celda.font = fuente_subtotal
            celda.number_format = "0.0%"
        fila_actual += 2

    if grupos:
        hoja.cell(row=fila_actual, column=1, value="Total").font = fuente_total
        for columna, valor in ((7, totales["venta"]), (8, totales["costo_total"]), (9, totales["renta_pesos"])):
            celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
            celda.font = fuente_total
            celda.number_format = '"$"#,##0'
        if totales["renta_pct"] is not None:
            celda = hoja.cell(row=fila_actual, column=10, value=round(float(totales["renta_pct"]) / 100, 4))
            celda.font = fuente_total
            celda.number_format = "0.0%"
        fila_actual += 2

    if no_calculables:
        hoja.cell(
            row=fila_actual, column=1,
            value=f"Quedaron AFUERA del cálculo {totales['no_calculables_casos']} artículos "
            f"({_formatear_numero(totales['no_calculables_bultos'])} bultos) — no sumaron como cero:",
        ).font = fuente_marca
        fila_actual += 1
        for columna, encabezado in enumerate(("Artículo", "Bultos", "Días", "Motivo"), start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1
        for entrada in no_calculables:
            hoja.cell(row=fila_actual, column=1, value=entrada["articulo_nombre"])
            hoja.cell(row=fila_actual, column=2, value=float(entrada["bultos"]))
            hoja.cell(row=fila_actual, column=3, value=entrada["dias"])
            hoja.cell(row=fila_actual, column=4, value=entrada["motivo_etiqueta"])
            fila_actual += 1

    for columna, ancho in enumerate((24, 9, 10, 9, 12, 12, 13, 13, 13, 10), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
