"""Genera Rentabilidad Real en PDF y Excel — puro, sin tocar la base.

Mismo criterio visual que el export de la teórica
(core/exportar_rentabilidad.py), con las columnas de la cuenta REAL
(venta = lo enviado, mercadería al costo FIFO, mermas del período) y —
SIEMPRE que exista — la sección "afuera del cálculo" POR MOTIVO, con sus
bultos y sus artículos: en el papel también es protagonista, no nota al
pie. grupos/totales/afuera_por_motivo: los que arma core.costo_real —
esto solo arma bytes en memoria, nunca guarda nada.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.exportar_rentabilidad import (
    ALTURA_ENCABEZADO,
    GRIS_FILA_ALTERNADA,
    GRIS_TEXTO_AYUDA,
    GRIS_TEXTO_AYUDA_HEX,
    OFFSET_LINEA,
    OFFSET_SUBTITULO,
    OFFSET_TITULO,
    ROJO_MARCA,
    ROJO_MARCA_HEX,
    VERDE_CLARO_ENCABEZADO_TABLA,
    VERDE_ENCABEZADO,
    VERDE_ENCABEZADO_HEX,
    _formatear_moneda,
    _formatear_numero,
    _formatear_pct,
)


def _armar_subtitulo_real(fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], cantidad_fechas: int) -> str:
    subtitulo = (
        f"Envíos del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} "
        f"({cantidad_fechas} día{'s' if cantidad_fechas != 1 else ''} con envíos) — "
        "la cuenta REAL: venta = lo ENVIADO × precio de lista vigente con las tasas del cliente; "
        "mercadería al costo FIFO del lote que salió; mermas del período a su costo; "
        "reproceso neutro; la segunda vale cero"
    )
    if filtros_texto:
        subtitulo += " — " + ", ".join(filtros_texto)
    return subtitulo


def _dibujar_encabezado_real(canvas, documento, subtitulo: str):
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Rentabilidad Real")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 8.5)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def _texto_articulos_afuera(resumen: dict) -> str:
    return " · ".join(f"{a['nombre']} ({_formatear_numero(a['bultos'])})" for a in resumen["articulos"])


def generar_pdf_rentabilidad_real(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], resultado: dict
) -> bytes:
    """Arma el PDF de Rentabilidad Real: el afuera por motivo primero, después una tabla por grupo y el total."""
    buffer = BytesIO()
    subtitulo = _armar_subtitulo_real(fecha_desde, fecha_hasta, filtros_texto, len(resultado["fechas_incluidas"]))
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado_real(canvas, doc, subtitulo)

    estilo_titulo_tabla = ParagraphStyle(
        "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
    )
    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=8.5, textColor=VERDE_ENCABEZADO
    )
    estilo_dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=8.5, textColor=colors.black)
    estilo_numero = ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.black)
    estilo_negativo = ParagraphStyle("negativo", fontName="Helvetica-Bold", fontSize=8.5, textColor=ROJO_MARCA)
    estilo_subtotal = ParagraphStyle("subtotal", fontName="Helvetica-Bold", fontSize=9, textColor=colors.black)
    estilo_total = ParagraphStyle("total", fontName="Helvetica-Bold", fontSize=13, textColor=colors.black)
    estilo_aviso = ParagraphStyle("aviso", fontName="Helvetica-Bold", fontSize=9.5, textColor=ROJO_MARCA)
    estilo_vacio = ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []
    grupos = resultado["grupos"]
    totales = resultado["totales"]
    afuera = resultado["afuera_por_motivo"]

    if not grupos and not afuera:
        elementos.append(Paragraph("Sin movimientos en este rango.", estilo_vacio))

    # El AFUERA primero, como en la pantalla: es la hoja de ruta.
    if afuera:
        elementos.append(
            Paragraph(
                f"AFUERA DEL CÁLCULO — no sumó como cero: {_formatear_numero(totales['afuera_bultos'])} bultos "
                f"por {totales['afuera_motivos']} motivo{'s' if totales['afuera_motivos'] != 1 else ''}:",
                estilo_aviso,
            )
        )
        elementos.append(Spacer(1, 6))
        datos_afuera = [
            [
                Paragraph("Motivo", estilo_encabezado_tabla),
                Paragraph("Bultos", estilo_encabezado_tabla),
                Paragraph("Artículos", estilo_encabezado_tabla),
            ]
        ]
        for resumen in afuera:
            datos_afuera.append(
                [
                    Paragraph(resumen["etiqueta"], estilo_dato),
                    Paragraph(_formatear_numero(resumen["bultos"]), estilo_numero),
                    Paragraph(_texto_articulos_afuera(resumen), estilo_dato),
                ]
            )
        tabla_afuera = Table(
            datos_afuera,
            colWidths=[ancho_util * 0.46, ancho_util * 0.1, ancho_util * 0.44],
            repeatRows=1,
        )
        tabla_afuera.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elementos.append(tabla_afuera)
        elementos.append(Spacer(1, 14))

    encabezados = ("Artículo", "Bultos env.", "Venta real", "Mercadería", "Envase", "Mermas", "Renta $", "Util. %")
    anchos = [
        ancho_util * 0.22, ancho_util * 0.1, ancho_util * 0.13, ancho_util * 0.13,
        ancho_util * 0.11, ancho_util * 0.11, ancho_util * 0.12, ancho_util * 0.08,
    ]
    for indice_grupo, grupo in enumerate(grupos):
        if indice_grupo > 0:
            elementos.append(Spacer(1, 14))

        datos_tabla = [
            [Paragraph(grupo["etiqueta"], estilo_titulo_tabla)] + [""] * (len(encabezados) - 1),
            [Paragraph(encabezado, estilo_encabezado_tabla) for encabezado in encabezados],
        ]
        estilos_filas = []
        for indice, fila in enumerate(grupo["filas"]):
            estilo_renta = estilo_negativo if fila["renta_pesos"] < 0 else estilo_numero
            datos_tabla.append(
                [
                    Paragraph(fila["articulo_nombre"], estilo_dato),
                    Paragraph(_formatear_numero(fila["bultos"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["venta_neta"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["costo_mercaderia"]), estilo_dato),
                    Paragraph(_formatear_moneda(fila["costo_envase"]), estilo_dato),
                    Paragraph(
                        f"{_formatear_moneda(fila['costo_mermas'])} ({_formatear_numero(fila['bultos_mermados'])}b)"
                        if fila["bultos_mermados"] else "—",
                        estilo_dato,
                    ),
                    Paragraph(_formatear_moneda(fila["renta_pesos"]), estilo_renta),
                    Paragraph(_formatear_pct(fila["utilidad_pct"]), estilo_renta),
                ]
            )
            if indice % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice + 2), (-1, indice + 2), GRIS_FILA_ALTERNADA))

        subtotal = grupo["subtotal"]
        indice_subtotal = len(datos_tabla)
        datos_tabla.append(
            [
                Paragraph("Subtotal", estilo_subtotal),
                Paragraph(_formatear_numero(subtotal["bultos"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["venta_neta"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["costo_mercaderia"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["costo_envase"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["costo_mermas"]), estilo_subtotal),
                Paragraph(_formatear_moneda(subtotal["renta_pesos"]), estilo_subtotal),
                Paragraph(_formatear_pct(subtotal["utilidad_pct"]), estilo_subtotal),
            ]
        )

        tabla = Table(datos_tabla, colWidths=anchos, repeatRows=2)
        tabla.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 2), (-1, -2), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("LINEABOVE", (0, indice_subtotal), (-1, indice_subtotal), 1, VERDE_ENCABEZADO),
                    ("TOPPADDING", (0, 0), (-1, 0), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

    if grupos:
        elementos.append(Spacer(1, 16))
        segunda = (
            f" Segunda generada: {_formatear_numero(totales['segunda_bultos'])} bultos (vale cero)."
            if totales["segunda_bultos"] else ""
        )
        devoluciones = (
            f" − devoluciones {_formatear_moneda(totales['devoluciones_venta'])}"
            f" ({_formatear_numero(totales['devoluciones_bultos'])} bultos que volvieron)"
            if totales.get("devoluciones_bultos") else ""
        )
        elementos.append(
            Paragraph(
                f"Total REAL: venta {_formatear_moneda(totales['venta_neta'])}{devoluciones} — mercadería {_formatear_moneda(totales['costo_mercaderia'])} "
                f"— envase {_formatear_moneda(totales['costo_envase'])} — mermas {_formatear_moneda(totales['costo_mermas'])} "
                + (
                    f"— rechazos perdidos {_formatear_moneda(totales['rechazos_perdidos'])} "
                    f"({_formatear_numero(totales['rechazos_bultos'])} bultos a segunda) "
                    if totales.get("rechazos_bultos") else ""
                )
                + "— "
                f"renta {_formatear_moneda(totales['renta_pesos'])} (utilidad {_formatear_pct(totales['utilidad_pct'])} sobre mercadería)."
                + segunda,
                estilo_total,
            )
        )

    documento.build(elementos, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_rentabilidad_real(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], resultado: dict
) -> bytes:
    """Arma el Excel de Rentabilidad Real: el afuera por motivo primero, secciones por grupo y el total."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Rentabilidad Real"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_grupo = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_marca = Font(bold=True, color=ROJO_MARCA_HEX, size=9)
    fuente_subtotal = Font(bold=True)
    fuente_total = Font(bold=True, size=13)

    grupos = resultado["grupos"]
    totales = resultado["totales"]
    afuera = resultado["afuera_por_motivo"]

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Rentabilidad Real")
    for columna in range(1, 12):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(
        row=fila_actual, column=1,
        value=_armar_subtitulo_real(fecha_desde, fecha_hasta, filtros_texto, len(resultado["fechas_incluidas"])),
    ).font = fuente_normal
    fila_actual += 2

    if not grupos and not afuera:
        hoja.cell(row=fila_actual, column=1, value="Sin movimientos en este rango.").font = fuente_normal

    if afuera:
        hoja.cell(
            row=fila_actual, column=1,
            value=f"AFUERA DEL CÁLCULO — no sumó como cero: {_formatear_numero(totales['afuera_bultos'])} bultos "
            f"por {totales['afuera_motivos']} motivo{'s' if totales['afuera_motivos'] != 1 else ''}:",
        ).font = fuente_marca
        fila_actual += 1
        for columna, encabezado in enumerate(("Motivo", "Bultos", "Artículos"), start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1
        for resumen in afuera:
            hoja.cell(row=fila_actual, column=1, value=resumen["etiqueta"])
            hoja.cell(row=fila_actual, column=2, value=float(resumen["bultos"]))
            hoja.cell(row=fila_actual, column=3, value=_texto_articulos_afuera(resumen))
            fila_actual += 1
        fila_actual += 1

    encabezados = (
        "Artículo", "Bultos enviados", "Unidades enviadas", "Venta real", "Mercadería (FIFO)",
        "Envase", "Mermas $", "Mermas bultos", "Segunda bultos", "Devol. bultos", "Devoluciones $",
        "Rechazos perdidos $", "Rechazos perdidos bultos", "Renta $", "Utilidad %",
    )
    for grupo in grupos:
        hoja.cell(row=fila_actual, column=1, value=grupo["etiqueta"]).font = fuente_grupo
        fila_actual += 1

        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for fila in grupo["filas"]:
            hoja.cell(row=fila_actual, column=1, value=fila["articulo_nombre"])
            hoja.cell(row=fila_actual, column=2, value=float(fila["bultos"]))
            hoja.cell(row=fila_actual, column=3, value=round(float(fila["unidades"]), 2))
            for columna, valor in ((4, fila["venta_neta"]), (5, fila["costo_mercaderia"]),
                                   (6, fila["costo_envase"]), (7, fila["costo_mermas"]),
                                   (11, fila["devoluciones_venta"]), (12, fila["rechazos_perdidos"]),
                                   (14, fila["renta_pesos"])):
                celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
                celda.number_format = '"$"#,##0'
            hoja.cell(row=fila_actual, column=8, value=float(fila["bultos_mermados"]))
            hoja.cell(row=fila_actual, column=9, value=float(fila["segunda_bultos"]))
            hoja.cell(row=fila_actual, column=10, value=float(fila["devoluciones_bultos"]))
            hoja.cell(row=fila_actual, column=13, value=float(fila["rechazos_bultos"]))
            if fila["utilidad_pct"] is not None:
                celda = hoja.cell(row=fila_actual, column=15, value=round(float(fila["utilidad_pct"]) / 100, 4))
                celda.number_format = "0.0%"
            if fila["renta_pesos"] < 0:
                hoja.cell(row=fila_actual, column=14).font = fuente_marca
            fila_actual += 1

        subtotal = grupo["subtotal"]
        hoja.cell(row=fila_actual, column=1, value="Subtotal").font = fuente_subtotal
        celda = hoja.cell(row=fila_actual, column=2, value=float(subtotal["bultos"]))
        celda.font = fuente_subtotal
        for columna, valor in ((4, subtotal["venta_neta"]), (5, subtotal["costo_mercaderia"]),
                               (6, subtotal["costo_envase"]), (7, subtotal["costo_mermas"]),
                               (11, subtotal["devoluciones_venta"]), (12, subtotal["rechazos_perdidos"]),
                               (14, subtotal["renta_pesos"])):
            celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
            celda.font = fuente_subtotal
            celda.number_format = '"$"#,##0'
        hoja.cell(row=fila_actual, column=13, value=float(subtotal["rechazos_bultos"])).font = fuente_subtotal
        if subtotal["utilidad_pct"] is not None:
            celda = hoja.cell(row=fila_actual, column=15, value=round(float(subtotal["utilidad_pct"]) / 100, 4))
            celda.font = fuente_subtotal
            celda.number_format = "0.0%"
        fila_actual += 2

    if grupos:
        hoja.cell(row=fila_actual, column=1, value="Total REAL").font = fuente_total
        for columna, valor in ((4, totales["venta_neta"]), (5, totales["costo_mercaderia"]),
                               (6, totales["costo_envase"]), (7, totales["costo_mermas"]),
                               (11, totales["devoluciones_venta"]), (12, totales["rechazos_perdidos"]),
                               (14, totales["renta_pesos"])):
            celda = hoja.cell(row=fila_actual, column=columna, value=round(float(valor), 2))
            celda.font = fuente_total
            celda.number_format = '"$"#,##0'
        hoja.cell(row=fila_actual, column=9, value=float(totales["segunda_bultos"])).font = fuente_total
        hoja.cell(row=fila_actual, column=10, value=float(totales["devoluciones_bultos"])).font = fuente_total
        hoja.cell(row=fila_actual, column=13, value=float(totales["rechazos_bultos"])).font = fuente_total
        if totales["utilidad_pct"] is not None:
            celda = hoja.cell(row=fila_actual, column=15, value=round(float(totales["utilidad_pct"]) / 100, 4))
            celda.font = fuente_total
            celda.number_format = "0.0%"

    for columna, ancho in enumerate((26, 13, 14, 13, 15, 11, 11, 12, 13, 13, 14, 17, 20, 13, 11), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
