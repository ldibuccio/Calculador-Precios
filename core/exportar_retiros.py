"""Genera el listado de Consultar Retiros en PDF y Excel — puro, sin tocar la base.

Mismo criterio que core/exportar_compras.py y core/exportar_vacios.py:
banda de encabezado repetida en cada página, tabla con encabezados que se
repiten si se corta (repeatRows), y los datos ya armados llegan de quien
llama (app/main.py) — esto solo arma bytes en memoria, nunca guarda nada.

El total de bultos va al FINAL con el mismo desglose que la pantalla:
total retirado, cuánto es anotado al retirar vs. tomado de la carga del
comprador, y — si hay — cuántos bultos no ingresaron al depósito con el
neto. Es el número con el que se liquida al carrero/cooperativa: los dos
números a la vista, nunca una resta en silencio.

filas: las mismas que arma /logistica/consultar (cada una con bultos,
usa_anotada y no_ingreso ya calculados). totales: dict con total_bultos,
total_anotados, total_del_comprador, total_no_ingresados y total_neto.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
ROJO_NO_INGRESO = colors.Color(0.6, 0.11, 0.11)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"
ROJO_NO_INGRESO_HEX = "991B1B"

ETIQUETAS_ESTADO_RETIRO = {"retirado": "Retirado", "cancelado": "Cancelado"}

OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm


def _formatear_numero(valor) -> str:
    return f"{float(valor):.2f}".rstrip("0").rstrip(".")


def _etiqueta_estado_retiro(fila: dict) -> str:
    return ETIQUETAS_ESTADO_RETIRO.get(fila.get("estado_retiro"), "Pendiente")


def _texto_bultos(fila: dict) -> str:
    """El número de bultos de la fila, con el mismo asterisco que la pantalla cuando viene de la carga."""
    return f"{_formatear_numero(fila['bultos'])}{'' if fila['usa_anotada'] else '*'}"


def _dibujar_encabezado(canvas, documento, subtitulo: str):
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, "Consultar Retiros")

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 10)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def _armar_subtitulo(fecha_desde: date, fecha_hasta: date, filtros_texto: list[str]) -> str:
    """"Del ... al ..." más los filtros aplicados (proveedor, artículo, tipo, estado), para que el archivo diga qué es."""
    subtitulo = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    if filtros_texto:
        subtitulo += " — " + ", ".join(filtros_texto)
    return subtitulo


def generar_pdf_listado_retiros(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], filas: list[dict], totales: dict
) -> bytes:
    """Arma el PDF de Consultar Retiros: la tabla del listado + el total desglosado al final."""
    buffer = BytesIO()
    subtitulo = _armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto)
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    ancho_util = documento.width

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado(canvas, doc, subtitulo)

    estilo_encabezado_tabla = ParagraphStyle(
        "encabezado_tabla", fontName="Helvetica-Bold", fontSize=9, textColor=VERDE_ENCABEZADO
    )
    estilo_dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=8.5, textColor=colors.black)
    estilo_dato_gris = ParagraphStyle("dato_gris", fontName="Helvetica", fontSize=8.5, textColor=GRIS_TEXTO_AYUDA)
    estilo_numero = ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.black)
    estilo_total = ParagraphStyle("total", fontName="Helvetica-Bold", fontSize=12, textColor=colors.black)
    estilo_desglose = ParagraphStyle("desglose", fontName="Helvetica", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)
    estilo_no_ingreso = ParagraphStyle(
        "no_ingreso", fontName="Helvetica-Bold", fontSize=10, textColor=ROJO_NO_INGRESO
    )
    estilo_vacio = ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA)

    elementos = []
    if not filas:
        elementos.append(Paragraph("No se encontraron retiros con estos filtros.", estilo_vacio))
    else:
        datos_tabla = [
            [
                Paragraph(encabezado, estilo_encabezado_tabla)
                for encabezado in ("Fecha", "Hora ret.", "Proveedor", "Artículo", "Bultos", "Tipo", "Estado")
            ]
        ]
        estilos_filas = []
        for indice, fila in enumerate(filas):
            estado = _etiqueta_estado_retiro(fila)
            if fila["no_ingreso"]:
                estado += "<br/><font color='#991B1B' size='7.5'><b>No ingresó al depósito</b></font>"
            datos_tabla.append(
                [
                    Paragraph(fila["fecha_operacion"].strftime("%d/%m"), estilo_dato_gris),
                    Paragraph(
                        fila["retiro_procesado_el"].strftime("%H:%M") if fila["retiro_procesado_el"] else "—",
                        estilo_dato_gris,
                    ),
                    Paragraph(fila["proveedor_nombre"], estilo_dato),
                    Paragraph(fila["articulo_nombre"], estilo_dato),
                    Paragraph(_texto_bultos(fila), estilo_numero),
                    Paragraph(fila["tipo_retiro"] or "—", estilo_dato),
                    Paragraph(estado, estilo_dato),
                ]
            )
            if indice % 2 == 1:
                estilos_filas.append(("BACKGROUND", (0, indice + 1), (-1, indice + 1), GRIS_FILA_ALTERNADA))

        tabla = Table(
            datos_tabla,
            colWidths=[
                ancho_util * 0.08,
                ancho_util * 0.09,
                ancho_util * 0.21,
                ancho_util * 0.2,
                ancho_util * 0.09,
                ancho_util * 0.12,
                ancho_util * 0.21,
            ],
            repeatRows=1,
        )
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), VERDE_CLARO_ENCABEZADO_TABLA),
                    ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    *estilos_filas,
                ]
            )
        )
        elementos.append(tabla)

        elementos.append(Spacer(1, 14))
        elementos.append(Paragraph(f"Total: {_formatear_numero(totales['total_bultos'])} bultos", estilo_total))
        elementos.append(Spacer(1, 3))
        elementos.append(
            Paragraph(
                f"{_formatear_numero(totales['total_anotados'])} anotados al retirar + "
                f"{_formatear_numero(totales['total_del_comprador'])}* de la carga del comprador "
                "(* sin cantidad anotada al retirar: se usa lo que cargó el comprador)",
                estilo_desglose,
            )
        )
        if totales["total_no_ingresados"] > 0:
            elementos.append(Spacer(1, 5))
            elementos.append(
                Paragraph(
                    f"{_formatear_numero(totales['total_no_ingresados'])} no ingresaron al depósito — "
                    f"Neto: {_formatear_numero(totales['total_neto'])} bultos",
                    estilo_no_ingreso,
                )
            )

    documento.build(elementos, onFirstPage=_encabezado_pagina, onLaterPages=_encabezado_pagina)
    return buffer.getvalue()


def generar_excel_listado_retiros(
    fecha_desde: date, fecha_hasta: date, filtros_texto: list[str], filas: list[dict], totales: dict
) -> bytes:
    """Arma el Excel de Consultar Retiros, mismas columnas que el PDF más el origen del número explícito."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Consultar Retiros"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_no_ingreso = Font(bold=True, color=ROJO_NO_INGRESO_HEX, size=9)
    fuente_total = Font(bold=True, size=12)
    fuente_neto = Font(bold=True, color=ROJO_NO_INGRESO_HEX)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Consultar Retiros")
    for columna in range(1, 9):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(row=fila_actual, column=1, value=_armar_subtitulo(fecha_desde, fecha_hasta, filtros_texto)).font = fuente_normal
    fila_actual += 2

    if not filas:
        hoja.cell(row=fila_actual, column=1, value="No se encontraron retiros con estos filtros.").font = fuente_normal
    else:
        encabezados = ("Fecha", "Hora retiro", "Proveedor", "Artículo", "Bultos", "Origen del número", "Tipo", "Estado")
        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for fila in filas:
            hoja.cell(row=fila_actual, column=1, value=fila["fecha_operacion"].strftime("%d/%m/%Y"))
            hoja.cell(
                row=fila_actual, column=2,
                value=fila["retiro_procesado_el"].strftime("%H:%M") if fila["retiro_procesado_el"] else "—",
            )
            hoja.cell(row=fila_actual, column=3, value=fila["proveedor_nombre"])
            hoja.cell(row=fila_actual, column=4, value=fila["articulo_nombre"])
            hoja.cell(row=fila_actual, column=5, value=float(fila["bultos"]))
            hoja.cell(
                row=fila_actual, column=6,
                value="anotado al retirar" if fila["usa_anotada"] else "carga del comprador",
            ).font = fuente_normal
            hoja.cell(row=fila_actual, column=7, value=fila["tipo_retiro"] or "—")
            celda_estado = hoja.cell(
                row=fila_actual, column=8,
                value=_etiqueta_estado_retiro(fila) + (" — No ingresó al depósito" if fila["no_ingreso"] else ""),
            )
            if fila["no_ingreso"]:
                celda_estado.font = fuente_no_ingreso
            fila_actual += 1

        fila_actual += 1
        hoja.cell(row=fila_actual, column=1, value="Total").font = fuente_total
        hoja.cell(row=fila_actual, column=5, value=float(totales["total_bultos"])).font = fuente_total
        fila_actual += 1
        hoja.cell(
            row=fila_actual, column=1,
            value=f"{_formatear_numero(totales['total_anotados'])} anotados al retirar + "
            f"{_formatear_numero(totales['total_del_comprador'])} de la carga del comprador",
        ).font = fuente_normal
        fila_actual += 1
        if totales["total_no_ingresados"] > 0:
            hoja.cell(row=fila_actual, column=1, value="No ingresaron al depósito").font = fuente_neto
            hoja.cell(row=fila_actual, column=5, value=float(totales["total_no_ingresados"])).font = fuente_neto
            fila_actual += 1
            hoja.cell(row=fila_actual, column=1, value="Neto").font = fuente_neto
            hoja.cell(row=fila_actual, column=5, value=float(totales["total_neto"])).font = fuente_neto

    for columna, ancho in enumerate((12, 11, 24, 22, 9, 19, 13, 30), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
