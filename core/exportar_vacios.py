"""Genera los exports de Vacíos (Movimientos y Stock del Sistema) en PDF y Excel — puro, sin tocar la base.

Mismo criterio que core/exportar_compras.py: banda de encabezado repetida en
cada página (dibujada en el canvas), tablas con su título embebido vía
repeatRows para que se repita solo si se cortan, y los datos ya armados
llegan de quien llama (app/main.py) — esto solo arma bytes en memoria,
nunca guarda nada.

Movimientos: las mismas tres secciones que la pantalla, en el mismo orden
(Salidas, Entradas, Ajustes) y con los anulados marcados, no escondidos —
el export es una foto del listado, y un movimiento anulado a la vista es
parte del control. Los ajustes llevan su motivo.

Stock: agrupado por proveedor con el detalle de las sumas y el total.
Para una fecha pasada se agrega la misma advertencia que la pantalla: ese
número puede cambiar si después se anula un movimiento anterior.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

VERDE_ENCABEZADO = colors.Color(0.18, 0.55, 0.34)
VERDE_CLARO_ENCABEZADO_TABLA = colors.Color(0.87, 0.94, 0.89)
GRIS_TEXTO_AYUDA = colors.Color(0.35, 0.35, 0.35)
GRIS_FILA_ALTERNADA = colors.Color(0.97, 0.97, 0.97)
ROJO_ANULADO = colors.Color(0.86, 0.15, 0.15)
VERDE_ENCABEZADO_HEX = "2E8C57"
GRIS_TEXTO_AYUDA_HEX = "595959"
ROJO_ANULADO_HEX = "DC2626"

AVISO_FECHA_PASADA = "El stock de fechas pasadas puede cambiar si se anulan movimientos anteriores."

OFFSET_TITULO = 15 * mm
OFFSET_SUBTITULO = 22 * mm
OFFSET_LINEA = 27 * mm
ALTURA_ENCABEZADO = 34 * mm


def _formatear_fecha_hora(valor) -> str:
    return valor.strftime("%d/%m/%Y %H:%M") if valor is not None else "—"


def _texto_anulado(fila: dict) -> str:
    return f"ANULADO el {_formatear_fecha_hora(fila['anulado_el'])}" if fila.get("anulado_el") else ""


def _texto_cantidad_ajuste(cantidad) -> str:
    cantidad = int(cantidad)
    return f"+{cantidad}" if cantidad > 0 else str(cantidad)


def _dibujar_encabezado(canvas, documento, titulo: str, subtitulo: str):
    """Banda de encabezado en TODAS las páginas, directo en el canvas (mismo criterio que exportar_compras)."""
    ancho_pagina, alto_pagina = A4
    x = documento.leftMargin
    x_derecha = ancho_pagina - documento.rightMargin
    canvas.saveState()

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(x, alto_pagina - OFFSET_TITULO, titulo)

    canvas.setFillColor(GRIS_TEXTO_AYUDA)
    canvas.setFont("Helvetica", 10)
    canvas.drawString(x, alto_pagina - OFFSET_SUBTITULO, subtitulo)

    canvas.setStrokeColor(VERDE_ENCABEZADO)
    canvas.setLineWidth(1)
    canvas.line(x, alto_pagina - OFFSET_LINEA, x_derecha, alto_pagina - OFFSET_LINEA)

    canvas.restoreState()


def _documento_pdf(buffer: BytesIO, titulo: str, subtitulo: str):
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=ALTURA_ENCABEZADO,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    def _encabezado_pagina(canvas, doc):
        _dibujar_encabezado(canvas, doc, titulo, subtitulo)

    return documento, _encabezado_pagina


def _estilos_pdf():
    return {
        "titulo_tabla": ParagraphStyle(
            "titulo_tabla", fontName="Helvetica-Bold", fontSize=11.5, textColor=VERDE_ENCABEZADO
        ),
        "encabezado_tabla": ParagraphStyle(
            "encabezado_tabla", fontName="Helvetica-Bold", fontSize=9.5, textColor=VERDE_ENCABEZADO
        ),
        "dato": ParagraphStyle("dato", fontName="Helvetica", fontSize=9.5, textColor=colors.black),
        "dato_gris": ParagraphStyle("dato_gris", fontName="Helvetica", fontSize=9, textColor=GRIS_TEXTO_AYUDA),
        "dato_anulado": ParagraphStyle("dato_anulado", fontName="Helvetica", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA),
        "marca_anulado": ParagraphStyle("marca_anulado", fontName="Helvetica-Bold", fontSize=8.5, textColor=ROJO_ANULADO),
        "numero": ParagraphStyle("numero", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.black),
        "aviso": ParagraphStyle("aviso", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA),
        "vacio": ParagraphStyle("vacio", fontName="Helvetica-Oblique", fontSize=9.5, textColor=GRIS_TEXTO_AYUDA),
    }


def _tabla_seccion_pdf(titulo: str, encabezados: list[str], filas_tabla: list[list], anchos: list[float], estilos) -> Table:
    """Una sección del PDF de Movimientos: título + encabezados embebidos (repeatRows=2) + filas ya armadas."""
    datos = [
        [Paragraph(titulo, estilos["titulo_tabla"])] + [""] * (len(encabezados) - 1),
        [Paragraph(encabezado, estilos["encabezado_tabla"]) for encabezado in encabezados],
    ]
    datos.extend(filas_tabla)

    estilos_filas = []
    for indice in range(len(filas_tabla)):
        if indice % 2 == 1:
            estilos_filas.append(("BACKGROUND", (0, indice + 2), (-1, indice + 2), GRIS_FILA_ALTERNADA))

    tabla = Table(datos, colWidths=anchos, repeatRows=2)
    tabla.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, 1), (-1, 1), VERDE_CLARO_ENCABEZADO_TABLA),
                ("LINEBELOW", (0, 2), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                ("TOPPADDING", (0, 0), (-1, 0), 0),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                *estilos_filas,
            ]
        )
    )
    return tabla


def _parrafo_movimiento(texto: str, fila: dict, estilos) -> Paragraph:
    """El detalle de un movimiento; los anulados van en gris con su marca en rojo debajo, nunca escondidos."""
    estilo = estilos["dato_anulado"] if fila.get("anulado_el") else estilos["dato"]
    marca = _texto_anulado(fila)
    if marca:
        return Paragraph(f"{texto}<br/><font color='#DC2626' size='8.5'><b>{marca}</b></font>", estilo)
    return Paragraph(texto, estilo)


def generar_pdf_movimientos_vacios(
    fecha_desde: date, fecha_hasta: date, devueltos: list[dict], recibidos: list[dict], ajustes: list[dict]
) -> bytes:
    """PDF de Movimientos de Vacíos: Salidas, Entradas y Ajustes (con motivo), en el mismo orden que la pantalla."""
    buffer = BytesIO()
    subtitulo = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    documento, encabezado_pagina = _documento_pdf(buffer, "Movimientos de Vacíos", subtitulo)
    ancho_util = documento.width
    estilos = _estilos_pdf()

    elementos = []

    filas_salidas = [
        [
            Paragraph(_formatear_fecha_hora(d["creado_en"]), estilos["dato_gris"]),
            _parrafo_movimiento(f"{d['cantidad']} × {d['tipo_nombre']} a {d['proveedor_nombre']}", d, estilos),
            Paragraph(str(d["cantidad"]), estilos["numero"]),
        ]
        for d in devueltos
    ]
    filas_entradas = [
        [
            Paragraph(_formatear_fecha_hora(r["creado_en"]), estilos["dato_gris"]),
            _parrafo_movimiento(
                f"{r['cantidad']} × {r['tipo_nombre']} de {r['proveedor_nombre']} — trajo {r['cliente_nombre']}",
                r,
                estilos,
            ),
            Paragraph(str(r["cantidad"]), estilos["numero"]),
        ]
        for r in recibidos
    ]
    filas_ajustes = [
        [
            Paragraph(_formatear_fecha_hora(a["creado_en"]), estilos["dato_gris"]),
            _parrafo_movimiento(f"{a['tipo_nombre']} de {a['proveedor_nombre']}", a, estilos),
            Paragraph(a["motivo"], estilos["dato_gris"]),
            Paragraph(_texto_cantidad_ajuste(a["cantidad"]), estilos["numero"]),
        ]
        for a in ajustes
    ]

    anchos_movimiento = [ancho_util * 0.2, ancho_util * 0.68, ancho_util * 0.12]
    anchos_ajuste = [ancho_util * 0.2, ancho_util * 0.35, ancho_util * 0.33, ancho_util * 0.12]

    secciones = [
        ("Salidas (devueltos al proveedor)", ["Fecha y hora", "Salida", "Cajones"], filas_salidas, anchos_movimiento),
        ("Entradas (recibidos de clientes)", ["Fecha y hora", "Entrada", "Cajones"], filas_entradas, anchos_movimiento),
        ("Ajustes (correcciones de stock)", ["Fecha y hora", "Ajuste", "Motivo", "Cantidad"], filas_ajustes, anchos_ajuste),
    ]
    for indice, (titulo, encabezados, filas_tabla, anchos) in enumerate(secciones):
        if indice > 0:
            elementos.append(Spacer(1, 14))
        if filas_tabla:
            elementos.append(_tabla_seccion_pdf(titulo, encabezados, filas_tabla, anchos, estilos))
        else:
            elementos.append(Paragraph(titulo, estilos["titulo_tabla"]))
            elementos.append(Spacer(1, 4))
            elementos.append(Paragraph("Sin movimientos en este rango.", estilos["vacio"]))

    documento.build(elementos, onFirstPage=encabezado_pagina, onLaterPages=encabezado_pagina)
    return buffer.getvalue()


def generar_excel_movimientos_vacios(
    fecha_desde: date, fecha_hasta: date, devueltos: list[dict], recibidos: list[dict], ajustes: list[dict]
) -> bytes:
    """Excel de Movimientos de Vacíos, mismas tres secciones que el PDF; los anulados con su propia columna."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Movimientos de Vacíos"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_seccion = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_anulado = Font(bold=True, color=ROJO_ANULADO_HEX, size=9)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Movimientos de Vacíos")
    for columna in range(1, 6):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    subtitulo = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    hoja.cell(row=fila_actual, column=1, value=subtitulo).font = fuente_normal
    fila_actual += 2

    def _seccion(titulo: str, encabezados: list[str], filas_datos: list[list], nonlocal_fila: int) -> int:
        hoja.cell(row=nonlocal_fila, column=1, value=titulo).font = fuente_seccion
        nonlocal_fila += 1
        if not filas_datos:
            hoja.cell(row=nonlocal_fila, column=1, value="Sin movimientos en este rango.").font = fuente_normal
            return nonlocal_fila + 2
        for columna, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=nonlocal_fila, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        nonlocal_fila += 1
        for fila_datos in filas_datos:
            for columna, valor in enumerate(fila_datos, start=1):
                celda = hoja.cell(row=nonlocal_fila, column=columna, value=valor)
                if columna == len(fila_datos) and isinstance(valor, str) and valor.startswith("Anulado"):
                    celda.font = fuente_anulado
            nonlocal_fila += 1
        return nonlocal_fila + 1

    fila_actual = _seccion(
        "Salidas (devueltos al proveedor)",
        ["Fecha y hora", "Tipo", "Proveedor", "Cajones", "Estado"],
        [
            [
                _formatear_fecha_hora(d["creado_en"]),
                d["tipo_nombre"],
                d["proveedor_nombre"],
                int(d["cantidad"]),
                f"Anulado el {_formatear_fecha_hora(d['anulado_el'])}" if d.get("anulado_el") else "",
            ]
            for d in devueltos
        ],
        fila_actual,
    )
    fila_actual = _seccion(
        "Entradas (recibidos de clientes)",
        ["Fecha y hora", "Tipo", "Proveedor", "Trajo", "Cajones", "Estado"],
        [
            [
                _formatear_fecha_hora(r["creado_en"]),
                r["tipo_nombre"],
                r["proveedor_nombre"],
                r["cliente_nombre"],
                int(r["cantidad"]),
                f"Anulado el {_formatear_fecha_hora(r['anulado_el'])}" if r.get("anulado_el") else "",
            ]
            for r in recibidos
        ],
        fila_actual,
    )
    fila_actual = _seccion(
        "Ajustes (correcciones de stock)",
        ["Fecha y hora", "Tipo", "Proveedor", "Motivo", "Cantidad", "Estado"],
        [
            [
                _formatear_fecha_hora(a["creado_en"]),
                a["tipo_nombre"],
                a["proveedor_nombre"],
                a["motivo"],
                int(a["cantidad"]),
                f"Anulado el {_formatear_fecha_hora(a['anulado_el'])}" if a.get("anulado_el") else "",
            ]
            for a in ajustes
        ],
        fila_actual,
    )

    for columna, ancho in enumerate((17, 20, 20, 24, 10, 22), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def generar_pdf_stock_vacios(fecha_consulta: date, es_pasada: bool, grupos: list[dict]) -> bytes:
    """PDF del Stock del Sistema a una fecha: una tabla por proveedor (tipo, recibidos, devueltos, ajustes, stock).

    grupos: los mismos que arma la pantalla (proveedor_nombre, tipos con
    recibidos/devueltos/ajustes/stock, total). Para una fecha pasada se
    agrega la advertencia de que ese número puede cambiar si se anulan
    movimientos anteriores.
    """
    buffer = BytesIO()
    subtitulo = f"Al {fecha_consulta.strftime('%d/%m/%Y')}"
    documento, encabezado_pagina = _documento_pdf(buffer, "Stock del Sistema — Vacíos", subtitulo)
    ancho_util = documento.width
    estilos = _estilos_pdf()

    elementos = []
    if es_pasada:
        elementos.append(Paragraph(AVISO_FECHA_PASADA, estilos["aviso"]))
        elementos.append(Spacer(1, 8))

    if not grupos:
        elementos.append(Paragraph("Sin movimientos hasta esta fecha.", estilos["vacio"]))

    encabezados = ["Tipo", "Recibidos", "Devueltos", "Ajustes", "Stock"]
    anchos = [ancho_util * 0.4, ancho_util * 0.15, ancho_util * 0.15, ancho_util * 0.15, ancho_util * 0.15]
    for indice, grupo in enumerate(grupos):
        if indice > 0:
            elementos.append(Spacer(1, 14))
        filas_tabla = [
            [
                Paragraph(t["tipo_nombre"], estilos["dato"]),
                Paragraph(str(int(t["recibidos"])), estilos["dato_gris"]),
                Paragraph(str(int(t["devueltos"])), estilos["dato_gris"]),
                Paragraph(_texto_cantidad_ajuste(t["ajustes"]) if int(t["ajustes"]) != 0 else "0", estilos["dato_gris"]),
                Paragraph(
                    f"<font color='#DC2626'>{t['stock']}</font>" if t["stock"] < 0 else str(t["stock"]),
                    estilos["numero"],
                ),
            ]
            for t in grupo["tipos"]
        ]
        if len(grupo["tipos"]) > 1:
            filas_tabla.append(
                [
                    Paragraph("Total", estilos["numero"]),
                    Paragraph("", estilos["dato"]),
                    Paragraph("", estilos["dato"]),
                    Paragraph("", estilos["dato"]),
                    Paragraph(
                        f"<font color='#DC2626'>{grupo['total']}</font>" if grupo["total"] < 0 else str(grupo["total"]),
                        estilos["numero"],
                    ),
                ]
            )
        elementos.append(_tabla_seccion_pdf(grupo["proveedor_nombre"], encabezados, filas_tabla, anchos, estilos))

    documento.build(elementos, onFirstPage=encabezado_pagina, onLaterPages=encabezado_pagina)
    return buffer.getvalue()


def generar_excel_stock_vacios(fecha_consulta: date, es_pasada: bool, grupos: list[dict]) -> bytes:
    """Excel del Stock del Sistema a una fecha, mismos grupos y columnas que el PDF."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Stock del Sistema"

    relleno_verde = PatternFill(start_color=VERDE_ENCABEZADO_HEX, end_color=VERDE_ENCABEZADO_HEX, fill_type="solid")
    relleno_verde_claro = PatternFill(start_color="DEEFE3", end_color="DEEFE3", fill_type="solid")
    fuente_blanca_titulo = Font(color="FFFFFF", bold=True, size=16)
    fuente_normal = Font(size=10, color=GRIS_TEXTO_AYUDA_HEX)
    fuente_proveedor = Font(bold=True, size=12, color=VERDE_ENCABEZADO_HEX)
    fuente_encabezado_tabla = Font(bold=True, color=VERDE_ENCABEZADO_HEX)
    fuente_stock = Font(bold=True)
    fuente_negativo = Font(bold=True, color=ROJO_ANULADO_HEX)

    fila_actual = 1
    hoja.cell(row=fila_actual, column=1, value="Stock del Sistema — Vacíos")
    for columna in range(1, 6):
        celda = hoja.cell(row=fila_actual, column=columna)
        celda.fill = relleno_verde
        if columna == 1:
            celda.font = fuente_blanca_titulo
    fila_actual += 1

    hoja.cell(row=fila_actual, column=1, value=f"Al {fecha_consulta.strftime('%d/%m/%Y')}").font = fuente_normal
    fila_actual += 1
    if es_pasada:
        hoja.cell(row=fila_actual, column=1, value=AVISO_FECHA_PASADA).font = fuente_normal
        fila_actual += 1
    fila_actual += 1

    if not grupos:
        hoja.cell(row=fila_actual, column=1, value="Sin movimientos hasta esta fecha.").font = fuente_normal

    for grupo in grupos:
        hoja.cell(row=fila_actual, column=1, value=grupo["proveedor_nombre"]).font = fuente_proveedor
        fila_actual += 1

        for columna, encabezado in enumerate(("Tipo", "Recibidos", "Devueltos", "Ajustes", "Stock"), start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = fuente_encabezado_tabla
            celda.fill = relleno_verde_claro
        fila_actual += 1

        for t in grupo["tipos"]:
            hoja.cell(row=fila_actual, column=1, value=t["tipo_nombre"])
            hoja.cell(row=fila_actual, column=2, value=int(t["recibidos"]))
            hoja.cell(row=fila_actual, column=3, value=int(t["devueltos"]))
            hoja.cell(row=fila_actual, column=4, value=int(t["ajustes"]))
            celda_stock = hoja.cell(row=fila_actual, column=5, value=int(t["stock"]))
            celda_stock.font = fuente_negativo if t["stock"] < 0 else fuente_stock
            fila_actual += 1

        if len(grupo["tipos"]) > 1:
            hoja.cell(row=fila_actual, column=1, value="Total").font = fuente_stock
            celda_total = hoja.cell(row=fila_actual, column=5, value=int(grupo["total"]))
            celda_total.font = fuente_negativo if grupo["total"] < 0 else fuente_stock
            fila_actual += 1

        fila_actual += 1  # renglón en blanco entre proveedores

    for columna, ancho in enumerate((26, 11, 11, 10, 9), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
