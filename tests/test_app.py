import base64
import io
import re
from contextlib import ExitStack
from datetime import date, datetime, time, timedelta, timezone
from unittest.mock import MagicMock, patch

import openpyxl
import pypdfium2 as pdfium
import pytest
from fastapi.testclient import TestClient
from PIL import Image

from app.db import (
    DATABASE_URL_ENV_VAR,
    obtener_conexion,
    RepartoDesactualizado,
    ReprocesoAnteriorAlCorte,
    StockInsuficienteParaReproceso,
)
from app.main import (
    SECTORES,
    _ICONO_INICIO,
    _fecha_de_corte_limpieza_fotos,
    _formatear_bytes,
    _formatear_fecha_corta,
    _formatear_kilos,
    _formatear_moneda,
    _formatear_numero,
    _generar_preview_foto,
    _tipo_retiro_default_desde_env,
    TOPE_FILAS_BUSQUEDA,
    _sufijo_unidad,
    app,
    templates,
)

cliente = TestClient(app)


def test_formatear_numero_saca_decimales_de_sobra():
    assert _formatear_numero(16.0) == "16"
    assert _formatear_numero(16) == "16"
    assert _formatear_numero(16.5) == "16.5"
    assert _formatear_numero(16.25) == "16.25"
    assert _formatear_numero(None) == ""


def test_formatear_fecha_corta_muestra_dd_mm():
    assert _formatear_fecha_corta(date(2026, 8, 6)) == "06/08"
    assert _formatear_fecha_corta(None) == ""


def test_formatear_moneda_usa_signo_pesos_y_puntos_para_miles():
    assert _formatear_moneda(20000) == "$20.000"
    assert _formatear_moneda(50000) == "$50.000"
    assert _formatear_moneda(1500000) == "$1.500.000"
    assert _formatear_moneda(500) == "$500"
    assert _formatear_moneda(None) == ""


def test_formatear_moneda_redondea_al_peso_entero():
    assert _formatear_moneda(45000.4) == "$45.000"
    assert _formatear_moneda(45000.6) == "$45.001"


def test_formatear_kilos_muestra_entero_sin_decimales_ni_separador():
    assert _formatear_kilos(1500.5) == "1500"
    assert _formatear_kilos(16.0) == "16"
    assert _formatear_kilos(16) == "16"
    assert _formatear_kilos(1500) == "1500"
    assert _formatear_kilos(None) == ""


def test_sufijo_unidad_devuelve_la_letra_corta():
    assert _sufijo_unidad("kilo") == "k"
    assert _sufijo_unidad("unidad") == "u"
    assert _sufijo_unidad("cubeta") == "c"
    assert _sufijo_unidad(None) == ""
    assert _sufijo_unidad("otracosa") == ""


def test_formatear_bytes_elige_la_unidad_que_queda_natural():
    assert _formatear_bytes(0) == "0 bytes"
    assert _formatear_bytes(500) == "500 bytes"
    assert _formatear_bytes(907397) == "886 KB"  # caso real confirmado en producción
    assert _formatear_bytes(356000000) == "339,5 MB"
    assert _formatear_bytes(1288490188) == "1,2 GB"
    assert _formatear_bytes(None) == ""


def test_fecha_de_corte_limpieza_fotos_es_3_anios_atras():
    with patch("app.main._hoy_argentina", return_value=date(2026, 8, 15)):
        assert _fecha_de_corte_limpieza_fotos() == date(2023, 8, 15)


def test_fecha_de_corte_limpieza_fotos_29_de_febrero_bisiesto():
    # Regresión: hace 3 años (2024, bisiesto) no siempre hay un 29/2 — 2021 no lo es.
    with patch("app.main._hoy_argentina", return_value=date(2024, 2, 29)):
        assert _fecha_de_corte_limpieza_fotos() == date(2021, 2, 28)


def test_raiz_devuelve_estado_ok():
    respuesta = cliente.get("/")
    assert respuesta.status_code == 200
    assert respuesta.json() == {"estado": "ok"}


def test_salud_db_devuelve_la_cantidad_de_articulos():
    with patch("app.main.contar_articulos", return_value=29):
        respuesta = cliente.get("/salud/db")

    assert respuesta.status_code == 200
    assert respuesta.json() == {"articulos": 29}


def test_salud_db_sin_database_url_devuelve_error_claro():
    with patch("app.main.contar_articulos", side_effect=RuntimeError(f"Falta configurar {DATABASE_URL_ENV_VAR}")):
        respuesta = cliente.get("/salud/db")

    assert respuesta.status_code == 500
    assert DATABASE_URL_ENV_VAR in respuesta.json()["detail"]


def test_salud_db_error_de_conexion_devuelve_500():
    with patch("app.main.contar_articulos", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/salud/db")

    assert respuesta.status_code == 500


def test_obtener_conexion_sin_database_url_lanza_error_claro(monkeypatch):
    monkeypatch.delenv(DATABASE_URL_ENV_VAR, raising=False)

    with pytest.raises(RuntimeError, match=DATABASE_URL_ENV_VAR):
        obtener_conexion()


ARTICULOS_DE_PRUEBA = [
    {"id": 1, "nombre": "Frutilla", "unidad_compra": None, "contenido_referencia": None, "grupo": "fruta"},
    {"id": 2, "nombre": "Mango", "unidad_compra": None, "contenido_referencia": None, "grupo": None},
]


def test_ver_articulos_lista_vacia():
    with patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.get("/articulos")

    assert respuesta.status_code == 200
    assert "No hay artículos cargados todavía." in respuesta.text


def test_ver_articulos_error_de_base_muestra_pagina_de_error_clara():
    with patch("app.main.listar_articulos", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/articulos")

    assert respuesta.status_code == 500
    assert "No se pudo leer el catálogo" in respuesta.text


def test_ver_articulos_muestra_solo_nombre():
    with patch("app.main.listar_articulos", return_value=ARTICULOS_DE_PRUEBA):
        respuesta = cliente.get("/articulos")

    assert respuesta.status_code == 200
    assert "Frutilla" in respuesta.text
    assert "Mango" in respuesta.text
    assert "/articulos/1/editar" in respuesta.text
    assert "/articulos/1/eliminar" in respuesta.text
    # codigo_interno es del cliente Día, no del artículo: no debe pedirse ni mostrarse acá
    assert "codigo_interno" not in respuesta.text
    # unidad_venta y envase ya no viven en articulos: no deben aparecer en la página
    assert "unidad_venta" not in respuesta.text
    assert "envase" not in respuesta.text.lower()
    # merma ya no se pide ni se muestra en esta pantalla
    assert "merma" not in respuesta.text.lower()


def test_ver_articulos_muestra_columna_grupo_con_sin_clasificar():
    with patch("app.main.listar_articulos", return_value=ARTICULOS_DE_PRUEBA):
        respuesta = cliente.get("/articulos")

    assert respuesta.status_code == 200
    assert "<th>Grupo</th>" in respuesta.text
    assert "Fruta" in respuesta.text
    assert "Sin clasificar" in respuesta.text


def test_ver_articulos_incluye_link_a_inicio():
    with patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.get("/articulos")

    assert respuesta.status_code == 200
    assert 'href="/inicio"' in respuesta.text


def test_agregar_articulo_exitoso_redirige_a_articulos():
    with patch("app.main.crear_articulo") as mock_crear:
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Kiwi", "unidad_compra": "unidad", "contenido_referencia": "10"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/articulos"
    mock_crear.assert_called_once_with("Kiwi", "unidad", 10.0, None)


def test_agregar_articulo_sin_contenido_referencia_guarda_none():
    with patch("app.main.crear_articulo") as mock_crear:
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Kiwi", "unidad_compra": "kilo", "contenido_referencia": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Kiwi", "kilo", None, None)


def test_agregar_articulo_con_grupo_valido_lo_guarda():
    with patch("app.main.crear_articulo") as mock_crear:
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Kiwi", "unidad_compra": "kilo", "contenido_referencia": "", "grupo": "fruta"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Kiwi", "kilo", None, "fruta")


def test_agregar_articulo_con_grupo_hoja_lo_guarda():
    with patch("app.main.crear_articulo") as mock_crear:
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Rúcula", "unidad_compra": "unidad", "contenido_referencia": "", "grupo": "hoja"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Rúcula", "unidad", None, "hoja")


def test_agregar_articulo_con_grupo_pesada_lo_guarda():
    with patch("app.main.crear_articulo") as mock_crear:
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Zapallo", "unidad_compra": "kilo", "contenido_referencia": "", "grupo": "pesada"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Zapallo", "kilo", None, "pesada")


def test_agregar_articulo_grupo_invalido_muestra_error():
    with patch("app.main.crear_articulo") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post(
            "/articulos/nuevo",
            data={"nombre": "Kiwi", "unidad_compra": "kilo", "grupo": "lacteo"},
        )

    assert respuesta.status_code == 400
    assert "grupo válido" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_articulo_nombre_vacio_muestra_error():
    with patch("app.main.crear_articulo") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post("/articulos/nuevo", data={"nombre": "   ", "unidad_compra": "kilo"})

    assert respuesta.status_code == 400
    assert "no puede estar vacío" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_articulo_nombre_string_vacio_muestra_error_prolijo_no_422():
    # Regresión: un campo de texto Form(...) vacío ("" y no solo espacios) hacía
    # que FastAPI devolviera un 422 crudo en vez de nuestro error prolijo.
    with patch("app.main.crear_articulo") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post("/articulos/nuevo", data={"nombre": "", "unidad_compra": "kilo"})

    assert respuesta.status_code == 400
    assert "no puede estar vacío" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_articulo_unidad_compra_invalida_muestra_error():
    with patch("app.main.crear_articulo") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post("/articulos/nuevo", data={"nombre": "Kiwi", "unidad_compra": "litro"})

    assert respuesta.status_code == 400
    assert "unidad de compra válida" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_articulo_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.crear_articulo", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_articulos", return_value=[]),
    ):
        respuesta = cliente.post("/articulos/nuevo", data={"nombre": "Kiwi", "unidad_compra": "kilo"})

    assert respuesta.status_code == 500
    assert "No se pudo guardar" in respuesta.text


ARTICULO_DE_PRUEBA = {
    "id": 1,
    "nombre": "Frutilla",
    "unidad_compra": "cubeta",
    "contenido_referencia": 12,
    "grupo": "fruta",
}


def test_ver_editar_articulo_muestra_datos_precargados():
    with patch("app.main.obtener_articulo", return_value=ARTICULO_DE_PRUEBA):
        respuesta = cliente.get("/articulos/1/editar")

    assert respuesta.status_code == 200
    assert "Frutilla" in respuesta.text
    assert 'action="/articulos/1/editar"' in respuesta.text
    assert "merma" not in respuesta.text.lower()
    assert '<option value="fruta" selected>Fruta</option>' in respuesta.text


def test_ver_editar_articulo_inexistente_da_404():
    with patch("app.main.obtener_articulo", return_value=None):
        respuesta = cliente.get("/articulos/999/editar")

    assert respuesta.status_code == 404


def test_ver_editar_articulo_error_de_base_da_500():
    with patch("app.main.obtener_articulo", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/articulos/1/editar")

    assert respuesta.status_code == 500


def test_editar_articulo_exitoso_redirige_a_articulos():
    with patch("app.main.actualizar_articulo") as mock_actualizar:
        respuesta = cliente.post(
            "/articulos/1/editar",
            data={"nombre": "Frutilla Premium", "unidad_compra": "cubeta", "contenido_referencia": "12"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/articulos"
    mock_actualizar.assert_called_once_with(1, "Frutilla Premium", "cubeta", 12.0, None)


def test_editar_articulo_con_grupo_valido_lo_guarda():
    with patch("app.main.actualizar_articulo") as mock_actualizar:
        respuesta = cliente.post(
            "/articulos/1/editar",
            data={"nombre": "Frutilla", "unidad_compra": "cubeta", "contenido_referencia": "12", "grupo": "hortaliza"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(1, "Frutilla", "cubeta", 12.0, "hortaliza")


def test_editar_articulo_grupo_invalido_muestra_error():
    with patch("app.main.actualizar_articulo") as mock_actualizar:
        respuesta = cliente.post(
            "/articulos/1/editar",
            data={"nombre": "Frutilla", "unidad_compra": "cubeta", "grupo": "lacteo"},
        )

    assert respuesta.status_code == 400
    assert "grupo válido" in respuesta.text
    mock_actualizar.assert_not_called()


def test_editar_articulo_nombre_vacio_muestra_error():
    with patch("app.main.actualizar_articulo") as mock_actualizar:
        respuesta = cliente.post("/articulos/1/editar", data={"nombre": "   ", "unidad_compra": "kilo"})

    assert respuesta.status_code == 400
    assert "no puede estar vacío" in respuesta.text
    mock_actualizar.assert_not_called()


def test_editar_articulo_unidad_compra_invalida_muestra_error():
    with patch("app.main.actualizar_articulo") as mock_actualizar:
        respuesta = cliente.post("/articulos/1/editar", data={"nombre": "Frutilla", "unidad_compra": "litro"})

    assert respuesta.status_code == 400
    assert "unidad de compra válida" in respuesta.text
    mock_actualizar.assert_not_called()


def test_editar_articulo_error_de_base_muestra_mensaje_claro():
    with patch("app.main.actualizar_articulo", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/articulos/1/editar", data={"nombre": "Frutilla", "unidad_compra": "kilo"})

    assert respuesta.status_code == 500
    assert "No se pudo guardar" in respuesta.text


def test_eliminar_articulo_exitoso_redirige_a_articulos():
    with patch("app.main.desactivar_articulo") as mock_desactivar:
        respuesta = cliente.post("/articulos/1/eliminar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/articulos"
    mock_desactivar.assert_called_once_with(1)


def test_eliminar_articulo_no_borra_la_fila_solo_marca_inactivo():
    # desactivar_articulo (borrado lógico) hace UPDATE activo=false, no DELETE.
    # Este test confirma que la ruta llama a esa función y no a otra.
    with patch("app.main.desactivar_articulo") as mock_desactivar:
        cliente.post("/articulos/1/eliminar", follow_redirects=False)

    mock_desactivar.assert_called_once()


def test_eliminar_articulo_error_de_base_da_500():
    with patch("app.main.desactivar_articulo", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/articulos/1/eliminar")

    assert respuesta.status_code == 500


CLIENTES_DE_PRUEBA = [
    {"id": 1, "nombre": "Día", "descuento": 23.0, "adicionales": 0.0, "utilidad_objetivo": 20.0},
    {"id": 2, "nombre": "Otro cliente", "descuento": 15.0, "adicionales": 10.5, "utilidad_objetivo": 10.0},
]


def test_ver_clientes_lista_vacia():
    with patch("app.main.listar_clientes", return_value=[]):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 200
    assert "No hay clientes cargados todavía." in respuesta.text


def test_ver_clientes_error_de_base_muestra_pagina_de_error_clara():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 500
    assert "No se pudo leer los clientes" in respuesta.text


def test_ver_clientes_muestra_nombre_descuentos_adicionales_y_utilidad_como_porcentaje():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 200
    assert "Día" in respuesta.text
    assert "Descuentos" in respuesta.text
    assert "Adicionales" in respuesta.text
    assert "23.0%" in respuesta.text
    assert "20.0%" in respuesta.text
    # "Otro cliente" tiene adicionales (10.5%, ej. IVA) además de descuentos.
    assert "10.5%" in respuesta.text
    assert "/clientes/1/editar" in respuesta.text
    assert "/clientes/1/eliminar" in respuesta.text


def test_ver_clientes_sin_adicionales_muestra_cero_por_ciento():
    # Regla explícita: si no tiene ninguna tasa de tipo suma, se muestra
    # "0%" (no una celda vacía ni un guión que confunda con "no se sabe").
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 200
    assert "0.0%" in respuesta.text


def test_ver_clientes_incluye_link_a_inicio():
    with patch("app.main.listar_clientes", return_value=[]):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 200
    assert 'href="/inicio"' in respuesta.text


def test_ver_agregar_cliente_muestra_formulario_vacio():
    respuesta = cliente.get("/clientes/nuevo")

    assert respuesta.status_code == 200
    assert 'action="/clientes/nuevo"' in respuesta.text
    # Regresión: los títulos y ejemplos de cada grupo de tasas, para que
    # quede claro qué va en cada uno.
    assert "Adicionales" in respuesta.text
    assert "IVA, premios" in respuesta.text
    assert "Descuentos" in respuesta.text
    assert "logística, flete" in respuesta.text
    assert 'id="utilidad_objetivo"' in respuesta.text
    assert '+ Agregar tasa Adicional' in respuesta.text
    assert '+ Agregar tasa de Descuento' in respuesta.text


def test_agregar_cliente_con_tasas_redirige_a_clientes():
    with patch("app.main.crear_cliente", return_value=5) as mock_crear:
        respuesta = cliente.post(
            "/clientes/nuevo",
            data={
                "nombre": "Vea",
                "utilidad_objetivo": "12",
                "cantidad_tasa_sumas": "1",
                "tasa_suma_0_nombre_original": "",
                "tasa_suma_0_valor_original": "",
                "tasa_suma_0_nombre": "IVA",
                "tasa_suma_0_valor": "21",
                "cantidad_tasa_restas": "1",
                "tasa_resta_0_nombre_original": "",
                "tasa_resta_0_valor_original": "",
                "tasa_resta_0_nombre": "Flete",
                "tasa_resta_0_valor": "4",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/clientes"
    mock_crear.assert_called_once_with(
        "Vea", [{"nombre": "IVA", "valor": 0.21}], [{"nombre": "Flete", "valor": 0.04}], 0.12
    )


def test_agregar_cliente_sin_tasas_redirige_a_clientes():
    # Las tasas son opcionales: un cliente se puede cargar solo con nombre
    # y utilidad, sin ninguna tasa todavía.
    with patch("app.main.crear_cliente", return_value=5) as mock_crear:
        respuesta = cliente.post(
            "/clientes/nuevo",
            data={"nombre": "Vea", "utilidad_objetivo": "12"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Vea", [], [], 0.12)


def test_agregar_cliente_nombre_vacio_muestra_error():
    with patch("app.main.crear_cliente") as mock_crear:
        respuesta = cliente.post("/clientes/nuevo", data={"nombre": "   ", "utilidad_objetivo": "12"})

    assert respuesta.status_code == 400
    assert "no puede estar vacío" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_cliente_utilidad_fuera_de_rango_muestra_error():
    with patch("app.main.crear_cliente") as mock_crear:
        respuesta = cliente.post("/clientes/nuevo", data={"nombre": "Vea", "utilidad_objetivo": "150"})

    assert respuesta.status_code == 400
    assert "entre 0 y 100" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_cliente_utilidad_no_numerica_muestra_error():
    with patch("app.main.crear_cliente") as mock_crear:
        respuesta = cliente.post("/clientes/nuevo", data={"nombre": "Vea", "utilidad_objetivo": "abc"})

    assert respuesta.status_code == 400
    assert "tiene que ser un número" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_cliente_tasa_con_nombre_sin_porcentaje_muestra_error():
    # Fila a medio completar (nombre sin %, o al revés): no se guarda nada
    # dudoso, se avisa para que la complete o la saque.
    with patch("app.main.crear_cliente") as mock_crear:
        respuesta = cliente.post(
            "/clientes/nuevo",
            data={
                "nombre": "Vea",
                "utilidad_objetivo": "12",
                "cantidad_tasa_sumas": "1",
                "tasa_suma_0_nombre": "IVA",
                "tasa_suma_0_valor": "",
            },
        )

    assert respuesta.status_code == 400
    assert "Completá el nombre y el porcentaje" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_cliente_fila_de_tasa_completamente_vacia_se_ignora():
    # El usuario tocó "+ Agregar tasa" pero no llegó a cargar nada — no
    # tiene que bloquear el guardado.
    with patch("app.main.crear_cliente", return_value=5) as mock_crear:
        respuesta = cliente.post(
            "/clientes/nuevo",
            data={
                "nombre": "Vea",
                "utilidad_objetivo": "12",
                "cantidad_tasa_sumas": "1",
                "tasa_suma_0_nombre": "",
                "tasa_suma_0_valor": "",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Vea", [], [], 0.12)


def test_agregar_cliente_error_de_base_muestra_mensaje_claro():
    with patch("app.main.crear_cliente", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/clientes/nuevo", data={"nombre": "Vea", "utilidad_objetivo": "12"})

    assert respuesta.status_code == 500
    assert "No se pudo guardar" in respuesta.text


CLIENTE_DE_PRUEBA = {"id": 1, "nombre": "Día"}

CONCEPTOS_EDITABLES_DE_PRUEBA = {
    "tasas_suma": [{"nombre": "IVA", "valor_pct": 21.0}],
    "tasas_resta": [{"nombre": "Flete", "valor_pct": 4.0}],
    "utilidad_pct": 20.0,
}


def test_ver_editar_cliente_muestra_datos_precargados():
    with (
        patch("app.main.obtener_cliente", return_value=CLIENTE_DE_PRUEBA),
        patch("app.main.listar_conceptos_editables_por_cliente", return_value=CONCEPTOS_EDITABLES_DE_PRUEBA),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
    ):
        respuesta = cliente.get("/clientes/1/editar")

    assert respuesta.status_code == 200
    assert "Día" in respuesta.text
    assert 'action="/clientes/1/editar"' in respuesta.text
    assert 'value="IVA"' in respuesta.text
    assert 'value="21.0"' in respuesta.text
    assert 'value="Flete"' in respuesta.text
    assert 'value="4.0"' in respuesta.text
    assert 'value="20.0"' in respuesta.text
    # Regresión: nombre_original/valor_original viajan ocultos, son el
    # punto de partida contra el que se compara al guardar.
    assert 'name="tasa_suma_0_nombre_original" value="IVA"' in respuesta.text
    assert 'name="tasa_resta_0_nombre_original" value="Flete"' in respuesta.text
    assert 'name="utilidad_original" value="20.0"' in respuesta.text
    assert "Dar de baja esta tasa" in respuesta.text


def test_ver_editar_cliente_inexistente_da_404():
    with patch("app.main.obtener_cliente", return_value=None):
        respuesta = cliente.get("/clientes/999/editar")

    assert respuesta.status_code == 404


def test_ver_editar_cliente_error_de_base_da_500():
    with patch("app.main.obtener_cliente", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/clientes/1/editar")

    assert respuesta.status_code == 500


def test_ver_editar_cliente_error_al_leer_conceptos_da_500():
    with (
        patch("app.main.obtener_cliente", return_value=CLIENTE_DE_PRUEBA),
        patch("app.main.listar_conceptos_editables_por_cliente", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.get("/clientes/1/editar")

    assert respuesta.status_code == 500


def _datos_editar_cliente(**overrides):
    datos = {
        "nombre": "Día",
        "utilidad_objetivo": "20",
        "utilidad_original": "20",
        "cantidad_tasa_sumas": "1",
        "tasa_suma_0_nombre_original": "IVA",
        "tasa_suma_0_valor_original": "21",
        "tasa_suma_0_nombre": "IVA",
        "tasa_suma_0_valor": "21",
        "cantidad_tasa_restas": "1",
        "tasa_resta_0_nombre_original": "Flete",
        "tasa_resta_0_valor_original": "4",
        "tasa_resta_0_nombre": "Flete",
        "tasa_resta_0_valor": "4",
    }
    datos.update(overrides)
    return datos


def test_editar_cliente_sin_ningun_cambio_no_genera_filas_nuevas():
    # Regresión explícita: si nada se tocó, no hay que agregar filas de
    # historial de más.
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post("/clientes/1/editar", data=_datos_editar_cliente(), follow_redirects=False)

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(1, "Día", [])


def test_editar_cliente_tasa_editada_genera_fila_nueva_sin_pisar_la_vieja():
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post(
            "/clientes/1/editar",
            data=_datos_editar_cliente(tasa_resta_0_valor="5"),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(1, "Día", [{"nombre_parametro": "Flete", "tipo": "resta", "valor": 0.05}])


def test_editar_cliente_tasa_nueva_se_agrega():
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post(
            "/clientes/1/editar",
            data=_datos_editar_cliente(
                cantidad_tasa_sumas="2",
                tasa_suma_1_nombre_original="",
                tasa_suma_1_valor_original="",
                tasa_suma_1_nombre="Premio",
                tasa_suma_1_valor="2",
            ),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(1, "Día", [{"nombre_parametro": "Premio", "tipo": "suma", "valor": 0.02}])


def test_editar_cliente_tasa_dada_de_baja_genera_fila_en_cero():
    # Regla de oro: nunca se borra el historial, se agrega una fila en 0
    # vigente desde hoy — los cálculos pasados siguen viendo el valor viejo.
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post(
            "/clientes/1/editar",
            data=_datos_editar_cliente(tasa_resta_0_baja="on"),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(1, "Día", [{"nombre_parametro": "Flete", "tipo": "resta", "valor": 0.0}])


def test_editar_cliente_tasa_renombrada_da_de_baja_la_vieja_y_alta_la_nueva():
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post(
            "/clientes/1/editar",
            data=_datos_editar_cliente(tasa_resta_0_nombre="Flete y logística"),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(
        1,
        "Día",
        [
            {"nombre_parametro": "Flete", "tipo": "resta", "valor": 0.0},
            {"nombre_parametro": "Flete y logística", "tipo": "resta", "valor": 0.04},
        ],
    )


def test_editar_cliente_utilidad_editada_genera_fila_nueva():
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post(
            "/clientes/1/editar",
            data=_datos_editar_cliente(utilidad_objetivo="25"),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar.assert_called_once_with(
        1, "Día", [{"nombre_parametro": "utilidad_objetivo", "tipo": "utilidad", "valor": 0.25}]
    )


def test_editar_cliente_nombre_vacio_muestra_error():
    with patch("app.main.actualizar_cliente") as mock_actualizar:
        respuesta = cliente.post("/clientes/1/editar", data=_datos_editar_cliente(nombre="   "))

    assert respuesta.status_code == 400
    assert "no puede estar vacío" in respuesta.text
    mock_actualizar.assert_not_called()


def test_editar_cliente_error_de_base_muestra_mensaje_claro():
    with patch("app.main.actualizar_cliente", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/clientes/1/editar", data=_datos_editar_cliente())

    assert respuesta.status_code == 500
    assert "No se pudo guardar" in respuesta.text


def test_eliminar_cliente_exitoso_redirige_a_clientes():
    with patch("app.main.desactivar_cliente") as mock_desactivar:
        respuesta = cliente.post("/clientes/1/eliminar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/clientes"
    mock_desactivar.assert_called_once_with(1)


def test_eliminar_cliente_error_de_base_da_500():
    with patch("app.main.desactivar_cliente", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/clientes/1/eliminar")

    assert respuesta.status_code == 500


CLIENTES_PARA_SELECTOR = [{"id": 1, "nombre": "Día"}, {"id": 2, "nombre": "Vea"}]

FICHAS_DE_PRUEBA = [
    {
        "id": 10,
        "articulo_nombre": "Mango",
        "envase_nombre": "Caja Chica Día",
        "contenido_caja": 10,
        "unidad_venta": "unidad",
        "envase_variable": True,
        "nombre_cliente": "MANGO",
        "codigo_cliente": "225863",
    },
    {
        "id": 11,
        "articulo_nombre": "Sandía",
        "envase_nombre": None,
        "contenido_caja": 18,
        "unidad_venta": "kilo",
        "envase_variable": False,
        "nombre_cliente": None,
        "codigo_cliente": None,
    },
]


def test_ver_fichas_sin_cliente_elegido_pide_elegir_uno():
    with patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR):
        respuesta = cliente.get("/fichas")

    assert respuesta.status_code == 200
    assert "Elegí un cliente" in respuesta.text
    assert "Día" in respuesta.text  # aparece como opción del selector


def test_ver_fichas_error_al_leer_clientes_muestra_error_claro():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/fichas")

    assert respuesta.status_code == 500
    assert "No se pudo leer los clientes" in respuesta.text


def test_ver_fichas_con_cliente_muestra_la_lista():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/fichas?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Mango" in respuesta.text
    assert "Caja Chica Día" in respuesta.text
    assert "Sandía" in respuesta.text
    assert "Sin envase" in respuesta.text
    assert "/fichas/10/editar" in respuesta.text
    assert "/fichas/nueva?cliente_id=1" in respuesta.text
    assert "Variable" in respuesta.text
    assert "Fijo" in respuesta.text
    # Alias fusionados desde la vieja /conversion: se muestran en la tabla.
    assert "<td>MANGO</td>" in respuesta.text
    assert "<td>225863</td>" in respuesta.text
    # Sandía no tiene alias cargado: se muestra "-", no vacío ni error.
    assert respuesta.text.count("<td>-</td>") >= 2


def test_ver_fichas_error_al_leer_fichas_muestra_error_claro():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.get("/fichas?cliente_id=1")

    assert respuesta.status_code == 500
    assert "No se pudieron leer las fichas" in respuesta.text


# El catálogo pelado, para lo que solo necesita id y nombre (compras).
ARTICULOS_DEL_CATALOGO = [{"id": 5, "nombre": "Kiwi"}]

# La lista que ofrece el form de ficha: TODOS los artículos activos, cada
# uno diciendo cuántas fichas ya tiene este cliente (0 = ninguna todavía).
ARTICULOS_PARA_FICHA = [
    {"id": 5, "nombre": "Kiwi", "fichas_existentes": 0},
    {"id": 6, "nombre": "Banana", "fichas_existentes": 1},
]
ENVASES_DEL_CLIENTE = [{"id": 100, "nombre": "Caja Chica Día"}]


def test_ver_nueva_ficha_muestra_formulario():
    with (
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.get("/fichas/nueva?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Kiwi" in respuesta.text
    assert "Caja Chica Día" in respuesta.text
    assert 'id="nombre_cliente" name="nombre_cliente"' in respuesta.text
    assert 'id="codigo_cliente" name="codigo_cliente"' in respuesta.text


def test_ver_nueva_ficha_sin_cliente_id_da_422():
    respuesta = cliente.get("/fichas/nueva")
    assert respuesta.status_code == 422


def test_ver_nueva_ficha_sin_articulos_en_el_catalogo_muestra_mensaje():
    # El único caso en que no hay nada para elegir es que el catálogo esté
    # vacío: que el artículo ya tenga ficha dejó de esconderlo (un cliente
    # puede tener dos fichas del mismo artículo).
    with (
        patch("app.main._articulos_para_ficha", return_value=[]),
        patch("app.main.listar_envases", return_value=[]),
    ):
        respuesta = cliente.get("/fichas/nueva?cliente_id=1")

    assert respuesta.status_code == 200
    assert "No hay artículos activos en el catálogo." in respuesta.text


def test_ver_nueva_ficha_ofrece_los_articulos_que_ya_tienen_ficha_avisando():
    # Lo que antes estaba prohibido ahora es un aviso: "Banana" aparece en
    # la lista aunque ya tenga ficha, porque cargar la segunda es el caso
    # que este cambio vino a habilitar.
    with (
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.get("/fichas/nueva?cliente_id=1")

    assert "Banana — ya tiene 1 ficha<" in respuesta.text
    assert ">Kiwi<" in respuesta.text  # sin ficha todavía: sin aviso
    assert "Banana Bolivia y Banana Ecuador" in respuesta.text


def test_ver_nueva_ficha_error_de_base_da_500():
    with patch("app.main._articulos_para_ficha", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/fichas/nueva?cliente_id=1")

    assert respuesta.status_code == 500


def test_agregar_ficha_exitosa_redirige_a_fichas_del_cliente():
    with patch("app.main.crear_ficha") as mock_crear:
        respuesta = cliente.post(
            "/fichas/nueva",
            data={
                "cliente_id": "1",
                "articulo_id": "5",
                "envase_id": "100",
                "contenido_caja": "10",
                "unidad_venta": "unidad",
                "envase_variable": "si",
                "nombre_cliente": "MANZ ROJ ELE",
                "codigo_cliente": "90039",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/fichas?cliente_id=1"
    mock_crear.assert_called_once_with(5, 1, 100, 10.0, "unidad", True, "MANZ ROJ ELE", "90039")


def test_agregar_ficha_sin_envase_con_contenido_caja_exitosa():
    with patch("app.main.crear_ficha") as mock_crear:
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "", "contenido_caja": "12", "unidad_venta": "cubeta"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Sin nombre_cliente/codigo_cliente en el form: quedan en None (son opcionales).
    mock_crear.assert_called_once_with(5, 1, None, 12.0, "cubeta", False, None, None)


def test_agregar_ficha_sin_envase_sin_contenido_caja_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "", "contenido_caja": "", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 400
    assert "obligatorio" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_sin_articulo_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "", "envase_id": "", "contenido_caja": "", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 400
    assert "Elegí un artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_con_envase_sin_contenido_caja_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "100", "contenido_caja": "", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 400
    assert "obligatorio" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_contenido_caja_no_numerico_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "100", "contenido_caja": "abc", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 400
    assert "tiene que ser un número" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_contenido_caja_cero_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "100", "contenido_caja": "0", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 400
    assert "mayor a cero" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_unidad_venta_invalida_muestra_error():
    with (
        patch("app.main.crear_ficha") as mock_crear,
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "", "contenido_caja": "", "unidad_venta": "litro"},
        )

    assert respuesta.status_code == 400
    assert "unidad de venta válida" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_ficha_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.crear_ficha", side_effect=Exception("no se pudo conectar")),
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/nueva",
            data={"cliente_id": "1", "articulo_id": "5", "envase_id": "", "contenido_caja": "10", "unidad_venta": "kilo"},
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la ficha" in respuesta.text


FICHA_DE_PRUEBA = {
    "id": 10,
    "cliente_id": 1,
    "cliente_nombre": "Día",
    "articulo_id": 5,
    "articulo_nombre": "Mango",
    "envase_id": 100,
    "contenido_caja": 10,
    "unidad_venta": "unidad",
    "envase_variable": True,
    "nombre_cliente": "MANGO",
    "codigo_cliente": "225863",
}


def test_ver_editar_ficha_muestra_datos_precargados():
    with (
        patch("app.main.obtener_ficha", return_value=FICHA_DE_PRUEBA),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
        patch("app.main._articulos_para_ficha", return_value=ARTICULOS_PARA_FICHA),
    ):
        respuesta = cliente.get("/fichas/10/editar")

    assert respuesta.status_code == 200
    assert "Mango" in respuesta.text
    assert 'action="/fichas/10/editar"' in respuesta.text
    assert "checked" in respuesta.text  # envase_variable=True precargado
    # Alias fusionados desde la vieja /conversion: editables en el mismo form.
    assert 'id="nombre_cliente" name="nombre_cliente"' in respuesta.text
    assert 'value="MANGO"' in respuesta.text
    assert 'id="codigo_cliente" name="codigo_cliente"' in respuesta.text
    assert 'value="225863"' in respuesta.text
    # Sección "Cambiar artículo": solo ofrece artículos SIN ficha del cliente.
    assert 'action="/fichas/10/cambiar-articulo"' in respuesta.text
    assert "Kiwi" in respuesta.text
    assert "Los precios ya negociados no cambian" in respuesta.text
    # El alias que se va a conservar queda a la vista y editable, con el
    # aviso: si el destino es otro producto, probablemente haya que cambiarlo.
    assert 'id="cambio_nombre_cliente" name="nombre_cliente"' in respuesta.text
    assert 'id="cambio_codigo_cliente" name="codigo_cliente"' in respuesta.text
    assert "probablemente haya que cambiar este alias" in respuesta.text


def test_ver_editar_ficha_inexistente_da_404():
    with patch("app.main.obtener_ficha", return_value=None):
        respuesta = cliente.get("/fichas/999/editar")

    assert respuesta.status_code == 404


def test_ver_editar_ficha_error_de_base_da_500():
    with patch("app.main.obtener_ficha", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/fichas/10/editar")

    assert respuesta.status_code == 500


def test_editar_ficha_exitosa_redirige_a_fichas_del_cliente():
    with patch("app.main.actualizar_ficha") as mock_actualizar:
        respuesta = cliente.post(
            "/fichas/10/editar",
            data={
                "cliente_id": "1",
                "articulo_nombre": "Mango",
                "envase_id": "100",
                "contenido_caja": "12",
                "unidad_venta": "unidad",
                "envase_variable": "si",
                "nombre_cliente": "MANGO",
                "codigo_cliente": "225863",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/fichas?cliente_id=1"
    mock_actualizar.assert_called_once_with(10, 100, 12.0, "unidad", True, "MANGO", "225863")


def test_editar_ficha_sin_tildar_envase_variable_guarda_false():
    with patch("app.main.actualizar_ficha") as mock_actualizar:
        respuesta = cliente.post(
            "/fichas/10/editar",
            data={
                "cliente_id": "1",
                "articulo_nombre": "Mango",
                "envase_id": "100",
                "contenido_caja": "12",
                "unidad_venta": "unidad",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Sin nombre_cliente/codigo_cliente en el form: quedan en None (son opcionales).
    mock_actualizar.assert_called_once_with(10, 100, 12.0, "unidad", False, None, None)


def test_editar_ficha_sin_contenido_caja_muestra_error():
    with (
        patch("app.main.actualizar_ficha") as mock_actualizar,
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/10/editar",
            data={
                "cliente_id": "1",
                "articulo_nombre": "Mango",
                "envase_id": "100",
                "contenido_caja": "",
                "unidad_venta": "unidad",
            },
        )

    assert respuesta.status_code == 400
    assert "obligatorio" in respuesta.text
    mock_actualizar.assert_not_called()


def test_editar_ficha_unidad_venta_invalida_muestra_error():
    with (
        patch("app.main.actualizar_ficha") as mock_actualizar,
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/10/editar",
            data={
                "cliente_id": "1",
                "articulo_nombre": "Mango",
                "envase_id": "100",
                "contenido_caja": "12",
                "unidad_venta": "litro",
            },
        )

    assert respuesta.status_code == 400
    assert "unidad de venta válida" in respuesta.text
    assert "Mango" in respuesta.text  # no se pierde el artículo al re-mostrar el error
    mock_actualizar.assert_not_called()


def test_editar_ficha_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.actualizar_ficha", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_envases", return_value=ENVASES_DEL_CLIENTE),
    ):
        respuesta = cliente.post(
            "/fichas/10/editar",
            data={
                "cliente_id": "1",
                "articulo_nombre": "Mango",
                "envase_id": "100",
                "contenido_caja": "12",
                "unidad_venta": "unidad",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la ficha" in respuesta.text


def test_eliminar_ficha_exitosa_redirige_a_fichas_del_cliente():
    with patch("app.main.eliminar_ficha") as mock_eliminar:
        respuesta = cliente.post("/fichas/10/eliminar", data={"cliente_id": "1"}, follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/fichas?cliente_id=1"
    mock_eliminar.assert_called_once_with(10)


def test_eliminar_ficha_error_de_base_da_500():
    with patch("app.main.eliminar_ficha", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/fichas/10/eliminar", data={"cliente_id": "1"})

    assert respuesta.status_code == 500


def test_cambiar_articulo_de_ficha_redirige_a_fichas_con_aviso():
    with patch("app.main.cambiar_articulo_de_ficha", return_value=33) as mock_cambiar:
        respuesta = cliente.post(
            "/fichas/10/cambiar-articulo",
            data={"cliente_id": "1", "articulo_nuevo_id": "5", "nombre_cliente": "ANCO", "codigo_cliente": "90200"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/fichas?")
    assert "cliente_id=1" in location
    assert "aviso=" in location
    # El alias editado en la pantalla viaja a la ficha nueva.
    mock_cambiar.assert_called_once_with(10, 5, "ANCO", "90200")


def test_cambiar_articulo_con_alias_vacio_lo_guarda_como_none():
    with patch("app.main.cambiar_articulo_de_ficha", return_value=33) as mock_cambiar:
        respuesta = cliente.post(
            "/fichas/10/cambiar-articulo",
            data={"cliente_id": "1", "articulo_nuevo_id": "5", "nombre_cliente": "  ", "codigo_cliente": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_cambiar.assert_called_once_with(10, 5, None, None)


def test_cambiar_articulo_de_ficha_inexistente_da_404():
    with patch("app.main.cambiar_articulo_de_ficha", return_value=None):
        respuesta = cliente.post(
            "/fichas/999/cambiar-articulo",
            data={"cliente_id": "1", "articulo_nuevo_id": "5"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 404


def test_cambiar_articulo_sin_elegir_articulo_vuelve_a_editar_con_error():
    with patch("app.main.cambiar_articulo_de_ficha") as mock_cambiar:
        respuesta = cliente.post(
            "/fichas/10/cambiar-articulo",
            data={"cliente_id": "1", "articulo_nuevo_id": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"].startswith("/fichas/10/editar?error=")
    mock_cambiar.assert_not_called()


def test_cambiar_articulo_con_error_de_base_vuelve_a_editar_con_error():
    # Cubre también la carrera del unique (dos pestañas): el error de la base
    # se muestra en Editar, y el rollback dejó la ficha original intacta.
    with patch("app.main.cambiar_articulo_de_ficha", side_effect=Exception("unique_violation")):
        respuesta = cliente.post(
            "/fichas/10/cambiar-articulo",
            data={"cliente_id": "1", "articulo_nuevo_id": "5"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"].startswith("/fichas/10/editar?error=")


HISTORIAL_FICHAS_DE_PRUEBA = [
    {
        "id": 3,
        "ficha_id": 33,
        "articulo_id": 5,
        "articulo_nombre": "Berenjena Elegida",
        "envase_id": 100,
        "envase_nombre": "Caja Chica Día",
        "contenido_caja": 6,
        "unidad_venta": "kilo",
        "envase_variable": False,
        "nombre_cliente": "BERENJENA",
        "codigo_cliente": None,
        "evento": "alta",
        "registrado_en": datetime(2026, 8, 21, 10, 30),
    },
    {
        "id": 2,
        "ficha_id": 10,
        "articulo_id": 4,
        "articulo_nombre": "Berenjena",
        "envase_id": 100,
        "envase_nombre": "Caja Chica Día",
        "contenido_caja": 6,
        "unidad_venta": "kilo",
        "envase_variable": False,
        "nombre_cliente": "BERENJENA",
        "codigo_cliente": None,
        "evento": "borrado",
        "registrado_en": datetime(2026, 8, 21, 10, 30),
    },
    {
        "id": 1,
        "ficha_id": 10,
        "articulo_id": 4,
        "articulo_nombre": "Berenjena",
        "envase_id": None,
        "envase_nombre": None,
        "contenido_caja": None,
        "unidad_venta": "kilo",
        "envase_variable": False,
        "nombre_cliente": None,
        "codigo_cliente": None,
        "evento": "foto_inicial",
        "registrado_en": datetime(2026, 8, 1, 9, 0),
    },
]


def test_ver_historial_fichas_muestra_los_eventos():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_historial_fichas_por_cliente", return_value=HISTORIAL_FICHAS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/fichas/historial?cliente_id=1")

    assert respuesta.status_code == 200
    # Un cambio de artículo se ve como borrado + alta: los dos artículos aparecen.
    assert "Berenjena Elegida" in respuesta.text
    assert "Alta" in respuesta.text
    assert "Borrado" in respuesta.text
    assert "Foto inicial" in respuesta.text
    assert "Caja Chica Día" in respuesta.text
    assert "sin envase" in respuesta.text


def test_ver_historial_fichas_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR):
        respuesta = cliente.get("/fichas/historial?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_fichas_muestra_el_aviso_del_redirect_y_el_link_al_historial():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/fichas?cliente_id=1&aviso=La+ficha+ahora+apunta+al+art%C3%ADculo+nuevo.")

    assert respuesta.status_code == 200
    assert "La ficha ahora apunta al artículo nuevo." in respuesta.text
    assert 'href="/fichas/historial?cliente_id=1"' in respuesta.text


def test_conversion_ya_no_existe_como_pantalla_propia():
    # Regresión: /conversion se fusionó dentro de las fichas de logística
    # del cliente (nombre_cliente/codigo_cliente ahora viven en la ficha).
    assert cliente.get("/conversion").status_code == 404
    assert cliente.get("/conversion/nueva?cliente_id=1").status_code == 404


HOY_DE_PRUEBA = date(2026, 8, 6)

COMPRAS_DE_PRUEBA = [
    {
        "id": 30,
        "fecha_operacion": HOY_DE_PRUEBA,
        "articulo_nombre": "Mzn Red",
        "proveedor_nombre": "Saturno",
        "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10,
        "contenido_por_cajon": 18,
        "unidad_compra": "kilo",
        "cantidad_kilos": 180,
        "cantidad_fraccion": None,
        "importe": 50000,
        "sena": None,
        "tipo_retiro": "Clark",
    },
    {
        "id": 31,
        "fecha_operacion": HOY_DE_PRUEBA - timedelta(days=1),
        "articulo_nombre": "Mango",
        "proveedor_nombre": "Frutamax",
        "proveedor_codigo_puesto": "L03P38",
        "cantidad_cajones": 5,
        "contenido_por_cajon": 10,
        "unidad_compra": "unidad",
        "cantidad_kilos": None,
        "cantidad_fraccion": 50,
        "importe": 15000,
        "sena": 2000,
        "tipo_retiro": "Carro",
    },
]


def test_ver_compras_muestra_solo_la_botonera():
    # /compras es solo el título y las dos botoneras — el listado vive en
    # /compras/buscar (Últimas Compras ya no existe como pantalla propia).
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert '<div class="barra-titulo">Compras</div>' in respuesta.text
    assert ">Cargar<" in respuesta.text
    assert ">Operaciones<" in respuesta.text
    assert respuesta.text.index(">Cargar<") < respuesta.text.index(">Operaciones<")
    assert "Buscar Compras" in respuesta.text
    assert "Últimas Compras" not in respuesta.text
    assert "<table" not in respuesta.text
    assert 'id="boton-borrar-seleccionadas"' not in respuesta.text


def test_ver_compras_no_muestra_nada_de_sistema_ahi():
    # Regresión: el indicador de espacio y el botón de limpieza de fotos se
    # movieron a /sistema — /compras es operativa, no tiene que mostrar
    # nada de eso (ni siquiera si esas funciones responden bien).
    with (
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 12, "bytes_totales": 907397}),
        patch("app.main.listar_fotos_para_limpiar", return_value=["2020-01-01/x.jpg"]),
    ):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "fotos guardadas" not in respuesta.text
    assert 'id="boton-limpiar-fotos-viejas"' not in respuesta.text
    assert 'href="/sistema"' in respuesta.text


def test_ver_compras_incluye_links_a_catalogo_y_a_inicio():
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert 'href="/articulos"' in respuesta.text
    assert 'href="/inicio"' in respuesta.text
    # Regresión: /conversion se fusionó dentro de las fichas de logística
    # del cliente, ya no existe como pantalla propia.
    assert 'href="/conversion"' not in respuesta.text


def test_ver_sistema_muestra_el_indicador_de_espacio_usado():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 12, "bytes_totales": 907397}),
        patch("app.main.listar_fotos_para_limpiar", return_value=[]),
    ):
        respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert "Sistema" in respuesta.text
    assert "12 fotos guardadas" in respuesta.text
    assert "886 KB" in respuesta.text


def test_ver_sistema_indicador_en_singular_con_una_sola_foto():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 1, "bytes_totales": 500}),
        patch("app.main.listar_fotos_para_limpiar", return_value=[]),
    ):
        respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert "1 foto guardada" in respuesta.text
    assert "fotos guardadas" not in respuesta.text


def test_ver_sistema_si_falla_el_uso_de_storage_no_muestra_el_indicador_ni_rompe():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_uso_storage_bucket", side_effect=Exception("permission denied for schema storage")),
        patch("app.main.listar_fotos_para_limpiar", side_effect=Exception("permission denied for schema storage")),
    ):
        respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert 'class="espacio-storage"' not in respuesta.text
    assert 'id="boton-limpiar-fotos-viejas"' not in respuesta.text


def test_ver_sistema_no_cuenta_las_fotos_viejas_solo_ofrece_revisarlas():
    # El conteo recorre todas las compras con foto: NO puede correr en
    # cada visita a /sistema por un numerito informativo — es bajo demanda.
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 12, "bytes_totales": 907397}),
        patch("app.main.listar_fotos_para_limpiar") as mock_listar,
    ):
        respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    mock_listar.assert_not_called()
    assert 'action="/sistema/revisar-fotos-viejas"' in respuesta.text
    assert 'id="boton-limpiar-fotos-viejas"' not in respuesta.text


def test_revisar_fotos_viejas_cuenta_bajo_demanda_y_ofrece_limpiar():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 12, "bytes_totales": 907397}),
        patch("app.main.listar_fotos_para_limpiar", return_value=["2020-01-01/x.jpg", "2020-02-02/y.jpg"]),
    ):
        respuesta = cliente.post("/sistema/revisar-fotos-viejas")

    assert respuesta.status_code == 200
    assert "2 fotos" in respuesta.text
    assert 'action="/sistema/limpiar-fotos-viejas"' in respuesta.text
    assert 'id="boton-limpiar-fotos-viejas"' in respuesta.text
    assert 'data-cantidad="2"' in respuesta.text


def test_ver_sistema_incluye_link_a_inicio():
    respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert 'href="/inicio"' in respuesta.text


def test_limpiar_fotos_viejas_borra_las_encontradas_y_limpia_foto_ruta():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fotos_para_limpiar", return_value=["2020-01-01/a.jpg", "2020-02-02/b.jpg"]),
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
        patch("app.main.limpiar_foto_ruta_de_compras") as mock_limpiar_ruta,
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 10, "bytes_totales": 500}),
    ):
        respuesta = cliente.post("/sistema/limpiar-fotos-viejas")

    assert respuesta.status_code == 200
    assert "Sistema" in respuesta.text
    assert "Se liberaron 2 fotos." in respuesta.text
    assert mock_borrar_foto.call_count == 2
    mock_borrar_foto.assert_any_call("2020-01-01/a.jpg")
    mock_borrar_foto.assert_any_call("2020-02-02/b.jpg")
    assert mock_limpiar_ruta.call_count == 2


def test_limpiar_fotos_viejas_sin_ninguna_para_borrar_no_rompe():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fotos_para_limpiar", return_value=[]),
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
    ):
        respuesta = cliente.post("/sistema/limpiar-fotos-viejas")

    assert respuesta.status_code == 200
    assert "No hay fotos de más de 3 años para borrar." in respuesta.text
    mock_borrar_foto.assert_not_called()


def test_limpiar_fotos_viejas_si_falla_una_sigue_con_las_demas():
    def borrar_side_effect(foto_ruta):
        if foto_ruta == "2020-01-01/a.jpg":
            raise Exception("Supabase Storage rechazó el borrado (404)")
        return None

    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch(
            "app.main.listar_fotos_para_limpiar",
            return_value=["2020-01-01/a.jpg", "2020-02-02/b.jpg"],
        ),
        patch("app.main.borrar_foto_comanda", side_effect=borrar_side_effect) as mock_borrar_foto,
        patch("app.main.limpiar_foto_ruta_de_compras") as mock_limpiar_ruta,
        patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 10, "bytes_totales": 500}),
    ):
        respuesta = cliente.post("/sistema/limpiar-fotos-viejas")

    assert respuesta.status_code == 200
    assert mock_borrar_foto.call_count == 2
    # Solo la que no falló llega a limpiar foto_ruta en la base.
    mock_limpiar_ruta.assert_called_once_with("2020-02-02/b.jpg")
    assert "Se liberaron 1 de 2 fotos" in respuesta.text


def test_limpiar_fotos_viejas_error_al_buscar_candidatas_da_500():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fotos_para_limpiar", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post("/sistema/limpiar-fotos-viejas")

    assert respuesta.status_code == 500
    assert "No se pudo revisar qué fotos limpiar" in respuesta.text


PROVEEDORES_DE_PRUEBA = [
    {"id": 200, "codigo_puesto": "N07P41", "nombre": "Saturno"},
    {"id": 201, "codigo_puesto": "L03P38", "nombre": "Frutamax"},
]

PROVEEDOR_DE_PRUEBA = {"id": 200, "codigo_puesto": "N07P41", "nombre": "Saturno"}

ARTICULOS_CON_UNIDAD_COMPRA = [
    {"id": 5, "nombre": "Kiwi", "unidad_compra": "kilo", "contenido_referencia": 18},
]

ARTICULO_KILO_DE_PRUEBA = {"id": 5, "nombre": "Kiwi", "unidad_compra": "kilo", "contenido_referencia": 18}
ARTICULO_UNIDAD_DE_PRUEBA = {"id": 6, "nombre": "Mango", "unidad_compra": "unidad", "contenido_referencia": 10}
ARTICULO_SIN_UNIDAD_COMPRA = {"id": 7, "nombre": "Kiwi", "unidad_compra": None, "contenido_referencia": None}


def test_ver_nueva_compra_sin_proveedor_muestra_formulario_de_proveedor():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva")

    assert respuesta.status_code == 200
    assert "Código de puesto" in respuesta.text
    assert "N07P41" in respuesta.text
    # Regresión: sin este cartel, leer la foto con la IA puede tardar varios
    # segundos sin ningún cambio visible y el comprador aprieta de nuevo.
    assert 'id="form-leer-comanda"' in respuesta.text
    assert "Leyendo comanda..." in respuesta.text
    # Regresión: en Safari/iOS, deshabilitar el botón sincrónicamente dentro
    # del evento submit cancela el envío del formulario.
    assert "setTimeout" in respuesta.text


def test_ver_nueva_compra_sin_proveedor_botones_confirmar_verde_y_cancelar_rojo():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva")

    assert respuesta.status_code == 200
    assert 'class="boton-exito" id="boton-confirmar-proveedor"' in respuesta.text
    assert 'class="boton boton-peligro" id="boton-cancelar"' in respuesta.text
    assert 'href="/compras"' in respuesta.text
    # Todavía no se cargó nada en esta pantalla: Cancelar no pide
    # confirmación, a diferencia del Cancelar de las pantallas siguientes.
    assert "confirm(" not in respuesta.text
    # Orden: Confirmar proveedor (verde) antes que Cancelar (rojo).
    assert respuesta.text.index('id="boton-confirmar-proveedor"') < respuesta.text.index('id="boton-cancelar"')


def test_ver_nueva_compra_muestra_autocompletado_en_los_dos_sentidos():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva")

    assert respuesta.status_code == 200
    # Código -> nombre (ya existía): el datalist del código y el objeto JS.
    assert 'list="lista_proveedores"' in respuesta.text
    assert "PROVEEDORES_CONOCIDOS" in respuesta.text
    assert 'oninput="sugerirNombre()"' in respuesta.text
    # Nombre -> código (nuevo): datalist propio del campo nombre, con los
    # nombres conocidos, y la función que arma el mapeo inverso.
    assert 'list="lista_nombres_proveedores"' in respuesta.text
    assert 'id="lista_nombres_proveedores"' in respuesta.text
    assert "Saturno" in respuesta.text
    assert "Frutamax" in respuesta.text
    assert 'oninput="sugerirCodigo()"' in respuesta.text
    assert "NOMBRES_A_CODIGOS" in respuesta.text
    assert "function sugerirCodigo" in respuesta.text


def test_ver_nueva_compra_sin_proveedor_error_de_base_da_500():
    with patch("app.main.listar_proveedores", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/nueva")

    assert respuesta.status_code == 500


def test_ver_nueva_compra_manual_muestra_proveedor_y_articulo_juntos():
    # UNA sola pantalla, como la carga de comandas: proveedor y artículo se
    # cargan juntos — se eliminó el paso intermedio de "confirmar proveedor".
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/nueva/manual")

    assert respuesta.status_code == 200
    assert "Código de puesto" in respuesta.text
    assert "N07P41" in respuesta.text
    assert "PROVEEDORES_LISTA" in respuesta.text
    # Los campos del artículo en la MISMA pantalla, con el mismo form.
    assert 'action="/compras/nueva/manual"' in respuesta.text
    assert "Kiwi" in respuesta.text
    assert "Cantidad de cajones" in respuesta.text
    # "Agregar más artículos" primero, "Guardar y cerrar" abajo de todo.
    assert "Agregar más artículos" in respuesta.text
    assert "Guardar y cerrar" in respuesta.text
    assert respuesta.text.index('id="boton-agregar-articulo"') < respuesta.text.index('id="boton-terminar-carga"')
    # No mezcla el flujo de foto: sin formulario de subida ni el link a
    # múltiples comandas.
    assert 'id="form-leer-comanda"' not in respuesta.text
    assert 'href="/compras/nueva/fotos"' not in respuesta.text


def test_ver_nueva_compra_manual_error_de_base_da_500():
    with patch("app.main.listar_proveedores", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/nueva/manual")

    assert respuesta.status_code == 500


def test_ver_nueva_compra_foto_muestra_solo_el_formulario_de_una_foto():
    respuesta = cliente.get("/compras/nueva/foto-una")

    assert respuesta.status_code == 200
    assert 'id="form-leer-comanda"' in respuesta.text
    assert "Leer una comanda" in respuesta.text
    assert 'action="/compras/nueva/foto"' in respuesta.text
    # Sigue ofreciendo el atajo a múltiples fotos desde acá.
    assert 'href="/compras/nueva/fotos"' in respuesta.text
    # No mezcla el flujo manual: sin el formulario de proveedor a mano.
    assert "Código de puesto" not in respuesta.text
    assert 'action="/compras/nueva/proveedor"' not in respuesta.text


def test_ver_nueva_compra_foto_un_solo_toque_dispara_selector_y_lee_directo():
    # Regresión: el botón grande no es un submit — abre el input file oculto
    # (que sí tiene el required, por si algo falla) y la lectura arranca
    # sola al elegir la foto, sin un segundo toque en un botón "Leer".
    respuesta = cliente.get("/compras/nueva/foto-una")

    assert respuesta.status_code == 200
    assert 'class="input-foto-oculto"' in respuesta.text
    assert 'type="button" id="boton-leer-comanda"' in respuesta.text
    assert 'botonLeerComanda.addEventListener("click", function () {\n        inputFoto.click();' in respuesta.text
    assert 'inputFoto.addEventListener("change"' in respuesta.text
    assert "formLeerComanda.requestSubmit()" in respuesta.text


def test_cambiar_proveedor_en_carga_manual_apunta_a_la_pantalla_manual():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    assert 'href="/compras/nueva/manual"' in respuesta.text


def test_ver_cargar_listado_de_compras_muestra_la_pantalla():
    respuesta = cliente.get("/compras/nueva/listado")

    assert respuesta.status_code == 200
    assert "Cargar listado de compras" in respuesta.text
    # Un solo toque: el botón abre el selector nativo, no hace falta un
    # segundo botón "Leer" — mismo criterio que en Múltiples comandas.
    assert 'class="input-foto-oculto"' in respuesta.text
    assert 'id="input-foto-listado"' in respuesta.text
    assert '>Subir plantilla compras<' in respuesta.text
    assert "inputFotoListado.click();" in respuesta.text
    assert 'inputFotoListado.addEventListener("change"' in respuesta.text
    assert 'id="boton-guardar-siguiente"' in respuesta.text
    assert 'id="boton-descartar-siguiente"' in respuesta.text
    assert 'id="boton-cancelar-multiple"' in respuesta.text
    # Regresión: mientras la IA lee la planilla (puede tardar, es una foto
    # con varios proveedores), tiene que quedar clarísimo que el sistema
    # está trabajando — spinner + cartel, no un cartelito chico gris — y el
    # botón no se puede volver a tocar (queda tapado Y deshabilitado).
    assert 'id="seccion-leyendo"' in respuesta.text
    assert 'class="spinner"' in respuesta.text
    assert "Leyendo la planilla..." in respuesta.text
    assert 'document.getElementById("boton-elegir-foto-listado").disabled = true;' in respuesta.text
    assert 'document.getElementById("seccion-leyendo").style.display = "flex";' in respuesta.text


def test_ver_cargar_listado_de_compras_incluye_los_proveedores_conocidos_para_sugerir():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva/listado")

    assert respuesta.status_code == 200
    assert '{ codigo: "N07P41", nombre: "Saturno" }' in respuesta.text
    assert '{ codigo: "L03P38", nombre: "Frutamax" }' in respuesta.text


# --- /compras/buscar: búsqueda de compras por rango de fechas, proveedor y artículo opcionales ---

COMPRAS_BUSQUEDA_DE_PRUEBA = [
    {
        "id": 1,
        "fecha_operacion": HOY_DE_PRUEBA,
        "articulo_nombre": "Tomate Cherry",
        "unidad_compra": "kilo",
        "proveedor_nombre": "Saturno",
        "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 40,
        "contenido_por_cajon": 20,
        "cantidad_kilos": 800,
        "cantidad_fraccion": None,
        "importe": 45000.0,
        "sena": None,
        "tipo_retiro": "Clark",
    },
    {
        "id": 2,
        "fecha_operacion": HOY_DE_PRUEBA - timedelta(days=1),
        "articulo_nombre": "Mango",
        "unidad_compra": "unidad",
        "proveedor_nombre": "Frutamax",
        "proveedor_codigo_puesto": "L03P38",
        "cantidad_cajones": 10,
        "contenido_por_cajon": 12,
        "cantidad_kilos": None,
        "cantidad_fraccion": 120,
        "importe": None,
        "sena": None,
        "tipo_retiro": "Carro",
    },
]


def test_ver_buscar_compras_sin_filtros_usa_las_ultimas_48hs():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA) as mock_buscar,
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 200
    mock_buscar.assert_called_once_with(
        HOY_DE_PRUEBA - timedelta(days=1), HOY_DE_PRUEBA, None, None, limite=TOPE_FILAS_BUSQUEDA + 1
    )
    assert f'value="{(HOY_DE_PRUEBA - timedelta(days=1)).isoformat()}"' in respuesta.text
    assert f'value="{HOY_DE_PRUEBA.isoformat()}"' in respuesta.text


def test_buscar_compras_cortada_por_el_tope_avisa_con_el_total():
    # Un rango ancho no puede tirar miles de filas al celular: se cortan
    # en el tope y el aviso dice cuántas había en total.
    muchas = [dict(COMPRAS_BUSQUEDA_DE_PRUEBA[0], id=i) for i in range(TOPE_FILAS_BUSQUEDA + 1)]
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=muchas),
        patch("app.main.contar_compras_buscadas", return_value=1234) as mock_contar,
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 200
    assert f"Se muestran las primeras {TOPE_FILAS_BUSQUEDA} compras de 1234" in respuesta.text
    mock_contar.assert_called_once()


def test_buscar_compras_sin_corte_no_cuenta_ni_avisa():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA),
        patch("app.main.contar_compras_buscadas") as mock_contar,
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 200
    assert "Se muestran las primeras" not in respuesta.text
    # El conteo extra solo se paga cuando de verdad hubo corte.
    mock_contar.assert_not_called()


def test_ver_buscar_compras_con_filtros_de_fecha_proveedor_y_articulo():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA) as mock_buscar,
    ):
        respuesta = cliente.get(
            "/compras/buscar?fecha_desde=2026-08-01&fecha_hasta=2026-08-06&proveedor_id=200&articulo_id=5"
        )

    assert respuesta.status_code == 200
    mock_buscar.assert_called_once_with(date(2026, 8, 1), date(2026, 8, 6), 200, 5, limite=TOPE_FILAS_BUSQUEDA + 1)
    # Los buscadores combinados quedan precargados con el nombre elegido.
    assert 'value="Saturno"' in respuesta.text
    assert 'value="Kiwi"' in respuesta.text
    assert "✕ Ver todos los proveedores" in respuesta.text
    assert "✕ Ver todos los artículos" in respuesta.text


def test_ver_buscar_compras_muestra_contador_y_tabla():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert "2 compras encontradas" in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "Saturno (N07P41)" in respuesta.text
    assert "40 cajones × 20k" in respuesta.text
    assert "$45.000" in respuesta.text
    # Sin importe -> "SIN PRECIO" en rojo, mismo criterio que Últimas Compras.
    assert "SIN PRECIO" in respuesta.text


def test_ver_buscar_compras_el_link_editar_lleva_los_filtros_activos():
    # Para que Guardar/Volver de la edición vuelvan a ESTA búsqueda y no a
    # la default de 48hs.
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA),
    ):
        respuesta = cliente.get("/compras/buscar?fecha_desde=2026-07-28&fecha_hasta=2026-07-29&proveedor_id=200")

    assert respuesta.status_code == 200
    assert "/editar?fecha_desde=2026-07-28&fecha_hasta=2026-07-29&proveedor_id=200" in respuesta.text


def test_pantallas_de_carga_multiple_cancelan_al_hub():
    # Cancelar = salir al hub /compras; terminar de cargar sí va a Buscar
    # (para revisar de un vistazo lo que la IA leyó).
    with patch("app.main.listar_proveedores", return_value=[]):
        multiples = cliente.get("/compras/nueva/fotos")
        listado = cliente.get("/compras/nueva/listado")

    for respuesta in (multiples, listado):
        assert respuesta.status_code == 200
        # El botón Cancelar va al hub...
        assert 'window.location = "/compras";' in respuesta.text
        # ...y el fin de la cola (terminaste todo) sigue yendo a Buscar.
        assert 'window.location = "/compras/buscar";' in respuesta.text


def test_ver_buscar_compras_sin_resultados_muestra_mensaje_y_no_boton_exportar():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert "0 compras encontradas" in respuesta.text
    assert "No se encontraron compras con estos filtros." in respuesta.text
    assert 'id="boton-exportar"' not in respuesta.text


def test_ver_buscar_compras_con_resultados_muestra_boton_exportar():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA),
    ):
        respuesta = cliente.get("/compras/buscar?fecha_desde=2026-08-01&fecha_hasta=2026-08-06")

    assert 'id="boton-exportar"' in respuesta.text
    assert "/compras/buscar/exportar-pdf?fecha_desde=2026-08-01&fecha_hasta=2026-08-06" in respuesta.text
    assert "/compras/buscar/exportar-excel?fecha_desde=2026-08-01&fecha_hasta=2026-08-06" in respuesta.text


def test_ver_buscar_compras_fecha_invalida_muestra_error_y_usa_default():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]) as mock_buscar,
    ):
        respuesta = cliente.get("/compras/buscar?fecha_desde=no-es-una-fecha")

    assert respuesta.status_code == 200
    assert "La fecha desde no es válida." in respuesta.text
    mock_buscar.assert_called_once_with(
        HOY_DE_PRUEBA - timedelta(days=1), HOY_DE_PRUEBA, None, None, limite=TOPE_FILAS_BUSQUEDA + 1
    )


def test_ver_buscar_compras_fecha_desde_posterior_a_hasta_muestra_error():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar?fecha_desde=2026-08-10&fecha_hasta=2026-08-01")

    assert "La fecha desde no puede ser posterior a la fecha hasta." in respuesta.text


def test_ver_buscar_compras_incluye_buscadores_combinados():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert '{ id: 200, codigo: "N07P41", nombre: "Saturno" }' in respuesta.text
    assert '{ id: 5, nombre: "Kiwi" }' in respuesta.text
    assert "actualizarListaProveedores" in respuesta.text
    assert "actualizarListaArticulos" in respuesta.text


def test_ver_buscar_compras_error_de_base_da_500():
    with patch("app.main.listar_proveedores", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 500


# --- /compras/buscar/exportar-pdf y exportar-excel ---


def test_exportar_listado_compras_pdf_devuelve_archivo_adjunto():
    with patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA) as mock_buscar:
        respuesta = cliente.get(
            "/compras/buscar/exportar-pdf?fecha_desde=2026-08-01&fecha_hasta=2026-08-06&proveedor_id=200&articulo_id=5"
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert "Listado_Compras_2026-08-01_a_2026-08-06" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"%PDF")
    # El export va SIN tope: un archivo cortado en silencio sería peor
    # que uno pesado.
    mock_buscar.assert_called_once_with(date(2026, 8, 1), date(2026, 8, 6), 200, 5)


def test_exportar_listado_compras_pdf_agrupa_por_fecha_y_proveedor():
    with patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA):
        respuesta = cliente.get("/compras/buscar/exportar-pdf?fecha_desde=2026-08-01&fecha_hasta=2026-08-06")

    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Listado de Compras" in texto
    # La fecha más reciente (HOY_DE_PRUEBA) va antes que la del día anterior.
    assert texto.index(HOY_DE_PRUEBA.strftime("%d/%m/%Y")) < texto.index(
        (HOY_DE_PRUEBA - timedelta(days=1)).strftime("%d/%m/%Y")
    )
    assert "Saturno (N07P41)" in texto
    assert "Frutamax (L03P38)" in texto


def test_exportar_listado_compras_pdf_fecha_invalida_da_400():
    respuesta = cliente.get("/compras/buscar/exportar-pdf?fecha_desde=no-es-fecha&fecha_hasta=2026-08-06")

    assert respuesta.status_code == 400


def test_exportar_listado_compras_excel_devuelve_archivo_adjunto():
    with patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA):
        respuesta = cliente.get("/compras/buscar/exportar-excel?fecha_desde=2026-08-01&fecha_hasta=2026-08-06")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip


def test_exportar_listado_compras_excel_fecha_invalida_da_400():
    respuesta = cliente.get("/compras/buscar/exportar-excel?fecha_desde=2026-08-01&fecha_hasta=no-es-fecha")

    assert respuesta.status_code == 400


def test_armar_listado_de_compras_muestra_en_construccion():
    respuesta = cliente.get("/compras/armar-listado")

    assert respuesta.status_code == 200
    assert "Armar listado de compras" in respuesta.text
    assert "En construcción" in respuesta.text


def test_ver_compras_muestra_la_botonera_de_cargar_y_operaciones():
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    # Grupo Cargar.
    assert 'href="/compras/nueva/manual"' in respuesta.text
    assert "Cargar Manual" in respuesta.text
    assert 'href="/compras/nueva/foto-una"' in respuesta.text
    assert "Cargar Foto" in respuesta.text
    assert 'href="/compras/nueva/fotos"' in respuesta.text
    assert "Cargar Múltiples Fotos" in respuesta.text
    assert 'href="/compras/nueva/listado"' in respuesta.text
    assert "Cargar Listado de Compras" in respuesta.text
    # Grupo Operaciones: Buscar Compras, Armar Listado, Compras sin precio,
    # Disponibles. "Últimas Compras" y "Enviar a Logística" ya no existen
    # (Buscar Compras + Logística los reemplazan).
    assert 'href="/compras/ultimas"' not in respuesta.text
    assert "Últimas Compras" not in respuesta.text
    assert 'href="/compras/buscar"' in respuesta.text
    assert 'href="/compras/armar-listado"' in respuesta.text
    assert 'href="/compras/pendientes"' in respuesta.text
    assert "Compras sin precio" in respuesta.text
    assert 'href="/compras/disponibles"' in respuesta.text
    assert "Disponibles" in respuesta.text
    assert 'href="/compras/enviar-logistica"' not in respuesta.text
    # El botón viejo "Agregar compra" queda retirado (lo cubre Cargar Manual).
    assert "Agregar compra" not in respuesta.text
    assert 'href="/compras/nueva"' not in respuesta.text


def test_ver_compras_botonera_diferencia_grupos_por_color():
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    # Cargar: azul, el color de acción normal (sin clase extra de color).
    assert 'class="boton" href="/compras/nueva/manual"' in respuesta.text
    # Operaciones: naranja — ni verde (=guardar) ni rojo (=borrar/cancelar).
    assert 'class="boton boton-naranja" href="/compras/buscar"' in respuesta.text
    assert 'class="boton boton-naranja" href="/compras/pendientes"' in respuesta.text
    assert ".boton-naranja { background: #ea580c; }" in respuesta.text
    # "Próximamente": color del grupo pero atenuado (mismo criterio en
    # Cargar y en Operaciones). Cargar Listado de Compras, Buscar Compras y
    # Disponibles ya no son "próximamente" — quedaron activos.
    assert 'class="boton" href="/compras/nueva/listado"' in respuesta.text
    assert 'class="boton boton-naranja" href="/compras/disponibles"' in respuesta.text
    assert "Disponibles (Próximamente)" not in respuesta.text
    assert ".boton-proximamente { opacity: 0.6; }" in respuesta.text


def test_ver_buscar_compras_boton_borrar_seleccionadas_es_tamano_normal():
    # Regresión: el botón no puede depender solo de .boton-eliminar (que no
    # trae padding/ancho propios) — necesita la clase .boton para tener un
    # área de toque cómoda en celular, no quedar chico/finito.
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_BUSQUEDA_DE_PRUEBA),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 200
    assert 'class="boton boton-eliminar" id="boton-borrar-seleccionadas"' in respuesta.text


def test_ver_buscar_compras_muestra_editar_y_ver_foto_solo_con_foto():
    compras = [
        dict(COMPRAS_BUSQUEDA_DE_PRUEBA[0], tiene_fotos=True),
        dict(COMPRAS_BUSQUEDA_DE_PRUEBA[1], tiene_fotos=False),
    ]
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=compras),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.text.count(">Editar<") == 2
    assert respuesta.text.count("Ver foto") == 1
    assert 'href="/compras/1/foto"' in respuesta.text
    assert 'href="/compras/2/foto"' not in respuesta.text
    assert 'href="/compras/1/editar?' in respuesta.text
    assert 'href="/compras/2/editar?' in respuesta.text
    # El botón Detalle aparece siempre, al lado de Editar, sin importar la foto.
    assert respuesta.text.count(">Detalle<") == 2
    assert 'href="/compras/1/detalle"' in respuesta.text
    assert 'href="/compras/2/detalle"' in respuesta.text


def test_ver_buscar_compras_muestra_el_aviso_cuando_viene_en_la_url():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar?aviso=3+compras+canceladas.")

    assert '<div class="aviso">3 compras canceladas.</div>' in respuesta.text


def _datos_compra_manual(**overrides):
    datos = {
        "codigo_puesto": "n07p41",
        "nombre": "Saturno",
        "accion": "agregar",
        "articulo_id": "5",
        "cantidad_cajones": "8",
        "contenido_por_cajon": "18",
        "importe": "3000",
        "sena": "",
        "tipo_retiro": "Clark",
    }
    datos.update(overrides)
    return datos


def test_agregar_compra_manual_guarda_proveedor_y_articulo_en_un_paso():
    hoy = date(2026, 8, 22)
    with (
        patch("app.main._hoy_argentina", return_value=hoy),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post("/compras/nueva/manual", data=_datos_compra_manual(), follow_redirects=False)

    assert respuesta.status_code == 303
    # Sigue en la pantalla clásica con el proveedor YA fijado: los
    # siguientes artículos no lo repiten.
    assert respuesta.headers["location"] == "/compras/nueva?proveedor_id=200"
    mock_proveedor.assert_called_once_with("N07P41", "Saturno")
    mock_crear.assert_called_once_with(hoy, 5, 200, 8.0, 18.0, 144.0, None, 3000.0, None, "Clark", None)


def test_agregar_compra_manual_con_guardar_termina_en_buscar():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual", data=_datos_compra_manual(accion="terminar"), follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_crear.assert_called_once()


def test_agregar_compra_manual_codigo_invalido_repuebla_todo_lo_tipeado():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post("/compras/nueva/manual", data=_datos_compra_manual(codigo_puesto="puesto15"))

    assert respuesta.status_code == 400
    assert "formato NNNPNN" in respuesta.text
    mock_proveedor.assert_not_called()
    mock_crear.assert_not_called()
    # Lo tipeado NO se pierde: el proveedor y el renglón vuelven cargados.
    assert 'value="puesto15"' in respuesta.text
    assert 'value="Saturno"' in respuesta.text
    assert 'value="8"' in respuesta.text
    assert 'value="3000"' in respuesta.text


def test_agregar_compra_manual_sin_nombre_muestra_error():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post("/compras/nueva/manual", data=_datos_compra_manual(nombre=""))

    assert respuesta.status_code == 400
    assert "El nombre no puede estar vacío" in respuesta.text
    mock_proveedor.assert_not_called()


def test_agregar_compra_manual_renglon_invalido_no_crea_el_proveedor():
    # El proveedor se resuelve recién con el renglón validado: un error de
    # tipeo en la compra no deja proveedores nuevos colgados.
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post("/compras/nueva/manual", data=_datos_compra_manual(articulo_id=""))

    assert respuesta.status_code == 400
    assert "Elegí un artículo" in respuesta.text
    mock_proveedor.assert_not_called()


def test_agregar_compra_manual_terminar_sin_nada_cargado_sale_sin_guardar():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data={"codigo_puesto": "N07P41", "nombre": "Saturno", "accion": "terminar", "articulo_id": "",
                  "cantidad_cajones": "", "contenido_por_cajon": "", "importe": "", "sena": "", "tipo_retiro": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_proveedor.assert_not_called()
    mock_crear.assert_not_called()


def test_agregar_compra_manual_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post("/compras/nueva/manual", data=_datos_compra_manual())

    assert respuesta.status_code == 500
    assert "No se pudo guardar el proveedor" in respuesta.text


def test_ver_nueva_compra_con_proveedor_muestra_formulario_de_renglon():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "N07P41" in respuesta.text
    assert "Kiwi" in respuesta.text
    assert "Cantidad de cajones" in respuesta.text
    assert "Contenido por cajón" in respuesta.text
    # Regresión: sin formnovalidate, el navegador bloquea "Terminar carga"
    # cuando el renglón está vacío (los campos "*" son required) y el POST
    # ni siquiera llega al servidor.
    assert 'name="accion" value="terminar" formnovalidate' in respuesta.text
    # Regresión: sin este cartel de "Guardando...", el comprador no ve
    # ningún cambio al apretar y termina cargando la compra duplicada.
    assert 'id="form-compra"' in respuesta.text
    assert "Guardando..." in respuesta.text
    # Regresión: en Safari/iOS, deshabilitar el botón sincrónicamente dentro
    # del evento submit cancela el envío del formulario.
    assert "setTimeout" in respuesta.text


def test_ver_nueva_compra_con_proveedor_botones_en_orden_agregar_guardar_cancelar():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    # Regresión: los 3 botones de abajo de todo, en este orden exacto y con
    # estos colores — "Agregar más artículos" primero (el gesto repetido) y
    # "Guardar y cerrar" abajo de todo, antes de Cancelar. Solo "Cancelar"
    # pide confirmación. Se busca por id, no por el texto visible.
    orden = ['id="boton-agregar-articulo"', 'id="boton-terminar-carga"', 'id="boton-cancelar-carga"']
    posiciones = [respuesta.text.index(texto) for texto in orden]
    assert posiciones == sorted(posiciones)
    assert "Agregar más artículos" in respuesta.text
    assert "Guardar y cerrar" in respuesta.text
    assert 'class="boton-exito" id="boton-terminar-carga"' in respuesta.text
    assert 'class="boton-peligro" id="boton-cancelar-carga"' in respuesta.text
    assert 'action="/compras/nueva/cancelar"' in respuesta.text
    assert "confirm('¿Seguro que querés cancelar? Se va a descartar todo lo cargado de este proveedor')" in respuesta.text
    assert 'name="proveedor_id" value="200"' in respuesta.text


def test_ver_nueva_compra_con_proveedor_muestra_cargado_hoy_con_formato_compacto():
    renglones_hoy = [
        {
            "id": 99,
            "articulo_nombre": "Kiwi",
            "cantidad_cajones": 10,
            "contenido_por_cajon": 18.6,
            "importe": 45000,
            "sena": None,
        },
    ]
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=renglones_hoy),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    # Regresión: la tabla de "Cargado hoy" no tenía NINGÚN filtro aplicado
    # (mostraba los números pelados, e importe/seña en None literal texto
    # "None" cuando la compra no tenía precio todavía).
    assert "<td>19</td>" in respuesta.text  # contenido_por_cajon redondeado (18.6 -> 19)
    assert "$45.000" in respuesta.text
    assert "None" not in respuesta.text


def test_ver_nueva_compra_con_proveedor_inexistente_da_404():
    with patch("app.main.obtener_proveedor", return_value=None):
        respuesta = cliente.get("/compras/nueva?proveedor_id=999")

    assert respuesta.status_code == 404


def test_ver_nueva_compra_con_proveedor_error_de_base_da_500():
    with patch("app.main.obtener_proveedor", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 500


def test_agregar_compra_exitosa_redirige_al_mismo_proveedor_calcula_kilos():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/nueva?proveedor_id=200"
    # 10 cajones × 18 kg = 180 kg (unidad_compra del artículo = kilo)
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 5, 200, 10.0, 18.0, 180.0, None, 50000.0, None, "Clark", None)


def test_agregar_compra_calcula_fraccion_para_articulo_por_unidad():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_UNIDAD_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "6",
                "cantidad_cajones": "5",
                "contenido_por_cajon": "10",
                "importe": "30000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # 5 cajones × 10 unidades = 50 unidades (unidad_compra del artículo = unidad)
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 6, 200, 5.0, 10.0, None, 50.0, 30000.0, None, "Carro", None)


def test_agregar_compra_proveedor_inexistente_da_404():
    with patch("app.main.obtener_proveedor", return_value=None):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "999",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 404


def test_agregar_compra_terminar_con_renglon_vacio_va_a_compras_sin_guardar():
    with (
        patch("app.main.obtener_proveedor") as mock_proveedor,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "",
                "cantidad_cajones": "",
                "contenido_por_cajon": "",
                "importe": "",
                "sena": "",
                "tipo_retiro": "",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_proveedor.assert_not_called()
    mock_crear.assert_not_called()


def test_agregar_compra_terminar_con_renglon_cargado_lo_guarda_y_va_a_compras():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 5, 200, 10.0, 18.0, 180.0, None, 50000.0, None, "Clark", None)


def test_agregar_compra_terminar_con_renglon_invalido_muestra_error_y_no_pierde_datos():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "Elegí un artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_sin_articulo_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "Elegí un artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_sin_cantidad_cajones_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "La cantidad de cajones es obligatoria" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_sin_contenido_por_cajon_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "El contenido por cajón es obligatorio" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_sin_importe_queda_pendiente():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "",
                "sena": "",
                "tipo_retiro": "Clark",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 5, 200, 10.0, 18.0, 180.0, None, None, None, "Clark", None)


def test_agregar_compra_importe_negativo_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "-5",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "El importe tiene que ser mayor a cero" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_tipo_retiro_invalido_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Camion",
            },
        )

    assert respuesta.status_code == 400
    assert "tipo de retiro válido" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_tipo_retiro_pases_se_acepta():
    # "Pases" es el tercer tipo de retiro, junto con Clark/Carro.
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Pases",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 5, 200, 10.0, 18.0, 180.0, None, 50000.0, None, "Pases", None)


def test_ver_nueva_compra_con_proveedor_muestra_las_tres_opciones_de_retiro():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    assert '<option value="Clark" selected>' in respuesta.text
    assert '<option value="Carro"' in respuesta.text
    assert '<option value="Pases"' in respuesta.text


def test_agregar_compra_articulo_sin_unidad_compra_configurada_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_SIN_UNIDAD_COMPRA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "7",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "no tiene la unidad de compra configurada" in respuesta.text
    mock_crear.assert_not_called()


def test_agregar_compra_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la compra" in respuesta.text


def test_cancelar_carga_proveedor_borra_todo_lo_de_hoy_y_va_al_hub():
    # Cancelar = salir: vuelve al hub /compras con el aviso, no a una
    # búsqueda que nadie pidió (terminar de cargar sí va a Buscar).
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.eliminar_compras_del_dia_por_proveedor", return_value={"borradas": 3, "protegidas": 0}) as mock_eliminar,
    ):
        respuesta = cliente.post(
            "/compras/nueva/cancelar",
            data={"proveedor_id": "200"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/compras?")
    assert "buscar" not in location
    assert "aviso=3+compras+canceladas." in location
    # Se borra TODO lo del proveedor en el día, no un renglón puntual: no
    # hace falta que el comprador haya cargado nada en el formulario para
    # que se descarte lo que ya estaba guardado de "Agregar artículo".
    mock_eliminar.assert_called_once_with(HOY_DE_PRUEBA, 200)


def test_cancelar_carga_proveedor_con_protegidas_avisa_sin_tecnicismos():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.eliminar_compras_del_dia_por_proveedor", return_value={"borradas": 3, "protegidas": 2}),
    ):
        respuesta = cliente.post(
            "/compras/nueva/cancelar",
            data={"proveedor_id": "200"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert "aviso=3+compras+canceladas.+2+no+se+pudieron+eliminar%3A+ya+fueron+retiradas+o+recepcionadas." in location


def test_cancelar_carga_proveedor_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.eliminar_compras_del_dia_por_proveedor", side_effect=Exception("no se pudo conectar")),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post("/compras/nueva/cancelar", data={"proveedor_id": "200"})

    assert respuesta.status_code == 500
    assert "No se pudo cancelar" in respuesta.text
    # No pierde de vista al proveedor: sigue mostrando su formulario, no una
    # página de error genérica.
    assert "Saturno" in respuesta.text


COMPRA_DE_PRUEBA = {
    "id": 30,
    "fecha_operacion": HOY_DE_PRUEBA,
    "articulo_id": 5,
    "articulo_nombre": "Mzn Red",
    "proveedor_id": 200,
    "proveedor_nombre": "Saturno",
    "proveedor_codigo_puesto": "N07P41",
    "cantidad_cajones": 10,
    "contenido_por_cajon": 18,
    "cantidad_kilos": 180,
    "cantidad_fraccion": None,
    "importe": 50000,
    "sena": None,
    "tipo_retiro": "Clark",
    "estado": "pendiente",
    "estado_retiro": "pendiente",
}

COMPRA_DETALLE_DE_PRUEBA = {
    "id": 30,
    "fecha_operacion": HOY_DE_PRUEBA,
    "cargado_el": datetime(2026, 8, 16, 12, 15, tzinfo=timezone.utc),
    "articulo_id": 5,
    "articulo_nombre": "Mzn Red",
    "unidad_compra": "kg",
    "proveedor_id": 200,
    "proveedor_nombre": "Saturno",
    "proveedor_codigo_puesto": "N07P41",
    "guia_id": 105,
    "guia_punto": 2,
    "cantidad_cajones": 10,
    "contenido_por_cajon": 18,
    "importe": 50000,
    "sena": None,
    "tipo_retiro": "Clark",
    "estado_retiro": "retirado",
    "retiro_procesado_el": datetime(2026, 8, 16, 13, 0, tzinfo=timezone.utc),
    "retiro_origen": "logistica",
    "cantidad_cajones_retirada": None,
    "estado": "recepcionado",
    "procesada_el": datetime(2026, 8, 16, 14, 0, tzinfo=timezone.utc),
    "cantidad_cajones_real": 10,
    "contenido_por_cajon_real": 18,
    "cantidad_fraccion_real": None,
    "cantidad_cajones_rechazada": None,
    "motivo_rechazo": None,
}


def test_ver_editar_compra_muestra_datos_precargados():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "N07P41" in respuesta.text
    assert 'action="/compras/30/editar"' in respuesta.text
    # Regresión: mostrar cajones y contenido por cajón por separado (no el
    # kilaje total), para poder corregir uno solo de los dos sin recalcular.
    assert 'id="cantidad_cajones" name="cantidad_cajones"' in respuesta.text
    assert 'value="10"' in respuesta.text
    assert 'id="contenido_por_cajon" name="contenido_por_cajon"' in respuesta.text
    assert 'value="18"' in respuesta.text


def test_ver_editar_compra_inexistente_da_404():
    with patch("app.main.obtener_compra", return_value=None):
        respuesta = cliente.get("/compras/999/editar")

    assert respuesta.status_code == 404


def test_ver_editar_compra_error_de_base_da_500():
    with patch("app.main.obtener_compra", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 500


def test_ver_detalle_compra_muestra_toda_la_historia():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "N07P41" in respuesta.text
    assert "Mzn Red" in respuesta.text
    assert "105.2" in respuesta.text
    assert "Retirado" in respuesta.text
    assert "Retirado por Logística" in respuesta.text
    assert "Recibido" in respuesta.text
    assert "Esta guía todavía no tiene fotos" in respuesta.text


def test_ver_detalle_compra_marca_la_diferencia_de_cajones_retirados():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, cantidad_cajones_retirada=8)
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Diferencia contra lo comprado" in respuesta.text


def test_ver_detalle_compra_marca_la_diferencia_de_recepcion():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, cantidad_cajones_real=9)
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Diferencia contra lo comprado" in respuesta.text


def test_ver_detalle_compra_con_fotos_muestra_la_galeria_de_la_guia():
    fotos = [
        {"id": 9, "foto_ruta": "2026-08-16/a.jpg", "creado_en": datetime(2026, 8, 16, 10, 0)},
        {"id": 10, "foto_ruta": "2026-08-16/b.pdf", "creado_en": datetime(2026, 8, 16, 11, 0)},
    ]
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.listar_fotos_de_guia", return_value=fotos),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    # Miniatura de la imagen y ficha del PDF, cada una con su link y su Borrar.
    assert 'src="/compras/30/fotos/9/ver"' in respuesta.text
    assert 'href="/compras/30/fotos/10/ver"' in respuesta.text
    assert 'action="/compras/30/fotos/9/borrar"' in respuesta.text
    # Y el botón para sumar otra (nunca reemplaza).
    assert 'action="/compras/30/fotos"' in respuesta.text
    assert "Agregar foto o archivo" in respuesta.text
    assert "Esta guía todavía no tiene fotos" not in respuesta.text


def test_ver_detalle_compra_inexistente_da_404():
    with (
        patch("app.main.obtener_detalle_compra", return_value=None),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/999/detalle")

    assert respuesta.status_code == 404


def test_ver_detalle_compra_error_de_base_da_500():
    with patch("app.main.obtener_detalle_compra", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 500


def test_ver_detalle_compra_ingreso_directo_muestra_etiqueta_propia():
    # Nunca 'deposito' (auto-retiro de algo que sí pasó por el puesto del
    # Mercado): 'ingreso_directo' es otra cosa, tiene su propia etiqueta.
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, retiro_origen="ingreso_directo")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Ingreso directo en Depósito" in respuesta.text


def test_ver_detalle_compra_recepcionada_muestra_boton_corregir_recepcion():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert 'href="/gerencia/compras/30/corregir-recepcion"' in respuesta.text


def test_ver_detalle_compra_no_recepcionada_no_muestra_boton_corregir_recepcion():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, estado="pendiente")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert 'href="/gerencia/compras/30/corregir-recepcion"' not in respuesta.text


def test_ver_detalle_compra_con_rechazo_parcial_muestra_el_registro():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, cantidad_cajones_real=8, cantidad_cajones_rechazada=2, motivo_rechazo="podrido")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Rechazo parcial: 2 bultos devueltos al proveedor — podrido" in respuesta.text
    assert "los aceptados" in respuesta.text


def test_ver_detalle_compra_sin_rechazo_parcial_no_muestra_el_registro():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert respuesta.status_code == 200
    assert "Rechazo parcial" not in respuesta.text


def test_ver_detalle_compra_muestra_el_aviso_cuando_viene_en_la_url():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle?aviso=Se+corrigi%C3%B3+la+recepci%C3%B3n+de+esta+compra.")

    assert respuesta.status_code == 200
    assert '<div class="aviso">Se corrigió la recepción de esta compra.</div>' in respuesta.text


@pytest.fixture(autouse=True)
def _puerta_de_gerencia_abierta(request):
    """Los tests de Corregir Recepción entran con la clave puesta.

    La pantalla vive en Gerencia y ESCRIBE, así que sin CLAVE_GERENCIA
    configurada no deja pasar (ver _puerta_de_gerencia_para_escribir). Esa
    puerta tiene sus propios tests, abajo; estos prueban la pantalla, así
    que la cruzan como la cruza una persona: con la clave y su cookie.
    """
    nombre = request.node.name.lower()
    # Los tests de LA PUERTA se cruzan solos: si esta fixture les pusiera la
    # cookie, probarían la pantalla en vez del control.
    if "corregir_recepcion" not in nombre or "clave" in nombre or "url_vieja" in nombre:
        yield
        return
    from app.main import _firma_acceso_gerencia
    # La lista de dependencias sale de la base: por default, lote sin usar.
    # Los tests que la miran la parchean ellos con lo que necesitan.
    with (
        patch("app.main._clave_gerencia", return_value="secreta"),
        patch("app.main.dependencias_del_lote_de_compra", return_value=None),
        patch("app.main.listar_clientes", return_value=[]),
    ):
        cliente.cookies.set("acceso_gerencia", _firma_acceso_gerencia("secreta"))
        try:
            yield
        finally:
            cliente.cookies.clear()


def test_ver_corregir_recepcion_compra_muestra_formulario_precargado():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, unidad_compra="kilo")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "Mzn Red" in respuesta.text
    assert 'action="/gerencia/compras/30/corregir-recepcion"' in respuesta.text
    assert 'id="cajones-real"' in respuesta.text
    assert 'value="10"' in respuesta.text
    # Por kilo: precarga con contenido_por_cajon_real (kilos de UN bulto).
    assert 'id="total-real"' in respuesta.text
    assert 'value="18"' in respuesta.text


def test_ver_corregir_recepcion_compra_por_unidad_precarga_por_cajon_no_el_total():
    compra = dict(
        COMPRA_DETALLE_DE_PRUEBA,
        unidad_compra="unidad",
        cantidad_cajones_real=30,
        contenido_por_cajon_real=3,
        cantidad_fraccion_real=90,
    )
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 200
    # Por unidad/cubeta también precarga con contenido_por_cajon_real (lo
    # que Depósito mira, un cajón por vez) -- nunca con el total
    # (cantidad_fraccion_real).
    assert 'id="total-real"' in respuesta.text
    assert 'value="3"' in respuesta.text


def test_ver_corregir_recepcion_compra_no_recepcionada_muestra_aviso_sin_formulario():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, estado="pendiente")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 200
    assert "no hay valores reales para corregir" in respuesta.text
    assert "<form" not in respuesta.text


def test_ver_corregir_recepcion_compra_inexistente_da_404():
    with (
        patch("app.main.obtener_detalle_compra", return_value=None),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/999/corregir-recepcion")

    assert respuesta.status_code == 404


def test_ver_corregir_recepcion_compra_error_de_base_da_500():
    with patch("app.main.obtener_detalle_compra", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 500


def test_corregir_recepcion_compra_ruta_guarda_y_redirige():
    with patch("app.main.corregir_recepcion_compra", return_value=None) as mock_corregir:
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "30", "cantidad_total_real": "2400"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/30/detalle?aviso=Se+corrigi%C3%B3+la+recepci%C3%B3n+de+esta+compra."
    mock_corregir.assert_called_once_with(30, 30.0, 2400.0, cantidad_cajones_rechazada=None, motivo_rechazo=None)


def test_ver_corregir_recepcion_muestra_los_campos_de_rechazo_parcial_precargados():
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, cantidad_cajones_real=8, cantidad_cajones_rechazada=2, motivo_rechazo="podrido")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 200
    assert 'name="cantidad_cajones_rechazada"' in respuesta.text
    assert 'value="2"' in respuesta.text
    assert 'name="motivo_rechazo"' in respuesta.text
    assert 'value="podrido"' in respuesta.text


def test_corregir_recepcion_compra_ruta_corrige_el_rechazo_parcial():
    with patch("app.main.corregir_recepcion_compra", return_value=None) as mock_corregir:
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={
                "cantidad_cajones_real": "7",
                "cantidad_total_real": "18",
                "cantidad_cajones_rechazada": "3",
                "motivo_rechazo": "golpeado",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_corregir.assert_called_once_with(30, 7.0, 18.0, cantidad_cajones_rechazada=3.0, motivo_rechazo="golpeado")


def test_corregir_recepcion_compra_ruta_con_rechazo_invalido_da_400():
    with (
        patch("app.main.corregir_recepcion_compra") as mock_corregir,
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={
                "cantidad_cajones_real": "7",
                "cantidad_total_real": "18",
                "cantidad_cajones_rechazada": "abc",
            },
        )

    assert respuesta.status_code == 400
    assert "La cantidad de bultos rechazados tiene que ser un número." in respuesta.text
    mock_corregir.assert_not_called()


def test_corregir_recepcion_compra_ruta_sin_datos_muestra_error_sin_guardar():
    with (
        patch("app.main.corregir_recepcion_compra") as mock_corregir,
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "", "cantidad_total_real": "2400"},
        )

    assert respuesta.status_code == 400
    assert "La cantidad de cajones real es obligatoria." in respuesta.text
    mock_corregir.assert_not_called()


def test_corregir_recepcion_compra_ruta_bloqueada_da_400():
    with (
        patch(
            "app.main.corregir_recepcion_compra",
            side_effect=ValueError("Esta compra no está recepcionada, no hay valores reales para corregir."),
        ),
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "30", "cantidad_total_real": "2400"},
        )

    assert respuesta.status_code == 400
    assert "no hay valores reales para corregir" in respuesta.text


def test_corregir_recepcion_compra_ruta_error_de_base_da_500():
    with (
        patch("app.main.corregir_recepcion_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "30", "cantidad_total_real": "2400"},
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la corrección" in respuesta.text


def test_editar_compra_exitosa_redirige_a_compras():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "55000",
                "sena": "1000",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    # COMPRA_DE_PRUEBA está pendiente/pendiente: ni cantidad ni precio
    # están bloqueados, se actualizan los dos.
    mock_actualizar_cantidad.assert_called_once_with(30, 5, 8.0, 15.0, 120.0, None, "Carro")
    mock_actualizar_precio.assert_called_once_with(30, 55000.0, 1000.0)


def test_editar_compra_recepcionada_guarda_precio_pero_no_cantidad():
    # Punto central del pedido: precio se puede editar aunque la
    # cantidad ya esté bloqueada por estar recepcionada.
    compra_recepcionada = dict(COMPRA_DE_PRUEBA, estado="recepcionado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_recepcionada),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "60000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_called_once_with(30, 60000.0, None)


def test_editar_compra_rechazada_no_guarda_ni_cantidad_ni_precio():
    # Rechazada por calidad: esa historia ya terminó y no entra al
    # costeo, así que queda bloqueada del todo (cantidad y precio), sin
    # importar el estado_retiro.
    compra_rechazada = dict(COMPRA_DE_PRUEBA, estado="rechazado", estado_retiro="cancelado")
    with (
        patch("app.main.obtener_compra", return_value=compra_rechazada),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "60000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()


def test_editar_compra_no_ingresada_no_guarda_ni_cantidad_ni_precio():
    # Nunca ingresó al depósito: mismo criterio que rechazada, queda
    # bloqueada del todo.
    compra_no_ingresada = dict(COMPRA_DE_PRUEBA, estado="no_ingresado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_no_ingresada),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "60000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()


def test_editar_compra_rechazada_y_retirada_no_guarda_nada():
    # Ambos bloqueados: el POST no rompe nada, simplemente no guarda ni
    # cantidad ni precio (el botón Guardar ya está deshabilitado en
    # pantalla para este caso, esto es el resguardo del backend).
    compra_bloqueada_del_todo = dict(COMPRA_DE_PRUEBA, estado="rechazado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_bloqueada_del_todo),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "60000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()


def test_editar_compra_agregar_articulo_crea_compra_nueva_en_la_misma_guia():
    # "Agregar artículo" NO actualiza el renglón 30: inserta una compra
    # nueva (mismo proveedor y fecha que la compra editada — crear_compra
    # arma el siguiente punto de esa guía solo). Se queda en la pantalla
    # de edición del renglón original, no va a Buscar Compras.
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "accion": "agregar",
                "articulo_id": "6",
                "cantidad_cajones": "5",
                "contenido_por_cajon": "10",
                "importe": "20000",
                "sena": "",
                "tipo_retiro": "Pases",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/30/editar"
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()
    mock_crear.assert_called_once_with(
        COMPRA_DE_PRUEBA["fecha_operacion"], 6, COMPRA_DE_PRUEBA["proveedor_id"], 5.0, 10.0, 50.0, None, 20000.0, None, "Pases"
    )


def test_editar_compra_agregar_articulo_funciona_aunque_el_renglon_original_este_bloqueado():
    # Punto 3: agregar no es editar. Aunque el renglón 30 ya esté
    # recepcionado/retirado (bloqueado para Guardar), "Agregar artículo"
    # tiene que seguir funcionando — no pasa por actualizar_cantidad_compra
    # ni actualizar_precio_compra.
    compra_bloqueada = dict(COMPRA_DE_PRUEBA, estado="recepcionado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_bloqueada),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "accion": "agregar",
                "articulo_id": "6",
                "cantidad_cajones": "5",
                "contenido_por_cajon": "10",
                "importe": "20000",
                "sena": "",
                "tipo_retiro": "Pases",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/30/editar"
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()
    mock_crear.assert_called_once()


def test_editar_compra_agregar_articulo_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "accion": "agregar",
                "articulo_id": "6",
                "cantidad_cajones": "5",
                "contenido_por_cajon": "10",
                "importe": "20000",
                "sena": "",
                "tipo_retiro": "Pases",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo agregar el artículo" in respuesta.text


def test_editar_compra_inexistente_da_404():
    with patch("app.main.obtener_compra", return_value=None):
        respuesta = cliente.post(
            "/compras/999/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "15",
                "importe": "55000",
                "sena": "",
                "tipo_retiro": "Carro",
            },
        )

    assert respuesta.status_code == 404


def test_editar_compra_sin_cantidad_de_cajones_muestra_error():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra") as mock_actualizar_cantidad,
        patch("app.main.actualizar_precio_compra") as mock_actualizar_precio,
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "",
                "contenido_por_cajon": "18",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "La cantidad de cajones es obligatoria" in respuesta.text
    mock_actualizar_cantidad.assert_not_called()
    mock_actualizar_precio.assert_not_called()


def test_editar_compra_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.actualizar_cantidad_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "12.5",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la compra" in respuesta.text


def test_editar_compra_ya_retirada_da_400_con_el_mensaje():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch(
            "app.main.actualizar_cantidad_compra",
            side_effect=ValueError("Esta compra ya fue retirada, no se puede editar la cantidad."),
        ),
    ):
        respuesta = cliente.post(
            "/compras/30/editar",
            data={
                "articulo_id": "5",
                "cantidad_cajones": "8",
                "contenido_por_cajon": "12.5",
                "importe": "50000",
                "sena": "",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "Esta compra ya fue retirada, no se puede editar la cantidad." in respuesta.text


def test_ver_editar_compra_recepcionada_bloquea_cantidad_pero_deja_precio_habilitado():
    # Recepcionada: cantidad bloqueada, precio NO (el comprador puede
    # seguir renegociando el precio con el proveedor). Guardar sigue
    # habilitado porque todavía hay algo que se puede guardar.
    compra_recepcionada = dict(COMPRA_DE_PRUEBA, estado="recepcionado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_recepcionada),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "La cantidad no se puede modificar: la compra ya fue recepcionada. El precio sí se puede corregir." in respuesta.text
    assert 'name="accion" value="guardar" disabled' not in respuesta.text
    # Los 4 campos de cantidad (artículo, cajones, contenido, retiro) grises;
    # importe/seña no.
    assert respuesta.text.count('class="campo-bloqueado"') == 4


def test_ver_editar_compra_no_ingresada_bloquea_todo_aunque_nunca_se_haya_retirado():
    # No ingresó: esa historia ya terminó y no entra al costeo, así que
    # queda bloqueada del todo (cantidad y precio) sin importar el
    # estado_retiro — ni siquiera hace falta que se haya retirado.
    compra_no_ingresada = dict(COMPRA_DE_PRUEBA, estado="no_ingresado", estado_retiro="pendiente")
    with (
        patch("app.main.obtener_compra", return_value=compra_no_ingresada),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "Esta compra nunca ingresó al depósito: no se puede modificar ni la cantidad ni el precio." in respuesta.text
    assert 'name="accion" value="guardar" disabled' in respuesta.text
    assert respuesta.text.count('class="campo-bloqueado"') == 6
    # "Agregar artículo" sigue habilitado siempre, incluso acá.
    assert 'name="accion" value="agregar" id="boton-agregar-articulo">Agregar artículo</button>' in respuesta.text


def test_ver_editar_compra_rechazada_bloquea_todo_aunque_nunca_se_haya_retirado():
    # Rechazada por calidad: mismo criterio que no_ingresado, bloquea
    # del todo sin importar el estado_retiro.
    compra_rechazada = dict(COMPRA_DE_PRUEBA, estado="rechazado", estado_retiro="cancelado")
    with (
        patch("app.main.obtener_compra", return_value=compra_rechazada),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "Esta compra tuvo un rechazo total: no se puede modificar ni la cantidad ni el precio." in respuesta.text
    assert 'name="accion" value="guardar" disabled' in respuesta.text
    assert respuesta.text.count('class="campo-bloqueado"') == 6
    # "Agregar artículo" sigue habilitado siempre, incluso acá.
    assert 'name="accion" value="agregar" id="boton-agregar-articulo">Agregar artículo</button>' in respuesta.text


def test_ver_editar_compra_sin_procesar_no_muestra_aviso_ni_deshabilita():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "no se puede editar" not in respuesta.text
    assert "no se puede modificar" not in respuesta.text
    assert 'name="accion" value="guardar" disabled' not in respuesta.text
    assert respuesta.text.count('class="campo-bloqueado"') == 0


def test_ver_editar_compra_muestra_boton_volver_rojo():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert '<a class="boton boton-peligro" href="/compras/buscar" id="boton-volver" onclick="return confirmarVolver()">Volver</a>' in respuesta.text
    assert "Volver a compras" not in respuesta.text
    assert "confirmarVolver" in respuesta.text


def test_ver_editar_compra_recepcionada_marca_la_bandera_js_para_el_aviso_de_precio():
    compra_recepcionada = dict(COMPRA_DE_PRUEBA, estado="recepcionado", estado_retiro="retirado")
    with (
        patch("app.main.obtener_compra", return_value=compra_recepcionada),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert "var compraRecepcionada = true;" in respuesta.text
    assert "cambia el costo del artículo" in respuesta.text


def test_ver_editar_compra_no_recepcionada_no_marca_la_bandera_js():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert "var compraRecepcionada = false;" in respuesta.text


def test_eliminar_compra_exitosa_redirige_a_compras():
    with (
        patch("app.main.eliminar_compra", return_value=[]) as mock_eliminar,
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
    ):
        respuesta = cliente.post("/compras/30/eliminar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_eliminar.assert_called_once_with(30)
    # Esta compra no tenía foto (eliminar_compra devolvió None): no hay
    # nada que borrar del Storage.
    mock_borrar_foto.assert_not_called()


def test_eliminar_compra_exitosa_conserva_los_filtros_de_la_busqueda():
    # Los filtros viajan como campos ocultos del form de la fila: al borrar
    # se vuelve a la MISMA búsqueda, no a la default de 48hs.
    with (
        patch("app.main.eliminar_compra", return_value=[]),
        patch("app.main.borrar_foto_comanda"),
    ):
        respuesta = cliente.post(
            "/compras/30/eliminar",
            data={"fecha_desde": "2026-07-28", "fecha_hasta": "2026-07-29", "proveedor_id": "7", "articulo_id": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/compras/buscar?")
    assert "fecha_desde=2026-07-28" in location
    assert "fecha_hasta=2026-07-29" in location
    assert "proveedor_id=7" in location
    # El filtro vacío no viaja (queda "Todos" en la búsqueda).
    assert "articulo_id" not in location


def test_editar_compra_exitosa_conserva_los_filtros_de_la_busqueda():
    # Los filtros viajan en el query de la URL de editar (los pone el link
    # "Editar" de Buscar): al guardar se vuelve a la MISMA búsqueda.
    datos = {
        "articulo_id": "5",
        "cantidad_cajones": "10",
        "contenido_por_cajon": "18",
        "importe": "5000",
        "sena": "",
        "tipo_retiro": "Clark",
    }
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULOS_CON_UNIDAD_COMPRA[0]),
        patch("app.main.actualizar_cantidad_compra"),
        patch("app.main.actualizar_precio_compra"),
    ):
        respuesta = cliente.post(
            "/compras/30/editar?fecha_desde=2026-07-28&fecha_hasta=2026-07-29&proveedor_id=7&articulo_id=",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/compras/buscar?")
    assert "fecha_desde=2026-07-28" in location
    assert "proveedor_id=7" in location


def test_ver_editar_compra_con_filtros_los_lleva_en_el_form_y_en_volver():
    with (
        patch("app.main.obtener_compra", return_value=COMPRA_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_DEL_CATALOGO),
    ):
        respuesta = cliente.get("/compras/30/editar?fecha_desde=2026-07-28&fecha_hasta=2026-07-29")

    assert respuesta.status_code == 200
    # El form postea a la misma URL con query (los filtros sobreviven al
    # guardado y a los reintentos por error de validación).
    assert 'action="/compras/30/editar?fecha_desde=2026-07-28&amp;fecha_hasta=2026-07-29"' in respuesta.text
    # Y el botón Volver vuelve a esa misma búsqueda.
    assert 'href="/compras/buscar?fecha_desde=2026-07-28&amp;fecha_hasta=2026-07-29"' in respuesta.text


def test_eliminar_compra_con_foto_que_era_la_unica_referencia_la_borra_tambien_del_storage():
    with (
        patch("app.main.eliminar_compra", return_value=["2026-08-13/n07p41-123-abcdef12.jpg"]),
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
    ):
        respuesta = cliente.post("/compras/30/eliminar", follow_redirects=False)

    assert respuesta.status_code == 303
    mock_borrar_foto.assert_called_once_with("2026-08-13/n07p41-123-abcdef12.jpg")


def test_eliminar_compra_si_falla_el_borrado_de_la_foto_igual_redirige_bien():
    # Regresión: borrar la foto es un extra — si falla, la compra ya se
    # borró y no debe tumbar la respuesta al usuario.
    with (
        patch("app.main.eliminar_compra", return_value=["2026-08-13/n07p41-123-abcdef12.jpg"]),
        patch("app.main.borrar_foto_comanda", side_effect=Exception("sin conexión con Storage")),
    ):
        respuesta = cliente.post("/compras/30/eliminar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"


def test_eliminar_compra_error_de_base_da_500():
    with patch("app.main.eliminar_compra", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.post("/compras/30/eliminar")

    assert respuesta.status_code == 500


def test_eliminar_compra_ya_recepcionada_muestra_cartel_en_buscar_compras():
    # Nada de JSON crudo: el rechazo vuelve a la pantalla de Buscar Compras
    # con el mensaje como cartel legible.
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_DE_PRUEBA),
        patch(
            "app.main.eliminar_compra",
            side_effect=ValueError("Esta compra ya fue recepcionada, no se puede eliminar."),
        ),
    ):
        respuesta = cliente.post("/compras/30/eliminar")

    assert respuesta.status_code == 400
    assert "Esta compra ya fue recepcionada, no se puede eliminar." in respuesta.text
    assert '<div class="aviso">' in respuesta.text


def test_eliminar_compra_no_ingresada_muestra_cartel_y_conserva_filtros():
    # Una compra marcada "No ingresó" por Depósito queda fija: el comprador
    # no la puede borrar desde Buscar Compras. El intento vuelve a la misma
    # búsqueda (los filtros viajan en el form) con el cartel.
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_DE_PRUEBA) as mock_buscar,
        patch(
            "app.main.eliminar_compra",
            side_effect=ValueError('Esta compra quedó registrada como "No ingresó" en Depósito, no se puede eliminar.'),
        ),
    ):
        respuesta = cliente.post(
            "/compras/30/eliminar",
            data={"fecha_desde": "2026-08-10", "fecha_hasta": "2026-08-12", "proveedor_id": "", "articulo_id": ""},
        )

    assert respuesta.status_code == 400
    assert "quedó registrada como &#34;No ingresó&#34; en Depósito" in respuesta.text
    mock_buscar.assert_called_once_with(date(2026, 8, 10), date(2026, 8, 12), None, None, limite=TOPE_FILAS_BUSQUEDA + 1)


def test_eliminar_varias_compras_exitosa_muestra_aviso_y_conserva_filtros():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_DE_PRUEBA),
        patch("app.main.eliminar_compra", return_value=[]) as mock_eliminar,
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
    ):
        respuesta = cliente.post(
            "/compras/eliminar-varias",
            data={
                "compra_id": ["30", "31"],
                "fecha_desde": "2026-08-01",
                "fecha_hasta": "2026-08-06",
                "proveedor_id": "200",
                "articulo_id": "",
            },
        )

    assert respuesta.status_code == 200
    assert mock_eliminar.call_count == 2
    mock_eliminar.assert_any_call(30)
    mock_eliminar.assert_any_call(31)
    mock_borrar_foto.assert_not_called()
    assert '<div class="aviso">Se eliminaron 2 compras.</div>' in respuesta.text
    # Conserva los filtros que estaban activos cuando se apretó el borrado.
    assert 'value="2026-08-01"' in respuesta.text
    assert 'value="2026-08-06"' in respuesta.text


def test_eliminar_varias_compras_sin_ninguna_seleccionada_no_hace_nada():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
        patch("app.main.eliminar_compra") as mock_eliminar,
    ):
        respuesta = cliente.post("/compras/eliminar-varias", data={})

    assert respuesta.status_code == 200
    mock_eliminar.assert_not_called()


def test_eliminar_varias_compras_una_falla_no_corta_el_lote_y_avisa_sin_tecnicismos():
    # id 30 (Mzn Red, Saturno) se borra bien; id 31 (Mango, Frutamax) falla
    # (ej. porque ya fue retirada o recepcionada). El mensaje al usuario no
    # debe mostrar ids ni el error crudo de Postgres, y sí debe nombrar el
    # renglón que no se pudo borrar de forma reconocible (artículo + proveedor).
    def eliminar_side_effect(compra_id):
        if compra_id == 31:
            raise Exception('update or delete on table "compras" violates foreign key constraint')
        return []

    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_DE_PRUEBA),
        patch("app.main.eliminar_compra", side_effect=eliminar_side_effect) as mock_eliminar,
        patch("app.main.borrar_foto_comanda") as mock_borrar_foto,
    ):
        respuesta = cliente.post(
            "/compras/eliminar-varias",
            data={"compra_id": ["30", "31"], "fecha_desde": "2026-08-01", "fecha_hasta": "2026-08-06"},
        )

    assert respuesta.status_code == 200
    assert mock_eliminar.call_count == 2
    mock_borrar_foto.assert_not_called()
    assert "Se eliminaron 1 de 2 compras" in respuesta.text
    assert "1 no se pudieron eliminar" in respuesta.text
    assert "ya fueron retiradas, recepcionadas o marcadas &#34;No ingresó&#34;" in respuesta.text
    # Identifica el renglón fallido por artículo+proveedor, no por id ni con
    # el texto crudo de Postgres. (El "31" en la fila de la tabla es el
    # value del checkbox, no el mensaje de error — no cuenta como fuga.)
    assert "Mango (Frutamax)" in respuesta.text
    assert "id 31" not in respuesta.text
    assert "foreign key" not in respuesta.text
    assert "violates" not in respuesta.text


def test_eliminar_varias_compras_todas_fallan_informa_las_dos():
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=COMPRAS_DE_PRUEBA),
        patch("app.main.eliminar_compra", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post(
            "/compras/eliminar-varias",
            data={"compra_id": ["30", "31"], "fecha_desde": "2026-08-01", "fecha_hasta": "2026-08-06"},
        )

    assert respuesta.status_code == 200
    assert "Se eliminaron 0 de 2 compras" in respuesta.text
    assert "2 no se pudieron eliminar" in respuesta.text
    assert "Mzn Red (Saturno)" in respuesta.text
    assert "Mango (Frutamax)" in respuesta.text


COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA = [
    {
        "id": 1, "guia_id": 105, "guia_punto": 1, "articulo_nombre": "Tomate Cherry", "unidad_compra": "kilo",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41", "fecha_operacion": HOY_DE_PRUEBA,
        "cantidad_cajones": 40, "contenido_por_cajon": 20, "cantidad_kilos": 800, "cantidad_fraccion": None,
    },
    {
        "id": 2, "guia_id": 105, "guia_punto": 2, "articulo_nombre": "Mango", "unidad_compra": "unidad",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "fecha_operacion": HOY_DE_PRUEBA - timedelta(days=3),
        "cantidad_cajones": 10, "contenido_por_cajon": 12, "cantidad_kilos": None, "cantidad_fraccion": 120,
    },
    {
        "id": 3, "guia_id": 106, "guia_punto": 1, "articulo_nombre": "Frutilla", "unidad_compra": "cubeta",
        "proveedor_nombre": "Don Pepe", "proveedor_codigo_puesto": "N01P02",
        "cantidad_cajones": 5, "contenido_por_cajon": 12, "cantidad_kilos": None, "cantidad_fraccion": 60,
    },
]


def test_ver_recepcion_agrupa_por_guia_y_muestra_estimado():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Guía 105" in respuesta.text
    assert "Guía 106" in respuesta.text
    assert "Saturno (N07P41)" in respuesta.text
    assert "Don Pepe (N01P02)" in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "Mango" in respuesta.text
    assert "Frutilla" in respuesta.text
    assert "40 cajones × 20k" in respuesta.text
    # Etiquetas según unidad_compra de cada artículo. Las tres piden por
    # cajón/bulto (Depósito mira un bulto por vez -- lo pesa o lo cuenta --
    # nunca toda la carga junta).
    assert "Kilos por cajón/bulto (real)" in respuesta.text
    assert "Unidades por cajón/bulto (real)" in respuesta.text
    assert "Cubetas por cajón/bulto (real)" in respuesta.text


def test_ver_recepcion_prellena_los_inputs_con_el_estimado():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert 'id="cajones-real-1"' in respuesta.text
    assert 'value="40"' in respuesta.text
    # El input real precarga siempre con el contenido por cajón estimado
    # (nunca con el total): 20 para Tomate Cherry (kilo), 12 para Mango
    # (unidad) y Frutilla (cubeta) -- las tres piden por cajón/bulto.
    assert 'id="total-real-1"' in respuesta.text
    assert 'id="total-real-2"' in respuesta.text
    assert 'id="total-real-3"' in respuesta.text
    assert 'value="20"' in respuesta.text
    assert 'value="12"' in respuesta.text


def test_ver_recepcion_muestra_el_proveedor_grande_y_la_guia_chica():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert '<h2>Saturno (N07P41)</h2>' in respuesta.text
    assert '<p class="guia-numero">Guía 105</p>' in respuesta.text


def test_ver_recepcion_muestra_los_tres_botones_con_sus_nombres():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert ">Recibir<" in respuesta.text
    assert ">Rechazo total<" in respuesta.text
    assert ">No ingresó<" in respuesta.text
    assert 'action="/deposito/recepcion/1/no-ingreso"' in respuesta.text


def test_ver_recepcion_confirmacion_es_en_el_lugar_no_confirm_nativo():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "onsubmit=\"return confirm(" not in respuesta.text
    assert "onclick=\"mostrarConfirmacion('1', 'recibir')\"" in respuesta.text
    assert "onclick=\"mostrarConfirmacion('1', 'rechazar')\"" in respuesta.text
    assert "onclick=\"mostrarConfirmacion('1', 'no-ingreso')\"" in respuesta.text
    assert "¿Recepcionar Tomate Cherry?" in respuesta.text
    # El rechazo total ya NO dice "No se puede deshacer": se puede, el
    # mismo día, y la confirmación tiene que decir dónde.
    bloque_rechazo = respuesta.text.split('id="confirmacion-rechazar-1"')[1].split("</div>")[0]
    assert "¿Rechazo total de Tomate Cherry?" in bloque_rechazo
    assert "No se puede deshacer" not in bloque_rechazo
    assert 'se puede deshacer hoy desde "Procesados hoy"' in bloque_rechazo
    # El rechazo PARCIAL sí sigue diciendo que no se puede deshacer: ese
    # recepciona el resto, crea lote y genera la devolución.
    bloque_parcial = respuesta.text.split('id="confirmacion-rechazo-parcial-1"')[1].split("</div>")[0]
    assert "No se puede deshacer." in bloque_parcial
    # Y el que confirma dice qué confirma: dos "Sí" idénticos en la misma
    # posición confirmaban cosas opuestas.
    assert ">Sí, rechazar todo<" in respuesta.text
    assert "¿Tomate Cherry nunca llegó al depósito?" in respuesta.text


def test_ver_recepcion_recibir_va_solo_y_los_otros_tres_abajo():
    # La causa del rechazo apretado por error no es que falte confirmación
    # -- ya son dos pasos en la propia pantalla -- sino que el destructivo
    # estaba pegado a Recibir, que es el que se aprieta casi siempre.
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert 'class="boton-exito boton-recibir"' in respuesta.text
    assert "Si algo salió mal:" in respuesta.text
    # Recibir queda FUERA de la fila donde están los otros tres.
    fila_problemas = respuesta.text.split("Si algo salió mal:")[1].split("</div>")[0]
    assert "Rechazo total" in fila_problemas
    assert "Rechazo parcial" in fila_problemas
    assert "No ingresó" in fila_problemas
    assert ">Recibir<" not in fila_problemas
    # button.clase, no .clase suelta: "button { }" de arriba le gana por
    # especificidad a una clase sola (ya nos pasó en Reproceso).
    assert ".acciones-renglon button.boton-ancho" in respuesta.text


def test_ver_recepcion_sin_pendientes_muestra_mensaje_vacio():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "No hay compras pendientes de recepción." in respuesta.text


def test_ver_recepcion_error_de_base_da_500():
    with patch("app.main.listar_compras_pendientes_recepcion", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 500


def test_recepcionar_compra_guarda_los_reales_y_redirige():
    with patch("app.main.recepcionar_compra", return_value=None) as mock_recepcionar:
        respuesta = cliente.post(
            "/deposito/recepcion/1/recepcionar",
            data={"cantidad_cajones_real": "38", "cantidad_total_real": "760"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=1"
    mock_recepcionar.assert_called_once_with(1, 38.0, 760.0)


def test_recepcionar_compra_con_aviso_de_retiro_lo_pasa_por_la_url():
    # Si Depósito recepciona algo que Logística ya había marcado 'cancelado',
    # recepcionar_compra devuelve un aviso — no se pisa el cancelado, pero
    # tampoco puede pasar callado.
    with patch(
        "app.main.recepcionar_compra", return_value="Esta compra figuraba cancelada en Logística."
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/recepcionar",
            data={"cantidad_cajones_real": "38", "cantidad_total_real": "760"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=1&aviso=Esta+compra+figuraba+cancelada+en+Log%C3%ADstica."


def test_recepcionar_compra_sin_datos_muestra_error_sin_guardar():
    with (
        patch("app.main.recepcionar_compra") as mock_recepcionar,
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/recepcionar",
            data={"cantidad_cajones_real": "", "cantidad_total_real": "760"},
        )

    assert respuesta.status_code == 400
    assert "La cantidad de cajones real es obligatoria." in respuesta.text
    mock_recepcionar.assert_not_called()


def test_recepcionar_compra_con_numero_invalido_muestra_error_sin_guardar():
    with (
        patch("app.main.recepcionar_compra") as mock_recepcionar,
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/recepcionar",
            data={"cantidad_cajones_real": "abc", "cantidad_total_real": "760"},
        )

    assert respuesta.status_code == 400
    assert "tiene que ser un número" in respuesta.text
    mock_recepcionar.assert_not_called()


def test_recepcionar_compra_error_de_base_muestra_mensaje():
    with (
        patch("app.main.recepcionar_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/recepcionar",
            data={"cantidad_cajones_real": "38", "cantidad_total_real": "760"},
        )

    assert respuesta.status_code == 500
    assert "No se pudo recepcionar la compra" in respuesta.text


def test_rechazar_compra_redirige_y_no_pide_datos():
    with patch("app.main.rechazar_compra", return_value=None) as mock_rechazar:
        respuesta = cliente.post("/deposito/recepcion/2/rechazar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=2"
    mock_rechazar.assert_called_once_with(2)


def test_rechazar_compra_con_aviso_de_retiro_lo_pasa_por_la_url():
    with patch(
        "app.main.rechazar_compra", return_value="Esta compra figuraba cancelada en Logística."
    ):
        respuesta = cliente.post("/deposito/recepcion/2/rechazar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=2&aviso=Esta+compra+figuraba+cancelada+en+Log%C3%ADstica."


def test_ver_recepcion_muestra_el_boton_y_el_panel_de_rechazo_parcial():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Rechazo parcial" in respuesta.text
    primera = COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA[0]
    assert f'action="/deposito/recepcion/{primera["id"]}/rechazo-parcial"' in respuesta.text
    assert 'name="cantidad_cajones_llegados"' in respuesta.text
    assert 'name="cantidad_cajones_rechazada"' in respuesta.text
    assert 'name="motivo_rechazo"' in respuesta.text


def test_rechazo_parcial_guarda_los_aceptados_y_el_registro():
    # Llegaron 10, rechaza 2: recepcionar_compra recibe 8 aceptados (los
    # que usa todo el costeo — importe por bulto, ninguna cuenta cambia)
    # más el registro del rechazo.
    with patch("app.main.recepcionar_compra", return_value=None) as mock_recepcionar:
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "2",
                "cantidad_total_real": "18",
                "motivo_rechazo": "podrido",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=1"
    mock_recepcionar.assert_called_once_with(
        1, 8.0, 18.0, cantidad_cajones_rechazada=2.0, motivo_rechazo="podrido"
    )


def test_rechazo_parcial_sin_motivo_guarda_none():
    with patch("app.main.recepcionar_compra", return_value=None) as mock_recepcionar:
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "2",
                "cantidad_total_real": "18",
                "motivo_rechazo": "   ",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_recepcionar.assert_called_once_with(
        1, 8.0, 18.0, cantidad_cajones_rechazada=2.0, motivo_rechazo=None
    )


def test_rechazo_parcial_que_rechaza_todo_da_error_y_manda_al_otro_boton():
    with (
        patch("app.main.recepcionar_compra") as mock_recepcionar,
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "10",
                "cantidad_total_real": "18",
            },
        )

    assert respuesta.status_code == 400
    assert "Si rechazás todo, usá Rechazo total." in respuesta.text
    mock_recepcionar.assert_not_called()


def test_rechazo_parcial_de_cero_bultos_da_error_y_manda_a_recibir():
    with (
        patch("app.main.recepcionar_compra") as mock_recepcionar,
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "0",
                "cantidad_total_real": "18",
            },
        )

    assert respuesta.status_code == 400
    assert "Si no rechazás nada, usá Recibir." in respuesta.text
    mock_recepcionar.assert_not_called()


def test_rechazo_parcial_sin_rechazados_da_error_sin_guardar():
    with (
        patch("app.main.recepcionar_compra") as mock_recepcionar,
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "",
                "cantidad_total_real": "18",
            },
        )

    assert respuesta.status_code == 400
    assert "La cantidad de bultos rechazados es obligatoria." in respuesta.text
    mock_recepcionar.assert_not_called()


def test_rechazo_parcial_error_de_base_muestra_mensaje():
    with (
        patch("app.main.recepcionar_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/recepcion/1/rechazo-parcial",
            data={
                "cantidad_cajones_llegados": "10",
                "cantidad_cajones_rechazada": "2",
                "cantidad_total_real": "18",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar el rechazo parcial" in respuesta.text


def test_ver_recepcion_muestra_el_aviso_cuando_viene_en_la_url():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion?aviso=Esta+compra+figuraba+cancelada+en+Log%C3%ADstica.")

    assert respuesta.status_code == 200
    assert '<div class="aviso">Esta compra figuraba cancelada en Logística.</div>' in respuesta.text


def test_rechazar_compra_error_de_base_muestra_mensaje():
    with (
        patch("app.main.rechazar_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post("/deposito/recepcion/2/rechazar")

    assert respuesta.status_code == 500
    assert "No se pudo rechazar la compra" in respuesta.text


def test_no_ingreso_compra_redirige_y_no_pide_datos():
    with patch("app.main.marcar_compra_no_ingresada", return_value=None) as mock_no_ingreso:
        respuesta = cliente.post("/deposito/recepcion/2/no-ingreso", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/recepcion?procesado=2"
    mock_no_ingreso.assert_called_once_with(2)


def test_no_ingreso_compra_error_de_base_muestra_mensaje():
    with (
        patch("app.main.marcar_compra_no_ingresada", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=COMPRAS_PENDIENTES_RECEPCION_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post("/deposito/recepcion/2/no-ingreso")

    assert respuesta.status_code == 500
    assert "No se pudo marcar la compra como no ingresada" in respuesta.text


PROCESADOS_HOY_RECEPCION_DE_PRUEBA = [
    {
        "id": 1, "articulo_nombre": "Tomate Cherry", "unidad_compra": "kilo",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 40, "contenido_por_cajon": 20,
        "cantidad_cajones_real": 38, "contenido_por_cajon_real": 19,
        "estado": "no_ingresado", "procesada_el": datetime(2026, 8, 17, 13, 0, tzinfo=timezone.utc),
    },
    {
        "id": 2, "articulo_nombre": "Mango", "unidad_compra": "unidad",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10, "contenido_por_cajon": 12,
        "cantidad_cajones_real": 10, "contenido_por_cajon_real": None,
        "estado": "recepcionado", "procesada_el": datetime(2026, 8, 17, 12, 30, tzinfo=timezone.utc),
    },
]


def test_ver_recepcion_con_procesado_no_ingresado_muestra_la_tarjeta_efimera():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=PROCESADOS_HOY_RECEPCION_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/recepcion?procesado=1")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' in respuesta.text
    assert 'Marcaste "No ingresó" en' in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "N07P41" in respuesta.text
    assert 'action="/deposito/recepcion/1/deshacer-no-ingreso"' in respuesta.text


def test_ver_recepcion_con_procesado_recepcionado_dice_recibiste_y_no_deja_deshacer():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=PROCESADOS_HOY_RECEPCION_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/recepcion?procesado=2")

    assert respuesta.status_code == 200
    assert "Recibiste" in respuesta.text
    assert "Mango" in respuesta.text
    # Recepcionar creó un lote de stock: eso no se deshace desde acá, y el
    # aviso dice por dónde se corrige en vez de dejarlo en la nada.
    assert "Ya se recepcionó: para corregirla hace falta Gerencia." in respuesta.text
    assert 'action="/deposito/recepcion/2/deshacer-no-ingreso"' not in respuesta.text


def test_ver_recepcion_con_procesado_rechazado_dice_rechazaste():
    procesado_rechazado = dict(PROCESADOS_HOY_RECEPCION_DE_PRUEBA[1], estado="rechazado")
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[procesado_rechazado]),
    ):
        respuesta = cliente.get("/deposito/recepcion?procesado=2")

    assert respuesta.status_code == 200
    assert "Rechazaste" in respuesta.text


def test_ver_recepcion_sin_procesado_no_muestra_tarjeta_efimera():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=PROCESADOS_HOY_RECEPCION_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' not in respuesta.text


def test_ver_recepcion_procesado_que_no_esta_en_hoy_no_muestra_tarjeta():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion?procesado=999")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' not in respuesta.text


def test_ver_recepcion_panel_procesados_hoy_muestra_hora_y_deshacer_solo_no_ingresado():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=PROCESADOS_HOY_RECEPCION_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Ver procesados hoy (2)" in respuesta.text
    assert "No ingresó a las 10:00" in respuesta.text  # 13:00 UTC -> 10:00 ARG
    assert "Recibido a las 09:30" in respuesta.text
    assert 'action="/deposito/recepcion/1/deshacer-no-ingreso"' in respuesta.text
    assert 'action="/deposito/recepcion/2/deshacer-no-ingreso"' not in respuesta.text


def test_ver_recepcion_panel_procesados_hoy_muestra_cantidad_y_kilos_recibidos():
    # Lo recepcionado muestra con qué números reales se recibió: cajones,
    # contenido por cajón y el total. Lo no ingresado (o sin contenido real
    # cargado) no muestra la línea — no hay nada recibido que contar.
    recibido = dict(PROCESADOS_HOY_RECEPCION_DE_PRUEBA[0], estado="recepcionado")
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[recibido, PROCESADOS_HOY_RECEPCION_DE_PRUEBA[1]]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Recibido: 38 cajones × 19k = 722k" in respuesta.text
    # El Mango se recepcionó sin contenido real: sin línea de Recibido.
    assert "Recibido: 10 cajones" not in respuesta.text


def test_ver_recepcion_panel_procesados_hoy_rechazado_deja_deshacer():
    # Una rechazada no escribió ningún valor real ni creó lote: deshacerla
    # no puede mover un número de stock ni un costo congelado.
    procesado_rechazado = dict(PROCESADOS_HOY_RECEPCION_DE_PRUEBA[1], estado="rechazado")
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[procesado_rechazado]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert 'action="/deposito/recepcion/2/deshacer-no-ingreso"' in respuesta.text
    assert "Ya se recepcionó" not in respuesta.text


def test_ver_recepcion_panel_procesados_hoy_recepcionado_no_deja_deshacer():
    procesado = dict(PROCESADOS_HOY_RECEPCION_DE_PRUEBA[1], estado="recepcionado")
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[procesado]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Ya se recepcionó: para corregirla hace falta Gerencia." in respuesta.text
    assert 'action="/deposito/recepcion/2/deshacer-no-ingreso"' not in respuesta.text


def test_ver_recepcion_sin_procesados_hoy_muestra_mensaje_vacio():
    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    assert "Ver procesados hoy" in respuesta.text
    assert "Todavía no se procesó nada hoy." in respuesta.text


def test_deshacer_no_ingreso_compra_ruta_redirige_sin_procesado():
    with patch("app.main.deshacer_procesado_compra", return_value="no_ingresado") as mock_deshacer:
        respuesta = cliente.post("/deposito/recepcion/1/deshacer-no-ingreso", follow_redirects=False)

    assert respuesta.status_code == 303
    # Sin tarjeta efímera (procesado): la compra se va del panel y
    # reaparece entre las pendientes, y el aviso dice que eso pasó.
    assert respuesta.headers["location"].startswith("/deposito/recepcion?aviso=")
    assert "procesado=" not in respuesta.headers["location"]
    mock_deshacer.assert_called_once_with(1)


def test_deshacer_de_un_rechazo_avisa_que_era_un_rechazo():
    with patch("app.main.deshacer_procesado_compra", return_value="rechazado"):
        respuesta = cliente.post("/deposito/recepcion/1/deshacer-no-ingreso", follow_redirects=False)

    assert respuesta.status_code == 303
    assert "Rechazo+deshecho" in respuesta.headers["location"]


def test_deshacer_no_ingreso_compra_ruta_bloqueado_da_400():
    with (
        patch(
            "app.main.deshacer_procesado_compra",
            side_effect=ValueError("Esta compra ya fue recepcionada o rechazada, no se puede deshacer."),
        ),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post("/deposito/recepcion/1/deshacer-no-ingreso")

    assert respuesta.status_code == 400
    assert "no se puede deshacer" in respuesta.text


def test_deshacer_no_ingreso_compra_ruta_error_de_base_da_500():
    with (
        patch("app.main.deshacer_procesado_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.post("/deposito/recepcion/1/deshacer-no-ingreso")

    assert respuesta.status_code == 500
    assert "No se pudo deshacer" in respuesta.text


def test_ver_foto_compra_redirige_a_la_url_firmada_de_la_primera_foto_de_la_guia():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    fotos = [{"id": 9, "foto_ruta": "2026-08-06/n07p41-123-abcdef12.jpg", "creado_en": datetime(2026, 8, 6, 10, 0)}]
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=fotos) as mock_fotos,
        patch(
            "app.main.obtener_url_foto",
            return_value="https://proyecto.supabase.co/storage/v1/object/sign/comandas/x.jpg?token=abc",
        ) as mock_url,
    ):
        respuesta = cliente.get("/compras/30/foto", follow_redirects=False)

    assert respuesta.status_code == 307
    assert respuesta.headers["location"] == "https://proyecto.supabase.co/storage/v1/object/sign/comandas/x.jpg?token=abc"
    mock_fotos.assert_called_once_with(105)
    mock_url.assert_called_once_with("2026-08-06/n07p41-123-abcdef12.jpg")


def test_ver_foto_compra_sin_fotos_en_la_guia_da_404():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/foto")

    assert respuesta.status_code == 404


def test_subir_foto_a_guia_comprime_y_la_suma():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    imagen = io.BytesIO()
    Image.new("RGB", (2000, 1500), (200, 100, 50)).save(imagen, format="JPEG")
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.subir_foto_comanda", return_value="2026-08-20/guia-105-abc.jpg") as mock_subir,
        patch("app.main.agregar_foto_guia") as mock_agregar,
    ):
        respuesta = cliente.post(
            "/compras/30/fotos",
            files={"archivo": ("comanda.jpg", imagen.getvalue(), "image/jpeg")},
            data={"volver": "editar", "query_filtros": "fecha_desde=2026-08-01"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/30/editar?fecha_desde=2026-08-01"
    # Se subió COMPRIMIDA (el pipeline de 1000px la deja mucho más chica
    # que el original) y se sumó a la guía, nunca reemplaza.
    bytes_subidos = mock_subir.call_args.args[0]
    assert Image.open(io.BytesIO(bytes_subidos)).width <= 1000
    mock_agregar.assert_called_once_with(105, "2026-08-20/guia-105-abc.jpg")


def test_subir_archivo_no_imagen_ni_pdf_da_400():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.agregar_foto_guia") as mock_agregar,
    ):
        respuesta = cliente.post(
            "/compras/30/fotos",
            files={"archivo": ("nota.txt", b"esto no es una imagen", "text/plain")},
            data={"volver": "editar"},
        )

    assert respuesta.status_code == 400
    mock_agregar.assert_not_called()


def test_borrar_foto_de_guia_borra_el_archivo_solo_si_nadie_mas_lo_usa():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    fotos = [{"id": 9, "foto_ruta": "2026/x.jpg", "creado_en": datetime(2026, 8, 6, 10, 0)}]
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=fotos),
        patch("app.main.borrar_foto_guia", return_value="2026/x.jpg") as mock_borrar,
        patch("app.main.borrar_foto_comanda") as mock_storage,
    ):
        respuesta = cliente.post(
            "/compras/30/fotos/9/borrar",
            data={"volver": "detalle"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/30/detalle"
    mock_borrar.assert_called_once_with(9)
    mock_storage.assert_called_once_with("2026/x.jpg")


def test_borrar_foto_ajena_a_la_guia_da_404_y_no_borra_nada():
    # Un foto_id que no es de la guía de ESTA compra no borra nada.
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
        patch("app.main.borrar_foto_guia") as mock_borrar,
    ):
        respuesta = cliente.post("/compras/30/fotos/999/borrar", data={"volver": "editar"})

    assert respuesta.status_code == 404
    mock_borrar.assert_not_called()


def test_ver_foto_de_guia_valida_que_la_foto_sea_de_esa_guia():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/fotos/999/ver")

    assert respuesta.status_code == 404


def test_ver_foto_compra_inexistente_da_404():
    with patch("app.main.obtener_compra", return_value=None):
        respuesta = cliente.get("/compras/999/foto")

    assert respuesta.status_code == 404


def test_ver_foto_compra_error_de_storage_da_500():
    compra_con_guia = dict(COMPRA_DE_PRUEBA, guia_id=105)
    fotos = [{"id": 9, "foto_ruta": "2026-08-06/n07p41-123-abcdef12.jpg", "creado_en": datetime(2026, 8, 6, 10, 0)}]
    with (
        patch("app.main.obtener_compra", return_value=compra_con_guia),
        patch("app.main.listar_fotos_de_guia", return_value=fotos),
        patch("app.main.obtener_url_foto", side_effect=RuntimeError("Supabase Storage no pudo firmar la URL (404)")),
    ):
        respuesta = cliente.get("/compras/30/foto")

    assert respuesta.status_code == 500


COMPRAS_PENDIENTES_DE_PRUEBA = [
    {
        "id": 40,
        "fecha_operacion": HOY_DE_PRUEBA,
        "articulo_nombre": "Mzn Red",
        "proveedor_nombre": "Saturno",
        "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10,
        "contenido_por_cajon": 18,
        "unidad_compra": "kilo",
    },
]


def test_ver_compras_pendientes_muestra_la_lista():
    with patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_DE_PRUEBA) as mock_listar:
        respuesta = cliente.get("/compras/pendientes")

    assert respuesta.status_code == 200
    assert "Mzn Red" in respuesta.text
    assert "Saturno" in respuesta.text
    mock_listar.assert_called_once_with()
    # Regresión: misma pantalla compacta que /compras — fecha dd/mm y
    # números sin decimales de sobra.
    assert "06/08" in respuesta.text
    assert "2026" not in respuesta.text
    assert "<td>10</td>" in respuesta.text
    # Regresión: letra de unidad pegada al contenido por cajón.
    assert "<td>18k</td>" in respuesta.text
    # Regresión: mismos encabezados compactos que /compras.
    assert "<th>Cant</th>" in respuesta.text
    assert "<th>K/U</th>" in respuesta.text
    assert "<th>$</th>" in respuesta.text


def test_ver_compras_pendientes_sin_pendientes_muestra_mensaje():
    with patch("app.main.listar_compras_sin_precio", return_value=[]):
        respuesta = cliente.get("/compras/pendientes")

    assert respuesta.status_code == 200
    assert "No hay compras pendientes de precio" in respuesta.text


def test_ver_compras_pendientes_error_de_base_da_500():
    with patch("app.main.listar_compras_sin_precio", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/pendientes")

    assert respuesta.status_code == 500
    assert "No se pudieron leer las compras pendientes" in respuesta.text


def test_completar_importe_compra_exitoso_redirige_a_pendientes():
    with patch("app.main.actualizar_importe_compra") as mock_actualizar:
        respuesta = cliente.post(
            "/compras/pendientes/40/importe",
            data={"importe": "50000"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/pendientes"
    mock_actualizar.assert_called_once_with(40, 50000.0)


def test_completar_importe_compra_vacio_muestra_error():
    with (
        patch("app.main.actualizar_importe_compra") as mock_actualizar,
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_DE_PRUEBA),
    ):
        respuesta = cliente.post("/compras/pendientes/40/importe", data={"importe": ""})

    assert respuesta.status_code == 400
    assert "El importe es obligatorio" in respuesta.text
    mock_actualizar.assert_not_called()


def test_completar_importe_compra_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.actualizar_importe_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_DE_PRUEBA),
    ):
        respuesta = cliente.post("/compras/pendientes/40/importe", data={"importe": "50000"})

    assert respuesta.status_code == 500
    assert "No se pudo guardar el importe" in respuesta.text


COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA = [
    {
        "id": 40,
        "fecha_operacion": HOY_DE_PRUEBA,
        "articulo_nombre": "Mzn Red",
        "proveedor_nombre": "Saturno",
        "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10,
        "contenido_por_cajon": 18,
    },
    {
        "id": 41,
        "fecha_operacion": HOY_DE_PRUEBA,
        "articulo_nombre": "Kiwi",
        "proveedor_nombre": "Frutamax",
        "proveedor_codigo_puesto": "L03P38",
        "cantidad_cajones": 5,
        "contenido_por_cajon": 16,
    },
]


def test_ver_compras_pendientes_muestra_boton_de_guardar_todos():
    with patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA):
        respuesta = cliente.get("/compras/pendientes")

    assert respuesta.status_code == 200
    assert 'id="boton-guardar-todos"' in respuesta.text
    assert "Guardar todos" in respuesta.text
    assert 'data-compra-id="40"' in respuesta.text
    assert 'data-compra-id="41"' in respuesta.text


def test_completar_importes_pendientes_guarda_solo_los_que_tienen_valor():
    with (
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA),
        patch("app.main.actualizar_importe_compra") as mock_actualizar,
    ):
        respuesta = cliente.post(
            "/compras/pendientes/guardar-todos",
            data={"importe_40": "50000", "importe_41": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/pendientes"
    mock_actualizar.assert_called_once_with(40, 50000.0)


def test_completar_importes_pendientes_ninguno_cargado_muestra_error():
    with (
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA),
        patch("app.main.actualizar_importe_compra") as mock_actualizar,
    ):
        respuesta = cliente.post(
            "/compras/pendientes/guardar-todos",
            data={"importe_40": "", "importe_41": ""},
        )

    assert respuesta.status_code == 400
    assert "Cargá al menos un importe para guardar" in respuesta.text
    mock_actualizar.assert_not_called()


def test_completar_importes_pendientes_importe_invalido_muestra_error_con_el_articulo():
    with (
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA),
        patch("app.main.actualizar_importe_compra") as mock_actualizar,
    ):
        respuesta = cliente.post(
            "/compras/pendientes/guardar-todos",
            data={"importe_40": "50000", "importe_41": "no-es-un-numero"},
        )

    assert respuesta.status_code == 400
    assert "Kiwi" in respuesta.text
    assert "El importe tiene que ser un número" in respuesta.text
    mock_actualizar.assert_not_called()


def test_completar_importes_pendientes_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.listar_compras_sin_precio", return_value=COMPRAS_PENDIENTES_MULTIPLES_DE_PRUEBA),
        patch("app.main.actualizar_importe_compra", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post(
            "/compras/pendientes/guardar-todos",
            data={"importe_40": "50000"},
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar el importe" in respuesta.text


COMANDA_LEIDA_DE_PRUEBA = {
    "proveedor": {"nombre": "Saturno", "tipo_pabellon": "nave", "numero_pabellon": "7", "puesto": "41"},
    "fecha": "2026-08-06",
    "items": [
        {
            "articulo": "Kiwi",
            "cantidad": 10,
            "importe": 5000,
            "sena": None,
            "nota_margen": "84",
            "confianza": "alta",
        },
        {
            "articulo": "completar artículo",
            "cantidad": 5,
            "importe": "completar importe",
            "sena": None,
            "nota_margen": "",
            "confianza": "baja",
        },
    ],
}


def test_subir_foto_compra_adivina_proveedor_y_articulo():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]) as mock_aprendizaje,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "N07P41" in respuesta.text
    assert "Kiwi" in respuesta.text
    assert "84" in respuesta.text  # nota al margen
    assert "⚠ revisar" in respuesta.text  # el segundo ítem, sin articulo ni importe
    mock_aprendizaje.assert_called_once_with(200)

    # Regresión: los 3 botones de abajo de todo, en este orden exacto y con
    # estos colores. Solo "Cancelar" pide confirmación.
    orden = ["Guardar", "Agregar Artículos", "Cancelar"]
    posiciones = [respuesta.text.index(texto) for texto in orden]
    assert posiciones == sorted(posiciones)
    assert 'value="agregar_articulos"' in respuesta.text
    assert 'class="boton-exito" id="boton-guardar"' in respuesta.text
    assert 'class="boton boton-peligro"' in respuesta.text
    assert "confirm('¿Seguro? Se pierde lo que cargaste de esta compra')" in respuesta.text
    # Tercer tipo de retiro (Pases), además de Clark/Carro. Clark viene
    # preseleccionado por defecto (mismo default que la carga manual).
    assert '<option value="Clark" selected>' in respuesta.text
    assert '<option value="Pases"' in respuesta.text
    # Regresión: la sugerencia puesto<->nombre de la carga manual también
    # está disponible acá (ayuda extra sobre lo que ya adivinó la IA).
    assert '{ codigo: "N07P41", nombre: "Saturno" }' in respuesta.text
    assert "oninput=\"actualizarListaProveedores('codigo_puesto')\"" in respuesta.text
    assert "oninput=\"actualizarListaProveedores('nombre')\"" in respuesta.text


def test_revision_de_foto_etiqueta_el_contenido_con_la_unidad_por_renglon():
    # El renglón matcheado a Kiwi (por kilo) arranca con "Kilos por cajón";
    # el renglón sin artículo queda con el genérico hasta que se elija uno
    # (ahí la actualiza el JS con la unidad del artículo nuevo).
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert "Kilos por cajón *" in respuesta.text
    assert "Contenido por cajón *" in respuesta.text
    assert 'id="item_0_label_contenido"' in respuesta.text


def test_subir_foto_compra_proveedor_nuevo_sin_proveedores_existentes_igual_arma_el_codigo():
    # Regresión: antes, si no había NINGÚN proveedor cargado todavía (o
    # ninguno matcheaba), se perdía el código ya interpretado de la foto y
    # el campo quedaba vacío para un proveedor nuevo. Ahora se sugiere igual.
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor") as mock_aprendizaje,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert 'value="N07P41"' in respuesta.text
    # El proveedor propuesto no existe todavía (id None): no tiene sentido
    # buscarle aprendizaje.
    mock_aprendizaje.assert_not_called()


def test_subir_foto_compra_sin_ningun_dato_de_proveedor_deja_codigo_vacio():
    comanda_sin_pabellon = {
        "proveedor": {"nombre": "", "tipo_pabellon": None, "numero_pabellon": "", "puesto": ""},
        "fecha": "2026-08-06",
        "items": [],
    }
    with (
        patch("app.main.extraer_comanda", return_value=comanda_sin_pabellon),
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert 'value="" autocomplete="off"' in respuesta.text
    # Con un solo renglón (o ninguno) no tiene sentido ofrecer "aplicar a
    # todos los demás".
    assert 'id="aplicar_retiro_a_todos"' not in respuesta.text


def test_subir_foto_compra_adivina_articulo_por_conversion():
    comanda = {
        "proveedor": {"nombre": "Saturno", "tipo_pabellon": "nave", "numero_pabellon": "7", "puesto": "41"},
        "fecha": "2026-08-06",
        "items": [
            {"articulo": "PG", "cantidad": 10, "importe": 5000, "sena": None, "nota_margen": "", "confianza": "alta"},
        ],
    }
    conversiones = [{"articulo_id": 5, "nombre_cliente": "MANZANA PG"}]
    with (
        patch("app.main.extraer_comanda", return_value=comanda),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=conversiones),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    # El artículo 5 (Kiwi, en ARTICULOS_CON_UNIDAD_COMPRA) queda seleccionado en el combo.
    assert "selected>Kiwi" in respuesta.text


def test_subir_foto_compra_error_del_lector_muestra_mensaje_claro():
    with patch("app.main.extraer_comanda", side_effect=Exception("no se pudo conectar con la API")):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 500
    assert "No se pudo leer la foto" in respuesta.text
    # Regresión: la pantalla de error vuelve a "cargar foto" (foco único),
    # no a la pantalla combinada vieja con el formulario manual mezclado.
    assert "Código de puesto" not in respuesta.text
    assert 'id="form-leer-comanda"' in respuesta.text


def _datos_confirmar_foto(descartar_item_1=True, codigo_puesto="N07P41", nombre="Saturno", foto_preview=None):
    datos = {
        "codigo_puesto": codigo_puesto,
        "nombre": nombre,
        "cantidad_renglones": "2",
        "item_0_texto_leido": "Kiwi",
        "item_0_articulo_id": "5",
        "item_0_cantidad_cajones": "10",
        "item_0_contenido_por_cajon": "18",
        "item_0_importe": "5000",
        "item_0_sena": "",
        "item_0_tipo_retiro": "Clark",
        "item_1_texto_leido": "completar artículo",
        "item_1_articulo_id": "",
        "item_1_cantidad_cajones": "5",
        "item_1_contenido_por_cajon": "",
        "item_1_importe": "",
        "item_1_sena": "",
        "item_1_tipo_retiro": "",
    }
    if descartar_item_1:
        datos["item_1_descartar"] = "on"
    if foto_preview is not None:
        datos["foto_preview"] = foto_preview
    return datos


RENGLON_KIWI_ESPERADO = {
    "articulo_id": 5,
    "cantidad_cajones": 10.0,
    "contenido_por_cajon": 18.0,
    "cantidad_kilos": 180.0,
    "cantidad_fraccion": None,
    "importe": 5000.0,
    "sena": None,
    "tipo_retiro": "Clark",
}


def test_confirmar_compra_foto_exitosa_guarda_solo_los_confirmados():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo") as mock_aprender,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=_datos_confirmar_foto(),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/nueva?proveedor_id=200"
    mock_proveedor.assert_called_once_with("N07P41", "Saturno")
    # Sin foto_preview en el form (estos datos de prueba no la mandan), no
    # hay nada que subir a Storage: foto_ruta queda en None. Y sin
    # carga_token (form viejo), viaja None: se guarda sin protección.
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 200, [RENGLON_KIWI_ESPERADO], None, None)
    mock_aprender.assert_called_once_with(200, "kiwi", 5)


def test_confirmar_compra_foto_no_aprende_de_los_placeholders_del_lector():
    # Bug real: cuando la IA no pudo leer el artículo, texto_leido llega como
    # "completar artículo" (el placeholder que le ordena el prompt, no texto
    # de la comanda). Al completar el renglón a mano, la compra se guarda
    # igual, pero NO se aprende nada: no hay texto de comanda que asociar.
    datos = {
        "codigo_puesto": "N07P41",
        "nombre": "Saturno",
        "cantidad_renglones": "1",
        "item_0_texto_leido": "completar artículo",
        "item_0_articulo_id": "5",
        "item_0_cantidad_cajones": "10",
        "item_0_contenido_por_cajon": "18",
        "item_0_importe": "5000",
        "item_0_sena": "",
        "item_0_tipo_retiro": "Clark",
    }
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo") as mock_aprender,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once()
    mock_aprender.assert_not_called()


def test_confirmar_compra_foto_accion_guardar_va_directo_al_resumen_y_guarda_igual():
    # El botón verde "Guardar" tiene que guardar exactamente lo mismo que
    # "Agregar Artículos" (misma llamada a crear_compras_de_comanda/
    # aprender_articulo), la única diferencia es a dónde redirige después.
    datos = _datos_confirmar_foto()
    datos["accion"] = "guardar"
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo") as mock_aprender,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_proveedor.assert_called_once_with("N07P41", "Saturno")
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 200, [RENGLON_KIWI_ESPERADO], None, None)
    mock_aprender.assert_called_once_with(200, "kiwi", 5)


FOTO_PREVIEW_DE_PRUEBA = "data:image/jpeg;base64,aGVsbG8="  # decodifica a b"hello"


def test_confirmar_compra_foto_sube_la_foto_una_vez_y_guarda_la_ruta_en_todos_los_renglones():
    # Dos renglones válidos (no uno descartado como en _datos_confirmar_foto
    # por defecto), para poder confirmar que la MISMA ruta de foto queda en
    # los dos — una comanda = una foto = varios renglones de compra.
    datos = {
        "codigo_puesto": "N07P41",
        "nombre": "Saturno",
        "foto_preview": FOTO_PREVIEW_DE_PRUEBA,
        "cantidad_renglones": "2",
        "item_0_texto_leido": "Kiwi",
        "item_0_articulo_id": "5",
        "item_0_cantidad_cajones": "10",
        "item_0_contenido_por_cajon": "18",
        "item_0_importe": "5000",
        "item_0_sena": "",
        "item_0_tipo_retiro": "Clark",
        "item_1_texto_leido": "Kiwi",
        "item_1_articulo_id": "5",
        "item_1_cantidad_cajones": "3",
        "item_1_contenido_por_cajon": "18",
        "item_1_importe": "1500",
        "item_1_sena": "",
        "item_1_tipo_retiro": "Clark",
    }
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.subir_foto_comanda", return_value="2026-08-06/n07p41-123-abcdef12.jpg") as mock_subir,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo"),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Se sube UNA sola vez (no una vez por renglón), con los bytes ya
    # decodificados del data URI y el código de puesto como base del nombre.
    mock_subir.assert_called_once_with(b"hello", "N07P41")
    # UNA sola llamada con los DOS renglones (todo-o-nada por comanda) y la
    # misma foto_ruta para toda la comanda.
    mock_crear.assert_called_once()
    _, _, renglones_guardados, foto_ruta_guardada, _ = mock_crear.call_args.args
    assert len(renglones_guardados) == 2
    assert foto_ruta_guardada == "2026-08-06/n07p41-123-abcdef12.jpg"


def test_confirmar_compra_foto_si_falla_la_subida_guarda_la_compra_igual_sin_foto():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.subir_foto_comanda", side_effect=RuntimeError("Supabase Storage rechazó la subida (403)")),
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo"),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=_datos_confirmar_foto(foto_preview=FOTO_PREVIEW_DE_PRUEBA),
            follow_redirects=False,
        )

    # NO es un error de "no se pudo guardar la compra": la foto es un
    # extra, la falla de Storage nunca puede bloquear la carga.
    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 200, [RENGLON_KIWI_ESPERADO], None, None)


def test_confirmar_compra_foto_codigo_puesto_invalido_muestra_error():
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.crear_compras_de_comanda") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=_datos_confirmar_foto(codigo_puesto="puesto15"),
        )

    assert respuesta.status_code == 400
    assert "formato NNNPNN" in respuesta.text
    mock_crear.assert_not_called()


def test_confirmar_compra_foto_renglon_invalido_muestra_error_con_numero():
    datos = _datos_confirmar_foto(descartar_item_1=False)
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.crear_compras_de_comanda") as mock_crear,
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert "Renglón 2" in respuesta.text
    assert "Elegí un artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_confirmar_compra_foto_error_incluye_sugerencia_de_proveedor_para_reintentar():
    # Regresión: al reintentar después de un error, la pantalla de revisión
    # sigue teniendo disponible la sugerencia puesto<->nombre (no solo en la
    # primera carga de la foto).
    datos = _datos_confirmar_foto(descartar_item_1=False)
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.crear_compras_de_comanda"),
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert '{ codigo: "N07P41", nombre: "Saturno" }' in respuesta.text


def test_confirmar_compra_foto_conserva_el_retiro_ya_elegido_al_reintentar():
    # Regresión: si un renglón fallaba, el "Retiro" de los DEMÁS renglones se
    # perdía al re-mostrar el formulario (siempre volvía a "Elegí..."),
    # obligando a re-elegirlo en todos de nuevo en cada intento.
    datos = _datos_confirmar_foto(descartar_item_1=False)
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.crear_compras_de_comanda"),
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert "selected>Clark" in respuesta.text


def test_confirmar_compra_foto_todo_descartado_muestra_error():
    datos = _datos_confirmar_foto()
    datos["item_0_descartar"] = "on"
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.crear_compras_de_comanda") as mock_crear,
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert "No hay ningún renglón para guardar" in respuesta.text
    mock_crear.assert_not_called()


def test_confirmar_compra_foto_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=_datos_confirmar_foto())

    assert respuesta.status_code == 500
    assert "No se pudo guardar la compra" in respuesta.text


def test_confirmar_compra_foto_con_token_pasa_el_token_al_guardado():
    datos = _datos_confirmar_foto()
    datos["carga_token"] = "token123"
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.comanda_ya_guardada", return_value=False) as mock_ya_guardada,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo"),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_ya_guardada.assert_called_once_with("token123")
    mock_crear.assert_called_once_with(HOY_DE_PRUEBA, 200, [RENGLON_KIWI_ESPERADO], None, "token123")


def test_confirmar_compra_foto_reintento_con_token_ya_usado_no_guarda_de_nuevo():
    # El corte de internet real: el server guardó y commiteó, pero el
    # teléfono nunca vio la respuesta y manda lo mismo otra vez. No se
    # guarda ni se sube nada de nuevo — se responde igual que el guardado
    # original, así la pantalla avanza a la comanda siguiente sin duplicar.
    datos = _datos_confirmar_foto(foto_preview=FOTO_PREVIEW_DE_PRUEBA)
    datos["carga_token"] = "token123"
    datos["accion"] = "guardar"
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.comanda_ya_guardada", return_value=True),
        patch("app.main.subir_foto_comanda") as mock_subir,
        patch("app.main.crear_compras_de_comanda") as mock_crear,
        patch("app.main.aprender_articulo") as mock_aprender,
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_subir.assert_not_called()
    mock_crear.assert_not_called()
    mock_aprender.assert_not_called()


def test_confirmar_compra_foto_si_falla_el_aprendizaje_la_comanda_queda_guardada_igual():
    # El aprendizaje es un extra: la comanda ya está commiteada cuando se
    # aprende — reportar "no se pudo guardar" acá sería mentira (y antes
    # pasaba exactamente eso).
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo", side_effect=Exception("se cortó la conexión")),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=_datos_confirmar_foto(),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once()


def test_confirmar_compra_foto_conserva_el_token_al_reintentar_por_error():
    # Si un renglón falla la validación, el form se re-muestra: el token
    # tiene que seguir viajando para que el guardado siguiente quede
    # protegido igual.
    datos = _datos_confirmar_foto(descartar_item_1=False)
    datos["carga_token"] = "token123"
    with patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert 'name="carga_token" value="token123"' in respuesta.text


def test_leer_foto_comanda_multiple_incluye_un_token_de_carga_en_el_fragmento():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert 'name="carga_token" value="' in datos["html"]
    # El token no puede salir vacío: es la protección anti-duplicado.
    assert 'name="carga_token" value=""' not in datos["html"]


def _imagen_de_prueba() -> bytes:
    buffer = io.BytesIO()
    Image.new("RGB", (20, 20), color="red").save(buffer, format="JPEG")
    return buffer.getvalue()


def test_generar_preview_foto_devuelve_data_uri_para_una_imagen_valida():
    preview = _generar_preview_foto(_imagen_de_prueba())
    assert preview.startswith("data:image/jpeg;base64,")


def test_generar_preview_foto_devuelve_vacio_si_no_es_una_imagen():
    assert _generar_preview_foto(b"esto no es una imagen") == ""


def _imagen_apaisada_con_orientacion_exif(orientacion: int) -> bytes:
    """Imagen rectangular (30x10) con el tag EXIF Orientation seteado, como la que graba un celular."""
    imagen = Image.new("RGB", (30, 10), color="red")
    exif = imagen.getexif()
    exif[274] = orientacion  # 274 = Orientation
    buffer = io.BytesIO()
    imagen.save(buffer, format="JPEG", exif=exif.tobytes())
    return buffer.getvalue()


def test_generar_preview_foto_aplica_la_rotacion_del_exif():
    # Orientation 6 = hay que rotar 90° para verse derecha: una foto de
    # celular "vertical" que el sensor graba apaisada (30x10) tiene que
    # terminar parada (alto > ancho) en el preview, no apaisada.
    imagen_con_exif = _imagen_apaisada_con_orientacion_exif(6)
    preview = _generar_preview_foto(imagen_con_exif)

    assert preview.startswith("data:image/jpeg;base64,")
    _, base64_texto = preview.split(";base64,", 1)
    imagen_resultante = Image.open(io.BytesIO(base64.standard_b64decode(base64_texto)))
    ancho, alto = imagen_resultante.size
    assert alto > ancho


def test_subir_foto_compra_incluye_preview_de_la_foto_para_el_boton_ver_foto():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert "data:image/jpeg;base64," in respuesta.text
    assert "Ver foto" in respuesta.text


def test_subir_foto_compra_muestra_cartel_de_guardando_para_evitar_duplicados():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert 'id="form-confirmar-foto"' in respuesta.text
    assert "Guardando..." in respuesta.text
    # Regresión: en Safari/iOS, deshabilitar el botón sincrónicamente dentro
    # del evento submit cancela el envío del formulario — tiene que ir
    # diferido en un setTimeout.
    assert "setTimeout" in respuesta.text
    assert "b.elemento.disabled = true;" in respuesta.text
    # Regresión: checkbox para aplicar el retiro del primer artículo a todos
    # los demás, y así no elegirlo renglón por renglón.
    assert 'id="aplicar_retiro_a_todos"' in respuesta.text
    assert "Usar este retiro para todos los artículos" in respuesta.text


def test_confirmar_compra_foto_conserva_la_foto_al_reintentar_por_error():
    datos = _datos_confirmar_foto(descartar_item_1=False)
    datos["foto_preview"] = "data:image/jpeg;base64,ABC123"
    with (
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.crear_compras_de_comanda"),
    ):
        respuesta = cliente.post("/compras/nueva/foto/confirmar", data=datos)

    assert respuesta.status_code == 400
    assert "data:image/jpeg;base64,ABC123" in respuesta.text


def test_ver_carga_comandas_multiples_muestra_la_pantalla():
    respuesta = cliente.get("/compras/nueva/fotos")

    assert respuesta.status_code == 200
    assert "Cargar varias comandas por foto" in respuesta.text
    assert 'id="input-fotos"' in respuesta.text
    assert "multiple" in respuesta.text
    assert 'id="boton-guardar-siguiente"' in respuesta.text
    # Regresión: "Descartar y siguiente" saltea la comanda actual sin
    # guardarla, distinto de "Guardar y siguiente" y de "Cancelar" (que sale
    # de todo el lote).
    assert 'id="boton-descartar-siguiente"' in respuesta.text
    assert "Descartar y siguiente" in respuesta.text
    # Regresión: en la última comanda del lote, "Descartar y siguiente"
    # pasa a decir "Descartar y terminar", igual que ya hace "Guardar y
    # siguiente" -> "Guardar y terminar". Misma condición esUltima, en la
    # misma función (actualizarBotonGuardar) que actualiza los dos botones.
    assert '"Descartar y terminar" : "Descartar y siguiente"' in respuesta.text
    # Regresión: la lectura de todas las fotos arranca de entrada, con un
    # límite de concurrencia (no una por una al guardar).
    assert "LIMITE_CONCURRENCIA_LECTURA" in respuesta.text
    assert "Leídas" in respuesta.text
    # Regresión: al mostrar una comanda nueva (guardar/descartar y pasar a
    # la siguiente) la pantalla vuelve al tope.
    assert "window.scrollTo(0, 0)" in respuesta.text
    # Regresión: buscador combinado de proveedor también en modo múltiple,
    # reusando la misma lógica que la carga manual.
    assert "actualizarListaProveedores" in respuesta.text
    assert "elegirProveedor" in respuesta.text
    # Regresión: al agregar un renglón en blanco, si el tilde "usar este
    # retiro para todos" está tildado, el renglón nuevo se precarga con ese
    # retiro en vez de quedar en "Elegí...".
    assert "aplicar_retiro_a_todos" in respuesta.text
    assert 'checkboxAplicarATodos.checked' in respuesta.text
    assert "retiroActivo ? document.getElementById" in respuesta.text
    # Regresión: un solo toque — "Elegir fotos" no valida ni arranca nada
    # por sí solo, solo abre el input file oculto; el flujo arranca recién
    # en el "change" del input, cuando ya hay fotos elegidas.
    assert 'class="input-foto-oculto"' in respuesta.text
    assert '>Elegir fotos<' in respuesta.text
    assert "inputFotos.click();" in respuesta.text
    assert 'inputFotos.addEventListener("change"' in respuesta.text


def test_ver_carga_comandas_multiples_incluye_los_proveedores_conocidos_para_sugerir():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva/fotos")

    assert respuesta.status_code == 200
    assert '{ codigo: "N07P41", nombre: "Saturno" }' in respuesta.text
    assert '{ codigo: "L03P38", nombre: "Frutamax" }' in respuesta.text


def test_link_multiples_comandas_en_pantalla_de_una_foto():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/compras/nueva")

    assert respuesta.status_code == 200
    assert 'href="/compras/nueva/fotos"' in respuesta.text
    # Regresión: distinguir claramente "Leer una comanda" (una sola foto) de
    # "Múltiples comandas" (el flujo nuevo).
    assert "Leer una comanda" in respuesta.text


def test_leer_foto_comanda_multiple_adivina_proveedor_y_articulo():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is True
    assert datos["cantidad_renglones"] == 2
    assert "Saturno" in datos["html"]
    assert "N07P41" in datos["html"]
    assert "Kiwi" in datos["html"]
    assert 'id="renglones-comanda"' in datos["html"]


def test_leer_foto_comanda_multiple_si_falla_la_ia_devuelve_renglon_en_blanco_para_completar_a_mano():
    with (
        patch("app.main.extraer_comanda", side_effect=Exception("no se pudo leer")),
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is True
    assert datos["cantidad_renglones"] == 1
    assert "⚠ revisar" in datos["html"]
    # No cortó la cola: la foto sigue disponible para el modal "Ver foto".
    assert "modal-foto" in datos["html"]


def test_leer_foto_comanda_multiple_sin_items_leidos_devuelve_renglon_en_blanco():
    comanda_sin_items = {
        "proveedor": {"nombre": "Saturno", "tipo_pabellon": "nave", "numero_pabellon": "7", "puesto": "41"},
        "fecha": "2026-08-06",
        "items": [],
    }
    with (
        patch("app.main.extraer_comanda", return_value=comanda_sin_items),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is True
    assert datos["cantidad_renglones"] == 1
    assert "⚠ revisar" in datos["html"]


def test_leer_foto_comanda_multiple_error_de_base_devuelve_ok_false():
    with (
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", side_effect=Exception("sin conexión")),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


RENGLONES_LISTADO_DE_PRUEBA = [
    {
        "es_idem": False,
        "proveedor_texto": "Saturno",
        "codigo": "",
        "articulo": "Kiwi",
        "cantidad": 10,
        "kg_x_bulto": 18,
        "importe": 5000,
        "nota_margen": "84",
        "confianza": "alta",
    },
    {
        "es_idem": True,
        "proveedor_texto": "",
        "codigo": "",
        "articulo": "Kiwi",
        "cantidad": 5,
        "kg_x_bulto": 20,
        "importe": 3000,
        "nota_margen": "",
        "confianza": "alta",
    },
    {
        "es_idem": False,
        "proveedor_texto": "Frutamax",
        "codigo": "",
        "articulo": "completar articulo",
        "cantidad": 3,
        "kg_x_bulto": None,
        "importe": None,
        "nota_margen": "",
        "confianza": "baja",
    },
]


def test_leer_listado_consolidado_agrupa_por_proveedor_y_arma_un_grupo_revisable_por_cada_uno():
    with (
        patch("app.main.extraer_listado_consolidado", return_value={"renglones": RENGLONES_LISTADO_DE_PRUEBA}),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is True
    # Saturno (2 renglones, el segundo es ídem) y Frutamax (1 renglón),
    # cada uno un grupo revisable aparte, en orden de primera aparición.
    assert len(datos["grupos"]) == 2
    assert datos["grupos"][0]["cantidad_renglones"] == 2
    assert "Saturno" in datos["grupos"][0]["html"]
    assert "N07P41" in datos["grupos"][0]["html"]
    assert datos["grupos"][1]["cantidad_renglones"] == 1
    assert "Frutamax" in datos["grupos"][1]["html"]
    # Opción A: renglón con articulo no reconocido queda marcado para
    # revisar a mano, nunca se guarda solo.
    assert "⚠ revisar" in datos["grupos"][1]["html"]
    # La foto de la planilla se manda en la respuesta para poder subirla
    # una sola vez desde el cliente (asegurarFotoSubida en el template).
    assert datos["foto_preview"].startswith("data:image/jpeg;base64,")


def test_leer_listado_consolidado_kg_x_bulto_de_la_planilla_pisa_el_de_catalogo():
    # Regresión: a diferencia de una comanda normal (contenido_por_cajon
    # sale del catálogo), en el listado cada renglón trae su propio
    # kg_x_bulto — el primer renglón de prueba trae 18 (coincide con el
    # catálogo) y el segundo (ídem) trae 20 (distinto al catálogo, 18):
    # tiene que quedar 20, no el del catálogo.
    with (
        patch("app.main.extraer_listado_consolidado", return_value={"renglones": RENGLONES_LISTADO_DE_PRUEBA}),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    datos = respuesta.json()
    assert 'value="20.0"' in datos["grupos"][0]["html"] or 'value="20"' in datos["grupos"][0]["html"]


def test_leer_listado_consolidado_proveedor_no_matcheado_queda_para_revisar():
    # Opción A también para proveedor no reconocido: el grupo queda armado
    # igual (con el nombre leído, id None), el comprador lo completa a mano
    # en el buscador combinado antes de guardar.
    renglones = [
        {
            "es_idem": False,
            "proveedor_texto": "Proveedor Desconocido SRL",
            "codigo": "",
            "articulo": "Kiwi",
            "cantidad": 10,
            "kg_x_bulto": 18,
            "importe": 5000,
            "nota_margen": "",
            "confianza": "alta",
        }
    ]
    with (
        patch("app.main.extraer_listado_consolidado", return_value={"renglones": renglones}),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    datos = respuesta.json()
    assert datos["ok"] is True
    assert len(datos["grupos"]) == 1
    assert "Proveedor Desconocido SRL" in datos["grupos"][0]["html"]


def test_leer_listado_consolidado_si_falla_la_ia_devuelve_ok_false():
    with (
        patch("app.main.extraer_listado_consolidado", side_effect=Exception("no se pudo leer la planilla")),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


def test_leer_listado_consolidado_error_de_base_devuelve_ok_false():
    with (
        patch("app.main.extraer_listado_consolidado", return_value={"renglones": RENGLONES_LISTADO_DE_PRUEBA}),
        patch("app.main.listar_proveedores", side_effect=Exception("sin conexión")),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


def test_leer_listado_consolidado_sin_renglones_leidos_devuelve_ok_false():
    with (
        patch("app.main.extraer_listado_consolidado", return_value={"renglones": []}),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/listado/leer",
            files={"foto": ("planilla.jpg", _imagen_de_prueba(), "image/jpeg")},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


def test_subir_foto_listado_sube_una_vez_y_devuelve_la_ruta():
    with patch("app.main.subir_foto_comanda", return_value="2026-08-15/listado-abc123.jpg") as mock_subir:
        respuesta = cliente.post(
            "/compras/nueva/listado/subir-foto",
            data={"foto_preview": FOTO_PREVIEW_DE_PRUEBA},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is True
    assert datos["foto_ruta"] == "2026-08-15/listado-abc123.jpg"
    mock_subir.assert_called_once_with(b"hello", "listado")


def test_subir_foto_listado_data_uri_invalida_devuelve_ok_false():
    respuesta = cliente.post(
        "/compras/nueva/listado/subir-foto",
        data={"foto_preview": "esto no es una data uri"},
    )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


def test_subir_foto_listado_si_falla_la_subida_devuelve_ok_false():
    with patch("app.main.subir_foto_comanda", side_effect=RuntimeError("Supabase Storage rechazó la subida (403)")):
        respuesta = cliente.post(
            "/compras/nueva/listado/subir-foto",
            data={"foto_preview": FOTO_PREVIEW_DE_PRUEBA},
        )

    assert respuesta.status_code == 200
    datos = respuesta.json()
    assert datos["ok"] is False
    assert "error" in datos


def test_confirmar_compra_foto_con_foto_ruta_ya_subida_no_vuelve_a_subir():
    # Regresión: en el listado consolidado, a partir del segundo proveedor
    # guardado el form manda foto_ruta_ya_subida — no hay que llamar a
    # subir_foto_comanda de nuevo, hay que usar esa ruta tal cual.
    datos = _datos_confirmar_foto()
    datos["foto_ruta_ya_subida"] = "2026-08-15/listado-abc123.jpg"
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.subir_foto_comanda") as mock_subir,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo"),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=datos,
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_subir.assert_not_called()
    mock_crear.assert_called_once_with(
        HOY_DE_PRUEBA, 200, [RENGLON_KIWI_ESPERADO], "2026-08-15/listado-abc123.jpg", None
    )


def test_confirmar_compra_foto_sin_foto_ruta_ya_subida_sigue_igual_que_antes():
    # Regresión: sin ese campo (comanda única y múltiples fotos, que nunca
    # lo mandan), el comportamiento no cambia — sigue subiendo a partir de
    # foto_preview como siempre.
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.subir_foto_comanda", return_value="2026-08-06/n07p41-123-abcdef12.jpg") as mock_subir,
        patch("app.main.crear_compras_de_comanda", return_value=True) as mock_crear,
        patch("app.main.aprender_articulo"),
    ):
        respuesta = cliente.post(
            "/compras/nueva/foto/confirmar",
            data=_datos_confirmar_foto(foto_preview=FOTO_PREVIEW_DE_PRUEBA),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_subir.assert_called_once_with(b"hello", "N07P41")
    # foto_ruta es el anteúltimo argumento (el último es carga_token).
    assert mock_crear.call_args.args[-2] == "2026-08-06/n07p41-123-abcdef12.jpg"


# --- /negociar: cuadro para negociar precios (Bajas / Subas / Resumen bajo objetivo), por cliente elegido ---

ARTICULOS_NEGOCIAR_DE_PRUEBA = [
    {
        "articulo_nombre": "Tomate Cherry",
        "fresco": True,
        "variacion": "bajo",
        "costo_anterior": 600.0,
        "costo_actual": 500.0,
        "precio_sugerido": 900.0,
        "precio_vigente": 950.0,  # vigente >= sugerido -> ✓
        "utilidad_aproximada": 0.30,  # por encima del objetivo, no entra al resumen
        "compras_sin_precio_excluidas": 0,
    },
    {
        "articulo_nombre": "Mango",
        "fresco": True,
        "variacion": "subio",
        "costo_anterior": 300.0,
        "costo_actual": 400.0,
        "precio_sugerido": 800.0,
        "precio_vigente": 700.0,  # vigente < sugerido -> 🔴
        "utilidad_aproximada": 0.10,  # bajo el objetivo (0.20) -> resumen
        "compras_sin_precio_excluidas": 0,
    },
    {
        "articulo_nombre": "Palta",
        "fresco": False,  # no fresco: no entra en Bajas ni Subas
        "variacion": None,
        "costo_anterior": None,
        "costo_actual": 900.0,
        "precio_sugerido": 1300.0,
        "precio_vigente": 800.0,
        "utilidad_aproximada": -0.05,  # peor utilidad -> primero en el resumen
        "compras_sin_precio_excluidas": 0,
    },
]

FICHAS_NEGOCIAR_DE_PRUEBA = [{"id": 1, "articulo_id": 1, "articulo_nombre": "Tomate Cherry"}]


def test_ver_negociar_sin_cliente_muestra_selector():
    # Mismo patrón que /fichas: sin cliente_id en la URL, solo el
    # selector — no se calcula nada todavía.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA) as mock_listar,
        patch("app.main.calcular_listado_para_negociar_precios") as mock_calcular,
    ):
        respuesta = cliente.get("/negociar")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para ver su cuadro de negociación." in respuesta.text
    assert '<option value="1"' in respuesta.text
    assert "Día" in respuesta.text
    assert "Otro cliente" in respuesta.text
    mock_listar.assert_called_once()
    mock_calcular.assert_not_called()


def test_ver_negociar_con_cliente_muestra_titulo_y_nombre_del_cliente():
    # La pantalla ya no es "de prueba": título fijo + nombre del cliente
    # elegido, sin ningún "(prueba)" en ningún lado.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Márgenes por Artículo" in respuesta.text
    assert "Cliente: <strong>Día</strong>" in respuesta.text
    assert "(prueba)" not in respuesta.text


def test_ver_negociar_bajas_incluye_fresco_que_bajo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA) as mock_calcular,
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    # El motor genérico recibe el cliente_id elegido, no uno fijo.
    assert mock_calcular.call_args[0][0] == 1
    assert "Bajas" in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "$950" in respuesta.text
    assert "✓" in respuesta.text


def test_ver_negociar_subas_incluye_fresco_que_subio():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Mango" in respuesta.text
    assert "🔴" in respuesta.text


def test_ver_negociar_no_fresco_no_aparece_en_bajas_ni_subas():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    import re

    bloque_bajas = re.search(r"<h2>Bajas.*?</h2>(.*?)<h2>Subas", respuesta.text, re.S).group(1)
    bloque_subas = re.search(r"<h2>Subas.*?</h2>(.*?)<h2>\s*Resumen", respuesta.text, re.S).group(1)
    assert "Palta" not in bloque_bajas
    assert "Palta" not in bloque_subas


def test_ver_negociar_resumen_ordena_de_peor_a_mejor_y_filtra_bajo_objetivo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    import re

    bloque_resumen = re.search(r"<h2>\s*Resumen.*?(?=<h2>Todos los artículos)", respuesta.text, re.S).group(0)
    # Palta (-5%, peor) tiene que aparecer antes que Mango (10%).
    pos_palta = bloque_resumen.index("Palta")
    pos_mango = bloque_resumen.index("Mango")
    assert pos_palta < pos_mango
    # Tomate Cherry (30%, por encima del objetivo) no entra al resumen.
    assert "Tomate Cherry" not in bloque_resumen
    assert "utilidad-negativa" in bloque_resumen
    assert "utilidad-baja" in bloque_resumen
    # Regresión: el encabezado de la columna era un literal fijo "P. Día",
    # ahora es genérico (la pantalla puede ser de cualquier cliente).
    assert "P. Día" not in respuesta.text
    assert "<th>Precio vigente</th>" in respuesta.text


def test_ver_negociar_todos_los_articulos_lista_todos_ordenados_por_utilidad_descendente():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    import re

    bloque_todos = re.search(r"<h2>Todos los artículos.*", respuesta.text, re.S).group(0)
    # Tomate Cherry (30%) > Mango (10%) > Palta (-5%) — a diferencia del
    # resumen bajo objetivo, acá SÍ tienen que aparecer los 3, incluido el
    # que está bien (Tomate Cherry).
    pos_tomate = bloque_todos.index("Tomate Cherry")
    pos_mango = bloque_todos.index("Mango")
    pos_palta = bloque_todos.index("Palta")
    assert pos_tomate < pos_mango < pos_palta


def test_ver_negociar_todos_los_articulos_utilidad_ok_sin_color_de_alerta():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    import re

    bloque_todos = re.search(r"<h2>Todos los artículos.*", respuesta.text, re.S).group(0)
    fila_tomate = re.search(r"<tr>\s*<td>Tomate Cherry</td>.*?</tr>", bloque_todos, re.S).group(0)
    # Tomate Cherry (30%, por encima del objetivo de 20%) no lleva ninguna
    # de las clases de alerta — solo Mango y Palta, que sí están mal.
    assert "utilidad-negativa" not in fila_tomate
    assert "utilidad-baja" not in fila_tomate
    assert "30,0%" in fila_tomate


def test_ver_negociar_todos_los_articulos_incluye_el_buscador():
    # El filtro es 100% client-side (ver filtrarTodosArticulos en
    # _cuadro_negociacion.html) — acá solo se verifica que el markup y el
    # script estén, no el comportamiento en el navegador.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="buscar-todos-articulos"' in respuesta.text
    assert 'placeholder="Buscar artículo"' in respuesta.text
    assert 'onclick="limpiarBusquedaTodosArticulos()"' in respuesta.text
    assert "function filtrarTodosArticulos()" in respuesta.text
    assert "Ningún artículo coincide." in respuesta.text


def test_ver_negociar_sin_articulos_no_muestra_el_buscador():
    # Sin nada para listar, el buscador no tiene sentido — no se renderiza.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="buscar-todos-articulos"' not in respuesta.text
    assert "No hay artículos para mostrar." in respuesta.text


def test_ver_negociar_con_precio_de_compra_sin_cerrar_en_fresco_muestra_advertencia():
    # Un artículo fresco (compra dentro de las últimas 48 hs) con alguna
    # compra sin precio de compra cargado todavía -> el costo puede estar
    # incompleto, se avisa con nombre y todo.
    articulos = [dict(a) for a in ARTICULOS_NEGOCIAR_DE_PRUEBA]
    articulos[0]["compras_sin_precio_excluidas"] = 1  # Tomate Cherry, fresco
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=articulos),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'class="aviso aviso-advertencia"' in respuesta.text
    assert "Hay artículos sin precio de compra cargado en las últimas 48 horas" in respuesta.text
    assert "Tomate Cherry" in respuesta.text.split('class="aviso aviso-advertencia"')[1][:300]
    assert 'href="/compras/pendientes"' in respuesta.text


def test_ver_negociar_precio_de_compra_sin_cerrar_en_no_fresco_no_muestra_advertencia():
    # Mismo caso pero en Palta, que no es fresca -- no cuenta para el aviso:
    # la ventana de 48 hs de la que se están sacando los datos no la incluye.
    articulos = [dict(a) for a in ARTICULOS_NEGOCIAR_DE_PRUEBA]
    articulos[2]["compras_sin_precio_excluidas"] = 1  # Palta, fresco=False
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=articulos),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'class="aviso aviso-advertencia"' not in respuesta.text


def test_ver_negociar_sin_precios_de_compra_sin_cerrar_no_muestra_advertencia():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIAR_DE_PRUEBA),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'class="aviso aviso-advertencia"' not in respuesta.text


def test_ver_negociar_sin_fichas_muestra_aviso_y_link_para_cargarlas():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "todavía no tiene fichas de logística cargadas" in respuesta.text
    assert 'href="/fichas?cliente_id=1"' in respuesta.text
    # Igual muestra la pantalla completa (secciones vacías, no en blanco).
    assert "Ningún artículo fresco bajó de costo" in respuesta.text


def test_ver_negociar_con_fichas_pero_sin_articulos_recientes_muestra_aviso_distinto():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "ningún artículo tuvo compra en los últimos 15 días" in respuesta.text
    assert "todavía no tiene fichas de logística cargadas" not in respuesta.text


def test_ver_negociar_sin_utilidad_objetivo_muestra_aviso():
    clientes_sin_utilidad = [{"id": 1, "nombre": "Cliente Nuevo", "descuento": 0.0, "adicionales": 0.0, "utilidad_objetivo": None}]
    with (
        patch("app.main.listar_clientes", return_value=clientes_sin_utilidad),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "no tiene utilidad objetivo cargada todavía" in respuesta.text
    assert 'href="/clientes/1/editar"' in respuesta.text


def test_ver_negociar_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/negociar?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_negociar_error_al_listar_clientes_da_500():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/negociar")

    assert respuesta.status_code == 500


def test_ver_negociar_error_de_base_da_500():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.get("/negociar?cliente_id=1")

    assert respuesta.status_code == 500


def test_ver_negociar_otro_cliente_no_muestra_datos_de_dia():
    # Regresión explícita: elegir un cliente que no es "Día" tiene que
    # calcular para ESE cliente_id, no para el 1 fijo de antes.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_NEGOCIAR_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]) as mock_calcular,
    ):
        respuesta = cliente.get("/negociar?cliente_id=2")

    assert respuesta.status_code == 200
    assert mock_calcular.call_args[0][0] == 2
    assert "Cliente: <strong>Otro cliente</strong>" in respuesta.text


def test_costeo_prueba_ya_no_existe():
    # La pantalla de depuración se retiró: la negociación real ahora se
    # accede eligiendo cliente en /negociar.
    respuesta = cliente.get("/costeo-prueba")

    assert respuesta.status_code == 404


# --- /precios: botonera de Lista de Precios ---


def test_ver_precios_muestra_la_botonera_con_los_seis_accesos_en_orden():
    respuesta = cliente.get("/precios")

    assert respuesta.status_code == 200
    assert 'href="/precios/cargar"' in respuesta.text
    assert 'href="/precios/cargar-foto"' in respuesta.text
    assert 'href="/precios/consultar"' in respuesta.text
    assert 'href="/negociar"' in respuesta.text
    assert 'href="/precios/resultado-negociacion"' in respuesta.text
    assert 'href="/precios/generar-listado"' in respuesta.text
    assert "Modificar Precios" in respuesta.text
    assert "Carga Foto Precios" in respuesta.text
    assert "Consultar Precios" in respuesta.text
    assert "Márgenes por Artículo" in respuesta.text
    assert "Resultado Negociación" in respuesta.text
    assert "Cargar Precios Nuevos" not in respuesta.text
    assert "Próximamente" in respuesta.text

    orden = [
        "Modificar Precios",
        "Carga Foto Precios",
        "Consultar Precios",
        "Márgenes por Artículo",
        "Resultado Negociación",
        "Generar Listado Actualizado",
    ]
    posiciones = [respuesta.text.index(texto) for texto in orden]
    assert posiciones == sorted(posiciones)


def test_ver_precios_guardado_muestra_mensaje_de_confirmacion():
    respuesta = cliente.get("/precios?guardado=3")

    assert respuesta.status_code == 200
    assert "Se cargaron 3 precios." in respuesta.text


def test_ver_precios_guardado_cero_muestra_mensaje_sin_cambios():
    respuesta = cliente.get("/precios?guardado=0")

    assert respuesta.status_code == 200
    assert "No se guardó ningún cambio" in respuesta.text


def test_ver_precios_sin_guardado_no_muestra_mensaje():
    respuesta = cliente.get("/precios")

    assert respuesta.status_code == 200
    assert "Se cargaron" not in respuesta.text
    assert "No se guardó ningún cambio" not in respuesta.text


def test_ver_precios_guardado_y_listado_muestra_mensaje_de_guardado_y_generado():
    respuesta = cliente.get("/precios?guardado=2&listado=1")

    assert respuesta.status_code == 200
    assert "Se guardaron 2 precios y se generó el listado." in respuesta.text


def test_ver_precios_guardado_cero_y_listado_muestra_mensaje_igual():
    respuesta = cliente.get("/precios?guardado=0&listado=1")

    assert respuesta.status_code == 200
    assert "Los precios ya estaban al día — se generó el listado igual." in respuesta.text


# --- /precios/consultar: consulta de precios vigentes (por cliente+fecha, o cliente+articulo+fecha) ---

FICHAS_PRECIOS_DE_PRUEBA = [
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Tomate Cherry"},
    {"id": 902, "articulo_id": 2, "articulo_nombre": "Mango"},
]

PRECIOS_VIGENTES_DE_PRUEBA = [
    {"ficha_id": 901, "articulo_id": 1, "precio": 500.0},
    {"ficha_id": 902, "articulo_id": 2, "precio": 350.0},
]


def test_ver_precios_consultar_sin_cliente_muestra_selector():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente") as mock_precios,
    ):
        respuesta = cliente.get("/precios/consultar")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para ver sus precios." in respuesta.text
    assert "Día" in respuesta.text
    mock_precios.assert_not_called()


def test_ver_precios_consultar_con_cliente_lista_todos_los_precios_vigentes():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA) as mock_precios,
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Tomate Cherry" in respuesta.text
    assert "$500" in respuesta.text
    assert "Mango" in respuesta.text
    assert "$350" in respuesta.text
    # Sin fecha en la URL, consulta a HOY.
    assert mock_precios.call_args[0] == (1, HOY_DE_PRUEBA)


def test_ver_precios_consultar_fecha_pasada_usa_esa_fecha():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]) as mock_precios,
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&fecha=2026-01-15")

    assert respuesta.status_code == 200
    assert mock_precios.call_args[0][1] == date(2026, 1, 15)
    assert "15/01/2026" in respuesta.text


def _precios_vigentes_con_fechas(vigente_desde_cherry, vigente_desde_mango):
    return [
        {"ficha_id": 901, "articulo_id": 1, "precio": 500.0, "vigente_desde": vigente_desde_cherry},
        {"ficha_id": 902, "articulo_id": 2, "precio": 350.0, "vigente_desde": vigente_desde_mango},
    ]


def _bloque_de_articulo(html: str, nombre: str) -> str:
    # El renglón del artículo: desde el arranque de su fila hasta la
    # siguiente. Alcanza para ver si ESE renglón quedó resaltado.
    inicio = html.rindex('<div class="fila-precio', 0, html.index(nombre))
    return html[inicio : html.index("</div>", html.index(nombre))]


def test_ver_precios_consultar_resalta_el_precio_que_empezo_a_regir_ese_dia():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_con_fechas(HOY_DE_PRUEBA, date(2026, 1, 1)),
        ),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert "nueva" in _bloque_de_articulo(respuesta.text, "Tomate Cherry")
    assert "nueva" not in _bloque_de_articulo(respuesta.text, "Mango")
    assert "Nuevo precio" in respuesta.text


def test_ver_precios_consultar_fecha_pasada_resalta_igual_lo_que_cambio_ese_dia():
    # El agujero que se arregló: consultando para atrás el resaltado se
    # perdía entero. Lo que empezó a regir el 15/01 tiene que verse en rojo
    # al consultar el 15/01, igual que si fuera hoy.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_con_fechas(date(2026, 1, 15), date(2026, 1, 1)),
        ),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&fecha=2026-01-15")

    assert "nueva" in _bloque_de_articulo(respuesta.text, "Tomate Cherry")
    assert "nueva" not in _bloque_de_articulo(respuesta.text, "Mango")
    assert "el precio que empezó a regir el 15/01/2026" in respuesta.text


def test_ver_precios_consultar_sin_cambios_ese_dia_no_resalta_ninguno():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_con_fechas(date(2026, 1, 1), date(2026, 1, 1)),
        ),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&fecha=2026-01-15")

    assert 'class="badge-nuevo"' not in respuesta.text
    assert "fila-precio nueva" not in respuesta.text
    assert "empezaron a regir" not in respuesta.text


def test_ver_precios_consultar_fecha_invalida_muestra_error():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&fecha=no-es-una-fecha")

    assert respuesta.status_code == 200
    assert "La fecha no es válida." in respuesta.text


def test_ver_precios_consultar_articulo_puntual_filtra_a_ese_solo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&ficha_id=902")

    assert respuesta.status_code == 200
    # El selector de artículo sigue listando todos (para poder elegir
    # otro), pero el RESULTADO tiene que quedar filtrado a uno solo.
    import re

    bloque_resultado = re.search(r'<h2>Precios vigentes.*?</div>\s*</div>', respuesta.text, re.S).group(0)
    assert "Mango" in bloque_resultado
    assert "Tomate Cherry" not in bloque_resultado


def test_ver_precios_consultar_articulo_puntual_sin_precio_vigente_muestra_mensaje():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&ficha_id=902")

    assert respuesta.status_code == 200
    assert "no tiene precio vigente" in respuesta.text


def test_ver_precios_consultar_cliente_sin_precios_muestra_mensaje():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "no tiene ningún precio vigente" in respuesta.text


def test_ver_precios_consultar_incluye_link_para_cargar():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert 'href="/precios/cargar?cliente_id=1"' in respuesta.text


def test_ver_precios_consultar_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/consultar?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_precios_consultar_error_de_base_da_500():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/precios/consultar")

    assert respuesta.status_code == 500


def test_ver_precios_consultar_articulo_id_vacio_trae_el_listado_completo():
    # Regresión: "Todos los artículos" manda articulo_id="" (el <select>
    # sin elegir nada puntual) — antes esto rompía con un 422 crudo de
    # FastAPI ("Input should be a valid integer") en vez de traer el
    # listado completo, que es lo que realmente significa un campo vacío.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&fecha=2026-08-15&articulo_id=")

    assert respuesta.status_code == 200
    assert "Tomate Cherry" in respuesta.text
    assert "Mango" in respuesta.text


def test_ver_precios_consultar_articulo_id_no_numerico_no_rompe_trae_el_listado_completo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&articulo_id=no-es-un-numero")

    assert respuesta.status_code == 200
    assert "Tomate Cherry" in respuesta.text
    assert "Mango" in respuesta.text


def test_ver_precios_consultar_cliente_id_vacio_muestra_el_selector_sin_romper():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/consultar?cliente_id=")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para ver sus precios." in respuesta.text


def test_ver_precios_consultar_url_no_repite_cliente_id():
    # Regresión: el form tenía un input hidden "cliente_id" duplicando el
    # <select> del mismo nombre, generando ?cliente_id=1&...&cliente_id=1.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert respuesta.text.count('name="cliente_id"') == 1


def test_ver_precios_consultar_incluye_buscador_de_articulo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert 'id="articulo_texto"' in respuesta.text
    assert "actualizarListaArticulos" in respuesta.text
    # El buscador lista FICHAS (901/902), no artículos: es la ficha la que
    # tiene precio, y dos del mismo artículo se llamarían igual.
    assert '{ id: 901, nombre: "Tomate Cherry" }' in respuesta.text
    assert '{ id: 902, nombre: "Mango" }' in respuesta.text


def test_ver_precios_consultar_articulo_elegido_muestra_boton_para_limpiar():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1&ficha_id=902")

    assert "Ver todos los artículos" in respuesta.text
    assert 'value="Mango"' in respuesta.text


def test_ver_precios_consultar_con_resultados_muestra_boton_exportar():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert 'id="boton-exportar"' in respuesta.text
    assert "/precios/consultar/exportar-pdf?cliente_id=1&fecha=" in respuesta.text
    assert "/precios/consultar/exportar-excel?cliente_id=1&fecha=" in respuesta.text


def test_ver_precios_consultar_boton_exportar_va_arriba_del_boton_cargar_precios():
    # El botón Exportar tiene que quedar justo debajo de "Ver" (el submit
    # del formulario de arriba), antes de "Cargar precios de este cliente"
    # y antes del listado — no al final de la pantalla.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert respuesta.text.index('id="boton-exportar"') < respuesta.text.index("Cargar precios de este cliente")


def test_ver_precios_consultar_sin_resultados_no_muestra_boton_exportar():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar?cliente_id=1")

    assert 'id="boton-exportar"' not in respuesta.text


# --- /precios/consultar/exportar-pdf y exportar-excel ---


def _texto_del_pdf_de_respuesta(pdf_bytes: bytes) -> str:
    documento = pdfium.PdfDocument(pdf_bytes)
    return "\n".join(pagina.get_textpage().get_text_range() for pagina in documento)


def _texto_sin_leyenda_de_respuesta(pdf_bytes: bytes) -> str:
    # La leyenda fija se repite en el encabezado de CADA página (para que
    # se vea aunque una sección se corte entre dos hojas) — se descarta acá
    # para contar los badges reales sin que la paginación lo altere.
    from core.exportar_precios import LEYENDA_PRECIO_NUEVO

    return _texto_del_pdf_de_respuesta(pdf_bytes).replace(LEYENDA_PRECIO_NUEVO, "")


FICHAS_EXPORTACION_DE_PRUEBA = [
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Tomate Cherry", "unidad_venta": "kilo"},
    {"id": 902, "articulo_id": 2, "articulo_nombre": "Mango", "unidad_venta": "unidad"},
]

ARTICULOS_EXPORTACION_DE_PRUEBA = [
    {"id": 1, "nombre": "Tomate Cherry", "grupo": "hortaliza"},
    {"id": 2, "nombre": "Mango", "grupo": "fruta"},
]


def _precios_vigentes_exportacion(vigente_desde_articulo_1):
    return [
        {"ficha_id": 901, "articulo_id": 1, "precio": 500.0, "vigente_desde": vigente_desde_articulo_1},
        {"ficha_id": 902, "articulo_id": 2, "precio": 350.0, "vigente_desde": date(2026, 1, 1)},
    ]


def test_exportar_precios_pdf_devuelve_archivo_adjunto():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(HOY_DE_PRUEBA),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get(f"/precios/consultar/exportar-pdf?cliente_id=1&fecha={HOY_DE_PRUEBA.isoformat()}")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "attachment" in respuesta.headers["content-disposition"]
    # Nombre de la empresa en el archivo: el cliente puede recibir listas
    # de más de una empresa y tiene que poder distinguirlas sin abrirlas.
    assert "Lista_Precios_Frutamax_D" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"%PDF")


def test_exportar_precios_pdf_hoy_resalta_el_que_cambio_hoy():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(HOY_DE_PRUEBA),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get(f"/precios/consultar/exportar-pdf?cliente_id=1&fecha={HOY_DE_PRUEBA.isoformat()}")

    texto = _texto_sin_leyenda_de_respuesta(respuesta.content)
    assert texto.count("Nuevo precio") == 1


def test_exportar_precios_pdf_fecha_pasada_resalta_lo_que_empezo_a_regir_ese_dia():
    # Consultar el 15/01 tiene que mostrar en rojo lo que arrancó el 15/01
    # (Tomate Cherry acá): es para lo que se consulta para atrás — facturar
    # sin tener que comparar contra la lista del día anterior. Mango viene
    # del 1/1 y queda en negro. Antes el resaltado se apagaba entero en
    # cuanto la fecha no era HOY.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(date(2026, 1, 15)),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=1&fecha=2026-01-15")

    texto = _texto_sin_leyenda_de_respuesta(respuesta.content)
    assert texto.count("Nuevo precio") == 1


def test_exportar_precios_pdf_separa_por_grupo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(date(2026, 1, 1)),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=1&fecha=2026-01-01")

    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert texto.index("FRUTA") < texto.index("HORTALIZA")
    assert "Mango" in texto[texto.index("FRUTA") : texto.index("HORTALIZA")]
    assert "Tomate Cherry" in texto[texto.index("HORTALIZA") :]


FICHAS_EXPORTACION_NOMBRE_CLIENTE_DE_PRUEBA = [
    # Con nombre_cliente cargado: la lista exportada tiene que usar ESE
    # nombre, no el interno del catálogo (articulo_nombre).
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Mzn Red", "unidad_venta": "kilo", "nombre_cliente": "Manzana Red Elegida"},
    # Sin nombre_cliente cargado: se cae al nombre del catálogo.
    {"id": 902, "articulo_id": 2, "articulo_nombre": "Anana", "unidad_venta": "unidad", "nombre_cliente": None},
]


def test_exportar_precios_pdf_usa_nombre_cliente_de_la_ficha_no_el_interno():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_NOMBRE_CLIENTE_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(date(2026, 1, 1)),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=1&fecha=2026-01-01")

    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Manzana Red Elegida" in texto
    assert "Mzn Red" not in texto
    # Sin nombre_cliente cargado, se cae al nombre del catálogo.
    assert "Anana" in texto


def test_exportar_precios_pdf_cliente_invalido_da_400():
    respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=abc&fecha=2026-08-16")

    assert respuesta.status_code == 400


def test_exportar_precios_pdf_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=999&fecha=2026-08-16")

    assert respuesta.status_code == 404


def test_exportar_precios_pdf_fecha_invalida_da_400():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/consultar/exportar-pdf?cliente_id=1&fecha=no-es-una-fecha")

    assert respuesta.status_code == 400


def test_exportar_precios_excel_devuelve_archivo_adjunto():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(HOY_DE_PRUEBA),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
    ):
        respuesta = cliente.get(f"/precios/consultar/exportar-excel?cliente_id=1&fecha={HOY_DE_PRUEBA.isoformat()}")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip


def test_exportar_precios_excel_usa_el_formato_de_planilla_del_dueno():
    # Formato pedido explícito (21/08/2026): fecha arriba, Producto |
    # Precio Anterior | Precio Desde HOY, cambios resaltados en naranja
    # con el precio en rojo, pie "PRECIOS POR KG - SIN IVA".
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(HOY_DE_PRUEBA),
        ),
        # Mango (articulo_id 2) tenía $300 antes; Tomate Cherry (1) nunca
        # tuvo un precio previo cargado.
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[{"ficha_id": 902, "articulo_id": 2, "precio": 300.0}]),
    ):
        respuesta = cliente.get(f"/precios/consultar/exportar-excel?cliente_id=1&fecha={HOY_DE_PRUEBA.isoformat()}")

    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    hoja = libro.active

    # Fecha arriba sin ceros a la izquierda; encabezados de la planilla.
    assert hoja.cell(row=1, column=1).value == f"{HOY_DE_PRUEBA.day}/{HOY_DE_PRUEBA.month}/{HOY_DE_PRUEBA.year}"
    assert hoja.cell(row=2, column=1).value == "Producto"
    assert hoja.cell(row=2, column=2).value == "Precio Anterior"
    assert hoja.cell(row=2, column=3).value == "Precio Desde HOY"

    # Cherry EMPEZÓ A REGIR hoy: naranja con el precio en rojo. Nunca tuvo
    # precio previo cargado, así que el anterior repite el vigente (figura
    # igual, pedido explícito) — lo que manda es la fecha, no la comparación.
    fila_cherry = next(fila for fila in hoja.iter_rows() if fila[0].value == "Tomate Cherry")
    assert fila_cherry[1].value == 500.0
    assert fila_cherry[2].value == 500.0
    assert fila_cherry[2].fill.start_color.rgb.endswith("FFC000")
    assert fila_cherry[2].font.color.rgb.endswith("C00000")
    assert fila_cherry[1].fill.start_color.rgb != fila_cherry[2].fill.start_color.rgb  # el anterior queda sin resaltar

    # Mango tiene un precio anterior distinto (300 -> 350) pero ese cambio
    # es del 1/1, no de hoy: NO se resalta. Antes sí, y era un falso aviso.
    # Se vende por unidad y el pie dice "POR KG", así que la unidad va
    # pegada al nombre.
    fila_mango = next(fila for fila in hoja.iter_rows() if fila[0].value == "Mango (por unidad)")
    assert fila_mango[1].value == 300.0
    assert fila_mango[2].value == 350.0
    assert fila_mango[2].fill.start_color.rgb != fila_cherry[2].fill.start_color.rgb

    # Pie de la planilla, en la última fila con contenido.
    valores = [celda.value for fila in hoja.iter_rows() for celda in fila if celda.value is not None]
    assert "PRECIOS POR KG - SIN IVA" in valores


def test_exportar_precios_excel_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/consultar/exportar-excel?cliente_id=999&fecha=2026-08-16")

    assert respuesta.status_code == 404


# --- /precios/cargar/guardar-y-exportar-pdf y guardar-y-exportar-excel: "Guardar y generar listado"
# de Carga Manual — guarda de verdad los pendientes y arma el PDF/Excel con lo recién guardado ---


def test_guardar_y_exportar_precios_cargar_manual_pdf_guarda_y_devuelve_archivo():
    precios_tras_guardar = [
        {"ficha_id": 901, "articulo_id": 1, "precio": 500.0, "vigente_desde": date(2026, 1, 1)},
        {"ficha_id": 902, "articulo_id": 2, "precio": 400.0, "vigente_desde": HOY_DE_PRUEBA},
    ]
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=precios_tras_guardar),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar/guardar-y-exportar-pdf",
            data={"cliente_id": "1", "pendiente_precio_902": "400", "pendiente_original_902": "350"},
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"%PDF")
    assert respuesta.headers["x-cantidad-guardada"] == "1"
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 902, "precio": 400.0}])

    # El archivo generado tiene que reflejar lo recién guardado: Mango
    # (el que se acaba de pactar) resaltado, Tomate Cherry (sin tocar) no.
    texto = _texto_sin_leyenda_de_respuesta(respuesta.content)
    bloque_frutas = texto[texto.index("FRUTA") : texto.index("HORTALIZA")]
    bloque_hortalizas = texto[texto.index("HORTALIZA") :]
    assert bloque_frutas.count("Nuevo precio") == 1
    assert "Nuevo precio" not in bloque_hortalizas


def test_guardar_y_exportar_precios_cargar_manual_excel_guarda_y_devuelve_archivo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(date(2026, 1, 1)),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar/guardar-y-exportar-excel",
            data={"cliente_id": "1", "pendiente_precio_902": "400", "pendiente_original_902": "350"},
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip
    assert respuesta.headers["x-cantidad-guardada"] == "1"
    mock_guardar.assert_called_once()


def test_guardar_y_exportar_precios_cargar_manual_pdf_usa_nombre_cliente_de_la_ficha():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_NOMBRE_CLIENTE_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(date(2026, 1, 1)),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
        patch("app.main.guardar_precios_cliente"),
    ):
        respuesta = cliente.post(
            "/precios/cargar/guardar-y-exportar-pdf",
            data={"cliente_id": "1", "pendiente_precio_901": "950", "pendiente_original_901": "890"},
        )

    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Manzana Red Elegida" in texto
    assert "Mzn Red" not in texto
    assert "Anana" in texto


def test_guardar_y_exportar_precios_cargar_manual_pdf_cliente_invalido_da_400():
    respuesta = cliente.post("/precios/cargar/guardar-y-exportar-pdf", data={"cliente_id": "abc"})

    assert respuesta.status_code == 400


def test_guardar_y_exportar_precios_cargar_manual_pdf_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post("/precios/cargar/guardar-y-exportar-pdf", data={"cliente_id": "999"})

    assert respuesta.status_code == 404


def test_guardar_y_exportar_precios_cargar_manual_pdf_precio_invalido_da_400_y_no_guarda():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar/guardar-y-exportar-pdf",
            data={"cliente_id": "1", "pendiente_precio_902": "no-es-un-numero"},
        )

    assert respuesta.status_code == 400
    mock_guardar.assert_not_called()


# --- /precios/cargar: carga de precios nuevos uno a la vez, con revisión antes de guardar ---


def test_ver_cargar_precios_sin_cliente_muestra_selector():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/cargar")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para cargarle precios nuevos." in respuesta.text


def test_ver_cargar_precios_embebe_el_catalogo_con_precio_vigente():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="articulo_texto"' in respuesta.text
    assert "actualizarListaArticulos" in respuesta.text
    # La lista embebida es de FICHAS (901/902), no de artículos (1/2):
    # el precio cuelga de la ficha.
    assert 'id: 901,\n        nombre: "Tomate Cherry",\n        precioVigente: 500.0,' in respuesta.text
    assert 'id: 902,\n        nombre: "Mango",\n        precioVigente: 350.0,' in respuesta.text


def test_ver_cargar_precios_articulo_sin_precio_previo_embebe_null():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id: 901,\n        nombre: "Tomate Cherry",\n        precioVigente: null,' in respuesta.text


def test_ver_cargar_precios_embebe_costo_y_denominador_para_simulacion():
    # Tomate Cherry (articulo_id 1) tiene costo reciente -> se embeben los
    # dos números que necesita "Simulación" para calcular sin servidor.
    # Mango (2) no aparece en calcular_listado_para_negociar_precios (sin
    # compra reciente) -> tiene que quedar en null, no romper nada.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIACION_DE_PRUEBA),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert (
        'costoProductoUnidadVenta: 280.0,\n        costoEnvaseUnidadVenta: 0.0,\n        denominadorTasas: 1.0,'
        in respuesta.text
    )
    assert (
        'id: 902,\n        nombre: "Mango",\n        precioVigente: 350.0,\n        '
        'costoProductoUnidadVenta: null,\n        costoEnvaseUnidadVenta: null,\n        denominadorTasas: null,'
    ) in respuesta.text


def test_ver_cargar_precios_incluye_boton_guardar_y_generar_listado():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="boton-guardar-y-generar"' in respuesta.text
    assert "Guardar y generar listado" in respuesta.text
    assert "guardarYGenerarListado" in respuesta.text
    assert "/precios/cargar/guardar-y-exportar-pdf" in respuesta.text
    assert "/precios/cargar/guardar-y-exportar-excel" in respuesta.text
    # El botón "Exportar sin guardar" que había antes ya no va.
    assert 'id="boton-exportar"' not in respuesta.text


def test_ver_cargar_precios_boton_cargar_otro_precio_va_en_azul_y_hay_cancelar():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    # "Cargar otro Precio" en azul sólido (clase "boton" a secas), no el
    # outline blanco-y-azul de "boton-secundario".
    assert '<button type="button" class="boton" onclick="cargarOtroPrecio()">Cargar otro Precio</button>' in respuesta.text
    # "Cancelar" junto a "Guardar y terminar", mismo criterio de
    # visibilidad (solo si hay algo cargado o un artículo elegido).
    assert 'class="boton-peligro" id="boton-cancelar-carga"' in respuesta.text
    assert 'onclick="cancelarTodo()"' in respuesta.text
    assert '"boton-cancelar-carga").style.display' in respuesta.text


def test_ver_cargar_precios_cliente_sin_fichas_muestra_mensaje():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "sin fichas no hay artículos a los que ponerle precio" in respuesta.text


# --- /precios/cargar: panel "Ver negociación" (estado oficial, con los precios ya vigentes) ---

ARTICULOS_NEGOCIACION_DE_PRUEBA = [
    {
        "articulo_id": 1,
        "articulo_nombre": "Tomate Cherry",
        "fresco": True,
        "variacion": "bajo",
        "costo_anterior": 300.0,
        "costo_actual": 280.0,
        "precio_sugerido": 420.0,
        "precio_vigente": 500.0,  # vigente >= sugerido -> ✓
        "utilidad_aproximada": 0.30,
        "compras_sin_precio_excluidas": 0,
        "costo_envase_unidad_venta": 0.0,
        "denominador_tasas": 1.0,
    },
]


def test_ver_cargar_precios_incluye_boton_y_panel_de_negociacion():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=ARTICULOS_NEGOCIACION_DE_PRUEBA) as mock_negociar,
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Ver negociación" in respuesta.text
    assert 'id="panel-negociacion"' in respuesta.text
    assert "abrirNegociacion" in respuesta.text
    assert "cerrarNegociacion" in respuesta.text
    # Reusa la misma función que /negociar, con el mismo cliente_id.
    assert mock_negociar.call_args[0][0] == 1
    # El cuadro embebido tiene que traer los datos reales de la negociación.
    assert "Bajas (frescos que bajaron de costo)" in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "$420" in respuesta.text
    # Mientras el panel está abierto, la barra de navegación (con el ícono
    # "Inicio") queda inhabilitada — evita que un toque se cuele por un
    # resquicio del overlay en Safari/iOS y navegue a /inicio en vez de
    # cerrar el panel.
    assert 'querySelector(".barra-navegacion").style.pointerEvents = "none"' in respuesta.text
    assert 'querySelector(".barra-navegacion").style.pointerEvents = ""' in respuesta.text


def test_ver_cargar_precios_incluye_el_recuadro_de_simulacion():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="recuadro-simulacion"' in respuesta.text
    assert 'id="precio_simulado"' in respuesta.text
    assert 'id="simulacion-resultado"' in respuesta.text
    assert "function actualizarSimulacion()" in respuesta.text


def test_ver_cargar_precios_panel_de_negociacion_no_usa_pendientes_sin_guardar():
    # El panel se arma UNA vez al cargar la pantalla, con lo que ya está
    # guardado — no hay forma de que use nada tipeado en el navegador
    # porque en ese momento todavía no se tipeó nada.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]) as mock_negociar,
    ):
        cliente.get("/precios/cargar?cliente_id=1")

    # Un solo cálculo, en el momento de armar la pantalla.
    mock_negociar.assert_called_once()


def test_ver_cargar_precios_sin_fichas_igual_muestra_boton_de_negociacion():
    # Aunque no haya fichas para cargar precios, el botón de negociación
    # sigue disponible (el panel tiene su propio aviso de "sin fichas").
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
    ):
        respuesta = cliente.get("/precios/cargar?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Ver negociación" in respuesta.text
    assert 'id="panel-negociacion"' in respuesta.text


def test_ver_cargar_precios_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/cargar?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_cargar_precios_error_de_base_da_500():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/precios/cargar")

    assert respuesta.status_code == 500


def _datos_pendientes_cargar_precios(**overrides):
    datos = {
        "cliente_id": "1",
        "pendiente_precio_902": "380",
        "pendiente_original_902": "350.0",
    }
    datos.update(overrides)
    return datos


def test_cargar_precios_guarda_los_pendientes_y_redirige_con_cantidad():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar", data=_datos_pendientes_cargar_precios(), follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/precios?guardado=1"
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 902, "precio": 380.0}])


def test_cargar_precios_varios_pendientes_se_guardan_todos_juntos():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar",
            data={
                "cliente_id": "1",
                "pendiente_precio_901": "520",
                "pendiente_original_901": "500.0",
                "pendiente_precio_902": "380",
                "pendiente_original_902": "350.0",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/precios?guardado=2"
    cambios_guardados = mock_guardar.call_args[0][1]
    assert {"ficha_id": 901, "precio": 520.0} in cambios_guardados
    assert {"ficha_id": 902, "precio": 380.0} in cambios_guardados


def test_cargar_precios_articulo_sin_precio_previo_genera_alta():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar",
            data=_datos_pendientes_cargar_precios(pendiente_original_2=""),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 902, "precio": 380.0}])


def test_cargar_precios_pendiente_igual_al_vigente_no_genera_fila():
    # No se revalida contra la base al guardar, pero si lo que se cargó es
    # igual al vigente que se snapshoteó al elegir el artículo (viaja como
    # "original"), no hace falta una fila nueva en el historial.
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar",
            data=_datos_pendientes_cargar_precios(pendiente_precio_902="350.0"),
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/precios?guardado=0"
    mock_guardar.assert_called_once_with(1, [])


def test_cargar_precios_sin_pendientes_no_guarda_nada():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post("/precios/cargar", data={"cliente_id": "1"}, follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/precios?guardado=0"
    mock_guardar.assert_called_once_with(1, [])


def test_cargar_precios_invalido_da_400():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar", data=_datos_pendientes_cargar_precios(pendiente_precio_902="abc")
        )

    assert respuesta.status_code == 400
    mock_guardar.assert_not_called()


def test_cargar_precios_cero_o_negativo_da_400():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar", data=_datos_pendientes_cargar_precios(pendiente_precio_902="0")
        )

    assert respuesta.status_code == 400
    mock_guardar.assert_not_called()


def test_cargar_precios_error_de_base_da_500():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.guardar_precios_cliente", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post("/precios/cargar", data=_datos_pendientes_cargar_precios())

    assert respuesta.status_code == 500


def test_cargar_precios_cliente_invalido_da_400():
    respuesta = cliente.post("/precios/cargar", data={"cliente_id": "abc"})

    assert respuesta.status_code == 400


def test_cargar_precios_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post("/precios/cargar", data={"cliente_id": "999"})

    assert respuesta.status_code == 404


# --- /precios/cargar-foto: carga de precios leyendo un archivo (foto, PDF o Excel) con IA ---

LISTADO_PRECIOS_LEIDO_DE_PRUEBA = {
    "items": [
        {"articulo": "Tomate Cherry", "precio": 520.0, "confianza": "alta"},
        {"articulo": "algo ilegible", "precio": 999.0, "confianza": "baja"},
    ]
}


def test_ver_cargar_foto_precios_sin_cliente_muestra_selector():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/cargar-foto")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para subirle un listado de precios." in respuesta.text


def test_ver_cargar_foto_precios_con_cliente_muestra_boton_de_subir():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/cargar-foto?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'id="archivo"' in respuesta.text
    assert 'accept="image/*,.pdf,.xlsx"' in respuesta.text


def test_ver_cargar_foto_precios_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/precios/cargar-foto?cliente_id=999")

    assert respuesta.status_code == 404


def test_leer_foto_precios_matchea_y_muestra_pantalla_de_revision():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.extraer_listado_precios_de_imagenes", return_value=LISTADO_PRECIOS_LEIDO_DE_PRUEBA) as mock_extraer,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.jpg", b"contenido falso de una foto", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert "Tomate Cherry" in respuesta.text
    assert 'value="520.0"' in respuesta.text
    assert "Precio vigente: $500" in respuesta.text
    assert "⚠ revisar" in respuesta.text  # el segundo ítem, no matcheado
    mock_extraer.assert_called_once_with([b"contenido falso de una foto"])


def test_leer_foto_precios_pantalla_revision_incluye_boton_guardar_y_generar_listado():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=PRECIOS_VIGENTES_DE_PRUEBA),
        patch("app.main.extraer_listado_precios_de_imagenes", return_value=LISTADO_PRECIOS_LEIDO_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.jpg", b"contenido falso de una foto", "image/jpeg")},
        )

    assert 'id="boton-guardar-y-generar"' in respuesta.text
    assert "Guardar y generar listado" in respuesta.text
    assert "guardarYGenerarListado" in respuesta.text
    assert "/precios/cargar-foto/guardar-y-exportar-pdf" in respuesta.text
    assert "/precios/cargar-foto/guardar-y-exportar-excel" in respuesta.text
    # El botón "Exportar sin guardar" que había antes ya no va.
    assert 'id="boton-exportar"' not in respuesta.text


def test_leer_foto_precios_pdf_convierte_paginas_a_imagenes():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.imagenes_desde_pdf", return_value=[b"pagina1", b"pagina2"]) as mock_pdf,
        patch("app.main.extraer_listado_precios_de_imagenes", return_value=LISTADO_PRECIOS_LEIDO_DE_PRUEBA) as mock_extraer,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.pdf", b"contenido falso de un pdf", "application/pdf")},
        )

    assert respuesta.status_code == 200
    mock_pdf.assert_called_once_with(b"contenido falso de un pdf")
    mock_extraer.assert_called_once_with([b"pagina1", b"pagina2"])


def test_leer_foto_precios_excel_convierte_a_texto():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.texto_desde_excel", return_value="Tomate Cherry | 520") as mock_excel,
        patch("app.main.extraer_listado_precios_de_texto", return_value=LISTADO_PRECIOS_LEIDO_DE_PRUEBA) as mock_extraer,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={
                "archivo": (
                    "lista.xlsx",
                    b"contenido falso de un excel",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert respuesta.status_code == 200
    mock_excel.assert_called_once_with(b"contenido falso de un excel")
    mock_extraer.assert_called_once_with("Tomate Cherry | 520")


def test_leer_foto_precios_extension_no_soportada_muestra_mensaje_claro():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.docx", b"contenido falso", "application/msword")},
        )

    assert respuesta.status_code == 400
    assert "No se pudo reconocer el tipo de archivo" in respuesta.text


def test_leer_foto_precios_error_de_lectura_no_rompe_muestra_mensaje_claro():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.extraer_listado_precios_de_imagenes", side_effect=RuntimeError("la IA no pudo leer nada")),
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 500
    assert "No se pudo leer el archivo" in respuesta.text
    assert "la IA no pudo leer nada" in respuesta.text


def test_leer_foto_precios_sin_ningun_articulo_muestra_mensaje_claro():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=[]),
        patch("app.main.extraer_listado_precios_de_imagenes", return_value={"items": []}),
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "1"},
            files={"archivo": ("lista.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 400
    assert "No se encontró ningún artículo" in respuesta.text


def test_leer_foto_precios_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/precios/cargar-foto",
            data={"cliente_id": "999"},
            files={"archivo": ("lista.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 404


def test_confirmar_carga_foto_precios_guarda_y_sube_el_archivo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.subir_archivo_comanda", return_value="2026-08-16/dia-123-abc.jpg") as mock_subir,
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto/confirmar",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "foto",
                "archivo_preview": "data:image/jpeg;base64,QUJD",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "500.0",
                "item_0_precio_nuevo": "520",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/precios?guardado=1"
    mock_subir.assert_called_once_with(b"ABC", "Día", "jpg", "image/jpeg")
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 901, "precio": 520.0}], foto_ruta="2026-08-16/dia-123-abc.jpg")


def test_confirmar_carga_foto_precios_renglon_descartado_no_se_guarda():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.subir_archivo_comanda") as mock_subir,
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto/confirmar",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "foto",
                "archivo_preview": "",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "500.0",
                "item_0_precio_nuevo": "520",
                "item_0_descartar": "on",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_subir.assert_not_called()
    mock_guardar.assert_called_once_with(1, [], foto_ruta=None)


def test_confirmar_carga_foto_precios_error_al_subir_archivo_guarda_igual_sin_archivo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.subir_archivo_comanda", side_effect=RuntimeError("Storage caído")),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto/confirmar",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "foto",
                "archivo_preview": "data:image/jpeg;base64,QUJD",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "500.0",
                "item_0_precio_nuevo": "520",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 901, "precio": 520.0}], foto_ruta=None)


def test_confirmar_carga_foto_precios_pdf_sube_con_extension_y_content_type_correctos():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PRECIOS_DE_PRUEBA),
        patch("app.main.subir_archivo_comanda", return_value="ruta.pdf") as mock_subir,
        patch("app.main.guardar_precios_cliente"),
    ):
        cliente.post(
            "/precios/cargar-foto/confirmar",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "pdf",
                "archivo_preview": "data:application/pdf;base64,QUJD",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "",
                "item_0_precio_nuevo": "520",
            },
            follow_redirects=False,
        )

    mock_subir.assert_called_once_with(b"ABC", "Día", "pdf", "application/pdf")


def test_confirmar_carga_foto_precios_cliente_invalido_da_400():
    respuesta = cliente.post("/precios/cargar-foto/confirmar", data={"cliente_id": "abc"})

    assert respuesta.status_code == 400


def test_confirmar_carga_foto_precios_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post("/precios/cargar-foto/confirmar", data={"cliente_id": "999"})

    assert respuesta.status_code == 404


# --- /precios/cargar-foto/guardar-y-exportar-pdf y guardar-y-exportar-excel: "Guardar y generar
# listado" de Carga Foto — guarda de verdad los renglones (subiendo la foto si corresponde) y arma el
# PDF/Excel con lo recién guardado ---


def test_guardar_y_exportar_precios_cargar_foto_pdf_guarda_y_devuelve_archivo():
    precios_tras_guardar = [{"ficha_id": 901, "articulo_id": 1, "precio": 520.0, "vigente_desde": HOY_DE_PRUEBA}]
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_precios_vigentes_por_cliente", return_value=precios_tras_guardar),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto/guardar-y-exportar-pdf",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "foto",
                "archivo_preview": "",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "500.0",
                "item_0_precio_nuevo": "520",
            },
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.headers["x-cantidad-guardada"] == "1"
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 901, "precio": 520.0}], foto_ruta=None)

    texto = _texto_sin_leyenda_de_respuesta(respuesta.content)
    assert "Tomate Cherry" in texto
    assert texto.count("Nuevo precio") == 1


def test_guardar_y_exportar_precios_cargar_foto_excel_sube_el_archivo_y_devuelve_excel():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_EXPORTACION_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_EXPORTACION_DE_PRUEBA),
        patch(
            "app.main.listar_precios_vigentes_por_cliente",
            return_value=_precios_vigentes_exportacion(HOY_DE_PRUEBA),
        ),
        patch("app.main.listar_precios_anteriores_por_cliente", return_value=[]),
        patch("app.main.subir_archivo_comanda", return_value="2026-08-16/dia-123-abc.jpg") as mock_subir,
        patch("app.main.guardar_precios_cliente") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/precios/cargar-foto/guardar-y-exportar-excel",
            data={
                "cliente_id": "1",
                "cantidad_renglones": "1",
                "tipo_archivo": "foto",
                "archivo_preview": "data:image/jpeg;base64,QUJD",
                "item_0_ficha_id": "901",
                "item_0_precio_original": "500.0",
                "item_0_precio_nuevo": "520",
            },
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in respuesta.headers["content-disposition"]
    assert respuesta.headers["x-cantidad-guardada"] == "1"
    mock_subir.assert_called_once_with(b"ABC", "Día", "jpg", "image/jpeg")
    mock_guardar.assert_called_once_with(1, [{"ficha_id": 901, "precio": 520.0}], foto_ruta="2026-08-16/dia-123-abc.jpg")


def test_guardar_y_exportar_precios_cargar_foto_pdf_cliente_invalido_da_400():
    respuesta = cliente.post("/precios/cargar-foto/guardar-y-exportar-pdf", data={"cliente_id": "abc"})

    assert respuesta.status_code == 400


def test_guardar_y_exportar_precios_cargar_foto_pdf_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post("/precios/cargar-foto/guardar-y-exportar-pdf", data={"cliente_id": "999"})

    assert respuesta.status_code == 404


# --- /precios/generar-listado: placeholder ---


def test_ver_generar_listado_precios_muestra_en_construccion_y_vuelve_a_precios():
    respuesta = cliente.get("/precios/generar-listado")

    assert respuesta.status_code == 200
    assert "En construcción" in respuesta.text
    assert 'href="/precios"' in respuesta.text


def test_ver_resultado_negociacion_muestra_en_construccion_y_vuelve_a_precios():
    respuesta = cliente.get("/precios/resultado-negociacion")

    assert respuesta.status_code == 200
    assert "En construcción" in respuesta.text
    assert 'href="/precios"' in respuesta.text


def test_ver_inicio_muestra_las_8_areas():
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert '<h1>Frutamax</h1>' in respuesta.text
    assert 'href="/compras"' in respuesta.text
    assert 'href="/comercial"' in respuesta.text
    assert 'href="/logistica"' in respuesta.text
    assert 'href="/deposito"' in respuesta.text
    assert 'href="/gerencia"' in respuesta.text
    assert 'href="/administracion"' in respuesta.text
    assert 'href="/puesto"' in respuesta.text
    assert 'href="/sistema"' in respuesta.text


def test_ver_inicio_usa_el_nombre_de_empresa_configurado():
    # Mismo código para varias empresas, cada una con su propia base: el
    # nombre que se ve en /inicio sale de la variable de entorno
    # NOMBRE_EMPRESA (ver app/main.py), no queda fijo en "Frutamax".
    with patch.dict(templates.env.globals, {"NOMBRE_EMPRESA": "Palmala"}):
        respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert "<h1>Palmala</h1>" in respuesta.text
    assert "<title>Palmala</title>" in respuesta.text
    assert "Frutamax" not in respuesta.text


def test_ver_inicio_deposito_y_logistica_ya_no_dicen_proximamente():
    # Depósito y Logística dejaron de ser "en construcción" (existen
    # Recepción y Retiro) — sus tarjetas en /inicio tienen que verse igual
    # de activas que Compras o Comercial, no atenuadas como Gerencia (que
    # sigue sin nada adentro).
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert "Depósito (Próximamente)" not in respuesta.text
    assert '<a class="boton-area" href="/deposito">' in respuesta.text
    assert "Logística (Próximamente)" not in respuesta.text
    assert '<a class="boton-area" href="/logistica">' in respuesta.text
    # Gerencia ya tiene contenido (Auditoría): dejó de ser placeholder.
    assert "Gerencia (Próximamente)" not in respuesta.text
    assert '<a class="boton-area" href="/gerencia">' in respuesta.text


def test_ver_inicio_puesto_es_el_primer_boton():
    # El Puesto es lo que más se usa en el día: va arriba de todo.
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert respuesta.text.index('href="/puesto"') < respuesta.text.index('href="/compras"')


def test_ver_inicio_administracion_y_puesto_son_modulos_plenos():
    # Administración ya es un módulo real (Ingresos a Depósito), igual que
    # Puesto: los dos van con botón pleno, sin "Próximamente".
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert '<a class="boton-area" href="/administracion">' in respuesta.text
    assert "Administración (Próximamente)" not in respuesta.text
    assert '<a class="boton-area" href="/puesto">' in respuesta.text
    assert "Puesto (Próximamente)" not in respuesta.text


def test_ver_inicio_usa_los_mismos_iconos_que_la_barra_de_navegacion():
    # Regresión: el mismo ícono que representa a cada sector en la barrita
    # de arriba tiene que estar en su tarjeta de la home, para un lenguaje
    # visual consistente.
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    for sector in ("compras", "comercial", "logistica", "deposito", "gerencia", "administracion", "puesto", "sistema"):
        assert SECTORES[sector]["icono"] in respuesta.text


def test_barra_navegacion_muestra_el_nombre_de_la_empresa():
    # Con dos empresas corriendo el mismo sistema, el cartelito de la
    # esquina dice siempre en cuál estás parado. Default: Frutamax.
    respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    assert 'class="barra-empresa">Frutamax</span>' in respuesta.text


def test_nombre_empresa_para_archivo_saca_acentos_y_caracteres_raros():
    from app.main import _nombre_empresa_para_archivo

    with patch("app.main.NOMBRE_EMPRESA", "Verdulería Sur"):
        assert _nombre_empresa_para_archivo() == "Verduleria_Sur"
    with patch("app.main.NOMBRE_EMPRESA", "Frutamax"):
        assert _nombre_empresa_para_archivo() == "Frutamax"


def test_nombre_archivo_disponibles_usa_el_nombre_de_empresa():
    from app.main import _nombre_archivo_disponibles

    with patch("app.main.NOMBRE_EMPRESA", "Palmala"):
        assert _nombre_archivo_disponibles(date(2026, 8, 14), 1) == "Disponibles_Palmala_14_Ago_2026.xlsx"
    # Default intacto: para Frutamax el nombre es exactamente el de siempre.
    assert _nombre_archivo_disponibles(date(2026, 8, 14), 1) == "Disponibles_Frutamax_14_Ago_2026.xlsx"


def test_ver_comercial_muestra_los_tres_accesos():
    respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    assert 'href="/precios"' in respuesta.text
    assert 'href="/clientes"' in respuesta.text
    assert 'href="/fichas"' in respuesta.text
    assert "Precios" in respuesta.text
    assert "Clientes" in respuesta.text
    assert "Fichas logísticas" in respuesta.text
    # Negociar precios y Lista de Precios ya no están sueltos acá — viven
    # ordenados adentro de la botonera de Precios.
    assert 'href="/negociar"' not in respuesta.text


def test_ver_comercial_muestra_en_su_banner_solo_las_alertas_que_le_tocan():
    # El criterio del mapa de módulos: la alerta aparece donde está la persona
    # que puede resolverla. Comercial factura, así que las compras sin precio
    # también son suyas; los retiros de Logística no, y no le ensucian el
    # banner — un banner que se llena de cosas ajenas se deja de mirar.
    foto = _foto_alertas({
        "compras_sin_precio": (4, date(2026, 7, 30)),
        "articulos_incotizables": (2, None),
        "retiros_sin_hacer": (9, date(2026, 8, 1)),
    })
    with patch("app.main.listar_estado_alertas", return_value=foto) as mock_foto:
        respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    # UNA sola consulta para todo el banner, tenga el registro 15 alertas o 100.
    mock_foto.assert_called_once_with()
    assert "Compras sin precio de compra cargado (4)" in respuesta.text
    assert 'href="/compras/pendientes"' in respuesta.text
    assert "sin ficha logística o sin precio de venta" in respuesta.text
    assert "Mercadería sin retirar" not in respuesta.text
    # Arriba de los tres botones, no mezclado ni después.
    assert respuesta.text.index("Compras sin precio") < respuesta.text.index('href="/precios"')


def test_ver_comercial_sin_alertas_no_muestra_banner():
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    assert 'class="banner-avisos"' not in respuesta.text
    assert "sin precio de compra cargado" not in respuesta.text


def test_ver_comercial_error_al_leer_las_alertas_no_rompe_la_pantalla():
    # El banner es un aviso, no algo crítico: si la consulta falla, la
    # pantalla sigue funcionando (sin banner), nunca un 500.
    with patch("app.main.listar_estado_alertas", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    assert 'class="banner-avisos"' not in respuesta.text
    assert 'href="/precios"' in respuesta.text
    assert 'href="/inicio"' in respuesta.text


def test_ver_compras_muestra_en_su_banner_solo_las_alertas_que_le_tocan():
    # El comprador ve apenas entra lo que él puede resolver: los precios que
    # faltan cargar y las Guías R cuyo costo no cierra porque falta un precio
    # de compra. Lo de Depósito no es suyo y no aparece.
    foto = _foto_alertas({
        "compras_sin_precio": (4, date(2026, 7, 30)),
        "guias_r_costo_incompleto": (1, date(2026, 8, 5)),
        "recepciones_pendientes": (3, date(2026, 8, 2)),
    })
    with patch("app.main.listar_estado_alertas", return_value=foto):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "Compras sin precio de compra cargado (4)" in respuesta.text
    assert 'href="/compras/pendientes"' in respuesta.text
    assert "Guías R esperando precio (1)" in respuesta.text
    assert "Mercadería sin recepcionar" not in respuesta.text
    # Arriba de los botones de carga, no mezclado ni después.
    assert respuesta.text.index("Compras sin precio") < respuesta.text.index('href="/compras/nueva/manual"')


def test_ver_compras_sin_alertas_no_muestra_banner():
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "sin precio de compra cargado" not in respuesta.text
    assert 'class="banner-avisos"' not in respuesta.text


def test_ver_compras_error_al_leer_las_alertas_no_rompe_la_pantalla():
    with patch("app.main.listar_estado_alertas", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert 'class="banner-avisos"' not in respuesta.text
    assert 'href="/compras/nueva/manual"' in respuesta.text


def test_banner_corre_y_duplica_el_contenido_para_el_loop():
    # La cinta se anima proporcional a cuántos avisos lleva, y el contenido va
    # dos veces (la copia oculta a lectores de pantalla) para que el corte del
    # loop no se note.
    foto = _foto_alertas({
        "compras_sin_precio": (4, date(2026, 7, 30)),
        "guias_r_costo_incompleto": (1, date(2026, 8, 5)),
    })
    with patch("app.main.listar_estado_alertas", return_value=foto):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert 'class="banner-cinta" style="animation-duration: 24s;"' in respuesta.text
    assert respuesta.text.count("Compras sin precio de compra cargado (4)") == 2
    assert '<span class="copia" aria-hidden="true">' in respuesta.text


def test_banner_avisa_cuando_la_foto_esta_vencida_aunque_no_haya_ninguna_alerta():
    # La misma trampa que en Auditoría: un banner vacío porque está todo bien
    # no puede verse igual que un banner vacío porque nadie calculó nada. Si
    # el cálculo automático se murió, el banner aparece igual y lo dice.
    vieja = datetime.now(ARGENTINA_TEST) - timedelta(hours=40)
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas(calculada_el=vieja)):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "no se actualizaron desde entonces" in respuesta.text
    assert 'href="/auditoria"' in respuesta.text


def test_banner_avisa_cuando_las_alertas_nunca_se_calcularon():
    with patch("app.main.listar_estado_alertas", return_value=[]):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "las alertas todavía no se calcularon nunca" in respuesta.text


def test_banner_muestra_la_alerta_que_no_se_pudo_calcular():
    # Que una alerta no se haya podido calcular ES la noticia: sale igual,
    # aunque su último conteo diera cero. Desaparecer en silencio es la peor
    # falla posible en un sistema de alertas.
    foto = _foto_alertas()
    for fila in foto:
        if fila["codigo"] == "compras_sin_precio":
            fila["error"] = "no se pudo conectar"
    with patch("app.main.listar_estado_alertas", return_value=foto):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "No se pudo calcular: Compras sin precio de compra cargado" in respuesta.text


def test_toda_alerta_con_modulo_se_ve_en_la_pantalla_de_ese_modulo():
    """La garantía que no se negocia: una alerta no puede quedar invisible.

    Recorre los módulos que declara EL REGISTRO —no una lista escrita a mano
    acá— y entra a la pantalla de cada uno con esa alerta en 1. Si un módulo
    nuevo se declara y su pantalla no incluye el banner, la alerta existiría
    en Auditoría y no la vería nunca el que puede resolverla. Este test es lo
    que lo impide.
    """
    from app.main import ALERTAS

    modulos = sorted({modulo for definicion in ALERTAS for modulo in definicion.modulos})
    assert modulos, "El registro se quedó sin ninguna alerta con módulo"

    for modulo in modulos:
        definicion = next(d for d in ALERTAS if modulo in d.modulos)
        foto = _foto_alertas({definicion.codigo: (1, date(2026, 8, 1))})
        with ExitStack() as pila:
            pila.enter_context(patch("app.main.listar_estado_alertas", return_value=foto))
            # Los datos que piden algunas de esas pantallas para poder abrir.
            pila.enter_context(patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA))
            pila.enter_context(patch("app.main.obtener_uso_storage_bucket", return_value=None))
            respuesta = cliente.get(f"/{modulo}")

        assert respuesta.status_code == 200, (
            f"/{modulo} no abre: o el módulo está mal escrito en la alerta "
            f"{definicion.codigo}, o su pantalla necesita datos que este test no le da"
        )
        assert 'class="banner-avisos"' in respuesta.text, (
            f"La alerta {definicion.codigo} dice ir a {modulo}, pero /{modulo} "
            f'no muestra el banner: le falta el include de "_banner_alertas.html"'
        )


def test_ver_compras_muestra_el_aviso_cuando_viene_en_la_url():
    # El aviso del Cancelar de la carga manual: cuántas se cancelaron y
    # cuántas no se pudieron (ya recepcionadas) — no puede perderse por
    # volver al hub en vez de a Buscar.
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get(
            "/compras?aviso=3+compras+canceladas.+2+no+se+pudieron+eliminar%3A+ya+fueron+retiradas+o+recepcionadas."
        )

    assert respuesta.status_code == 200
    assert "3 compras canceladas. 2 no se pudieron eliminar: ya fueron retiradas o recepcionadas." in respuesta.text


def test_ver_compras_sin_aviso_no_muestra_cartel_vacio():
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert "canceladas" not in respuesta.text


def test_ver_logistica_muestra_solo_clark_y_consultar():
    # Clark es el único que entra a tildar. Carro y Cooperativa nacen
    # retirados solos, y Pases lo retira Depósito desde su propio botón —
    # ninguno de los tres lleva botón acá (las rutas siguen vivas).
    respuesta = cliente.get("/logistica")

    assert respuesta.status_code == 200
    assert "Logística" in respuesta.text
    assert 'href="/logistica/retiro/Clark"' in respuesta.text
    assert 'href="/logistica/consultar"' in respuesta.text
    assert 'href="/logistica/retiro/Carro"' not in respuesta.text
    assert 'href="/logistica/retiro/Pases"' not in respuesta.text
    assert "Cooperativa" not in respuesta.text
    assert "En construcción" not in respuesta.text
    assert 'href="/inicio"' in respuesta.text


# --- /logistica/retiro/{tipo_retiro}: retiro de mercadería en el Mercado Central ---

COMPRAS_PENDIENTES_RETIRO_DE_PRUEBA = [
    {
        "id": 1, "guia_id": 105, "guia_punto": 1, "articulo_nombre": "Tomate Cherry", "unidad_compra": "kilo",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41", "fecha_operacion": HOY_DE_PRUEBA,
        "cantidad_cajones": 40, "contenido_por_cajon": 20, "cantidad_kilos": 800, "cantidad_fraccion": None,
    },
    {
        "id": 2, "guia_id": 105, "guia_punto": 2, "articulo_nombre": "Mango", "unidad_compra": "unidad",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "fecha_operacion": HOY_DE_PRUEBA - timedelta(days=3),
        "cantidad_cajones": 10, "contenido_por_cajon": 12, "cantidad_kilos": None, "cantidad_fraccion": 120,
    },
]


def test_ver_logistica_retiro_muestra_la_fecha_y_resalta_las_viejas():
    # El que retira tiene que ver de cuándo es lo que levanta: la fecha va
    # en cada renglón, y con más de un día se resalta para que se note.
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_compras_pendientes_retiro", return_value=COMPRAS_PENDIENTES_RETIRO_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    # La de hoy: fecha visible, sin resaltar.
    assert "Compra del 06/08" in respuesta.text
    # La de hace 3 días: resaltada con advertencia.
    assert "⚠ Compra del 03/08" in respuesta.text
    assert 'class="fecha-compra fecha-vieja"' in respuesta.text


def test_ver_logistica_retiro_agrupa_por_guia_sin_mostrar_el_numero():
    # A diferencia de Recepción en Depósito, acá no interesa el número de
    # guía — el título de cada tarjeta es directamente el proveedor.
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=COMPRAS_PENDIENTES_RETIRO_DE_PRUEBA) as mock_listar,
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    mock_listar.assert_called_once_with("Clark")
    assert "Guía" not in respuesta.text
    assert "105" not in respuesta.text
    # Puesto grande primero, proveedor chico debajo (busca antes por dónde
    # ir que confirma quién es).
    assert '<p class="puesto-nombre">N07P41</p>' in respuesta.text
    assert '<p class="proveedor-nombre">Saturno</p>' in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "Mango" in respuesta.text
    # Sin ?origen en la URL, cae en logistica (el default).
    assert 'action="/logistica/retiro/Clark/1/retirar?origen=logistica"' in respuesta.text
    assert 'action="/logistica/retiro/Clark/1/cancelar?origen=logistica"' in respuesta.text


def test_ver_logistica_retiro_sin_origen_muestra_el_icono_y_el_volver_de_logistica():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    assert 'href="/logistica" aria-label="Ir a Logística"' in respuesta.text
    assert '<a class="volver" href="/logistica">Volver a Logística</a>' in respuesta.text


def test_ver_logistica_retiro_con_origen_deposito_muestra_el_icono_y_el_volver_de_deposito():
    # Entrando desde el botón "Retirar Mercadería" de Depósito: la barrita
    # y el "Volver" tienen que ser de Depósito, no de Logística.
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Pases?origen=deposito")

    assert respuesta.status_code == 200
    assert 'href="/deposito" aria-label="Ir a Depósito"' in respuesta.text
    assert '<a class="volver" href="/deposito">Volver a Depósito</a>' in respuesta.text
    assert "Volver a Logística" not in respuesta.text


def test_ver_logistica_retiro_origen_invalido_cae_en_logistica():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark?origen=algo-raro")

    assert respuesta.status_code == 200
    assert '<a class="volver" href="/logistica">Volver a Logística</a>' in respuesta.text


def test_ver_logistica_retiro_tipo_invalido_da_404():
    respuesta = cliente.get("/logistica/retiro/Moto")

    assert respuesta.status_code == 404


def test_ver_logistica_retiro_sin_pendientes_muestra_mensaje():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Pases")

    assert respuesta.status_code == 200
    assert "No hay compras pendientes de retiro por Pases." in respuesta.text


def test_retirar_compra_marca_retirada_y_redirige():
    with patch("app.main.marcar_compra_retirada", return_value=None) as mock_marcar:
        respuesta = cliente.post("/logistica/retiro/Clark/1/retirar", follow_redirects=False)

    assert respuesta.status_code == 303
    # Sin origen en el POST, el redirect vuelve con el default (logistica),
    # y con ?procesado=1 para que la tarjeta efímera aparezca en la próxima carga.
    assert respuesta.headers["location"] == "/logistica/retiro/Clark?origen=logistica&procesado=1"
    mock_marcar.assert_called_once_with(1, "logistica", None)


def test_retirar_compra_conserva_el_origen_deposito_en_el_redirect():
    # El botón "Retirar Mercadería" de Depósito entra con ?origen=deposito
    # — al tocar "Retirado" el redirect tiene que conservarlo, si no la
    # persona vuelve a Logística en vez de a Depósito.
    with patch("app.main.marcar_compra_retirada", return_value=None):
        respuesta = cliente.post(
            "/logistica/retiro/Pases/1/retirar?origen=deposito", follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Pases?origen=deposito&procesado=1"


def test_retirar_compra_origen_invalido_cae_en_logistica():
    with patch("app.main.marcar_compra_retirada", return_value=None):
        respuesta = cliente.post(
            "/logistica/retiro/Clark/1/retirar?origen=cualquier-cosa", follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Clark?origen=logistica&procesado=1"


def test_retirar_compra_con_cantidad_cajones_retirada_la_pasa_a_marcar_compra_retirada():
    with patch("app.main.marcar_compra_retirada", return_value=None) as mock_marcar:
        respuesta = cliente.post(
            "/logistica/retiro/Clark/1/retirar",
            data={"cantidad_cajones_retirada": "8.5"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(1, "logistica", 8.5)


def test_retirar_compra_con_cantidad_no_numerica_da_error_sin_llamar_a_marcar():
    with (
        patch("app.main.marcar_compra_retirada", return_value=None) as mock_marcar,
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.post(
            "/logistica/retiro/Clark/1/retirar",
            data={"cantidad_cajones_retirada": "no-es-un-numero"},
        )

    assert respuesta.status_code == 400
    assert "tiene que ser un número" in respuesta.text
    mock_marcar.assert_not_called()


def test_retirar_compra_con_cantidad_negativa_da_error_sin_llamar_a_marcar():
    with (
        patch("app.main.marcar_compra_retirada", return_value=None) as mock_marcar,
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.post(
            "/logistica/retiro/Clark/1/retirar",
            data={"cantidad_cajones_retirada": "-3"},
        )

    assert respuesta.status_code == 400
    assert "no puede ser negativa" in respuesta.text
    mock_marcar.assert_not_called()


def test_cancelar_retiro_compra_marca_cancelada_y_redirige():
    with patch("app.main.marcar_compra_cancelada", return_value=None) as mock_marcar:
        respuesta = cliente.post("/logistica/retiro/Carro/1/cancelar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Carro?origen=logistica&procesado=1"
    mock_marcar.assert_called_once_with(1, "logistica")


def test_cancelar_retiro_compra_conserva_el_origen_deposito_en_el_redirect():
    with patch("app.main.marcar_compra_cancelada", return_value=None):
        respuesta = cliente.post(
            "/logistica/retiro/Pases/1/cancelar?origen=deposito", follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Pases?origen=deposito&procesado=1"


def test_retirar_compra_error_de_base_muestra_mensaje():
    with (
        patch("app.main.marcar_compra_retirada", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.post("/logistica/retiro/Clark/1/retirar")

    assert respuesta.status_code == 500
    assert "No se pudo marcar como retirada" in respuesta.text


def test_ver_logistica_retiro_confirmacion_es_en_el_lugar_no_confirm_nativo():
    # Punto central: nada de confirm() del navegador. El único confirm()
    # que puede quedar en el HTML es el que arma Playwright/el usuario en
    # otro lado — acá se verifica que el markup usa las funciones JS
    # propias (mostrarConfirmacion/ocultarConfirmacion), no un onsubmit
    # con confirm(...).
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=COMPRAS_PENDIENTES_RETIRO_DE_PRUEBA),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    # El único confirm() nativo que quedaba (onsubmit de Cancelado) se
    # saca del todo — la comprobación pasa a ser en el lugar, con las
    # funciones JS propias.
    assert "onsubmit=\"return confirm(" not in respuesta.text
    assert "onclick=\"mostrarConfirmacion('1', 'retirar')\"" in respuesta.text
    assert "onclick=\"mostrarConfirmacion('1', 'cancelar')\"" in respuesta.text
    assert '¿Retirar Tomate Cherry?' in respuesta.text
    assert '¿Cancelar el retiro de Tomate Cherry?' in respuesta.text


PROCESADOS_HOY_DE_PRUEBA = [
    {
        "id": 1, "articulo_nombre": "Tomate Cherry", "unidad_compra": "kilo",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 40, "contenido_por_cajon": 20, "cantidad_cajones_retirada": 38,
        "estado_retiro": "retirado", "retiro_procesado_el": datetime(2026, 8, 17, 13, 0, tzinfo=timezone.utc),
        "estado": "pendiente",
    },
    {
        "id": 2, "articulo_nombre": "Mango", "unidad_compra": "unidad",
        "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10, "contenido_por_cajon": 12, "cantidad_cajones_retirada": None,
        "estado_retiro": "cancelado", "retiro_procesado_el": datetime(2026, 8, 17, 12, 30, tzinfo=timezone.utc),
        "estado": "pendiente",
    },
]


def test_ver_logistica_retiro_con_procesado_muestra_la_tarjeta_efimera():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=PROCESADOS_HOY_DE_PRUEBA),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark?procesado=1")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' in respuesta.text
    assert "Retiraste" in respuesta.text
    assert "Tomate Cherry" in respuesta.text
    assert "N07P41" in respuesta.text
    assert 'action="/logistica/retiro/Clark/1/deshacer?origen=logistica"' in respuesta.text


def test_ver_logistica_retiro_con_procesado_cancelado_dice_cancelaste():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=PROCESADOS_HOY_DE_PRUEBA),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark?procesado=2")

    assert respuesta.status_code == 200
    assert "Cancelaste" in respuesta.text
    assert "Mango" in respuesta.text


def test_ver_logistica_retiro_sin_procesado_no_muestra_tarjeta_efimera():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=PROCESADOS_HOY_DE_PRUEBA),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' not in respuesta.text


def test_ver_logistica_retiro_procesado_que_no_esta_en_hoy_no_muestra_tarjeta():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark?procesado=999")

    assert respuesta.status_code == 200
    assert 'id="tarjeta-efimera"' not in respuesta.text


def test_ver_logistica_retiro_panel_procesados_hoy_muestra_hora_y_deshacer():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=PROCESADOS_HOY_DE_PRUEBA),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    assert "Ver procesados hoy (2)" in respuesta.text
    assert "Retirado a las 10:00" in respuesta.text  # 13:00 UTC -> 10:00 ARG
    assert "Cancelado a las 09:30" in respuesta.text
    assert 'action="/logistica/retiro/Clark/1/deshacer?origen=logistica"' in respuesta.text
    assert 'action="/logistica/retiro/Clark/2/deshacer?origen=logistica"' in respuesta.text


def test_ver_logistica_retiro_panel_procesados_hoy_bloqueado_muestra_aviso_no_boton():
    procesado_recepcionado = dict(PROCESADOS_HOY_DE_PRUEBA[0], estado="recepcionado")
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[procesado_recepcionado]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    assert "no se puede deshacer" in respuesta.text
    assert 'action="/logistica/retiro/Clark/1/deshacer' not in respuesta.text


def test_ver_logistica_retiro_sin_procesados_hoy_muestra_mensaje_vacio():
    with (
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/retiro/Clark")

    assert respuesta.status_code == 200
    assert "Ver procesados hoy" in respuesta.text
    assert "Todavía no se procesó nada hoy." in respuesta.text


def test_deshacer_retiro_compra_ruta_redirige_sin_procesado():
    with patch("app.main.deshacer_retiro_compra", return_value=None) as mock_deshacer:
        respuesta = cliente.post("/logistica/retiro/Clark/1/deshacer", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Clark?origen=logistica"
    mock_deshacer.assert_called_once_with(1)


def test_deshacer_retiro_compra_ruta_conserva_origen():
    with patch("app.main.deshacer_retiro_compra", return_value=None):
        respuesta = cliente.post(
            "/logistica/retiro/Pases/1/deshacer?origen=deposito", follow_redirects=False
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/logistica/retiro/Pases?origen=deposito"


def test_deshacer_retiro_compra_ruta_bloqueado_da_400():
    with (
        patch("app.main.deshacer_retiro_compra", side_effect=ValueError("Esta compra ya fue procesada en Depósito, no se puede deshacer el retiro.")),
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.post("/logistica/retiro/Clark/1/deshacer")

    assert respuesta.status_code == 400
    assert "no se puede deshacer el retiro" in respuesta.text


def test_deshacer_retiro_compra_ruta_error_de_base_da_500():
    with (
        patch("app.main.deshacer_retiro_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_compras_pendientes_retiro", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_retiro", return_value=[]),
    ):
        respuesta = cliente.post("/logistica/retiro/Clark/1/deshacer")

    assert respuesta.status_code == 500
    assert "No se pudo deshacer" in respuesta.text


def test_ver_deposito_muestra_el_acceso_a_recepcion():
    # /deposito dejó de ser "en construcción": ahora es un hub, como
    # /compras o /comercial, con Recepción como primer acceso real.
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert 'href="/deposito/recepcion"' in respuesta.text
    assert "Recepción" in respuesta.text
    assert "En construcción" not in respuesta.text


def test_ver_deposito_muestra_el_acceso_a_retirar_mercaderia():
    # El personal de depósito retira puntualmente por Pases (no Clark ni
    # Carro, que son de Logística) — el botón va directo a esa pantalla,
    # sin pasar por el hub de /logistica con los 3 botones.
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    # Con ?origen=deposito, para que la barrita y el "Volver" de esa
    # pantalla sean de Depósito (de donde realmente se entró), no Logística.
    assert 'href="/logistica/retiro/Pases?origen=deposito"' in respuesta.text
    assert "Retirar Mercadería" in respuesta.text


def test_ver_deposito_muestra_el_acceso_a_ingresar_mercaderia():
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert 'href="/deposito/ingresar"' in respuesta.text
    assert "Ingresar Mercadería" in respuesta.text


def test_ver_deposito_muestra_el_aviso_cuando_viene_en_la_url():
    respuesta = cliente.get("/deposito?aviso=Ingresada+sin+precio.+El+comprador+tiene+que+cargar+el+costo.")

    assert respuesta.status_code == 200
    assert '<div class="aviso">Ingresada sin precio. El comprador tiene que cargar el costo.</div>' in respuesta.text


def test_ver_deposito_sin_aviso_no_muestra_el_cartel():
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert '<div class="aviso">' not in respuesta.text


# --- /deposito/ingresar: ingreso directo de mercadería, sin Logística ni Recepción ---


def test_ver_ingresar_mercaderia_sin_proveedor_muestra_formulario_de_proveedor():
    with patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA):
        respuesta = cliente.get("/deposito/ingresar")

    assert respuesta.status_code == 200
    assert "Código de puesto" in respuesta.text
    assert "N07P41" in respuesta.text
    assert 'action="/deposito/ingresar/proveedor"' in respuesta.text
    assert "PROVEEDORES_LISTA" in respuesta.text
    # Sin nada del flujo de foto (eso es del comprador, no de esta pantalla).
    assert 'id="form-leer-comanda"' not in respuesta.text


def test_ver_ingresar_mercaderia_sin_proveedor_error_de_base_da_500():
    with patch("app.main.listar_proveedores", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/deposito/ingresar")

    assert respuesta.status_code == 500


def test_ver_ingresar_mercaderia_con_proveedor_muestra_formulario_de_renglon():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=200")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "N07P41" in respuesta.text
    assert "Kiwi" in respuesta.text
    assert "Cantidad de cajones" in respuesta.text
    assert "Contenido por cajón" in respuesta.text
    # Sin campo de precio/costo: eso lo carga el comprador después.
    assert 'name="importe"' not in respuesta.text
    assert 'name="sena"' not in respuesta.text
    assert ">Importe<" not in respuesta.text
    assert ">Seña<" not in respuesta.text
    assert 'action="/deposito/ingresar"' in respuesta.text


def test_ver_ingresar_mercaderia_con_proveedor_tipo_retiro_clark_por_default():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=200")

    assert respuesta.status_code == 200
    assert '<option value="Clark" selected>Clark</option>' in respuesta.text


# --- TIPO_RETIRO_DEFAULT: tipo de logística preseleccionado por empresa ---


def test_tipo_retiro_default_sin_variable_es_clark():
    # Frutamax no setea la variable: nada cambia para ese deploy.
    with patch.dict("os.environ", {}, clear=True):
        assert _tipo_retiro_default_desde_env() == "Clark"


def test_tipo_retiro_default_toma_el_valor_de_la_variable():
    with patch.dict("os.environ", {"TIPO_RETIRO_DEFAULT": "Pases"}):
        assert _tipo_retiro_default_desde_env() == "Pases"


def test_tipo_retiro_default_invalido_cae_a_clark():
    # Un valor mal escrito no puede romper las pantallas de carga.
    with patch.dict("os.environ", {"TIPO_RETIRO_DEFAULT": "Camion"}):
        assert _tipo_retiro_default_desde_env() == "Clark"


def test_form_de_compra_manual_preselecciona_el_default_de_la_empresa():
    with (
        patch.dict("app.main.templates.env.globals", {"TIPO_RETIRO_DEFAULT": "Pases"}),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    assert '<option value="Pases" selected>Pases</option>' in respuesta.text
    assert '<option value="Clark" selected>' not in respuesta.text


def test_ingreso_directo_preselecciona_el_default_de_la_empresa():
    with (
        patch.dict("app.main.templates.env.globals", {"TIPO_RETIRO_DEFAULT": "Pases"}),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=200")

    assert respuesta.status_code == 200
    assert '<option value="Pases" selected>Pases</option>' in respuesta.text


def test_editar_compra_conserva_el_retiro_guardado_aunque_cambie_el_default():
    # El default de la empresa es solo para compras NUEVAS: una compra que
    # ya tiene su tipo guardado lo conserva, nunca se lo pisa el default.
    compra = dict(COMPRA_DE_PRUEBA, tipo_retiro="Carro")
    with (
        patch.dict("app.main.templates.env.globals", {"TIPO_RETIRO_DEFAULT": "Pases"}),
        patch("app.main.obtener_compra", return_value=compra),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert '<option value="Carro" selected>Carro</option>' in respuesta.text
    assert '<option value="Pases" selected>' not in respuesta.text


def test_renglones_de_comanda_leida_preseleccionan_el_default_de_la_empresa():
    with (
        patch.dict("app.main.templates.env.globals", {"TIPO_RETIRO_DEFAULT": "Pases"}),
        patch("app.main.extraer_comanda", return_value=COMANDA_LEIDA_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_todas_las_conversiones", return_value=[]),
        patch("app.main.listar_aprendizaje_articulos_por_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva/fotos/leer",
            files={"foto": ("comanda.jpg", b"contenido falso", "image/jpeg")},
        )

    assert respuesta.status_code == 200
    assert '<option value="Pases" selected>Pases</option>' in respuesta.json()["html"]


def test_renglon_en_blanco_de_multiples_fotos_arranca_en_el_default_de_la_empresa():
    # El JS de "agregar renglón" tiene que usar el mismo default que los
    # selects (antes tenía "Clark" hardcodeado).
    with (
        patch.dict("app.main.templates.env.globals", {"TIPO_RETIRO_DEFAULT": "Pases"}),
        patch("app.main.listar_proveedores", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva/fotos")

    assert respuesta.status_code == 200
    assert ': "Pases";' in respuesta.text
    assert ': "Clark";' not in respuesta.text


def test_ver_ingresar_mercaderia_con_proveedor_muestra_cargado_hoy():
    renglones_hoy = [
        {"id": 99, "articulo_nombre": "Kiwi", "cantidad_cajones": 10, "contenido_por_cajon": 18.6},
    ]
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=renglones_hoy),
    ):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=200")

    assert respuesta.status_code == 200
    assert "Kiwi" in respuesta.text
    assert "<td>19</td>" in respuesta.text  # contenido_por_cajon redondeado (18.6 -> 19)


def test_ver_ingresar_mercaderia_con_proveedor_inexistente_da_404():
    with patch("app.main.obtener_proveedor", return_value=None):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=999")

    assert respuesta.status_code == 404


def test_ver_ingresar_mercaderia_con_proveedor_error_de_base_da_500():
    with patch("app.main.obtener_proveedor", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/deposito/ingresar?proveedor_id=200")

    assert respuesta.status_code == 500


def test_elegir_proveedor_ingreso_directo_exitoso_redirige_con_proveedor_id():
    # Mismo mecanismo que /compras/nueva/proveedor: Depósito tiene que
    # poder cargar un proveedor nuevo por código de puesto (mercadería
    # que entra fuera de hora puede venir de un proveedor que nunca se
    # compró), no solo elegir uno ya existente.
    with patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor:
        respuesta = cliente.post(
            "/deposito/ingresar/proveedor",
            data={"codigo_puesto": "n07p41", "nombre": "Saturno"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito/ingresar?proveedor_id=200"
    mock_proveedor.assert_called_once_with("N07P41", "Saturno")


def test_elegir_proveedor_ingreso_directo_codigo_invalido_muestra_error():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main.listar_proveedores", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar/proveedor",
            data={"codigo_puesto": "puesto15", "nombre": "Saturno"},
        )

    assert respuesta.status_code == 400
    assert "formato NNNPNN" in respuesta.text
    mock_proveedor.assert_not_called()


def test_elegir_proveedor_ingreso_directo_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_proveedores", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar/proveedor",
            data={"codigo_puesto": "N07P41", "nombre": "Saturno"},
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar el proveedor" in respuesta.text


def test_ingresar_mercaderia_exitoso_agregar_redirige_con_aviso():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == (
        "/deposito/ingresar?proveedor_id=200&aviso=Ingresada+sin+precio."
        "+El+comprador+tiene+que+cargar+el+costo."
    )
    # importe/sena van None siempre -- ingreso_directo_deposito=True hace
    # el resto (estado/estado_retiro/reales/retiro_origen).
    mock_crear.assert_called_once_with(
        HOY_DE_PRUEBA, 5, 200, 10.0, 18.0, 180.0, None, None, None, "Clark",
        ingreso_directo_deposito=True,
    )


def test_ingresar_mercaderia_terminar_redirige_a_deposito_con_aviso():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == (
        "/deposito?aviso=Ingresada+sin+precio.+El+comprador+tiene+que+cargar+el+costo."
    )
    mock_crear.assert_called_once()


def test_ingresar_mercaderia_terminar_con_renglon_vacio_va_a_deposito_sin_guardar():
    with (
        patch("app.main.obtener_proveedor") as mock_proveedor,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "accion": "terminar",
                "articulo_id": "",
                "cantidad_cajones": "",
                "contenido_por_cajon": "",
                "tipo_retiro": "",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/deposito"
    mock_proveedor.assert_not_called()
    mock_crear.assert_not_called()


def test_ingresar_mercaderia_proveedor_inexistente_da_404():
    with patch("app.main.obtener_proveedor", return_value=None):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "999",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 404


def test_ingresar_mercaderia_sin_articulo_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "articulo_id": "",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "Elegí un artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_ingresar_mercaderia_sin_cantidad_cajones_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "La cantidad de cajones es obligatoria" in respuesta.text
    mock_crear.assert_not_called()


def test_ingresar_mercaderia_articulo_sin_unidad_compra_configurada_muestra_error():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_SIN_UNIDAD_COMPRA),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "articulo_id": "7",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 400
    assert "no tiene la unidad de compra configurada" in respuesta.text
    mock_crear.assert_not_called()


def test_ingresar_mercaderia_error_de_base_muestra_mensaje_claro():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.crear_compra", side_effect=Exception("no se pudo conectar")),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/ingresar",
            data={
                "proveedor_id": "200",
                "articulo_id": "5",
                "cantidad_cajones": "10",
                "contenido_por_cajon": "18",
                "tipo_retiro": "Clark",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo guardar la compra" in respuesta.text


def test_ver_gerencia_es_el_hub_del_dinero_sin_auditoria():
    # Con Auditoría afuera (sector propio), en Gerencia queda solo el
    # manejo del dinero: las dos rentabilidades y lo que venga.
    respuesta = cliente.get("/gerencia")

    assert respuesta.status_code == 200
    assert 'href="/gerencia/rentabilidad"' in respuesta.text
    assert 'href="/gerencia/rentabilidad-real"' in respuesta.text
    assert "/auditoria" not in respuesta.text
    assert 'href="/inicio"' in respuesta.text


def test_auditoria_tiene_sector_propio_y_la_url_vieja_redirige():
    respuesta = cliente.get("/gerencia/auditoria", follow_redirects=False)
    assert respuesta.status_code == 301
    assert respuesta.headers["location"] == "/auditoria"

    # Y la tarjeta propia en Inicio, junto a los demás sectores.
    inicio = cliente.get("/inicio")
    assert 'href="/auditoria"' in inicio.text


# --- La clave de Gerencia: la zona del dinero pide su PROPIA clave ---


def test_gerencia_sin_clave_configurada_no_tiene_puerta():
    # Sin CLAVE_GERENCIA cargada en Railway no hay puerta: el deploy no
    # traba nada hasta que la variable exista.
    with patch("app.main._clave_gerencia", return_value=None):
        assert cliente.get("/gerencia").status_code == 200


def test_gerencia_con_clave_configurada_pide_clave_en_toda_la_zona():
    with patch("app.main._clave_gerencia", return_value="secreta"):
        for url in (
            "/gerencia",
            "/gerencia/rentabilidad",
            "/gerencia/rentabilidad-real",
            "/gerencia/costos-fijos",
            "/gerencia/costos-fijos/plan",
            "/gerencia/costos-fijos/indices",
        ):
            respuesta = cliente.get(url)
            assert respuesta.status_code == 401, url
            assert "Gerencia pide clave" in respuesta.text
            assert 'action="/gerencia/clave"' in respuesta.text
        # Los exports sin acceso redirigen a su pantalla (que pide la clave).
        for url in (
            "/gerencia/rentabilidad/exportar-pdf",
            "/gerencia/rentabilidad/exportar-excel",
            "/gerencia/rentabilidad-real/exportar-pdf",
            "/gerencia/rentabilidad-real/exportar-excel",
        ):
            respuesta = cliente.get(url, follow_redirects=False)
            assert respuesta.status_code == 303, url


def test_clave_gerencia_correcta_deja_cookie_y_bloquear_la_corta():
    try:
        with patch("app.main._clave_gerencia", return_value="secreta"):
            respuesta = cliente.post(
                "/gerencia/clave",
                data={"clave": "secreta", "volver": "/gerencia/rentabilidad-real"},
                follow_redirects=False,
            )
            assert respuesta.status_code == 303
            assert respuesta.headers["location"] == "/gerencia/rentabilidad-real"
            assert "acceso_gerencia" in respuesta.headers.get("set-cookie", "")

            # Con la cookie, la zona abre y el botón Bloquear está a mano.
            adentro = cliente.get("/gerencia")
            assert adentro.status_code == 200
            assert 'action="/gerencia/bloquear"' in adentro.text

            # Bloquear borra la cookie: la zona vuelve a pedir la clave.
            bloqueo = cliente.post("/gerencia/bloquear", follow_redirects=False)
            assert bloqueo.status_code == 303
            assert cliente.get("/gerencia").status_code == 401
    finally:
        cliente.cookies.clear()


def test_clave_gerencia_incorrecta_no_entra_y_el_destino_es_solo_gerencia():
    with patch("app.main._clave_gerencia", return_value="secreta"):
        respuesta = cliente.post("/gerencia/clave", data={"clave": "nope", "volver": "/gerencia"})
        assert respuesta.status_code == 401
        assert "Clave incorrecta" in respuesta.text

        # Un "volver" que apunta fuera de Gerencia se pisa: nada de usar la
        # puerta como redirector abierto.
        respuesta = cliente.post(
            "/gerencia/clave", data={"clave": "secreta", "volver": "/compras"}, follow_redirects=False
        )
        assert respuesta.headers["location"] == "/gerencia"
    cliente.cookies.clear()


def test_la_firma_del_puesto_no_abre_gerencia():
    # Misma clave en las dos zonas ≠ mismas cookies: los mensajes firmados
    # son distintos, así que una cookie del control del Puesto jamás valida
    # en Gerencia (ni al revés).
    from app.main import _firma_acceso_control, _firma_acceso_gerencia

    assert _firma_acceso_control("secreta") != _firma_acceso_gerencia("secreta")


def test_las_urls_viejas_de_facturacion_redirigen_no_dan_404():
    """El link viejo sigue en el historial del celular.

    Mismo criterio que /gerencia/auditoria -> /auditoria cuando Auditoría
    salió de Gerencia: 301, no 404 un martes a las 6 de la mañana.
    """
    respuesta = cliente.get("/facturacion", follow_redirects=False)
    assert respuesta.status_code == 301
    assert respuesta.headers["location"] == "/administracion"


def test_la_url_vieja_de_ingresos_conserva_los_filtros_al_redirigir():
    # Un link guardado con fechas y proveedor no puede perder los filtros
    # en el camino: llegaría a la pantalla mostrando otra cosa.
    respuesta = cliente.get(
        "/facturacion/ingresos?fecha_desde=2026-08-01&fecha_hasta=2026-08-28&proveedor_id=5",
        follow_redirects=False,
    )
    assert respuesta.status_code == 301
    destino = respuesta.headers["location"]
    assert destino.startswith("/administracion/ingresos?")
    assert "fecha_desde=2026-08-01" in destino
    assert "proveedor_id=5" in destino


def test_ver_administracion_es_un_hub_con_ingresos_a_deposito():
    respuesta = cliente.get("/administracion")

    assert respuesta.status_code == 200
    assert "Administración" in respuesta.text
    assert 'href="/administracion/ingresos"' in respuesta.text
    assert "Ingresos a Depósito" in respuesta.text
    assert "En construcción" not in respuesta.text
    assert 'href="/inicio"' in respuesta.text


INGRESOS_DEPOSITO_DE_PRUEBA = [
    # Saturno: una recepción normal, un rechazo parcial (8 aceptados, 2
    # devueltos) y una recepcionada SIN precio cargado todavía.
    {"id": 1, "fecha_operacion": date(2026, 8, 18), "procesada_el": datetime(2026, 8, 18, 14, 30),
     "guia_id": 105, "guia_punto": 1, "estado": "recepcionado",
     "cantidad_cajones_real": 8.0, "contenido_por_cajon_real": 20.0,
     "cantidad_cajones_rechazada": None, "motivo_rechazo": None, "importe": 5000.0, "sena": 500.0,
     "articulo_nombre": "Kiwi", "unidad_compra": "kilo",
     "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41"},
    {"id": 2, "fecha_operacion": date(2026, 8, 18), "procesada_el": datetime(2026, 8, 18, 15, 0),
     "guia_id": 105, "guia_punto": 2, "estado": "recepcionado",
     "cantidad_cajones_real": 8.0, "contenido_por_cajon_real": 10.0,
     "cantidad_cajones_rechazada": 2.0, "motivo_rechazo": "Podrido", "importe": 3000.0, "sena": None,
     "articulo_nombre": "Mango", "unidad_compra": "unidad",
     "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41"},
    {"id": 3, "fecha_operacion": date(2026, 8, 18), "procesada_el": datetime(2026, 8, 18, 15, 30),
     "guia_id": 105, "guia_punto": 3, "estado": "recepcionado",
     "cantidad_cajones_real": 5.0, "contenido_por_cajon_real": 6.0,
     "cantidad_cajones_rechazada": None, "motivo_rechazo": None, "importe": None, "sena": None,
     "articulo_nombre": "Palta", "unidad_compra": "kilo",
     "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41"},
    {"id": 4, "fecha_operacion": date(2026, 8, 18), "procesada_el": datetime(2026, 8, 18, 16, 0),
     "guia_id": 106, "guia_punto": 1, "estado": "recepcionado",
     "cantidad_cajones_real": 10.0, "contenido_por_cajon_real": 16.0,
     "cantidad_cajones_rechazada": None, "motivo_rechazo": None, "importe": 2000.0, "sena": None,
     "articulo_nombre": "Limón", "unidad_compra": "kilo",
     "proveedor_nombre": "Verdurin", "proveedor_codigo_puesto": "N03P12"},
]


def test_ver_ingresos_deposito_agrupa_por_proveedor_con_subtotales_y_total():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_ingresos_deposito", return_value=INGRESOS_DEPOSITO_DE_PRUEBA) as mock_buscar,
    ):
        respuesta = cliente.get("/administracion/ingresos")

    assert respuesta.status_code == 200
    # Default: últimas 48hs y SOLO lo que hay que pagar (recepcionadas,
    # rechazo parcial incluido — es una recepción por los bultos aceptados).
    mock_buscar.assert_called_once_with(
        HOY_DE_PRUEBA - timedelta(days=1), HOY_DE_PRUEBA, None, None, "recepcionado",
        limite=TOPE_FILAS_BUSQUEDA + 1,
    )
    # Cantidades REALES (las de Depósito): 8 × 20k, no lo que cargó el comprador.
    assert "8 × 20k" in respuesta.text
    # El rechazo parcial se ve aparte, para explicar la diferencia: se
    # facturan los 8 aceptados y se muestra que 2 se devolvieron.
    assert "Rechazo parcial (2 rech.)" in respuesta.text
    # El subtotal es lo que hay que DEPOSITARLE al proveedor: mercadería
    # (8×5000 + 8×3000 = 64000) + las señas de sus cajones (8×500 = 4000).
    assert "Subtotal Saturno" in respuesta.text
    assert "$68.000" in respuesta.text
    assert "mercadería $64.000 + señas $4.000" in respuesta.text
    # Verdurin no tiene señas: su subtotal es la mercadería sola.
    assert "Subtotal Verdurin" in respuesta.text
    assert "$20.000" in respuesta.text
    # Total al final: 84000 de mercadería + 4000 de señas (la compra sin
    # precio no suma).
    assert "Total a depositar: $88.000" in respuesta.text
    assert "mercadería $84.000 + señas de los cajones $4.000" in respuesta.text
    # La compra sin precio queda marcada: hay que completarla antes de facturar.
    assert "Sin precio" in respuesta.text
    assert "1 compra sin precio cargado" in respuesta.text
    # La seña: por cajón y, al lado, lo señado por los cajones reales
    # (500 × 8 = 4000). Las filas sin seña van con "—".
    assert "<th>Seña</th>" in respuesta.text
    assert "$500 ($4.000)" in respuesta.text
    # Export con los mismos filtros.
    assert "/administracion/ingresos/exportar-pdf?" in respuesta.text
    assert "/administracion/ingresos/exportar-excel?" in respuesta.text


def test_ver_ingresos_deposito_estado_todas_pasa_none_a_la_consulta():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_ingresos_deposito", return_value=[]) as mock_buscar,
    ):
        respuesta = cliente.get(
            "/administracion/ingresos?fecha_desde=2026-08-10&fecha_hasta=2026-08-12&estado=todas&proveedor_id=7"
        )

    assert respuesta.status_code == 200
    mock_buscar.assert_called_once_with(
        date(2026, 8, 10), date(2026, 8, 12), 7, None, None, limite=TOPE_FILAS_BUSQUEDA + 1
    )
    assert "No se encontraron ingresos" in respuesta.text


def test_ver_ingresos_deposito_cortada_por_el_tope_avisa_y_oculta_totales():
    muchos = [dict(INGRESOS_DEPOSITO_DE_PRUEBA[0], id=i) for i in range(TOPE_FILAS_BUSQUEDA + 1)]
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_ingresos_deposito", return_value=muchos),
        patch("app.main.contar_ingresos_deposito", return_value=1200) as mock_contar,
    ):
        respuesta = cliente.get("/administracion/ingresos")

    assert respuesta.status_code == 200
    assert f"Se muestran los primeros {TOPE_FILAS_BUSQUEDA} ingresos de 1200" in respuesta.text
    assert "Total general:" not in respuesta.text
    mock_contar.assert_called_once()


def test_exportar_ingresos_deposito_pdf_sin_tope_con_subtotales_y_marcas():
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_ingresos_deposito", return_value=INGRESOS_DEPOSITO_DE_PRUEBA) as mock_buscar,
    ):
        respuesta = cliente.get("/administracion/ingresos/exportar-pdf?fecha_desde=2026-08-17&fecha_hasta=2026-08-18")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "Ingresos_Deposito_2026-08-17_a_2026-08-18" in respuesta.headers["content-disposition"]
    # SIN tope: el export no pasa limite.
    mock_buscar.assert_called_once_with(date(2026, 8, 17), date(2026, 8, 18), None, None, "recepcionado")
    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Ingresos a Depósito" in texto
    assert "Saturno (N07P41)" in texto
    # El subtotal del PDF es lo a depositar, con las dos partes al lado.
    assert "Subtotal" in texto
    assert "$68.000" in texto
    assert "mercadería $64.000 + señas $4.000" in " ".join(texto.split())
    assert "Total a depositar: $88.000" in texto
    assert "SIN PRECIO" in texto
    # El texto puede quebrarse en dos renglones dentro de la columna: se
    # compara con los espacios normalizados.
    assert "Rechazo parcial (2 rech.)" in " ".join(texto.split())
    assert "105.2" in texto  # número de guía y punto
    # La seña del Kiwi: por cajón y el total señado.
    assert "$500 ($4.000)" in " ".join(texto.split())


def test_exportar_ingresos_deposito_excel_devuelve_archivo_adjunto():
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_ingresos_deposito", return_value=INGRESOS_DEPOSITO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/administracion/ingresos/exportar-excel?fecha_desde=2026-08-17&fecha_hasta=2026-08-18")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Ingresos_Deposito_2026-08-17_a_2026-08-18" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip

    # La seña viaja en sus dos columnas propias: por bulto y el total señado.
    from io import BytesIO

    from openpyxl import load_workbook

    hoja = load_workbook(BytesIO(respuesta.content)).active
    valores = [celda.value for fila in hoja.iter_rows() for celda in fila]
    assert "Seña por bulto" in valores
    assert "Total seña" in valores
    assert 500.0 in valores
    assert 4000.0 in valores


def test_exportar_ingresos_deposito_pdf_fecha_invalida_da_400():
    respuesta = cliente.get("/administracion/ingresos/exportar-pdf?fecha_desde=no-es-fecha&fecha_hasta=2026-08-18")

    assert respuesta.status_code == 400


def test_barra_navegacion_en_compras_apunta_a_compras_y_a_inicio():
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert '<header class="barra-navegacion">' in respuesta.text
    assert f'href="/inicio" aria-label="Ir a Inicio">{_ICONO_INICIO}</a>' in respuesta.text
    assert f'href="/compras" aria-label="Ir a Compras">{SECTORES["compras"]["icono"]}</a>' in respuesta.text
    assert '<div class="barra-titulo">Compras</div>' in respuesta.text


def test_barra_navegacion_tiene_boton_de_volver_atras():
    # Tercer botón junto a la casita y al módulo: SUBE en la jerarquía del
    # sistema (padre declarado por pantalla) — un hub sube a Inicio. El
    # history.back() del navegador queda para el gesto nativo.
    respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert '<a class="barra-boton" href="/inicio" aria-label="Volver atrás">' in respuesta.text
    assert 'onclick="history.back()"' not in respuesta.text


def test_el_titulo_de_la_barra_se_autoajusta_y_no_se_recorta():
    # Cada pantalla tiene un nombre de largo distinto: el script busca el
    # tamaño más grande al que entra en una línea ("Compras" grande, los
    # largos más chicos o a dos líneas), y nada se corta con "…". El
    # 1.2rem del CSS es solo el arranque/fallback sin JS.
    respuesta = cliente.get("/compras")

    assert "text-overflow" not in respuesta.text
    assert "font-size: 1.2rem" in respuesta.text
    assert "function ajustarTitulo()" in respuesta.text


def test_barra_navegacion_en_comercial_usa_icono_distinto_de_compras():
    # Regresión: Compras y Comercial arrancan las dos con "C" — se
    # distinguen con íconos distintos, no con la inicial.
    with patch("app.main.listar_clientes", return_value=[]):
        respuesta = cliente.get("/clientes")

    assert respuesta.status_code == 200
    assert f'href="/comercial" aria-label="Ir a Comercial">{SECTORES["comercial"]["icono"]}</a>' in respuesta.text
    assert SECTORES["compras"]["icono"] not in respuesta.text


def test_barra_navegacion_en_sistema():
    respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert f'href="/sistema" aria-label="Ir a Sistema">{SECTORES["sistema"]["icono"]}</a>' in respuesta.text


ARGENTINA_TEST = timezone(timedelta(hours=-3))


def _foto_alertas(valores=None, calculada_el=None):
    """La foto de alertas_estado: todas las del registro en cero, salvo las que se pasen.

    Auditoría ya NO calcula: lee esta foto. Que los conteos se hagan con las
    ventanas correctas es lo que prueba
    test_recalcular_alertas_usa_las_ventanas_de_cada_control.
    """
    from app.main import ALERTAS

    valores = valores or {}
    momento = calculada_el or datetime.now(ARGENTINA_TEST)
    filas = []
    for definicion in ALERTAS:
        casos, mas_viejo = valores.get(definicion.codigo, (0, None))
        filas.append({
            "codigo": definicion.codigo, "casos": casos, "mas_viejo": mas_viejo,
            "calculada_el": momento, "duracion_ms": 5, "error": None,
        })
    return filas


def test_ver_auditoria_lista_las_alertas_con_casos_y_el_mas_viejo():
    # El tablero es una LISTA de alertas con la misma forma (título + casos +
    # más viejo + link): agregar la número diecisiete es sumar una definición
    # al registro, no tocar el diseño. Y ahora LEE la foto: una sola consulta,
    # no quince conteos en vivo.
    foto = _foto_alertas({
        "compras_sin_precio": (2, date(2026, 7, 30)),
        "retiros_sin_hacer": (7, date(2026, 8, 1)),
        "recepciones_pendientes": (3, date(2026, 8, 2)),
        "stock_vacios_negativo": (1, None),
        "stock_deposito_negativo": (2, None),
        "guias_r_costo_incompleto": (1, date(2026, 8, 5)),
        "articulos_incotizables": (4, None),
        "senas_vacios_pendientes": (5, date(2026, 7, 28)),
        "pedidos_sin_identificar": (2, date(2026, 8, 3)),
        "pedidos_incompletos": (1, date(2026, 8, 5)),
        "mails_sin_confirmar": (2, date(2026, 8, 6)),
        "mails_leidos_con_ia": (1, date(2026, 8, 7)),
        "pedido_faltante": (1, date(2026, 8, 5)),
        "casilla_sin_revisar": (1, None),
    })
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_estado_alertas", return_value=foto) as mock_foto,
    ):
        respuesta = cliente.get("/auditoria")

    assert respuesta.status_code == 200
    # UNA sola consulta para toda la pantalla: ese es el punto del cambio.
    mock_foto.assert_called_once_with()
    # Las alertas, con sin-precio primera (la más importante). Su título ya no
    # habla de horas: esa alerta no tiene ventana, avisa el mismo día.
    assert respuesta.text.index("Compras sin precio de compra cargado") < respuesta.text.index("Mercadería sin retirar hace")
    assert "Mercadería sin recepcionar hace más de 48 horas" in respuesta.text
    assert "Stock de vacíos negativo" in respuesta.text
    assert "Stock de depósito en negativo (salidas sin explicar)" in respuesta.text
    # Al Remanente, que es donde quedaron los negativos: Stock del Sistema se
    # borró el 06/09 y la alerta habría llevado a un 404.
    assert 'href="/administracion/stock/remanente"' in respuesta.text
    assert "Guías R esperando el precio de una compra" in respuesta.text
    assert 'href="/administracion/stock/guias-r"' in respuesta.text
    assert "sin ficha logística o sin precio de venta" in respuesta.text
    assert "Señas de vacíos pendientes hace más de 7 días" in respuesta.text
    assert "Pedidos con renglones sin identificar" in respuesta.text
    assert "Pedidos con renglones incompletos" in respuesta.text
    assert "Mails de pedido sin confirmar" in respuesta.text
    assert "leídos con IA (el parser de estructura no pudo" in respuesta.text
    assert "Falta el pedido de un día esperado" in respuesta.text
    assert "La casilla de pedidos no se pudo revisar" in respuesta.text
    assert 'href="/sistema/casilla-pedidos"' in respuesta.text
    assert "el más viejo es del 01/08/2026" in respuesta.text
    assert "el más viejo es del 28/07/2026" in respuesta.text
    # Los links al detalle. El de retiros se arma con el dato de la foto.
    assert "/logistica/consultar?fecha_desde=2026-08-01&amp;fecha_hasta=2026-08-04&amp;estado=pendiente" in respuesta.text
    assert 'href="/deposito/recepcion"' in respuesta.text
    assert 'href="/compras/pendientes"' in respuesta.text
    assert 'href="/puesto/envases/stock"' in respuesta.text
    assert 'href="/fichas"' in respuesta.text
    assert 'href="/puesto/envases/pendientes"' in respuesta.text


def test_recalcular_alertas_usa_las_ventanas_de_cada_control():
    """Las ventanas de cada alerta ("más de 48 horas", "7 días") ahora se prueban acá.

    Antes esto lo verificaba la vista, porque calculaba en vivo. Ahora calcula
    el recálculo, y la garantía tiene que seguir estando.
    """
    from app.main import ALERTAS, recalcular

    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.alertas.candado_alertas") as candado,
        patch("app.alertas.guardar_estado_alerta"),
        patch("app.main.contar_compras_sin_precio", return_value={"casos": 0, "mas_viejo": None}) as sin_precio,
        patch("app.main.contar_retiros_pendientes_viejos", return_value={"casos": 0, "mas_viejo": None}) as retiros,
        patch("app.main.contar_recepciones_pendientes_viejas", return_value={"casos": 0, "mas_viejo": None}) as recepciones,
        patch("app.main.contar_articulos_comprados_incotizables", return_value=0) as incotizables,
        patch("app.main.contar_senas_pendientes_viejas", return_value={"casos": 0, "mas_viejo": None}) as senas,
        patch("app.main.contar_mails_pedido_leidos_con_ia", return_value={"casos": 0, "mas_viejo": None}) as leidos_ia,
        patch("app.main.contar_stock_vacios_negativos", return_value=0),
        patch("app.main.contar_stock_deposito_negativo", return_value=0),
        patch("app.main.contar_reprocesos_costo_incompleto", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.contar_pedidos_con_renglones_sin_identificar", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.contar_pedidos_incompletos", return_value={"casos": 0, "mas_viejo": None}) as incompletos,
        patch("app.main.contar_mails_pedido_sin_procesar", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.contar_pedidos_faltantes", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.contar_casillas_sin_revisar", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        candado.return_value.__enter__.return_value = True
        resumen = recalcular(ALERTAS)

    assert resumen["corrio"] is True and resumen["fallaron"] == 0
    # "Más de 48 horas" = de anteayer para atrás; señas y comprados, 7 días.
    # Compras sin precio va SIN ventana: se avisa el mismo día y no deja de
    # avisar por vieja. Pedidos incompletos SÍ lleva ventana (7 días): un
    # pedido que ya salió incompleto no se puede completar después.
    sin_precio.assert_called_once_with()
    incompletos.assert_called_once_with(date(2026, 7, 30))
    retiros.assert_called_once_with(date(2026, 8, 4))
    recepciones.assert_called_once_with(date(2026, 8, 4))
    incotizables.assert_called_once_with(date(2026, 7, 30), HOY_DE_PRUEBA)
    senas.assert_called_once_with(date(2026, 7, 30))
    leidos_ia.assert_called_once_with(date(2026, 7, 30))


def test_ver_auditoria_sin_casos_muestra_todo_en_orden():
    # Un control sin casos NO aparece; si ninguno tiene, cartel verde.
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/auditoria")

    assert respuesta.status_code == 200
    assert "Todo en orden" in respuesta.text
    assert "Mercadería sin retirar" not in respuesta.text
    assert "Mercadería sin recepcionar" not in respuesta.text


def test_auditoria_muestra_siempre_de_cuando_es_la_foto():
    """De cuándo son las alertas se muestra SIEMPRE, aunque no haya ninguna con casos."""
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/auditoria")
    assert "Calculadas" in respuesta.text
    assert "Recalcular ahora" in respuesta.text


def test_auditoria_avisa_fuerte_cuando_el_calculo_automatico_se_murio():
    """La trampa del 25/08: "no hay problemas" y "no se está calculando" no pueden verse iguales."""
    vieja = datetime.now(ARGENTINA_TEST) - timedelta(hours=40)
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas(calculada_el=vieja)):
        respuesta = cliente.get("/auditoria")
    assert "el cálculo automático no está corriendo" in respuesta.text


def test_auditoria_sin_ninguna_foto_lo_dice():
    with patch("app.main.listar_estado_alertas", return_value=[]):
        respuesta = cliente.get("/auditoria")
    assert "todavía no se calcularon" in respuesta.text


def test_boton_de_recalcular_corre_las_alertas_y_vuelve_con_el_resultado():
    with patch("app.main.recalcular", return_value={"corrio": True, "ok": 16, "fallaron": 0}) as mock_recalcular:
        respuesta = cliente.post("/auditoria/recalcular", follow_redirects=False)
    mock_recalcular.assert_called_once()
    assert respuesta.status_code == 303
    assert "16+controles+recalculados" in respuesta.headers["location"]


def test_si_ya_se_estan_recalculando_el_boton_avisa_en_vez_de_duplicar():
    with patch("app.main.recalcular", return_value={"corrio": False, "ok": 0, "fallaron": 0}):
        respuesta = cliente.post("/auditoria/recalcular", follow_redirects=False)
    assert respuesta.status_code == 303
    assert "Ya+se+estaban+recalculando" in respuesta.headers["location"]


def test_barra_navegacion_en_placeholders_logistica_gerencia():
    casos = [("/logistica", "logistica", "Logística"), ("/gerencia", "gerencia", "Gerencia")]
    for url, sector, nombre in casos:
        respuesta = cliente.get(url)
        assert respuesta.status_code == 200
        assert f'href="{url}" aria-label="Ir a {nombre}">{SECTORES[sector]["icono"]}</a>' in respuesta.text


def test_barra_navegacion_en_deposito_apunta_a_deposito_y_a_inicio():
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert f'href="/inicio" aria-label="Ir a Inicio">{_ICONO_INICIO}</a>' in respuesta.text
    assert f'href="/deposito" aria-label="Ir a Depósito">{SECTORES["deposito"]["icono"]}</a>' in respuesta.text
    assert '<div class="barra-titulo">Depósito</div>' in respuesta.text


def test_ver_inicio_no_tiene_barra_de_navegacion():
    # /inicio ya es el punto de partida de todo: no tiene adónde volver.
    respuesta = cliente.get("/inicio")

    assert respuesta.status_code == 200
    assert "barra-navegacion" not in respuesta.text


def test_emojis_de_colores_reemplazados_por_iconos_svg():
    # Regresión: los emojis de colores (🏠🛒💼📦🚚📊⚙️) se reemplazaron por
    # íconos SVG minimalistas de línea (heroicons), consistentes entre sí.
    emojis_viejos = ["🏠", "🛒", "💼", "📦", "🚚", "📊", "⚙️"]
    with patch("app.main.listar_clientes", return_value=[]):
        respuesta_clientes = cliente.get("/clientes")
    for url, respuesta in [("/inicio", cliente.get("/inicio")), ("/compras", cliente.get("/compras")), ("/clientes", respuesta_clientes)]:
        assert respuesta.status_code == 200
        for emoji in emojis_viejos:
            assert emoji not in respuesta.text, f"{emoji} todavía aparece en {url}"


def test_titulo_grande_del_cuerpo_ya_no_aparece_en_ninguna_pantalla():
    # Regresión: el título grande en el CUERPO de la pantalla (duplicado con
    # el de la barrita) se sacó por completo — ni la regla CSS .titulo-sector
    # ni el <h1> quedan en ningún lado.
    urls = [
        "/compras",
        "/comercial",
        "/sistema",
        "/logistica",
        "/deposito",
        "/gerencia",
        "/compras/nueva/listado",
    ]
    for url in urls:
        respuesta = cliente.get(url)
        assert respuesta.status_code == 200
        assert "titulo-sector" not in respuesta.text, f"'titulo-sector' todavía aparece en {url}"

    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/nueva/manual")
    assert respuesta.status_code == 200
    assert "titulo-sector" not in respuesta.text

    with (
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")
    assert respuesta.status_code == 200
    assert "titulo-sector" not in respuesta.text

    with (
        # Buscar Compras es un FILTRO: usa la lista completa, de baja incluidos.
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar")
    assert respuesta.status_code == 200
    assert "titulo-sector" not in respuesta.text


def test_barra_navegacion_muestra_el_titulo_del_sector_en_las_pantallas_principales():
    # El único título que queda es el de la barrita — se agrandó, pero
    # sigue siendo el mismo <div class="barra-titulo">.
    casos = [("/compras", "Compras"), ("/comercial", "Comercial"), ("/sistema", "Sistema")]
    for url, nombre in casos:
        respuesta = cliente.get(url)
        assert respuesta.status_code == 200
        assert f'<div class="barra-titulo">{nombre}</div>' in respuesta.text


# --- /compras/disponibles ---

FICHAS_DISPONIBLES_DE_PRUEBA = [
    {
        "id": 1, "articulo_id": 10, "articulo_nombre": "Tomate Cherry", "articulo_grupo": "hortaliza",
        "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
        "envase_variable": False, "nombre_cliente": "TOM CHERRY", "codigo_cliente": "90039",
    },
    {
        "id": 2, "articulo_id": 11, "articulo_nombre": "Manzana Roja", "articulo_grupo": "fruta",
        "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
        "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "20044",
    },
    {
        "id": 3, "articulo_id": 12, "articulo_nombre": "Cubeta Pesada", "articulo_grupo": "pesada",
        "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "cubeta",
        "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "10021",
    },
    {
        # Sin código ni nombre de cliente: nunca se le armó la ficha de
        # logística completa para este cliente todavía — no se precarga
        # sola, hay que agregarla a mano desde "Agregar desde el catálogo".
        "id": 4, "articulo_id": 13, "articulo_nombre": "Sin Alias Todavía", "articulo_grupo": "fruta",
        "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
        "envase_variable": False, "nombre_cliente": None, "codigo_cliente": None,
    },
]

BORRADOR_DISPONIBLE_DE_PRUEBA = {
    "id": 30, "cliente_id": 1, "fecha_desde": date(2026, 8, 14), "fecha_hasta": date(2026, 8, 14),
    "estado": "borrador", "version": None,
}

DETALLE_DISPONIBLE_DE_PRUEBA = [
    {"id": 1, "articulo_id": 10, "codigo": "90039", "nombre": "TOM CHERRY", "cantidad": 38.0, "orden": 1},
    {"id": 2, "articulo_id": None, "codigo": None, "nombre": "Frutilla", "cantidad": 12.0, "orden": 2},
]


def test_ver_disponibles_sin_cliente_muestra_selector():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/compras/disponibles")

    assert respuesta.status_code == 200
    assert "Elegí un cliente para armarle su Disponible." in respuesta.text
    assert "Día" in respuesta.text


def test_ver_disponibles_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/compras/disponibles?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_disponibles_error_de_base_da_500():
    with patch("app.main.listar_clientes", side_effect=Exception("no se pudo conectar")):
        respuesta = cliente.get("/compras/disponibles")

    assert respuesta.status_code == 500


def test_ver_disponibles_precarga_desde_borrador_existente():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_DISPONIBLES_DE_PRUEBA),
        patch("app.main.obtener_borrador_disponible", return_value=BORRADOR_DISPONIBLE_DE_PRUEBA) as mock_borrador,
        patch("app.main.listar_detalle_disponible", return_value=DETALLE_DISPONIBLE_DE_PRUEBA) as mock_detalle,
        patch("app.main.obtener_ultimo_disponible_cliente") as mock_ultimo,
    ):
        respuesta = cliente.get("/compras/disponibles?cliente_id=1")

    assert respuesta.status_code == 200
    mock_borrador.assert_called_once_with(1)
    mock_detalle.assert_called_once_with(30)
    mock_ultimo.assert_not_called()
    # Los renglones del borrador (con lo ya cargado) quedan embebidos para editar.
    assert '"nombre": "TOM CHERRY"' in respuesta.text
    assert '"cantidad": 38.0' in respuesta.text
    assert '"nombre": "Frutilla"' in respuesta.text
    assert "2026-08-14" in respuesta.text


def test_ver_disponibles_precarga_desde_ultimo_disponible_si_no_hay_borrador():
    ultimo = {"id": 29, "cliente_id": 1, "fecha_desde": date(2026, 8, 10), "fecha_hasta": date(2026, 8, 10)}
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_DISPONIBLES_DE_PRUEBA),
        patch("app.main.obtener_borrador_disponible", return_value=None),
        patch("app.main.obtener_ultimo_disponible_cliente", return_value=ultimo) as mock_ultimo,
        patch("app.main.listar_detalle_disponible", return_value=DETALLE_DISPONIBLE_DE_PRUEBA) as mock_detalle,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 16)),
    ):
        respuesta = cliente.get("/compras/disponibles?cliente_id=1")

    assert respuesta.status_code == 200
    mock_ultimo.assert_called_once_with(1)
    mock_detalle.assert_called_once_with(29)
    # Nuevo Disponible (sin id todavía) pero con los renglones del último, y
    # la fecha por default es HOY, no la del último Disponible (10/08).
    assert "null" in respuesta.text  # DISPONIBLE_ID = null
    assert 'value="2026-08-16"' in respuesta.text
    assert '"nombre": "TOM CHERRY"' in respuesta.text


def test_ver_disponibles_precarga_desde_fichas_si_nunca_tuvo_uno():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_DISPONIBLES_DE_PRUEBA),
        patch("app.main.obtener_borrador_disponible", return_value=None),
        patch("app.main.obtener_ultimo_disponible_cliente", return_value=None),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 16)),
    ):
        respuesta = cliente.get("/compras/disponibles?cliente_id=1")

    assert respuesta.status_code == 200
    # Orden fruta -> hortaliza -> pesada dentro de "renglones" (la precarga
    # editable) — no dentro de CATALOGO, que lista las fichas tal cual
    # vienen de la base, sin reordenar. Se busca a partir de "let renglones"
    # para no confundirse con el orden de CATALOGO, que aparece antes.
    texto_renglones = respuesta.text[respuesta.text.index("let renglones"):]
    indice_manzana = texto_renglones.index("Manzana Roja")
    indice_tomate = texto_renglones.index("TOM CHERRY")
    indice_cubeta = texto_renglones.index("Cubeta Pesada")
    assert indice_manzana < indice_tomate < indice_cubeta
    # Sin nombre_cliente cargado (Manzana Roja), cae al nombre interno del artículo.
    assert '"nombre": "Manzana Roja"' in respuesta.text
    # Cantidad en 0: primera vez, no hay nada previo que sugerir.
    assert '"cantidad": 0.0}' in respuesta.text
    # Sin código NI nombre de cliente todavía: no se precarga sola.
    assert '"nombre": "Sin Alias Todavía"' not in texto_renglones


def test_guardar_disponible_ruta_guarda_y_redirige():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.guardar_disponible", return_value=30) as mock_guardar,
    ):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_articulo_id_0": "10",
                "renglon_codigo_0": "90039",
                "renglon_nombre_0": "TOM CHERRY",
                "renglon_cantidad_0": "38",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/disponibles?cliente_id=1&guardado=1"
    mock_guardar.assert_called_once_with(
        None, 1, date(2026, 8, 14), date(2026, 8, 14),
        [{"articulo_id": 10, "codigo": "90039", "nombre": "TOM CHERRY", "cantidad": 38.0}],
    )


def test_guardar_disponible_ruta_con_disponible_id_actualiza_el_mismo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.guardar_disponible", return_value=30) as mock_guardar,
    ):
        cliente.post(
            "/compras/disponibles/guardar",
            data={
                "cliente_id": "1",
                "disponible_id": "30",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_articulo_id_0": "",
                "renglon_codigo_0": "",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "12",
            },
        )

    mock_guardar.assert_called_once_with(
        30, 1, date(2026, 8, 14), date(2026, 8, 14),
        [{"articulo_id": None, "codigo": None, "nombre": "Frutilla", "cantidad": 12.0}],
    )


def test_guardar_disponible_ruta_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={"cliente_id": "999", "disponible_id": "", "fecha_desde": "2026-08-14", "fecha_hasta": "2026-08-14"},
        )

    assert respuesta.status_code == 404


def test_guardar_disponible_ruta_fecha_hasta_anterior_a_desde_da_400():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-10",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "12",
            },
        )

    assert respuesta.status_code == 400
    assert "no puede ser anterior" in respuesta.text


def test_guardar_disponible_ruta_sin_renglones_da_400():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={"cliente_id": "1", "disponible_id": "", "fecha_desde": "2026-08-14", "fecha_hasta": "2026-08-14"},
        )

    assert respuesta.status_code == 400
    assert "Agregá al menos un artículo" in respuesta.text


def test_guardar_disponible_ruta_cantidad_no_numerica_da_400():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "no-es-un-numero",
            },
        )

    assert respuesta.status_code == 400
    assert "tiene que ser un número" in respuesta.text


def test_guardar_disponible_ruta_cantidad_negativa_da_400():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.post(
            "/compras/disponibles/guardar",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "-5",
            },
        )

    assert respuesta.status_code == 400
    assert "no puede ser negativa" in respuesta.text


def test_guardar_y_exportar_disponible_excel_cierra_y_devuelve_excel():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.guardar_disponible", return_value=30),
        patch("app.main.cerrar_disponible_generado", return_value=1) as mock_cerrar,
    ):
        respuesta = cliente.post(
            "/compras/disponibles/guardar-y-exportar-excel",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_articulo_id_0": "10",
                "renglon_codigo_0": "90039",
                "renglon_nombre_0": "TOM CHERRY",
                "renglon_cantidad_0": "38",
            },
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="Disponibles_Frutamax_14_Ago_2026.xlsx"' in respuesta.headers["content-disposition"]
    mock_cerrar.assert_called_once_with(30, 1, date(2026, 8, 14))


def test_guardar_y_exportar_disponible_excel_con_reenvio_agrega_sufijo_de_version():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.guardar_disponible", return_value=31),
        patch("app.main.cerrar_disponible_generado", return_value=2),
    ):
        respuesta = cliente.post(
            "/compras/disponibles/guardar-y-exportar-excel",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "12",
            },
        )

    assert respuesta.status_code == 200
    assert 'filename="Disponibles_Frutamax_14_Ago_2026_v2.xlsx"' in respuesta.headers["content-disposition"]


def test_guardar_y_exportar_disponible_excel_error_de_base_da_500():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.guardar_disponible", return_value=30),
        patch("app.main.cerrar_disponible_generado", side_effect=Exception("no se pudo conectar")),
    ):
        respuesta = cliente.post(
            "/compras/disponibles/guardar-y-exportar-excel",
            data={
                "cliente_id": "1",
                "disponible_id": "",
                "fecha_desde": "2026-08-14",
                "fecha_hasta": "2026-08-14",
                "renglon_nombre_0": "Frutilla",
                "renglon_cantidad_0": "12",
            },
        )

    assert respuesta.status_code == 500
    assert "No se pudo generar el Excel" in respuesta.text


# --- /compras/objetivo: Objetivo de Compra (la Rutina A al revés) ---

OBJETIVOS_DE_PRUEBA = {
    "articulos": [
        {
            "id": 901, "articulo_id": 1,
            "articulo_nombre": "Manzana Roja",
            "unidad_venta": "kilo",
            "fecha_ultima_compra": date(2026, 8, 10),
            "precio_bulto_ultima": 20000.0,
            "contenido_ultima": 18.0,
            "utilidad_actual": -0.34,
            "precio_vigente": 900.0,
            "entra_por_unidad": 760.5,
            "envase_por_unidad": 40.625,
            "umbral_envase": None,
            "envase_variable": False,
            "objetivo_por_unidad": 599.8958,
            "objetivo_bulto_ultima": 10798.125,
        }
    ],
    "sin_precio_vigente": [{"articulo_nombre": "Pera", "fecha_ultima_compra": date(2026, 8, 10)}],
    "sin_ficha": ["Morrón Rojo"],
    "utilidad_objetivo": 0.2,
}


def test_ver_objetivo_de_compra_sin_cliente_muestra_solo_el_selector():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/compras/objetivo")

    assert respuesta.status_code == 200
    assert "Elegí un cliente" in respuesta.text
    assert "Manzana" not in respuesta.text


def test_ver_objetivo_de_compra_con_cliente_muestra_los_articulos_bajo_objetivo():
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.calcular_objetivos_de_compra", return_value=OBJETIVOS_DE_PRUEBA) as mock_objetivos,
    ):
        respuesta = cliente.get("/compras/objetivo?cliente_id=1")

    assert respuesta.status_code == 200
    mock_objetivos.assert_called_once_with(1)
    assert "Manzana Roja" in respuesta.text
    # La utilidad objetivo del cliente, arriba de todo.
    assert "20.0%" in respuesta.text
    # Última compra y objetivo inicial (con el kilaje de esa compra).
    assert "$20.000" in respuesta.text
    assert "$10.798" in respuesta.text
    # Los números que necesita el recálculo en vivo viajan en data-attrs, y
    # el input de kilos viene precargado con el contenido de la última compra.
    assert 'data-entra="760.5"' in respuesta.text
    assert 'data-envase-unidad="40.625"' in respuesta.text
    assert 'class="input-kilos"' in respuesta.text
    assert 'value="18"' in respuesta.text
    # Las listas aparte, nunca en silencio.
    assert "Pera" in respuesta.text
    assert "Morrón Rojo" in respuesta.text


def test_objetivo_de_compra_etiqueta_el_input_con_la_unidad_de_la_ficha():
    # El input de contenido lleva SU unidad real (de la ficha logística del
    # cliente) + "por cajón": sirve de control — si en Palta dijera
    # "Kilos", la ficha está mal cargada. Nada de "Kilo/Unidad/Cubeta"
    # barreado: el sistema ya sabe cuál es.
    base = dict(OBJETIVOS_DE_PRUEBA["articulos"][0])
    objetivos = dict(
        OBJETIVOS_DE_PRUEBA,
        articulos=[
            dict(base, articulo_nombre="Manzana Roja", unidad_venta="kilo"),
            dict(base, articulo_nombre="Palta", unidad_venta="unidad"),
            dict(base, articulo_nombre="Frutilla", unidad_venta="cubeta"),
        ],
    )
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.calcular_objetivos_de_compra", return_value=objetivos),
    ):
        respuesta = cliente.get("/compras/objetivo?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Kilos por cajón" in respuesta.text
    assert "Unidades por cajón" in respuesta.text
    assert "Cubetas por cajón" in respuesta.text
    assert "por bulto" not in respuesta.text


def test_ver_objetivo_de_compra_cliente_sin_utilidad_objetivo_avisa():
    objetivos = {"articulos": [], "sin_precio_vigente": [], "sin_ficha": [], "utilidad_objetivo": None}
    with (
        patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA),
        patch("app.main.calcular_objetivos_de_compra", return_value=objetivos),
    ):
        respuesta = cliente.get("/compras/objetivo?cliente_id=1")

    assert respuesta.status_code == 200
    assert "no tiene una utilidad objetivo vigente" in respuesta.text


def test_ver_objetivo_de_compra_cliente_inexistente_da_404():
    with patch("app.main.listar_clientes", return_value=CLIENTES_DE_PRUEBA):
        respuesta = cliente.get("/compras/objetivo?cliente_id=999")

    assert respuesta.status_code == 404


def test_ver_compras_tiene_el_boton_objetivo_de_compra():
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/compras")

    assert respuesta.status_code == 200
    assert 'href="/compras/objetivo"' in respuesta.text
    assert "Objetivo de Compra" in respuesta.text


# --- /envases: costos de envase por cliente, con historial ---

ENVASES_CON_COSTO_DE_PRUEBA = [
    {"id": 7, "nombre": "Caja Chica Día", "costo": 650.0, "vigente_desde": date(2026, 8, 1), "fichas_que_lo_usan": 3},
    {"id": 8, "nombre": "Caja Grande Día", "costo": None, "vigente_desde": None, "fichas_que_lo_usan": 0},
]
HISTORIAL_ENVASES_DE_PRUEBA = [
    {"envase_id": 7, "costo": 650.0, "vigente_desde": date(2026, 8, 1)},
    {"envase_id": 7, "costo": 500.0, "vigente_desde": date(2026, 7, 1)},
]


def _patches_envases():
    return (
        patch("app.main.listar_envases_con_costo", return_value=ENVASES_CON_COSTO_DE_PRUEBA),
        patch("app.main.listar_historial_costos_envases", return_value=HISTORIAL_ENVASES_DE_PRUEBA),
    )


def test_ver_envases_muestra_costo_vigente_historial_y_advertencia():
    # El catálogo es compartido: la pantalla lista TODO, sin elegir cliente.
    parches = _patches_envases()
    with parches[0], parches[1]:
        respuesta = cliente.get("/envases")

    assert respuesta.status_code == 200
    assert "Caja Chica Día" in respuesta.text
    assert "$650" in respuesta.text
    assert "Vigente desde el 01/08/2026" in respuesta.text
    assert "lo usan 3 artículos entre todos los clientes" in respuesta.text
    # Envase sin costo cargado: se dice, no se inventa un cero.
    assert "Sin costo" in respuesta.text
    # El historial completo, para ver la evolución.
    assert "$500" in respuesta.text
    assert "01/07/2026" in respuesta.text
    # La advertencia pedida antes de dejar cambiar un costo.
    assert "cambia el precio sugerido de TODOS los artículos, de todos los clientes" in respuesta.text
    # El catálogo es compartido y la regla de oro, explicadas en la pantalla.
    assert "Catálogo único de envases" in respuesta.text
    assert "nunca pisan el" in respuesta.text


def test_cambiar_costo_envase_registra_fila_nueva_y_vuelve_con_aviso():
    with patch("app.main.registrar_costo_envase") as mock_registrar:
        respuesta = cliente.post("/envases/7/costo", data={"costo": "800"}, follow_redirects=False)

    assert respuesta.status_code == 303
    mock_registrar.assert_called_once_with(7, 800.0)
    assert "/envases?" in respuesta.headers["location"]
    assert "vigente+desde+hoy" in respuesta.headers["location"]


def test_cambiar_costo_envase_invalido_muestra_error_sin_registrar():
    parches = _patches_envases()
    with parches[0], parches[1], patch("app.main.registrar_costo_envase") as mock_registrar:
        respuesta = cliente.post("/envases/7/costo", data={"costo": "-5"})

    assert respuesta.status_code == 400
    assert "El costo tiene que ser mayor a cero." in respuesta.text
    mock_registrar.assert_not_called()


def test_dar_de_baja_envase_registra_costo_cero_desde_hoy():
    with patch("app.main.registrar_costo_envase") as mock_registrar:
        respuesta = cliente.post("/envases/7/baja", follow_redirects=False)

    assert respuesta.status_code == 303
    mock_registrar.assert_called_once_with(7, 0)


def test_agregar_envase_crea_y_vuelve_con_aviso():
    with patch("app.main.crear_envase") as mock_crear:
        respuesta = cliente.post("/envases/nuevo", data={"nombre": "Caja Mediana", "costo": "700"}, follow_redirects=False)

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("Caja Mediana", 700.0)


def test_agregar_envase_con_nombre_repetido_muestra_el_error():
    parches = _patches_envases()
    with (
        parches[0], parches[1],
        patch("app.main.crear_envase", side_effect=ValueError("Ya existe un envase con ese nombre.")),
    ):
        respuesta = cliente.post("/envases/nuevo", data={"nombre": "Caja Chica Día", "costo": "700"})

    assert respuesta.status_code == 400
    assert "Ya existe un envase con ese nombre" in respuesta.text


def test_ver_comercial_tiene_el_boton_envases():
    with patch("app.main.listar_estado_alertas", return_value=_foto_alertas()):
        respuesta = cliente.get("/comercial")

    assert respuesta.status_code == 200
    assert 'href="/envases"' in respuesta.text
    assert "Envases" in respuesta.text


# --- tipo_retiro Cooperativa ---


def test_form_de_compra_manual_ofrece_cooperativa():
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.get("/compras/nueva?proveedor_id=200")

    assert respuesta.status_code == 200
    assert 'value="Cooperativa"' in respuesta.text


def test_logistica_no_tiene_boton_cooperativa():
    # Las compras Cooperativa nacen retiradas: no hay nada que procesar en
    # Logística, así que no existe pantalla ni botón para ese tipo.
    respuesta = cliente.get("/logistica")

    assert respuesta.status_code == 200
    assert "Cooperativa" not in respuesta.text


def test_detalle_tiene_etiqueta_para_el_origen_cooperativa():
    from app.main import ORIGENES_RETIRO_LABELS

    assert ORIGENES_RETIRO_LABELS["automatico_cooperativa"] == "Retiro a cargo de la Cooperativa (automático)"
    assert ORIGENES_RETIRO_LABELS["automatico_carro"] == "Retiro a cargo del Carrero (automático)"


# --- /logistica/consultar: el histórico de retiros ---

RETIROS_DE_PRUEBA = [
    {"id": 1, "fecha_operacion": date(2026, 8, 16), "retiro_procesado_el": datetime(2026, 8, 16, 7, 30),
     "tipo_retiro": "Carro", "estado_retiro": "retirado", "estado": "pendiente", "cantidad_cajones": 10.0,
     "cantidad_cajones_retirada": None, "proveedor_nombre": "Saturno",
     "proveedor_codigo_puesto": "N07P41", "articulo_nombre": "Kiwi"},
    {"id": 2, "fecha_operacion": date(2026, 8, 16), "retiro_procesado_el": datetime(2026, 8, 16, 8, 15),
     "tipo_retiro": "Clark", "estado_retiro": "retirado", "estado": "recepcionado", "cantidad_cajones": 8.0,
     "cantidad_cajones_retirada": 7.0, "proveedor_nombre": "Saturno",
     "proveedor_codigo_puesto": "N07P41", "articulo_nombre": "Mango"},
    {"id": 3, "fecha_operacion": date(2026, 8, 16), "retiro_procesado_el": None,
     "tipo_retiro": "Clark", "estado_retiro": "pendiente", "estado": "pendiente", "cantidad_cajones": 5.0,
     "cantidad_cajones_retirada": None, "proveedor_nombre": "Crefu",
     "proveedor_codigo_puesto": "N03P12", "articulo_nombre": "Palta"},
]


def test_consultar_retiros_cortada_por_el_tope_avisa_y_oculta_los_totales():
    # Con la lista cortada, el total de bultos NO se muestra: un total
    # parcial usado para liquidarle al carrero sería un número falso.
    muchos = [dict(RETIROS_DE_PRUEBA[0], id=i) for i in range(TOPE_FILAS_BUSQUEDA + 1)]
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=muchos),
        patch("app.main.contar_retiros_buscados", return_value=2000) as mock_contar,
    ):
        respuesta = cliente.get("/logistica/consultar")

    assert respuesta.status_code == 200
    assert f"Se muestran los primeros {TOPE_FILAS_BUSQUEDA} retiros de 2000" in respuesta.text
    assert "Total:" not in respuesta.text
    mock_contar.assert_called_once()


def test_consultar_retiros_default_48hs_y_total_desglosado():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=RETIROS_DE_PRUEBA) as mock_buscar,
    ):
        respuesta = cliente.get("/logistica/consultar")

    assert respuesta.status_code == 200
    mock_buscar.assert_called_once_with(
        HOY_DE_PRUEBA - timedelta(days=1), HOY_DE_PRUEBA, None, None, None, None, limite=TOPE_FILAS_BUSQUEDA + 1
    )
    # El total para liquidar al carrero/cooperativa, desglosado: lo anotado
    # al retirar (7) + lo tomado de la carga del comprador (10 + 5).
    assert "Total: 22 bultos" in respuesta.text
    assert "7 anotados al retirar + 15* de la carga del comprador" in respuesta.text
    assert "se usa lo que cargó el comprador" in respuesta.text
    # En la tabla: con asterisco lo que viene de la carga, sin asterisco lo anotado.
    assert "10*" in respuesta.text
    assert "5*" in respuesta.text
    assert "7*" not in respuesta.text
    # Pendiente: sin hora de retiro.
    assert "Pendiente" in respuesta.text
    # Sin ninguna compra "no ingresó", el renglón del neto no aparece.
    assert "no ingresaron al depósito" not in respuesta.text


def test_consultar_retiros_marca_las_no_ingresadas_y_desglosa_el_neto():
    # Carro/Cooperativa nacen retiradas solas y el total se usa para
    # liquidarle a terceros — pero si Depósito marcó "no ingresó", esa
    # mercadería nunca llegó. NUNCA se resta en silencio: el total sigue
    # siendo el retirado (22), y aparte se ve cuánto no ingresó (10) y el
    # neto (12) — los dos números a la vista, el dueño decide qué paga.
    retiros = [dict(r) for r in RETIROS_DE_PRUEBA]
    retiros[0]["estado"] = "no_ingresado"  # la de Carro de 10 bultos
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=retiros),
    ):
        respuesta = cliente.get("/logistica/consultar")

    assert respuesta.status_code == 200
    # El total NO cambia: sigue mostrando todo lo retirado.
    assert "Total: 22 bultos" in respuesta.text
    assert "10 no ingresaron al depósito — Neto: 12 bultos" in respuesta.text
    # Y la fila queda marcada con la etiqueta visible.
    assert "No ingresó al depósito" in respuesta.text


def test_consultar_retiros_pasa_los_filtros_a_la_consulta():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=[]) as mock_buscar,
    ):
        respuesta = cliente.get(
            "/logistica/consultar?fecha_desde=2026-08-10&fecha_hasta=2026-08-12"
            "&proveedor_id=7&articulo_id=5&tipo=Cooperativa&estado=pendiente"
        )

    assert respuesta.status_code == 200
    mock_buscar.assert_called_once_with(
        date(2026, 8, 10), date(2026, 8, 12), 7, 5, "Cooperativa", "pendiente", limite=TOPE_FILAS_BUSQUEDA + 1
    )
    assert "No se encontraron retiros" in respuesta.text


def test_consultar_retiros_ofrece_exportar_con_todos_los_filtros():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=RETIROS_DE_PRUEBA),
    ):
        respuesta = cliente.get(
            "/logistica/consultar?fecha_desde=2026-08-10&fecha_hasta=2026-08-12&tipo=Carro&estado=retirado"
        )

    assert respuesta.status_code == 200
    assert (
        "/logistica/consultar/exportar-pdf?fecha_desde=2026-08-10&fecha_hasta=2026-08-12"
        "&proveedor_id=&articulo_id=&tipo=Carro&estado=retirado"
    ) in respuesta.text
    assert "/logistica/consultar/exportar-excel?fecha_desde=2026-08-10" in respuesta.text


def test_exportar_retiros_pdf_sin_tope_y_con_el_total_desglosado():
    retiros = [dict(r) for r in RETIROS_DE_PRUEBA]
    retiros[0]["estado"] = "no_ingresado"  # la de Carro de 10 bultos
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=retiros) as mock_buscar,
    ):
        respuesta = cliente.get(
            "/logistica/consultar/exportar-pdf?fecha_desde=2026-08-10&fecha_hasta=2026-08-12&tipo=Carro"
        )

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "Retiros_2026-08-10_a_2026-08-12" in respuesta.headers["content-disposition"]
    # SIN tope: el export no pasa limite, aunque la pantalla corte en 500.
    mock_buscar.assert_called_once_with(date(2026, 8, 10), date(2026, 8, 12), None, None, "Carro", None)
    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Consultar Retiros" in texto
    assert "tipo Carro" in texto  # los filtros aplicados viajan al subtítulo
    # El total para liquidar, con el mismo desglose que la pantalla.
    assert "Total: 22 bultos" in texto
    assert "10 no ingresaron al depósito — Neto: 12 bultos" in texto
    assert "No ingresó al depósito" in texto  # la fila marcada


def test_exportar_retiros_excel_devuelve_archivo_adjunto():
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=RETIROS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/logistica/consultar/exportar-excel?fecha_desde=2026-08-10&fecha_hasta=2026-08-12")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Retiros_2026-08-10_a_2026-08-12" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip


def test_exportar_retiros_pdf_fecha_invalida_da_400():
    respuesta = cliente.get("/logistica/consultar/exportar-pdf?fecha_desde=no-es-fecha&fecha_hasta=2026-08-12")

    assert respuesta.status_code == 400


def test_consultar_retiros_ofrece_cooperativa_en_el_filtro_de_tipo():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_retiros", return_value=[]),
    ):
        respuesta = cliente.get("/logistica/consultar")

    assert 'value="Cooperativa"' in respuesta.text
    assert 'value="Carro"' in respuesta.text


def test_editar_compra_etiqueta_el_contenido_con_la_unidad_del_articulo():
    # Mismo criterio que Objetivo de Compra: el input de contenido dice su
    # unidad real ("Kilos por cajón"), no "Contenido" a secas — se edita
    # el contenido de UN bulto y sin la unidad se confunde en el Mercado.
    compra = {
        "id": 30, "fecha_operacion": HOY_DE_PRUEBA, "articulo_id": 5, "articulo_nombre": "Kiwi",
        "proveedor_id": 200, "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10, "contenido_por_cajon": 18, "cantidad_kilos": 180, "cantidad_fraccion": None,
        "importe": 5000, "sena": None, "tipo_retiro": "Clark", "foto_ruta": None,
        "estado": None, "estado_retiro": "pendiente",
    }
    with (
        patch("app.main.obtener_compra", return_value=compra),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    # Kiwi es por kilo: la etiqueta lo dice desde el arranque (sin JS).
    assert "Kilos por cajón *" in respuesta.text
    assert 'id="label-contenido-por-cajon"' in respuesta.text


def test_editar_compra_muestra_las_fotos_de_la_guia_con_subir_y_borrar():
    compra = {
        "id": 30, "fecha_operacion": HOY_DE_PRUEBA, "articulo_id": 5, "articulo_nombre": "Kiwi",
        "proveedor_id": 200, "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "guia_id": 105,
        "cantidad_cajones": 10, "contenido_por_cajon": 18, "cantidad_kilos": 180, "cantidad_fraccion": None,
        "importe": 5000, "sena": None, "tipo_retiro": "Clark",
        "estado": None, "estado_retiro": "pendiente",
    }
    fotos = [{"id": 9, "foto_ruta": "2026/x.jpg", "creado_en": datetime(2026, 8, 6, 10, 0)}]
    with (
        patch("app.main.obtener_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=fotos),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/30/editar?fecha_desde=2026-08-01")

    assert respuesta.status_code == 200
    # La miniatura de la foto, el borrar por foto, y el botón de sumar otra.
    assert 'src="/compras/30/fotos/9/ver"' in respuesta.text
    assert 'action="/compras/30/fotos/9/borrar"' in respuesta.text
    assert 'action="/compras/30/fotos"' in respuesta.text
    assert "Agregar foto o archivo" in respuesta.text
    # Los filtros de Buscar viajan ocultos para volver a la misma búsqueda.
    assert 'name="query_filtros" value="fecha_desde=2026-08-01"' in respuesta.text


def test_editar_compra_sin_foto_no_muestra_el_boton():
    compra = {
        "id": 30, "fecha_operacion": HOY_DE_PRUEBA, "articulo_id": 5, "articulo_nombre": "Kiwi",
        "proveedor_id": 200, "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
        "cantidad_cajones": 10, "contenido_por_cajon": 18, "cantidad_kilos": 180, "cantidad_fraccion": None,
        "importe": 5000, "sena": None, "tipo_retiro": "Clark", "foto_ruta": None,
        "estado": None, "estado_retiro": "pendiente",
    }
    with (
        patch("app.main.obtener_compra", return_value=compra),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.get("/compras/30/editar")

    assert respuesta.status_code == 200
    assert "Ver foto de la comanda" not in respuesta.text


# --- Vacíos (Envases Puesto): cajones de proveedores que entran y salen del puesto ---

# Las MISMAS claves que devuelve listar_tipos_envase_puesto: id,
# proveedor_id, nombre y proveedor_nombre. Traía además un codigo_puesto
# que esa consulta nunca devolvió — nadie lo leía, pero una fixture que
# miente es la que dejó pasar el 500 de Guías R.
TIPOS_ENVASE_PUESTO_DE_PRUEBA = [
    {"id": 1, "proveedor_id": 200, "nombre": "cajón plástico negro", "proveedor_nombre": "Saturno"},
    {"id": 2, "proveedor_id": 200, "nombre": "torito", "proveedor_nombre": "Saturno"},
    {"id": 3, "proveedor_id": 201, "nombre": "cajón madera", "proveedor_nombre": "Don Pepe"},
]

CLIENTES_PUESTO_DE_PRUEBA = [
    {"id": 10, "nombre": "Juan Pérez"},
    {"id": 11, "nombre": "Marta"},
]


def test_ver_puesto_es_un_hub_con_envases_puesto():
    respuesta = cliente.get("/puesto")

    assert respuesta.status_code == 200
    assert "En construcción" not in respuesta.text
    assert 'href="/puesto/envases"' in respuesta.text
    assert "Envases Puesto" in respuesta.text


def test_ver_envases_puesto_separa_vacios_del_resto():
    respuesta = cliente.get("/puesto/envases")

    assert respuesta.status_code == 200
    assert 'href="/puesto/envases/vacios"' in respuesta.text
    assert 'href="/puesto/envases/stock"' in respuesta.text
    assert 'href="/puesto/envases/cotejo"' in respuesta.text
    assert 'href="/puesto/envases/pendientes"' in respuesta.text
    assert 'href="/puesto/envases/movimientos"' in respuesta.text
    assert 'href="/puesto/envases/tipos"' in respuesta.text
    # Sin botón de Clientes del Puesto: los clientes se crean solos al
    # tipear en Recibir — la ruta sigue viva, pero sin botón en el hub.
    assert 'href="/puesto/envases/clientes"' not in respuesta.text


def test_ver_vacios_muestra_las_tres_pantallas_del_empleado():
    respuesta = cliente.get("/puesto/envases/vacios")

    assert respuesta.status_code == 200
    assert 'href="/puesto/envases/vacios/recibir"' in respuesta.text
    assert 'href="/puesto/envases/vacios/devolver"' in respuesta.text
    assert 'href="/puesto/envases/vacios/stock-fisico"' in respuesta.text
    # El hub del empleado NO linkea al stock del sistema.
    assert 'href="/puesto/envases/stock"' not in respuesta.text


def test_ver_vacios_muestra_el_aviso_del_redirect():
    # La confirmación de Recibir/Devolver viaja por query string y se
    # muestra en el hub, adonde vuelve el empleado después de guardar.
    respuesta = cliente.get("/puesto/envases/vacios", params={"aviso": "Recibidos 12 cajones (Cajón manzana) de Saturno."})

    assert respuesta.status_code == 200
    assert "Recibidos 12 cajones (Cajón manzana) de Saturno." in respuesta.text


def test_ver_recibir_vacios_arma_el_form_con_lista_cerrada():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_clientes_puesto", return_value=CLIENTES_PUESTO_DE_PRUEBA),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/vacios/recibir")

    assert respuesta.status_code == 200
    # Proveedores derivados de los tipos (lista cerrada): uno por proveedor.
    # Sin código de puesto: son proveedores del PUESTO, no los de Compras.
    assert respuesta.text.count('value="200"') == 1
    assert "Saturno" in respuesta.text
    assert "Don Pepe" in respuesta.text
    assert "N07P41" not in respuesta.text
    # Clientes conocidos como sugerencias al tipear.
    assert 'list="clientes-conocidos"' in respuesta.text
    assert "Juan Pérez" in respuesta.text
    # Tipos por proveedor para el JS que preselecciona el primero.
    assert "TIPOS_POR_PROVEEDOR" in respuesta.text
    assert "cajón plástico negro" in respuesta.text


def test_ver_recibir_vacios_sin_tipos_explica_como_habilitar():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=[]),
        patch("app.main.listar_clientes_puesto", return_value=[]),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/vacios/recibir")

    assert respuesta.status_code == 200
    assert "Todavía no hay tipos de envase cargados" in respuesta.text


def test_recibir_vacios_guarda_con_cliente_normalizado():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_cliente_puesto", return_value=10) as mock_cliente,
        patch("app.main.crear_vacio_recibido") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/recibir",
            data={
                "cliente_nombre": "  JUAN   Pérez ",
                "proveedor_id": "200",
                "tipo_envase_id": "2",
                "cantidad": "12",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Después de guardar se vuelve al hub de Vacíos, con la confirmación ahí.
    assert respuesta.headers["location"].startswith("/puesto/envases/vacios?aviso=")
    # El nombre se limpia (espacios) y se normaliza (minúsculas, sin acentos)
    # para que "Juan", "juan " y "JUAN" sean EL MISMO cliente.
    mock_cliente.assert_called_once_with("JUAN Pérez", "juan perez")
    mock_crear.assert_called_once_with(10, 200, 2, 12)


def test_recibir_vacios_rechaza_tipo_que_no_es_del_proveedor():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_clientes_puesto", return_value=[]),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
        patch("app.main.crear_vacio_recibido") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/recibir",
            data={
                "cliente_nombre": "Juan",
                "proveedor_id": "201",
                "tipo_envase_id": "1",  # el tipo 1 es de Saturno, no de Don Pepe
                "cantidad": "12",
            },
        )

    assert respuesta.status_code == 400
    assert "Elegí un proveedor y un tipo de envase válidos." in respuesta.text
    mock_crear.assert_not_called()


def test_recibir_vacios_sin_nombre_de_cliente_da_error():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_clientes_puesto", return_value=[]),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
        patch("app.main.crear_vacio_recibido") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/recibir",
            data={"cliente_nombre": "   ", "proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "12"},
        )

    assert respuesta.status_code == 400
    assert "El nombre del cliente es obligatorio." in respuesta.text
    mock_crear.assert_not_called()


def test_recibir_vacios_con_cantidad_invalida_da_error():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_clientes_puesto", return_value=[]),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
        patch("app.main.crear_vacio_recibido") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/recibir",
            data={"cliente_nombre": "Juan", "proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "0"},
        )

    assert respuesta.status_code == 400
    assert "mayor a cero" in respuesta.text
    mock_crear.assert_not_called()


def test_anular_vacio_recibido_redirige_a_recibir():
    with patch("app.main.anular_vacio_recibido") as mock_anular:
        respuesta = cliente.post("/puesto/envases/vacios/recibidos/5/anular", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases/vacios/recibir"
    mock_anular.assert_called_once_with(5)


def test_devolver_vacios_con_stock_suficiente_avisa_sin_advertencia():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_vacio_devuelto", return_value=40) as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/devolver",
            data={"proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "30"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    # Después de guardar se vuelve al hub de Vacíos, con la confirmación ahí.
    assert location.startswith("/puesto/envases/vacios?aviso=")
    assert "Ojo" not in location
    mock_crear.assert_called_once_with(200, 1, 30)


def test_devolver_vacios_que_supera_el_stock_guarda_igual_y_lo_dice():
    # Nunca se bloquea: el camión se lleva los cajones aunque el sistema
    # esté atrasado. La diferencia queda grabada (stock_sistema en la
    # fila) y el aviso lo dice.
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_vacio_devuelto", return_value=40) as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/devolver",
            data={"proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "50"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert "seg%C3%BAn+el+sistema+hab%C3%ADa+40" in location
    mock_crear.assert_called_once_with(200, 1, 50)


def test_anular_vacio_devuelto_redirige_a_devolver():
    with patch("app.main.anular_vacio_devuelto") as mock_anular:
        respuesta = cliente.post("/puesto/envases/vacios/devueltos/7/anular", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases/vacios/devolver"
    mock_anular.assert_called_once_with(7)


def test_ver_stock_vacios_agrupa_por_proveedor_y_marca_negativos():
    filas = [
        {"proveedor_id": 200, "proveedor_nombre": "Saturno",
         "tipo_envase_id": 1, "tipo_nombre": "cajón plástico negro",
         "recibidos": 50, "devueltos": 30, "ajustes": 0, "stock": 20},
        {"proveedor_id": 200, "proveedor_nombre": "Saturno",
         "tipo_envase_id": 2, "tipo_nombre": "torito",
         "recibidos": 10, "devueltos": 15, "ajustes": 0, "stock": -5},
        {"proveedor_id": 201, "proveedor_nombre": "Don Pepe",
         "tipo_envase_id": 3, "tipo_nombre": "cajón madera",
         "recibidos": 8, "devueltos": 0, "ajustes": 0, "stock": 8},
    ]
    with patch("app.main.stock_vacios", return_value=filas):
        respuesta = cliente.get("/puesto/envases/stock")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert "Don Pepe" in respuesta.text
    # Con ajustes en 0 el desglose no los menciona.
    assert "50 recibidos − 30 devueltos)" in respuesta.text
    # El negativo va marcado en rojo (clase negativo).
    assert 'class="numero negativo">-5<' in respuesta.text
    # Total del proveedor con más de un tipo: 20 + (-5) = 15.
    assert ">Total</span>" in respuesta.text
    assert ">15</span>" in respuesta.text
    # Por default es el stock de HOY: sin aviso de fecha pasada, y con
    # el selector de fecha y el botón de exportar a la vista.
    assert "puede cambiar si se anulan movimientos anteriores" not in respuesta.text
    assert 'name="fecha"' in respuesta.text
    assert "/puesto/envases/stock/exportar-pdf?fecha=" in respuesta.text


STOCK_VACIOS_DE_PRUEBA = [
    {"proveedor_id": 200, "proveedor_nombre": "Saturno",
     "tipo_envase_id": 1, "tipo_nombre": "cajón plástico negro",
     "recibidos": 50, "devueltos": 30, "ajustes": -5, "stock": 15},
]


def test_ver_stock_vacios_a_fecha_pasada_avisa_que_puede_cambiar():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.stock_vacios", return_value=STOCK_VACIOS_DE_PRUEBA) as mock_stock,
    ):
        respuesta = cliente.get("/puesto/envases/stock?fecha=2026-08-10")

    assert respuesta.status_code == 200
    # El stock se calcula A ESA FECHA (la consulta recibe la fecha elegida).
    mock_stock.assert_called_once_with(date(2026, 8, 10))
    # Aviso pedido a propósito: anular un movimiento lo borra del stock de
    # TODOS los días (desde siempre), así que una fecha pasada puede dar
    # distinto en dos consultas — no es una falla del sistema.
    assert "puede cambiar si se anulan movimientos anteriores" in respuesta.text
    assert "Stock al 10/08/2026" in respuesta.text
    # Los exports llevan la misma fecha consultada.
    assert "/puesto/envases/stock/exportar-pdf?fecha=2026-08-10" in respuesta.text
    assert "/puesto/envases/stock/exportar-excel?fecha=2026-08-10" in respuesta.text


def test_ver_stock_vacios_fecha_invalida_o_futura_cae_a_hoy():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.stock_vacios", return_value=STOCK_VACIOS_DE_PRUEBA) as mock_stock,
    ):
        respuesta = cliente.get("/puesto/envases/stock?fecha=2027-01-01")

    assert respuesta.status_code == 200
    mock_stock.assert_called_once_with(date(2026, 8, 19))
    assert "puede cambiar si se anulan movimientos anteriores" not in respuesta.text


def test_exportar_stock_vacios_pdf_lleva_fecha_y_aviso_si_es_pasada():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.stock_vacios", return_value=STOCK_VACIOS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/puesto/envases/stock/exportar-pdf?fecha=2026-08-10")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "Stock_Vacios_2026-08-10" in respuesta.headers["content-disposition"]
    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    assert "Stock del Sistema" in texto
    assert "Al 10/08/2026" in texto
    assert "puede cambiar si se anulan movimientos anteriores" in texto
    assert "Saturno" in texto


def test_exportar_stock_vacios_excel_devuelve_archivo_adjunto():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.stock_vacios", return_value=STOCK_VACIOS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/puesto/envases/stock/exportar-excel?fecha=2026-08-19")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Stock_Vacios_2026-08-19" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip


PROVEEDORES_PUESTO_DE_PRUEBA = [
    {"id": 200, "nombre": "Saturno"},
    {"id": 201, "nombre": "Don Pepe"},
]


VALORES_SENA_DE_PRUEBA = {
    1: {"tipo_envase_id": 1, "tipo_nombre": "cajón plástico negro", "proveedor_nombre": "Saturno",
        "monto": 700, "vigente_desde": date(2026, 8, 20), "ultima_vigencia": date(2026, 8, 20)},
    2: {"tipo_envase_id": 2, "tipo_nombre": "torito", "proveedor_nombre": "Saturno",
        "monto": None, "vigente_desde": None, "ultima_vigencia": None},
}


def _pantalla_tipos(tipos=None, valores=None, historiales=None):
    """Los cuatro parches que necesita la pantalla de Tipos de Envase.

    Desde que la seña vive acá, el render pide también el valor vigente de
    cada tipo y el historial de todos (en una consulta, no una por
    renglón).
    """
    tipos = TIPOS_ENVASE_PUESTO_DE_PRUEBA if tipos is None else tipos
    return (
        patch("app.main.listar_tipos_envase_puesto", return_value=tipos),
        patch("app.main.listar_proveedores_puesto", return_value=PROVEEDORES_PUESTO_DE_PRUEBA),
        patch("app.main.listar_valores_sena",
              return_value=list((VALORES_SENA_DE_PRUEBA if valores is None else valores).values())),
        patch("app.main.listar_historiales_valores_sena", return_value=historiales or {}),
    )


def test_ver_tipos_envase_puesto_lista_y_ofrece_alta():
    # El select de proveedores es de proveedores_puesto (cajera), NUNCA los
    # de Compras.
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        respuesta = cliente.get("/puesto/envases/tipos")

    assert respuesta.status_code == 200
    assert 'action="/puesto/envases/tipos/nuevo"' in respuesta.text
    assert "cajón plástico negro" in respuesta.text
    assert 'action="/puesto/envases/tipos/1/baja"' in respuesta.text


def test_crear_tipo_envase_puesto_limpia_el_nombre_y_redirige():
    with patch("app.main.crear_tipo_envase_puesto") as mock_crear:
        respuesta = cliente.post(
            "/puesto/envases/tipos/nuevo",
            data={"proveedor_id": "200", "nombre": "  cajón   negro "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(200, "cajón negro")


def test_el_alta_de_un_tipo_acepta_la_sena_pero_NO_la_exige():
    """Opcional a propósito.

    Si fuera obligatoria, el día que no se sepa cuánto vale la seña van a
    inventar un número para poder seguir, y un número inventado es peor
    que "sin valor cargado": se paga.
    """
    with (
        patch("app.main.crear_tipo_envase_puesto", return_value=9) as mock_crear,
        patch("app.main.cargar_valor_sena") as mock_cargar,
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/nuevo",
            data={"proveedor_id": "200", "nombre": "cajón negro", "monto": "", "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(200, "cajón negro")
    # Sin monto, el tipo nace igual y sin seña.
    mock_cargar.assert_not_called()


def test_el_alta_con_sena_se_la_cuelga_al_tipo_recien_creado():
    with (
        patch("app.main.crear_tipo_envase_puesto", return_value=9),
        patch("app.main.cargar_valor_sena") as mock_cargar,
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/nuevo",
            data={"proveedor_id": "200", "nombre": "cajón negro", "monto": "500",
                  "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Al id que devolvió el alta, no a otro.
    mock_cargar.assert_called_once_with(9, 500.0, date(2026, 8, 28))


def test_el_hub_ya_no_tiene_el_boton_de_valor_de_la_sena():
    # La pantalla se eliminó: no puede quedar un link a un 404.
    respuesta = cliente.get("/puesto/envases")

    assert respuesta.status_code == 200
    assert "/puesto/envases/senas" not in respuesta.text
    assert "Valor de la Seña" not in respuesta.text


def test_la_ruta_vieja_del_valor_de_la_sena_ya_no_existe():
    assert cliente.get("/puesto/envases/senas").status_code == 404
    assert cliente.post("/puesto/envases/senas/cargar", data={}).status_code == 404


def test_crear_tipo_envase_puesto_sin_nombre_da_error():
    with ExitStack() as pila:
        for parche in _pantalla_tipos(tipos=[]):
            pila.enter_context(parche)
        mock_crear = pila.enter_context(patch("app.main.crear_tipo_envase_puesto"))
        respuesta = cliente.post("/puesto/envases/tipos/nuevo", data={"proveedor_id": "200", "nombre": "  "})

    assert respuesta.status_code == 400
    assert "El nombre del tipo de envase es obligatorio." in respuesta.text
    mock_crear.assert_not_called()


def test_dar_de_baja_tipo_envase_puesto_redirige():
    with patch("app.main.desactivar_tipo_envase_puesto") as mock_baja:
        respuesta = cliente.post("/puesto/envases/tipos/3/baja", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases/tipos"
    mock_baja.assert_called_once_with(3)


def test_dar_de_baja_un_tipo_con_saldo_se_niega_y_lo_dice_con_el_numero():
    # Una cuenta abierta no es una falla del sistema: es algo para cerrar.
    # Se muestra en la pantalla con el número adentro (400, no 500).
    with ExitStack() as pila:
        for parche in _pantalla_tipos(tipos=[]):
            pila.enter_context(parche)
        pila.enter_context(patch(
            "app.main.desactivar_tipo_envase_puesto",
            side_effect=ValueError("'cajón madera' todavía tiene 12 en stock. Devolvelos o ajustá a cero antes de darlo de baja.")))
        respuesta = cliente.post("/puesto/envases/tipos/3/baja", follow_redirects=False)

    assert respuesta.status_code == 400
    assert "todavía tiene 12 en stock" in respuesta.text
    assert "ajustá a cero antes de darlo de baja" in respuesta.text


def test_el_renglon_del_tipo_muestra_la_sena_vigente_y_el_que_no_tiene_lo_dice():
    # La mudanza: la seña se ve en el renglón del tipo, no en una pantalla
    # aparte. Sin valor cargado NO es $0 — un cero se lee como dato real y
    # se paga.
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        respuesta = cliente.get("/puesto/envases/tipos")

    assert respuesta.status_code == 200
    cuerpo = respuesta.text.split("</style>")[-1]
    assert "$700" in cuerpo
    assert "desde el 20/08" in cuerpo
    assert "Sin valor cargado" in cuerpo
    assert "$0" not in cuerpo


def test_el_bloque_de_edicion_del_tipo_trae_nombre_Y_sena():
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        respuesta = cliente.get("/puesto/envases/tipos")

    assert respuesta.status_code == 200
    assert 'action="/puesto/envases/tipos/1/editar"' in respuesta.text
    # El nombre precargado, y los dos campos de la seña en el mismo bloque.
    assert 'value="cajón plástico negro"' in respuesta.text
    assert 'name="monto"' in respuesta.text
    assert 'name="vigente_desde"' in respuesta.text


def test_el_historial_de_la_sena_va_plegado_y_tacha_lo_reemplazado():
    historiales = {1: [
        {"monto": 700, "vigente_desde": date(2026, 8, 20), "creado_en": None, "reemplazada": False},
        {"monto": 7000, "vigente_desde": date(2026, 8, 20), "creado_en": None, "reemplazada": True},
    ], 2: [], 3: []}
    with ExitStack() as pila:
        for parche in _pantalla_tipos(historiales=historiales):
            pila.enter_context(parche)
        respuesta = cliente.get("/puesto/envases/tipos")

    assert respuesta.status_code == 200
    cuerpo = respuesta.text.split("</style>")[-1]
    # Plegado: no es lo que se mira todos los días.
    assert "<details" in cuerpo and "Historial de la seña (2)" in cuerpo
    # El que quedó atrás sigue a la vista, tachado y marcado.
    assert "$7.000" in cuerpo
    assert 'class="historial-fila reemplazada"' in cuerpo
    assert "reemplazado" in cuerpo
    # Y el tipo sin historial no muestra un desplegable vacío.
    assert cuerpo.count("<details") == 1


def test_editar_el_tipo_sin_monto_solo_renombra_y_NO_toca_la_sena():
    # El monto vacío significa "no toques la seña", no cero. Quien entra a
    # arreglar un tipeo no tiene por qué cargar un valor.
    with (
        patch("app.main.renombrar_tipo_envase_puesto") as mock_renombrar,
        patch("app.main.cargar_valor_sena") as mock_cargar,
        patch("app.main.contar_senas_afectadas_por_valor") as mock_contar,
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/3/editar",
            data={"nombre": "  cajón   negro ", "monto": "", "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_renombrar.assert_called_once_with(3, "cajón negro")
    mock_cargar.assert_not_called()
    # Sin monto no hay nada retroactivo que avisar.
    mock_contar.assert_not_called()


def test_editar_el_tipo_con_monto_renombra_Y_carga_la_sena():
    with (
        patch("app.main.renombrar_tipo_envase_puesto") as mock_renombrar,
        patch("app.main.cargar_valor_sena") as mock_cargar,
        patch("app.main.contar_senas_afectadas_por_valor", return_value=0),
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/3/editar",
            data={"nombre": "cajón negro", "monto": "700", "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_renombrar.assert_called_once_with(3, "cajón negro")
    mock_cargar.assert_called_once_with(3, 700.0, date(2026, 8, 28))


def test_el_cero_explicito_de_sena_SI_se_permite():
    # "Este envase no lleva seña" es un dato, y es distinto de no tener
    # fila. Por eso el corte es en el negativo, no en el cero.
    with (
        patch("app.main.renombrar_tipo_envase_puesto"),
        patch("app.main.cargar_valor_sena") as mock_cargar,
        patch("app.main.contar_senas_afectadas_por_valor", return_value=0),
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/3/editar",
            data={"nombre": "cajón negro", "monto": "0", "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_cargar.assert_called_once_with(3, 0.0, date(2026, 8, 28))


def test_LA_FECHA_RETROACTIVA_SIGUE_AVISANDO_despues_de_la_mudanza():
    """Lo que no se puede perder al mudar la pantalla.

    Si el aviso se pierde, cargar una fecha vieja mueve plata de señas ya
    recibidas en silencio. Y AVISA, NO TRABA: ofrece seguir.

    Se chequea además que NO escriba nada todavía — ni el nombre ni la
    seña — para que no quede la mitad hecha esperando una confirmación.
    """
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        pila.enter_context(patch("app.main.contar_senas_afectadas_por_valor", return_value=3))
        mock_cargar = pila.enter_context(patch("app.main.cargar_valor_sena"))
        mock_renombrar = pila.enter_context(patch("app.main.renombrar_tipo_envase_puesto"))
        respuesta = cliente.post(
            "/puesto/envases/tipos/1/editar",
            data={"nombre": "cajón plástico negro", "monto": "600", "vigente_desde": "2026-08-10"},
        )

    assert respuesta.status_code == 200
    assert "cambia el valor de 3 señas ya recibidas" in respuesta.text
    mock_cargar.assert_not_called()
    mock_renombrar.assert_not_called()
    # Ofrece seguir: es aviso, no traba.
    assert "Guardar igual" in respuesta.text
    assert 'name="confirmado" value="1"' in respuesta.text
    # Y dice QUÉ se está por guardar.
    assert "$600" in respuesta.text and "10/08/2026" in respuesta.text


def test_el_boton_de_la_advertencia_no_se_pinta_como_el_de_confirmar():
    """Ya pasó con "Dar de baja", que salía verde igual que "Cargar".

    button[type="submit"] es más específico que una clase suelta y le gana.
    Un botón del que la pantalla acaba de advertir no puede verse igual que
    el de la operación normal.
    """
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        pila.enter_context(patch("app.main.contar_senas_afectadas_por_valor", return_value=3))
        pila.enter_context(patch("app.main.cargar_valor_sena"))
        pila.enter_context(patch("app.main.renombrar_tipo_envase_puesto"))
        respuesta = cliente.post(
            "/puesto/envases/tipos/1/editar",
            data={"nombre": "cajón plástico negro", "monto": "600", "vigente_desde": "2026-08-10"},
        )

    estilos = respuesta.text.split("</style>")[0]
    assert "button.boton-igual" in estilos, "la regla tiene que ganarle a button[type=submit]"


def test_una_sola_sena_afectada_se_dice_en_singular():
    with ExitStack() as pila:
        for parche in _pantalla_tipos():
            pila.enter_context(parche)
        pila.enter_context(patch("app.main.contar_senas_afectadas_por_valor", return_value=1))
        pila.enter_context(patch("app.main.cargar_valor_sena"))
        pila.enter_context(patch("app.main.renombrar_tipo_envase_puesto"))
        respuesta = cliente.post(
            "/puesto/envases/tipos/1/editar",
            data={"nombre": "cajón plástico negro", "monto": "600", "vigente_desde": "2026-08-10"},
        )

    assert "cambia el valor de 1 seña ya recibida" in respuesta.text


def test_confirmada_la_retroactiva_guarda_las_dos_cosas_sin_volver_a_preguntar():
    with (
        patch("app.main.contar_senas_afectadas_por_valor") as mock_contar,
        patch("app.main.renombrar_tipo_envase_puesto") as mock_renombrar,
        patch("app.main.cargar_valor_sena") as mock_cargar,
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/1/editar",
            data={"nombre": "cajón plástico negro", "monto": "600",
                  "vigente_desde": "2026-08-10", "confirmado": "1"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_renombrar.assert_called_once_with(1, "cajón plástico negro")
    mock_cargar.assert_called_once_with(1, 600.0, date(2026, 8, 10))
    mock_contar.assert_not_called()


def test_editar_el_tipo_sin_nombre_da_error():
    with ExitStack() as pila:
        for parche in _pantalla_tipos(tipos=[]):
            pila.enter_context(parche)
        mock_renombrar = pila.enter_context(patch("app.main.renombrar_tipo_envase_puesto"))
        respuesta = cliente.post("/puesto/envases/tipos/3/editar", data={"nombre": "  "})

    assert respuesta.status_code == 400
    assert "El nombre del tipo de envase es obligatorio." in respuesta.text
    mock_renombrar.assert_not_called()


def test_editar_el_tipo_con_un_monto_invalido_da_400_y_no_escribe():
    for monto, esperado in (("cinco", "El monto de la seña tiene que ser un número."),
                            ("-5", "El monto de la seña no puede ser negativo.")):
        with ExitStack() as pila:
            for parche in _pantalla_tipos(tipos=[]):
                pila.enter_context(parche)
            mock_renombrar = pila.enter_context(patch("app.main.renombrar_tipo_envase_puesto"))
            mock_cargar = pila.enter_context(patch("app.main.cargar_valor_sena"))
            respuesta = cliente.post(
                "/puesto/envases/tipos/3/editar",
                data={"nombre": "cajón", "monto": monto, "vigente_desde": "2026-08-28"},
            )

        assert respuesta.status_code == 400
        assert esperado in respuesta.text
        mock_renombrar.assert_not_called()
        mock_cargar.assert_not_called()


def test_renombrar_tipo_a_un_nombre_repetido_da_400_con_el_nombre_nunca_500():
    # Chocar con uno que ya existe es dato mal cargado, no una falla del
    # sistema: se muestra en la pantalla nombrando al que ya está.
    with ExitStack() as pila:
        for parche in _pantalla_tipos(tipos=[]):
            pila.enter_context(parche)
        pila.enter_context(patch(
            "app.main.renombrar_tipo_envase_puesto",
            side_effect=ValueError("Ese proveedor ya tiene un tipo llamado 'cajón madera'.")))
        respuesta = cliente.post("/puesto/envases/tipos/3/editar", data={"nombre": "cajón madera"})

    assert respuesta.status_code == 400
    assert "ya tiene un tipo llamado" in respuesta.text
    assert "cajón madera" in respuesta.text


# --- Vacíos tanda 2: Stock Físico, Cotejo, Pendientes de Pago, Movimientos, Clientes ---


def test_ver_stock_fisico_no_muestra_ningun_numero_del_sistema():
    # El control cruzado depende de esto: la pantalla del empleado no puede
    # traer el stock del sistema ni escondido en el HTML.
    contados = [
        {"id": 1, "cantidad": 38, "creado_en": datetime(2026, 8, 19, 14, 0, tzinfo=timezone.utc),
         "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
    ]
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_conteos_vacios_de_fecha", return_value=contados) as mock_contados,
    ):
        respuesta = cliente.get("/puesto/envases/vacios/stock-fisico")

    assert respuesta.status_code == 200
    assert "Cajones contados" in respuesta.text
    assert "38 × cajón plástico negro" in respuesta.text
    # La fuente de datos es la función que NO trae stock_sistema.
    mock_contados.assert_called_once()
    assert "stock_sistema" not in respuesta.text


def test_cargar_stock_fisico_guarda_el_conteo_sin_devolver_el_stock():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_conteo_vacios") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/stock-fisico",
            data={"proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "0"},
            follow_redirects=False,
        )

    # Contar 0 es válido: contó y no hay ninguno.
    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/puesto/envases/vacios/stock-fisico?aviso=")
    # El aviso repite lo contado, jamás el número del sistema.
    assert "Conteo+guardado%3A+0" in location
    mock_crear.assert_called_once_with(200, 1, 0)


def test_cargar_stock_fisico_con_cantidad_negativa_da_error():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.listar_conteos_vacios_de_fecha", return_value=[]),
        patch("app.main.crear_conteo_vacios") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/vacios/stock-fisico",
            data={"proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "-3"},
        )

    assert respuesta.status_code == 400
    assert "no puede ser negativa" in respuesta.text
    mock_crear.assert_not_called()


def _conteo_cotejo(**cambios):
    """Un conteo del Cotejo con lo mínimo, para variar solo lo que prueba cada caso."""
    base = {
        "id": 1, "cantidad": 35, "stock_sistema": 40,
        "proveedor_id": 200, "tipo_envase_id": 1,
        "creado_en": datetime(2026, 8, 19, 14, 0, tzinfo=timezone.utc),
        "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro",
        "proveedor_activo": True, "tipo_activo": True,
        "ajustes_posteriores": 0, "stock_actual": 40,
    }
    return {**base, **cambios}


CONTEOS_COTEJO_DE_PRUEBA = [
    # Diferencia sin ajustar: lo único que hay para hacer hoy.
    _conteo_cotejo(),
    # Contó igual que el sistema: nada que hacer.
    _conteo_cotejo(id=2, cantidad=8, stock_sistema=8, stock_actual=8,
                   proveedor_id=201, tipo_envase_id=3,
                   creado_en=datetime(2026, 8, 19, 14, 5, tzinfo=timezone.utc),
                   proveedor_nombre="Don Pepe", tipo_nombre="cajón madera"),
]


def test_ver_cotejo_compara_contra_la_foto_del_conteo():
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=CONTEOS_COTEJO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    # Diferencia -5, resaltada; la tarjeta con diferencia lleva borde.
    assert "-5" in respuesta.text
    assert "con-diferencia" in respuesta.text
    assert "diferencia-cero" in respuesta.text


def test_ver_cotejo_ofrece_ajustar_a_lo_contado_solo_con_diferencia():
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=CONTEOS_COTEJO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    # La fila con diferencia (Saturno, contado 35 vs sistema 40) lleva el
    # botón con toda la precarga; la que dio igual, no.
    assert respuesta.text.count("Ajustar a lo contado") == 1
    assert "/puesto/envases/ajustar?" in respuesta.text
    assert "proveedor_id=200" in respuesta.text
    assert "tipo_envase_id=1" in respuesta.text
    assert "contado=35" in respuesta.text
    assert "stock_conteo=40" in respuesta.text
    assert "fecha_conteo=2026-08-19" in respuesta.text


def test_ver_cotejo_una_diferencia_ya_ajustada_se_pone_en_verde():
    """El bug que hacía que el módulo se abandonara.

    "Ajustar a lo contado" escribe un ajuste, NO un conteo nuevo. Como la
    tarjeta comparaba lo contado contra la foto congelada del conteo, la
    misma diferencia quedaba en rojo para siempre, con el botón invitando a
    ajustar de nuevo. Ordenabas todo y la pantalla seguía igual de roja.
    """
    ajustado = _conteo_cotejo(ajustes_posteriores=-5, stock_actual=35)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[ajustado]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert 'class="tarjeta con-diferencia"' not in respuesta.text
    assert "Ajustar a lo contado" not in respuesta.text
    assert "la diferencia se ajustó después del conteo" in respuesta.text
    # La diferencia histórica NO desaparece: es el dato de auditoría de ese día.
    assert "-5" in respuesta.text


def test_ver_cotejo_un_ajuste_posterior_a_un_conteo_que_COINCIDIO_no_inventa_un_faltante():
    """El bug del 28/08: el sin-explicar daba el ajuste con el signo cambiado.

    Restar los ajustes posteriores a secas asume que TODO ajuste posterior
    viene a tapar la diferencia del conteo. Si el conteo coincidió, no hay
    ninguna diferencia que tapar: un ajuste hecho después por otro motivo
    (aparecieron cajones) salía como "queda sin explicar" por su monto con
    el signo invertido — 657/-657, 9/-9, 257/-257.

    Y era peligroso, no solo feo: la tarjeta ofrecía "Ajustar a lo contado",
    y apretarlo habría borrado los cajones que de verdad aparecieron, para
    después ponerse en verde tapando el daño.
    """
    coincidio = _conteo_cotejo(cantidad=657, stock_sistema=657,
                               ajustes_posteriores=657, stock_actual=1314)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[coincidio]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert 'class="tarjeta con-diferencia"' not in respuesta.text
    assert "Ajustar a lo contado" not in respuesta.text
    assert "sin explicar" not in respuesta.text


def test_ver_cotejo_un_ajuste_en_la_direccion_CONTRARIA_no_agranda_el_faltante():
    # Faltan 5 y el ajuste posterior suma 100 por otro motivo: eso no tapa
    # nada, así que siguen faltando 5 — no 105.
    contrario = _conteo_cotejo(ajustes_posteriores=100, stock_actual=140)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[contrario]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert 'class="tarjeta con-diferencia"' in respuesta.text
    assert "sin explicar" not in respuesta.text  # sigue siendo la diferencia entera
    assert "Ajustar a lo contado" in respuesta.text


def test_ver_cotejo_un_ajuste_parcial_deja_el_resto_a_la_vista():
    # Ajustó 3 de los 5: quedan 2 sin explicar, y eso sigue siendo tarea.
    parcial = _conteo_cotejo(ajustes_posteriores=-3, stock_actual=37)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[parcial]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert 'class="tarjeta con-diferencia"' in respuesta.text
    assert "Ajustar a lo contado" in respuesta.text
    assert "ya se ajustó -3" in respuesta.text
    assert "queda sin explicar <strong>-2</strong>" in respuesta.text


def test_ver_cotejo_los_movimientos_normales_posteriores_no_inventan_una_alarma():
    """Si después del conteo entraron cajones, el stock ya no coincide con lo
    contado — y eso NO es un problema. Por eso se mide contra los ajustes
    posteriores y no contra el stock de hoy: medir así pondría en rojo algo
    que está bien, pidiendo un ajuste que sería incorrecto.
    """
    con_movimientos = _conteo_cotejo(cantidad=8, stock_sistema=8, stock_actual=25,
                                     ajustes_posteriores=0)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[con_movimientos]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert 'class="tarjeta con-diferencia"' not in respuesta.text
    assert "Ajustar a lo contado" not in respuesta.text


def test_ver_cotejo_un_par_cerrado_va_en_gris_al_final_y_sin_boton():
    # Proveedor dado de baja y saldo cero: no hay nada que cotejar. En gris
    # y al final — pero NO escondido: si desapareciera, no habría forma de
    # ver que existió ni de notar que se cerró con una diferencia.
    cerrado = _conteo_cotejo(id=9, proveedor_activo=False, stock_actual=0,
                             cantidad=3, stock_sistema=5,
                             proveedor_nombre="Gómez", tipo_nombre="cajón viejo")
    with patch("app.main.listar_ultimos_conteos_vacios",
               return_value=[cerrado, _conteo_cotejo()]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert "Cerrado" in respuesta.text
    assert 'class="tarjeta cerrada"' in respuesta.text
    assert "ya no hay nada que ajustar" in respuesta.text
    assert "Gómez" in respuesta.text
    # El cerrado va DESPUÉS del que sí tiene algo para hacer.
    assert respuesta.text.index("Saturno") < respuesta.text.index("Gómez")
    # Y sin botón: solo el de la tarjeta viva.
    assert respuesta.text.count("Ajustar a lo contado") == 1


def test_ver_cotejo_un_par_de_baja_con_saldo_pide_cerrar_la_cuenta():
    # El estado a medio camino que la validación de la baja ya no deja
    # crear. Los que quedaron de antes se muestran, y lo único que ofrecen
    # es cerrar: ajustarlo "a lo contado" sería devolverle vida a algo
    # que ya no se puede mover.
    a_medias = _conteo_cotejo(id=8, tipo_activo=False, stock_actual=4,
                              cantidad=10, stock_sistema=10)
    with patch("app.main.listar_ultimos_conteos_vacios", return_value=[a_medias]):
        respuesta = cliente.get("/puesto/envases/cotejo")

    assert respuesta.status_code == 200
    assert "Dado de baja con saldo" in respuesta.text
    assert "le quedaron <strong>4</strong> en el sistema" in respuesta.text
    assert "Cerrar la cuenta en cero" in respuesta.text
    assert "Ajustar a lo contado" not in respuesta.text
    assert "contado=0" in respuesta.text and "cierre=1" in respuesta.text


def test_ver_pendientes_de_pago_ofrece_los_tres_cierres():
    pendientes = [
        {"id": 5, "cantidad": 12, "creado_en": datetime(2026, 8, 19, 10, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Juan Pérez", "proveedor_nombre": "Saturno",
         "tipo_nombre": "cajón plástico negro", "monto_unitario": 500},
    ]
    with (
        patch("app.main.listar_senas_pendientes", return_value=pendientes),
        patch("app.main.listar_senas_resueltas", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/pendientes")

    assert respuesta.status_code == 200
    assert "Juan Pérez" in respuesta.text
    assert 'action="/puesto/envases/pendientes/5/pagar"' in respuesta.text
    assert 'action="/puesto/envases/pendientes/5/vale"' in respuesta.text
    assert 'action="/puesto/envases/pendientes/5/anular-sena"' in respuesta.text
    # Los tres con confirmación en dos toques.
    assert "Sí, pagó" in respuesta.text
    assert "Sí, se hizo vale" in respuesta.text
    assert "Borrar Operación" in respuesta.text
    assert "Sí, borrar la operación" in respuesta.text
    # "Borrar Operación" va separado y chico (clase propia), no como tercer botón
    # grande al lado de "Pagó" — la pantalla se usa apurada.
    assert 'class="boton-anular-sena"' in respuesta.text
    assert 'class="fila-anular"' in respuesta.text


def test_ver_pendientes_el_historial_distingue_los_tres_cierres():
    resueltas = [
        {"id": 4, "cantidad": 3, "creado_en": datetime(2026, 8, 18, 10, 0, tzinfo=timezone.utc),
         "cierre": "pagada", "cerrada_el": datetime(2026, 8, 18, 12, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Marta", "proveedor_nombre": "Saturno",
         "tipo_nombre": "torito", "monto_unitario": 500},
        {"id": 3, "cantidad": 5, "creado_en": datetime(2026, 8, 17, 10, 0, tzinfo=timezone.utc),
         "cierre": "vale", "cerrada_el": datetime(2026, 8, 17, 12, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Juan Pérez", "proveedor_nombre": "Saturno",
         "tipo_nombre": "torito", "monto_unitario": None},
        {"id": 2, "cantidad": 7, "creado_en": datetime(2026, 8, 16, 10, 0, tzinfo=timezone.utc),
         "cierre": "anulada", "cerrada_el": datetime(2026, 8, 16, 12, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Pedro", "proveedor_nombre": "Don Pepe",
         "tipo_nombre": "cajón madera", "monto_unitario": 300},
    ]
    with (
        patch("app.main.listar_senas_pendientes", return_value=[]),
        patch("app.main.listar_senas_resueltas", return_value=resueltas),
    ):
        respuesta = cliente.get("/puesto/envases/pendientes")

    assert respuesta.status_code == 200
    assert "<details>" in respuesta.text
    assert 'class="etiqueta-cierre cierre-pagada">Pagada<' in respuesta.text
    assert 'class="etiqueta-cierre cierre-vale">Vale<' in respuesta.text
    assert 'class="etiqueta-cierre cierre-anulada">Anulada<' in respuesta.text


def test_los_tres_cierres_de_sena_llaman_a_cerrar_sena_y_redirigen():
    for url, cierre in (
        ("/puesto/envases/pendientes/5/pagar", "pagada"),
        ("/puesto/envases/pendientes/5/vale", "vale"),
        ("/puesto/envases/pendientes/5/anular-sena", "anulada"),
    ):
        with patch("app.main.cerrar_sena") as mock_cerrar:
            respuesta = cliente.post(url, follow_redirects=False)

        assert respuesta.status_code == 303, url
        assert respuesta.headers["location"] == "/puesto/envases/pendientes"
        mock_cerrar.assert_called_once_with(5, cierre)


def test_ver_movimientos_usa_los_ultimos_7_dias_por_defecto():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=[]) as mock_recibidos,
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=[]) as mock_devueltos,
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=[]) as mock_ajustes,
    ):
        respuesta = cliente.get("/puesto/envases/movimientos")

    assert respuesta.status_code == 200
    mock_recibidos.assert_called_once_with(date(2026, 8, 12), date(2026, 8, 19))
    mock_devueltos.assert_called_once_with(date(2026, 8, 12), date(2026, 8, 19))
    mock_ajustes.assert_called_once_with(date(2026, 8, 12), date(2026, 8, 19))


def test_ver_movimientos_permite_anular_de_cualquier_fecha_conservando_filtros():
    recibidos = [
        {"id": 9, "cantidad": 12, "creado_en": datetime(2026, 8, 10, 10, 0, tzinfo=timezone.utc),
         "anulado_el": None, "cliente_nombre": "Juan Pérez", "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=recibidos),
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=[]),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos?fecha_desde=2026-08-09&fecha_hasta=2026-08-11")

    assert respuesta.status_code == 200
    assert 'action="/puesto/envases/movimientos/recibidos/9/anular"' in respuesta.text
    # Los filtros viajan ocultos para volver al mismo rango tras anular.
    assert 'name="fecha_desde" value="2026-08-09"' in respuesta.text


def test_anular_desde_movimientos_redirige_al_mismo_rango():
    with patch("app.main.anular_vacio_recibido") as mock_anular:
        respuesta = cliente.post(
            "/puesto/envases/movimientos/recibidos/9/anular",
            data={"fecha_desde": "2026-08-09", "fecha_hasta": "2026-08-11"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/puesto/envases/movimientos?")
    assert "fecha_desde=2026-08-09" in location
    mock_anular.assert_called_once_with(9)


MOVIMIENTOS_EXPORT_RECIBIDOS = [
    {"id": 9, "cantidad": 12, "creado_en": datetime(2026, 8, 10, 10, 0, tzinfo=timezone.utc),
     "anulado_el": None, "cliente_nombre": "Juan Pérez", "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
]
MOVIMIENTOS_EXPORT_DEVUELTOS = [
    {"id": 4, "cantidad": 30, "stock_sistema": 40, "creado_en": datetime(2026, 8, 10, 12, 0, tzinfo=timezone.utc),
     "anulado_el": datetime(2026, 8, 11, 9, 0, tzinfo=timezone.utc), "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
]
MOVIMIENTOS_EXPORT_AJUSTES = [
    {"id": 30, "cantidad": -5, "motivo": "Se rompieron dos", "stock_sistema": 40,
     "creado_en": datetime(2026, 8, 10, 15, 0, tzinfo=timezone.utc),
     "anulado_el": None, "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
]


def test_ver_movimientos_ofrece_exportar_con_los_mismos_filtros():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=MOVIMIENTOS_EXPORT_RECIBIDOS),
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=[]),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos?fecha_desde=2026-08-09&fecha_hasta=2026-08-11")

    assert respuesta.status_code == 200
    assert "/puesto/envases/movimientos/exportar-pdf?fecha_desde=2026-08-09&fecha_hasta=2026-08-11" in respuesta.text
    assert "/puesto/envases/movimientos/exportar-excel?fecha_desde=2026-08-09&fecha_hasta=2026-08-11" in respuesta.text


def test_exportar_movimientos_vacios_pdf_trae_las_tres_secciones_y_los_motivos():
    with (
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=MOVIMIENTOS_EXPORT_RECIBIDOS) as mock_r,
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=MOVIMIENTOS_EXPORT_DEVUELTOS),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=MOVIMIENTOS_EXPORT_AJUSTES),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos/exportar-pdf?fecha_desde=2026-08-09&fecha_hasta=2026-08-11")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert "Movimientos_Vacios_2026-08-09_a_2026-08-11" in respuesta.headers["content-disposition"]
    # El export usa los mismos filtros que la pantalla, sin tope.
    mock_r.assert_called_once_with(date(2026, 8, 9), date(2026, 8, 11))
    texto = _texto_del_pdf_de_respuesta(respuesta.content)
    # Las tres secciones, en el mismo orden que la pantalla.
    assert texto.index("Salidas") < texto.index("Entradas") < texto.index("Ajustes")
    assert "Juan Pérez" in texto  # quién trajo, en Entradas
    assert "Se rompieron dos" in texto  # el motivo del ajuste viaja
    assert "ANULADO" in texto  # los anulados van marcados, no escondidos


def test_exportar_movimientos_vacios_excel_devuelve_archivo_adjunto():
    with (
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=MOVIMIENTOS_EXPORT_RECIBIDOS),
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=MOVIMIENTOS_EXPORT_DEVUELTOS),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=MOVIMIENTOS_EXPORT_AJUSTES),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos/exportar-excel?fecha_desde=2026-08-09&fecha_hasta=2026-08-11")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Movimientos_Vacios_2026-08-09_a_2026-08-11" in respuesta.headers["content-disposition"]
    assert respuesta.content.startswith(b"PK")  # xlsx es un zip


# --- Vacíos: ajustes de stock (cajera) ---


AJUSTES_VACIOS_DE_PRUEBA = [
    {"id": 30, "cantidad": -5, "motivo": "Se rompieron dos y tres desaparecieron",
     "stock_sistema": 40, "creado_en": datetime(2026, 8, 19, 15, 0, tzinfo=timezone.utc),
     "anulado_el": None, "proveedor_nombre": "Saturno", "tipo_nombre": "cajón plástico negro"},
    {"id": 31, "cantidad": 3, "motivo": "Conteo inicial mal cargado",
     "stock_sistema": 10, "creado_en": datetime(2026, 8, 18, 11, 0, tzinfo=timezone.utc),
     "anulado_el": datetime(2026, 8, 19, 9, 0, tzinfo=timezone.utc),
     "proveedor_nombre": "Don Pepe", "tipo_nombre": "cajón madera"},
]


def test_ver_movimientos_pone_las_salidas_antes_que_las_entradas():
    # Las salidas son pocas (un camión cada tanto) y abajo se perderían
    # entre montones de entradas: van primero.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=[]),
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=[]),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos")

    assert respuesta.status_code == 200
    assert respuesta.text.index("Salidas (devueltos al proveedor)") < respuesta.text.index(
        "Entradas (recibidos de clientes)"
    )


def test_ver_movimientos_muestra_los_ajustes_claramente_distintos():
    # Pedido explícito: los ajustes (correcciones) no se pueden confundir con
    # entradas/salidas reales al auditar — tarjeta y pill propias.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 19)),
        patch("app.main.listar_vacios_recibidos_por_rango", return_value=[]),
        patch("app.main.listar_vacios_devueltos_por_rango", return_value=[]),
        patch("app.main.listar_ajustes_vacios_por_rango", return_value=AJUSTES_VACIOS_DE_PRUEBA),
    ):
        respuesta = cliente.get("/puesto/envases/movimientos")

    assert respuesta.status_code == 200
    assert "tarjeta-ajustes" in respuesta.text
    assert respuesta.text.count('class="pill-ajuste"') == 2
    # Cantidad siempre con signo y el motivo a la vista.
    assert "-5" in respuesta.text
    assert "+3" in respuesta.text
    assert "Se rompieron dos y tres desaparecieron" in respuesta.text
    assert "el sistema decía 40" in respuesta.text
    # Solo el activo ofrece Anular (con los filtros ocultos); el anulado queda marcado.
    assert 'action="/puesto/envases/movimientos/ajustes/30/anular"' in respuesta.text
    assert 'action="/puesto/envases/movimientos/ajustes/31/anular"' not in respuesta.text
    assert "Anulado el" in respuesta.text


def test_anular_ajuste_desde_movimientos_redirige_al_mismo_rango():
    with patch("app.main.anular_ajuste_vacios") as mock_anular:
        respuesta = cliente.post(
            "/puesto/envases/movimientos/ajustes/30/anular",
            data={"fecha_desde": "2026-08-09", "fecha_hasta": "2026-08-11"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/puesto/envases/movimientos?")
    assert "fecha_desde=2026-08-09" in location
    mock_anular.assert_called_once_with(30)


def test_ver_ajustar_stock_en_blanco():
    with patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/ajustar")

    assert respuesta.status_code == 200
    assert 'action="/puesto/envases/ajustar"' in respuesta.text
    # La cantidad admite negativos: number sin min.
    assert 'name="cantidad"' in respuesta.text
    assert 'min=' not in respuesta.text.split('name="cantidad"')[1].split(">")[0]
    # El motivo es obligatorio también en el HTML.
    assert 'id="motivo" name="motivo" required' in respuesta.text
    # Sin precarga no hay cartel amarillo del conteo.
    assert "aviso-conteo" not in respuesta.text.replace(".aviso-conteo", "")


def test_ver_ajustar_stock_precarga_contra_el_stock_actual():
    # Contado 35, stock actual 40 (igual a la foto del conteo): precarga -5
    # y sin aviso, porque nada se movió desde el conteo.
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.stock_vacios_de_tipo", return_value=40) as mock_stock,
    ):
        respuesta = cliente.get(
            "/puesto/envases/ajustar?proveedor_id=200&tipo_envase_id=1"
            "&contado=35&stock_conteo=40&fecha_conteo=2026-08-19"
        )

    assert respuesta.status_code == 200
    mock_stock.assert_called_once_with(200, 1)
    assert 'value="-5"' in respuesta.text
    assert "Ajuste a lo contado: conteo del 2026-08-19 (35 contados)" in respuesta.text
    assert "Ojo: el conteo fue del" not in respuesta.text


def test_ver_ajustar_stock_avisa_si_el_stock_cambio_desde_el_conteo():
    # El caso del usuario: conteo viejo (35 contados, sistema decía 40),
    # pero desde entonces hubo movimientos y el stock actual es 52. El
    # ajuste sugerido es 35 − 52 = -17, NO la diferencia del Cotejo (-5),
    # y la pantalla lo explica con todos los números.
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.stock_vacios_de_tipo", return_value=52),
    ):
        respuesta = cliente.get(
            "/puesto/envases/ajustar?proveedor_id=200&tipo_envase_id=1"
            "&contado=35&stock_conteo=40&fecha_conteo=2026-08-19"
        )

    assert respuesta.status_code == 200
    assert 'value="-17"' in respuesta.text
    assert "Ojo: el conteo fue del 2026-08-19 con 35 contados" in respuesta.text
    assert "el sistema decía 40" in respuesta.text
    assert "el stock actual es 52" in respuesta.text
    assert "-17 (no la diferencia que viste en el Cotejo)" in respuesta.text


def test_guardar_ajuste_negativo_con_motivo_redirige_con_el_stock_resultante():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_ajuste_vacios", return_value=15) as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/ajustar",
            data={"proveedor_id": "200", "tipo_envase_id": "1",
                  "cantidad": "-5", "motivo": "  Se rompieron   dos "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    location = respuesta.headers["location"]
    assert location.startswith("/puesto/envases/ajustar?")
    assert "El+stock+qued%C3%B3+en+15" in location
    mock_crear.assert_called_once_with(200, 1, -5, "Se rompieron dos")


def test_guardar_ajuste_sin_motivo_no_guarda():
    # La regla dura del usuario: sin motivo no se guarda, ni con espacios.
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_ajuste_vacios") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/ajustar",
            data={"proveedor_id": "200", "tipo_envase_id": "1",
                  "cantidad": "-5", "motivo": "   "},
        )

    assert respuesta.status_code == 400
    assert "El motivo es obligatorio: sin motivo no se guarda el ajuste." in respuesta.text
    mock_crear.assert_not_called()


def test_guardar_ajuste_de_cero_rechazado():
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_ajuste_vacios") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/ajustar",
            data={"proveedor_id": "200", "tipo_envase_id": "1",
                  "cantidad": "0", "motivo": "porque sí"},
        )

    assert respuesta.status_code == 400
    assert "Un ajuste de 0 no ajusta nada." in respuesta.text
    mock_crear.assert_not_called()


def test_guardar_ajuste_con_par_proveedor_tipo_invalido_da_error():
    # El tipo 3 es de Don Pepe (201), no de Saturno (200): la lista es cerrada.
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
        patch("app.main.crear_ajuste_vacios") as mock_crear,
    ):
        respuesta = cliente.post(
            "/puesto/envases/ajustar",
            data={"proveedor_id": "200", "tipo_envase_id": "3",
                  "cantidad": "4", "motivo": "prueba"},
        )

    assert respuesta.status_code == 400
    assert "Elegí un proveedor y un tipo de envase válidos." in respuesta.text
    mock_crear.assert_not_called()


def test_ver_stock_vacios_muestra_los_ajustes_en_el_desglose():
    filas = [
        {"proveedor_id": 200, "proveedor_nombre": "Saturno",
         "tipo_envase_id": 1, "tipo_nombre": "cajón plástico negro",
         "recibidos": 50, "devueltos": 30, "ajustes": -5, "stock": 15},
    ]
    with patch("app.main.stock_vacios", return_value=filas):
        respuesta = cliente.get("/puesto/envases/stock")

    assert respuesta.status_code == 200
    assert "50 recibidos − 30 devueltos − 5 ajustes" in respuesta.text
    assert ">15</span>" in respuesta.text


def test_hub_envases_puesto_tiene_el_boton_de_ajustar_stock():
    respuesta = cliente.get("/puesto/envases")

    assert respuesta.status_code == 200
    assert 'href="/puesto/envases/ajustar"' in respuesta.text
    assert "Ajustar Stock" in respuesta.text


# --- Clave de la zona de control del Puesto ---


def _con_clave_control(valor="1234"):
    """La clave se lee SIEMPRE por _clave_control_puesto (un solo lugar): patchear ahí es patchear el origen."""
    return patch("app.main._clave_control_puesto", return_value=valor)


def _cliente_destrabado(clave="1234"):
    """Un TestClient propio (la cookie no puede contaminar al cliente compartido) ya con la clave puesta."""
    cliente_propio = TestClient(app)
    respuesta = cliente_propio.post(
        "/puesto/envases/clave", data={"clave": clave, "volver": "/puesto/envases"}, follow_redirects=False
    )
    assert respuesta.status_code == 303
    return cliente_propio


def test_sin_clave_configurada_no_hay_puerta():
    # El comportamiento de siempre: sin la variable, todo abierto como hoy.
    filas_stock = [
        {"proveedor_id": 200, "proveedor_nombre": "Saturno",
         "tipo_envase_id": 1, "tipo_nombre": "torito",
         "recibidos": 5, "devueltos": 0, "ajustes": 0, "stock": 5},
    ]
    with patch("app.main.stock_vacios", return_value=filas_stock):
        respuesta = cliente.get("/puesto/envases/stock")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    # Sin clave tampoco hay botón Bloquear (no hay nada que bloquear).
    assert "Bloquear" not in respuesta.text


def test_con_clave_las_pantallas_de_control_piden_clave():
    with _con_clave_control():
        for url in (
            "/puesto/envases/stock",
            "/puesto/envases/cotejo",
            "/puesto/envases/movimientos",
            "/puesto/envases/ajustar",
            "/puesto/envases/proveedores",
            "/puesto/envases/tipos",
            "/puesto/envases/clientes",
        ):
            respuesta = cliente.get(url)
            assert respuesta.status_code == 401, url
            assert 'action="/puesto/envases/clave"' in respuesta.text, url
            assert f'name="volver" value="{url}"' in respuesta.text, url


def test_cambiar_el_valor_de_la_sena_sin_destrabar_NO_toca_la_base():
    """El POST, que es donde de verdad se mueve plata.

    Que la pantalla pida clave no alcanza: el que importa es el POST, que
    se puede mandar sin pasar por la pantalla. Cuánto vale la seña define
    cuánto se le paga a cada cliente.
    """
    with (
        _con_clave_control(),
        patch("app.main.cargar_valor_sena") as mock_cargar,
        patch("app.main.renombrar_tipo_envase_puesto") as mock_renombrar,
    ):
        respuesta = cliente.post(
            "/puesto/envases/tipos/1/editar",
            data={"nombre": "cajón", "monto": "500", "vigente_desde": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_cargar.assert_not_called()
    mock_renombrar.assert_not_called()


def test_la_pantalla_de_clave_no_deja_que_el_navegador_la_recuerde():
    # Si el celular guarda o autocompleta la clave, la traba desaparece:
    # el form va con autocomplete off y el campo como one-time-code.
    with _con_clave_control():
        respuesta = cliente.get("/puesto/envases/stock")

    assert respuesta.status_code == 401
    assert 'autocomplete="off"' in respuesta.text
    assert 'autocomplete="one-time-code"' in respuesta.text
    assert 'type="password"' in respuesta.text


def test_el_volver_de_la_puerta_conserva_la_query():
    # Venir del Cotejo con precarga y chocar la puerta no puede perder los
    # parámetros: tras la clave se llega a la MISMA pantalla precargada.
    with _con_clave_control():
        respuesta = cliente.get("/puesto/envases/ajustar?proveedor_id=200&tipo_envase_id=1&contado=35")

    assert respuesta.status_code == 401
    assert 'value="/puesto/envases/ajustar?proveedor_id=200&amp;tipo_envase_id=1&amp;contado=35"' in respuesta.text


def test_clave_correcta_desbloquea_toda_la_zona_una_sola_vez():
    with _con_clave_control():
        cliente_propio = TestClient(app)
        respuesta = cliente_propio.post(
            "/puesto/envases/clave",
            data={"clave": "1234", "volver": "/puesto/envases/cotejo"},
            follow_redirects=False,
        )
        assert respuesta.status_code == 303
        assert respuesta.headers["location"] == "/puesto/envases/cotejo"
        assert "acceso_control_puesto" in respuesta.headers.get("set-cookie", "")
        # La clave NUNCA viaja en la cookie: solo la firma.
        assert "1234" not in respuesta.headers.get("set-cookie", "")

        # Con esa única clave, varias pantallas distintas sin repetirla.
        with patch("app.main.listar_ultimos_conteos_vacios", return_value=[]):
            assert cliente_propio.get("/puesto/envases/cotejo").status_code == 200
        with patch("app.main.stock_vacios", return_value=[]):
            respuesta_stock = cliente_propio.get("/puesto/envases/stock")
        assert respuesta_stock.status_code == 200
        # Destrabado, aparece el botón para volver a bloquear.
        assert 'action="/puesto/envases/bloquear"' in respuesta_stock.text


def test_clave_incorrecta_no_deja_cookie_y_avisa():
    with _con_clave_control():
        cliente_propio = TestClient(app)
        respuesta = cliente_propio.post(
            "/puesto/envases/clave", data={"clave": "9999", "volver": "/puesto/envases/stock"}
        )

    assert respuesta.status_code == 401
    assert "Clave incorrecta." in respuesta.text
    assert "acceso_control_puesto" not in respuesta.headers.get("set-cookie", "")


def test_el_volver_no_puede_salir_de_envases_puesto():
    # La puerta redirige solo adentro de la zona: un volver ajeno cae al hub.
    with _con_clave_control():
        respuesta = TestClient(app).post(
            "/puesto/envases/clave",
            data={"clave": "1234", "volver": "https://otro-lado.com/x"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases"


def test_los_post_protegidos_sin_clave_no_ejecutan_nada():
    with _con_clave_control(), patch("app.main.anular_ajuste_vacios") as mock_anular, patch(
        "app.main.crear_ajuste_vacios"
    ) as mock_crear, patch("app.main.obtener_o_crear_proveedor_puesto") as mock_proveedor:
        r1 = cliente.post(
            "/puesto/envases/movimientos/ajustes/30/anular",
            data={"fecha_desde": "2026-08-09", "fecha_hasta": "2026-08-11"},
            follow_redirects=False,
        )
        r2 = cliente.post(
            "/puesto/envases/ajustar",
            data={"proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "-5", "motivo": "x"},
            follow_redirects=False,
        )
        r3 = cliente.post("/puesto/envases/proveedores/nuevo", data={"nombre": "Colado"}, follow_redirects=False)

    assert r1.status_code == 303 and r2.status_code == 303 and r3.status_code == 303
    mock_anular.assert_not_called()
    mock_crear.assert_not_called()
    mock_proveedor.assert_not_called()


def test_bloquear_borra_la_cookie_y_vuelve_a_pedir_clave():
    with _con_clave_control():
        cliente_propio = _cliente_destrabado()
        with patch("app.main.stock_vacios", return_value=[]):
            assert cliente_propio.get("/puesto/envases/stock").status_code == 200

        respuesta = cliente_propio.post("/puesto/envases/bloquear", follow_redirects=False)
        assert respuesta.status_code == 303
        assert respuesta.headers["location"] == "/puesto/envases"

        # Bloqueado: la misma pantalla pide clave de nuevo.
        assert cliente_propio.get("/puesto/envases/stock").status_code == 401


def test_pendientes_y_pantallas_del_empleado_quedan_sin_clave():
    # Pendientes de Pago se usa todo el día con el cliente adelante, y las
    # pantallas del empleado son la operación: nada de clave ahí.
    with _con_clave_control():
        with (
            patch("app.main.listar_senas_pendientes", return_value=[]),
            patch("app.main.listar_senas_resueltas", return_value=[]),
        ):
            assert cliente.get("/puesto/envases/pendientes").status_code == 200
        assert cliente.get("/puesto/envases/vacios").status_code == 200
        with (
            patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA),
            patch("app.main.listar_clientes_puesto", return_value=CLIENTES_PUESTO_DE_PRUEBA),
            patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
        ):
            assert cliente.get("/puesto/envases/vacios/recibir").status_code == 200


def test_el_alta_automatica_de_cliente_al_recibir_no_pide_clave():
    # El empleado tipea un nombre nuevo en Recibir y el cliente se crea
    # solo: ese camino es operación, NO catálogo — sin puerta.
    with _con_clave_control(), (
        patch("app.main.listar_tipos_envase_puesto", return_value=TIPOS_ENVASE_PUESTO_DE_PRUEBA)
    ), patch("app.main.obtener_o_crear_cliente_puesto", return_value=10) as mock_cliente, patch(
        "app.main.crear_vacio_recibido"
    ) as mock_crear:
        respuesta = cliente.post(
            "/puesto/envases/vacios/recibir",
            data={"cliente_nombre": "Doña Rosa", "proveedor_id": "200", "tipo_envase_id": "1", "cantidad": "4"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_cliente.assert_called_once()
    mock_crear.assert_called_once()


def test_ver_clientes_puesto_lista_con_alta_y_baja():
    with patch("app.main.listar_clientes_puesto", return_value=CLIENTES_PUESTO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/clientes")

    assert respuesta.status_code == 200
    assert "Juan Pérez" in respuesta.text
    assert 'action="/puesto/envases/clientes/nuevo"' in respuesta.text
    assert 'action="/puesto/envases/clientes/10/baja"' in respuesta.text


def test_crear_cliente_puesto_desde_la_pantalla_normaliza():
    with patch("app.main.obtener_o_crear_cliente_puesto", return_value=12) as mock_crear:
        respuesta = cliente.post(
            "/puesto/envases/clientes/nuevo",
            data={"nombre": "  DOÑA   Rosa "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("DOÑA Rosa", "dona rosa")


def test_dar_de_baja_cliente_puesto_redirige():
    with patch("app.main.desactivar_cliente_puesto") as mock_baja:
        respuesta = cliente.post("/puesto/envases/clientes/10/baja", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases/clientes"
    mock_baja.assert_called_once_with(10)


def test_ver_proveedores_puesto_lista_con_alta_y_baja():
    with patch("app.main.listar_proveedores_puesto", return_value=PROVEEDORES_PUESTO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/proveedores")

    assert respuesta.status_code == 200
    assert "Saturno" in respuesta.text
    assert 'action="/puesto/envases/proveedores/nuevo"' in respuesta.text
    assert 'action="/puesto/envases/proveedores/200/baja"' in respuesta.text
    # Ni rastro de los proveedores de Compras (circuitos separados).
    assert "codigo_puesto" not in respuesta.text


def test_crear_proveedor_puesto_normaliza_y_redirige():
    with patch("app.main.obtener_o_crear_proveedor_puesto", return_value=7) as mock_crear:
        respuesta = cliente.post(
            "/puesto/envases/proveedores/nuevo",
            data={"nombre": "  EL   Cajónero "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("EL Cajónero", "el cajonero")


def test_crear_proveedor_puesto_sin_nombre_da_error():
    with (
        patch("app.main.listar_proveedores_puesto", return_value=[]),
        patch("app.main.obtener_o_crear_proveedor_puesto") as mock_crear,
    ):
        respuesta = cliente.post("/puesto/envases/proveedores/nuevo", data={"nombre": "   "})

    assert respuesta.status_code == 400
    assert "El nombre del proveedor es obligatorio." in respuesta.text
    mock_crear.assert_not_called()


def test_dar_de_baja_proveedor_puesto_redirige():
    with patch("app.main.desactivar_proveedor_puesto") as mock_baja:
        respuesta = cliente.post("/puesto/envases/proveedores/200/baja", follow_redirects=False)

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/puesto/envases/proveedores"
    mock_baja.assert_called_once_with(200)


def test_dar_de_baja_un_proveedor_con_saldo_se_niega_nombrando_sus_tipos():
    with (
        patch("app.main.desactivar_proveedor_puesto",
              side_effect=ValueError("Gómez todavía tiene stock (cajón madera: 12, cajón plástico: -3). Devolvelos o ajustá a cero antes de darlo de baja.")),
        patch("app.main.listar_proveedores_puesto", return_value=[]),
    ):
        respuesta = cliente.post("/puesto/envases/proveedores/200/baja", follow_redirects=False)

    assert respuesta.status_code == 400
    assert "cajón madera: 12" in respuesta.text
    assert "cajón plástico: -3" in respuesta.text


def test_ver_proveedores_puesto_ofrece_editar_el_nombre():
    with patch("app.main.listar_proveedores_puesto", return_value=PROVEEDORES_PUESTO_DE_PRUEBA):
        respuesta = cliente.get("/puesto/envases/proveedores")

    assert respuesta.status_code == 200
    assert 'action="/puesto/envases/proveedores/200/renombrar"' in respuesta.text
    assert 'value="Saturno"' in respuesta.text


def test_renombrar_proveedor_puesto_manda_TAMBIEN_el_normalizado():
    # Si la ruta mandara solo el nombre, el normalizado quedaría con el
    # viejo adentro y la próxima alta duplicaría el proveedor.
    with patch("app.main.renombrar_proveedor_puesto") as mock_renombrar:
        respuesta = cliente.post(
            "/puesto/envases/proveedores/200/renombrar",
            data={"nombre": "  EL   Cajónero "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_renombrar.assert_called_once_with(200, "EL Cajónero", "el cajonero")


def test_renombrar_proveedor_puesto_sin_nombre_da_error():
    with (
        patch("app.main.listar_proveedores_puesto", return_value=[]),
        patch("app.main.renombrar_proveedor_puesto") as mock_renombrar,
    ):
        respuesta = cliente.post("/puesto/envases/proveedores/200/renombrar", data={"nombre": "   "})

    assert respuesta.status_code == 400
    assert "El nombre del proveedor es obligatorio." in respuesta.text
    mock_renombrar.assert_not_called()


def test_renombrar_proveedor_a_un_nombre_repetido_da_400_con_el_nombre_nunca_500():
    with (
        patch("app.main.renombrar_proveedor_puesto",
              side_effect=ValueError("Ya existe un proveedor llamado 'Gómez'.")),
        patch("app.main.listar_proveedores_puesto", return_value=[]),
    ):
        respuesta = cliente.post("/puesto/envases/proveedores/200/renombrar", data={"nombre": "GOMEZ"})

    assert respuesta.status_code == 400
    assert "Ya existe un proveedor llamado" in respuesta.text
    assert "Gómez" in respuesta.text


def test_pendientes_muestra_el_monto_en_su_renglon_y_el_total_por_cantidad():
    pendientes = [
        {"id": 5, "cantidad": 12, "creado_en": datetime(2026, 8, 19, 10, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Juan Pérez", "proveedor_nombre": "Saturno",
         "tipo_nombre": "cajón plástico negro", "monto_unitario": 500},
        {"id": 6, "cantidad": 3, "creado_en": datetime(2026, 8, 19, 11, 0, tzinfo=timezone.utc),
         "cliente_nombre": "Marta", "proveedor_nombre": "Saturno",
         "tipo_nombre": "torito", "monto_unitario": None},
    ]
    with (
        patch("app.main.listar_senas_pendientes", return_value=pendientes),
        patch("app.main.listar_senas_resueltas", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/pendientes")

    assert respuesta.status_code == 200
    # El total grande (500 × 12) y el unitario chico al lado.
    assert "$6.000" in respuesta.text
    assert 'class="monto-total"' in respuesta.text
    assert 'class="monto-unitario"' in respuesta.text
    # Y el que no tiene valor lo dice con palabras, sin ningún número.
    assert "Sin valor de seña cargado" in respuesta.text
    assert "$0" not in respuesta.text.split("</style>")[-1]


def test_hub_envases_puesto_tiene_proveedores_del_puesto():
    respuesta = cliente.get("/puesto/envases")

    assert respuesta.status_code == 200
    assert 'href="/puesto/envases/proveedores"' in respuesta.text


# --- Pedidos de clientes (etapa 1: cargar pegando el texto del mail) ---


PEDIDO_VIGENTE_DE_PRUEBA = {
    "id": 50, "cliente_id": 1, "fecha_operacion": date(2026, 8, 21), "origen": "texto",
    "recibido_el": None, "reemplaza_a_pedido_id": None, "creado_en": datetime(2026, 8, 21, 12, 10),
    "reemplazado_creado_en": None, "armado_cerrado_el": None,
}

SUCURSALES_PEDIDO_DE_PRUEBA = [
    {"id": 1, "sucursal": "VL", "orden_compra": "1257673", "total_bultos_declarado": 235.0},
    {"id": 2, "sucursal": "BZ", "orden_compra": "1257642", "total_bultos_declarado": 40.0},
]

RENGLONES_PEDIDO_DE_PRUEBA = [
    {"id": 10, "sucursal": "VL", "articulo_id": None, "articulo_nombre": None, "ficha_id": None, "nombre_venta": None,
     "texto_codigo": "99999", "texto_descripcion": "RADICHETA", "cantidad": 5.0, "armado_el": None, "cantidad_armada": None,
     "kilos_enviados": None, "anulado_el": None},
    {"id": 11, "sucursal": "VL", "articulo_id": 1, "articulo_nombre": "Banana", "ficha_id": 901, "nombre_venta": "Banana",
     "texto_codigo": "90101", "texto_descripcion": "BANANA", "cantidad": 225.0, "armado_el": None, "cantidad_armada": None,
     "kilos_enviados": None, "anulado_el": None},
    {"id": 12, "sucursal": "BZ", "articulo_id": 2, "articulo_nombre": "Batata", "ficha_id": 902, "nombre_venta": "Batata",
     "texto_codigo": "90102", "texto_descripcion": "BATATA", "cantidad": 40.0, "armado_el": None, "cantidad_armada": None,
     "kilos_enviados": None, "anulado_el": None},
]

FICHAS_PEDIDO_DE_PRUEBA = [
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Banana", "articulo_grupo": "fruta",
     "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": "BANANA", "codigo_cliente": "90101"},
    {"id": 902, "articulo_id": 2, "articulo_nombre": "Batata", "articulo_grupo": "hortaliza",
     "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "90102"},
    {"id": 903, "articulo_id": 3, "articulo_nombre": "Zapallo", "articulo_grupo": "hortaliza",
     "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": None, "codigo_cliente": None},
]


def test_ver_deposito_muestra_los_accesos_a_pedido():
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert 'href="/deposito/pedido"' in respuesta.text
    assert 'href="/deposito/pedido/armar"' in respuesta.text
    # Buscar Pedidos se mudó a Administración: es consulta, no operación.
    assert "/administracion/pedidos/buscar" not in respuesta.text
    assert "Buscar Pedidos" not in respuesta.text
    # Y Cargar Pedido también: el depósito arma lo que ya está cargado, no
    # transcribe el mail del cliente.
    assert "Cargar Pedido" not in respuesta.text


def test_administracion_tiene_cargar_pedido_en_la_tarjeta_de_pedidos():
    with patch("app.main._banner_alertas", return_value=None):
        respuesta = cliente.get("/administracion")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert 'href="/deposito/pedido/cargar"' in cuerpo
    # Va en Pedidos, junto a Buscar, y no en otra tarjeta.
    pedidos = cuerpo.split("<h2>Pedidos</h2>")[1].split("</div>")[0]
    assert "Cargar Pedido" in pedidos
    assert "Buscar Pedidos" in pedidos


def test_la_pantalla_de_cargar_pedido_vuelve_a_ADMINISTRACION():
    # Si el botón se muda pero la barra no, el atrás devuelve a un Depósito
    # donde ese botón ya no está — el mismo resto que dejó la fase 2.
    with patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR):
        respuesta = cliente.get("/deposito/pedido/cargar")

    assert respuesta.status_code == 200
    assert 'href="/deposito"' not in respuesta.text
    assert "/administracion" in respuesta.text


def test_ver_pedido_sin_cliente_muestra_solo_el_selector():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.get("/deposito/pedido")

    assert respuesta.status_code == 200
    assert "Elegí un cliente" in respuesta.text


def test_ver_pedido_muestra_sucursales_con_oc_declarado_informativo_y_sin_identificar():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_PEDIDO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fotos_pedido", return_value=[]),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    # Cada sucursal con su número de OC bien visible.
    assert "OC 1257673" in respuesta.text
    assert "OC 1257642" in respuesta.text
    # El total declarado es solo un DATO informativo (Día se equivoca
    # sumando — decisión del 24/08/2026): NADA de alarma de descuadre,
    # aunque VL declare 235 y lo leído sume 230.
    assert "pero el mail dice" not in respuesta.text
    assert 'class="descuadre"' not in respuesta.text
    assert "(mail: 235)" in respuesta.text
    # El renglón sin identificar, arriba y con su texto crudo, asignable.
    assert "Sin identificar (1)" in respuesta.text
    assert "99999" in respuesta.text
    assert "RADICHETA" in respuesta.text
    assert f'action="/deposito/pedido/50/renglones/10/asignar"' in respuesta.text
    assert "Guardar este código en la ficha" in respuesta.text


def test_leer_pedido_pegado_elige_el_bloque_de_esta_empresa_y_marca_los_sin_match():
    # El mail trae los DOS bloques juntos: el sistema se queda con el de
    # esta empresa (el encabezado la nombra) y los sin match van arriba.
    datos_leidos = {
        "bloques": [
            {
                "empresa": "9582 FRUTAMAX",
                "sucursales": [
                    {"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 235},
                    {"sucursal": "BZ", "orden_compra": "1257642", "total_bultos": 40},
                ],
                "renglones": [
                    {"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225, "BZ": None}, "confianza": "alta"},
                    {"codigo": "99999", "descripcion": "RADICHETA", "cantidades": {"VL": 10, "BZ": 40}, "confianza": "alta"},
                ],
            },
            {
                "empresa": "11344 PALMALA",
                "sucursales": [{"sucursal": "VL", "orden_compra": "9999", "total_bultos": 10}],
                "renglones": [{"codigo": "555", "descripcion": "OTRA COSA", "cantidades": {"VL": 10}, "confianza": "alta"}],
            },
        ]
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos) as mock_extraer,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-21", "texto": "9582 FRUTAMAX ... el mail ..."},
        )

    assert respuesta.status_code == 200
    mock_extraer.assert_called_once_with("9582 FRUTAMAX ... el mail ...")
    # Se quedó con el bloque de Frutamax, no con el de Palmala.
    assert "Bloque leído: 9582 FRUTAMAX" in respuesta.text
    assert "OTRA COSA" not in respuesta.text
    # El sin match va marcado, con su texto crudo, y arriba del que matcheó.
    assert "Sin match — asignar a mano" in respuesta.text
    assert respuesta.text.index("RADICHETA") < respuesta.text.index("BANANA")
    # Banana matcheó por código exacto de la ficha: llega preseleccionada.
    assert "Guardar pedido" in respuesta.text


def test_confirmar_pedido_expande_por_sucursal_y_guarda_el_alias_pedido():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.obtener_pedido_vigente", return_value=None),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido", return_value=51) as mock_crear,
        patch("app.main.guardar_alias_en_ficha") as mock_alias,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-21", "texto_original": "el mail",
                "cantidad_sucursales": "2",
                "sucursal_0_nombre": "VL", "sucursal_0_oc": "1257673", "sucursal_0_total": "235",
                "sucursal_1_nombre": "BZ", "sucursal_1_oc": "1257642", "sucursal_1_total": "40",
                "cantidad_renglones": "3",
                # Banana: cantidades en las dos sucursales -> DOS renglones.
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "225", "renglon_0_cant_1": "40",
                # Radicheta: sin match, asignada a mano con "guardar alias".
                "renglon_1_codigo": "99999", "renglon_1_descripcion": "RADICHETA",
                "renglon_1_ficha_id": "903", "renglon_1_guardar_alias": "si",
                "renglon_1_cant_0": "10", "renglon_1_cant_1": "",
                # Renglón sin ninguna cantidad: se guarda igual, en 0.
                "renglon_2_codigo": "77777", "renglon_2_descripcion": "SIN CANTIDAD",
                "renglon_2_ficha_id": "", "renglon_2_cant_0": "", "renglon_2_cant_1": "",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "/deposito/pedido?" in respuesta.headers["location"]
    (cliente_arg, fecha_arg, origen_arg, texto_arg, sucursales_arg, renglones_arg), kwargs = mock_crear.call_args
    assert cliente_arg == 1 and fecha_arg == date(2026, 8, 21) and origen_arg == "texto"
    assert texto_arg == "el mail"
    assert sucursales_arg[0] == {"sucursal": "VL", "orden_compra": "1257673", "total_bultos_declarado": 235.0}
    # Banana en dos sucursales = dos renglones; el sin cantidad se guarda
    # igual (invariante: nada del mail se pierde).
    # Cada renglón lleva la FICHA con la que se vende (901 = Banana para
    # este cliente) y el artículo que sale de ella. El sin identificar no
    # tiene ninguna de las dos.
    assert {"sucursal": "VL", "articulo_id": 1, "ficha_id": 901, "texto_codigo": "90101", "texto_descripcion": "BANANA", "cantidad": 225.0} in renglones_arg
    assert {"sucursal": "BZ", "articulo_id": 1, "ficha_id": 901, "texto_codigo": "90101", "texto_descripcion": "BANANA", "cantidad": 40.0} in renglones_arg
    assert {"sucursal": None, "articulo_id": None, "ficha_id": None, "texto_codigo": "77777", "texto_descripcion": "SIN CANTIDAD", "cantidad": 0} in renglones_arg
    assert kwargs == {"reemplaza_a_pedido_id": None, "mail_message_id": None, "recibido_el": None}
    # El alias va a LA FICHA con la que se pidió (903), no a (cliente, artículo).
    mock_alias.assert_called_once_with(903, "99999", "RADICHETA")


def test_confirmar_pedido_con_uno_vigente_lo_reemplaza():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido", return_value=52) as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-21", "texto_original": "mail corregido",
                "cantidad_sucursales": "1", "sucursal_0_nombre": "VL", "sucursal_0_oc": "", "sucursal_0_total": "",
                "cantidad_renglones": "1",
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "60",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "Reemplaza+al+anterior" in respuesta.headers["location"]
    assert mock_crear.call_args.kwargs == {"reemplaza_a_pedido_id": 50, "mail_message_id": None, "recibido_el": None}


def test_asignar_renglon_de_pedido_guarda_la_ficha_y_opcionalmente_el_alias():
    # Se asigna la FICHA (907), no el artículo: el artículo lo deriva la
    # base de ella, y el alias se guarda en ESA ficha — con dos fichas del
    # mismo artículo, guardarlo por (cliente, artículo) se lo pegaba a las
    # dos.
    with (
        patch("app.main.asignar_ficha_a_renglon_pedido") as mock_asignar,
        patch("app.main.guardar_alias_en_ficha") as mock_alias,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/10/asignar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "ficha_id": "907",
                  "guardar_alias": "si", "texto_codigo": "99999", "texto_descripcion": "RADICHETA"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "/deposito/pedido?" in respuesta.headers["location"]
    mock_asignar.assert_called_once_with(10, 907)
    mock_alias.assert_called_once_with(907, "99999", "RADICHETA")


def test_leer_pedido_con_capturas_lee_los_originales_y_arrastra_el_respaldo():
    # Desde el celular: capturas en vez de texto. La IA lee los ORIGINALES
    # (mejor resolución) y las versiones comprimidas viajan a la revisión
    # para quedar como respaldo al confirmar.
    datos_leidos = {
        "bloques": [{
            "empresa": "9582 FRUTAMAX",
            "sucursales": [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 235}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-comprimido") as mock_comprimir,
        patch("app.main.extraer_pedido_de_imagenes", return_value=datos_leidos) as mock_extraer,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-21", "texto": ""},
            files=[
                ("imagenes", ("parte1.png", b"bytes-originales-1", "image/png")),
                ("imagenes", ("parte2.png", b"bytes-originales-2", "image/png")),
            ],
        )

    assert respuesta.status_code == 200
    # Las DOS capturas van juntas a la misma lectura, con los bytes originales.
    mock_extraer.assert_called_once_with([b"bytes-originales-1", b"bytes-originales-2"])
    assert mock_comprimir.call_count == 2
    # Las comprimidas viajan escondidas a la revisión, listas para el respaldo.
    assert 'name="cantidad_fotos" value="2"' in respuesta.text
    assert 'name="foto_data_0" value="data:image/jpeg;base64,' in respuesta.text
    assert "al guardar quedan como respaldo del pedido" in respuesta.text


def test_leer_pedido_sin_texto_ni_capturas_pide_alguno_de_los_dos():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-21", "texto": "   "},
        )

    assert respuesta.status_code == 400
    assert "Pegá el texto del mail o subí una captura" in respuesta.text


def test_confirmar_pedido_sube_las_capturas_como_respaldo():
    import base64
    foto_data = "data:image/jpeg;base64," + base64.standard_b64encode(b"jpeg-comprimido").decode("ascii")
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.obtener_pedido_vigente", return_value=None),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido", return_value=60),
        patch("app.main.subir_foto_comanda", return_value="2026/pedido-60-abc.jpg") as mock_subir,
        patch("app.main.agregar_foto_pedido") as mock_foto,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-21", "texto_original": "",
                "cantidad_fotos": "1", "foto_data_0": foto_data,
                "cantidad_sucursales": "1", "sucursal_0_nombre": "VL", "sucursal_0_oc": "", "sucursal_0_total": "",
                "cantidad_renglones": "1",
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "225",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_subir.assert_called_once_with(b"jpeg-comprimido", "pedido-60")
    mock_foto.assert_called_once_with(60, "2026/pedido-60-abc.jpg")


# --- Pedidos etapa 2: armado con tildes por sucursal ---


RENGLONES_ARMADO_DE_PRUEBA = [
    {"id": 11, "sucursal": "VL", "articulo_id": 1, "articulo_nombre": "Banana", "ficha_id": 901, "nombre_venta": "Banana",
     "texto_codigo": "90101", "texto_descripcion": "BANANA", "cantidad": 15.0,
     "armado_el": None, "cantidad_armada": None, "kilos_enviados": None, "anulado_el": None},
    {"id": 12, "sucursal": "VL", "articulo_id": 2, "articulo_nombre": "Batata", "ficha_id": 902, "nombre_venta": "Batata",
     "texto_codigo": "90102", "texto_descripcion": "BATATA", "cantidad": 20.0,
     "armado_el": datetime(2026, 8, 21, 13, 0), "cantidad_armada": 12.0, "kilos_enviados": 120.0, "anulado_el": None},
    {"id": 13, "sucursal": "BZ", "articulo_id": 1, "articulo_nombre": "Banana", "ficha_id": 901, "nombre_venta": "Banana",
     "texto_codigo": "90101", "texto_descripcion": "BANANA", "cantidad": 40.0,
     "armado_el": datetime(2026, 8, 21, 13, 5), "cantidad_armada": None, "kilos_enviados": None, "anulado_el": None},
]


def test_armar_pedido_muestra_las_sucursales_con_progreso_y_oc():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert respuesta.status_code == 200
    # Una sucursal a la vez: botones grandes con progreso y OC.
    assert "VL — 1 de 2 artículos armados" in respuesta.text
    assert "BZ — 1 de 1 artículos armados" in respuesta.text
    assert "OC 1257673" in respuesta.text
    assert "1 incompleto" in respuesta.text


def test_armar_pedido_en_una_sucursal_separa_pendientes_de_armados():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21&sucursal=VL")

    assert respuesta.status_code == 200
    assert "VL: 1 de 2 artículos armados" in respuesta.text
    # Banana pendiente arriba con su tilde; Batata armada abajo, atenuada,
    # con la marca de incompleto y el botón para destildar.
    assert "Falta armar (1)" in respuesta.text
    assert 'action="/deposito/pedido/50/renglones/11/armar"' in respuesta.text
    assert "Ya armado (1)" in respuesta.text
    assert "armó 12 de 20" in respuesta.text
    assert 'action="/deposito/pedido/50/renglones/12/desarmar"' in respuesta.text
    # El gesto secundario para cargar una cantidad distinta a la pedida.
    # Ya no dice "Armé menos": desde el 05/09 también sirve para más.
    assert "Armé otra cantidad" in respuesta.text


def test_armar_el_atras_de_la_barra_sube_en_la_jerarquia_no_en_el_historial():
    # Cada tilde es un POST+redirect que mete una entrada en el historial
    # del navegador: history.back() devolvía al tilde anterior. El atrás de
    # Armar ahora es un LINK al padre del sistema, calculado por estado.
    ancla = '<a class="barra-boton" href="{destino}" aria-label="Volver atrás">'
    patches = dict(
        _hoy_argentina=date(2026, 8, 21),
        listar_clientes=CLIENTES_PARA_SELECTOR,
        listar_pedidos_vigentes_con_armado=[],
        obtener_pedido_vigente=PEDIDO_VIGENTE_DE_PRUEBA,
        listar_sucursales_pedido=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA],
        listar_renglones_pedido=RENGLONES_ARMADO_DE_PRUEBA,
        listar_fichas_por_cliente=FICHAS_PEDIDO_DE_PRUEBA,
        listar_mails_pedido_sin_procesar_de_cliente=[],
        fichas_con_cajas_armadas=set(),
    )
    with ExitStack() as stack:
        for nombre, valor in patches.items():
            stack.enter_context(patch(f"app.main.{nombre}", return_value=valor))

        # Armando la sucursal VL → atrás = el pedido (lista de sucursales).
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21&sucursal=VL")
        assert ancla.format(destino="/deposito/pedido/armar?cliente_id=1&amp;fecha=2026-08-21") in respuesta.text

        # En el pedido → atrás = el listado de pedidos.
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")
        assert ancla.format(destino="/deposito/pedido/armar?cliente_id=1") in respuesta.text

        # En el listado → atrás = el hub de Depósito.
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1")
        assert ancla.format(destino="/deposito") in respuesta.text

    # Los hubs también declaran su padre: Depósito sube a Inicio.
    respuesta = cliente.get("/deposito")
    assert ancla.format(destino="/inicio") in respuesta.text


def test_armar_pedido_corregido_muestra_el_diff_contra_el_anterior():
    pedido_corregido = dict(PEDIDO_VIGENTE_DE_PRUEBA, id=52, reemplaza_a_pedido_id=50,
                            reemplazado_creado_en=datetime(2026, 8, 21, 12, 10))
    renglones_nuevos = [
        dict(RENGLONES_ARMADO_DE_PRUEBA[0], cantidad=60.0),  # Banana VL: 15 -> 60
        RENGLONES_ARMADO_DE_PRUEBA[2],  # Banana BZ idéntica
    ]

    def _renglones(pedido_id):
        return renglones_nuevos if pedido_id == 52 else RENGLONES_ARMADO_DE_PRUEBA

    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.obtener_pedido_vigente", return_value=pedido_corregido),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", side_effect=_renglones),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert respuesta.status_code == 200
    assert "los cambiados quedaron sin tildar" in respuesta.text
    assert "cambió: Banana VL 15 → 60" in respuesta.text
    assert "ya no está: Batata VL" in respuesta.text


def test_armar_renglon_completo_no_guarda_cantidad():
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "sucursal=VL" in respuesta.headers["location"]
    mock_marcar.assert_called_once_with(11, None, None)


def test_armar_renglon_parcial_guarda_la_cantidad_real():
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_armada": "12", "cantidad_pedida": "15"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(11, 12.0, None)


def test_armar_renglon_con_todo_lo_pedido_cuenta_como_completo():
    # EXACTAMENTE lo pedido: se guarda como armado normal, sin el número
    # redundante. NULL significa "armó justo lo que pedían".
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_armada": "15", "cantidad_pedida": "15"},
            follow_redirects=False,
        )

    mock_marcar.assert_called_once_with(11, None, None)


def test_armar_DE_MAS_guarda_los_bultos_de_verdad():
    """El camión ya salió con 80. Negar el registro no des-entrega nada.

    Hasta el 05/09 un 80 sobre 50 se guardaba como None —o sea 50— y los 30
    de más salían del galpón sin quedar en ningún lado: ni en el stock, ni
    en la factura, ni en una pantalla. Ahora se guarda el 80.
    """
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_armada": "80", "cantidad_pedida": "50"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(11, 80.0, None)


def test_los_kilos_de_un_armado_DE_MAS_salen_por_los_bultos_de_verdad():
    """Lo que se factura son los kilos que se mandaron.

    Antes el kilaje se multiplicaba por lo PEDIDO, así que 80 bultos a 16 kg
    se facturaban como 50 × 16 = 800. Ahora son 80 × 16 = 1280.
    """
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_armada": "80", "cantidad_pedida": "50",
                  "kilos_por_bulto": "16"},
            follow_redirects=False,
        )

    mock_marcar.assert_called_once_with(11, 80.0, 1280.0)


def test_el_renglon_armado_de_MAS_no_lleva_el_ambar_del_incompleto():
    """El ámbar significa "queda algo por hacer" y acá no queda nada.

    Es el `!=` que quedó vivo en DOS plantillas cuando el resto del código
    pasó a `<`. Estaba dormido porque el server nunca guardaba una cantidad
    mayor a la pedida — y este cambio es justamente lo que lo despierta: sin
    esto, un renglón donde SOBRÓ mercadería salía en ámbar pidiendo
    atención.
    """
    renglones = [
        dict(RENGLONES_ARMADO_DE_PRUEBA[1], id=12, cantidad=50.0, cantidad_armada=80.0),
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=renglones),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21&sucursal=VL")

    texto = respuesta.text
    # Se VE, que es lo que hoy no pasa: el 80 no existía en ninguna pantalla.
    assert '<span class="marca-de-mas">armó 80 de 50</span>' in texto
    # Y NO con la clase del ámbar, que dice "queda algo por hacer".
    assert '<span class="marca-incompleto">armó 80 de 50</span>' not in texto
    # Tampoco cuenta como incompleto en el progreso de la sucursal.
    assert "incompleto" not in texto.split("VL: ")[1].split("<")[0]


# --- El Remanente: la vista del depósito, una porción por renglón (06/09) ---

REMANENTE_FILAS = [
    {"articulo_id": 1, "nombre": "Mandarina", "stock": 35.0, "segunda": 3.0, "grupo": "fruta"},
    {"articulo_id": 2, "nombre": "Pomelo", "stock": 31.0, "segunda": 0.0, "grupo": "fruta"},
    {"articulo_id": 3, "nombre": "Mzn Gob", "stock": 20.0, "segunda": 0.0, "grupo": None},
    {"articulo_id": 5, "nombre": "Berenjena", "stock": 20.0, "segunda": 0.0, "grupo": "hortaliza"},
]
REMANENTE_CAJAS = {(1, 11): 15.0, (2, 12): 15.0, (5, 13): 20.0}
# `nombre_cliente` es el CÓDIGO CON EL QUE EL CLIENTE nombra su producto, y acá
# va como es en la base: en mayúscula y sin parecerse al nombre del artículo.
# La fixture vieja tenía "Berenjena Caja Día" —el nombre que queríamos ver, no
# el que hay— y por eso los tests pasaban en verde mientras la pantalla mostraba
# "BERENJENA G". Una fixture escrita con el resultado esperado adentro no prueba
# nada: repite la hipótesis.
REMANENTE_FICHAS = [
    {"id": 11, "articulo_id": 1, "cliente_id": 1, "articulo_nombre": "Mandarina",
     "nombre_cliente": "MANDA COM X 10", "contenido_caja": 10.0, "unidad_venta": "kilo"},
    {"id": 12, "articulo_id": 2, "cliente_id": 1, "articulo_nombre": "Pomelo",
     "nombre_cliente": "POMELO ROSADO G", "contenido_caja": 15.0, "unidad_venta": "kilo"},
    {"id": 13, "articulo_id": 5, "cliente_id": 1, "articulo_nombre": "Berenjena",
     "nombre_cliente": "BERENJENA G", "contenido_caja": 8.0, "unidad_venta": "kilo"},
]
REMANENTE_CLIENTES = [{"id": 1, "nombre": "Día"}, {"id": 2, "nombre": "Vea"}]


def _remanente(filas=None, cajas=None, fichas=None, reingresos=0, clientes=None,
               url="/administracion/stock/remanente"):
    with (
        patch("app.main.stock_deposito_por_articulo",
              return_value=REMANENTE_FILAS if filas is None else filas),
        patch("app.main.cajas_armadas_por_ficha",
              return_value=REMANENTE_CAJAS if cajas is None else cajas),
        patch("app.main.listar_fichas_de_todos_los_clientes",
              return_value=REMANENTE_FICHAS if fichas is None else fichas),
        patch("app.main.listar_clientes",
              return_value=REMANENTE_CLIENTES if clientes is None else clientes),
        patch("app.main.total_reingresos_rechazo", return_value=reingresos),
        # El corte, que es el piso de la pantalla: antes de esa fecha no
        # puede contestar y manda a hoy con un aviso.
        patch("app.main.fecha_corte", return_value=date(2026, 9, 5)),
        patch("app.main._hoy_argentina", return_value=date(2026, 9, 6)),
    ):
        return cliente.get(url)


def _leer_excel_remanente(**kwargs):
    """Lee el Excel del Remanente y CLASIFICA sus filas, en un solo lugar.

    Antes cada test separaba porciones de títulos por "tiene número en la
    columna Sistema", y eso dejó de alcanzar el día que el subtotal —que
    tiene número— se sumó. La clasificación real es visual y es la misma que
    ve el que imprime: lo que va con relleno gris es CHROME (título o
    subtotal) y lo que va sin relleno es MERCADERÍA.

    Devuelve {"titulos", "porciones", "subtotales", "total", "por_seccion"}.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    kwargs.setdefault("url", "/administracion/stock/remanente/exportar-excel")
    hoja = load_workbook(BytesIO(_remanente(**kwargs).content)).active

    leido = {"titulos": [], "porciones": [], "subtotales": [], "total": None, "por_seccion": {}}
    seccion = None
    for fila in hoja.iter_rows(min_row=5, max_col=3):
        nombre, bultos, contado = (c.value for c in fila)
        relleno = fila[0].fill.start_color.rgb if fila[0].fill.start_color else None
        es_chrome = relleno is not None and "D9E2F3" in str(relleno)
        if str(nombre or "").startswith("TOTAL"):
            leido["total"] = (nombre, bultos, contado)
        elif es_chrome and bultos is None:
            seccion = nombre
            leido["titulos"].append(nombre)
        elif es_chrome:
            leido["subtotales"].append((nombre, bultos, contado))
        else:
            leido["porciones"].append((nombre, bultos, contado))
            leido["por_seccion"].setdefault(seccion, []).append(nombre)
    return leido


def _porciones_en_pantalla(texto):
    return re.findall(r'<span class="que">([^<]+)</span>\s*<span class="cuanto">([^<]+)</span>', texto)


def test_el_remanente_es_una_porcion_por_renglon_y_alfabetico():
    """En el piso no hay "un artículo con un total": hay pilas distintas.

    Saber que hay 9 limones sumando 4 sueltos más 5 en caja no le sirve a
    nadie; lo que hace falta es cuántas cajas hay de cada cosa. Y el orden
    agrupa por ARTÍCULO, no por el texto que se muestra.
    """
    respuesta = _remanente()

    assert respuesta.status_code == 200
    assert _porciones_en_pantalla(respuesta.text) == [
        ("Berenjena Caja Día", "20"),
        ("Mandarina", "20"),
        ("Mandarina Caja Día", "15"),
        ("Mandarina Segunda", "3"),
        ("Mzn Gob", "20"),
        ("Pomelo", "16"),
        ("Pomelo Caja Día", "15"),
    ]


def test_el_remanente_NO_dice_la_palabra_suelto_ni_totales_por_articulo():
    """El nombre pelado es la mercadería como viene del puesto, que es como el
    depósito la llama. "Caja Día" y "Segunda" son las que necesitan
    aclaración porque son otra cosa."""
    texto = _remanente().text

    assert "suelto" not in texto.lower()
    # Mandarina son 20 + 15 + 3: el 38 no aparece en ningún lado.
    assert "38" not in texto


def test_el_remanente_manda_a_Stock_Fisico_para_contar():
    """Muestra números del sistema, que es lo contrario del criterio del
    depósito. Dice dónde se cuenta de verdad, y por qué ahí no se ven."""
    texto = _remanente().text

    assert "Esto es para mirar y para exportar" in texto
    assert '/deposito/stock/fisico' in texto


def test_una_caja_se_llama_ARTICULO_MAS_CAJA_CLIENTE_no_con_el_codigo_del_cliente():
    """`nombre_cliente` es el código con el que EL CLIENTE nombra su producto
    ("BERENJENA G", "LIMA X 1KG"). Sirve donde alguien elige una ficha PARA
    ese cliente —el armado, la guía R—, porque es lo que está impreso en la
    caja. Acá no: el que lee el remanente mira su propio depósito.

    Y con el nombre del artículo adelante las tres porciones caen juntas al
    ordenar, que era la idea del orden desde el principio.
    """
    nombres = [n for n, _ in _porciones_en_pantalla(_remanente().text)]

    assert "Berenjena Caja Día" in nombres
    # El código del cliente NO puede aparecer en el caso normal.
    for codigo in ("BERENJENA G", "MANDA COM X 10", "POMELO ROSADO G"):
        assert codigo not in nombres


def test_dos_cajas_del_MISMO_cliente_se_distinguen_por_el_kilaje():
    """Un cliente puede tener varias fichas del mismo artículo — fue el motivo
    de que la clave de venta pasara de artículo a ficha. Sin desempate,
    "Lima Caja Día" saldría dos veces y no se sabría cuál es cuál."""
    filas = [{"articulo_id": 1, "nombre": "Lima", "stock": 12.0, "segunda": 0.0}]
    fichas = [
        {"id": 11, "articulo_id": 1, "cliente_id": 1, "nombre_cliente": "LIMA X 1KG",
         "contenido_caja": 5.0, "unidad_venta": "kilo"},
        {"id": 12, "articulo_id": 1, "cliente_id": 1, "nombre_cliente": "LIMA CHICA",
         "contenido_caja": 10.0, "unidad_venta": "kilo"},
    ]
    nombres = [n for n, _ in _porciones_en_pantalla(
        _remanente(filas=filas, cajas={(1, 11): 4.0, (1, 12): 8.0}, fichas=fichas).text)]

    # El orden es alfabético, así que "10kg" cae antes que "5kg". Se acepta:
    # los dos renglones quedan pegados y con el kilaje a la vista, que es lo
    # que hacía falta. Ordenar por kilaje obligaría a ordenar las fichas por
    # número también cuando son de clientes distintos, y ahí lo que se quiere
    # es el orden por cliente.
    assert nombres == ["Lima Caja Día 10kg", "Lima Caja Día 5kg"]


def test_dos_cajas_de_CLIENTES_distintos_no_necesitan_kilaje():
    """El caso normal se lee limpio: el nombre del cliente ya las distingue."""
    filas = [{"articulo_id": 1, "nombre": "Lima", "stock": 12.0, "segunda": 0.0}]
    fichas = [
        {"id": 11, "articulo_id": 1, "cliente_id": 1, "nombre_cliente": "LIMA X 1KG",
         "contenido_caja": 5.0, "unidad_venta": "kilo"},
        {"id": 12, "articulo_id": 1, "cliente_id": 2, "nombre_cliente": "LIMA G",
         "contenido_caja": 10.0, "unidad_venta": "kilo"},
    ]
    nombres = [n for n, _ in _porciones_en_pantalla(
        _remanente(filas=filas, cajas={(1, 11): 4.0, (1, 12): 8.0}, fichas=fichas).text)]

    assert nombres == ["Lima Caja Día", "Lima Caja Vea"]


def test_dos_cajas_del_mismo_cliente_y_MISMO_kilaje_caen_al_codigo_del_cliente():
    """El modelo tampoco prohíbe esto. El kilaje ya no desempata, así que se
    usa lo único que seguro las distingue. Feo, pero solo en el caso feo: es
    preferible a dos renglones idénticos con números distintos."""
    filas = [{"articulo_id": 1, "nombre": "Lima", "stock": 12.0, "segunda": 0.0}]
    fichas = [
        {"id": 11, "articulo_id": 1, "cliente_id": 1, "nombre_cliente": "LIMA X 1KG",
         "contenido_caja": 5.0, "unidad_venta": "kilo"},
        {"id": 12, "articulo_id": 1, "cliente_id": 1, "nombre_cliente": "LIMA CHICA",
         "contenido_caja": 5.0, "unidad_venta": "kilo"},
    ]
    nombres = [n for n, _ in _porciones_en_pantalla(
        _remanente(filas=filas, cajas={(1, 11): 4.0, (1, 12): 8.0}, fichas=fichas).text)]

    assert nombres == ["Lima Caja Día 5kg", "Lima Caja Día 5kg"]


# --- La fecha del Remanente ---


def test_el_remanente_sin_fecha_muestra_HOY_y_no_se_marca_como_pasado():
    cuerpo = _remanente().text.split("</style>")[-1]

    assert 'value="2026-09-06"' in cuerpo
    assert "No es el stock de ahora" not in cuerpo
    assert "Volver a hoy" not in cuerpo


def test_una_fecha_pedida_llega_a_LAS_TRES_cuentas_y_al_link_del_Excel():
    """Las tres se cortan a la misma fecha o los renglones no cierran entre sí:
    los sueltos salen de restarle a la primera las cajas de la segunda."""
    with (
        patch("app.main.stock_deposito_por_articulo",
              return_value=[{"articulo_id": 1, "nombre": "Lima", "stock": 4.0, "segunda": 0.0}]) as filas,
        patch("app.main.cajas_armadas_por_ficha", return_value={}) as cajas,
        patch("app.main.total_reingresos_rechazo", return_value=0) as reingresos,
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=REMANENTE_CLIENTES),
        patch("app.main.fecha_corte", return_value=date(2026, 9, 5)),
        patch("app.main._hoy_argentina", return_value=date(2026, 9, 10)),
    ):
        respuesta = cliente.get("/administracion/stock/remanente?fecha=2026-09-08")

    pedida = date(2026, 9, 8)
    filas.assert_called_once_with(pedida)
    cajas.assert_called_once_with(pedida)
    reingresos.assert_called_once_with(pedida)
    cuerpo = respuesta.text.split("</style>")[-1]
    # Y hay que poder VER que no es hoy, no solo la fecha en gris chico.
    assert "Así estaba el depósito al cierre del 08/09/2026" in cuerpo
    assert "No es el stock de ahora" in cuerpo
    assert "Volver a hoy" in cuerpo
    # El Excel que baja tiene que ser el de lo que se está mirando.
    assert "exportar-excel?fecha=2026-09-08" in cuerpo


def test_el_excel_se_baja_a_la_fecha_pedida_y_el_archivo_la_lleva_en_el_nombre():
    """Dos exports de días distintos no se pueden pisar en la carpeta."""
    from io import BytesIO

    from openpyxl import load_workbook

    with (
        patch("app.main.stock_deposito_por_articulo",
              return_value=[{"articulo_id": 1, "nombre": "Lima", "stock": 4.0, "segunda": 0.0}]),
        patch("app.main.cajas_armadas_por_ficha", return_value={}),
        patch("app.main.total_reingresos_rechazo", return_value=0),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=REMANENTE_CLIENTES),
        patch("app.main.fecha_corte", return_value=date(2026, 9, 5)),
        patch("app.main._hoy_argentina", return_value=date(2026, 9, 10)),
    ):
        respuesta = cliente.get(
            "/administracion/stock/remanente/exportar-excel?fecha=2026-09-08")

    assert 'filename="Remanente_08_09_2026.xlsx"' in respuesta.headers["content-disposition"]
    hoja = load_workbook(BytesIO(respuesta.content)).active
    assert hoja["A2"].value == "Al 08/09/2026"


def test_una_fecha_ANTERIOR_AL_CORTE_no_se_contesta_y_dice_por_que():
    """No es que falten datos: las tres cuentas serían de dos épocas. El total
    lo rebasea el compensatorio y las cajas por ficha arrancan EN el corte, así
    que el 20/08 daría el total del modelo viejo junto a cero cajas."""
    respuesta = _remanente(url="/administracion/stock/remanente?fecha=2026-08-20")
    cuerpo = respuesta.text.split("</style>")[-1]

    assert "anterior al corte del modelo" in cuerpo
    assert "Se muestra el día de hoy" in cuerpo
    # Y muestra HOY de verdad, no una tabla vacía ni un error.
    assert 'value="2026-09-06"' in cuerpo
    assert "No es el stock de ahora" not in cuerpo


def test_una_fecha_FUTURA_o_ROTA_cae_a_hoy_con_aviso():
    for pedida, esperado in (("2026-12-25", "Todavía no pasó ese día"),
                             ("mañana", "no se entendió")):
        cuerpo = _remanente(
            url=f"/administracion/stock/remanente?fecha={pedida}").text.split("</style>")[-1]
        assert esperado in cuerpo
        assert 'value="2026-09-06"' in cuerpo


def test_los_negativos_van_ABAJO_Y_APARTE_de_las_porciones():
    """La lista de arriba es lo que HAY; los negativos son un problema a
    resolver. Si se mezclan como un renglón más, vuelve a ser la pantalla
    que se borró."""
    filas = [
        {"articulo_id": 1, "nombre": "Berenjena", "stock": 46.0, "reproceso_primera": 0.0, "segunda": 0.0},
        {"articulo_id": 2, "nombre": "Palta", "stock": -45.0, "reproceso_primera": 0.0, "segunda": 0.0},
    ]
    cuerpo = _remanente(filas=filas, cajas={}, fichas=[]).text.split("</style>")[-1]

    # Palta NO está entre las porciones: no hay pila que contar.
    assert [n for n, _ in _porciones_en_pantalla(cuerpo)] == ["Berenjena"]
    # Está en su propio bloque, y el bloque va DESPUÉS de la lista.
    assert 'class="bloque-negativos"' in cuerpo
    assert cuerpo.index('class="porcion"') < cuerpo.index('class="bloque-negativos"')
    # Y linkea a Stock por Guía, que es la única puerta que le queda.
    assert 'href="/administracion/stock/sistema/2"' in cuerpo


def test_un_negativo_dice_QUE_PASO_no_cuanto_hay():
    """Un artículo en −45 no tiene menos cuarenta y cinco cajones: tiene 45
    bultos que salieron sin que ninguna guía los cubra. Mostrar "−45" en una
    pantalla de cantidades invita a leerlo como stock y restarlo de algo — es
    el mismo defecto del "sin procesar −95" que esta pantalla vino a
    reemplazar."""
    filas = [{"articulo_id": 2, "nombre": "Palta", "stock": -45.0,
              "reproceso_primera": 0.0, "segunda": 0.0}]
    cuerpo = _remanente(filas=filas, cajas={}, fichas=[]).text.split("</style>")[-1]

    assert "Faltan explicar 45 bultos" in cuerpo
    # El número NUNCA en la columna de la derecha, donde van los bultos que hay.
    assert "-45" not in cuerpo and "−45" not in cuerpo
    assert '<span class="cuanto">' not in cuerpo


def test_sin_negativos_el_bloque_no_aparece():
    filas = [{"articulo_id": 1, "nombre": "Berenjena", "stock": 46.0,
              "reproceso_primera": 0.0, "segunda": 0.0}]
    cuerpo = _remanente(filas=filas, cajas={}, fichas=[]).text.split("</style>")[-1]

    assert 'class="bloque-negativos"' not in cuerpo
    assert "Faltan explicar" not in cuerpo


def test_el_remanente_muestra_los_dos_pools_aparte_y_en_chico():
    """Reingresos y segunda van APARTE del stock normal, y sobrevivieron a
    Stock del Sistema. En chico: son contexto, no lo que se viene a mirar."""
    filas = [{"articulo_id": 1, "nombre": "Berenjena", "stock": 46.0,
              "reproceso_primera": 0.0, "segunda": 7.0}]
    cuerpo = _remanente(filas=filas, cajas={}, fichas=[], reingresos=12).text.split("</style>")[-1]

    assert "12 bultos reingresados por rechazo" in cuerpo
    assert "7 bultos de segunda esperando el remito al Puesto" in cuerpo


def test_el_REMANENTE_NO_se_ofrece_desde_el_deposito():
    """LA REGLA, y por eso tiene test propio: al operario no se le muestran
    los números del sistema de lo que después tiene que contar. Con la lista
    a mano, el conteo de Stock Físico se transcribe en vez de contarse, y un
    conteo transcripto confirma al sistema en vez de controlarlo.

    Estuvo un día en Depósito, en una tarjeta aparte y con un cartel que
    pedía no usarlo para eso. Una tarjeta no es una frontera: los dos
    botones estaban en la misma pantalla. Es la misma regla que la del
    armado ("si lo ve, arma contra el sistema en vez de contra el piso"),
    que ahí se cumple ESTRUCTURALMENTE — el número no llega al HTML.
    """
    deposito = cliente.get("/deposito/stock").text
    assert "remanente" not in deposito.lower()

    # Y donde sí vive: en Administración, primera de "Control de stock".
    administracion = cliente.get("/administracion").text
    assert '/administracion/stock/remanente' in administracion
    assert (administracion.index('/administracion/stock/remanente')
            < administracion.index('/administracion/stock/cotejo'))


def test_un_articulo_que_no_se_reprocesa_es_un_solo_renglon_pelado():
    """Manzana, pera, arándano: sin guías R con ficha, el nombre pelado se
    lleva el total entero. Sale del modelo, sin ninguna excepción escrita."""
    nombres = [n for n, _ in _porciones_en_pantalla(_remanente().text)]

    assert nombres.count("Mzn Gob") == 1
    assert not any(n.startswith("Mzn Gob ") for n in nombres)


def test_una_porcion_en_CERO_no_es_un_renglon():
    """Berenjena tiene 20 de stock y 20 en cajas: no hay sueltos. Un renglón
    en cero sería mandar al que arma a buscar una pila vacía."""
    nombres = [n for n, _ in _porciones_en_pantalla(_remanente().text)]

    assert "Berenjena Caja Día" in nombres
    assert "Berenjena" not in nombres


def test_el_excel_del_remanente_sale_con_la_columna_CONTADO_vacia():
    """Vacía a propósito: es para escribir a mano contra el conteo. Si saliera
    precargada, el que cuenta transcribe en vez de contar."""
    from io import BytesIO

    from openpyxl import load_workbook

    respuesta = _remanente(url="/administracion/stock/remanente/exportar-excel")

    assert respuesta.status_code == 200
    assert 'filename="Remanente_06_09_2026.xlsx"' in respuesta.headers["content-disposition"]
    hoja = load_workbook(BytesIO(respuesta.content)).active
    assert hoja.title == "Remanente"
    assert [c.value for c in hoja[4]] == ["Producto", "Sistema", "Contado"]
    datos = [(f[0], f[1], f[2]) for f in hoja.iter_rows(min_row=5, max_col=3, values_only=True)]
    # La tercera columna SIEMPRE vacía, el total incluido: si el que cuenta ve
    # un total del sistema al pie, tiene contra qué cuadrar sin haber contado.
    assert all(fila[2] is None for fila in datos)
    # SIN PLATA: ninguna celda con un número que parezca un costo.
    assert hoja.max_column == 3


def test_el_excel_del_remanente_sale_en_EL_MISMO_ORDEN_que_la_pantalla():
    """Dos criterios de orden se van separando. Acá hay UNO: las porciones se
    ordenan en app/main.py y el exportador no reordena nada.

    Se compara contra lo que la PANTALLA muestra de verdad, no contra una
    lista escrita a mano: una lista a mano es otra copia del criterio, y el
    día que el orden cambie los dos tendrían que cambiar juntos.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    pantalla = [n for n, _ in _porciones_en_pantalla(_remanente().text)]
    excel = [nombre for nombre, _b, _c in _leer_excel_remanente()["porciones"]]

    # Mismas porciones, sin perder ni inventar ninguna.
    assert sorted(excel) == sorted(pantalla)
    # Y dentro de cada sección, el mismo orden relativo que la pantalla.
    for porciones in _leer_excel_remanente()["por_seccion"].values():
        assert porciones == [n for n in pantalla if n in porciones]
    # Las porciones de un artículo caen juntas: es para lo que sirve el orden.
    assert pantalla[:3] == ["Berenjena Caja Día", "Mandarina", "Mandarina Caja Día"]


def test_el_excel_del_remanente_cierra_con_UN_total_al_pie():
    """Cuántos renglones y cuántos bultos, para ver de un vistazo si el archivo
    impreso está completo. UNO solo al pie, no uno por artículo: el total por
    artículo es justo la suma que el dueño pidió no mostrar."""
    from io import BytesIO

    from openpyxl import load_workbook

    leido = _leer_excel_remanente()
    porciones, total = leido["porciones"], leido["total"]

    # Cuenta PORCIONES, no filas escritas: ni los títulos ni los subtotales
    # son renglones que alguien tenga que ir a contar.
    assert total[0] == f"TOTAL — {len(porciones)} renglones"
    assert total[1] == round(sum(p[1] for p in porciones), 2)
    assert total[2] is None
    # Y el total del pie es la suma de los subtotales: las dos cuentas tienen
    # que cerrar entre sí o el que imprime no sabe a cuál creerle.
    assert total[1] == round(sum(s[1] for s in leido["subtotales"]), 2)


def test_el_excel_agrupa_por_GRUPO_y_deja_las_CAJAS_PROCESADAS_al_final():
    """Es para caminar el depósito: primero cada grupo de artículo en el orden
    fijo del sistema, y al final las cajas armadas, que son una pila aparte y
    se cuentan en otro momento."""
    from io import BytesIO

    from openpyxl import load_workbook

    leido = _leer_excel_remanente()

    # Hortaliza NO sale, y está bien: la Berenjena de la fixture tiene 20 de
    # stock y 20 en caja, así que sus sueltos son cero y su única porción es
    # la caja, que va al final.
    assert leido["titulos"] == ["Fruta", "Sin grupo", "Cajas Procesadas"]
    assert leido["por_seccion"]["Fruta"] == ["Mandarina", "Mandarina Segunda", "Pomelo"]
    assert leido["por_seccion"]["Sin grupo"] == ["Mzn Gob"]
    assert leido["por_seccion"]["Cajas Procesadas"] == [
        "Berenjena Caja Día", "Mandarina Caja Día", "Pomelo Caja Día"]


def test_cada_seccion_CIERRA_con_su_subtotal():
    """Con la hoja impresa, saber cuántos bultos hay en una sección antes de
    pasar a la siguiente acota DÓNDE buscar si algo no cierra, sin rehacer la
    suma entera."""
    leido = _leer_excel_remanente()

    etiquetas = [nombre for nombre, _b, _c in leido["subtotales"]]
    assert etiquetas == [
        "Subtotal Fruta — 3 renglones",
        "Subtotal Sin grupo — 1 renglón",
        "Subtotal Cajas Procesadas — 3 renglones",
    ]
    # Cada uno suma lo suyo, y solo lo suyo.
    por_etiqueta = {n: b for n, b, _c in leido["subtotales"]}
    por_nombre = {n: b for n, b, _c in leido["porciones"]}
    for titulo, nombres in leido["por_seccion"].items():
        esperado = round(sum(por_nombre[n] for n in nombres), 2)
        assert por_etiqueta[f"Subtotal {titulo} — {len(nombres)} "
                           f"{'renglón' if len(nombres) == 1 else 'renglones'}"] == esperado


def test_el_subtotal_sigue_LAS_MISMAS_TRES_REGLAS_que_el_total():
    """Valor y no fórmula, "Contado" vacía, y que no se pueda confundir con un
    renglón de mercadería: impreso tiene que leerse como cierre de sección."""
    from io import BytesIO

    from openpyxl import load_workbook

    hoja = load_workbook(BytesIO(
        _remanente(url="/administracion/stock/remanente/exportar-excel").content)).active

    filas_subtotal = [f for f in hoja.iter_rows(min_row=5, max_col=3)
                      if str(f[0].value or "").startswith("Subtotal ")]
    assert filas_subtotal

    for nombre, bultos, contado in filas_subtotal:
        # 1. VALOR, no fórmula: importa lo que quedó impreso en el papel.
        assert isinstance(bultos.value, (int, float))
        assert not str(bultos.value).startswith("=")
        # 2. "Contado" vacía: si el que cuenta ve un subtotal del sistema, ya
        #    tiene contra qué cuadrar sin haber contado.
        assert contado.value is None
        # 3. Se lee como CHROME, no como mercadería: mismo relleno que el
        #    título de su sección y en negrita. Impreso, la sección queda
        #    encerrada entre dos franjas grises.
        for celda in (nombre, bultos, contado):
            assert celda.font.bold
            assert "D9E2F3" in str(celda.fill.start_color.rgb)


def test_la_SEGUNDA_no_es_una_caja_procesada():
    """Son bultos sueltos de calidad menor esperando el remito al Puesto, no
    cajas armadas para un cliente. Se queda con su artículo."""
    from io import BytesIO

    from openpyxl import load_workbook

    por_seccion = _leer_excel_remanente()["por_seccion"]

    assert "Mandarina Segunda" in por_seccion["Fruta"]
    assert "Mandarina Segunda" not in por_seccion["Cajas Procesadas"]


def test_una_seccion_sin_nada_NO_sale_en_el_excel():
    """Una hoja impresa con "HOJA" y ningún renglón abajo hace dudar de si
    falta algo o no hay."""
    from io import BytesIO

    from openpyxl import load_workbook

    leido = _leer_excel_remanente(
        filas=[{"articulo_id": 1, "nombre": "Lima", "stock": 4.0,
                "segunda": 0.0, "grupo": "fruta"}], cajas={}, fichas=[])

    assert leido["titulos"] == ["Fruta"]
    for ausente in ("Hortaliza", "Hoja", "Pesada", "Sin grupo", "Cajas Procesadas"):
        assert ausente not in leido["titulos"]


def test_un_grupo_que_el_sistema_no_conoce_cae_en_SIN_GRUPO_y_no_desaparece():
    """Si mañana alguien carga "bolsa" en la base sin agregarlo a ORDEN_GRUPOS,
    el artículo tiene que seguir apareciendo: se cuenta igual."""
    from io import BytesIO

    from openpyxl import load_workbook

    leido = _leer_excel_remanente(
        filas=[{"articulo_id": 1, "nombre": "Papa Bolsa", "stock": 4.0,
                "segunda": 0.0, "grupo": "bolsa"}], cajas={}, fichas=[])

    assert leido["titulos"] == ["Sin grupo"]
    assert leido["por_seccion"]["Sin grupo"] == ["Papa Bolsa"]
    assert leido["total"][1] == 4


def test_el_boton_de_exportar_esta_pegado_al_de_la_fecha():
    """Lo que baja es el archivo de la fecha del campo de arriba: juntos, eso
    se ve; al pie de la lista, no."""
    cuerpo = _remanente().text.split("</style>")[-1]

    assert (cuerpo.index("Ver esa fecha")
            < cuerpo.index("Exportar Excel")
            < cuerpo.index("Qué hay en el depósito"))


def test_el_total_del_excel_dice_renglon_en_singular_con_uno_solo():
    from io import BytesIO

    from openpyxl import load_workbook

    filas = [{"articulo_id": 1, "nombre": "Lima", "stock": 4.0, "segunda": 0.0}]
    respuesta = _remanente(filas=filas, cajas={}, fichas=[],
                           url="/administracion/stock/remanente/exportar-excel")
    hoja = load_workbook(BytesIO(respuesta.content)).active
    ultima = list(hoja.iter_rows(min_row=5, max_col=2, values_only=True))[-1]

    assert ultima == ("TOTAL — 1 renglón", 4)


def test_desarmar_renglon_destilda():
    with patch("app.main.desmarcar_renglon_armado") as mock_desmarcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/12/desarmar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_desmarcar.assert_called_once_with(12)


def test_ver_pedido_muestra_el_incompleto_y_el_armado_real_por_sucursal():
    # La pantalla Pedido es la del que factura o atiende el reclamo de
    # Día: el "armó 12 de 20" y el total real armado tienen que estar acá.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fotos_pedido", return_value=[]),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    assert "armó 12 de 20" in respuesta.text
    # VL: solo Batata armada (12 reales de 20) -> armado real 12, 1 incompleto.
    assert "Armado real (1 de 2 renglones)" in respuesta.text
    assert "12 bultos — 1 incompleto" in respuesta.text
    # Botón para ir a armar.
    assert 'href="/deposito/pedido/armar?cliente_id=1' in respuesta.text


# --- Casilla de pedidos (etapa 3, tramo 1) ---

from core.casilla_pedidos import ErrorCasilla  # noqa: E402

ARGENTINA_TEST = timezone(timedelta(hours=-3))

CASILLA_DE_PRUEBA = {
    "id": 3, "direccion": "casilla@empresa.com", "servidor_imap": "imap.gmail.com",
    "cliente_id": 1, "asunto_filtro": "Pedido Dia", "remitentes_permitidos": "pedidos@dia.com.ar",
    "activa": True, "fecha_activacion": datetime(2026, 8, 22, 11, 0, tzinfo=ARGENTINA_TEST),
    "auto_confirmar": False, "ultima_revision_el": None, "ultima_revision_automatica_el": None,
    "ultimo_error": None, "ultimo_error_el": None,
    # Horario de revisión sin configurar: rigen los defaults (12-15, cada 15).
    "revision_desde": None, "revision_hasta": None, "revision_cada_minutos": None,
    "cliente_nombre": "Día",
}

MAIL_PEDIDO_DE_PRUEBA = {
    "id": 9, "casilla_id": 3, "cliente_id": 1, "message_id": "<pedido-1@dia.com.ar>",
    "remitente": "pedidos@dia.com.ar", "asunto": "Pedido del dia",
    "recibido_el": datetime(2026, 8, 22, 12, 5, tzinfo=ARGENTINA_TEST),
    "cuerpo_crudo": "cuerpo crudo de prueba (texto plano)", "cuerpo_texto": "9582 FRUTAMAX\nVL\t235",
    "estado": "pendiente", "motivo": None, "pedido_id": None, "procesado_el": None,
    "creado_en": datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST),
    "cliente_nombre": "Día",
}


def test_ver_sistema_muestra_el_acceso_a_la_casilla_de_pedidos():
    with patch("app.main.obtener_uso_storage_bucket", return_value={"cantidad": 12, "bytes_totales": 907397}):
        respuesta = cliente.get("/sistema")

    assert respuesta.status_code == 200
    assert 'href="/sistema/casilla-pedidos"' in respuesta.text


def test_ver_casilla_pedidos_sin_casillas_muestra_el_alta_y_avisa_que_falta_la_clave():
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value=None),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    # El alta, con el filtro de remitente obligatorio.
    assert 'action="/sistema/casilla-pedidos/guardar"' in respuesta.text
    assert "Remitentes permitidos" in respuesta.text
    # La clave falta y la pantalla lo dice, apuntando a la variable de Railway.
    assert "Falta la clave" in respuesta.text
    assert "CLAVE_CASILLA_PEDIDOS" in respuesta.text
    # Las garantías, dichas de frente.
    assert "solo lectura estricta" in respuesta.text


def test_ver_casilla_pedidos_muestra_el_estado_y_el_boton_revisar_ahora():
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.listar_mails_pedido", return_value=[dict(MAIL_PEDIDO_DE_PRUEBA)]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    assert "casilla@empresa.com" in respuesta.text
    assert ">Activa<" in respuesta.text
    assert 'action="/sistema/casilla-pedidos/3/revisar"' in respuesta.text
    assert "está configurada" in respuesta.text
    # El mail pendiente, con sus dos acciones.
    assert "Pedido del dia" in respuesta.text
    assert 'href="/deposito/pedido/mails/9/revisar"' in respuesta.text
    assert 'action="/sistema/casilla-pedidos/mails/9/ignorar"' in respuesta.text


def test_ver_casilla_pedidos_destaca_un_error_mas_nuevo_que_la_ultima_revision_ok():
    casilla = dict(
        CASILLA_DE_PRUEBA,
        ultima_revision_el=datetime(2026, 8, 21, 12, 0, tzinfo=ARGENTINA_TEST),
        ultimo_error="login fallido",
        ultimo_error_el=datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_TEST),
    )
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert "La última revisión falló" in respuesta.text
    assert "login fallido" in respuesta.text


def test_ver_casilla_pedidos_muestra_el_error_recuperado_sin_destacarlo():
    # Falló a las 12:00 pero a las 12:15 pudo: el error queda a la VISTA
    # (para diagnosticar), pero sin el bloque rojo de "está fallando".
    casilla = dict(
        CASILLA_DE_PRUEBA,
        ultima_revision_el=datetime(2026, 8, 22, 12, 15, tzinfo=ARGENTINA_TEST),
        ultimo_error="caída momentánea",
        ultimo_error_el=datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_TEST),
    )
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert "La última revisión falló" not in respuesta.text
    assert "Último error (ya recuperado)" in respuesta.text
    assert "caída momentánea" in respuesta.text


def test_guardar_casilla_nueva_normaliza_y_redirige():
    with patch("app.main.crear_casilla_pedidos", return_value=3) as mock_crear:
        respuesta = cliente.post(
            "/sistema/casilla-pedidos/guardar",
            data={"direccion": " Casilla@Empresa.com ", "servidor_imap": "",
                  "cliente_id": "1", "asunto_filtro": "  Pedido   Dia ",
                  "remitentes_permitidos": " Pedidos@dia.com.ar , otro@dia.com.ar "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        "casilla@empresa.com", "imap.gmail.com", 1, "Pedido Dia", "pedidos@dia.com.ar, otro@dia.com.ar"
    )


def test_guardar_casilla_sin_asunto_no_guarda():
    # El asunto es EL filtro obligatorio: sin él no se lee nada del buzón.
    with (
        patch("app.main.crear_casilla_pedidos") as mock_crear,
        patch("app.main.listar_casillas_pedidos", return_value=[]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.post(
            "/sistema/casilla-pedidos/guardar",
            data={"direccion": "casilla@empresa.com", "cliente_id": "1", "asunto_filtro": "  ",
                  "remitentes_permitidos": "pedidos@dia.com.ar"},
        )

    assert respuesta.status_code == 400
    assert "Falta el filtro de asunto" in respuesta.text
    mock_crear.assert_not_called()


def test_guardar_casilla_sin_remitentes_guarda_con_cualquier_remitente():
    # El remitente es OPCIONAL: vacío = cualquiera, así el pedido no se
    # pierde porque cambió quién lo manda.
    with patch("app.main.crear_casilla_pedidos", return_value=3) as mock_crear:
        respuesta = cliente.post(
            "/sistema/casilla-pedidos/guardar",
            data={"direccion": "casilla@empresa.com", "cliente_id": "1",
                  "asunto_filtro": "Pedido Dia", "remitentes_permitidos": "  "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with("casilla@empresa.com", "imap.gmail.com", 1, "Pedido Dia", None)


def test_activar_casilla_sin_fecha_activa_desde_ahora():
    with patch("app.main.activar_casilla_pedidos") as mock_activar:
        respuesta = cliente.post(
            "/sistema/casilla-pedidos/3/activar", data={"fecha_activacion": ""}, follow_redirects=False
        )

    assert respuesta.status_code == 303
    (casilla_id, fecha), _ = mock_activar.call_args
    assert casilla_id == 3
    # "Ahora" en hora argentina, con zona: solo correos posteriores a este instante.
    assert fecha.tzinfo is not None


def test_revisar_ahora_reporta_el_detalle_de_lo_que_encontro():
    mail_leido = {
        "message_id": "<pedido-1@dia.com.ar>", "remitente": "pedidos@dia.com.ar",
        "asunto": "Pedido del dia", "recibido_el": datetime(2026, 8, 22, 12, 5, tzinfo=ARGENTINA_TEST),
        "cuerpo_crudo": "<html>...</html>", "cuerpo_texto": "texto",
    }
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=dict(CASILLA_DE_PRUEBA)),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.revisar_casilla",
              return_value={"total_desde": 12, "candidatos": 3, "con_asunto": 1, "mails": [mail_leido]}) as mock_revisar,
        patch("app.main.registrar_mail_pedido", return_value=9) as mock_registrar,
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    # La conexión usa la config guardada y la clave de la variable.
    mock_revisar.assert_called_once_with(
        "casilla@empresa.com", "clave", "imap.gmail.com",
        CASILLA_DE_PRUEBA["fecha_activacion"], "Pedido Dia", ["pedidos@dia.com.ar"],
    )
    mock_registrar.assert_called_once()
    # El botón MANUAL no sella la revisión automática (la alerta no lo cuenta).
    mock_revision.assert_called_once_with(3)
    # El detalle completo en el mensaje: total, por remitente, por asunto, nuevos.
    destino = respuesta.headers["location"]
    assert "Conectado+OK+a+casilla%40empresa.com" in destino
    assert "12+mails+desde+la+activaci%C3%B3n" in destino
    assert "3+de+remitentes+permitidos" in destino
    assert "1+con+el+asunto" in destino
    assert "Pedido+Dia" in destino
    assert "1+nuevo+por+confirmar" in destino


def test_revisar_ahora_sin_remitentes_no_menciona_ese_filtro():
    casilla = dict(CASILLA_DE_PRUEBA, remitentes_permitidos=None)
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=casilla),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.revisar_casilla",
              return_value={"total_desde": 12, "candidatos": 12, "con_asunto": 0, "mails": []}) as mock_revisar,
        patch("app.main.registrar_revision_casilla"),
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    mock_revisar.assert_called_once_with(
        "casilla@empresa.com", "clave", "imap.gmail.com",
        CASILLA_DE_PRUEBA["fecha_activacion"], "Pedido Dia", [],
    )
    destino = respuesta.headers["location"]
    assert "remitentes+permitidos" not in destino
    # Hay mails pero ninguno con el asunto: el aviso para afinar el filtro.
    assert "0+con+el+asunto" in destino
    assert "ninguno+contiene+ese+asunto" in destino


def test_revisar_ahora_avisa_si_hay_mails_pero_ninguno_pasa_el_filtro_de_remitente():
    # El remitente mal escrito se ve DE UNA, y el aviso recuerda que se
    # puede dejar vacío.
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=dict(CASILLA_DE_PRUEBA)),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.revisar_casilla",
              return_value={"total_desde": 12, "candidatos": 0, "con_asunto": 0, "mails": []}),
        patch("app.main.registrar_revision_casilla"),
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    assert "ninguno+pasa+el+filtro+de+remitente" in respuesta.headers["location"]


def test_revisar_ahora_sin_asunto_configurado_pide_configurarlo():
    casilla = dict(CASILLA_DE_PRUEBA, asunto_filtro=None)
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=casilla),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.revisar_casilla") as mock_revisar,
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert "Falta+el+filtro+de+asunto" in respuesta.headers["location"]
    mock_revisar.assert_not_called()


def test_revisar_ahora_sin_clave_no_conecta_y_apunta_a_railway():
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=dict(CASILLA_DE_PRUEBA)),
        patch("app.main.clave_casilla_configurada", return_value=None),
        patch("app.main.revisar_casilla") as mock_revisar,
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert "CLAVE_CASILLA_PEDIDOS" in respuesta.headers["location"]
    mock_revisar.assert_not_called()


def test_revisar_ahora_con_error_imap_lo_registra_y_lo_muestra():
    with (
        patch("app.main.obtener_casilla_pedidos", return_value=dict(CASILLA_DE_PRUEBA)),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.revisar_casilla", side_effect=ErrorCasilla("IMAP deshabilitado")),
        patch("app.main.registrar_mail_pedido") as mock_registrar,
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        respuesta = cliente.post("/sistema/casilla-pedidos/3/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    # El error queda GRABADO (después alimenta la alerta de Auditoría) y a la vista.
    mock_revision.assert_called_once_with(3, error="IMAP deshabilitado")
    assert "IMAP+deshabilitado" in respuesta.headers["location"]
    mock_registrar.assert_not_called()


# --- "Buscar pedido" en Armar Pedido: la misma revisión de la casilla, sin ir a Sistema ---


def test_armar_lista_muestra_el_boton_buscar_pedido_y_los_mails_pendientes():
    # Si el que arma no ve el pedido del día, lo busca él mismo desde acá;
    # y un mail pendiente se confirma acá mismo, sin ir a otra pantalla.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente",
              return_value=[dict(MAIL_PEDIDO_DE_PRUEBA)]) as mock_mails,
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1")

    assert respuesta.status_code == 200
    mock_mails.assert_called_once_with(1)
    texto = respuesta.text
    assert 'action="/deposito/pedido/armar/buscar-pedido"' in texto
    assert "Buscar pedido en la casilla" in texto
    # El mail pendiente, visible y confirmable ahí mismo.
    assert "Mails de pedido por confirmar (1)" in texto
    assert 'href="/deposito/pedido/mails/9/revisar"' in texto
    assert "pedido estimado del 22/08/2026" in texto


def test_armar_sin_pedido_para_la_fecha_ofrece_buscarlo_en_la_casilla():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.obtener_pedido_vigente", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert respuesta.status_code == 200
    texto = respuesta.text
    assert 'action="/deposito/pedido/armar/buscar-pedido"' in texto
    # El botón vuelve a ESTA fecha después de revisar.
    assert '<input type="hidden" name="fecha" value="2026-08-21">' in texto


def test_buscar_pedido_revisa_solo_las_casillas_activas_del_cliente_y_vuelve_a_armar():
    mail_leido = {
        "message_id": "<pedido-1@dia.com.ar>", "remitente": "pedidos@dia.com.ar",
        "asunto": "Pedido del dia", "recibido_el": datetime(2026, 8, 22, 12, 5, tzinfo=ARGENTINA_TEST),
        "cuerpo_crudo": "<html>...</html>", "cuerpo_texto": "texto",
    }
    casillas = [
        dict(CASILLA_DE_PRUEBA),
        # De OTRO cliente y una inactiva: no se tocan.
        dict(CASILLA_DE_PRUEBA, id=4, cliente_id=2),
        dict(CASILLA_DE_PRUEBA, id=5, activa=False),
    ]
    with (
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.listar_casillas_pedidos", return_value=casillas),
        patch("app.main.revisar_casilla",
              return_value={"total_desde": 3, "candidatos": 1, "con_asunto": 1, "mails": [mail_leido]}) as mock_revisar,
        patch("app.main.registrar_mail_pedido", return_value=9),
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/armar/buscar-pedido", data={"cliente_id": "1"}, follow_redirects=False
        )

    assert respuesta.status_code == 303
    # Una sola revisión: la casilla activa del cliente 1 (misma conexión
    # que "Revisar ahora", solo lectura).
    mock_revisar.assert_called_once_with(
        "casilla@empresa.com", "clave", "imap.gmail.com",
        CASILLA_DE_PRUEBA["fecha_activacion"], "Pedido Dia", ["pedidos@dia.com.ar"],
    )
    # Manual: NO sella la revisión automática (la alerta no lo cuenta).
    mock_revision.assert_called_once_with(3)
    destino = respuesta.headers["location"]
    assert destino.startswith("/deposito/pedido/armar?")
    assert "cliente_id=1" in destino
    assert "1+mail+nuevo+por+confirmar" in destino


def test_buscar_pedido_sin_novedades_lo_dice_y_vuelve_a_donde_estaba():
    with (
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.revisar_casilla",
              return_value={"total_desde": 3, "candidatos": 0, "con_asunto": 0, "mails": []}),
        patch("app.main.registrar_revision_casilla"),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/armar/buscar-pedido",
            data={"cliente_id": "1", "fecha": "2026-08-22", "sucursal": "VL"},
            follow_redirects=False,
        )

    destino = respuesta.headers["location"]
    # Vuelve EXACTAMENTE a donde estaba: misma fecha y misma sucursal.
    assert "fecha=2026-08-22" in destino
    assert "sucursal=VL" in destino
    assert "sin+novedades" in destino


def test_buscar_pedido_sin_casilla_activa_del_cliente_avisa_sin_conectar():
    casillas = [dict(CASILLA_DE_PRUEBA, cliente_id=2), dict(CASILLA_DE_PRUEBA, id=5, activa=False)]
    with (
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.listar_casillas_pedidos", return_value=casillas),
        patch("app.main.revisar_casilla") as mock_revisar,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/armar/buscar-pedido", data={"cliente_id": "1"}, follow_redirects=False
        )

    assert respuesta.status_code == 303
    mock_revisar.assert_not_called()
    assert "ninguna+casilla+activa" in respuesta.headers["location"]


def test_buscar_pedido_con_error_imap_lo_registra_y_lo_muestra_en_armar():
    with (
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.revisar_casilla", side_effect=ErrorCasilla("IMAP deshabilitado")),
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/armar/buscar-pedido", data={"cliente_id": "1"}, follow_redirects=False
        )

    assert respuesta.status_code == 303
    # El error queda GRABADO en la casilla (alimenta la alerta), como siempre.
    mock_revision.assert_called_once_with(3, error="IMAP deshabilitado")
    destino = respuesta.headers["location"]
    assert destino.startswith("/deposito/pedido/armar?")
    assert "IMAP+deshabilitado" in destino


def test_revisar_mail_pendiente_precarga_la_revision_con_el_cuerpo_guardado():
    datos_leidos = {
        "bloques": [
            {
                "empresa": "9582 FRUTAMAX",
                "sucursales": [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 235}],
                "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 235}, "confianza": "alta"}],
            }
        ]
    }
    with (
        patch("app.main.obtener_mail_pedido", return_value=dict(MAIL_PEDIDO_DE_PRUEBA)),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos) as mock_extraer,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    # La IA lee el cuerpo YA guardado (pasado a texto), no vuelve al buzón.
    mock_extraer.assert_called_once_with(MAIL_PEDIDO_DE_PRUEBA["cuerpo_texto"])
    # Es la misma revisión de siempre, pero atada al mail: al guardar lo confirma.
    assert 'name="mail_id" value="9"' in respuesta.text
    assert "Leído del mail" in respuesta.text
    assert "pedidos@dia.com.ar" in respuesta.text
    # La fecha del pedido es el día del mail (22/08, hora argentina).
    assert 'name="fecha" value="2026-08-22"' in respuesta.text


def test_revisar_mail_ya_procesado_no_deja_confirmar_dos_veces():
    mail_confirmado = dict(MAIL_PEDIDO_DE_PRUEBA, estado="confirmado")
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail_confirmado),
        patch("app.main.extraer_pedido_de_texto") as mock_extraer,
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert "/sistema/casilla-pedidos" in respuesta.headers["location"]
    mock_extraer.assert_not_called()


def test_confirmar_pedido_desde_mail_lo_marca_confirmado_y_guarda_el_origen():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_mail_pedido", return_value=dict(MAIL_PEDIDO_DE_PRUEBA)),
        patch("app.main.obtener_pedido_vigente", return_value=None),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido", return_value=61) as mock_crear,
        patch("app.main.marcar_mail_pedido_confirmado") as mock_confirmar,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-22", "texto_original": "el cuerpo", "mail_id": "9",
                "cantidad_sucursales": "1", "sucursal_0_nombre": "VL", "sucursal_0_oc": "1257673", "sucursal_0_total": "235",
                "cantidad_renglones": "1",
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "235",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # El pedido nace con origen mail y su Message-ID (la idempotencia).
    assert mock_crear.call_args.args[2] == "mail"
    assert mock_crear.call_args.kwargs == {
        "reemplaza_a_pedido_id": None,
        "mail_message_id": "<pedido-1@dia.com.ar>",
        "recibido_el": MAIL_PEDIDO_DE_PRUEBA["recibido_el"],
    }
    mock_confirmar.assert_called_once_with(9, 61)


def test_confirmar_pedido_de_un_mail_ya_procesado_no_guarda_nada():
    mail_confirmado = dict(MAIL_PEDIDO_DE_PRUEBA, estado="confirmado")
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_mail_pedido", return_value=mail_confirmado),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-22", "texto_original": "el cuerpo", "mail_id": "9",
                "cantidad_sucursales": "1", "sucursal_0_nombre": "VL", "sucursal_0_oc": "", "sucursal_0_total": "",
                "cantidad_renglones": "1",
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "235",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "/sistema/casilla-pedidos" in respuesta.headers["location"]
    mock_crear.assert_not_called()


def test_ignorar_mail_pedido_lo_marca_y_redirige():
    with patch("app.main.marcar_mail_pedido_ignorado") as mock_ignorar:
        respuesta = cliente.post("/sistema/casilla-pedidos/mails/9/ignorar", follow_redirects=False)

    assert respuesta.status_code == 303
    mock_ignorar.assert_called_once_with(9, "Marcado a mano desde Sistema")


def test_leer_pedido_pegado_manda_a_la_ia_solo_el_bloque_de_esta_empresa():
    # El mail trae los dos bloques: a la IA va solo el de esta empresa (la
    # mitad de la salida y del tiempo); el texto_original guarda TODO igual.
    texto_completo = (
        "9582 FRUTAMAX\nCod\tProd\tVL\n90101\tBANANA\t225\n"
        "11344 PALMALA\nCod\tProd\tVL\n555\tOTRA COSA\t10\n"
    )
    datos_leidos = {
        "bloques": [{
            "empresa": "9582 FRUTAMAX",
            "sucursales": [{"sucursal": "VL", "orden_compra": None, "total_bultos": 225}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.parsear_pedido_estructurado", return_value=None),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos) as mock_extraer,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-21", "texto": texto_completo},
        )

    assert respuesta.status_code == 200
    texto_mandado = mock_extraer.call_args.args[0]
    assert "FRUTAMAX" in texto_mandado and "BANANA" in texto_mandado
    assert "PALMALA" not in texto_mandado
    # El respaldo completo sigue viajando entero a la revisión.
    assert "PALMALA" in respuesta.text


def test_revisar_mail_con_fallo_del_lector_graba_el_error_en_el_mail():
    # Cuando la revisión corra sola a las 12:00 esto va a pasar igual: la
    # falla queda en el mail (estado error + motivo) y alimenta la alerta
    # de Auditoría — no se pierde en un redirect que nadie vio.
    with (
        patch("app.main.obtener_mail_pedido", return_value=dict(MAIL_PEDIDO_DE_PRUEBA)),
        patch("app.main.extraer_pedido_de_texto", side_effect=RuntimeError("se cortó la lectura")),
        patch("app.main.marcar_mail_pedido_error") as mock_error,
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar", follow_redirects=False)

    assert respuesta.status_code == 303
    assert "No+se+pudo+leer+el+pedido" in respuesta.headers["location"]
    mock_error.assert_called_once_with(9, "La lectura falló: se cortó la lectura")


def test_revisar_mail_en_estado_error_se_puede_reintentar():
    mail_con_error = dict(MAIL_PEDIDO_DE_PRUEBA, estado="error", motivo="La lectura falló: se cortó")
    datos_leidos = {
        "bloques": [{
            "empresa": "9582 FRUTAMAX",
            "sucursales": [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 235}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 235}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail_con_error),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    assert 'name="mail_id" value="9"' in respuesta.text


def test_confirmar_pedido_de_un_mail_con_error_previo_lo_confirma_igual():
    # El reintento anduvo: el error previo no bloquea la confirmación.
    mail_con_error = dict(MAIL_PEDIDO_DE_PRUEBA, estado="error")
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_mail_pedido", return_value=mail_con_error),
        patch("app.main.obtener_pedido_vigente", return_value=None),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.crear_pedido", return_value=61) as mock_crear,
        patch("app.main.marcar_mail_pedido_confirmado") as mock_confirmar,
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/confirmar",
            data={
                "cliente_id": "1", "fecha": "2026-08-22", "texto_original": "el cuerpo", "mail_id": "9",
                "cantidad_sucursales": "1", "sucursal_0_nombre": "VL", "sucursal_0_oc": "", "sucursal_0_total": "",
                "cantidad_renglones": "1",
                "renglon_0_codigo": "90101", "renglon_0_descripcion": "BANANA",
                "renglon_0_ficha_id": "901", "renglon_0_cant_0": "235",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once()
    mock_confirmar.assert_called_once_with(9, 61)


def test_ver_casilla_muestra_las_acciones_en_un_mail_con_error():
    mail_con_error = dict(MAIL_PEDIDO_DE_PRUEBA, estado="error", motivo="La lectura falló: se cortó")
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.listar_mails_pedido", return_value=[mail_con_error]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    # El error se ve (pill + motivo) y el mail sigue accionable: reintentar o ignorar.
    assert ">Error<" in respuesta.text
    assert "La lectura falló: se cortó" in respuesta.text
    assert 'href="/deposito/pedido/mails/9/revisar"' in respuesta.text
    assert 'action="/sistema/casilla-pedidos/mails/9/ignorar"' in respuesta.text


TEXTO_PEDIDO_ESTRUCTURADO = (
    "9582 FRUTAMAX\n"
    "Código\tProducto\tVL\tBZ\n"
    "\tOC\t1257673\t1257642\n"
    "90101\tBANANA\t225\t\n"
    "90102\tBATATA\t\t40\n"
    "\tTOTAL BULTOS\t225\t40\n"
)


def test_leer_pedido_pegado_estructurado_no_llama_a_la_ia():
    # El camino principal: el parser por estructura lee la tabla ENTERA sin
    # IA — las cantidades por posición, la OC y el total por columna.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.extraer_pedido_de_texto") as mock_ia,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-22", "texto": TEXTO_PEDIDO_ESTRUCTURADO},
        )

    assert respuesta.status_code == 200
    mock_ia.assert_not_called()
    # El banner lo dice de un vistazo.
    assert "Leído por estructura" in respuesta.text
    # Banana matcheó por código de ficha; la OC y el total llegaron a la revisión.
    assert "OC 1257673" in respuesta.text
    assert "declara 225 bultos" in respuesta.text


def test_leer_pedido_pegado_sin_estructura_cae_a_ia_y_lo_dice():
    datos_leidos = {
        "bloques": [{
            "empresa": "",
            "sucursales": [{"sucursal": "VL", "orden_compra": None, "total_bultos": None}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos) as mock_ia,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-22", "texto": "pedido sin estructura 90101 BANANA 225"},
        )

    assert respuesta.status_code == 200
    mock_ia.assert_called_once()
    assert "Leído con IA — el parser de estructura no pudo" in respuesta.text


def test_revisar_mail_estructurado_lee_sin_ia_y_graba_el_metodo():
    mail = dict(MAIL_PEDIDO_DE_PRUEBA, cuerpo_texto=TEXTO_PEDIDO_ESTRUCTURADO)
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail),
        patch("app.main.extraer_pedido_de_texto") as mock_ia,
        patch("app.main.marcar_lectura_mail_pedido") as mock_marcar,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    mock_ia.assert_not_called()
    mock_marcar.assert_called_once_with(9, leido_con_ia=False)
    assert "Leído por estructura" in respuesta.text


def test_revisar_mail_sin_estructura_graba_el_fallback_a_ia():
    # El fallback VISIBLE: la marca alimenta la alerta "leídos con IA" — si
    # Día cambia el formato, se ve ese mismo día.
    datos_leidos = {
        "bloques": [{
            "empresa": "9582 FRUTAMAX",
            "sucursales": [{"sucursal": "VL", "orden_compra": None, "total_bultos": None}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 235}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main.obtener_mail_pedido", return_value=dict(MAIL_PEDIDO_DE_PRUEBA)),
        patch("app.main.extraer_pedido_de_texto", return_value=datos_leidos),
        patch("app.main.marcar_lectura_mail_pedido") as mock_marcar,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    mock_marcar.assert_called_once_with(9, leido_con_ia=True)
    assert "Leído con IA — el parser de estructura no pudo" in respuesta.text


def test_leer_pedido_desde_capturas_dice_que_fue_con_ia():
    datos_leidos = {
        "bloques": [{
            "empresa": "",
            "sucursales": [{"sucursal": "VL", "orden_compra": None, "total_bultos": None}],
            "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225}, "confianza": "alta"}],
        }]
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-comprimido"),
        patch("app.main.extraer_pedido_de_imagenes", return_value=datos_leidos),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.post(
            "/deposito/pedido/cargar/leer",
            data={"cliente_id": "1", "fecha": "2026-08-22"},
            files=[("imagenes", ("captura.jpg", b"bytes-de-imagen", "image/jpeg"))],
        )

    assert respuesta.status_code == 200
    assert "Leído con IA desde capturas" in respuesta.text


def test_revisar_mail_toma_la_fecha_del_asunto_no_de_la_llegada():
    # El mail del mediodía es para el DÍA SIGUIENTE: llegó el 22 con asunto
    # "23-08" -> el pedido es del 23. Si no, dos mails de días distintos
    # caerían en la misma fecha y el segundo pisaría al primero.
    mail = dict(MAIL_PEDIDO_DE_PRUEBA, asunto="Pedido Dia 23-08 Domingo",
                cuerpo_texto=TEXTO_PEDIDO_ESTRUCTURADO)
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail),
        patch("app.main.marcar_lectura_mail_pedido"),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None) as mock_vigente,
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    assert 'name="fecha" value="2026-08-23"' in respuesta.text
    # El "pedido vigente" (y el reemplazo del tramo 2) se compara contra la
    # fecha REAL del pedido, la del asunto.
    mock_vigente.assert_called_once_with(1, date(2026, 8, 23))
    assert "aviso" not in respuesta.text.split("Renglones")[0].lower() or "quedó con la fecha de llegada" not in respuesta.text


def test_revisar_mail_sin_fecha_en_el_asunto_usa_la_llegada_y_avisa():
    mail = dict(MAIL_PEDIDO_DE_PRUEBA, asunto="Pedido del dia", cuerpo_texto=TEXTO_PEDIDO_ESTRUCTURADO)
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail),
        patch("app.main.marcar_lectura_mail_pedido"),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    assert 'name="fecha" value="2026-08-22"' in respuesta.text
    assert "El asunto del mail no trae fecha" in respuesta.text
    assert "22/08/2026" in respuesta.text


def test_revisar_mail_con_fecha_del_asunto_lejana_avisa_posible_tipeo():
    mail = dict(MAIL_PEDIDO_DE_PRUEBA, asunto="Pedido Dia 15-09", cuerpo_texto=TEXTO_PEDIDO_ESTRUCTURADO)
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail),
        patch("app.main.marcar_lectura_mail_pedido"),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    # La fecha del asunto igual manda (15/09), pero con el aviso bien visible.
    assert 'name="fecha" value="2026-09-15"' in respuesta.text
    assert "puede ser un error de tipeo" in respuesta.text


def test_revisar_mail_reconvierte_el_crudo_html_y_lee_por_estructura():
    # El formato REAL de Día: colspan en la fila de empresa, columnas de
    # margen, OC sin etiqueta. El cuerpo_texto guardado está VIEJO (de una
    # conversión anterior): la relectura va al crudo con la conversión
    # vigente, así las mejoras aplican a los mails ya registrados.
    crudo = (
        '<div dir="ltr"><table><tbody>'
        "<tr><td></td><td></td><td></td><td></td><td></td><td>VL</td><td>BZ</td></tr>"
        "<tr><td></td><td></td><td></td><td></td><td></td><td>1257673</td><td>1257642</td></tr>"
        '<tr><td>9582</td><td colspan="2">FRUTAMAX</td><td></td><td></td><td>225</td><td>40</td></tr>'
        "<tr><td></td><td></td><td>90101</td><td>BANANA</td><td></td><td>225</td><td></td></tr>"
        "<tr><td></td><td></td><td>90102</td><td>BATATA</td><td></td><td></td><td>40</td></tr>"
        "</tbody></table></div>"
    )
    mail = dict(MAIL_PEDIDO_DE_PRUEBA, cuerpo_crudo=crudo, cuerpo_texto="foto vieja de la conversión")
    with (
        patch("app.main.obtener_mail_pedido", return_value=mail),
        patch("app.main.extraer_pedido_de_texto") as mock_ia,
        patch("app.main.marcar_lectura_mail_pedido") as mock_marcar,
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido/mails/9/revisar")

    assert respuesta.status_code == 200
    mock_ia.assert_not_called()
    mock_marcar.assert_called_once_with(9, leido_con_ia=False)
    assert "Leído por estructura" in respuesta.text
    assert "OC 1257673" in respuesta.text
    # Banana matcheó por el código de la ficha y la cantidad quedó en VL.
    assert "declara 225 bultos" in respuesta.text


LISTADO_PEDIDOS_DE_PRUEBA = [
    # Desordenados a propósito: el orden lo decide la pantalla.
    {"id": 48, "fecha_operacion": date(2026, 8, 20), "origen": "texto", "creado_en": datetime(2026, 8, 19, 12, 0),
     "renglones_totales": 30, "renglones_armados": 30, "sin_identificar": 0},
    {"id": 51, "fecha_operacion": date(2026, 8, 23), "origen": "mail", "creado_en": datetime(2026, 8, 22, 12, 0),
     "renglones_totales": 28, "renglones_armados": 0, "sin_identificar": 2},
    {"id": 50, "fecha_operacion": date(2026, 8, 22), "origen": "mail", "creado_en": datetime(2026, 8, 21, 12, 10),
     "renglones_totales": 32, "renglones_armados": 18, "sin_identificar": 0},
]


def test_armar_pedido_sin_fecha_lista_los_pedidos_hoy_primero():
    # El del sábado se puede ir armando el viernes: se listan TODOS los
    # vigentes (próximos y última semana), hoy primero, con su estado.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=list(LISTADO_PEDIDOS_DE_PRUEBA)) as mock_listar,
        patch("app.main.obtener_pedido_vigente") as mock_vigente,
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1")

    assert respuesta.status_code == 200
    # Ventana: últimos 7 días (los futuros van sin tope).
    mock_listar.assert_called_once_with(1, date(2026, 8, 15))
    mock_vigente.assert_not_called()
    texto = respuesta.text
    # Hoy primero, después el próximo, después el pasado.
    assert texto.index("Pedido del 22/08/2026 — HOY") < texto.index("Pedido del 23/08/2026 — próximo") < texto.index("Pedido del 20/08/2026")
    # El estado de un vistazo, sin entrar: armados, completo, sin identificar.
    assert "18 de 32 artículos armados" in texto
    assert "30 de 30 artículos armados · COMPLETO ✔" in texto
    assert "0 de 28 artículos armados" in texto and "2 sin identificar" in texto
    # Cada uno entra a su fecha.
    assert 'href="/deposito/pedido/armar?cliente_id=1&fecha=2026-08-23"' in texto


def test_armar_pedido_con_fecha_muestra_ese_pedido_con_su_fecha_visible():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert respuesta.status_code == 200
    # La fecha bien visible (se puede estar armando el de mañana) y la
    # vuelta al listado a mano.
    assert "Pedido del 21/08/2026 — elegí la sucursal" in respuesta.text
    assert 'href="/deposito/pedido/armar?cliente_id=1">Volver a los pedidos' in respuesta.text


def test_ver_pedido_lista_todos_los_cargados_ademas_del_abierto():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=list(LISTADO_PEDIDOS_DE_PRUEBA)),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1&fecha=2026-08-24")

    assert respuesta.status_code == 200
    texto = respuesta.text
    # Aunque la fecha abierta no tenga pedido, el listado muestra los que hay.
    assert "No hay pedido cargado para el 24/08/2026" in texto
    assert "Pedido del 22/08/2026 — HOY" in texto
    assert "Pedido del 23/08/2026 — próximo" in texto
    # El pasado completo sigue consultable, marcado, sin desaparecer.
    assert "Pedido del 20/08/2026" in texto and "COMPLETO ✔" in texto
    assert 'href="/deposito/pedido?cliente_id=1&fecha=2026-08-20"' in texto


# --- Etapa 3, tramo 2: revisión automática, alertas y auto-confirmar ---

from app.main import (  # noqa: E402
    _dias_esperados_como_numeros,
    _casilla_en_ventana,
    _le_toca_revision,
    _fechas_esperadas_sin_pedido,
    _intentar_auto_confirmar,
    contar_casillas_sin_revisar,
    contar_pedidos_faltantes,
    revisar_casillas_activas,
)

ARGENTINA_APP = timezone(timedelta(hours=-3))


def test_dias_esperados_como_numeros_ignora_basura():
    assert _dias_esperados_como_numeros("1,2,3,4,5,6") == {1, 2, 3, 4, 5, 6}
    assert _dias_esperados_como_numeros(" 1 , 7 ") == {1, 7}
    assert _dias_esperados_como_numeros("0,8,x,") == set()
    assert _dias_esperados_como_numeros(None) == set()


def test_fechas_esperadas_sin_pedido_no_cuenta_hoy_antes_de_las_15():
    # Viernes 21/08 a las 14:00: el mail de hoy todavía puede llegar — hoy
    # no cuenta como faltante. El jueves 20 sin pedido, sí.
    ahora = datetime(2026, 8, 21, 14, 0, tzinfo=ARGENTINA_APP)
    with (
        patch("app.main.listar_fechas_con_pedido_vigente", return_value=[date(2026, 8, 19)]),
        patch("app.main.listar_dias_sin_pedido", return_value=[]),
    ):
        resultado = _fechas_esperadas_sin_pedido(1, "1,2,3,4,5", ahora)

    assert date(2026, 8, 21) not in resultado["faltantes"]
    assert date(2026, 8, 20) in resultado["faltantes"]
    # El 19 tiene pedido vivo: no falta.
    assert date(2026, 8, 19) not in resultado["faltantes"]


def test_fechas_esperadas_sin_pedido_cuenta_hoy_desde_las_15():
    ahora = datetime(2026, 8, 21, 15, 0, tzinfo=ARGENTINA_APP)
    with (
        patch("app.main.listar_fechas_con_pedido_vigente", return_value=[]),
        patch("app.main.listar_dias_sin_pedido", return_value=[]),
    ):
        resultado = _fechas_esperadas_sin_pedido(1, "5", ahora)  # solo viernes

    assert resultado["faltantes"] == [date(2026, 8, 14), date(2026, 8, 21)]


def test_fechas_esperadas_sin_pedido_respeta_la_marca_y_el_pedido_manda():
    # El 20 está marcado "no hubo pedido": deja de faltar y aparece en
    # marcados. El 19 está marcado PERO después se cargó un pedido: el
    # pedido manda y la marca no se muestra (queda de registro).
    ahora = datetime(2026, 8, 21, 16, 0, tzinfo=ARGENTINA_APP)
    marcas = [
        {"id": 1, "fecha": date(2026, 8, 19), "motivo": None, "registrado_en": None},
        {"id": 2, "fecha": date(2026, 8, 20), "motivo": "Feriado", "registrado_en": None},
    ]
    with (
        patch("app.main.listar_fechas_con_pedido_vigente", return_value=[date(2026, 8, 19), date(2026, 8, 21)]),
        patch("app.main.listar_dias_sin_pedido", return_value=marcas),
    ):
        resultado = _fechas_esperadas_sin_pedido(1, "3,4,5", ahora)  # mié-jue-vie

    assert date(2026, 8, 20) not in resultado["faltantes"]
    assert [m["fecha"] for m in resultado["marcados"]] == [date(2026, 8, 20)]


def test_fechas_esperadas_sin_pedido_esporadico_no_alerta():
    ahora = datetime(2026, 8, 21, 16, 0, tzinfo=ARGENTINA_APP)
    with (
        patch("app.main.listar_fechas_con_pedido_vigente", return_value=[]),
        patch("app.main.listar_dias_sin_pedido", return_value=[]),
    ):
        resultado = _fechas_esperadas_sin_pedido(1, None, ahora)

    assert resultado["faltantes"] == []


def test_contar_pedidos_faltantes_suma_todos_los_clientes_con_dias():
    condiciones = [
        {"cliente_id": 1, "dias_esperados": "1,2,3,4,5", "cliente_nombre": "Día"},
        {"cliente_id": 2, "dias_esperados": "6", "cliente_nombre": "Otro"},
    ]

    def _faltantes(cliente_id, dias, ahora):
        if cliente_id == 1:
            return {"faltantes": [date(2026, 8, 20), date(2026, 8, 21)], "marcados": []}
        return {"faltantes": [date(2026, 8, 15)], "marcados": []}

    with (
        patch("app.main.listar_condiciones_pedido", return_value=condiciones),
        patch("app.main._fechas_esperadas_sin_pedido", side_effect=_faltantes),
    ):
        resultado = contar_pedidos_faltantes()

    assert resultado == {"casos": 3, "mas_viejo": date(2026, 8, 15)}


def test_contar_casillas_sin_revisar_no_alerta_por_un_fallo_que_se_recupero_solo():
    # Falló a las 12:00 pero a las 12:15 pudo: caída momentánea, el sistema
    # se recuperó solo. NO alerta — una alerta que grita por eso se deja de
    # mirar en una semana. El error igual queda visible en la pantalla de
    # la casilla.
    casillas = [
        dict(CASILLA_DE_PRUEBA, id=1, ultima_revision_el=datetime(2026, 8, 22, 12, 15, tzinfo=ARGENTINA_TEST),
             ultima_revision_automatica_el=datetime(2026, 8, 22, 12, 15, tzinfo=ARGENTINA_TEST),
             ultimo_error_el=datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_TEST)),
        # Y un error DESPUÉS del éxito de hoy tampoco alerta: los mails del
        # día ya se registraron a las 12:15.
        dict(CASILLA_DE_PRUEBA, id=2, ultima_revision_el=datetime(2026, 8, 22, 12, 15, tzinfo=ARGENTINA_TEST),
             ultima_revision_automatica_el=datetime(2026, 8, 22, 12, 15, tzinfo=ARGENTINA_TEST),
             ultimo_error_el=datetime(2026, 8, 22, 14, 15, tzinfo=ARGENTINA_TEST)),
    ]
    with patch("app.main.listar_casillas_pedidos", return_value=casillas):
        resultado = contar_casillas_sin_revisar(ahora=datetime(2026, 8, 22, 14, 30, tzinfo=ARGENTINA_APP))

    assert resultado == {"casos": 0, "mas_viejo": None}


def test_contar_casillas_sin_revisar_alerta_desde_las_14_si_hoy_nunca_pudo():
    casillas = [
        # Viene fallando todo el día (último éxito AYER): problema real.
        dict(CASILLA_DE_PRUEBA, id=1, ultima_revision_el=datetime(2026, 8, 21, 13, 0, tzinfo=ARGENTINA_TEST),
             ultima_revision_automatica_el=datetime(2026, 8, 21, 13, 0, tzinfo=ARGENTINA_TEST),
             ultimo_error_el=datetime(2026, 8, 22, 13, 45, tzinfo=ARGENTINA_TEST)),
        # Nunca corrió nada hoy (ni error registrado): también es problema —
        # el task puede estar muerto.
        dict(CASILLA_DE_PRUEBA, id=2, ultima_revision_el=None, ultimo_error_el=None),
        # Inactiva: nunca cuenta, aunque esté fallando.
        dict(CASILLA_DE_PRUEBA, id=3, activa=False, ultima_revision_el=None,
             ultimo_error_el=datetime(2026, 8, 22, 13, 0, tzinfo=ARGENTINA_TEST)),
    ]
    with patch("app.main.listar_casillas_pedidos", return_value=casillas):
        resultado = contar_casillas_sin_revisar(ahora=datetime(2026, 8, 22, 14, 0, tzinfo=ARGENTINA_APP))

    assert resultado == {"casos": 2, "mas_viejo": None}

    # A las 13:59 todavía NO, aunque venga fallando: quedan ticks por
    # delante y puede recuperarse solo. Avisar recién cuando "pasó un rato
    # largo y nunca pudo".
    with patch("app.main.listar_casillas_pedidos", return_value=casillas):
        assert contar_casillas_sin_revisar(
            ahora=datetime(2026, 8, 22, 13, 59, tzinfo=ARGENTINA_APP)
        ) == {"casos": 0, "mas_viejo": None}


def test_contar_casillas_sin_revisar_respeta_el_horario_configurado():
    # Casilla con ventana propia 09:00-18:00: el umbral de alerta es 17:00
    # (una hora antes de SU cierre), no las 14:00 del default.
    casillas = [
        dict(CASILLA_DE_PRUEBA, id=1, ultima_revision_el=None, ultimo_error_el=None,
             revision_desde=time(9, 0), revision_hasta=time(18, 0), revision_cada_minutos=30),
    ]
    with patch("app.main.listar_casillas_pedidos", return_value=casillas):
        antes = contar_casillas_sin_revisar(ahora=datetime(2026, 8, 22, 16, 59, tzinfo=ARGENTINA_APP))
        despues = contar_casillas_sin_revisar(ahora=datetime(2026, 8, 22, 17, 0, tzinfo=ARGENTINA_APP))

    assert antes == {"casos": 0, "mas_viejo": None}
    assert despues == {"casos": 1, "mas_viejo": None}


def test_editar_cliente_muestra_los_dias_de_pedido_tildados():
    with (
        patch("app.main.obtener_cliente", return_value=CLIENTE_DE_PRUEBA),
        patch("app.main.listar_conceptos_editables_por_cliente", return_value=CONCEPTOS_EDITABLES_DE_PRUEBA),
        patch("app.main.obtener_condiciones_pedido", return_value={"cliente_id": 1, "dias_esperados": "1,4"}),
    ):
        respuesta = cliente.get("/clientes/1/editar")

    assert respuesta.status_code == 200
    assert "Días de pedido" in respuesta.text
    assert 'action="/clientes/1/dias-pedido"' in respuesta.text
    assert 'value="1" checked' in respuesta.text
    assert 'value="4" checked' in respuesta.text
    assert 'value="2" checked' not in respuesta.text


def test_guardar_dias_pedido_ordena_y_filtra_los_dias():
    with patch("app.main.guardar_condiciones_pedido") as mock_guardar:
        respuesta = cliente.post(
            "/clientes/1/dias-pedido",
            data={"dia": ["6", "1", "3", "9", "x"]},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/clientes/1/editar"
    mock_guardar.assert_called_once_with(1, "1,3,6")


def test_guardar_dias_pedido_sin_dias_deja_al_cliente_esporadico():
    with patch("app.main.guardar_condiciones_pedido") as mock_guardar:
        respuesta = cliente.post("/clientes/1/dias-pedido", data={}, follow_redirects=False)

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(1, None)


def test_ver_pedido_muestra_los_dias_faltantes_con_sus_dos_cierres():
    estado_dias = {
        "faltantes": [date(2026, 8, 20)],
        "marcados": [{"id": 2, "fecha": date(2026, 8, 19), "motivo": "Feriado", "registrado_en": None}],
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value={"cliente_id": 1, "dias_esperados": "3,4,5"}),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main._fechas_esperadas_sin_pedido", return_value=estado_dias),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    texto = respuesta.text
    # El faltante, con sus DOS cierres: cargar el pedido o marcar que no hubo.
    assert "Falta el pedido del 20/08/2026" in texto
    assert 'href="/deposito/pedido/cargar?cliente_id=1&fecha=2026-08-20"' in texto
    assert 'action="/deposito/pedido/dias-sin-pedido"' in texto
    assert "No hubo pedido" in texto
    # El marcado, atenuado con su motivo y el deshacer.
    assert "19/08/2026 — no hubo pedido (Feriado)" in texto
    assert 'action="/deposito/pedido/dias-sin-pedido/deshacer"' in texto
    assert "Deshacer" in texto


def test_marcar_dia_sin_pedido_guarda_con_motivo_opcional():
    with patch("app.main.marcar_dia_sin_pedido") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/dias-sin-pedido",
            data={"cliente_id": "1", "fecha": "2026-08-20", "motivo": "  Feriado  "},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(1, date(2026, 8, 20), "Feriado")


def test_deshacer_dia_sin_pedido_borra_la_marca():
    with patch("app.main.borrar_dia_sin_pedido") as mock_borrar:
        respuesta = cliente.post(
            "/deposito/pedido/dias-sin-pedido/deshacer",
            data={"cliente_id": "1", "fecha": "2026-08-20"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_borrar.assert_called_once_with(1, date(2026, 8, 20))


def test_ver_pedido_avisa_cuando_se_confirmo_automaticamente():
    pedido_de_mail = dict(PEDIDO_VIGENTE_DE_PRUEBA, origen="mail")
    mail_del_pedido = {
        "id": 9, "remitente": "pedidos@dia.com.ar", "asunto": "Pedido Dia 21-08",
        "recibido_el": datetime(2026, 8, 20, 12, 5, tzinfo=ARGENTINA_TEST),
        "motivo": "Confirmado automáticamente",
        "procesado_el": datetime(2026, 8, 20, 12, 10, tzinfo=ARGENTINA_TEST),
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=pedido_de_mail),
        patch("app.main.obtener_mail_de_pedido", return_value=mail_del_pedido),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_PEDIDO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fotos_pedido", return_value=[]),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Confirmado automáticamente" in respuesta.text
    assert "Nadie lo revisó a mano" in respuesta.text


def test_ver_pedido_confirmado_a_mano_desde_mail_no_muestra_el_aviso_auto():
    pedido_de_mail = dict(PEDIDO_VIGENTE_DE_PRUEBA, origen="mail")
    mail_del_pedido = {
        "id": 9, "remitente": "pedidos@dia.com.ar", "asunto": "Pedido Dia 21-08",
        "recibido_el": datetime(2026, 8, 20, 12, 5, tzinfo=ARGENTINA_TEST),
        "motivo": None,
        "procesado_el": datetime(2026, 8, 20, 12, 10, tzinfo=ARGENTINA_TEST),
    }
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=pedido_de_mail),
        patch("app.main.obtener_mail_de_pedido", return_value=mail_del_pedido),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_PEDIDO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fotos_pedido", return_value=[]),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_PEDIDO_DE_PRUEBA),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    assert "Nadie lo revisó a mano" not in respuesta.text


# --- Auto-confirmar: los candados, uno por uno ---

MAIL_AUTO_DE_PRUEBA = dict(
    MAIL_PEDIDO_DE_PRUEBA,
    asunto="Pedido Dia 23-08 Domingo",
    cuerpo_texto=TEXTO_PEDIDO_ESTRUCTURADO,
)


def _patches_auto_confirmar(**overrides):
    """Los patches del camino FELIZ del auto-confirmar; cada test rompe un candado."""
    valores = {
        "listar_fichas_por_cliente": {"return_value": FICHAS_PEDIDO_DE_PRUEBA},
        "obtener_pedido_vigente": {"return_value": None},
        "crear_pedido": {"return_value": 77},
        "marcar_mail_pedido_confirmado": {},
        "marcar_lectura_mail_pedido": {},
    }
    valores.update(overrides)
    return {nombre: patch(f"app.main.{nombre}", **kwargs) for nombre, kwargs in valores.items()}


def test_auto_confirmar_camino_feliz_crea_el_pedido_y_confirma_el_mail():
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"] as mock_vigente,
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"] as mock_confirmar,
        patches["marcar_lectura_mail_pedido"] as mock_lectura,
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

    assert resultado is True
    # La fecha sale del ASUNTO (llegó el 22, es para el 23).
    mock_vigente.assert_called_once_with(1, date(2026, 8, 23))
    args = mock_crear.call_args
    assert args.args[0] == 1
    assert args.args[1] == date(2026, 8, 23)
    assert args.args[2] == "mail"
    # Renglones expandidos por sucursal con cantidad, identificados.
    #
    # Se compara el renglón COMPLETO, no tres campos elegidos. La versión
    # vieja miraba (sucursal, articulo_id, cantidad) y por eso pasó NUEVE
    # DÍAS en verde mientras el auto-confirmado guardaba todos los pedidos
    # sin ficha_id (27/08 al 04/09/2026): el campo que faltaba no estaba
    # entre los que el test miraba. Si mañana el renglón gana un campo,
    # este test tiene que fallar hasta que alguien decida qué va ahí.
    renglones = args.args[5]
    assert sorted(renglones, key=lambda r: r["sucursal"]) == [
        {
            "sucursal": "BZ",
            "articulo_id": 2,
            "ficha_id": 902,
            "texto_codigo": "90102",
            "texto_descripcion": "BATATA",
            "cantidad": 40.0,
        },
        {
            "sucursal": "VL",
            "articulo_id": 1,
            "ficha_id": 901,
            "texto_codigo": "90101",
            "texto_descripcion": "BANANA",
            "cantidad": 225.0,
        },
    ]
    assert args.kwargs["mail_message_id"] == MAIL_AUTO_DE_PRUEBA["message_id"]
    mock_confirmar.assert_called_once_with(9, 77, motivo="Confirmado automáticamente")
    mock_lectura.assert_called_once_with(9, leido_con_ia=False)


def test_auto_confirmar_no_toca_un_mail_con_error_previo():
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA, estado="error"))

    assert resultado is False
    mock_crear.assert_not_called()


def test_auto_confirmar_sin_fecha_en_el_asunto_no_adivina():
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA, asunto="Pedido del dia"))

    assert resultado is False
    mock_crear.assert_not_called()


def test_auto_confirmar_con_fecha_lejana_del_asunto_no_confirma():
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA, asunto="Pedido Dia 15-10"))

    assert resultado is False
    mock_crear.assert_not_called()


def test_auto_confirmar_jamas_usa_la_ia_si_el_parser_no_puede():
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"] as mock_lectura,
        patch("app.main.extraer_pedido_de_texto") as mock_ia,
    ):
        resultado = _intentar_auto_confirmar(
            dict(MAIL_AUTO_DE_PRUEBA, cuerpo_texto="pedido sin estructura 90101 BANANA 225")
        )

    assert resultado is False
    mock_ia.assert_not_called()
    mock_crear.assert_not_called()
    # El mail queda PENDIENTE sin marca de lectura: lo va a leer una persona.
    mock_lectura.assert_not_called()


def test_auto_confirmar_con_un_renglon_sin_identificar_no_confirma():
    solo_banana = [f for f in FICHAS_PEDIDO_DE_PRUEBA if f["articulo_id"] == 1]
    patches = _patches_auto_confirmar(listar_fichas_por_cliente={"return_value": solo_banana})
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
        patch("app.main.adivinar_articulo", return_value=None),
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

    assert resultado is False
    mock_crear.assert_not_called()


def test_auto_confirmar_con_match_por_sugerencia_no_confirma():
    # La sugerencia difusa alcanza para PRECARGAR la revisión, nunca para
    # confirmar solo: elegir es una decisión humana.
    solo_banana = [f for f in FICHAS_PEDIDO_DE_PRUEBA if f["articulo_id"] == 1]
    patches = _patches_auto_confirmar(listar_fichas_por_cliente={"return_value": solo_banana})
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
        patch("app.main.adivinar_articulo", return_value=2),
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

    assert resultado is False
    mock_crear.assert_not_called()


def test_auto_confirmar_ya_no_frena_por_totales_declarados():
    # El candado de "sumas exactas contra el total declarado" SE SACÓ
    # (24/08/2026): Día se equivoca sumando sus propios totales, y contra
    # lecturas rotas la red es el parser por estructura. Sin total
    # declarado, con descuadre, o con una sucursal solo en cantidades: si
    # los CINCO candados restantes cierran, se confirma igual.
    casos = [
        # Sin total declarado.
        [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": None}],
        # Total declarado que NO cuadra con la suma (Día sumó mal: 300 vs 225).
        [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 300}],
    ]
    for sucursales_bloque in casos:
        datos = {
            "bloques": [{
                "empresa": "9582 FRUTAMAX",
                "sucursales": sucursales_bloque,
                "renglones": [{"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225}, "confianza": "alta"}],
            }]
        }
        patches = _patches_auto_confirmar()
        with (
            patches["listar_fichas_por_cliente"],
            patches["obtener_pedido_vigente"],
            patches["crear_pedido"] as mock_crear,
            patches["marcar_mail_pedido_confirmado"] as mock_confirmar,
            patches["marcar_lectura_mail_pedido"],
            patch("app.main.parsear_pedido_estructurado", return_value=datos),
        ):
            resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

        assert resultado is True, sucursales_bloque
        mock_crear.assert_called_once()
        # El total declarado se sigue GUARDANDO como dato (viaja en las
        # sucursales del pedido): solo dejó de frenar.
        sucursales_guardadas = mock_crear.call_args.args[4]
        assert sucursales_guardadas[0]["total_bultos_declarado"] == sucursales_bloque[0]["total_bultos"]
        mock_confirmar.assert_called_once_with(9, 77, motivo="Confirmado automáticamente")


def test_auto_confirmar_con_sucursal_solo_en_cantidades_confirma_igual():
    # Una sucursal que aparece en cantidades pero no vino declarada arriba
    # ya no frena: se agrega sin OC ni total (nada se pierde) y se confirma.
    datos = {
        "bloques": [{
            "empresa": "9582 FRUTAMAX",
            "sucursales": [{"sucursal": "VL", "orden_compra": "1257673", "total_bultos": 225}],
            "renglones": [
                {"codigo": "90101", "descripcion": "BANANA", "cantidades": {"VL": 225, "GR": 10}, "confianza": "alta"},
            ],
        }]
    }
    patches = _patches_auto_confirmar()
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"],
        patches["marcar_lectura_mail_pedido"],
        patch("app.main.parsear_pedido_estructurado", return_value=datos),
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

    assert resultado is True
    sucursales_guardadas = mock_crear.call_args.args[4]
    assert {s["sucursal"] for s in sucursales_guardadas} == {"VL", "GR"}


def test_auto_confirmar_no_reemplaza_un_pedido_vigente():
    patches = _patches_auto_confirmar(obtener_pedido_vigente={"return_value": PEDIDO_VIGENTE_DE_PRUEBA})
    with (
        patches["listar_fichas_por_cliente"],
        patches["obtener_pedido_vigente"],
        patches["crear_pedido"] as mock_crear,
        patches["marcar_mail_pedido_confirmado"] as mock_confirmar,
        patches["marcar_lectura_mail_pedido"],
    ):
        resultado = _intentar_auto_confirmar(dict(MAIL_AUTO_DE_PRUEBA))

    assert resultado is False
    mock_crear.assert_not_called()
    mock_confirmar.assert_not_called()


# --- El tick de la revisión automática ---

MAILS_DEL_BUZON_DE_PRUEBA = [
    {
        "message_id": "<pedido-n@dia.com.ar>", "remitente": "pedidos@dia.com.ar",
        "asunto": "Pedido Dia 23-08", "recibido_el": datetime(2026, 8, 22, 12, 5, tzinfo=ARGENTINA_TEST),
        "cuerpo_crudo": "<html>...</html>", "cuerpo_texto": "9582 FRUTAMAX...",
    },
    {
        "message_id": "<pedido-viejo@dia.com.ar>", "remitente": "pedidos@dia.com.ar",
        "asunto": "Pedido Dia 22-08", "recibido_el": datetime(2026, 8, 21, 12, 5, tzinfo=ARGENTINA_TEST),
        "cuerpo_crudo": "<html>...</html>", "cuerpo_texto": "9582 FRUTAMAX...",
    },
]


def test_tick_revisa_las_activas_y_auto_confirma_solo_los_nuevos():
    casilla_auto = dict(CASILLA_DE_PRUEBA, auto_confirmar=True)
    casilla_inactiva = dict(CASILLA_DE_PRUEBA, id=4, activa=False)
    resultado_buzon = {"total_desde": 2, "candidatos": 2, "con_asunto": 2, "mails": MAILS_DEL_BUZON_DE_PRUEBA}
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla_auto, casilla_inactiva]),
        patch("app.main.clave_casilla_configurada", return_value="clave-app"),
        patch("app.main.revisar_casilla", return_value=resultado_buzon) as mock_revisar,
        # El primero es NUEVO (id 11); el segundo ya estaba registrado (None).
        patch("app.main.registrar_mail_pedido", side_effect=[11, None]),
        patch("app.main.registrar_revision_casilla") as mock_revision,
        patch("app.main.obtener_mail_pedido", return_value=dict(MAIL_AUTO_DE_PRUEBA, id=11)) as mock_obtener,
        patch("app.main._intentar_auto_confirmar", return_value=True) as mock_auto,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST))

    # Solo la activa se revisó (una conexión, no dos).
    mock_revisar.assert_called_once()
    mock_revision.assert_called_once_with(3, automatica=True)
    # Auto-confirmar SOLO el mail nuevo, no el ya registrado.
    mock_obtener.assert_called_once_with(11)
    mock_auto.assert_called_once()


def test_tick_sin_auto_confirmar_solo_registra():
    resultado_buzon = {"total_desde": 1, "candidatos": 1, "con_asunto": 1, "mails": MAILS_DEL_BUZON_DE_PRUEBA[:1]}
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.clave_casilla_configurada", return_value="clave-app"),
        patch("app.main.revisar_casilla", return_value=resultado_buzon),
        patch("app.main.registrar_mail_pedido", return_value=11),
        patch("app.main.registrar_revision_casilla"),
        patch("app.main._intentar_auto_confirmar") as mock_auto,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST))

    mock_auto.assert_not_called()


def test_tick_sin_clave_registra_el_error_en_la_casilla():
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.clave_casilla_configurada", return_value=None),
        patch("app.main.revisar_casilla") as mock_revisar,
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST))

    mock_revisar.assert_not_called()
    args = mock_revision.call_args
    assert args.args == (3,)
    assert "CLAVE_CASILLA_PEDIDOS" in args.kwargs["error"]


def test_tick_con_error_de_buzon_lo_registra_y_sigue():
    # El buzón falla corriendo solo a las 12: el error queda GRABADO en la
    # casilla (alimenta la alerta de Auditoría), no en un log que nadie ve.
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.clave_casilla_configurada", return_value="clave-app"),
        patch("app.main.revisar_casilla", side_effect=ErrorCasilla("No se pudo entrar al buzón")),
        patch("app.main.registrar_revision_casilla") as mock_revision,
        patch("app.main._intentar_auto_confirmar") as mock_auto,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST))

    args = mock_revision.call_args
    assert args.args == (3,)
    assert "No se pudo entrar al buzón" in args.kwargs["error"]
    mock_auto.assert_not_called()


def test_tick_una_casilla_sin_fecha_de_activacion_se_saltea():
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA, fecha_activacion=None)]),
        patch("app.main.clave_casilla_configurada", return_value="clave-app"),
        patch("app.main.revisar_casilla") as mock_revisar,
        patch("app.main.registrar_revision_casilla") as mock_revision,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 12, 30, tzinfo=ARGENTINA_TEST))

    mock_revisar.assert_not_called()
    mock_revision.assert_not_called()


def test_ventana_de_revision_default_es_de_12_a_15_argentina():
    # Sin horario configurado rigen los defaults de siempre: 12:00 a 15:00.
    casilla = dict(CASILLA_DE_PRUEBA)
    assert not _casilla_en_ventana(casilla, datetime(2026, 8, 22, 11, 59, tzinfo=ARGENTINA_APP))
    assert _casilla_en_ventana(casilla, datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_APP))
    assert _casilla_en_ventana(casilla, datetime(2026, 8, 22, 14, 59, tzinfo=ARGENTINA_APP))
    assert not _casilla_en_ventana(casilla, datetime(2026, 8, 22, 15, 0, tzinfo=ARGENTINA_APP))


def test_ventana_de_revision_configurada_por_casilla():
    casilla = dict(CASILLA_DE_PRUEBA, revision_desde=time(9, 30), revision_hasta=time(18, 0))
    assert not _casilla_en_ventana(casilla, datetime(2026, 8, 22, 9, 29, tzinfo=ARGENTINA_APP))
    assert _casilla_en_ventana(casilla, datetime(2026, 8, 22, 9, 30, tzinfo=ARGENTINA_APP))
    assert _casilla_en_ventana(casilla, datetime(2026, 8, 22, 17, 59, tzinfo=ARGENTINA_APP))
    assert not _casilla_en_ventana(casilla, datetime(2026, 8, 22, 18, 0, tzinfo=ARGENTINA_APP))


def test_le_toca_revision_respeta_la_cadencia_configurada():
    # Cada 30 minutos: a los 29 del último intento no toca, a los 30 sí.
    # Un ERROR también cuenta como intento (el reintento espera su turno).
    casilla = dict(
        CASILLA_DE_PRUEBA, revision_cada_minutos=30,
        ultima_revision_el=datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_TEST),
        ultimo_error_el=datetime(2026, 8, 22, 12, 10, tzinfo=ARGENTINA_TEST),
    )
    assert not _le_toca_revision(casilla, datetime(2026, 8, 22, 12, 39, tzinfo=ARGENTINA_TEST))
    assert _le_toca_revision(casilla, datetime(2026, 8, 22, 12, 40, tzinfo=ARGENTINA_TEST))
    # Sin ningún intento previo, toca ya mismo.
    assert _le_toca_revision(dict(CASILLA_DE_PRUEBA), datetime(2026, 8, 22, 12, 0, tzinfo=ARGENTINA_TEST))


def test_tick_no_revisa_fuera_de_la_ventana_de_la_casilla_ni_antes_de_la_cadencia():
    # A las 16:00 la casilla default (12-15) queda afuera; la de horario
    # extendido se revisa. Y la que ya se revisó hace 5 minutos espera.
    casilla_default = dict(CASILLA_DE_PRUEBA, id=1)
    casilla_extendida = dict(CASILLA_DE_PRUEBA, id=2, revision_hasta=time(18, 0))
    casilla_recien_revisada = dict(
        CASILLA_DE_PRUEBA, id=4, revision_hasta=time(18, 0),
        ultima_revision_el=datetime(2026, 8, 22, 15, 55, tzinfo=ARGENTINA_TEST),
    )
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla_default, casilla_extendida, casilla_recien_revisada]),
        patch("app.main._revisar_casilla_automaticamente") as mock_revisar,
    ):
        revisar_casillas_activas(ahora=datetime(2026, 8, 22, 16, 0, tzinfo=ARGENTINA_TEST))

    assert [llamada.args[0]["id"] for llamada in mock_revisar.call_args_list] == [2]


# --- Rentabilidad de Pedidos (Gerencia) ---

RENGLONES_RENTABILIDAD_DE_PRUEBA = [
    {"fecha_operacion": date(2026, 8, 21), "ficha_id": 901, "articulo_id": 1,
     "articulo_nombre": "Banana", "articulo_grupo": "fruta", "bultos": 10.0},
    {"fecha_operacion": date(2026, 8, 22), "ficha_id": 901, "articulo_id": 1,
     "articulo_nombre": "Banana", "articulo_grupo": "fruta", "bultos": 5.0},
    {"fecha_operacion": date(2026, 8, 22), "ficha_id": None, "articulo_id": None,
     "articulo_nombre": None, "articulo_grupo": None, "bultos": 3.0},
]

FICHAS_RENTABILIDAD_DE_PRUEBA = [
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Banana", "articulo_grupo": "fruta",
     "envase_id": None, "envase_nombre": None, "contenido_caja": 20.0, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "90101"},
]


def _patches_rentabilidad():
    # La fila de Márgenes por Artículo que Rentabilidad consume tal cual:
    # lista $100, denominador de tasas 0.8 (neta $80), costo $60 + $4 de
    # envase por kilo.
    fila_margenes = {
        "ficha_id": 901, "articulo_id": 1, "articulo_nombre": "Banana", "unidad_venta": "kilo",
        "precio_vigente": 100.0, "costo_actual": 60.0,
        "costo_envase_unidad_venta": 4.0, "denominador_tasas": 0.8,
        "utilidad_aproximada": 0.2666, "precio_sugerido": None, "fresco": True,
    }
    return (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_RENTABILIDAD_DE_PRUEBA),
        patch("app.main.listar_renglones_pedidos_vigentes", return_value=list(RENGLONES_RENTABILIDAD_DE_PRUEBA)),
        patch("app.main.calcular_listados_para_negociar_precios",
              side_effect=lambda cliente_id, momentos: {m: [dict(fila_margenes)] for m in momentos}),
    )


def test_ver_rentabilidad_sin_cliente_muestra_solo_el_selector():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_renglones_pedidos_vigentes") as mock_renglones,
    ):
        respuesta = cliente.get("/gerencia/rentabilidad")

    assert respuesta.status_code == 200
    assert "Elegí un cliente" in respuesta.text
    mock_renglones.assert_not_called()
    # La regla de la medición SIEMPRE a la vista.
    assert "lo pedido" in respuesta.text


def test_ver_rentabilidad_usa_el_listado_de_margenes_por_fecha():
    p1, p2, p3, p4, p5 = _patches_rentabilidad()
    with p1, p2, p3, p4, p5 as mock_margenes:
        respuesta = cliente.get("/gerencia/rentabilidad?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")

    assert respuesta.status_code == 200
    # El listado de Márgenes por Artículo sigue anclado a CADA fecha con
    # pedido (nunca el margen de hoy retroactivo) — pero se pide UNA sola
    # vez con todas las fechas juntas, no una llamada por fecha.
    mock_margenes.assert_called_once()
    assert mock_margenes.call_args.args[0] == 1
    fechas_margenes = sorted(momento.date() for momento in mock_margenes.call_args.args[1])
    assert fechas_margenes == [date(2026, 8, 21), date(2026, 8, 22)]

    texto = respuesta.text
    assert "Fruta" in texto
    assert "Banana" in texto
    # 15 bultos × 20k = 300k. Venta NETA (con tasas, no la lista cruda):
    # 300 × $80 = 24.000. Costo: 300 × ($60 + $4 envase) = 19.200. Renta
    # 4.800, utilidad SOLO sobre mercadería: 4.800 / 18.000 = 26.7%.
    assert "$24.000" in texto
    assert "$19.200" in texto
    assert "$4.800" in texto
    assert "26.7%" in texto
    # La lista cruda ($100) y la neta ($80) se ven por unidad, no se mezclan.
    assert "lista $100" in texto
    assert "neta $80" in texto
    # El sin identificar AFUERA, con su peso — no sumó como cero.
    assert "Afuera del cálculo (1)" in texto
    assert "Sin identificar — 3 bultos" in texto
    # Export con los mismos filtros.
    assert "/gerencia/rentabilidad/exportar-pdf?cliente_id=1" in texto
    assert "/gerencia/rentabilidad/exportar-excel?cliente_id=1" in texto


def test_exportar_rentabilidad_pdf_y_excel_bajan_con_sus_nombres():
    p1, p2, p3, p4, p5 = _patches_rentabilidad()
    with p1, p2, p3, p4, p5, patch("app.main.obtener_cliente", return_value=dict(CLIENTE_DE_PRUEBA)):
        pdf = cliente.get("/gerencia/rentabilidad/exportar-pdf?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")

    assert pdf.status_code == 200
    assert pdf.headers["content-type"] == "application/pdf"
    assert 'filename="Rentabilidad_Pedidos_2026-08-15_a_2026-08-22.pdf"' in pdf.headers["content-disposition"]
    assert pdf.content.startswith(b"%PDF")

    p1, p2, p3, p4, p5 = _patches_rentabilidad()
    with p1, p2, p3, p4, p5, patch("app.main.obtener_cliente", return_value=dict(CLIENTE_DE_PRUEBA)):
        excel = cliente.get("/gerencia/rentabilidad/exportar-excel?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")

    assert excel.status_code == 200
    assert "spreadsheetml" in excel.headers["content-type"]
    assert 'filename="Rentabilidad_Pedidos_2026-08-15_a_2026-08-22.xlsx"' in excel.headers["content-disposition"]


def test_exportar_rentabilidad_sin_cliente_da_400():
    respuesta = cliente.get("/gerencia/rentabilidad/exportar-pdf?fecha_desde=2026-08-15&fecha_hasta=2026-08-22")
    assert respuesta.status_code == 400


def test_ver_rentabilidad_fechas_invalidas_avisa_sin_calcular():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_renglones_pedidos_vigentes") as mock_renglones,
    ):
        respuesta = cliente.get("/gerencia/rentabilidad?cliente_id=1&fecha_desde=2026-08-25&fecha_hasta=2026-08-22")

    assert respuesta.status_code == 200
    assert "La fecha desde no puede ser posterior a la fecha hasta." in respuesta.text
    mock_renglones.assert_not_called()


def test_gerencia_tiene_el_acceso_a_rentabilidad():
    respuesta = cliente.get("/gerencia")
    assert respuesta.status_code == 200
    assert 'href="/gerencia/rentabilidad"' in respuesta.text


# --- Adjuntar comanda a la carga manual (foto sin analizar) ---

def test_agregar_compra_manual_con_comanda_adjunta_la_cuelga_de_la_guia():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-chico") as mock_comprimir,
        patch("app.main.subir_foto_comanda", return_value="2026-08-22/n07p41-abc.jpg") as mock_subir,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data=_datos_compra_manual(),
            files={"comanda_foto": ("comanda.jpg", b"bytes-originales", "image/jpeg")},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # SIEMPRE comprimida antes de subir (baja resolución, nunca el original).
    mock_comprimir.assert_called_once_with(b"bytes-originales")
    mock_subir.assert_called_once_with(b"jpeg-chico", "N07P41")
    # La ruta viaja a crear_compra: la foto cuelga de la guía del día.
    assert mock_crear.call_args.args[10] == "2026-08-22/n07p41-abc.jpg"


def test_agregar_compra_manual_foto_ilegible_avisa_sin_crear_nada():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)) as mock_proveedor,
        patch("app.main._comprimir_foto_jpeg", return_value=None),
        patch("app.main.crear_compra") as mock_crear,
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data=_datos_compra_manual(),
            files={"comanda_foto": ("comanda.jpg", b"no-es-imagen", "image/jpeg")},
        )

    assert respuesta.status_code == 400
    assert "no es una imagen legible" in respuesta.text
    mock_proveedor.assert_not_called()
    mock_crear.assert_not_called()


def test_agregar_compra_manual_si_storage_falla_la_compra_se_guarda_igual():
    # La foto es un extra: un fallo de Storage jamás pierde la compra.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-chico"),
        patch("app.main.subir_foto_comanda", side_effect=RuntimeError("sin conexión")),
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data=_datos_compra_manual(),
            files={"comanda_foto": ("comanda.jpg", b"bytes", "image/jpeg")},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert mock_crear.call_args.args[10] is None


def test_agregar_compra_manual_error_con_foto_avisa_que_hay_que_readjuntarla():
    # El navegador no repuebla un input de archivo: el error lo dice.
    with (
        patch("app.main.listar_proveedores", return_value=[]),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data=_datos_compra_manual(codigo_puesto="puesto15"),
            files={"comanda_foto": ("comanda.jpg", b"bytes", "image/jpeg")},
        )

    assert respuesta.status_code == 400
    assert "volvé a adjuntarla" in respuesta.text


def test_agregar_compra_clasica_con_comanda_adjunta_la_cuelga_de_la_guia():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.obtener_articulo", return_value=ARTICULO_KILO_DE_PRUEBA),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-chico"),
        patch("app.main.subir_foto_comanda", return_value="2026-08-06/n07p41-abc.jpg") as mock_subir,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={"proveedor_id": "200", "accion": "agregar", "articulo_id": "5", "cantidad_cajones": "10",
                  "contenido_por_cajon": "18", "importe": "50000", "sena": "", "tipo_retiro": "Clark"},
            files={"comanda_foto": ("comanda.jpg", b"bytes", "image/jpeg")},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_subir.assert_called_once_with(b"jpeg-chico", "N07P41")
    assert mock_crear.call_args.args[10] == "2026-08-06/n07p41-abc.jpg"


def test_terminar_solo_con_la_comanda_adjunta_la_cuelga_y_cierra():
    # El "cerrar la comanda": todos los artículos ya cargados, se adjunta
    # la foto al final y Guardar la cuelga de la guía del día — sin
    # renglón nuevo.
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-chico"),
        patch("app.main.subir_foto_comanda", return_value="2026-08-06/n07p41-abc.jpg"),
        patch("app.main.agregar_foto_guia_del_dia", return_value=True) as mock_colgar,
        patch("app.main.crear_compra") as mock_crear,
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={"proveedor_id": "200", "accion": "terminar", "articulo_id": "", "cantidad_cajones": "",
                  "contenido_por_cajon": "", "importe": "", "sena": "", "tipo_retiro": ""},
            files={"comanda_foto": ("comanda.jpg", b"bytes", "image/jpeg")},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert respuesta.headers["location"] == "/compras/buscar"
    mock_colgar.assert_called_once_with(HOY_DE_PRUEBA, 200, "2026-08-06/n07p41-abc.jpg")
    mock_crear.assert_not_called()


def test_terminar_solo_con_foto_sin_guia_del_dia_avisa():
    with (
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main._comprimir_foto_jpeg", return_value=b"jpeg-chico"),
        patch("app.main.subir_foto_comanda", return_value="ruta.jpg"),
        patch("app.main.agregar_foto_guia_del_dia", return_value=False),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        respuesta = cliente.post(
            "/compras/nueva",
            data={"proveedor_id": "200", "accion": "terminar", "articulo_id": "", "cantidad_cajones": "",
                  "contenido_por_cajon": "", "importe": "", "sena": "", "tipo_retiro": ""},
            files={"comanda_foto": ("comanda.jpg", b"bytes", "image/jpeg")},
        )

    assert respuesta.status_code == 400
    assert "se guarda junto con un artículo" in respuesta.text


def test_boton_adjuntar_comanda_en_las_dos_pantallas_de_carga_manual():
    with (
        patch("app.main.listar_proveedores", return_value=PROVEEDORES_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
    ):
        combinada = cliente.get("/compras/nueva/manual")
    with (
        patch("app.main.obtener_proveedor", return_value=PROVEEDOR_DE_PRUEBA),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.listar_compras_por_fecha_y_proveedor", return_value=[]),
    ):
        clasica = cliente.get("/compras/nueva?proveedor_id=200")

    for respuesta in (combinada, clasica):
        assert respuesta.status_code == 200
        assert "Adjuntar comanda" in respuesta.text
        assert 'class="boton-adjuntar"' in respuesta.text
        assert 'name="comanda_foto"' in respuesta.text
        assert 'accept="image/*"' in respuesta.text
        assert 'enctype="multipart/form-data"' in respuesta.text


def test_ver_recepcion_muestra_la_fecha_de_cada_partida_y_marca_las_viejas():
    # La fecha SIEMPRE a la vista para ver si la partida es de un día
    # anterior; con más de un día, marca fuerte (mismo criterio que
    # Retirar Mercadería). La guía es por proveedor y día: una fecha por tarjeta.
    pendientes = [
        {
            "id": 1, "guia_id": 105, "guia_punto": 1, "articulo_nombre": "Tomate", "unidad_compra": "kilo",
            "proveedor_nombre": "Saturno", "proveedor_codigo_puesto": "N07P41",
            "fecha_operacion": date(2026, 8, 22),
            "cantidad_cajones": 40, "contenido_por_cajon": 20, "cantidad_kilos": 800, "cantidad_fraccion": None,
        },
        {
            "id": 2, "guia_id": 106, "guia_punto": 1, "articulo_nombre": "Mango", "unidad_compra": "unidad",
            "proveedor_nombre": "Don Pepe", "proveedor_codigo_puesto": "N01P02",
            "fecha_operacion": date(2026, 8, 20),
            "cantidad_cajones": 10, "contenido_por_cajon": 12, "cantidad_kilos": None, "cantidad_fraccion": 120,
        },
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=pendientes),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")

    assert respuesta.status_code == 200
    # La de hoy: fecha visible, sin marca.
    assert '<p class="fecha-guia">Compra del 22/08</p>' in respuesta.text
    # La de anteayer: con la marca fuerte.
    assert '<p class="fecha-guia fecha-vieja">⚠ Compra del 20/08</p>' in respuesta.text


def test_cambiar_horario_revision_guarda_y_confirma():
    with patch("app.main.guardar_horario_revision_casilla") as mock_guardar:
        respuesta = cliente.post(
            "/sistema/casilla-pedidos/3/horario",
            data={"revision_desde": "12:00", "revision_hasta": "13:00", "revision_cada_minutos": "30"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(3, time(12, 0), time(13, 0), 30)
    assert "de+12%3A00+a+13%3A00+cada+30+minutos" in respuesta.headers["location"]


def test_cambiar_horario_revision_valida_la_ventana_y_la_cadencia():
    casos = [
        # Desde después de hasta: no hay ventana.
        ({"revision_desde": "15:00", "revision_hasta": "12:00", "revision_cada_minutos": "15"},
         "La hora de inicio tiene que ser anterior a la de fin."),
        # Cadencia fuera de rango.
        ({"revision_desde": "12:00", "revision_hasta": "15:00", "revision_cada_minutos": "3"},
         "entre 5 y 240"),
        # Hora ilegible.
        ({"revision_desde": "doce", "revision_hasta": "15:00", "revision_cada_minutos": "15"},
         "El horario de revisión no es válido."),
    ]
    for datos, mensaje in casos:
        with (
            patch("app.main.guardar_horario_revision_casilla") as mock_guardar,
            patch("app.main.listar_casillas_pedidos", return_value=[]),
            patch("app.main.listar_mails_pedido", return_value=[]),
            patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
            patch("app.main.clave_casilla_configurada", return_value="clave"),
        ):
            respuesta = cliente.post("/sistema/casilla-pedidos/3/horario", data=datos)

        assert respuesta.status_code == 400, datos
        assert mensaje in respuesta.text
        mock_guardar.assert_not_called()


def test_ver_casilla_muestra_el_horario_de_revision_y_su_formulario():
    casilla = dict(CASILLA_DE_PRUEBA, revision_desde=time(12, 0), revision_hasta=time(13, 0), revision_cada_minutos=30)
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    assert "de 12:00 a 13:00,\n      cada 30 min" in respuesta.text
    assert 'action="/sistema/casilla-pedidos/3/horario"' in respuesta.text
    assert 'value="12:00" required' in respuesta.text
    assert 'value="13:00" required' in respuesta.text
    assert "Cambiar horario de revisión automática" in respuesta.text


# --- Armado: kilos reales, anular, terminar, Buscar Pedidos ---

FICHAS_ARMADO_CON_CONTENIDO = [
    {"id": 1, "articulo_id": 1, "articulo_nombre": "Banana", "articulo_grupo": "fruta",
     "envase_id": None, "envase_nombre": None, "contenido_caja": 20.0, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": "BANANA", "codigo_cliente": "90101"},
    {"id": 2, "articulo_id": 2, "articulo_nombre": "Batata", "articulo_grupo": "hortaliza",
     "envase_id": None, "envase_nombre": None, "contenido_caja": None, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "90102"},
]


def _get_armar_sucursal(renglones, pedido=None):
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=pedido or PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=renglones),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
        patch("app.main.listar_fichas_por_cliente", return_value=FICHAS_ARMADO_CON_CONTENIDO),
    ):
        return cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21&sucursal=VL")


def test_armar_muestra_el_kilaje_por_bulto_precargado_de_la_ficha():
    respuesta = _get_armar_sucursal(RENGLONES_ARMADO_DE_PRUEBA)

    assert respuesta.status_code == 200
    # Banana pendiente: el input es el POR BULTO (20 kg de ficha, editable),
    # con la unidad real del artículo en la etiqueta, y el total al lado
    # como dato informativo.
    assert 'name="kilos_por_bulto"' in respuesta.text
    assert 'value="20"' in respuesta.text
    assert "Kilos por cajón" in respuesta.text
    assert "Manda 15 bultos × 20 kg = 300 kg." in respuesta.text


def test_armar_sin_contenido_en_la_ficha_pide_el_por_bulto_a_mano():
    renglon_batata = [dict(RENGLONES_ARMADO_DE_PRUEBA[0], articulo_id=2, articulo_nombre="Batata")]
    respuesta = _get_armar_sucursal(renglon_batata)

    assert respuesta.status_code == 200
    assert "La ficha no tiene contenido por caja: cargá cuánto va por cajón a mano" in respuesta.text


def test_armar_muestra_los_kilos_enviados_y_marca_el_editado_a_mano():
    renglones = [
        # Armado con kilos = ficha (12 armados × 20 = 240): sin marca.
        dict(RENGLONES_ARMADO_DE_PRUEBA[1], articulo_id=1, articulo_nombre="Banana",
             cantidad_armada=12.0, kilos_enviados=240.0),
        # Armado con kilos DISTINTOS de la ficha: marca "editado a mano".
        dict(RENGLONES_ARMADO_DE_PRUEBA[1], id=14, articulo_id=1, articulo_nombre="Banana2",
             cantidad_armada=None, kilos_enviados=350.0),
        # Armado SIN kilaje: se dice, no se calcula.
        dict(RENGLONES_ARMADO_DE_PRUEBA[1], id=15, articulo_id=2, articulo_nombre="Batata",
             cantidad_armada=None, kilos_enviados=None),
    ]
    respuesta = _get_armar_sucursal(renglones)

    assert respuesta.status_code == 200
    assert "Mandó 240 kg" in respuesta.text
    assert "Mandó 350 kg" in respuesta.text
    assert respuesta.text.count("kilaje editado a mano") == 1
    assert "sin kilaje cargado" in respuesta.text


def test_armar_renglon_guarda_el_total_calculado_por_el_server():
    # Se carga el POR BULTO; el total enviado (lo que se factura) lo
    # calcula el server al tildar y queda congelado: 16 × 15 pedidos = 240.
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_pedida": "15", "cantidad_armada": "", "kilos_por_bulto": "16"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(11, None, 240.0)


def test_armar_renglon_incompleto_calcula_el_total_con_los_bultos_armados():
    # "Armé menos" cambia los bultos REALES del tilde: el total sale de
    # esos, no de lo pedido: 16 × 12 armados = 192.
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_pedida": "15", "cantidad_armada": "12", "kilos_por_bulto": "16"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_marcar.assert_called_once_with(11, 12.0, 192.0)


def test_armar_renglon_kilos_invalidos_da_400():
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "cantidad_pedida": "15", "kilos_por_bulto": "-3"},
        )

    assert respuesta.status_code == 400
    mock_marcar.assert_not_called()


def test_armar_renglon_con_kilaje_pero_sin_bultos_da_400():
    # Sin cantidad pedida ni armada no hay con qué multiplicar el por
    # bulto: mejor un error claro que guardar cualquier cosa.
    with patch("app.main.marcar_renglon_armado") as mock_marcar:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/armar",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL",
                  "kilos_por_bulto": "16"},
        )

    assert respuesta.status_code == 400
    mock_marcar.assert_not_called()


def test_anular_renglon_lo_registra_y_desanular_lo_repone():
    with patch("app.main.anular_renglon_pedido") as mock_anular:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/anular",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_anular.assert_called_once_with(11)

    with patch("app.main.desanular_renglon_pedido") as mock_desanular:
        respuesta = cliente.post(
            "/deposito/pedido/50/renglones/11/desanular",
            data={"cliente_id": "1", "fecha": "2026-08-21", "sucursal": "VL"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_desanular.assert_called_once_with(11)


def test_armar_separa_anulados_en_su_seccion_y_los_saca_del_progreso():
    renglones = [
        dict(RENGLONES_ARMADO_DE_PRUEBA[0]),  # Banana pendiente
        dict(RENGLONES_ARMADO_DE_PRUEBA[0], id=19, articulo_nombre="Anulada",
             anulado_el=datetime(2026, 8, 21, 13, 0)),
    ]
    respuesta = _get_armar_sucursal(renglones)

    assert respuesta.status_code == 200
    # El anulado NO está en pendientes ni cuenta en el progreso.
    assert "Falta armar (1)" in respuesta.text
    assert "VL: 0 de 1 artículos armados" in respuesta.text
    # Sección propia, distinta del completado, con Reponer.
    assert "Anulados (1)" in respuesta.text
    assert "anulado (no se armó)" in respuesta.text
    assert 'action="/deposito/pedido/50/renglones/19/desanular"' in respuesta.text
    assert "Reponer" in respuesta.text
    # La cruz en los pendientes.
    assert 'action="/deposito/pedido/50/renglones/11/anular"' in respuesta.text


def test_armar_los_armados_van_plegados_abajo_como_en_vacios_y_retiros():
    respuesta = _get_armar_sucursal(RENGLONES_ARMADO_DE_PRUEBA)

    assert respuesta.status_code == 200
    # Pendientes arriba y corto; lo terminado abajo, plegado, con deshacer.
    assert "Falta armar (1)" in respuesta.text
    assert "<summary>Ya armado (1)</summary>" in respuesta.text
    assert "Destildar" in respuesta.text


def test_terminar_pedido_cierra_y_muestra_el_banner_con_reabrir():
    with patch("app.main.cerrar_armado_pedido") as mock_cerrar:
        respuesta = cliente.post(
            "/deposito/pedido/50/terminar",
            data={"cliente_id": "1", "fecha": "2026-08-21"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_cerrar.assert_called_once_with(50)
    assert "terminado" in respuesta.headers["location"]

    # El pedido cerrado muestra el banner con Reabrir (y sin el botón Terminar).
    pedido_cerrado = dict(PEDIDO_VIGENTE_DE_PRUEBA, armado_cerrado_el=datetime(2026, 8, 21, 18, 0, tzinfo=timezone.utc))
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=pedido_cerrado),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        vista = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert vista.status_code == 200
    assert "Pedido terminado" in vista.text
    assert "Reabrir pedido" in vista.text
    assert "Terminar pedido (" not in vista.text

    with patch("app.main.reabrir_armado_pedido") as mock_reabrir:
        respuesta = cliente.post(
            "/deposito/pedido/50/reabrir",
            data={"cliente_id": "1", "fecha": "2026-08-21"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_reabrir.assert_called_once_with(50)


def test_armar_muestra_terminar_pedido_con_los_sin_tildar_avisados():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=RENGLONES_ARMADO_DE_PRUEBA),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    assert respuesta.status_code == 200
    # Un pendiente (Banana): el botón lo dice y el confirm avisa, sin impedir.
    assert "Terminar pedido (1 sin tildar)" in respuesta.text
    assert "sin tildar. ¿Terminar el pedido igual?" in respuesta.text
    assert 'action="/deposito/pedido/50/terminar"' in respuesta.text


def test_terminar_pedido_no_cuenta_renglones_sin_sucursal_como_pendientes():
    # Bug del 25/08: "VL 21 de 21, BZ 22 de 22, GR 19 de 19 — Terminar
    # pedido (10 sin tildar)". Los 10 eran renglones identificados SIN
    # sucursal (vinieron sin cantidades en el mail, cantidad 0): no
    # aparecen en ninguna sucursal y no se pueden tildar. El botón se suma
    # DE LOS MISMOS conteos por sucursal — no puede contradecirlos.
    renglones = [dict(r) for r in RENGLONES_ARMADO_DE_PRUEBA] + [
        {"id": 14, "sucursal": None, "articulo_id": 3, "articulo_nombre": "Kiwi", "ficha_id": 903, "nombre_venta": "Kiwi",
         "texto_codigo": "90103", "texto_descripcion": "KIWI", "cantidad": 0,
         "armado_el": None, "cantidad_armada": None, "kilos_enviados": None, "anulado_el": None},
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=renglones),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    # Sigue habiendo UN solo pendiente real (Banana VL): el Kiwi sin
    # sucursal no infla el contador.
    assert "Terminar pedido (1 sin tildar)" in respuesta.text


RENGLONES_BUSCAR_DE_PRUEBA = [
    {"fecha_operacion": date(2026, 8, 21), "id": 11, "sucursal": "VL", "articulo_id": 1,
     "articulo_nombre": "Banana", "cantidad": 15.0, "cantidad_armada": 12.0,
     "kilos_enviados": 240.0, "armado_el": datetime(2026, 8, 21, 13, 0), "anulado_el": None},
    {"fecha_operacion": date(2026, 8, 21), "id": 12, "sucursal": "BZ", "articulo_id": 2,
     "articulo_nombre": "Batata", "cantidad": 40.0, "cantidad_armada": None,
     "kilos_enviados": None, "armado_el": None, "anulado_el": None},
    {"fecha_operacion": date(2026, 8, 20), "id": 13, "sucursal": "VL", "articulo_id": 1,
     "articulo_nombre": "Banana", "cantidad": 10.0, "cantidad_armada": None,
     "kilos_enviados": None, "armado_el": None, "anulado_el": datetime(2026, 8, 20, 13, 0)},
]


def test_buscar_pedidos_muestra_los_kilos_enviados_nunca_los_de_ficha():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.buscar_renglones_pedidos", return_value=list(RENGLONES_BUSCAR_DE_PRUEBA)),
    ):
        respuesta = cliente.get("/administracion/pedidos/buscar?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")

    assert respuesta.status_code == 200
    texto = respuesta.text
    assert "Pedido del 21/08/2026" in texto
    # Los kilos REALES del depósito (240, de 12 bultos armados) — y el sin
    # armar dice "sin kilaje", jamás un cálculo de ficha.
    assert "240 kg" in texto
    assert "sin kilaje" in texto
    assert "sin armar" in texto
    # El anulado, visible como anulado (registrado, no borrado) y sin sumar.
    assert "anulado" in texto
    assert "1 anulado" in texto
    # Export con los mismos filtros.
    assert "/administracion/pedidos/buscar/exportar-pdf?cliente_id=1" in texto
    assert "/administracion/pedidos/buscar/exportar-excel?cliente_id=1" in texto


def test_buscar_pedidos_sin_cliente_muestra_solo_el_selector():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.buscar_renglones_pedidos") as mock_buscar,
    ):
        respuesta = cliente.get("/administracion/pedidos/buscar")

    assert respuesta.status_code == 200
    assert "Elegí un cliente" in respuesta.text
    mock_buscar.assert_not_called()


def test_exportar_buscar_pedidos_pdf_y_excel():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.obtener_cliente", return_value=dict(CLIENTE_DE_PRUEBA)),
        patch("app.main.buscar_renglones_pedidos", return_value=list(RENGLONES_BUSCAR_DE_PRUEBA)),
    ):
        pdf = cliente.get("/administracion/pedidos/buscar/exportar-pdf?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")
        excel = cliente.get("/administracion/pedidos/buscar/exportar-excel?cliente_id=1&fecha_desde=2026-08-15&fecha_hasta=2026-08-22")

    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")
    assert 'filename="Pedidos_2026-08-15_a_2026-08-22.pdf"' in pdf.headers["content-disposition"]
    assert excel.status_code == 200
    assert "spreadsheetml" in excel.headers["content-type"]


def test_armar_esconde_los_terminados_en_una_seccion_plegada():
    # El que arma ve SOLO lo pendiente. Terminado = tocó "Terminar pedido"
    # (armado_cerrado_el), NO "todo tildado": el completo sin cerrar sigue
    # arriba. Los terminados quedan plegados abajo, entrables para reabrir.
    listado = [
        # Completo pero SIN cerrar: sigue en la lista principal.
        {"id": 48, "fecha_operacion": date(2026, 8, 20), "origen": "texto",
         "creado_en": datetime(2026, 8, 19, 12, 0), "armado_cerrado_el": None,
         "renglones_totales": 30, "renglones_armados": 30, "sin_identificar": 0},
        # CERRADO (aunque incompleto): va plegado abajo.
        {"id": 50, "fecha_operacion": date(2026, 8, 21), "origen": "mail",
         "creado_en": datetime(2026, 8, 21, 12, 0),
         "armado_cerrado_el": datetime(2026, 8, 21, 18, 0),
         "renglones_totales": 32, "renglones_armados": 18, "sin_identificar": 0},
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=listado),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1")

    texto = respuesta.text
    # El pendiente (completo sin cerrar) arriba; el terminado, solo en la
    # sección plegada con su acceso para reabrir.
    assert "Ver terminados (1)" in texto
    assert texto.index("Pedido del 20/08/2026") < texto.index("Ver terminados (1)")
    assert texto.index("Ver terminados (1)") < texto.index("Pedido del 21/08/2026")
    assert "COMPLETO ✔" in texto
    # El cerrado de este caso quedó a 18 de 32: TERMINADO sin tilde, con lo
    # que le falta. El tilde es solo para el que salió completo — antes el
    # verde y el ✔ iban fijos y le decían al que mira que no hay nada que
    # hacer.
    assert "TERMINADO · 14 sin armar — entrá para reabrirlo" in texto
    assert "TERMINADO ✔" not in texto

    # Todo terminado: la lista principal lo dice y apunta a la sección.
    todos_cerrados = [dict(listado[1])]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=todos_cerrados),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1")
    assert "Nada pendiente de armar" in respuesta.text
    assert "Ver terminados (1)" in respuesta.text


def test_ver_pedido_tiene_el_boton_buscar_pedidos_y_el_badge_terminado():
    listado = [{
        "id": 50, "fecha_operacion": date(2026, 8, 21), "origen": "texto",
        "creado_en": datetime(2026, 8, 21, 12, 10), "armado_cerrado_el": datetime(2026, 8, 21, 18, 0),
        "renglones_totales": 3, "renglones_armados": 2, "sin_identificar": 0,
    }]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=listado),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    assert 'href="/administracion/pedidos/buscar?cliente_id=1"' in respuesta.text
    # Cerrado explícitamente y a 2 de 3: TERMINADO, pero SIN tilde y diciendo
    # qué falta. "Terminado" es que alguien tocó el botón, no que salió
    # completo.
    assert "TERMINADO · 1 sin armar" in respuesta.text
    assert "TERMINADO ✔" not in respuesta.text


def test_ver_pedido_muestra_los_mails_trabados_del_cliente_con_revisar():
    # El mail con problemas NO queda estacionado solo en Sistema: aparece
    # en la pantalla de Pedido del cliente, con su estado y el botón de
    # revisar (que para un error de lectura es también el reintento).
    mails = [
        {"id": 9, "remitente": "pedidos@dia.com.ar", "asunto": "Pedido Dia 23-08 Domingo",
         "recibido_el": datetime(2026, 8, 22, 15, 8, tzinfo=timezone.utc), "estado": "error",
         "motivo": "La lectura falló: se cortó la respuesta"},
        {"id": 8, "remitente": "pedidos@dia.com.ar", "asunto": "Pedido Dia 22-08 Sabado",
         "recibido_el": datetime(2026, 8, 21, 15, 8, tzinfo=timezone.utc), "estado": "pendiente", "motivo": None},
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_condiciones_pedido", return_value=None),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=mails) as mock_mails,
        patch("app.main.obtener_pedido_vigente", return_value=None),
    ):
        respuesta = cliente.get("/deposito/pedido?cliente_id=1")

    assert respuesta.status_code == 200
    mock_mails.assert_called_once_with(1)
    texto = respuesta.text
    assert "Mails de pedido por confirmar (2)" in texto
    # El error con su motivo a la vista, y la fecha ESTIMADA del asunto.
    assert "Error de lectura" in texto
    assert "La lectura falló: se cortó la respuesta" in texto
    assert "pedido estimado del 23/08/2026" in texto
    assert "pedido estimado del 22/08/2026" in texto
    # El botón lleva directo a la revisión (mismo circuito de siempre).
    assert 'href="/deposito/pedido/mails/9/revisar"' in texto
    assert 'href="/deposito/pedido/mails/8/revisar"' in texto
    # Con un error, la tarjeta se marca fuerte.
    assert "tarjeta-mails-trabados con-error" in texto


# --- Stock del Depósito (tanda 1) ---

FILAS_STOCK_DE_PRUEBA = [
    {"articulo_id": 1, "nombre": "Banana", "entradas": 40.0, "salidas": 15.0,
     "reingresos": 2.0, "ajustes": -3.0, "reproceso_primera": 0.0,
     "reproceso_tomados": 0.0, "segunda": 0.0, "stock": 24.0},
    {"articulo_id": 2, "nombre": "Anco", "entradas": 0.0, "salidas": 5.0,
     "reingresos": 0.0, "ajustes": 0.0, "reproceso_primera": 0.0,
     "reproceso_tomados": 0.0, "segunda": 0.0, "stock": -5.0},
]


def test_ver_deposito_muestra_el_acceso_a_stock():
    respuesta = cliente.get("/deposito")

    assert respuesta.status_code == 200
    assert 'href="/deposito/stock"' in respuesta.text


def test_el_hub_de_stock_queda_solo_con_la_carga_del_operario():
    respuesta = cliente.get("/deposito/stock")

    assert respuesta.status_code == 200
    # Solo la CARGA del operario: el control se mudó a Administración.
    assert 'href="/deposito/stock/fisico"' in respuesta.text
    assert 'href="/deposito/stock/merma"' in respuesta.text
    assert 'href="/deposito/stock/reingreso"' in respuesta.text
    assert 'href="/deposito/stock/reproceso"' in respuesta.text
    assert 'href="/deposito/stock/remito-segunda"' in respuesta.text
    assert "Carga del depósito" in respuesta.text
    # Y NO quedan botones apuntando a otro módulo: mudar la pantalla y
    # dejar el botón sería mudar a medias.
    assert "/administracion/" not in respuesta.text
    # Módulo completo: no queda nada "Próximamente".
    assert "(Próximamente)" not in respuesta.text


# Stock del Sistema (el listado) se BORRÓ el 06/09: lo miraron tres veces y
# ninguna sirvió. Lo que aquellos tests protegían no se borró con la pantalla
# —se mudó, y los tests con ello—: la línea de las seis patas está ahora en
# Stock por Guía, y los negativos y los dos totales globales en el Remanente.
# Ver los tests de abajo, que dicen dónde quedó cada cosa.


def test_stock_por_guia_muestra_LAS_SEIS_PATAS_de_donde_sale_el_numero():
    """Vivían en el listado borrado. Es lo único que dice QUÉ pata movió
    cuando un total no cuadra, y ésta es la pantalla a la que se viene justo
    a preguntarse eso."""
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(0.0))),
        patch("app.main.stock_deposito_por_articulo",
              return_value=[dict(f) for f in FILAS_STOCK_DE_PRUEBA]),
    ):
        respuesta = cliente.get("/administracion/stock/sistema/1")

    assert respuesta.status_code == 200
    cuerpo = respuesta.text.split("</style>")[-1]
    assert "40</b> entraron" in cuerpo
    assert "15</b> salieron" in cuerpo
    assert "2</b> reingresados" in cuerpo
    assert "-3</b> ajustados" in cuerpo


def test_stock_articulo_reparte_fifo_y_muestra_lo_que_queda_por_lote():
    entradas = [
        {"fecha_orden": date(2026, 8, 20), "momento_orden": datetime(2026, 8, 20, 10), "orden": (date(2026, 8, 20), datetime(2026, 8, 20, 10)),
         "tipo_lote": "guia", "fecha_lote": date(2026, 8, 20), "detalle": "Norte 15",
         "motivo": None, "cantidad": 8.0},
        {"fecha_orden": date(2026, 8, 22), "momento_orden": datetime(2026, 8, 22, 10), "orden": (date(2026, 8, 22), datetime(2026, 8, 22, 10)),
         "tipo_lote": "guia", "fecha_lote": date(2026, 8, 22), "detalle": "Norte 15",
         "motivo": None, "cantidad": 10.0},
    ]
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=(entradas, _salidas_fifo(11.0))),
        patch("app.main.stock_deposito_por_articulo", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/sistema/1")

    assert respuesta.status_code == 200
    # 11 salieron: la guía del 20 (8) se agotó y la del 22 quedó en 7.
    assert "Banana: 7 bultos" in respuesta.text
    assert "Guía 22/08 · Norte 15" in respuesta.text
    assert "7 <span class=\"de\">de 10</span>" in respuesta.text
    assert "Lotes agotados (1)" in respuesta.text
    assert "Salieron 11 bultos en total" in respuesta.text


def test_stock_articulo_negativo_muestra_los_bultos_sin_lote():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Anco"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(5.0))),
        patch("app.main.stock_deposito_por_articulo", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/sistema/2")

    assert respuesta.status_code == 200
    assert "Anco: -5 bultos" in respuesta.text
    # Sin lote, dicho tal cual: no se cuelga de ninguna guía.
    assert "5 bultos salieron sin lote" in respuesta.text


def test_stock_articulo_inexistente_da_404():
    with patch("app.main.obtener_articulo", return_value=None):
        respuesta = cliente.get("/administracion/stock/sistema/999")

    assert respuesta.status_code == 404


def test_ajustar_stock_guarda_el_movimiento_y_precarga_el_motivo_en_cadena():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_movimiento_stock", return_value=12.0) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/ajustar",
            data={"articulo_id": "1", "cantidad": "12", "motivo": "stock inicial"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, "ajuste", 12.0, "stock inicial", date(2026, 8, 25))
    # El motivo vuelve en la URL para encadenar varios ajustes con el mismo
    # motivo sin retipearlo. (El stock inicial del corte ya NO se carga por
    # acá: tiene pantalla y tipo propios.)
    assert "motivo=stock+inicial" in respuesta.headers["location"]
    assert "El+stock+qued%C3%B3+en+12" in respuesta.headers["location"]


def test_ajustar_stock_sin_motivo_da_400():
    with patch("app.main.crear_movimiento_stock") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post(
            "/administracion/stock/ajustar",
            data={"articulo_id": "1", "cantidad": "5", "motivo": "   "},
        )

    assert respuesta.status_code == 400
    assert "El motivo es obligatorio" in respuesta.text
    mock_crear.assert_not_called()


def test_ajustar_stock_cantidad_cero_da_400():
    with patch("app.main.crear_movimiento_stock") as mock_crear, patch("app.main.listar_articulos", return_value=[]):
        respuesta = cliente.post(
            "/administracion/stock/ajustar",
            data={"articulo_id": "1", "cantidad": "0", "motivo": "nada"},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_ajustar_stock_muestra_el_motivo_precargado():
    with patch("app.main.listar_articulos", return_value=[{"id": 1, "nombre": "Banana"}]):
        respuesta = cliente.get("/administracion/stock/ajustar?motivo=stock+inicial")

    assert respuesta.status_code == 200
    assert 'value="stock inicial"' in respuesta.text


# --- Stock del Depósito (tanda 2): Merma, Reingreso, Movimientos ---


def test_merma_guarda_negativa_y_el_aviso_no_muestra_el_stock():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_movimiento_stock", return_value=17.0) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/merma",
            data={"articulo_id": "1", "cantidad": "3", "motivo": "podrido"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # Siempre negativa: el signo lo pone el tipo, no la persona.
    # Sin lote elegido: el FIFO de siempre (lote en None).
    mock_crear.assert_called_once_with(
        1, "merma", -3.0, "podrido", date(2026, 8, 25), lote_tipo=None, lote_origen_id=None
    )
    # Pantalla de OPERARIO: el aviso repite lo cargado, JAMÁS el stock
    # resultante (17 no puede aparecer).
    destino = respuesta.headers["location"]
    assert "Merma+guardada%3A+3+bultos+de+Banana" in destino
    assert "17" not in destino


def test_merma_sin_motivo_da_400():
    with (
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(0.0))),
    ):
        respuesta = cliente.post(
            "/deposito/stock/merma",
            data={"articulo_id": "1", "cantidad": "3", "motivo": " "},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_merma_cantidad_negativa_da_400():
    # El operario carga bultos tirados (positivo): el signo no es cosa suya.
    with (
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(0.0))),
    ):
        respuesta = cliente.post(
            "/deposito/stock/merma",
            data={"articulo_id": "1", "cantidad": "-3", "motivo": "podrido"},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


# El renglón armado que el reingreso vincula: 25 bultos armados con 500 kg
# enviados (20 kg por bulto), 5 ya devueltos — el tope del server es 20.
RENGLON_REINGRESO_DE_PRUEBA = {
    "id": 77, "pedido_id": 40, "sucursal": "VL", "articulo_id": 2, "ficha_id": 902,
    "articulo_nombre": "Anco", "cliente_id": 1, "cliente_nombre": "Día",
    "fecha_pedido": date(2026, 8, 24), "orden_compra": "1257673",
    "bultos_armados": 25.0, "kilos_enviados": 500.0, "ya_devuelto": 5.0,
}


def test_reingreso_a_segunda_manda_los_bultos_al_pool_y_lo_dice_en_el_aviso():
    # El destino se elige al cargar: "pasa a segunda tal cual" manda los
    # mismos bultos al pool, con su caja y su kilaje.
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main._costo_congelado_para_reingreso", return_value=2000.0),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "4", "motivo": "rechazado por calidad",
                  "fecha": "2026-08-24", "destino": "segunda"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        2, "reingreso_rechazo", 4.0, "rechazado por calidad", date(2026, 8, 24),
        cliente_id=1, pedido_renglon_id=77, costo_por_bulto=2000.0,
        destino_rechazo="segunda", bultos_segunda=4.0,
    )
    destino = respuesta.headers["location"]
    assert "Pas%C3%B3+a+segunda+tal+cual" in destino
    assert "4+bultos+al+pool" in destino


def test_reingreso_a_cajon_grande_carga_los_cajones_que_salieron():
    # Las palabras del dueño: yo cargo cuántos cajones grandes salieron, y
    # eso entra al pool de segunda.
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main._costo_congelado_para_reingreso", return_value=2000.0),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "20", "motivo": "rechazado",
                  "fecha": "2026-08-24", "destino": "reproceso", "cajones": "6"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert mock_crear.call_args.kwargs["destino_rechazo"] == "reproceso"
    assert mock_crear.call_args.kwargs["bultos_segunda"] == 6.0
    destino = respuesta.headers["location"]
    assert "Volvi%C3%B3+a+caj%C3%B3n+grande" in destino
    assert "salieron+6+cajones" in destino
    assert "pool+de+segunda" in destino


def test_reingreso_a_cajon_grande_sin_cajones_da_400():
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "20", "motivo": "rechazado",
                  "fecha": "2026-08-24", "destino": "reproceso", "cajones": ""},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_el_form_del_reingreso_ofrece_los_tres_destinos():
    with patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)):
        respuesta = cliente.get("/deposito/stock/reingreso?renglon_id=77")

    assert respuesta.status_code == 200
    assert "¿Qué se hace con esta mercadería?" in respuesta.text
    assert 'value="stock"' in respuesta.text and "Queda en stock" in respuesta.text
    assert 'value="segunda"' in respuesta.text and "Pasa a segunda tal cual" in respuesta.text
    assert 'value="reproceso"' in respuesta.text and "Vuelve a cajón grande" in respuesta.text
    assert "¿Cuántos cajones grandes salieron?" in respuesta.text


def test_el_form_del_reingreso_pide_la_fecha_de_la_vuelta_no_la_del_pedido():
    # Son dos fechas distintas y las dos están a la vista en la pantalla:
    # la etiqueta tiene que decir cuál se está pidiendo, y la ayuda la
    # contrasta con la del pedido (24/08 en este renglón).
    with patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)):
        respuesta = cliente.get("/deposito/stock/reingreso?renglon_id=77")

    assert "¿Qué día entró al depósito?" in respuesta.text
    assert "El pedido es del 24/08. Esta fecha es la de la vuelta, y ordena el stock." in respuesta.text
    assert "Fecha en que volvió" not in respuesta.text


LOTES_MERMA_DE_PRUEBA = [
    {"fecha_orden": date(2026, 8, 20), "momento_orden": datetime(2026, 8, 20, 10), "orden": (date(2026, 8, 20), datetime(2026, 8, 20, 10)),
     "tipo_lote": "guia", "origen_id": 55, "fecha_lote": date(2026, 8, 20),
     "detalle": "Norte 15", "motivo": None, "cantidad": 80.0, "costo_bulto": 500.0,
     "cliente_lote_id": None},
    {"fecha_orden": date(2026, 8, 24), "momento_orden": datetime(2026, 8, 24, 10), "orden": (date(2026, 8, 24), datetime(2026, 8, 24, 10)),
     "tipo_lote": "reproceso", "origen_id": 9, "fecha_lote": date(2026, 8, 24),
     "detalle": "Día", "motivo": None, "cantidad": 40.0, "costo_bulto": 1800.0,
     "cliente_lote_id": 1},
]


def test_merma_ofrece_los_lotes_del_articulo_con_el_mas_viejo_por_default():
    # El operario no tiene que pensar salvo que sepa cuál se pudrió: el
    # default es "el más viejo (como siempre)".
    with (
        patch("app.main.listar_articulos", return_value=[{"id": 2, "nombre": "Tomate Perita"}]),
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Tomate Perita"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=(LOTES_MERMA_DE_PRUEBA, _salidas_fifo(0.0))),
    ):
        respuesta = cliente.get("/deposito/stock/merma?articulo_id=2")

    assert respuesta.status_code == 200
    assert "El más viejo (como siempre)" in respuesta.text
    # Cada lote se describe con lo que el operario reconoce en el piso.
    assert 'value="guia:55"' in respuesta.text
    assert "Compra de Norte 15" in respuesta.text
    assert 'value="reproceso:9"' in respuesta.text
    assert "Guía R9 armada para Día" in respuesta.text
    assert "quedan 40" in respuesta.text


def test_merma_lista_un_lote_de_compra_sin_guia_con_la_fecha_del_hecho():
    # Una compra cargada sin guía no tiene fecha de guía: se muestra con
    # la de su recepción, en vez de romper la pantalla.
    sin_guia = [dict(LOTES_MERMA_DE_PRUEBA[0], fecha_lote=None)]
    with (
        patch("app.main.listar_articulos", return_value=[{"id": 2, "nombre": "Tomate Perita"}]),
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Tomate Perita"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=(sin_guia, _salidas_fifo(0.0))),
    ):
        respuesta = cliente.get("/deposito/stock/merma?articulo_id=2")

    assert respuesta.status_code == 200
    assert "Compra de Norte 15 — 20/08" in respuesta.text


def test_merma_dirigida_a_un_lote_guarda_el_lote_y_lo_dice_en_el_aviso():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Tomate Perita"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=(LOTES_MERMA_DE_PRUEBA, _salidas_fifo(0.0))),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 26)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/merma",
            data={"articulo_id": "2", "cantidad": "10", "motivo": "se pudrió", "lote": "reproceso:9"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        2, "merma", -10.0, "se pudrió", date(2026, 8, 26),
        lote_tipo="reproceso", lote_origen_id=9,
    )
    assert "Salieron+de%3A+Gu%C3%ADa+R9+armada+para+D%C3%ADa" in respuesta.headers["location"]


def test_merma_a_un_lote_que_ya_no_tiene_bultos_da_400():
    with (
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Tomate Perita"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=(LOTES_MERMA_DE_PRUEBA, _salidas_fifo(200.0))),
        patch("app.main.crear_movimiento_stock") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/stock/merma",
            data={"articulo_id": "2", "cantidad": "10", "motivo": "se pudrió", "lote": "reproceso:9"},
        )

    assert respuesta.status_code == 400
    assert "Ese lote ya no tiene bultos" in respuesta.text
    mock_crear.assert_not_called()


def test_reingreso_paso_1_lista_pedidos_para_elegir_sin_numeros_del_sistema():
    pedidos = [{"pedido_id": 40, "fecha_operacion": date(2026, 8, 24), "cliente_nombre": "Día",
                "sucursal": "VL", "orden_compra": "1257673", "renglones_armados": 3}]
    with (
        patch("app.main.listar_pedidos_para_reingreso", return_value=pedidos) as mock_pedidos,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/reingreso")

    assert respuesta.status_code == 200
    mock_pedidos.assert_called_once_with(oc=None)
    texto = respuesta.text
    assert "¿De qué pedido volvió la mercadería?" in texto
    assert "Pedido del 24/08/2026 — VL" in texto
    assert "OC 1257673" in texto
    assert "3 renglones armados" in texto
    # Búsqueda por OC, y nada de costos: pantalla de operario.
    assert 'name="oc"' in texto


def test_reingreso_paso_1_busca_por_oc():
    with (
        patch("app.main.listar_pedidos_para_reingreso", return_value=[]) as mock_pedidos,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/reingreso?oc=1257673")

    assert respuesta.status_code == 200
    mock_pedidos.assert_called_once_with(oc="1257673")
    assert "No hay ningún pedido con renglones armados con esa OC" in respuesta.text


def test_reingreso_paso_2_muestra_renglones_armados_con_su_tope():
    pedidos = [{"pedido_id": 40, "fecha_operacion": date(2026, 8, 24), "cliente_nombre": "Día",
                "sucursal": "VL", "orden_compra": "1257673", "renglones_armados": 2}]
    renglones = [
        {"id": 77, "articulo_nombre": "Anco", "bultos_armados": 25.0, "ya_devuelto": 5.0},
        {"id": 78, "articulo_nombre": "Banana", "bultos_armados": 10.0, "ya_devuelto": 10.0},
    ]
    with (
        patch("app.main.listar_pedidos_para_reingreso", return_value=pedidos),
        patch("app.main.listar_renglones_para_reingreso", return_value=renglones) as mock_renglones,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/reingreso?pedido_id=40&sucursal=VL")

    assert respuesta.status_code == 200
    mock_renglones.assert_called_once_with(40, "VL")
    texto = respuesta.text
    assert "Armaste 25 bultos" in texto and "ya devolviste 5" in texto
    assert 'href="/deposito/stock/reingreso?renglon_id=77"' in texto
    # El renglón ya devuelto entero no es tocable: no queda nada por devolver.
    assert "no queda nada por devolver" in texto
    assert 'href="/deposito/stock/reingreso?renglon_id=78"' not in texto


def test_reingreso_paso_3_precarga_el_tope_y_hoy_con_cliente_y_articulo_del_pedido():
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/reingreso?renglon_id=77")

    assert respuesta.status_code == 200
    texto = respuesta.text
    # El cliente y el artículo salen del pedido: son texto fijo, no selects.
    assert "Anco — pedido del 24/08/2026 (VL)" in texto
    assert "<strong>Día</strong>" in texto
    assert "<select" not in texto
    # Default = lo que queda por devolver (armado − ya devuelto), editable
    # con el tope a la vista; fecha hoy por default y nunca futura.
    assert "podés devolver hasta <strong>20</strong>" in texto
    assert 'value="20"' in texto
    assert 'max="20.0"' in texto
    assert 'value="2026-08-25"' in texto
    assert 'max="2026-08-25"' in texto


def test_reingreso_guarda_vinculado_con_costo_congelado_y_fecha_editable():
    # El costo congelado sale de la fila de la FICHA con la que se vendió.
    listado = [{"ficha_id": 902, "articulo_id": 2, "costo_actual": 100.0, "precio_vigente": 150.0}]
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=listado) as mock_listado,
        patch("app.main.crear_movimiento_stock", return_value=9.0) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "4", "motivo": "rechazado por calidad", "fecha": "2026-08-24"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # El listado se ancla a la fecha del PEDIDO de origen (24/08), y el
    # costo por unidad (100) se pasa a bultos con el kilaje real del
    # renglón (500 kg / 25 bultos = 20): 2000 por bulto, congelado.
    fecha_ancla = mock_listado.call_args.args[1]
    assert fecha_ancla.date() == date(2026, 8, 24)
    mock_crear.assert_called_once_with(
        2, "reingreso_rechazo", 4.0, "rechazado por calidad", date(2026, 8, 24),
        cliente_id=1, pedido_renglon_id=77, costo_por_bulto=2000.0,
        destino_rechazo="stock", bultos_segunda=None,
    )
    destino = respuesta.headers["location"]
    # El aviso repite lo cargado (fecha REAL del hecho incluida) y JAMÁS
    # el costo ni el stock: pantalla de operario.
    assert "devolvi%C3%B3+D%C3%ADa+el+24%2F08" in destino
    assert "2000" not in destino
    assert "qued" not in destino


def test_reingreso_tope_duro_del_server_no_deja_devolver_mas_de_lo_armado():
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "21", "motivo": "rechazo", "fecha": ""},
        )

    # Armados 25, ya devueltos 5: 21 pasa el tope de 20 — 400 del server,
    # aunque el HTML se toque a mano.
    assert respuesta.status_code == 400
    assert "No se puede devolver más de lo armado" in respuesta.text
    assert "el tope es 20" in respuesta.text
    mock_crear.assert_not_called()


def test_reingreso_sin_fecha_usa_hoy_y_sin_costo_posible_guarda_sin_costo():
    # Sin costo en el listado a la fecha del pedido: el reingreso se guarda
    # igual, con costo None — el lote queda visible como "sin costo" en la
    # Real, nunca un número inventado.
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main.calcular_listado_para_negociar_precios", return_value=[]),
        patch("app.main.crear_movimiento_stock", return_value=9.0) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "4", "motivo": "rechazo", "fecha": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        2, "reingreso_rechazo", 4.0, "rechazo", date(2026, 8, 25),
        cliente_id=1, pedido_renglon_id=77, costo_por_bulto=None,
        destino_rechazo="stock", bultos_segunda=None,
    )


def test_reingreso_con_fecha_futura_da_400():
    with (
        patch("app.main.obtener_renglon_para_reingreso", return_value=dict(RENGLON_REINGRESO_DE_PRUEBA)),
        patch("app.main.crear_movimiento_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reingreso",
            data={"renglon_id": "77", "cantidad": "4", "motivo": "x", "fecha": "2026-08-26"},
        )

    assert respuesta.status_code == 400
    assert "no puede ser futura" in respuesta.text
    mock_crear.assert_not_called()


def test_reingreso_de_renglon_no_disponible_da_404():
    # Un renglón de un pedido reemplazado/anulado (o un id inventado) no
    # está disponible: su armado no aportó al stock.
    with patch("app.main.obtener_renglon_para_reingreso", return_value=None):
        respuesta = cliente.post(
            "/deposito/stock/reingreso", data={"renglon_id": "999", "cantidad": "4", "motivo": "x"}
        )
    assert respuesta.status_code == 404


MOVIMIENTOS_STOCK_DE_PRUEBA = [
    {"id": 31, "tipo": "ajuste", "cantidad": 20.0, "motivo": "stock inicial",
     "fecha_operacion": date(2026, 8, 25), "stock_sistema": -5.0,
     "creado_en": datetime(2026, 8, 25, 10, 0), "anulado_el": None,
     "articulo_nombre": "Anco", "cliente_nombre": None},
    {"id": 32, "tipo": "merma", "cantidad": -3.0, "motivo": "podrido",
     "fecha_operacion": date(2026, 8, 25), "stock_sistema": 15.0,
     "creado_en": datetime(2026, 8, 25, 11, 0), "anulado_el": None,
     "articulo_nombre": "Banana", "cliente_nombre": None},
    {"id": 33, "tipo": "reingreso_rechazo", "cantidad": 4.0, "motivo": "rechazo calidad",
     "fecha_operacion": date(2026, 8, 24), "stock_sistema": 12.0,
     "creado_en": datetime(2026, 8, 25, 12, 0), "anulado_el": datetime(2026, 8, 25, 13, 0),
     "articulo_nombre": "Banana", "cliente_nombre": "Día"},
]


def test_movimientos_stock_muestra_los_tipos_con_pill_y_la_foto_del_sistema():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_movimientos_stock_por_rango",
              return_value=[dict(m) for m in MOVIMIENTOS_STOCK_DE_PRUEBA]) as mock_listar,
        patch("app.main.listar_remitos_segunda_por_rango", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/movimientos")

    assert respuesta.status_code == 200
    mock_listar.assert_called_once_with(date(2026, 8, 18), date(2026, 8, 25))
    # El tipo de un vistazo, como en Vacíos.
    assert 'class="pill pill-ajuste"' in respuesta.text
    assert 'class="pill pill-merma"' in respuesta.text
    assert 'class="pill pill-reingreso_rechazo"' in respuesta.text
    assert ">Reingreso</span>" in respuesta.text
    # Pantalla de CONTROL: acá sí se ve la foto del sistema.
    assert "el sistema decía -5" in respuesta.text
    assert "Devolvió Día" in respuesta.text
    # El anulado queda visible y marcado, sin botón de anular.
    assert "ANULADO" in respuesta.text
    assert respuesta.text.count('class="boton-anular"') == 2


def test_anular_movimiento_stock_vuelve_al_rango():
    with patch("app.main.anular_movimiento_stock") as mock_anular:
        respuesta = cliente.post(
            "/administracion/stock/movimientos/32/anular",
            data={"fecha_desde": "2026-08-18", "fecha_hasta": "2026-08-25"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_anular.assert_called_once_with(32)
    assert "fecha_desde=2026-08-18" in respuesta.headers["location"]


# --- Stock del Depósito (tanda 3): Stock Físico, Cotejo, alerta ---


def test_stock_fisico_guarda_el_conteo_y_el_aviso_no_muestra_el_sistema():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_conteo_stock") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "1", "cantidad": "12", "que_conto": "sueltos"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, 12.0, ficha_id=None)
    # Pantalla de OPERARIO: el aviso repite SOLO lo contado.
    destino = respuesta.headers["location"]
    assert "Conteo+guardado%3A+12+bultos+sueltos+de+Banana" in destino
    assert "sistema" not in destino
    # El artículo vuelve puesto: de un mismo artículo se cuentan los
    # sueltos y después las cajas de cada ficha, uno atrás de otro.
    assert "articulo_id=1" in destino


def test_stock_fisico_guarda_el_conteo_de_las_cajas_de_una_ficha():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 7, "nombre": "Banana"}),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.crear_conteo_stock") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "7", "cantidad": "12", "que_conto": "11"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(7, 12.0, ficha_id=11)
    assert "cajas+de+Vea" in respuesta.headers["location"]


def test_stock_fisico_no_deja_contar_una_ficha_de_OTRO_articulo():
    # Un conteo raro se guarda (es declarativo), pero una ficha que no es
    # de este artículo no es un conteo raro: es un dato roto.
    with (
        patch("app.main.obtener_articulo", return_value={"id": 9, "nombre": "Pera"}),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_conteos_stock_de_fecha", return_value=[]),
        patch("app.main.crear_conteo_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "9", "cantidad": "12", "que_conto": "11"},
        )

    assert respuesta.status_code == 400
    assert "no es de este artículo" in respuesta.text
    mock_crear.assert_not_called()


def test_stock_fisico_sin_decir_que_conto_da_400():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_conteos_stock_de_fecha", return_value=[]),
        patch("app.main.crear_conteo_stock") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "1", "cantidad": "12", "que_conto": ""},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_stock_fisico_los_sueltos_van_PRIMEROS_y_al_mismo_nivel_que_las_fichas():
    """Es lo contrario de "sin asignar" en la guía R, y a propósito.

    Allá es la excepción del que no sabe y va en un optgroup aparte. Acá
    es el caso más común —la mayoría de la mercadería está sin procesar—
    así que esconderlo sería hacer más difícil justo lo que más se hace.
    """
    with (
        patch("app.main.listar_articulos", return_value=[{"id": 7, "nombre": "Banana"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_conteos_stock_de_fecha", return_value=[]),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/fisico?articulo_id=7")

    cuerpo = respuesta.text.split("</style>")[-1]
    selector = cuerpo.split('name="que_conto"')[1].split("</select>")[0]
    assert "optgroup" not in selector
    # Primera opción real, antes de cualquier ficha.
    assert selector.index("Bultos sueltos") < selector.index("Cajas de")


def test_stock_fisico_acepta_cero_pero_no_negativos():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_conteo_stock") as mock_crear,
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "1", "cantidad": "0", "que_conto": "sueltos"},
            follow_redirects=False,
        )
    # 0 vale: contó y no hay ninguno.
    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, 0.0, ficha_id=None)

    with (
        patch("app.main.crear_conteo_stock") as mock_crear,
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_conteos_stock_de_fecha", return_value=[]),
    ):
        respuesta = cliente.post(
            "/deposito/stock/fisico",
            data={"articulo_id": "1", "cantidad": "-2", "que_conto": "sueltos"},
        )
    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_stock_fisico_muestra_lo_contado_hoy_sin_numeros_del_sistema():
    contados = [{"id": 5, "cantidad": 12.0, "creado_en": datetime(2026, 8, 25, 10, 30),
                 "articulo_nombre": "Banana", "ficha_id": None,
                 "ficha_nombre": None, "ficha_cliente": None}]
    with (
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_conteos_stock_de_fecha", return_value=contados),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/fisico")

    assert respuesta.status_code == 200
    assert "Contado hoy" in respuesta.text
    assert "Banana" in respuesta.text
    # Ni rastro del stock del sistema en el HTML del operario.
    assert "stock_sistema" not in respuesta.text
    assert "Sistema" not in respuesta.text


def test_cotejo_stock_compara_contra_la_foto_congelada_y_arma_el_link_de_ajuste():
    conteos = [
        {"id": 5, "articulo_id": 1, "cantidad": 12.0, "stock_sistema": 15.0,
         "creado_en": datetime(2026, 8, 25, 10, 30), "articulo_nombre": "Banana",
         "ficha_id": None, "ficha_nombre": None, "ficha_cliente": None},
        {"id": 6, "articulo_id": 2, "cantidad": 7.0, "stock_sistema": 7.0,
         "creado_en": datetime(2026, 8, 25, 11, 0), "articulo_nombre": "Anco",
         "ficha_id": None, "ficha_nombre": None, "ficha_cliente": None},
    ]
    with patch("app.main.listar_ultimos_conteos_stock", return_value=conteos):
        respuesta = cliente.get("/administracion/stock/cotejo")

    assert respuesta.status_code == 200
    # La diferencia es contra la FOTO del conteo, resaltada.
    assert "-3" in respuesta.text
    assert 'class="tarjeta con-diferencia"' in respuesta.text
    # Con diferencia: botón al ajuste precargado; sin diferencia, no.
    assert "articulo_id=1&amp;contado=12.0&amp;stock_conteo=15.0&amp;fecha_conteo=2026-08-25" in respuesta.text
    assert respuesta.text.count("Ajustar a lo contado") == 1
    assert "diferencia-cero" in respuesta.text


def test_ajustar_desde_cotejo_precarga_contra_el_stock_actual_y_avisa_si_se_movio():
    # El conteo fue con el sistema en 15, pero el stock ACTUAL es 18: el
    # ajuste para dejarlo en lo contado (12) es -6, no el -3 del cotejo —
    # y la pantalla lo explica ANTES de guardar.
    with (
        patch("app.main.stock_deposito_de_articulo", return_value=18.0),
        patch("app.main.listar_articulos", return_value=[{"id": 1, "nombre": "Banana"}]),
    ):
        respuesta = cliente.get(
            "/administracion/stock/ajustar?articulo_id=1&contado=12.0&stock_conteo=15.0&fecha_conteo=2026-08-25"
        )

    assert respuesta.status_code == 200
    assert 'value="-6.0"' in respuesta.text
    assert "Ajuste a lo contado: conteo del 2026-08-25 (12 contados)" in respuesta.text
    assert "el stock actual es 18" in respuesta.text
    assert "el ajuste sugerido para dejarlo en lo contado es -6" in respuesta.text


def test_ajustar_desde_cotejo_sin_movimientos_no_avisa():
    with (
        patch("app.main.stock_deposito_de_articulo", return_value=15.0),
        patch("app.main.listar_articulos", return_value=[{"id": 1, "nombre": "Banana"}]),
    ):
        respuesta = cliente.get(
            "/administracion/stock/ajustar?articulo_id=1&contado=12.0&stock_conteo=15.0&fecha_conteo=2026-08-25"
        )

    assert respuesta.status_code == 200
    assert 'value="-3.0"' in respuesta.text
    assert "Desde entonces hubo movimientos" not in respuesta.text


def _salidas_fifo(total, fecha=None):
    """Las salidas del FIFO de stock, ya fechadas (desde E4 van una por una).

    Los tests que solo miran cuánto salió pasan el total y les alcanza con una
    salida; la fecha por default es bien posterior a los lotes de prueba, para
    que la regla de "un lote posterior no cubre la salida" no cambie lo que el
    test estaba midiendo.
    """
    if not total:
        return []
    return [{"orden": (fecha or date(2030, 1, 1), 0), "cantidad": total}]


# --- Reproceso (tanda 1): pantalla de operario y Guías R ---


def _get_reproceso(articulos_stock=None, fichas=None, espia_total=None):
    # articulos_stock son los artículos CON SUELTOS que devuelve la base ya
    # filtrados: desde el 31/08 el filtro vive en listar_articulos_para_reproceso
    # (probado en test_db) y no en la ruta.
    # Las fichas se piden TODAS de una y cada una dice de qué cliente es: acá
    # se le dan las mismas a los dos clientes del selector.
    todas = [dict(f, cliente_id=c["id"]) for c in CLIENTES_PARA_SELECTOR for f in (fichas or [])]
    with (
        patch("app.main.listar_articulos_para_reproceso", return_value=articulos_stock or []),
        patch("app.main.stock_deposito_por_articulo",
              side_effect=espia_total or AssertionError(
                  "El reproceso NO puede filtrar por el total del artículo: toma SUELTOS")),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=todas),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        # El piso del selector sale de corte_modelo, no de una constante.
        # Acá va una fecha distinta de la real a propósito: si alguien
        # clava el 31/08 en el código, este test lo agarra.
        patch("app.main.fecha_corte", return_value=date(2026, 8, 20)),
    ):
        return cliente.get("/deposito/stock/reproceso")


def test_reproceso_lista_los_articulos_con_sueltos_por_nombre_y_sin_numeros():
    """El selector muestra lo que hay para TOMAR, que son los bultos sueltos.

    Que la ruta no pueda tocar stock_deposito_por_articulo lo garantiza el
    helper: si vuelve a filtrar por el total del artículo, revienta. Es la
    falla del 31/08 en producción — el total baja cuando salen cajas ya
    armadas, y el artículo desaparecía con la pila suelta intacta en el piso.
    """
    respuesta = _get_reproceso(articulos_stock=[{"id": 1, "nombre": "Banana"}])

    assert respuesta.status_code == 200
    # Lo que hay para tomar, POR NOMBRE...
    assert "Banana" in respuesta.text
    # ...y ningún artículo que la base no haya devuelto.
    assert "Anco" not in respuesta.text
    assert "Kiwi" not in respuesta.text


def test_reproceso_no_muestra_ninguna_cantidad_al_operario():
    """Criterio Vacíos: el número del sistema no viaja a la pantalla del operario
    ni escondido en el HTML. Por eso listar_articulos_para_reproceso devuelve id y
    nombre y nada más — acá se comprueba que la ruta tampoco agrega ninguno."""
    respuesta = _get_reproceso(articulos_stock=[{"id": 1, "nombre": "Banana"}])

    assert respuesta.status_code == 200
    assert "37.5" not in respuesta.text and "37,5" not in respuesta.text


def test_reproceso_pide_el_cliente_primero_y_la_ayuda_es_solo_de_su_ficha():
    # El cliente va PRIMERO (la primera se arma para alguien) y la ayuda
    # de kilaje muestra SOLO la ficha de ese cliente — todas juntas no
    # servían (pedido del dueño 25/08).
    fichas = [{"id": 901, "articulo_id": 1, "contenido_caja": 6.0, "unidad_venta": "kilo",
               "articulo_nombre": "Tomate Perita", "nombre_cliente": None}]
    respuesta = _get_reproceso(
        articulos_stock=[{"id": 1, "nombre": "Tomate Perita"}],
        fichas=fichas,
    )

    assert respuesta.status_code == 200
    texto = respuesta.text
    assert 'name="cliente_id"' in texto
    assert "Para qué cliente armo" in texto
    # El selector de cliente aparece ANTES que el de artículo.
    assert texto.index('name="cliente_id"') < texto.index('name="articulo_id"')
    # Las ayudas viajan por (cliente, artículo): el JS muestra solo la del elegido.
    # Los acentos van tal cual, no como \\u00XX: |tojson corre con
    # ensure_ascii=False (ver app/main.py) para que el HTML se pueda leer.
    assert '"1:1": "6 kg por caja, según la ficha de Día."' in texto
    assert '"2:1": "6 kg por caja, según la ficha de Vea."' in texto


def test_reproceso_guarda_con_cliente_y_el_aviso_repite_solo_lo_cargado():
    with (
        patch("app.main.obtener_cliente", return_value={"id": 1, "nombre": "Día"}),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Tomate Perita"}),
        patch("app.main.crear_reproceso", return_value=12) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "30",
                  "bultos_primera": "20", "bultos_segunda": "5", "bultos_merma": "5",
                  "fecha": "2026-08-25"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # El cliente queda en la guía R como DATO (el stock sigue sin dueño).
    mock_crear.assert_called_once_with(1, 30.0, 20.0, 5.0, 5.0, date(2026, 8, 25), cliente_id=1, ficha_id=None,
                                       reparto=None)
    destino = respuesta.headers["location"]
    # "Guía R12: tomé 30... para Día..." — lo cargado, jamás costos ni stock.
    assert "Gu%C3%ADa+R12" in destino
    assert "tom%C3%A9+30+bultos" in destino
    assert "para+D%C3%ADa" in destino
    assert "costo" not in destino.lower()
    assert "qued" not in destino


def test_el_reproceso_guarda_A_QUE_FICHA_fueron_las_cajas():
    """Lo que la etapa 1 vino a arreglar.

    Sin esto, la única forma de saber a qué ficha fue una guía R era
    derivarla de (cliente, artículo) — y esa derivación es ambigua POR
    DISEÑO: un cliente pide Banana Bolivia y recibe Banana Ecuador, así
    que puede tener dos fichas del mismo artículo. No es algo que se rompa
    el día que se crea una ficha nueva: ya está roto.
    """
    with (
        patch("app.main.obtener_cliente", return_value={"id": 1, "nombre": "Día"}),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_reproceso", return_value=13) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "30",
                  "bultos_primera": "20", "bultos_segunda": "5", "bultos_merma": "5",
                  "fecha": "2026-08-25", "ficha_id": "901"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        1, 30.0, 20.0, 5.0, 5.0, date(2026, 8, 25), cliente_id=1, ficha_id=901, reparto=None
    )
    # Asignada, el aviso no dice nada de "sin asignar".
    assert "sin+asignar" not in respuesta.headers["location"]


def test_SIN_ASIGNAR_es_una_eleccion_y_el_aviso_lo_dice():
    """"Sin asignar" tiene que ser una opción que se elige, no el resultado
    de no contestar: es lo que después deja distinguir "no lo sabía" de
    "me lo salteé". Y el aviso lo repite, para que no pase inadvertido.
    """
    with (
        patch("app.main.obtener_cliente", return_value={"id": 1, "nombre": "Día"}),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_reproceso", return_value=14) as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "30",
                  "bultos_primera": "20", "bultos_segunda": "0", "bultos_merma": "0",
                  "fecha": "2026-08-25", "ficha_id": "sin_asignar"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(
        1, 30.0, 20.0, 0.0, 0.0, date(2026, 8, 25), cliente_id=1, ficha_id=None, reparto=None
    )
    assert "sin+asignar" in respuesta.headers["location"]


def _pantalla_de_reproceso_con(datos, **parches):
    """El POST del reproceso con el cliente y el artículo ya resueltos."""
    contexto = {
        "app.main.obtener_cliente": {"id": 1, "nombre": "Día"},
        "app.main.obtener_articulo": {"id": 1, "nombre": "Zapallito"},
        "app.main.listar_articulos_para_reproceso": [{"id": 1, "nombre": "Zapallito"}],
        "app.main.listar_clientes": [{"id": 1, "nombre": "Día"}],
        "app.main._ayudas_ficha_por_cliente_y_articulo": {},
        "app.main._fichas_por_cliente_y_articulo": {},
    }
    with ExitStack() as pila:
        for destino, valor in contexto.items():
            pila.enter_context(patch(destino, return_value=valor))
        pila.enter_context(patch("app.main._hoy_argentina", return_value=date(2026, 9, 2)))
        if "app.main.fecha_corte" not in parches:
            pila.enter_context(patch("app.main.fecha_corte", return_value=date(2026, 8, 31)))
        if "app.main.contar_guias_r_afectadas_por_fecha" not in parches:
            # Sin guías R posteriores no hay aviso: el camino de siempre.
            pila.enter_context(
                patch("app.main.contar_guias_r_afectadas_por_fecha", return_value=0)
            )
        for destino, parche in parches.items():
            pila.enter_context(patch(destino, **parche))
        return cliente.post("/deposito/stock/reproceso", data=datos, follow_redirects=False)


def test_el_freno_devuelve_la_pantalla_ENTERA_y_dice_QUE_HACER():
    """No hay salida de escape: el reproceso es 100% o nada.

    Lo que la pared tiene que dejar en claro no es que está trabado —eso ya
    lo sabe— sino qué ir a cargar y quién lo carga. Y todo lo que tipeó
    tiene que volver puesto: la pared no se puede pagar con volver a cargar
    la pantalla entera.
    """
    freno = StockInsuficienteParaReproceso(40.0, 25.0, [
        {"tipo_lote": "guia", "origen_id": 101, "fecha_lote": date(2026, 8, 28),
         "detalle": "Norte 15", "restante": 13.0},
        {"tipo_lote": "guia", "origen_id": 102, "fecha_lote": date(2026, 9, 2),
         "detalle": "Vitale", "restante": 12.0},
    ])
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "40",
         "bultos_primera": "30", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-09-02", "ficha_id": "901"},
        **{"app.main.crear_reproceso": {"side_effect": freno}},
    )

    assert respuesta.status_code == 400
    texto = respuesta.text
    # Qué declaró y qué había: los dos números, sin vueltas.
    assert "No alcanza para 40 bultos de Zapallito el 02/09." in texto
    assert "Declaraste 40 y ese día había 25." in texto
    # El detalle por lote: fecha y cantidad.
    assert "28/08" in texto and "13" in texto
    # QUÉ HACER, y quién lo hace.
    assert "falta cargar algo antes" in texto
    assert 'href="/deposito/recepcion"' in texto
    assert 'href="/deposito/ingresar"' in texto
    assert "esa la carga Administración" in texto
    # Y NADA de cargarlo igual.
    assert "igual" not in texto.split("Las tres las cargás vos")[1].split("</div>")[0]
    # Todo lo que cargó, de vuelta en su lugar.
    assert 'value="40"' in texto and 'value="30"' in texto
    assert 'value="2026-09-02"' in texto


def test_el_freno_con_CERO_lotes_ese_dia_igual_explica_que_hacer():
    """El caso más áspero: no hay ni un lote. El mensaje no puede quedar en
    una lista vacía."""
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "5",
         "bultos_primera": "5", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-09-02", "ficha_id": "901"},
        **{"app.main.crear_reproceso": {"side_effect": StockInsuficienteParaReproceso(5.0, 0.0, [])}},
    )

    assert respuesta.status_code == 400
    assert "Ese día no había ningún lote de este artículo." in respuesta.text
    assert "falta cargar algo antes" in respuesta.text


def test_el_desglose_que_edito_el_operario_viaja_al_server():
    with patch("app.main.crear_reproceso", return_value=31) as mock_crear:
        respuesta = _pantalla_de_reproceso_con(
            {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
             "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "2",
             "fecha": "2026-09-02", "ficha_id": "901",
             "reparto": '[{"tipo_lote": "guia", "origen_id": 101, "bultos": 3}]'},
        )

    assert respuesta.status_code == 303
    assert mock_crear.call_args.kwargs["reparto"] == [
        {"tipo_lote": "guia", "origen_id": 101, "bultos": 3.0}
    ]


def test_un_reparto_ilegible_NO_se_guarda_a_medias():
    """Si la pantalla mandó algo y no se entiende, se frena. Guardar "lo que
    se pudo leer" sería inventarle al operario un reparto que no eligió."""
    with patch("app.main.crear_reproceso") as mock_crear:
        respuesta = _pantalla_de_reproceso_con(
            {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
             "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "2",
             "fecha": "2026-09-02", "ficha_id": "901", "reparto": "{roto"},
        )

    assert respuesta.status_code == 400
    assert "No se entendió de qué lotes sale" in respuesta.text
    mock_crear.assert_not_called()


def test_el_reparto_desactualizado_vuelve_a_la_pantalla_sin_guardar():
    """Entre el desglose y el Guardar alguien movió el stock."""
    with patch("app.main.crear_reproceso",
               side_effect=RepartoDesactualizado("De uno de los lotes pediste más de lo que quedaba.")):
        respuesta = _pantalla_de_reproceso_con(
            {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
             "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "2",
             "fecha": "2026-09-02", "ficha_id": "901",
             "reparto": '[{"tipo_lote": "guia", "origen_id": 101, "bultos": 10}]'},
        )

    assert respuesta.status_code == 400
    assert "Cambió el stock mientras cargabas" in respuesta.text
    assert 'value="10"' in respuesta.text


def test_el_desglose_muestra_el_proveedor_SOLO_para_desempatar_el_mismo_dia():
    """En 390px un renglón cargado de datos se vuelve ilegible y el operario
    deja de mirarlo. La regla es de la pantalla, no del dato: el detalle
    completo sigue en Guías R."""
    lotes = [
        {"tipo_lote": "guia", "origen_id": 101, "fecha_lote": date(2026, 8, 28),
         "detalle": "Norte 15", "restante": 13.0},
        {"tipo_lote": "guia", "origen_id": 102, "fecha_lote": date(2026, 9, 2),
         "detalle": "Vitale", "restante": 12.0},
        {"tipo_lote": "guia", "origen_id": 103, "fecha_lote": date(2026, 9, 2),
         "detalle": "Rodríguez", "restante": 5.0},
    ]

    with patch("app.main.lotes_para_reproceso",
               return_value={"lotes": lotes, "sin_lote": 0, "stock": 30}):
        datos = cliente.get("/deposito/stock/reproceso/desglose"
                            "?articulo_id=1&fecha=2026-09-02&bultos=20").json()

    # El 28/08 está solo ese día: no necesita proveedor.
    assert datos["lotes"][0] == {"clave": "guia:101", "tipo_lote": "guia", "origen_id": 101,
                                 "fecha": "28/08", "restante": 13.0, "detalle": ""}
    # Los dos del 02/09 sí, o no se distinguen.
    assert datos["lotes"][1]["detalle"] == "Vitale"
    assert datos["lotes"][2]["detalle"] == "Rodríguez"


def test_el_desglose_propone_del_mas_viejo_primero_y_avisa_si_no_alcanza():
    lotes = [
        {"tipo_lote": "guia", "origen_id": 101, "fecha_lote": date(2026, 8, 28),
         "detalle": "Norte 15", "restante": 13.0},
        {"tipo_lote": "guia", "origen_id": 102, "fecha_lote": date(2026, 9, 2),
         "detalle": "Vitale", "restante": 12.0},
    ]
    with patch("app.main.lotes_para_reproceso",
               return_value={"lotes": lotes, "sin_lote": 0, "stock": 25}):
        alcanza = cliente.get("/deposito/stock/reproceso/desglose"
                              "?articulo_id=1&fecha=2026-09-02&bultos=20").json()
        no_alcanza = cliente.get("/deposito/stock/reproceso/desglose"
                                 "?articulo_id=1&fecha=2026-09-02&bultos=40").json()

    assert alcanza["disponible"] == 25.0
    assert alcanza["alcanza"] is True
    assert alcanza["propuesta"] == {"guia:101": 13.0, "guia:102": 7.0}
    # El mismo número que va a usar el freno: es una sola definición.
    assert no_alcanza["alcanza"] is False
    assert no_alcanza["disponible"] == 25.0


def test_los_lotes_VACIOS_no_llegan_a_la_pantalla():
    """Un lote sin nada que dar no es una opción: solo ocupa renglón."""
    lotes = [
        {"tipo_lote": "guia", "origen_id": 101, "fecha_lote": date(2026, 8, 28),
         "detalle": "Norte 15", "restante": 0.0},
        {"tipo_lote": "guia", "origen_id": 102, "fecha_lote": date(2026, 9, 2),
         "detalle": "Vitale", "restante": 12.0},
    ]
    with patch("app.main.lotes_para_reproceso",
               return_value={"lotes": lotes, "sin_lote": 0, "stock": 12}):
        datos = cliente.get("/deposito/stock/reproceso/desglose"
                            "?articulo_id=1&fecha=2026-09-02&bultos=5").json()

    assert [lote["clave"] for lote in datos["lotes"]] == ["guia:102"]


def test_la_pantalla_de_reproceso_ofrece_las_fichas_del_cliente_y_del_articulo():
    # El select se llena en el navegador con este mapa: puede haber DOS
    # fichas del mismo cliente y artículo, y elegir por él sería adivinar.
    fichas = [
        {"id": 901, "articulo_id": 1, "contenido_caja": 6.0, "unidad_venta": "kilo",
         "articulo_nombre": "Banana", "nombre_cliente": "Banana Bolivia"},
        {"id": 902, "articulo_id": 1, "contenido_caja": 8.0, "unidad_venta": "kilo",
         "articulo_nombre": "Banana", "nombre_cliente": "Banana Ecuador"},
    ]
    respuesta = _get_reproceso(
        articulos_stock=[{"id": 1, "nombre": "Banana"}],
        fichas=fichas,
    )

    assert respuesta.status_code == 200
    assert 'name="ficha_id"' in respuesta.text
    # Las dos fichas viajan bajo la misma clave cliente:articulo.
    assert '"1:1"' in respuesta.text
    assert "Banana Bolivia" in respuesta.text
    assert "Banana Ecuador" in respuesta.text


def test_reproceso_sin_cliente_da_400():
    with (
        patch("app.main.obtener_cliente", return_value=None),
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Tomate Perita"}),
        patch("app.main.crear_reproceso") as mock_crear,
        patch("app.main.listar_articulos_para_reproceso", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.fecha_corte", return_value=date(2026, 8, 20)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"cliente_id": "", "articulo_id": "1", "bultos_tomados": "30",
                  "bultos_primera": "20", "bultos_segunda": "", "bultos_merma": ""},
        )

    assert respuesta.status_code == 400
    assert "Elegí para qué cliente estás armando" in respuesta.text
    mock_crear.assert_not_called()


def test_reproceso_sin_nada_producido_da_400():
    with patch("app.main.crear_reproceso") as mock_crear, patch("app.main.listar_articulos_para_reproceso", return_value=[]), patch("app.main.listar_clientes", return_value=[]), patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]), patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)), patch("app.main.fecha_corte", return_value=date(2026, 8, 20)):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"articulo_id": "1", "bultos_tomados": "5", "bultos_primera": "",
                  "bultos_segunda": "", "bultos_merma": ""},
        )

    assert respuesta.status_code == 400
    assert "Cargá en qué se transformó" in respuesta.text
    mock_crear.assert_not_called()


def test_reproceso_fecha_futura_da_400():
    with patch("app.main.crear_reproceso") as mock_crear, patch("app.main.listar_articulos_para_reproceso", return_value=[]), patch("app.main.listar_clientes", return_value=[]), patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]), patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)), patch("app.main.fecha_corte", return_value=date(2026, 8, 20)):
        respuesta = cliente.post(
            "/deposito/stock/reproceso",
            data={"articulo_id": "1", "bultos_tomados": "5", "bultos_primera": "3",
                  "fecha": "2026-08-26"},
        )

    assert respuesta.status_code == 400
    assert "no puede ser futura" in respuesta.text
    mock_crear.assert_not_called()


# --- El piso de la fecha y el aviso de la fecha hacia atrás (04/09) ---
# Los tres caminos del corte: nada antes de la fecha de corte, aviso con
# segundo toque cuando la fecha vieja mueve el reparto de guías R que ya
# están, y el camino de siempre intacto cuando la fecha es la de hoy.


def test_el_piso_de_la_fecha_sale_del_CORTE_y_no_de_una_constante():
    """La fecha de corte se mueve cada vez que se hace un corte nuevo.

    Por eso el rechazo dice la fecha que devuelve fecha_corte() y no un
    31/08 escrito a mano: acá el corte es el 20/08 y el mensaje tiene que
    decir 20/08. Si alguien clava la constante, este test lo agarra.
    """
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
         "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-08-15", "ficha_id": "sin_asignar"},
        **{"app.main.fecha_corte": {"return_value": date(2026, 8, 20)},
           "app.main.crear_reproceso": {
               "side_effect": ReprocesoAnteriorAlCorte(date(2026, 8, 15), date(2026, 8, 20))}},
    )

    assert respuesta.status_code == 400
    assert "no puede ser anterior al 20/08/2026" in respuesta.text
    # La fecha real de hoy no puede aparecer: sería la constante clavada.
    assert "31/08/2026" not in respuesta.text


def test_la_fecha_hacia_atras_AVISA_cuantas_guias_R_quedan_desactualizadas():
    """El molde de las señas: contar antes de escribir, mostrar y pedir el segundo toque.

    Y la frase del costo es OBLIGATORIA. Es la que evita el susto: lo
    primero que se piensa al leer "puede dejar de coincidir" es que se
    movió plata, y no se movió — el costo está congelado.
    """
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
         "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-08-31", "ficha_id": "sin_asignar"},
        **{"app.main.contar_guias_r_afectadas_por_fecha": {"return_value": 3},
           "app.main.crear_reproceso": {}},
    )

    assert respuesta.status_code == 200
    texto = respuesta.text
    assert "Hay 3" in texto and "guías R" in texto
    assert "Zapallito" in texto
    assert "31/08/2026" in texto
    # LA FRASE. Sin ella el aviso asusta por la razón equivocada.
    assert "Su costo no cambia." in texto
    assert "Guardar igual" in texto


def test_el_aviso_de_la_fecha_NO_escribe_nada():
    """Contar va ANTES de escribir: el primer toque no puede dejar media guía R."""
    with patch("app.main.crear_reproceso") as mock_crear:
        _pantalla_de_reproceso_con(
            {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
             "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
             "fecha": "2026-08-31", "ficha_id": "sin_asignar"},
            **{"app.main.contar_guias_r_afectadas_por_fecha": {"return_value": 3}},
        )
    mock_crear.assert_not_called()


def test_el_segundo_toque_guarda_con_la_fecha_vieja():
    """Confirmado, entra con la fecha que el operario dijo: la fecha real es un dato, no un permiso."""
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
         "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-08-31", "ficha_id": "sin_asignar", "confirmado": "1"},
        **{"app.main.contar_guias_r_afectadas_por_fecha": {"return_value": 3},
           "app.main.crear_reproceso": {"return_value": 77}},
    )

    assert respuesta.status_code == 303
    assert "R77" in respuesta.headers["location"]


def test_sin_guias_R_posteriores_NO_avisa_y_guarda_derecho():
    """Un aviso que aparece cuando no hay nada que avisar es un aviso que se deja de leer."""
    with patch("app.main.contar_guias_r_afectadas_por_fecha", return_value=0) as mock_contar:
        respuesta = _pantalla_de_reproceso_con(
            {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
             "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
             "fecha": "2026-08-31", "ficha_id": "sin_asignar"},
            **{"app.main.contar_guias_r_afectadas_por_fecha": {"return_value": 0},
               "app.main.crear_reproceso": {"return_value": 78}},
        )

    assert respuesta.status_code == 303


def test_con_la_fecha_de_HOY_no_se_cuenta_nada():
    """El camino normal no paga un toque de más: con la fecha de hoy no hay
    ninguna guía R posterior que se pueda desactualizar, así que ni se
    pregunta."""
    espia = MagicMock(return_value=9)
    respuesta = _pantalla_de_reproceso_con(
        {"cliente_id": "1", "articulo_id": "1", "bultos_tomados": "10",
         "bultos_primera": "8", "bultos_segunda": "0", "bultos_merma": "0",
         "fecha": "2026-09-02", "ficha_id": "sin_asignar"},
        **{"app.main.contar_guias_r_afectadas_por_fecha": {"new": espia},
           "app.main.crear_reproceso": {"return_value": 79}},
    )

    assert respuesta.status_code == 303
    espia.assert_not_called()


def test_el_selector_de_fecha_tiene_el_piso_del_corte_puesto():
    """Comodidad, no la regla: que no pueda elegir una fecha que va a rebotar."""
    respuesta = _get_reproceso(articulos_stock=[{"id": 1, "nombre": "Banana"}])

    assert 'min="2026-08-20"' in respuesta.text


GUIAS_R_DE_PRUEBA = [
    {"id": 12, "articulo_id": 1, "fecha_operacion": date(2026, 8, 25), "bultos_tomados": 30.0,
     "bultos_primera": 20.0, "bultos_segunda": 5.0, "bultos_merma": 5.0,
     "costo_total": 33000.0, "costo_por_bulto_primera": 1650.0,
     "creado_en": datetime(2026, 8, 25, 15, 0), "anulado_el": None,
     "articulo_nombre": "Tomate Perita",
     "consumos": [
         {"origen": "compra", "origen_id": 101, "bultos": 20.0, "costo_por_bulto": 1000.0,
          "guia_fecha": date(2026, 8, 24), "proveedor_nombre": "Norte 15"},
         {"origen": "compra", "origen_id": 102, "bultos": 10.0, "costo_por_bulto": 1300.0,
          "guia_fecha": date(2026, 8, 25), "proveedor_nombre": "Sur 3"},
     ]},
    {"id": 13, "articulo_id": 2, "fecha_operacion": date(2026, 8, 25), "bultos_tomados": 4.0,
     "bultos_primera": 6.0, "bultos_segunda": 0.0, "bultos_merma": 0.0,
     "costo_total": None, "costo_por_bulto_primera": None,
     "creado_en": datetime(2026, 8, 25, 16, 0), "anulado_el": datetime(2026, 8, 25, 17, 0),
     "articulo_nombre": "Anco",
     "consumos": [
         {"origen": "ajuste", "origen_id": 1, "bultos": 4.0, "costo_por_bulto": None,
          "guia_fecha": None, "proveedor_nombre": None},
     ]},
]


# La que SÍ se puede completar: lo único que le falta es el precio de una
# COMPRA, que es lo que alguien puede ir a cargar. La R13 de arriba es el otro
# caso —consumió un ajuste— y ésa no se puede cerrar nunca.
GUIA_R_ESPERANDO_PRECIO = {
    "id": 30, "articulo_id": 1, "fecha_operacion": date(2026, 8, 25), "bultos_tomados": 10.0,
    "bultos_primera": 8.0, "bultos_segunda": 2.0, "bultos_merma": 0.0,
    "costo_total": None, "costo_por_bulto_primera": None,
    "creado_en": datetime(2026, 8, 25, 18, 0), "anulado_el": None,
    "articulo_nombre": "Tomate Perita",
    "consumos": [
        {"origen": "compra", "origen_id": 103, "bultos": 10.0, "costo_por_bulto": None,
         "guia_fecha": date(2026, 8, 24), "proveedor_nombre": "Norte 15"},
    ],
}


def test_guias_r_muestra_trazabilidad_costo_y_marca_incompleto():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_reprocesos_por_rango", return_value=[dict(g) for g in GUIAS_R_DE_PRUEBA]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    assert respuesta.status_code == 200
    assert "Guía R12 — Tomate Perita" in respuesta.text
    assert "Tomó 30 bultos" in respuesta.text
    # La trazabilidad hacia atrás: de qué guías de compra consumió.
    assert "De la guía 24/08 · Norte 15" in respuesta.text
    assert "De la guía 25/08 · Sur 3" in respuesta.text
    # El costo, TODO a la primera: $33.000 total, $1.650 por caja.
    assert "$33.000 total" in respuesta.text
    assert "$1.650 por caja de primera" in respuesta.text
    # La anulada queda visible, marcada y sin botón de anular.
    assert "ANULADA" in respuesta.text
    assert respuesta.text.count("Anular guía") == 1
    # Y la única sin costo de esta fixture es la ANULADA: su cartel y su
    # detalle de lotes NO salen (ver los dos tests de abajo).
    assert "COSTO INCOMPLETO" not in respuesta.text
    assert "De un ajuste (ej. stock inicial)" not in respuesta.text


def test_guias_r_una_guia_VIGENTE_sin_costo_si_muestra_el_cartel_y_el_detalle():
    # El cartel tiene que seguir estando donde SÍ hay algo que hacer: le falta
    # el precio de una COMPRA, y ese precio puede llegar.
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=[dict(GUIA_R_ESPERANDO_PRECIO)]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "COSTO INCOMPLETO" in cuerpo
    assert "De la guía 24/08 · Norte 15" in cuerpo
    # Y el botón que lo resuelve.
    assert "Completar costo" in cuerpo


def test_guias_r_una_guia_SIN_COSTO_POSIBLE_no_ofrece_el_boton_que_no_puede_hacer_nada():
    """Un ajuste, un stock inicial sin costo o el compensatorio del corte no
    tienen precio ni lo van a tener: "Completar costo" no tiene de dónde
    sacarlo y siempre devolvía "sigue con costo incompleto". El cartel gritaba
    sin ninguna acción detrás, que es lo que esta pantalla ya le había sacado
    a las anuladas.
    """
    guia = dict(GUIAS_R_DE_PRUEBA[1], id=20, anulado_el=None)
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=[guia]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "SIN COSTO POSIBLE" in cuerpo
    assert "De un ajuste (ej. stock inicial)" in cuerpo
    # El detalle se sigue viendo. Lo que NO va es el botón ni el cartel que
    # promete que se arregla cargando algo.
    assert "Completar costo" not in cuerpo
    assert "COSTO INCOMPLETO" not in cuerpo


def test_guias_r_el_consumo_del_compensatorio_no_se_puede_completar():
    """El lote fantasma del corte: la base PROHÍBE que tenga costo."""
    guia = dict(GUIAS_R_DE_PRUEBA[1], id=22, anulado_el=None, consumos=[
        {"origen": "cierre_modelo_viejo", "origen_id": 9, "bultos": 4.0, "costo_por_bulto": None,
         "guia_fecha": None, "proveedor_nombre": None},
    ])
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=[guia]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "Del cierre del modelo viejo (sin costo posible)" in cuerpo
    assert "SIN COSTO POSIBLE" in cuerpo
    assert "Completar costo" not in cuerpo


def test_guias_r_muestra_el_total_de_las_que_no_se_pueden_cerrar_nunca():
    """El número que NO va al banner: se ve acá, y con el "no hay nada que
    cargar" al lado para que nadie lo lea como un pendiente."""
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=[]),
        patch("app.main.contar_reprocesos_sin_costo_posible",
              return_value={"casos": 7, "mas_viejo": date(2026, 8, 31)}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "7 guías sin costo posible" in cuerpo
    assert "31/08" in cuerpo
    assert "No hay nada que cargar para arreglarlas" in cuerpo


def test_guias_r_una_guia_ANULADA_no_grita_nada():
    """En una anulada no hay nada que completar ni nada que opinar.

    Lo tomado ya volvió a sus lotes y la primera salió del stock. El
    cartel de COSTO INCOMPLETO gritaba sin ninguna acción detrás: el
    botón "Completar costo" ya estaba escondido en las anuladas.
    """
    guia = dict(GUIAS_R_DE_PRUEBA[1], id=21,
                anulado_el=datetime(2026, 8, 25, 17, 0), cliente_nombre="Día")
    cruces = [{"reproceso_id": 21, "cliente_salida_nombre": "Vea", "bultos": 3.0}]
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=[guia]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=cruces),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    # Sigue visible y marcada: no se esconde la guía, se esconde el ruido.
    assert "Guía R21" in cuerpo
    assert "ANULADA" in cuerpo
    assert "Tomó 4 bultos" in cuerpo
    # Nada de esto va en una anulada.
    assert "COSTO INCOMPLETO" not in cuerpo
    assert "CRUCE DE CLIENTE" not in cuerpo
    assert "De un ajuste (ej. stock inicial)" not in cuerpo
    assert "Completar costo" not in cuerpo
    assert "Anular guía" not in cuerpo


def test_anular_guia_r_vuelve_al_rango():
    with patch("app.main.anular_reproceso") as mock_anular:
        respuesta = cliente.post(
            "/administracion/stock/guias-r/12/anular",
            data={"fecha_desde": "2026-08-18", "fecha_hasta": "2026-08-25"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_anular.assert_called_once_with(12)
    assert "fecha_desde=2026-08-18" in respuesta.headers["location"]


def test_el_control_de_stock_vive_en_administracion():
    # Las pantallas de control, en el hub que les corresponde.
    respuesta = cliente.get("/administracion")

    assert respuesta.status_code == 200
    for destino in ("/administracion/stock/remanente", "/administracion/stock/cotejo",
                    "/administracion/stock/ajustar", "/administracion/stock/movimientos",
                    "/administracion/stock/guias-r", "/administracion/pedidos/buscar"):
        assert f'href="{destino}"' in respuesta.text, destino
    # Stock del Sistema se BORRÓ el 06/09: el Remanente ocupa su lugar. El
    # botón no puede quedar apuntando a una ruta que ya no responde.
    assert 'href="/administracion/stock/sistema"' not in respuesta.text
    # El reproceso NO se muda: es carga del depósito.
    assert "/deposito/stock/reproceso" not in respuesta.text


# --- Reproceso (tanda 2): segunda, remito al Puesto, completar costo ---


# Los tres tests del desglose de Stock del Sistema (sin procesar / armado por
# guía R / segunda / total por artículo) murieron con la pantalla el 06/09, y
# no se mudaron a ningún lado A PROPÓSITO: ese desglose es justamente lo que
# confundía. El "sin procesar" era la cuenta 3 —el FIFO rejugado, que daba −95
# sin que eso fuera una cantidad— y el total por artículo era la suma que el
# dueño ya había pedido no mostrar ("80 cajones sin procesar + 40 cajas armadas
# no son 120 bultos"). El Remanente muestra una porción por renglón, sin total,
# y la trazabilidad por guía R vive en Stock por Guía.
#
# El total global de segunda SÍ sobrevive, en el Remanente: ver
# test_el_remanente_muestra_los_dos_pools_aparte_y_en_chico.


def test_remito_segunda_lista_solo_articulos_con_segunda_y_guarda():
    filas = [
        dict(FILAS_STOCK_DE_PRUEBA[0], segunda=3.0),
        dict(FILAS_STOCK_DE_PRUEBA[1], segunda=0.0),
    ]
    with (
        patch("app.main.stock_deposito_por_articulo", return_value=filas),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.get("/deposito/stock/remito-segunda")

    assert respuesta.status_code == 200
    assert "Banana" in respuesta.text
    assert "Anco" not in respuesta.text

    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Banana"}),
        patch("app.main.crear_remito_segunda") as mock_crear,
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    ):
        respuesta = cliente.post(
            "/deposito/stock/remito-segunda",
            data={"articulo_id": "1", "cantidad": "2", "fecha": "2026-08-24"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, 2.0, date(2026, 8, 24))
    destino = respuesta.headers["location"]
    assert "2+bultos+de+segunda+de+Banana" in destino
    assert "al+Puesto" in destino


def test_movimientos_incluye_los_remitos_de_segunda_con_su_anular():
    remitos = [{"id": 7, "bultos": 2.0, "fecha_operacion": date(2026, 8, 25),
                "creado_en": datetime(2026, 8, 25, 14, 0), "anulado_el": None,
                "articulo_nombre": "Banana"}]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_movimientos_stock_por_rango", return_value=[]),
        patch("app.main.listar_remitos_segunda_por_rango", return_value=remitos),
    ):
        respuesta = cliente.get("/administracion/stock/movimientos")

    assert respuesta.status_code == 200
    assert "Remito 2ª" in respuesta.text
    assert "Segunda remitida al Puesto" in respuesta.text
    assert 'action="/administracion/stock/movimientos/remitos/7/anular"' in respuesta.text
    # Sale del pool: cantidad negativa; y no tiene foto del sistema.
    assert "-2" in respuesta.text
    assert "el sistema decía" not in respuesta.text


def test_anular_remito_segunda_vuelve_al_rango():
    with patch("app.main.anular_remito_segunda") as mock_anular:
        respuesta = cliente.post(
            "/administracion/stock/movimientos/remitos/7/anular",
            data={"fecha_desde": "2026-08-18", "fecha_hasta": "2026-08-25"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_anular.assert_called_once_with(7)
    assert "fecha_desde=2026-08-18" in respuesta.headers["location"]


def test_completar_costo_avisa_si_completo_o_que_sigue_incompleto():
    with patch("app.main.completar_costo_reproceso", return_value={"completado": True, "sin_precio": 0}) as mock_completar:
        respuesta = cliente.post(
            "/administracion/stock/guias-r/12/completar-costo",
            data={"fecha_desde": "2026-08-18", "fecha_hasta": "2026-08-25"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_completar.assert_called_once_with(12)
    assert "completado" in respuesta.headers["location"]

    with patch("app.main.completar_costo_reproceso", return_value={"completado": False, "sin_precio": 2}):
        respuesta = cliente.post(
            "/administracion/stock/guias-r/13/completar-costo",
            data={"fecha_desde": "2026-08-18", "fecha_hasta": "2026-08-25"},
            follow_redirects=False,
        )
    assert "sigue+con+costo+incompleto" in respuesta.headers["location"]
    assert "2+consumos+sin+precio" in respuesta.headers["location"]


def test_guias_r_muestra_el_boton_completar_solo_en_incompletas():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_reprocesos_por_rango", return_value=[dict(g) for g in GUIAS_R_DE_PRUEBA]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    assert respuesta.status_code == 200
    # La R12 tiene costo completo y la R13 está ANULADA: ninguna lo lleva.
    assert "Completar costo" not in respuesta.text

    # La que sí lo lleva es la que espera el precio de una COMPRA. La R13
    # consumió un ajuste: aunque esté vigente, no hay nada que completar.
    con_incompleta = [dict(GUIA_R_ESPERANDO_PRECIO, id=13)]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_reprocesos_por_rango", return_value=con_incompleta),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    # Se cuenta el BOTÓN, no la frase: el cartel de arriba también nombra a
    # "Completar costo" para decir que con eso se cierra.
    assert respuesta.text.count('class="boton-completar"') == 1
    assert 'action="/administracion/stock/guias-r/13/completar-costo"' in respuesta.text


def test_hub_stock_tiene_remitir_segunda():
    respuesta = cliente.get("/deposito/stock")

    assert respuesta.status_code == 200
    assert 'href="/deposito/stock/remito-segunda"' in respuesta.text


# --- Rentabilidad Real (tanda 1) ---


def test_gerencia_muestra_el_boton_de_rentabilidad_real_aparte():
    respuesta = cliente.get("/gerencia")

    assert respuesta.status_code == 200
    # Botón APARTE: la teórica sigue con el suyo, intacta.
    assert 'href="/gerencia/rentabilidad"' in respuesta.text
    assert 'href="/gerencia/rentabilidad-real"' in respuesta.text


def test_rentabilidad_real_sin_cliente_muestra_solo_los_filtros():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
    ):
        respuesta = cliente.get("/gerencia/rentabilidad-real")

    assert respuesta.status_code == 200
    assert "Elegí un cliente..." in respuesta.text
    assert "Afuera del cálculo" not in respuesta.text


RESULTADO_REAL_DE_PRUEBA = {
    "grupos": [{
        "grupo": "fruta", "etiqueta": "Fruta",
        "filas": [{
            "articulo_id": 1, "articulo_nombre": "Banana", "grupo": "fruta",
            "bultos": 10.0, "unidades": 160.0, "venta_neta": 14400.0,
            "costo_mercaderia": 5000.0, "costo_envase": 320.0,
            # Los $1.500 de merma, abiertos: 2 bultos crudos a $350 y 1 ya
            # trabajado a $800 — el trabajado sale más caro por bulto.
            "costo_mermas": 1500.0, "bultos_mermados": 3.0,
            "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
            "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
            "segunda_bultos": 2.0,
            "devoluciones_bultos": 0.0, "devoluciones_venta": 0.0,
            "rechazos_perdidos": 0.0, "rechazos_bultos": 0.0,
            "costo_total": 6820.0, "renta_pesos": 7580.0, "utilidad_pct": 151.6,
        }],
        "subtotal": {"bultos": 10.0, "venta_neta": 14400.0, "costo_mercaderia": 5000.0,
                     "costo_envase": 320.0, "costo_mermas": 1500.0, "costo_total": 6820.0, "bultos_mermados": 3.0,
                     "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
                     "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
                     "devoluciones_bultos": 0.0, "devoluciones_venta": 0.0,
                     "rechazos_perdidos": 0.0, "rechazos_bultos": 0.0,
                     "renta_pesos": 7580.0, "utilidad_pct": 151.6},
    }],
    "totales": {"bultos": 10.0, "venta_neta": 14400.0, "costo_mercaderia": 5000.0,
                "costo_envase": 320.0, "costo_mermas": 1500.0, "segunda_bultos": 2.0, "bultos_mermados": 3.0,
                "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
                "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
                "devoluciones_bultos": 0.0, "devoluciones_venta": 0.0,
                "rechazos_perdidos": 0.0, "rechazos_bultos": 0.0,
                "afuera_bultos": 42.0, "afuera_motivos": 2, "costo_total": 6820.0,
                "renta_pesos": 7580.0, "utilidad_pct": 151.6},
    "afuera_por_motivo": [
        {"motivo": "ajuste_sin_costo", "etiqueta": "Consumió un ajuste (sin costo posible)",
         "bultos": 30.0, "articulos": [{"nombre": "Banana", "bultos": 30.0}]},
        {"motivo": "sin_kilaje", "etiqueta": "Renglón armado sin kilaje cargado",
         "bultos": 12.0, "articulos": [{"nombre": "Anco", "bultos": 12.0}]},
    ],
    "fechas_incluidas": [date(2026, 8, 25)],
}


def test_rentabilidad_real_muestra_la_cuenta_y_el_afuera_protagonista():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main._datos_rentabilidad_real", return_value=dict(RESULTADO_REAL_DE_PRUEBA)) as mock_datos,
        patch("app.main._datos_rentabilidad", return_value={"grupos": [], "totales": {
            "bultos": 0, "venta_neta": 0, "costo_mercaderia": 0, "costo_envase": 0, "costo_total": 0,
            "renta_pesos": 0, "utilidad_pct": None, "no_calculables_casos": 0, "no_calculables_bultos": 0,
        }, "no_calculables": [], "fechas_incluidas": []}),
    ):
        respuesta = cliente.get("/gerencia/rentabilidad-real?cliente_id=1")

    assert respuesta.status_code == 200
    mock_datos.assert_called_once_with(1, date(2026, 8, 18), date(2026, 8, 25), None, None)
    # La cuenta real: renta, mermas y segunda a la vista.
    assert "$7.580" in respuesta.text
    assert "mermas $1.500" in respuesta.text
    assert "Segunda generada: 2 bultos (vale cero)." in respuesta.text
    # El AFUERA protagonista: total en grande, bultos por motivo, artículos.
    assert "Afuera del cálculo — no sumó como cero" in respuesta.text
    assert "42 bultos afuera, por 2 motivos" in respuesta.text
    assert "Consumió un ajuste (sin costo posible)" in respuesta.text
    assert "Banana (30)" in respuesta.text
    assert "Anco (12)" in respuesta.text
    # Y aparece ANTES que la tabla de grupos (arriba, no nota al pie).
    assert respuesta.text.index("Afuera del cálculo") < respuesta.text.index("Subtotal Fruta")


def test_rentabilidad_real_abre_la_merma_en_cruda_y_trabajada():
    # Lo que se quiere saber mirando la pantalla no es cuánto se tiró, sino
    # QUÉ se tiró: materia prima o trabajo. El total sigue siendo el mismo.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main._datos_rentabilidad_real", return_value=dict(RESULTADO_REAL_DE_PRUEBA)),
        patch("app.main._datos_rentabilidad", return_value={"grupos": [], "totales": {
            "bultos": 0, "venta_neta": 0, "costo_mercaderia": 0, "costo_envase": 0, "costo_total": 0,
            "renta_pesos": 0, "utilidad_pct": None, "no_calculables_casos": 0, "no_calculables_bultos": 0,
        }, "no_calculables": [], "fechas_incluidas": []}),
    ):
        respuesta = cliente.get("/gerencia/rentabilidad-real?cliente_id=1")

    assert "Mermas: ¿materia prima o trabajo?" in respuesta.text
    assert "Mercadería cruda" in respuesta.text
    assert "Mercadería ya trabajada" in respuesta.text
    # $700 en 2 bultos crudos y $800 en 1 trabajado, del mismo total de $1.500.
    assert "$700" in respuesta.text and "$800" in respuesta.text
    assert "$1.500 en total, abierto por lo que se tiró." in respuesta.text
    # Y el artículo lleva un chip por cada tipo, con sus propios bultos.
    assert "merma cruda 2 bultos · $700" in respuesta.text
    assert "merma trabajada 1 bulto · $800" in respuesta.text


def test_rentabilidad_real_sin_mermas_no_muestra_el_bloque_de_mermas():
    # Nada de renglones vacíos: si no se tiró nada, el bloque no va.
    sin_mermas = dict(RESULTADO_REAL_DE_PRUEBA)
    sin_mermas["totales"] = dict(RESULTADO_REAL_DE_PRUEBA["totales"], costo_mermas=0.0,
                                 costo_mermas_cruda=0.0, costo_mermas_trabajada=0.0)
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main._datos_rentabilidad_real", return_value=sin_mermas),
        patch("app.main._datos_rentabilidad", return_value={"grupos": [], "totales": {
            "bultos": 0, "venta_neta": 0, "costo_mercaderia": 0, "costo_envase": 0, "costo_total": 0,
            "renta_pesos": 0, "utilidad_pct": None, "no_calculables_casos": 0, "no_calculables_bultos": 0,
        }, "no_calculables": [], "fechas_incluidas": []}),
    ):
        respuesta = cliente.get("/gerencia/rentabilidad-real?cliente_id=1")

    assert "Mermas: ¿materia prima o trabajo?" not in respuesta.text


def test_rentabilidad_real_junta_historia_completa_y_ancla_precios_por_fecha():
    entradas = [{"fecha_orden": date(2026, 8, 20), "momento_orden": datetime(2026, 8, 20, 10), "orden": (date(2026, 8, 20), datetime(2026, 8, 20, 10)),
                 "tipo_lote": "guia", "origen_id": 9, "fecha_lote": date(2026, 8, 20),
                 "detalle": "Norte", "motivo": None, "cantidad": 10.0, "costo_bulto": 500.0}]
    # La salida de armado lleva su FICHA: es lo que ancla el precio.
    salidas = [{"fecha_orden": date(2026, 8, 25), "momento_orden": datetime(2026, 8, 25, 13), "orden": (date(2026, 8, 25), datetime(2026, 8, 25, 13)),
                "tipo": "armado", "fecha": date(2026, 8, 25), "cantidad": 4.0,
                "unidades": 64.0, "cliente_id": 1, "motivo": None, "bultos_segunda": None,
                "ficha_id": 901}]
    with (
        patch("app.main.articulos_con_salidas_stock",
              return_value=[{"articulo_id": 1, "nombre": "Banana", "grupo": "fruta"}]) as mock_articulos,
        patch("app.main.devoluciones_vinculadas_por_rango", return_value=[]),
        patch("app.main.entradas_y_salidas_stock_articulos", return_value={1: (entradas, _salidas_fifo(4.0))}),
        patch("app.main.salidas_stock_articulos", return_value={1: salidas}),
        patch("app.main.calcular_listados_para_negociar_precios",
              side_effect=lambda cliente_id, momentos: {
                  momento: [{"ficha_id": 901, "articulo_id": 1, "precio_vigente": 100.0,
                             "costo_actual": 60.0, "costo_envase_unidad_venta": 2.0,
                             "denominador_tasas": 0.9}] for momento in momentos
              }) as mock_listado,
    ):
        resultado = __import__("app.main", fromlist=["x"])._datos_rentabilidad_real(
            1, date(2026, 8, 18), date(2026, 8, 25), None, None
        )

    mock_articulos.assert_called_once_with(1, date(2026, 8, 18), date(2026, 8, 25))
    # El listado sigue anclado a cada fecha con armados del rango, pero se
    # pide UNA sola vez con todas juntas.
    mock_listado.assert_called_once()
    assert [m.date() for m in mock_listado.call_args.args[1]] == [date(2026, 8, 25)]
    fila = resultado["grupos"][0]["filas"][0]
    assert fila["venta_neta"] == 64.0 * 100.0 * 0.9
    assert fila["costo_mercaderia"] == 4 * 500.0


# --- Rentabilidad Real (tanda 2): comparativa y exports ---

TOTALES_TEORICA_DE_PRUEBA = {
    "bultos": 20.0, "venta_neta": 20000.0, "costo_mercaderia": 12000.0, "costo_envase": 500.0,
    "costo_total": 12500.0, "renta_pesos": 7500.0, "utilidad_pct": 62.5,
    "no_calculables_casos": 0, "no_calculables_bultos": 0.0,
}


def test_rentabilidad_real_muestra_la_comparativa_contra_la_teorica():
    teorica = {"grupos": [], "totales": dict(TOTALES_TEORICA_DE_PRUEBA), "no_calculables": [],
               "fechas_incluidas": []}
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main._datos_rentabilidad_real", return_value=dict(RESULTADO_REAL_DE_PRUEBA)),
        patch("app.main._datos_rentabilidad", return_value=teorica) as mock_teorica,
    ):
        respuesta = cliente.get("/gerencia/rentabilidad-real?cliente_id=1")

    assert respuesta.status_code == 200
    # La teórica se calcula con LOS MISMOS filtros y rango.
    mock_teorica.assert_called_once_with(1, date(2026, 8, 18), date(2026, 8, 25), None, None, [])
    # Las tres celdas: teórica, real y la diferencia (real 7.580 − teórica
    # 7.500 = +80) — "la lista de cosas a explicar".
    assert "Teórica (lo pedido)" in respuesta.text
    assert "Real (lo enviado)" in respuesta.text
    assert "$7.500" in respuesta.text
    assert "$7.580" in respuesta.text
    assert "+$80" in respuesta.text
    assert "la lista de cosas a explicar" in respuesta.text
    # Los botones de export de la Real.
    assert "/gerencia/rentabilidad-real/exportar-pdf?cliente_id=1" in respuesta.text
    assert "/gerencia/rentabilidad-real/exportar-excel?cliente_id=1" in respuesta.text


def test_exportar_rentabilidad_real_pdf_y_excel():
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.obtener_cliente", return_value={"id": 1, "nombre": "Día"}),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main._datos_rentabilidad_real", return_value=dict(RESULTADO_REAL_DE_PRUEBA)),
    ):
        pdf = cliente.get("/gerencia/rentabilidad-real/exportar-pdf?cliente_id=1")
        excel = cliente.get("/gerencia/rentabilidad-real/exportar-excel?cliente_id=1")

    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")
    assert "Rentabilidad_Real_" in pdf.headers["content-disposition"]
    assert excel.status_code == 200
    assert len(excel.content) > 1000


def test_exportar_rentabilidad_real_sin_cliente_da_400():
    respuesta = cliente.get("/gerencia/rentabilidad-real/exportar-pdf")
    assert respuesta.status_code == 400


# --- Fix del tick de la casilla: latido, origen automático y lifespan ---


def test_la_app_usa_lifespan_y_no_quedan_handlers_on_event():
    import app.main as m

    # El on_event deprecado se fue: el bucle arranca por lifespan, con
    # referencia FUERTE al task (pitfall de create_task documentado).
    assert not m.app.router.on_startup
    assert m.app.router.lifespan_context is not None


def test_casilla_muestra_el_latido_vivo_del_bucle():
    casilla = dict(CASILLA_DE_PRUEBA, ultima_revision_automatica_el=datetime(2026, 8, 25, 12, 30, tzinfo=ARGENTINA_TEST))
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[casilla]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.obtener_ultimo_tick_revision",
              return_value=datetime(2026, 8, 25, 12, 30, 30, tzinfo=ARGENTINA_TEST)),
        patch("app.main.datetime") as mock_dt,
    ):
        mock_dt.now.return_value = datetime(2026, 8, 25, 12, 31, tzinfo=ARGENTINA_TEST)
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    assert "Bucle de revisión automática vivo: último tick" in respuesta.text
    # Y la última AUTOMÁTICA por casilla, separada de la manual.
    assert "Última revisión AUTOMÁTICA OK" in respuesta.text
    assert "el botón manual no la mueve" in respuesta.text


def test_casilla_grita_en_rojo_si_el_bucle_esta_muerto():
    # "Sin novedades" y "bucle muerto" no pueden verse iguales: con el
    # latido viejo (más de 3 ticks + el tope de un tick colgado), la
    # pantalla lo dice sin vueltas.
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.obtener_ultimo_tick_revision",
              return_value=datetime(2026, 8, 25, 9, 0, tzinfo=ARGENTINA_TEST)),
        patch("app.main.datetime") as mock_dt,
    ):
        mock_dt.now.return_value = datetime(2026, 8, 25, 12, 31, tzinfo=ARGENTINA_TEST)
        respuesta = cliente.get("/sistema/casilla-pedidos")

    assert respuesta.status_code == 200
    assert "El bucle de revisión automática NO está corriendo" in respuesta.text

    # Sin ningún tick registrado jamás, también en rojo.
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[dict(CASILLA_DE_PRUEBA)]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.obtener_ultimo_tick_revision", return_value=None),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")
    assert "nunca registró un tick" in respuesta.text


def test_el_bucle_abandona_un_tick_colgado_y_sigue():
    import asyncio
    import app.main as m

    # Un tick que se cuelga (el bug del 25/08: IMAP en blackhole) tiene
    # que morir por el timeout envolvente SIN matar el bucle.
    async def probar():
        colgado = asyncio.Event()

        def revision_colgada():
            import time as t
            t.sleep(10)

        with (
            patch.object(m, "SEGUNDOS_TIMEOUT_TICK", 0.2),
            patch.object(m, "SEGUNDOS_TICK_REVISION", 0.05),
            patch.object(m, "revisar_casillas_activas", revision_colgada),
            patch.object(m, "registrar_tick_revision", lambda: None),
        ):
            tarea = asyncio.create_task(m._bucle_revision_casillas())
            await asyncio.sleep(1)
            # Tras varios ciclos con la revisión colgada, el bucle sigue vivo.
            assert not tarea.done()
            tarea.cancel()

    asyncio.run(probar())


# --- Costos Fijos (Gerencia): el gasto de operar, derivado siempre ---

GRUPOS_CF_DE_PRUEBA = [
    {"id": 1, "numero": 10, "nombre": "Sueldos", "creado_en": datetime(2026, 8, 25, 12, 0), "baja_el": None},
    {"id": 2, "numero": 40, "nombre": "Ocupación", "creado_en": datetime(2026, 8, 25, 12, 0), "baja_el": None},
]
SUBCUENTAS_CF_DE_PRUEBA = [
    {"id": 1, "grupo_id": 1, "numero": 1, "nombre": "Sdo Christian", "creado_en": datetime(2026, 8, 25, 12, 0),
     "baja_desde": None, "grupo_numero": 10, "grupo_nombre": "Sueldos"},
    {"id": 2, "grupo_id": 2, "numero": 2, "nombre": "Luz", "creado_en": datetime(2026, 8, 25, 12, 0),
     "baja_desde": None, "grupo_numero": 40, "grupo_nombre": "Ocupación"},
]
IMPORTES_CF_DE_PRUEBA = [
    {"id": 1, "subcuenta_id": 1, "mes_desde": date(2026, 7, 1), "importe": 1000.0,
     "alcance": "en_adelante", "creado_en": datetime(2026, 7, 1, 12, 0)},
    {"id": 2, "subcuenta_id": 2, "mes_desde": date(2026, 8, 1), "importe": 200.0,
     "alcance": "en_adelante", "creado_en": datetime(2026, 8, 1, 12, 0)},
]
INDICES_CF_DE_PRUEBA = [
    {"mes": date(2026, 8, 1), "porcentaje": 10.0, "actualizado_en": datetime(2026, 8, 1, 12, 0)},
]


def _patches_costos_fijos(indices=None):
    return (
        patch("app.main.listar_grupos_costos_fijos", return_value=[dict(g) for g in GRUPOS_CF_DE_PRUEBA]),
        patch("app.main.listar_subcuentas_costos_fijos", return_value=[dict(s) for s in SUBCUENTAS_CF_DE_PRUEBA]),
        patch("app.main.listar_importes_costos_fijos", return_value=[dict(i) for i in IMPORTES_CF_DE_PRUEBA]),
        patch("app.main.listar_indices_inflacion",
              return_value=[dict(i) for i in (INDICES_CF_DE_PRUEBA if indices is None else indices)]),
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
    )


def test_costos_fijos_muestra_el_mes_derivado_con_desglose_por_grupo():
    with ExitStack() as stack:
        for parche in _patches_costos_fijos():
            stack.enter_context(parche)
        respuesta = cliente.get("/gerencia/costos-fijos")

    assert respuesta.status_code == 200
    texto = respuesta.text
    # Agosto: la foto de julio (1000) inflada por el índice de agosto
    # (10%) = 1100, más la foto de agosto tal cual (200) = 1300.
    assert "$1.300" in texto
    assert "Costo fijo de 08/2026" in texto
    assert "10.1" in texto and "Sdo Christian" in texto and "$1.100" in texto
    assert "foto de 07/2026" in texto
    assert "40.2" in texto and "$200" in texto
    # Nada de INCOMPLETO ni tarjeta roja: el índice está.
    assert "INCOMPLETO" not in texto
    # Los accesos del módulo y el botón en el hub de Gerencia.
    assert 'href="/gerencia/costos-fijos/plan"' in texto
    assert 'href="/gerencia/costos-fijos/indices"' in texto
    hub = cliente.get("/gerencia")
    assert 'href="/gerencia/costos-fijos"' in hub.text


def test_costos_fijos_sin_indice_grita_y_marca_el_total_incompleto():
    with ExitStack() as stack:
        for parche in _patches_costos_fijos(indices=[]):
            stack.enter_context(parche)
        respuesta = cliente.get("/gerencia/costos-fijos")

    texto = respuesta.text
    # Sin el índice de agosto, la foto de julio no se proyecta: tarjeta
    # roja con el mes que falta, la subcuenta afectada y el link a cargar.
    assert "FALTA EL ÍNDICE DE INFLACIÓN DE 08/2026" in texto
    assert "10.1 Sdo Christian" in texto
    assert "INCOMPLETO" in texto
    assert 'href="/gerencia/costos-fijos/indices"' in texto
    # La foto de agosto (200) sí entra: el total parcial se muestra COMO parcial.
    assert "$200" in texto


def test_costos_fijos_filtra_por_grupo():
    with ExitStack() as stack:
        for parche in _patches_costos_fijos():
            stack.enter_context(parche)
        respuesta = cliente.get("/gerencia/costos-fijos?grupo=40")

    assert "Luz" in respuesta.text
    assert "Sdo Christian" not in respuesta.text


def test_plan_de_cuentas_crea_grupo_con_numero_del_dueno_y_rechaza_repetido():
    with (
        patch("app.main.listar_grupos_costos_fijos", return_value=[dict(g) for g in GRUPOS_CF_DE_PRUEBA]),
        patch("app.main.crear_grupo_costos_fijos", return_value=9) as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/grupos", data={"numero": "90", "nombre": "Informática"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(90, "Informática")

    # El número lo elige el dueño, pero repetido no: error con el dueño actual.
    with (
        patch("app.main.listar_grupos_costos_fijos", return_value=[dict(g) for g in GRUPOS_CF_DE_PRUEBA]),
        patch("app.main.crear_grupo_costos_fijos") as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/grupos", data={"numero": "10", "nombre": "Otro"},
            follow_redirects=False,
        )
    mock_crear.assert_not_called()
    assert "ya+es+de" in respuesta.headers["location"]


def test_plan_de_cuentas_crea_subcuenta_y_rechaza_decimal_repetido():
    with (
        patch("app.main.listar_grupos_costos_fijos", return_value=[dict(g) for g in GRUPOS_CF_DE_PRUEBA]),
        patch("app.main.listar_subcuentas_costos_fijos", return_value=[dict(s) for s in SUBCUENTAS_CF_DE_PRUEBA]),
        patch("app.main.crear_subcuenta_costos_fijos", return_value=9) as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/subcuentas",
            data={"grupo_id": "1", "numero": "2", "nombre": "Sdo Camila"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, 2, "Sdo Camila")

    with (
        patch("app.main.listar_grupos_costos_fijos", return_value=[dict(g) for g in GRUPOS_CF_DE_PRUEBA]),
        patch("app.main.listar_subcuentas_costos_fijos", return_value=[dict(s) for s in SUBCUENTAS_CF_DE_PRUEBA]),
        patch("app.main.crear_subcuenta_costos_fijos") as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/subcuentas",
            data={"grupo_id": "1", "numero": "1", "nombre": "Otro"},
            follow_redirects=False,
        )
    mock_crear.assert_not_called()
    assert "10.1+ya+es+de" in respuesta.headers["location"]


def test_indice_de_inflacion_se_guarda_por_mes_y_acepta_negativo():
    with patch("app.main.guardar_indice_inflacion") as mock_guardar:
        respuesta = cliente.post(
            "/gerencia/costos-fijos/indices", data={"mes": "2026-09", "porcentaje": "-1,5"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(date(2026, 9, 1), -1.5)

    with patch("app.main.guardar_indice_inflacion") as mock_guardar:
        respuesta = cliente.post(
            "/gerencia/costos-fijos/indices", data={"mes": "2026-09", "porcentaje": "tres"},
            follow_redirects=False,
        )
    mock_guardar.assert_not_called()
    assert "porcentaje" in respuesta.headers["location"]


def test_cargar_importe_guarda_la_foto_y_avisa_que_infla_de_ahi_en_adelante():
    with (
        patch("app.main.obtener_subcuenta_costos_fijos", return_value=dict(SUBCUENTAS_CF_DE_PRUEBA[0])),
        patch("app.main.crear_importe_costos_fijos") as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/subcuentas/1/cargar",
            data={"importe": "1500,50", "mes": "2026-08"},
            follow_redirects=False,
        )
    assert respuesta.status_code == 303
    mock_crear.assert_called_once_with(1, date(2026, 8, 1), 1500.5)
    destino = respuesta.headers["location"]
    assert destino.startswith("/gerencia/costos-fijos?")
    assert "de+ah%C3%AD+en+adelante+infla" in destino


def test_cargar_importe_invalido_no_guarda():
    with (
        patch("app.main.obtener_subcuenta_costos_fijos", return_value=dict(SUBCUENTAS_CF_DE_PRUEBA[0])),
        patch("app.main.crear_importe_costos_fijos") as mock_crear,
    ):
        respuesta = cliente.post(
            "/gerencia/costos-fijos/subcuentas/1/cargar",
            data={"importe": "mil", "mes": "2026-08"},
            follow_redirects=False,
        )
    mock_crear.assert_not_called()
    assert "error" in respuesta.headers["location"]


# --- El cruce de primera de reproceso: armada para uno, salió a otro ---


def _entrada_reproceso_cruce(reproceso_id, cliente_id, cliente_nombre, bultos):
    # "orden" incluido: desde E4 lo arma la base, no la pantalla, y el fixture
    # tiene que traer lo mismo que la consulta de verdad.
    return {
        "fecha_orden": date(2026, 8, 24), "momento_orden": datetime(2026, 8, 24, 10),
        "orden": (date(2026, 8, 24), datetime(2026, 8, 24, 10)),
        "tipo_lote": "reproceso", "origen_id": reproceso_id, "fecha_lote": date(2026, 8, 24),
        "detalle": cliente_nombre, "motivo": None, "cantidad": bultos, "costo_bulto": 1000.0,
        "cliente_lote_id": cliente_id,
    }


def _salida_armado_cruce(cliente_id, bultos):
    return {
        "fecha_orden": date(2026, 8, 25), "momento_orden": datetime(2026, 8, 25, 13),
        "orden": (date(2026, 8, 25), datetime(2026, 8, 25, 13)),
        "tipo": "armado", "fecha": date(2026, 8, 25), "cantidad": bultos,
        "unidades": None, "cliente_id": cliente_id, "motivo": None, "bultos_segunda": None,
    }


def test_cruce_de_primera_detecta_bultos_que_salieron_a_otro_cliente():
    # La guía R12 se armó para Día (10 cajas) y un pedido de Vea se llevó
    # 4 por FIFO: el cruce se delata con guía, clientes, bultos y fecha.
    with (
        patch("app.main.listar_articulos_con_primera_de_cliente",
              return_value=[{"articulo_id": 1, "articulo_nombre": "Tomate Perita"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.entradas_y_salidas_stock_articulos",
              return_value={1: ([_entrada_reproceso_cruce(12, 1, "Día", 10.0)], _salidas_fifo(4.0))}),
        patch("app.main.salidas_stock_articulos", return_value={1: [_salida_armado_cruce(2, 4.0)]}),
    ):
        cruces = __import__("app.main", fromlist=["x"])._cruces_primera_reproceso()

    assert len(cruces) == 1
    cruce = cruces[0]
    assert cruce["reproceso_id"] == 12
    assert cruce["cliente_lote_nombre"] == "Día"
    assert cruce["cliente_salida_nombre"] == "Vea"
    assert cruce["bultos"] == 4.0
    assert cruce["fecha"] == date(2026, 8, 25)


def test_cruce_de_primera_no_avisa_si_sale_al_mismo_cliente():
    with (
        patch("app.main.listar_articulos_con_primera_de_cliente",
              return_value=[{"articulo_id": 1, "articulo_nombre": "Tomate Perita"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.entradas_y_salidas_stock_articulos",
              return_value={1: ([_entrada_reproceso_cruce(12, 1, "Día", 10.0)], _salidas_fifo(4.0))}),
        patch("app.main.salidas_stock_articulos", return_value={1: [_salida_armado_cruce(1, 4.0)]}),
    ):
        cruces = __import__("app.main", fromlist=["x"])._cruces_primera_reproceso()
    assert cruces == []


def test_guias_r_muestra_la_ficha_y_deja_completar_la_que_no_tiene():
    guias = [
        dict(GUIAS_R_DE_PRUEBA[0], id=1, articulo_id=5, ficha_id=901,
             ficha_nombre="Banana Bolivia", anulado_el=None),
        dict(GUIAS_R_DE_PRUEBA[0], id=2, articulo_id=5, ficha_id=None,
             ficha_nombre=None, anulado_el=None),
    ]
    # Las mismas claves que devuelve listar_fichas_de_todos_los_clientes:
    # cliente_id, no cliente_nombre. La fixture vieja inventaba una columna
    # que esa consulta nunca devolvió, y por eso pasaba en verde mientras
    # la pantalla real tiraba 500.
    fichas = [
        {"id": 901, "articulo_id": 5, "cliente_id": 1, "nombre_cliente": "Banana Bolivia",
         "articulo_nombre": "Banana"},
        {"id": 902, "articulo_id": 5, "cliente_id": 1, "nombre_cliente": "Banana Ecuador",
         "articulo_nombre": "Banana"},
    ]
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=guias),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=fichas),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    assert respuesta.status_code == 200
    cuerpo = respuesta.text.split("</style>")[-1]
    # La asignada muestra su ficha; la otra dice que le falta.
    assert "Banana Bolivia" in cuerpo
    assert "Sin asignar a una ficha" in cuerpo
    # Y la que falta se completa desde acá, con LAS DOS fichas del artículo.
    assert 'action="/administracion/stock/guias-r/2/asignar-ficha"' in cuerpo
    assert "Banana Ecuador" in cuerpo


def test_guias_r_marca_las_guias_donde_el_reparto_lo_eligio_el_OPERARIO():
    """La otra mitad del desglose editable: que se pueda saber después.

    Un costo congelado contra los lotes que eligió una persona no es lo
    mismo que uno que salió del FIFO, y la única forma de distinguirlos
    dentro de tres meses es que la guía lo diga.
    """
    guias = [
        dict(GUIAS_R_DE_PRUEBA[0], id=1, ficha_id=None, ficha_nombre=None,
             anulado_el=None, consumos_editados=True),
        dict(GUIAS_R_DE_PRUEBA[0], id=2, ficha_id=None, ficha_nombre=None,
             anulado_el=None, consumos_editados=False),
    ]
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=guias),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert cuerpo.count("lo eligió el operario, no el FIFO") == 1


def test_SIN_ASIGNAR_va_en_su_propio_grupo_no_al_lado_de_las_cajas():
    """Al mismo nivel que las fichas, elegirlo cuesta lo mismo que elegir
    bien, y se vuelve el default de hecho: un toque menos que pensar.
    Separado se lee como la excepción que es. Y sigue SIEMPRE presente:
    el que no sabe tiene que poder decirlo.
    """
    guias = [dict(GUIAS_R_DE_PRUEBA[0], id=2, articulo_id=5, ficha_id=None,
                  ficha_nombre=None, anulado_el=None)]
    # Las MISMAS claves que devuelve listar_fichas_de_todos_los_clientes:
    # cliente_id, no cliente_nombre. La fixture vieja traía un
    # cliente_nombre que esa consulta nunca devolvió, y por eso los tests
    # pasaban mientras la pantalla real tiraba 500 en producción.
    fichas = [{"id": 901, "articulo_id": 5, "cliente_id": 1,
               "nombre_cliente": "Banana Bolivia", "articulo_nombre": "Banana"}]
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=guias),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=fichas),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert '<optgroup label="Cajas de este artículo">' in cuerpo
    assert '<optgroup label="Si no sabés">' in cuerpo
    # La caja y la excepción NO comparten grupo.
    assert cuerpo.index("Banana Bolivia") < cuerpo.index('label="Si no sabés"')
    assert "Sin asignar" in cuerpo
    # La opción dice de qué cliente es la ficha: dos fichas con el mismo
    # nombre de dos clientes distintos serían indistinguibles.
    assert "Banana Bolivia (Día)" in cuerpo


def test_una_guia_anulada_no_ofrece_asignar_ficha():
    guias = [dict(GUIAS_R_DE_PRUEBA[0], id=3, articulo_id=5, ficha_id=None,
                  ficha_nombre=None, anulado_el=datetime(2026, 8, 26, 10, 0))]
    with (
        patch("app.main.listar_reprocesos_por_rango", return_value=guias),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main._cruces_primera_reproceso", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    assert respuesta.status_code == 200
    assert "/guias-r/3/asignar-ficha" not in respuesta.text


def test_asignar_una_ficha_de_otro_articulo_se_muestra_en_pantalla_nunca_500():
    with patch("app.main.asignar_ficha_a_reproceso",
               side_effect=ValueError("Esa ficha es de otro artículo: no puede ser la de esta guía R.")):
        respuesta = cliente.post(
            "/administracion/stock/guias-r/2/asignar-ficha",
            data={"ficha_id": "901", "fecha_desde": "2026-08-20", "fecha_hasta": "2026-08-28"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "otro+art" in respuesta.headers["location"]
    # Y vuelve al mismo rango que estaba mirando.
    assert "fecha_desde=2026-08-20" in respuesta.headers["location"]


def test_desasignar_desde_guias_r_manda_None():
    with patch("app.main.asignar_ficha_a_reproceso") as mock_asignar:
        respuesta = cliente.post(
            "/administracion/stock/guias-r/2/asignar-ficha",
            data={"ficha_id": "sin_asignar", "fecha_desde": "", "fecha_hasta": ""},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_asignar.assert_called_once_with(2, None)
    assert "sin+asignar" in respuesta.headers["location"]


def test_borrar_una_ficha_con_guias_R_lo_dice_en_la_pantalla_y_NO_da_500():
    with patch("app.main.eliminar_ficha",
               side_effect=ValueError("Esa ficha tiene 3 guías R cargadas: no se puede borrar. "
                                      "Reasignalas a otra ficha desde Guías R si hace falta.")):
        respuesta = cliente.post("/fichas/901/eliminar", data={"cliente_id": "1"},
                                 follow_redirects=False)

    assert respuesta.status_code == 303
    destino = respuesta.headers["location"]
    assert "/fichas?" in destino
    assert "3+gu%C3%ADas+R" in destino
    assert "cliente_id=1" in destino


def test_guias_r_muestran_para_quien_y_el_cruce_con_datos():
    guia = {
        "id": 12, "articulo_id": 1, "fecha_operacion": date(2026, 8, 24),
        "bultos_tomados": 12.0, "bultos_primera": 10.0, "bultos_segunda": 1.0,
        "bultos_merma": 1.0, "costo_total": 12000.0, "costo_por_bulto_primera": 1200.0,
        "creado_en": datetime(2026, 8, 24, 10, 0), "anulado_el": None,
        "articulo_nombre": "Tomate Perita", "cliente_id": 1, "cliente_nombre": "Día",
        "consumos": [],
    }
    guia_vieja = dict(guia, id=9, cliente_id=None, cliente_nombre=None)
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_reprocesos_por_rango", return_value=[guia, guia_vieja]),
        patch("app.main.contar_reprocesos_sin_costo_posible", return_value={"casos": 0, "mas_viejo": None}),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_articulos_con_primera_de_cliente",
              return_value=[{"articulo_id": 1, "articulo_nombre": "Tomate Perita"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.entradas_y_salidas_stock_articulos",
              return_value={1: ([_entrada_reproceso_cruce(12, 1, "Día", 10.0)], _salidas_fifo(4.0))}),
        patch("app.main.salidas_stock_articulos", return_value={1: [_salida_armado_cruce(2, 4.0)]}),
    ):
        respuesta = cliente.get("/administracion/stock/guias-r")

    assert respuesta.status_code == 200
    texto = respuesta.text
    assert "Primera armada para <strong>Día</strong>" in texto
    # La guía vieja queda marcada, no acusada.
    assert "Sin cliente (guía vieja, anterior al dato)" in texto
    # El cruce, con los datos: aviso — jamás traba.
    assert "CRUCE DE CLIENTE" in texto
    assert "4 bultos salieron en pedidos de Vea" in texto


def test_auditoria_incluye_la_alerta_de_cruce_de_primera():
    foto = _foto_alertas({"cruce_primera_reproceso": (1, date(2026, 8, 25))})
    with patch("app.main.listar_estado_alertas", return_value=foto):
        respuesta = cliente.get("/auditoria")

    assert respuesta.status_code == 200
    assert "Primera de reproceso armada para un cliente salió en pedidos de otro" in respuesta.text
    assert 'href="/administracion/stock/guias-r"' in respuesta.text


def test_el_cruce_de_primera_se_cuenta_con_su_fecha_mas_vieja():
    """El conteo del cruce sigue siendo el de siempre: cantidad y fecha del más viejo."""
    from app.main import _contar_cruces_primera_reproceso

    cruces = [
        {"reproceso_id": 12, "cliente_salida_nombre": "Vea", "bultos": 4.0, "fecha": date(2026, 8, 25)},
        {"reproceso_id": 13, "cliente_salida_nombre": "Día", "bultos": 2.0, "fecha": date(2026, 8, 20)},
    ]
    with patch("app.main._cruces_primera_reproceso", return_value=cruces):
        assert _contar_cruces_primera_reproceso() == {"casos": 2, "mas_viejo": date(2026, 8, 20)}


def test_el_atras_jerarquico_esta_declarado_en_todo_el_sistema():
    # Muestra representativa de cada zona: el botón de atrás es un LINK al
    # padre del sistema (nunca history.back(), que queda para el gesto
    # nativo del navegador).
    ancla = '<a class="barra-boton" href="{destino}" aria-label="Volver atrás">'

    # Hubs → Inicio; hijos → su hub; nietos → su padre intermedio.
    directas = [
        ("/puesto", "/inicio"),
        ("/logistica", "/inicio"),
        ("/puesto/envases", "/puesto"),
        ("/deposito/stock", "/deposito"),
    ]
    for url, padre in directas:
        respuesta = cliente.get(url)
        assert ancla.format(destino=padre) in respuesta.text, url
        assert 'onclick="history.back()"' not in respuesta.text, url

    # Vacíos: el hub del empleado cuelga de Envases Puesto; Recibir, de Vacíos.
    respuesta = cliente.get("/puesto/envases/vacios")
    assert ancla.format(destino="/puesto/envases") in respuesta.text
    with (
        patch("app.main.listar_tipos_envase_puesto", return_value=[]),
        patch("app.main.listar_clientes_puesto", return_value=[]),
        patch("app.main.listar_vacios_recibidos_de_fecha", return_value=[]),
    ):
        respuesta = cliente.get("/puesto/envases/vacios/recibir")
    assert ancla.format(destino="/puesto/envases/vacios") in respuesta.text

    # Retiro: el padre es el hub DESDE el que se entró (origen en la URL).
    with patch("app.main.listar_compras_pendientes_retiro", return_value=[]):
        desde_logistica = cliente.get("/logistica/retiro/Clark")
        desde_deposito = cliente.get("/logistica/retiro/Pases?origen=deposito")
    assert ancla.format(destino="/logistica") in desde_logistica.text
    assert ancla.format(destino="/deposito") in desde_deposito.text

    # Recepción → Depósito.
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 25)),
        patch("app.main.listar_compras_pendientes_recepcion", return_value=[]),
        patch("app.main.listar_compras_procesadas_hoy_recepcion", return_value=[]),
    ):
        respuesta = cliente.get("/deposito/recepcion")
    assert ancla.format(destino="/deposito") in respuesta.text

    # El detalle FIFO cuelga del Remanente, que es de donde se llega.
    with (
        patch("app.main.obtener_articulo", return_value={"id": 2, "nombre": "Anco"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(0.0))),
        patch("app.main.stock_deposito_por_articulo", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/sistema/2")
    assert ancla.format(destino="/administracion/stock/remanente") in respuesta.text

    # La casilla cuelga de Sistema.
    with (
        patch("app.main.listar_casillas_pedidos", return_value=[]),
        patch("app.main.listar_mails_pedido", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.clave_casilla_configurada", return_value="clave"),
        patch("app.main.obtener_ultimo_tick_revision", return_value=None),
    ):
        respuesta = cliente.get("/sistema/casilla-pedidos")
    assert ancla.format(destino="/sistema") in respuesta.text


# --- Etapa 2: el stock inicial del corte ---
# La pantalla que se usa UNA vez, el día antes del corte, para que lo que
# hay en el piso pase a existir para el sistema. Va en Administración y no
# en Depósito porque lleva costo.

FICHAS_STOCK_INICIAL = [
    {"id": 10, "cliente_id": 1, "articulo_id": 7, "articulo_nombre": "Banana",
     "contenido_caja": 18, "unidad_venta": "kilo", "envase_variable": False,
     "nombre_cliente": "Banana Bolivia", "codigo_cliente": None},
    {"id": 11, "cliente_id": 2, "articulo_id": 7, "articulo_nombre": "Banana",
     "contenido_caja": 18, "unidad_venta": "kilo", "envase_variable": False,
     "nombre_cliente": "Banana Ecuador", "codigo_cliente": None},
]

CARGADO_VACIO = {"sueltos": [], "armadas": [], "total_bultos": 0, "total_pesos": 0}


def _pantalla_stock_inicial(url="/administracion/stock/inicial", cargado=None, fichas=None):
    with (
        patch("app.main.listar_articulos", return_value=[{"id": 7, "nombre": "Banana"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes",
              return_value=FICHAS_STOCK_INICIAL if fichas is None else fichas),
        patch("app.main.listar_stock_inicial", return_value=cargado or CARGADO_VACIO),
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        return cliente.get(url)


def test_stock_inicial_muestra_la_fecha_de_corte_leida_de_la_base():
    respuesta = _pantalla_stock_inicial()

    assert respuesta.status_code == 200
    # La fecha NO está escrita en el código: sale de corte_modelo.
    assert "31/08/2026" in respuesta.text


def test_stock_inicial_sin_articulo_no_muestra_los_formularios_de_carga():
    respuesta = _pantalla_stock_inicial()

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "Elegí el artículo" in cuerpo
    assert "Guardar bultos sueltos" not in cuerpo
    assert "Guardar cajas armadas" not in cuerpo


def test_stock_inicial_con_articulo_muestra_las_dos_formas_de_cargar():
    respuesta = _pantalla_stock_inicial("/administracion/stock/inicial?articulo_id=7")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "Guardar bultos sueltos" in cuerpo
    assert "Guardar cajas armadas" in cuerpo
    # El artículo queda elegido en el selector: no hay que reelegirlo.
    assert 'value="7" selected' in " ".join(cuerpo.split())


def test_stock_inicial_las_fichas_se_eligen_por_articulo_y_dicen_de_que_cliente_son():
    respuesta = _pantalla_stock_inicial("/administracion/stock/inicial?articulo_id=7")

    cuerpo = respuesta.text.split("</style>")[-1]
    # Sin el cliente adentro del nombre, las dos fichas de Banana de dos
    # clientes distintos serían dos opciones idénticas.
    assert "Día — Banana Bolivia" in cuerpo
    assert "Vea — Banana Ecuador" in cuerpo
    # Y NO hay un "sin asignar" como en la guía R: una caja que está en el
    # piso se puede ir a mirar.
    assert "Sin asignar" not in cuerpo


def test_stock_inicial_sin_fichas_del_articulo_lo_dice_en_vez_de_dejar_cargar():
    respuesta = _pantalla_stock_inicial("/administracion/stock/inicial?articulo_id=7", fichas=[])

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "no tiene ninguna ficha cargada" in cuerpo
    assert "Guardar cajas armadas" not in cuerpo


def test_stock_inicial_sueltos_guarda_con_la_fecha_de_corte_y_vuelve_al_mismo_articulo():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 7, "nombre": "Banana"}),
        patch("app.main.crear_stock_inicial", return_value=40.0) as mock_crear,
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/inicial/sueltos",
            data={"articulo_id": "7", "bultos": "40", "costo_por_bulto": "1500"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # La fecha del movimiento es la del corte, no la de hoy: se carga el
    # domingo lo que hay para el lunes.
    mock_crear.assert_called_once_with(7, 40.0, 1500.0, date(2026, 8, 31))
    # Vuelve con el MISMO artículo puesto y al formulario, no arriba de
    # todo: son muchos renglones seguidos.
    assert "articulo_id=7" in respuesta.headers["location"]
    assert respuesta.headers["location"].endswith("#carga")
    # El importe del aviso se escribe igual que el del listado: $1.500, no
    # $1500. Verlos distintos en la misma pantalla hace dudar del número.
    assert "%241.500" in respuesta.headers["location"]


def test_stock_inicial_sueltos_sin_costo_da_400_y_no_guarda():
    with (
        patch("app.main.crear_stock_inicial") as mock_crear,
        patch("app.main.listar_articulos", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=[]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_stock_inicial", return_value=CARGADO_VACIO),
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/inicial/sueltos",
            data={"articulo_id": "7", "bultos": "40", "costo_por_bulto": ""},
        )

    assert respuesta.status_code == 400
    # Es lo único que este stock no puede recuperar después.
    assert "sin costear" in respuesta.text
    mock_crear.assert_not_called()


def test_stock_inicial_armadas_guarda_un_reproceso_inicial_con_su_ficha():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 7, "nombre": "Banana"}),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.crear_reproceso_inicial", return_value=99) as mock_crear,
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/inicial/armadas",
            data={"articulo_id": "7", "ficha_id": "11", "cajas": "20", "costo_por_caja": "2200"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    # El cliente sale de la ficha elegida, no se pregunta aparte.
    mock_crear.assert_called_once_with(
        7, 20.0, 2200.0, date(2026, 8, 31), ficha_id=11, cliente_id=2
    )
    assert "articulo_id=7" in respuesta.headers["location"]


def test_stock_inicial_armadas_sin_ficha_da_400_y_no_guarda():
    with (
        patch("app.main.obtener_articulo", return_value={"id": 7, "nombre": "Banana"}),
        patch("app.main.listar_articulos", return_value=[{"id": 7, "nombre": "Banana"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_stock_inicial", return_value=CARGADO_VACIO),
        patch("app.main.crear_reproceso_inicial") as mock_crear,
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/inicial/armadas",
            data={"articulo_id": "7", "ficha_id": "", "cajas": "20", "costo_por_caja": "2200"},
        )

    assert respuesta.status_code == 400
    assert "siempre es de alguna" in respuesta.text
    mock_crear.assert_not_called()


def test_stock_inicial_armadas_con_ficha_de_otro_articulo_no_guarda():
    # La ficha 10 es de Banana (artículo 7): pedirla para el artículo 9
    # tiene que rebotar, aunque el id exista.
    with (
        patch("app.main.obtener_articulo", return_value={"id": 9, "nombre": "Pera"}),
        patch("app.main.listar_articulos", return_value=[{"id": 9, "nombre": "Pera"}]),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_fichas_de_todos_los_clientes", return_value=FICHAS_STOCK_INICIAL),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_stock_inicial", return_value=CARGADO_VACIO),
        patch("app.main.crear_reproceso_inicial") as mock_crear,
        patch("app.main.fecha_corte", return_value=date(2026, 8, 31)),
    ):
        respuesta = cliente.post(
            "/administracion/stock/inicial/armadas",
            data={"articulo_id": "9", "ficha_id": "10", "cajas": "20", "costo_por_caja": "2200"},
        )

    assert respuesta.status_code == 400
    mock_crear.assert_not_called()


def test_stock_inicial_muestra_lo_cargado_con_el_total_de_las_dos_formas():
    cargado = {
        "sueltos": [{"id": 1, "articulo_id": 7, "articulo_nombre": "Banana",
                     "cantidad": 40, "costo_por_bulto": 1500,
                     "fecha_operacion": date(2026, 8, 31), "creado_en": None}],
        "armadas": [{"id": 99, "articulo_id": 7, "articulo_nombre": "Banana",
                     "bultos_primera": 20, "costo_por_bulto_primera": 2200,
                     "ficha_id": 11, "ficha_nombre": "Banana Ecuador",
                     "cliente_id": 2, "cliente_nombre": "Vea",
                     "fecha_operacion": date(2026, 8, 31), "creado_en": None}],
        "total_bultos": 60,
        "total_pesos": 104000.0,
    }
    respuesta = _pantalla_stock_inicial(cargado=cargado)

    cuerpo = respuesta.text.split("</style>")[-1]
    # Para saber por dónde va y no cargar dos veces lo mismo.
    assert "40 bultos × $1.500" in cuerpo
    assert "20 cajas × $2.200" in cuerpo
    assert "Banana Ecuador" in cuerpo
    # El total, de las dos formas juntas.
    assert "60</span>" in cuerpo
    assert "$104.000" in cuerpo


def test_stock_inicial_anular_pasa_la_clase_y_vuelve_al_mismo_articulo():
    with patch("app.main.anular_renglon_stock_inicial") as mock_anular:
        respuesta = cliente.post(
            "/administracion/stock/inicial/anular",
            data={"clase": "armadas", "renglon_id": "99", "articulo_id": "7"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_anular.assert_called_once_with("armadas", 99)
    assert "articulo_id=7" in respuesta.headers["location"]


def test_stock_inicial_anular_con_clase_desconocida_no_llama_a_la_base():
    with patch("app.main.anular_renglon_stock_inicial") as mock_anular:
        respuesta = cliente.post(
            "/administracion/stock/inicial/anular",
            data={"clase": "cualquiera", "renglon_id": "99", "articulo_id": "7"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_anular.assert_not_called()


def test_administracion_tiene_el_boton_del_stock_inicial():
    with patch("app.main._banner_alertas", return_value=None):
        respuesta = cliente.get("/administracion")

    assert 'href="/administracion/stock/inicial"' in respuesta.text


def test_stock_inicial_los_importes_grandes_no_salen_en_notacion_cientifica():
    """120 bultos a $8.500 son más de un millón, y ahí es donde se rompía.

    La primera versión formateaba a mano con %g y mostraba "$1.02e+06" en
    la columna de plata: el número que hay que mirar para saber si la
    carga va bien, ilegible justo cuando la carga ya avanzó.
    """
    cargado = {
        "sueltos": [{"id": 1, "articulo_id": 7, "articulo_nombre": "Tomate Redondo",
                     "cantidad": 120, "costo_por_bulto": 8500,
                     "fecha_operacion": date(2026, 8, 31), "creado_en": None}],
        "armadas": [],
        "total_bultos": 120,
        "total_pesos": 1020000.0,
    }
    respuesta = _pantalla_stock_inicial(cargado=cargado)

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "e+" not in cuerpo
    assert "$1.020.000" in cuerpo
    # Y el detalle usa el mismo formato que el total: $8.500, no $8500.
    assert "120 bultos × $8.500" in cuerpo


def test_cotejo_dice_de_que_porcion_es_cada_tarjeta():
    # Un artículo puede tener varias tarjetas. Sin decir cuál es cuál,
    # "Banana 12 / Banana 40" no se entiende.
    conteos = [
        {"id": 5, "articulo_id": 1, "cantidad": 40.0, "stock_sistema": 40.0,
         "creado_en": datetime(2026, 8, 25, 10, 30), "articulo_nombre": "Banana",
         "ficha_id": None, "ficha_nombre": None, "ficha_cliente": None},
        {"id": 6, "articulo_id": 1, "cantidad": 12.0, "stock_sistema": 12.0,
         "creado_en": datetime(2026, 8, 25, 11, 0), "articulo_nombre": "Banana",
         "ficha_id": 11, "ficha_nombre": "Banana Bolivia", "ficha_cliente": "Día"},
    ]
    with patch("app.main.listar_ultimos_conteos_stock", return_value=conteos):
        respuesta = cliente.get("/administracion/stock/cotejo")

    cuerpo = respuesta.text.split("</style>")[-1]
    assert "Bultos sueltos" in cuerpo
    assert "Cajas de Banana Bolivia" in cuerpo
    assert "Día" in cuerpo


def test_cotejo_no_ofrece_ajustar_stock_en_una_diferencia_de_FICHA():
    """El ajuste es por ARTÍCULO: mueve el total, no reparte entre fichas.

    Si sobran cajas de Bolivia y faltan de Ecuador, el total del artículo
    está bien y ajustarlo lo rompería. Lo que hay que corregir es a qué
    ficha fue una guía R.
    """
    conteos = [
        {"id": 6, "articulo_id": 1, "cantidad": 12.0, "stock_sistema": 20.0,
         "creado_en": datetime(2026, 8, 25, 11, 0), "articulo_nombre": "Banana",
         "ficha_id": 11, "ficha_nombre": "Banana Bolivia", "ficha_cliente": "Día"},
    ]
    with patch("app.main.listar_ultimos_conteos_stock", return_value=conteos):
        respuesta = cliente.get("/administracion/stock/cotejo")

    cuerpo = respuesta.text.split("</style>")[-1]
    # La diferencia se muestra igual: -8.
    assert "-8" in cuerpo
    assert "Ajustar a lo contado" not in cuerpo
    # Y manda donde sí se arregla.
    assert "/administracion/stock/guias-r" in cuerpo


def test_cotejo_si_ofrece_ajustar_stock_en_una_diferencia_de_SUELTOS():
    conteos = [
        {"id": 5, "articulo_id": 1, "cantidad": 12.0, "stock_sistema": 15.0,
         "creado_en": datetime(2026, 8, 25, 10, 30), "articulo_nombre": "Banana",
         "ficha_id": None, "ficha_nombre": None, "ficha_cliente": None},
    ]
    with patch("app.main.listar_ultimos_conteos_stock", return_value=conteos):
        respuesta = cliente.get("/administracion/stock/cotejo")

    assert "Ajustar a lo contado" in respuesta.text
    assert "articulo_id=1&amp;contado=12.0" in respuesta.text


def test_cotejo_no_inventa_renglones_de_fichas_que_nunca_se_contaron():
    """Solo sale de conteos_stock, de ningún otro lado.

    Si listara todas las fichas de todos los clientes en cero, la pantalla
    se vuelve ilegible y se deja de mirar.
    """
    with patch("app.main.listar_ultimos_conteos_stock", return_value=[]) as mock_listar:
        respuesta = cliente.get("/administracion/stock/cotejo")

    assert respuesta.status_code == 200
    assert "Todavía no hay conteos físicos cargados" in respuesta.text
    mock_listar.assert_called_once_with()


def test_las_pantallas_que_se_mudaron_vuelven_a_ADMINISTRACION():
    # Resto de la etapa 0: se cambiaron los links de ida pero el "atrás"
    # seguía apuntando al hub de Depósito, donde esos botones ya no están.
    with (
        patch("app.main.listar_ultimos_conteos_stock", return_value=[]),
        patch("app.main.listar_articulos", return_value=[]),
    ):
        cotejo = cliente.get("/administracion/stock/cotejo")
        ajustar = cliente.get("/administracion/stock/ajustar")

    for respuesta in (cotejo, ajustar):
        assert 'href="/deposito/stock"' not in respuesta.text
        assert "/administracion" in respuesta.text


# --- Etapa 5: el pedido descuenta del formato de la ficha, con tolerancia ---
# La regla del envase quedó afuera a propósito: el sistema no captura en
# qué envase vino un bulto (ver el pendiente con nombre propio en
# docs/diseno_base_datos.md).

FICHAS_E5 = [
    {"id": 901, "articulo_id": 1, "articulo_nombre": "Banana", "articulo_grupo": "fruta",
     "envase_id": None, "envase_nombre": None, "contenido_caja": 15, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": "BANANA", "codigo_cliente": "90101"},
    {"id": 902, "articulo_id": 2, "articulo_nombre": "Batata", "articulo_grupo": "hortaliza",
     "envase_id": None, "envase_nombre": None, "contenido_caja": 10, "unidad_venta": "kilo",
     "envase_variable": False, "nombre_cliente": None, "codigo_cliente": "90102"},
]


def _armar_e5(renglones, con_cajas=frozenset({901, 902}), fichas=None):
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 21)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value=PEDIDO_VIGENTE_DE_PRUEBA),
        patch("app.main.listar_sucursales_pedido", return_value=[dict(s) for s in SUCURSALES_PEDIDO_DE_PRUEBA]),
        patch("app.main.listar_renglones_pedido", return_value=renglones),
        patch("app.main.listar_fichas_por_cliente", return_value=fichas if fichas is not None else FICHAS_E5),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.fichas_con_cajas_armadas", return_value=set(con_cajas)),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21&sucursal=VL")
    # Los espacios se normalizan: el aviso está partido en varias líneas en
    # el template, y buscar la frase entera fallaría por el salto.
    return " ".join(respuesta.text.split("</style>")[-1].split())


def _renglon(**cambios):
    base = {"id": 11, "sucursal": "VL", "articulo_id": 1, "articulo_nombre": "Banana",
            "ficha_id": 901, "nombre_venta": "Banana", "texto_codigo": "90101",
            "texto_descripcion": "BANANA", "cantidad": 10.0, "armado_el": None,
            "cantidad_armada": None, "kilos_enviados": None, "anulado_el": None}
    base.update(cambios)
    return base


def test_armar_con_cajas_de_la_ficha_no_avisa_nada():
    # El caso normal no tiene que estorbar.
    cuerpo = _armar_e5([_renglon()])

    assert "Banana" in cuerpo
    assert "No hay cajas armadas de esta ficha" not in cuerpo


def test_armar_sin_cajas_de_la_ficha_avisa_SIN_decir_ningun_numero():
    """El pedido sale del stock de LA FICHA, no del artículo.

    Y el aviso no dice cuántas hay: esta pantalla es de operario y el
    número del sistema no puede viajar acá ni escondido en el HTML. Si lo
    ve, arma contra el sistema en vez de contra el piso.
    """
    cuerpo = _armar_e5([_renglon()], con_cajas=set())

    assert "No hay cajas armadas de esta ficha" in cuerpo
    assert "Fijate si hay que reprocesar" in cuerpo
    # Ni el número ni la palabra: no dice "0 cajas" ni "hay 0".
    assert "0 cajas" not in cuerpo
    # Y NO traba: el botón de tildar sigue estando.
    assert "Listo" in cuerpo or "armar-11" in cuerpo


def test_armar_avisa_por_la_FICHA_y_no_por_el_articulo():
    # Dos renglones del mismo artículo con fichas distintas: solo el que
    # no tiene cajas de SU ficha se marca. Que haya bananas no quiere
    # decir que haya cajas de Banana Bolivia.
    fichas = FICHAS_E5 + [
        {"id": 903, "articulo_id": 1, "articulo_nombre": "Banana", "articulo_grupo": "fruta",
         "envase_id": None, "envase_nombre": None, "contenido_caja": 18, "unidad_venta": "kilo",
         "envase_variable": False, "nombre_cliente": "BANANA ECUADOR", "codigo_cliente": "90109"},
    ]
    renglones = [
        _renglon(id=11, ficha_id=901, nombre_venta="Banana Bolivia"),
        _renglon(id=12, ficha_id=903, nombre_venta="Banana Ecuador"),
    ]
    cuerpo = _armar_e5(renglones, con_cajas={901}, fichas=fichas)

    assert cuerpo.count("No hay cajas armadas de esta ficha") == 1
    # El que avisa es el de Ecuador, que es el que no tiene.
    bolivia, ecuador = cuerpo.split("Banana Ecuador", 1)
    assert "No hay cajas armadas de esta ficha" not in bolivia


def test_armar_un_renglon_sin_ficha_no_se_marca():
    # Renglón viejo o ficha borrada: no hay ficha de la cual mirar el
    # stock, así que no se inventa un aviso.
    cuerpo = _armar_e5([_renglon(ficha_id=None)], con_cajas=set())

    assert "No hay cajas armadas de esta ficha" not in cuerpo


def test_tolerancia_el_kilaje_pasado_avisa_DESPUES_de_tildar():
    # Ficha de 15 kg, cargó 21 por bulto: 6 de más, fuera de los 3.
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad_armada=None,
        cantidad=10.0, kilos_enviados=210.0)])

    assert "Cargaste 21 kg por bulto y la ficha dice 15" in cuerpo
    assert "6 kg de más por bulto" in cuerpo
    # Queda armado igual: avisa, no traba.
    assert "Destildar" in cuerpo


def test_tolerancia_dentro_de_los_3_kg_no_avisa_pero_marca_editado_a_mano():
    # Ficha de 15, cargó 17: 2 de más, dentro de tolerancia.
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0, kilos_enviados=170.0)])

    assert "por bulto y la ficha dice" not in cuerpo
    assert "kilaje editado a mano" in cuerpo


def test_tolerancia_es_por_BULTO_y_no_por_renglon():
    """20 bultos con 1 kg de más son 20 kg en el renglón y están DENTRO.

    Lo que se controla es cómo se llenó cada caja, no cuánto suma el
    renglón: si fuera por renglón, un pedido grande daría fuera de
    tolerancia por acumulación aunque cada caja esté perfecta.
    """
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=20.0, kilos_enviados=320.0)])

    assert "por bulto y la ficha dice" not in cuerpo

    # Y 2 bultos con 5 kg de más son 10 kg en el renglón y están FUERA.
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=2.0, kilos_enviados=40.0)])

    assert "Cargaste 20 kg por bulto y la ficha dice 15" in cuerpo


def test_tolerancia_usa_los_bultos_REALMENTE_armados_si_armo_menos():
    # Pidió 10, armó 4, mandó 84 kg: son 21 por bulto, no 8,4.
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0,
        cantidad_armada=4.0, kilos_enviados=84.0)])

    assert "Cargaste 21 kg por bulto" in cuerpo


def test_el_aviso_de_tolerancia_TAPA_la_marca_de_editado_a_mano():
    """Las dos dicen lo mismo con distinta fuerza: no se apilan.

    Pero el aviso rojo tiene que dejar claro que el kilaje lo cargó una
    persona — si no, se leería como un error de cálculo del sistema, que
    es lo contrario de lo que pasó.
    """
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0, kilos_enviados=210.0)])

    assert "kilaje editado a mano" not in cuerpo
    assert "Cargaste" in cuerpo


def test_tolerancia_no_aplica_a_fichas_que_no_son_por_kilo():
    """±3 no significa nada en una ficha por unidad o por cubeta.

    3 unidades y 3 cubetas son cosas distintas y nadie definió un
    equivalente: inventarlo sería inventar una regla.
    """
    fichas = [dict(FICHAS_E5[0], unidad_venta="unidad", contenido_caja=15)]
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0, kilos_enviados=210.0)],
        fichas=fichas)

    assert "por bulto y la ficha dice" not in cuerpo


def test_tolerancia_sin_contenido_en_la_ficha_no_inventa_una_comparacion():
    fichas = [dict(FICHAS_E5[0], contenido_caja=None)]
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0, kilos_enviados=210.0)],
        fichas=fichas)

    assert "por bulto y la ficha dice" not in cuerpo


def test_tolerancia_avisa_tambien_cuando_mando_de_MENOS():
    # Ficha de 15, cargó 9: 6 de menos. Un cajón a medio llenar también
    # es un problema, y del lado que se factura.
    cuerpo = _armar_e5([_renglon(
        armado_el=datetime(2026, 8, 21, 13, 0), cantidad=10.0, kilos_enviados=90.0)])

    assert "Cargaste 9 kg por bulto y la ficha dice 15" in cuerpo
    assert "6 kg de menos por bulto" in cuerpo


def test_deposito_ordena_los_botones_como_pasan_las_cosas():
    """Retirar → recepcionar → ingresar (si algo falló) → armar → stock.

    El orden es el del día, no el de los módulos: Recepción estaba primera
    de cuando la pantalla se armó por dónde vivía cada ruta, pero lo
    primero que hace el depósito es traer la mercadería del puesto.
    """
    respuesta = cliente.get("/deposito")
    cuerpo = respuesta.text.split("</style>")[-1]

    orden = ["Retirar Mercadería", "Recepción Compras", "Ingresar Mercadería",
             "Armar Pedido", "Stock"]
    posiciones = [cuerpo.index(t) for t in orden]
    assert posiciones == sorted(posiciones), dict(zip(orden, posiciones))


def test_stock_por_guia_nombra_los_bultos_tomados_por_guias_R_en_las_salidas():
    """total_salidas suma CUATRO cosas, no tres.

    El texto se escribió cuando los reprocesos no existían y nombraba solo
    pedidos, mermas y ajustes. Desde que existen, los bultos tomados para
    armar cajas quedaban adentro del número sin figurar en ningún lado:
    quien tomó 300 bultos para una guía R no los encontraba y creía que el
    sistema los había perdido.
    """
    with (
        patch("app.main.obtener_articulo", return_value={"id": 1, "nombre": "Tomate Redondo"}),
        patch("app.main.entradas_y_salidas_stock_articulo", return_value=([], _salidas_fifo(648.0))),
        patch("app.main.stock_deposito_por_articulo", return_value=[]),
    ):
        respuesta = cliente.get("/administracion/stock/sistema/1")

    cuerpo = " ".join(respuesta.text.split("</style>")[-1].split())
    assert "Salieron 648 bultos en total" in cuerpo
    assert "bultos tomados por guías R" in cuerpo


# --- Entrega 3: de qué lote salió un renglón armado ---


def test_el_desglose_de_un_renglon_armado_trae_la_propuesta_y_los_lotes():
    desglose = {
        "articulo_id": 1,
        "armado": 15.0,
        "editado": False,
        "lotes": [
            {"tipo_lote": "guia", "origen_id": 101, "fecha_lote": date(2026, 8, 28),
             "detalle": "Norte 15", "restante": 13.0},
            {"tipo_lote": "guia", "origen_id": 102, "fecha_lote": date(2026, 9, 2),
             "detalle": "Vitale", "restante": 8.0},
        ],
        "propuesta": {"guia:101": 13.0, "guia:102": 2.0},
    }
    with patch("app.main.desglose_de_renglon_armado", return_value=desglose):
        datos = cliente.get("/deposito/pedido/renglones/55/lotes").json()

    assert datos["armado"] == 15.0
    assert datos["editado"] is False
    assert datos["propuesta"] == {"guia:101": 13.0, "guia:102": 2.0}
    # Fecha y cantidad; el proveedor no viaja porque cada día tiene un lote solo.
    assert datos["lotes"][0] == {"clave": "guia:101", "tipo_lote": "guia", "origen_id": 101,
                                 "fecha": "28/08", "restante": 13.0, "detalle": ""}


def test_un_renglon_SIN_TILDAR_no_tiene_desglose_que_mostrar():
    """Antes del tilde esta pantalla no muestra números del sistema: el que
    arma todavía no declaró nada, y si los ve arma contra el sistema en vez
    de contra el piso."""
    with patch("app.main.desglose_de_renglon_armado", return_value=None):
        respuesta = cliente.get("/deposito/pedido/renglones/55/lotes")

    assert respuesta.status_code == 404


def test_guardar_de_donde_salio_manda_solo_lo_que_eligio():
    with patch("app.main.guardar_lotes_elegidos") as mock_guardar:
        respuesta = cliente.post(
            "/deposito/pedido/7/renglones/55/lotes",
            data={"cliente_id": "1", "fecha": "2026-09-02", "sucursal": "VL",
                  "reparto": '[{"tipo_lote": "guia", "origen_id": 101, "bultos": 5}]'},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(
        55, [{"lote_tipo": "guia", "lote_origen_id": 101, "bultos": 5.0}]
    )


def test_aceptar_la_propuesta_BORRA_la_correccion_y_vuelve_al_FIFO():
    """Se guarda SOLO la excepción: un reparto vacío no deja ninguna fila, y
    un renglón sin filas se reparte por FIFO como siempre. El default nunca
    se escribe, así que no puede quedar viejo cuando cambie el stock."""
    with patch("app.main.guardar_lotes_elegidos") as mock_guardar:
        respuesta = cliente.post(
            "/deposito/pedido/7/renglones/55/lotes",
            data={"cliente_id": "1", "fecha": "2026-09-02", "sucursal": "VL", "reparto": "[]"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once_with(55, [])


def test_guardar_de_donde_salio_NO_TRABA_si_no_llega_a_lo_armado():
    """La diferencia con el reproceso, y es toda la entrega: acá el camión ya
    salió. Lo que no se reparta cae al FIFO o a sin_lote, como hasta hoy."""
    with patch("app.main.guardar_lotes_elegidos") as mock_guardar:
        respuesta = cliente.post(
            "/deposito/pedido/7/renglones/55/lotes",
            data={"cliente_id": "1", "fecha": "2026-09-02", "sucursal": "VL",
                  "reparto": '[{"tipo_lote": "guia", "origen_id": 101, "bultos": 2}]'},
            follow_redirects=False,
        )

    # 2 sobre un renglón de 15: se guarda igual, sin un solo freno.
    assert respuesta.status_code == 303
    mock_guardar.assert_called_once()


# --- La puerta de Corregir Recepción (Gerencia) ---


def test_corregir_recepcion_SIN_CLAVE_CONFIGURADA_no_deja_pasar():
    """El default de Gerencia se da vuelta acá, y es lo que hace que la
    pantalla sirva: las de consulta se abren igual mientras la variable no
    esté cargada, para que un deploy no trabe nada. Esta ESCRIBE — puede
    dejar sin explicación el costo congelado de una guía R— así que sin
    puerta no entra nadie.

    Si esto se rompiera, la pantalla quedaría abierta a cualquiera que sepa
    la dirección en la empresa a la que le falte la variable, y nadie se
    enteraría: se vería normal.
    """
    with patch("app.main._clave_gerencia", return_value=None):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 503
    # Y dice QUÉ falta y DÓNDE cargarlo: un "no autorizado" pelado manda a
    # buscar un permiso que no existe.
    assert "CLAVE_GERENCIA" in respuesta.text
    assert "Railway" in respuesta.text


def test_corregir_recepcion_SIN_CLAVE_tampoco_deja_GUARDAR():
    """La puerta va en las dos rutas. Solo en el GET, cualquiera podría
    guardar mandando el POST directo."""
    with patch("app.main._clave_gerencia", return_value=None), \
         patch("app.main.corregir_recepcion_compra") as mock_guardar:
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "10", "cantidad_total_real": "40"},
        )

    assert respuesta.status_code == 503
    mock_guardar.assert_not_called()


def test_corregir_recepcion_con_clave_pero_sin_cookie_pide_la_clave():
    with patch("app.main._clave_gerencia", return_value="secreta"):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert respuesta.status_code == 401
    assert "clave" in respuesta.text.lower()


def test_la_url_vieja_de_corregir_recepcion_redirige_a_la_de_gerencia():
    """La pantalla se mudó: un link guardado en el celular de alguien no
    puede terminar en un 404. Mismo criterio que /gerencia/auditoria."""
    respuesta = cliente.get("/compras/30/corregir-recepcion", follow_redirects=False)

    assert respuesta.status_code == 301
    assert respuesta.headers["location"] == "/gerencia/compras/30/corregir-recepcion"


def test_desde_el_detalle_se_llega_a_corregir_y_la_clave_devuelve_A_ESA_COMPRA():
    """El que va a corregir viene mirando esa compra, no entrando por el hub.

    Así que el botón sigue estando en el detalle, y la puerta tiene que
    devolverlo AL MISMO renglón: si la clave lo dejara en el hub de
    Gerencia, tendría que volver a buscar la compra — y con el número mal
    ya adentro de la cabeza, que es justo cuando se busca otra.
    """
    with patch("app.main._clave_gerencia", return_value="secreta"):
        # 1. Sin cookie, la pantalla de la compra pide la clave...
        puerta = cliente.get("/gerencia/compras/30/corregir-recepcion")
        assert puerta.status_code == 401
        # ...y se acuerda de a dónde volver.
        assert 'value="/gerencia/compras/30/corregir-recepcion"' in puerta.text

        # 2. Con la clave correcta, vuelve a ESA compra y no al hub.
        try:
            entrada = cliente.post(
                "/gerencia/clave",
                data={"clave": "secreta", "volver": "/gerencia/compras/30/corregir-recepcion"},
                follow_redirects=False,
            )
            assert entrada.status_code == 303
            assert entrada.headers["location"] == "/gerencia/compras/30/corregir-recepcion"
        finally:
            cliente.cookies.clear()


def test_el_detalle_de_la_compra_sigue_teniendo_el_boton_y_avisa_que_pide_clave():
    """Que el botón diga Gerencia no es decoración: sin eso, el que lo toca
    se encuentra una pantalla de clave sin entender por qué."""
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, estado="recepcionado")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.listar_fotos_de_guia", return_value=[]),
    ):
        respuesta = cliente.get("/compras/30/detalle")

    assert 'href="/gerencia/compras/30/corregir-recepcion"' in respuesta.text
    assert "Gerencia" in respuesta.text


# --- Corregir Recepción: la lista de dependencias y el segundo toque ---

SIN_USAR = {"entraron": 40.0, "guias_r": [], "renglones": [], "salieron": 0.0,
            "sin_lote_de_mas": 0.0, "guias_rotas": []}


def _dependencias_usadas(**extra):
    base = {
        "entraron": 40.0,
        "guias_r": [{"reproceso_id": 38, "fecha": date(2026, 9, 1), "bultos": 12.0},
                    {"reproceso_id": 41, "fecha": date(2026, 9, 2), "bultos": 18.0}],
        "renglones": [{"cliente_id": 1, "fecha": date(2026, 9, 2), "bultos": 2.0,
                       "elegido_a_mano": False},
                      {"cliente_id": 1, "fecha": date(2026, 9, 1), "bultos": 2.0,
                       "elegido_a_mano": True}],
        "salieron": 34.0,
        "sin_lote_de_mas": 0.0,
        "guias_rotas": [],
    }
    base.update(extra)
    return base


def test_corregir_recepcion_con_el_lote_SIN_USAR_lo_dice_en_una_linea_y_sin_cartel():
    """El amarillo se guarda para cuando algo se rompe. Si aparece igual, en
    dos meses se deja de leer."""
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, unidad_compra="kilo")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.dependencias_del_lote_de_compra", return_value=SIN_USAR),
        patch("app.main.listar_clientes", return_value=[]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    assert "Nada todavía" in respuesta.text
    # El cartel amarillo ni se dibuja (el CSS está siempre; el bloque no).
    assert '<div class="advertencia">' not in respuesta.text


def test_corregir_recepcion_lista_las_guias_R_y_los_renglones_por_separado():
    """Son dos clases de dato distintas y la pantalla lo dice: las guías R
    están congeladas, los renglones los recalcula el FIFO cada vez."""
    compra = dict(COMPRA_DETALLE_DE_PRUEBA, unidad_compra="kilo")
    with (
        patch("app.main.obtener_detalle_compra", return_value=compra),
        patch("app.main.dependencias_del_lote_de_compra", return_value=_dependencias_usadas()),
        patch("app.main.listar_clientes", return_value=[{"id": 1, "nombre": "Día"}]),
    ):
        respuesta = cliente.get("/gerencia/compras/30/corregir-recepcion")

    texto = respuesta.text
    assert "ya salieron 34" in texto
    assert "R41" in texto and "R38" in texto
    assert "Día" in texto
    # El que corrigió el reparto a mano se marca: es el que decidió una persona.
    assert "lo eligió a mano" in texto
    # Y la asimetría, escrita: sin esto, dos aperturas con números distintos
    # se leen como un bug.
    assert "pueden cambiar solos" in texto


def test_corregir_recepcion_bajar_sin_romper_nada_guarda_DE_UNA():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.dependencias_del_lote_de_compra", return_value=_dependencias_usadas()),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.corregir_recepcion_compra") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "35", "cantidad_total_real": "40"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once()


def test_corregir_recepcion_bajando_de_mas_pide_el_SEGUNDO_TOQUE():
    impacto = _dependencias_usadas(sin_lote_de_mas=5.0)
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.dependencias_del_lote_de_compra", return_value=impacto),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.corregir_recepcion_compra") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "25", "cantidad_total_real": "40"},
        )

    assert respuesta.status_code == 400
    assert "5 bultos que hoy tienen lote se quedan sin lote" in respuesta.text
    assert "De 40 bultos a 25." in respuesta.text
    # Y abajo queda EL NÚMERO QUE ESCRIBIÓ, no el viejo: ver 40 otra vez se
    # leería como que no se guardó nada.
    assert 'value="25"' in respuesta.text
    mock_guardar.assert_not_called()


def test_corregir_recepcion_el_aviso_NOMBRA_la_guia_R_rota():
    """"Algunas" deja mirando el cartel: lo que hay que hacer es anular ESAS."""
    impacto = _dependencias_usadas(
        sin_lote_de_mas=5.0,
        guias_rotas=[{"reproceso_id": 41, "fecha": date(2026, 9, 2), "bultos": 18.0}],
    )
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.dependencias_del_lote_de_compra", return_value=impacto),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.corregir_recepcion_compra"),
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "25", "cantidad_total_real": "40"},
        )

    assert "la guía R41 va" in respuesta.text
    # Y le dice la verdad al que está por apretar Guardar igual: esto no se
    # ve en ningún lado después.
    assert "Nadie te lo va a volver a avisar" in respuesta.text


def test_corregir_recepcion_confirmado_SI_guarda_sin_volver_a_simular():
    with (
        patch("app.main.obtener_detalle_compra", return_value=COMPRA_DETALLE_DE_PRUEBA),
        patch("app.main.dependencias_del_lote_de_compra") as mock_impacto,
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main.corregir_recepcion_compra") as mock_guardar,
    ):
        respuesta = cliente.post(
            "/gerencia/compras/30/corregir-recepcion",
            data={"cantidad_cajones_real": "25", "cantidad_total_real": "40",
                  "confirmado": "1"},
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    mock_guardar.assert_called_once()
    # Con el segundo toque ni se vuelve a simular: ya decidió.
    mock_impacto.assert_not_called()


# --- El color de la tarjeta: lo que queda por hacer, no lo que se tocó ---


def _pedido_listado(**extra):
    base = {"id": 60, "fecha_operacion": date(2026, 8, 21), "origen": "mail",
            "creado_en": datetime(2026, 8, 21, 12, 0), "armado_cerrado_el": None,
            "renglones_totales": 53, "renglones_armados": 53, "sin_identificar": 0,
            "renglones_cortos": 0}
    base.update(extra)
    return base


def _armar_con(listado):
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_pedidos_vigentes_con_armado", return_value=listado),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
    ):
        return cliente.get("/deposito/pedido/armar?cliente_id=1")


def test_un_pedido_TERMINADO_con_renglones_cortos_NO_va_en_verde():
    """El verde que mentía. "53 de 53" cuenta renglones TILDADOS, y uno
    tildado con "armé 12 de 15" cuenta como armado: el pedido salía verde con
    un tilde y le decía al que mira que no hay nada que hacer."""
    respuesta = _armar_con([_pedido_listado(
        armado_cerrado_el=datetime(2026, 8, 21, 18, 0), renglones_cortos=4)])

    texto = respuesta.text
    assert "atencion" in texto
    assert "4 armados de menos" in texto
    # Y no queda ni el verde ni el tilde.
    assert "boton-sucursal completa" not in texto
    assert "TERMINADO ✔" not in texto


def test_un_pedido_TERMINADO_con_renglones_SIN_IDENTIFICAR_tampoco_va_en_verde():
    """Los sin identificar ni siquiera entran en renglones_totales: un pedido
    con 5 decía "53 de 53 · TERMINADO ✔" en verde. Es el verde que miente en
    su forma más pura."""
    respuesta = _armar_con([_pedido_listado(
        armado_cerrado_el=datetime(2026, 8, 21, 18, 0), sin_identificar=5)])

    assert "atencion" in respuesta.text
    assert "5 sin identificar" in respuesta.text
    assert "TERMINADO ✔" not in respuesta.text


def test_un_pedido_TERMINADO_dejando_renglones_sin_armar_tampoco():
    respuesta = _armar_con([_pedido_listado(
        armado_cerrado_el=datetime(2026, 8, 21, 18, 0), renglones_armados=48)])

    assert "5 sin armar" in respuesta.text
    assert "TERMINADO ✔" not in respuesta.text


def test_el_pedido_que_SI_salio_completo_sigue_en_verde_con_su_tilde():
    """El amarillo tiene que significar algo: si apareciera igual en los que
    están bien, en dos meses se deja de mirar."""
    respuesta = _armar_con([_pedido_listado(armado_cerrado_el=datetime(2026, 8, 21, 18, 0))])

    texto = respuesta.text
    assert "TERMINADO ✔" in texto
    # El CSS del amarillo está siempre; la clase en la tarjeta, no.
    assert 'class="boton-sucursal atencion"' not in texto


def test_armar_de_MAS_no_cuenta_como_incompleto_en_las_sucursales():
    """El bug del "!=": la alerta de Auditoría ya lo tenía con "<" y la
    pantalla no. Un renglón de 18 sobre 15 pedidos figuraba como incompleto,
    y con eso el color habría dicho que falta mercadería donde sobró."""
    renglones = [
        {"id": 1, "sucursal": "VL", "articulo_id": 1, "ficha_id": 901, "nombre_venta": "Banana",
         "cantidad": 15, "armado_el": datetime(2026, 8, 21, 9, 0), "cantidad_armada": 18,
         "kilos_enviados": 180.0, "anulado_el": None, "texto_codigo": None,
         "texto_descripcion": None},
    ]
    with (
        patch("app.main._hoy_argentina", return_value=date(2026, 8, 22)),
        patch("app.main.listar_clientes", return_value=CLIENTES_PARA_SELECTOR),
        patch("app.main.listar_mails_pedido_sin_procesar_de_cliente", return_value=[]),
        patch("app.main.obtener_pedido_vigente", return_value={
            "id": 7, "cliente_id": 1, "fecha_operacion": date(2026, 8, 21),
            "reemplaza_a_pedido_id": None, "armado_cerrado_el": None, "origen": "mail",
            "recibido_el": datetime(2026, 8, 21, 12, 0)}),
        patch("app.main.listar_sucursales_pedido", return_value=[{"sucursal": "VL", "oc": None}]),
        patch("app.main.listar_renglones_pedido", return_value=renglones),
        patch("app.main.listar_fichas_por_cliente", return_value=[]),
        patch("app.main.fichas_con_cajas_armadas", return_value=set()),
    ):
        respuesta = cliente.get("/deposito/pedido/armar?cliente_id=1&fecha=2026-08-21")

    # Armó 18 de 15: está completo, no incompleto. (El "incompleto" del
    # comentario del CSS no cuenta: se mira la tarjeta.)
    assert "· 1 incompleto" not in respuesta.text
    assert "boton-sucursal completa" in respuesta.text


def test_el_banner_dice_DESDE_CUANDO_y_con_el_titulo_corto():
    """"Pedidos incompletos (3)" no dice si es de hoy o de la semana pasada,
    y esa es la diferencia entre ir ahora o después. El título va corto
    porque el banner es una cinta que corre en 390px; Auditoría muestra el
    largo, que ahí sobra lugar."""
    from app.alertas import DefinicionAlerta, unir

    definicion = DefinicionAlerta(
        codigo="pedidos_incompletos",
        titulo="Pedidos con renglones incompletos (se armó menos de lo pedido)",
        titulo_corto="Pedidos incompletos",
        url="/deposito/pedido",
        texto_link="Ver en Pedido",
        contar=lambda: {"casos": 3, "mas_viejo": date(2026, 8, 28)},
        modulos=("deposito",),
    )
    estado = [{"codigo": "pedidos_incompletos", "casos": 3, "mas_viejo": date(2026, 8, 28),
               "calculada_el": None, "error": None, "duracion_ms": None}]

    (alerta,) = unir([definicion], estado)

    assert alerta["titulo_corto"] == "Pedidos incompletos"
    assert alerta["titulo"].startswith("Pedidos con renglones incompletos")
    assert alerta["mas_viejo"] == date(2026, 8, 28)


def test_una_alerta_sin_titulo_corto_usa_el_largo():
    """Vacío = el título entero, que es lo normal: 14 de las 16 alertas no
    necesitan uno corto y no hay que ponérselo."""
    from app.alertas import DefinicionAlerta, unir

    definicion = DefinicionAlerta(
        codigo="compras_sin_precio", titulo="Compras sin precio de compra cargado",
        url="/compras/pendientes", texto_link="Ver", contar=lambda: 1, modulos=("compras",),
    )

    (alerta,) = unir([definicion], [])

    assert alerta["titulo_corto"] == "Compras sin precio de compra cargado"


def test_reproceso_tiene_CANCELAR_que_solo_sale_al_hub_de_stock():
    """Descartar el formulario en curso, no anular una guía guardada (eso es
    "Anular guía" en Guías R).

    Es un ENLACE y no un botón a propósito: cancelar no le manda nada al
    server, así que no hay forma de que borre algo. Lo que había cargado
    —incluido un reparto editado en el desglose— se descarta solo, porque
    nunca se escribió.
    """
    with (
        patch("app.main.listar_articulos_para_reproceso", return_value=[]),
        patch("app.main.listar_clientes", return_value=[]),
        patch("app.main._ayudas_ficha_por_cliente_y_articulo", return_value={}),
        patch("app.main._fichas_por_cliente_y_articulo", return_value={}),
        patch("app.main.fecha_corte", return_value=date(2026, 8, 20)),
    ):
        respuesta = cliente.get("/deposito/stock/reproceso")

    texto = respuesta.text
    assert '<a class="boton-cancelar" href="/deposito/stock">Cancelar</a>' in texto
    # Sin confirmación: el que aprieta Cancelar quiere salir, no discutir.
    assert "confirm(" not in texto.split("boton-cancelar")[1][:400]
    # Y Guardar se lleva el ancho: los dos juntos, pero no del mismo peso.
    assert "flex: 2;" in texto and "flex: 1;" in texto


# --- ABM de proveedores de compras ---

PROVEEDORES_ABM_DE_PRUEBA = [
    {"id": 1, "codigo_puesto": "N01P02", "nombre": "Don Pedro", "activo": True, "compras": 12},
    {"id": 2, "codigo_puesto": "L03P11", "nombre": "La Rosa", "activo": True, "compras": 1},
    {"id": 3, "codigo_puesto": "N99P99", "nombre": "Tipeo Mal", "activo": False, "compras": 0},
]


def test_abm_proveedores_compras_muestra_codigo_estado_y_cuantas_compras():
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.get("/compras/proveedores")

    assert respuesta.status_code == 200
    assert "N01P02" in respuesta.text
    # El conteo es lo que separa el fantasma de un código mal tipeado del
    # proveedor de verdad que alguien está por esconder sin querer.
    assert "12 compras" in respuesta.text
    assert "1 compra" in respuesta.text
    assert "sin compras cargadas" in respuesta.text
    assert "De baja" in respuesta.text


def test_abm_proveedores_compras_no_ofrece_alta_de_proveedor_nuevo():
    # Un proveedor de compras nace solo al cargar una compra con un código
    # nuevo. Un alta acá crearía códigos sin ninguna compra detrás.
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.get("/compras/proveedores")

    assert 'action="/compras/proveedores/nuevo"' not in respuesta.text
    assert "No se cargan de acá" in respuesta.text


def test_abm_proveedores_compras_el_activo_ofrece_baja_y_el_de_baja_ofrece_alta():
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.get("/compras/proveedores")

    assert 'action="/compras/proveedores/1/baja"' in respuesta.text
    assert 'action="/compras/proveedores/1/alta"' not in respuesta.text
    # La baja no es un camino de ida.
    assert 'action="/compras/proveedores/3/alta"' in respuesta.text
    assert 'action="/compras/proveedores/3/baja"' not in respuesta.text


def test_abm_proveedores_compras_la_confirmacion_de_baja_dice_cuantas_compras_tiene():
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.get("/compras/proveedores")

    assert "Tiene 12 compras cargadas, que quedan como están." in respuesta.text
    assert "Solo deja de aparecer para elegir al cargar una compra." in respuesta.text


def test_abm_proveedores_compras_no_deja_editar_el_codigo():
    # codigo_puesto es la identidad: cambiarlo movería todas las compras a
    # otro proveedor. El formulario de edición manda solo el nombre.
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.get("/compras/proveedores")

    bloque = respuesta.text.split('action="/compras/proveedores/1/renombrar"')[1].split("</form>")[0]
    assert 'name="nombre"' in bloque
    assert 'name="codigo_puesto"' not in bloque


def test_renombrar_proveedor_compras_redirige_con_aviso():
    with patch("app.main.renombrar_proveedor") as mock_renombrar:
        respuesta = cliente.post(
            "/compras/proveedores/7/renombrar", data={"nombre": "  Don   Pedro  "}, follow_redirects=False
        )

    assert respuesta.status_code == 303
    # El nombre se normaliza igual que en el ABM del puesto.
    mock_renombrar.assert_called_once_with(7, "Don Pedro")
    assert "Proveedor+renombrado" in respuesta.headers["location"]


def test_renombrar_proveedor_compras_sin_nombre_da_400():
    with patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA):
        respuesta = cliente.post("/compras/proveedores/7/renombrar", data={"nombre": "   "})

    assert respuesta.status_code == 400
    assert "El nombre del proveedor es obligatorio." in respuesta.text


def test_baja_y_alta_de_proveedor_compras_usan_la_misma_funcion():
    with patch("app.main.cambiar_actividad_proveedor") as mock_cambiar:
        respuesta = cliente.post("/compras/proveedores/7/baja", follow_redirects=False)
    assert respuesta.status_code == 303
    mock_cambiar.assert_called_once_with(7, False)
    assert "dado+de+baja" in respuesta.headers["location"]

    with patch("app.main.cambiar_actividad_proveedor") as mock_cambiar:
        respuesta = cliente.post("/compras/proveedores/7/alta", follow_redirects=False)
    assert respuesta.status_code == 303
    mock_cambiar.assert_called_once_with(7, True)
    assert "dado+de+alta" in respuesta.headers["location"]


def test_baja_de_proveedor_compras_que_ya_no_existe_da_400():
    with (
        patch("app.main.cambiar_actividad_proveedor", side_effect=ValueError("Ese proveedor ya no existe.")),
        patch("app.main.listar_proveedores_para_abm", return_value=PROVEEDORES_ABM_DE_PRUEBA),
    ):
        respuesta = cliente.post("/compras/proveedores/7/baja")

    assert respuesta.status_code == 400
    assert "Ese proveedor ya no existe." in respuesta.text


def test_el_hub_de_compras_lleva_al_abm_de_proveedores():
    respuesta = cliente.get("/compras")
    assert respuesta.status_code == 200
    assert 'href="/compras/proveedores"' in respuesta.text


def test_cargar_una_compra_con_el_codigo_de_un_proveedor_de_baja_lo_dice():
    # La reactivación es correcta (si llegó mercadería, el proveedor
    # existe), pero no puede ser silenciosa: el que lo dio de baja tiene
    # que poder enterarse de que volvió.
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, True)),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_articulo", return_value=ARTICULOS_CON_UNIDAD_COMPRA[0]),
        patch("app.main.crear_compra"),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data={
                "codigo_puesto": "N01P02",
                "nombre": "Don Pedro",
                "articulo_id": str(ARTICULOS_CON_UNIDAD_COMPRA[0]["id"]),
                "cantidad_cajones": "10",
                "contenido_por_cajon": "20",
                "importe": "1000",
                "sena": "0",
                "tipo_retiro": "Clark",
                "accion": "agregar",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "aviso=" in respuesta.headers["location"]
    assert "volvi%C3%B3+a+quedar+activo" in respuesta.headers["location"]


def test_cargar_una_compra_de_un_proveedor_activo_no_avisa_nada():
    with (
        patch("app.main.obtener_o_crear_proveedor_por_codigo", return_value=(200, False)),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.obtener_articulo", return_value=ARTICULOS_CON_UNIDAD_COMPRA[0]),
        patch("app.main.crear_compra"),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
    ):
        respuesta = cliente.post(
            "/compras/nueva/manual",
            data={
                "codigo_puesto": "N01P02",
                "nombre": "Don Pedro",
                "articulo_id": str(ARTICULOS_CON_UNIDAD_COMPRA[0]["id"]),
                "cantidad_cajones": "10",
                "contenido_por_cajon": "20",
                "importe": "1000",
                "sena": "0",
                "tipo_retiro": "Clark",
                "accion": "agregar",
            },
            follow_redirects=False,
        )

    assert respuesta.status_code == 303
    assert "aviso=" not in respuesta.headers["location"]


# --- El nombre con & escapado dos veces ---

PROVEEDOR_CON_AMPERSAND = [{"id": 1, "codigo_puesto": "N01P02", "nombre": "RIO URUGUAY & GOLOSO"}]


def test_el_nombre_con_ampersand_no_sale_escapado_dos_veces_en_el_autocompletar():
    # Jinja escapaba el & dentro de una CADENA DE JAVASCRIPT, y el JS lo pone
    # con textContent, que no desescapa: en pantalla se leía "&amp;". El JS
    # estaba bien; sobraba el escape de la plantilla. Con |tojson el & viaja
    # como \u0026 y el navegador lo devuelve a "&" al parsear el literal.
    with (
        patch("app.main.listar_todos_los_proveedores", return_value=PROVEEDOR_CON_AMPERSAND),
        patch("app.main.listar_articulos", return_value=ARTICULOS_CON_UNIDAD_COMPRA),
        patch("app.main.buscar_compras", return_value=[]),
    ):
        respuesta = cliente.get("/compras/buscar")

    assert respuesta.status_code == 200
    assert "&amp;amp;" not in respuesta.text
    assert 'nombre: "RIO URUGUAY &amp; GOLOSO"' not in respuesta.text
    assert "RIO URUGUAY \\u0026 GOLOSO" in respuesta.text


def test_ninguna_plantilla_mete_un_nombre_a_mano_dentro_de_una_cadena_de_javascript():
    # El corolario: cuando se arregla una copia hay que ir a buscar la otra.
    # Eran 14 lugares en 13 plantillas con el mismo patrón; arreglar solo el
    # que se vio en una captura es exactamente el error que ya nos costó tres
    # veces esta semana.
    import glob
    import re

    patron = re.compile(r'(nombre|codigo):\s*"\{\{')
    culpables = []
    for ruta in sorted(glob.glob("templates/*.html")):
        with open(ruta, encoding="utf-8") as archivo:
            for numero, linea in enumerate(archivo, 1):
                if patron.search(linea):
                    culpables.append(f"{ruta}:{numero}")

    assert culpables == [], "Escapan de más (usar |tojson): " + ", ".join(culpables)


# --- Los dos sietes ---

def test_la_ventana_de_pedidos_incompletos_sale_del_listado_y_no_de_un_7_suelto():
    # Eran dos sietes que coincidian por casualidad y no se nombraban: el de
    # la alerta y el del listado de /deposito/pedido. El dia que alguien
    # cambiara uno, el banner habria contado pedidos que la pantalla adonde
    # lleva no muestra — el link diciendo "3" y la pantalla mostrando 2, sin
    # ningun error. Este test falla si vuelven a separarse.
    from app.main import ALERTAS, DIAS_PASADOS_LISTADO_PEDIDOS

    alerta = next(a for a in ALERTAS if a.codigo == "pedidos_incompletos")
    # Un valor DISTINTO de 7 a proposito: si el lambda tuviera el numero
    # escrito a mano, este test pasaría igual con 7 y no probaría nada.
    with (
        patch("app.main.DIAS_PASADOS_LISTADO_PEDIDOS", 3),
        patch("app.main._hoy_argentina", return_value=HOY_DE_PRUEBA),
        patch("app.main.contar_pedidos_incompletos", return_value={"casos": 0, "mas_viejo": None}) as mock_contar,
    ):
        alerta.contar()

    mock_contar.assert_called_once_with(HOY_DE_PRUEBA - timedelta(days=3))
    # Y el default sigue siendo el de la pantalla, sin sorpresas.
    assert DIAS_PASADOS_LISTADO_PEDIDOS == 7

