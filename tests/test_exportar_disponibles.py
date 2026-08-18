from datetime import date
from io import BytesIO

import openpyxl

from core.exportar_disponibles import generar_excel_disponibles

FILAS_DE_PRUEBA = [
    {"codigo": "90039", "nombre": "Manzana Roja", "cantidad": 40.0},
    {"codigo": None, "nombre": "Frutilla", "cantidad": 0.0},
    {"codigo": "10021", "nombre": "Uva", "cantidad": 0.0},
]


def _cargar_excel(excel_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(excel_bytes))


def test_generar_excel_disponibles_hoja_y_titulo():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja.title == "Stock Actualizado"
    assert hoja["A1"].value == "FRUTAMAX - Disponibilidad de Stock"
    assert "A1:C1" in [str(rango) for rango in hoja.merged_cells.ranges]


def test_generar_excel_disponibles_titulo_usa_el_nombre_de_empresa_recibido():
    # Mismo código para varias empresas: el título sale del parámetro, en
    # mayúsculas — nada de "Frutamax" fijo adentro del módulo.
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Palmalá")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja["A1"].value == "PALMALÁ - Disponibilidad de Stock"


def test_generar_excel_disponibles_fecha_un_solo_dia():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja["A2"].value == "Fecha: 14/08/2026"
    assert "A2:C2" in [str(rango) for rango in hoja.merged_cells.ranges]


def test_generar_excel_disponibles_fecha_rango_de_dos_dias():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 15), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja["A2"].value == "Fecha: 14/08/2026 al 15/08/2026"


def test_generar_excel_disponibles_filas_vacias_y_encabezado():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    for fila in (3, 4, 5):
        assert hoja.cell(row=fila, column=1).value is None
        assert hoja.cell(row=fila, column=2).value is None
        assert hoja.cell(row=fila, column=3).value is None

    assert hoja.cell(row=6, column=1).value == "Código"
    assert hoja.cell(row=6, column=2).value == "Producto"
    assert hoja.cell(row=6, column=3).value == "Stock"


def test_generar_excel_disponibles_datos_desde_fila_7_en_el_orden_dado():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja.cell(row=7, column=1).value == "90039"
    assert hoja.cell(row=7, column=2).value == "Manzana Roja"
    assert hoja.cell(row=7, column=3).value == 40.0

    assert hoja.cell(row=8, column=2).value == "Frutilla"
    assert hoja.cell(row=9, column=2).value == "Uva"


def test_generar_excel_disponibles_codigo_vacio_deja_la_celda_vacia():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    # Frutilla (segunda fila de datos, fila 8) no tiene código.
    assert hoja.cell(row=8, column=1).value is None


def test_generar_excel_disponibles_cantidad_cero_se_escribe_no_se_omite():
    excel_bytes = generar_excel_disponibles(date(2026, 8, 14), date(2026, 8, 14), FILAS_DE_PRUEBA, "Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    # Uva (fila 9) tiene cantidad 0 — tiene que figurar como 0.0, no vacío.
    assert hoja.cell(row=9, column=3).value == 0.0
