from datetime import date
from io import BytesIO

import openpyxl
import pypdfium2 as pdfium

from core.exportar_precios import LEYENDA_PRECIO_NUEVO, generar_excel_lista_precios, generar_pdf_lista_precios

FILAS_DE_PRUEBA = [
    {"articulo_nombre": "Mango", "grupo": "fruta", "precio": 350.0, "precio_anterior": 300.0, "unidad": "unidad", "es_nuevo": False},
    {"articulo_nombre": "Manzana Red", "grupo": "fruta", "precio": 890.0, "unidad": "kilo", "es_nuevo": True},
    {"articulo_nombre": "Morrón Rojo", "grupo": "hortaliza", "precio": 1200.0, "unidad": "kilo", "es_nuevo": False},
    {"articulo_nombre": "Tomate Cherry", "grupo": "hortaliza", "precio": 500.0, "unidad": "kilo", "es_nuevo": True},
    {"articulo_nombre": "Rúcula", "grupo": "hoja", "precio": 700.0, "unidad": "unidad", "es_nuevo": False},
    {"articulo_nombre": "Cubeta X", "grupo": "pesada", "precio": 15000.0, "unidad": "cubeta", "es_nuevo": False},
    {"articulo_nombre": "Articulo Raro", "grupo": None, "precio": 100.0, "unidad": None, "es_nuevo": False},
]


def _texto_del_pdf(pdf_bytes: bytes) -> str:
    documento = pdfium.PdfDocument(pdf_bytes)
    textos = []
    for pagina in documento:
        textos.append(pagina.get_textpage().get_text_range())
    return "\n".join(textos)


def _textos_por_pagina(pdf_bytes: bytes) -> list[str]:
    documento = pdfium.PdfDocument(pdf_bytes)
    return [pagina.get_textpage().get_text_range() for pagina in documento]


def _texto_sin_leyenda(pdf_bytes: bytes) -> str:
    # La leyenda fija se repite en el encabezado de CADA página (para que
    # se vea aunque una sección se corte entre dos hojas) — se descarta acá
    # para poder contar los badges reales de las filas sin que la cantidad
    # de páginas altere el resultado.
    return _texto_del_pdf(pdf_bytes).replace(LEYENDA_PRECIO_NUEVO, "")


# --- generar_pdf_lista_precios ---


def test_generar_pdf_incluye_encabezado_y_pie():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    # El título lleva la empresa: el mismo cliente puede recibir listas de
    # más de una y tiene que poder distinguirlas.
    assert "Lista de Precios — Frutamax" in texto
    assert "Cliente: Día" in texto
    assert "Vigencia: 16/08/2026 · Precios + IVA" in texto
    assert "Nuevo precio indica los productos cuyo precio fue actualizado." in texto
    assert "Todos los precios están expresados en pesos y no incluyen IVA." in texto


def test_generar_pdf_titulo_usa_el_nombre_de_empresa_recibido():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Palmala")
    texto = _texto_del_pdf(pdf_bytes)

    assert "Lista de Precios — Palmala" in texto
    assert "Frutamax" not in texto


def test_generar_pdf_secciones_en_el_orden_fruta_hortaliza_hoja_pesada_sin_clasificar():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    posiciones = [texto.index(titulo) for titulo in ("FRUTA", "HORTALIZA", "HOJA", "PESADA", "SIN CLASIFICAR")]
    assert posiciones == sorted(posiciones)


def test_generar_pdf_articulo_sin_grupo_cae_en_sin_clasificar():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    bloque_sin_clasificar = texto[texto.index("SIN CLASIFICAR") :]
    assert "Articulo Raro" in bloque_sin_clasificar


def test_generar_pdf_precios_formateados_con_signo_pesos_y_separador_de_miles():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    assert "$ 1.200" in texto
    assert "$ 15.000" in texto


def test_generar_pdf_unidad_segun_la_ficha():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    assert "por kilo" in texto
    assert "por unidad" in texto
    assert "por cubeta" in texto


def test_generar_pdf_hoy_muestra_badge_nuevo_precio_solo_en_los_nuevos():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_sin_leyenda(pdf_bytes)

    # Solo Manzana Red y Tomate Cherry vinieron con es_nuevo=True.
    assert texto.count("Nuevo precio") == 2


def test_generar_pdf_fecha_pasada_resalta_igual_lo_que_empezo_a_regir_ese_dia():
    # Lo que se resalta es "empezó a regir en la fecha de esta lista", y eso
    # no depende de que la fecha sea HOY: consultar para atrás es justamente
    # para ver qué cambió ESE día (sirve para facturar sin comparar contra la
    # lista anterior). Antes el badge se apagaba en toda fecha pasada.
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 1, 15), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_sin_leyenda(pdf_bytes)

    assert texto.count("Nuevo precio") == 2


def test_generar_pdf_badge_nuevo_precio_va_en_su_propia_columna_a_la_derecha():
    # El badge va en una columna aparte, después de Precio y Unidad — no
    # debajo del producto ni partiendo la fila Precio/Unidad.
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    bloque_fruta = texto[texto.index("FRUTA") : texto.index("HORTALIZA")]
    posicion_nombre = bloque_fruta.index("Manzana Red")
    posicion_precio = bloque_fruta.index("$ 890")
    posicion_unidad = bloque_fruta.index("por kilo")
    posicion_badge = bloque_fruta.index("Nuevo precio")
    assert posicion_nombre < posicion_precio < posicion_unidad < posicion_badge


def test_generar_pdf_badge_nuevo_precio_tiene_puntito():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    assert "• Nuevo precio" in texto


def test_generar_pdf_leyenda_tiene_puntito_rojo():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    assert "• Nuevo precio indica los productos cuyo precio fue actualizado." in texto


def test_generar_pdf_cada_grupo_empieza_en_su_propia_pagina():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    paginas = _textos_por_pagina(pdf_bytes)

    assert len(paginas) == 5  # fruta, hortaliza, hoja, pesada, sin clasificar
    assert "FRUTA" in paginas[0] and "HORTALIZA" not in paginas[0]
    assert "HORTALIZA" in paginas[1] and "FRUTA" not in paginas[1] and "PESADA" not in paginas[1]
    assert "HOJA" in paginas[2] and "HORTALIZA" not in paginas[2] and "PESADA" not in paginas[2]
    assert "PESADA" in paginas[3] and "HOJA" not in paginas[3]
    assert "SIN CLASIFICAR" in paginas[4]


def test_generar_pdf_repite_encabezado_en_cada_pagina():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, nombre_empresa="Frutamax")
    paginas = _textos_por_pagina(pdf_bytes)

    assert len(paginas) > 1  # si no hay más de una página, esto no prueba nada
    for texto_pagina in paginas:
        assert "Lista de Precios" in texto_pagina
        assert "Cliente: Día" in texto_pagina
        assert "Vigencia: 16/08/2026 · Precios + IVA" in texto_pagina
        assert LEYENDA_PRECIO_NUEVO in texto_pagina


def test_generar_pdf_sin_filas_no_rompe():
    pdf_bytes = generar_pdf_lista_precios("Día", date(2026, 8, 16), [], nombre_empresa="Frutamax")
    texto = _texto_del_pdf(pdf_bytes)

    assert "Lista de Precios" in texto
    assert "FRUTA" not in texto


# --- generar_excel_lista_precios ---


def _cargar_excel(excel_bytes: bytes):
    return openpyxl.load_workbook(BytesIO(excel_bytes))


# El Excel usa el formato de planilla que el dueño ya manejaba a mano
# (pedido explícito del 21/08/2026): fecha arriba, Producto | Precio
# Anterior | Precio Desde HOY, cambios en naranja con el precio en rojo,
# pie "PRECIOS POR KG - SIN IVA". El PDF conserva su propio diseño.


def test_generar_excel_usa_el_formato_de_la_planilla_del_dueno():
    excel_bytes = generar_excel_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, es_hoy=True, nombre_empresa="Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    assert hoja.cell(row=1, column=1).value == "16/8/2026"  # sin ceros a la izquierda
    assert hoja.cell(row=2, column=1).value == "Producto"
    assert hoja.cell(row=2, column=2).value == "Precio Anterior"
    assert hoja.cell(row=2, column=3).value == "Precio Desde HOY"

    valores = [celda.value for fila in hoja.iter_rows() for celda in fila if celda.value is not None]
    assert "PRECIOS POR KG - SIN IVA" in valores
    # La planilla original no lleva título de empresa ni secciones por grupo.
    assert not any(isinstance(v, str) and "Lista de Precios" in v for v in valores)
    assert "FRUTA" not in valores and "HORTALIZA" not in valores

    # Una sola tabla alfabética; lo que no se vende por kilo lleva la
    # unidad pegada al nombre, porque el pie dice "POR KG".
    nombres = [fila[0].value for fila in hoja.iter_rows() if fila[0].value and fila[0].row > 2][:-1]
    assert nombres == sorted(nombres, key=str.lower)
    assert "Mango (por unidad)" in nombres
    assert "Cubeta X (por cubeta)" in nombres
    assert "Articulo Raro" in nombres  # sin unidad conocida, sin sufijo


def _nombres_con_cambio_resaltado(hoja):
    from core.exportar_precios import NARANJA_CAMBIO_HEX

    # Columna 3 (índice 2, 0-based) es "Precio Desde HOY".
    return [
        fila[0].value
        for fila in hoja.iter_rows()
        if fila[2].fill
        and fila[2].fill.start_color
        and fila[2].fill.start_color.rgb == f"00{NARANJA_CAMBIO_HEX}"
    ]


def test_generar_excel_resalta_los_que_empezaron_a_regir_en_la_fecha_de_la_lista():
    excel_bytes = generar_excel_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, es_hoy=True, nombre_empresa="Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    # Manzana Red y Tomate Cherry son los que empezaron a regir ese día
    # (es_nuevo). Mango NO se resalta aunque su precio anterior sea distinto
    # (300 -> 350): ese cambio es de otro día y resaltarlo era un falso aviso.
    assert _nombres_con_cambio_resaltado(hoja) == ["Manzana Red", "Tomate Cherry"]
    fila_manzana = next(fila for fila in hoja.iter_rows() if fila[0].value == "Manzana Red")
    assert fila_manzana[2].font.color.rgb.endswith("C00000")
    # El precio queda como número (sumar/filtrar sigue andando).
    assert fila_manzana[2].value == 890.0
    assert "Mango (por unidad)" not in _nombres_con_cambio_resaltado(hoja)


def test_generar_excel_fecha_pasada_resalta_igual_lo_que_empezo_a_regir_ese_dia():
    # El mismo resaltado que en el día: consultar el 15/1 para atrás tiene
    # que mostrar en rojo lo que arrancó el 15/1. Antes esta columna solo
    # comparaba contra el precio anterior, sin mirar la fecha.
    excel_bytes = generar_excel_lista_precios("Día", date(2026, 1, 15), FILAS_DE_PRUEBA, es_hoy=False, nombre_empresa="Frutamax")
    libro = _cargar_excel(excel_bytes)

    assert _nombres_con_cambio_resaltado(libro.active) == ["Manzana Red", "Tomate Cherry"]


def test_generar_excel_el_precio_anterior_figura_siempre():
    excel_bytes = generar_excel_lista_precios("Día", date(2026, 8, 16), FILAS_DE_PRUEBA, es_hoy=True, nombre_empresa="Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    # Manzana Red nunca tuvo precio previo cargado: el anterior repite el
    # vigente (siempre figura, pedido explícito).
    fila_manzana = next(fila for fila in hoja.iter_rows() if fila[0].value == "Manzana Red")
    assert fila_manzana[1].value == 890.0
    assert fila_manzana[2].value == 890.0


def test_generar_excel_fecha_pasada_cambia_el_encabezado_del_precio():
    excel_bytes = generar_excel_lista_precios("Día", date(2026, 1, 15), FILAS_DE_PRUEBA, es_hoy=False, nombre_empresa="Frutamax")
    libro = _cargar_excel(excel_bytes)
    hoja = libro.active

    # "Desde HOY" mentiría en una consulta histórica.
    assert hoja.cell(row=2, column=3).value == "Precio al 15/1"
