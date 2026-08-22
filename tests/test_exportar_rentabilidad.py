from datetime import date
from io import BytesIO

import openpyxl

from core.exportar_rentabilidad import generar_excel_rentabilidad, generar_pdf_rentabilidad

RESULTADO_DE_PRUEBA = {
    "grupos": [
        {
            "grupo": "fruta",
            "etiqueta": "Fruta",
            "filas": [
                {
                    "articulo_id": 1, "articulo_nombre": "Banana", "grupo": "fruta", "unidad_venta": "kilo",
                    "bultos": 15.0, "unidades": 300.0, "venta": 30000.0, "costo_total": 24000.0,
                    "precio_promedio": 100.0, "costo_promedio": 80.0, "renta_pesos": 6000.0, "renta_pct": 20.0,
                }
            ],
            "subtotal": {"bultos": 15.0, "venta": 30000.0, "costo_total": 24000.0, "renta_pesos": 6000.0, "renta_pct": 20.0},
        }
    ],
    "totales": {
        "bultos": 15.0, "venta": 30000.0, "costo_total": 24000.0, "renta_pesos": 6000.0, "renta_pct": 20.0,
        "no_calculables_casos": 1, "no_calculables_bultos": 3.0,
    },
    "no_calculables": [
        {"motivo": "sin_identificar", "motivo_etiqueta": "Renglón sin identificar (no matchea ninguna ficha)",
         "articulo_id": None, "articulo_nombre": "Sin identificar", "bultos": 3.0, "dias": 1},
    ],
    "fechas_incluidas": [date(2026, 8, 21), date(2026, 8, 22)],
}


def test_generar_pdf_rentabilidad_arma_un_pdf():
    pdf = generar_pdf_rentabilidad(date(2026, 8, 15), date(2026, 8, 22), ["cliente Día"], RESULTADO_DE_PRUEBA)
    assert pdf.startswith(b"%PDF")


def test_generar_excel_rentabilidad_grupos_totales_y_no_calculables():
    excel = generar_excel_rentabilidad(date(2026, 8, 15), date(2026, 8, 22), ["cliente Día"], RESULTADO_DE_PRUEBA)
    hoja = openpyxl.load_workbook(BytesIO(excel)).active

    assert hoja.title == "Rentabilidad de Pedidos"
    valores = [str(celda.value) for fila in hoja.iter_rows() for celda in fila if celda.value is not None]
    texto = "\n".join(valores)
    # El subtítulo lleva la regla de la medición, para que el archivo se
    # explique solo cuando alguien lo abra en tres meses.
    assert "bultos = lo PEDIDO" in texto
    assert "2 días con pedido" in texto
    assert "cliente Día" in texto
    assert "Fruta" in texto and "Banana" in texto
    assert "Subtotal" in texto and "Total" in texto
    # Los no calculables NUNCA se caen del archivo.
    assert "Quedaron AFUERA del cálculo 1 artículos (3 bultos)" in texto
    assert "Sin identificar" in texto


def test_generar_excel_rentabilidad_sin_datos_lo_dice():
    vacio = {
        "grupos": [], "no_calculables": [], "fechas_incluidas": [],
        "totales": {"bultos": 0, "venta": 0, "costo_total": 0, "renta_pesos": 0, "renta_pct": None,
                    "no_calculables_casos": 0, "no_calculables_bultos": 0},
    }
    excel = generar_excel_rentabilidad(date(2026, 8, 15), date(2026, 8, 22), [], vacio)
    hoja = openpyxl.load_workbook(BytesIO(excel)).active
    valores = [str(celda.value) for fila in hoja.iter_rows() for celda in fila if celda.value is not None]
    assert any("No se encontraron pedidos" in v for v in valores)
