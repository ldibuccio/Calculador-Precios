from datetime import date
from io import BytesIO

import openpyxl

from core.exportar_rentabilidad_real import generar_excel_rentabilidad_real, generar_pdf_rentabilidad_real

RESULTADO_REAL = {
    "grupos": [
        {
            "grupo": "fruta",
            "etiqueta": "Fruta",
            "filas": [
                {
                    "articulo_id": 1, "articulo_nombre": "Banana", "grupo": "fruta",
                    "bultos": 10.0, "unidades": 160.0, "venta_neta": 14400.0,
                    "costo_mercaderia": 5000.0, "costo_envase": 320.0,
                    # Los $1.500 de merma, abiertos: 2 bultos crudos a $350
                    # y 1 ya trabajado a $800 — el trabajado sale más caro,
                    # que es justo lo que el desglose tiene que dejar ver.
                    "costo_mermas": 1500.0, "bultos_mermados": 3.0,
                    "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
                    "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
                    "segunda_bultos": 2.0,
                    "devoluciones_bultos": 5.0, "devoluciones_venta": 4500.0,
                    "rechazos_perdidos": 900.0, "rechazos_bultos": 2.0,
                    "costo_total": 6820.0, "renta_pesos": 3080.0, "utilidad_pct": 61.6,
                }
            ],
            "subtotal": {"bultos": 10.0, "venta_neta": 14400.0, "costo_mercaderia": 5000.0,
                         "costo_envase": 320.0, "costo_mermas": 1500.0, "costo_total": 6820.0, "bultos_mermados": 3.0,
                         "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
                         "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
                         "devoluciones_bultos": 5.0, "devoluciones_venta": 4500.0,
                         "rechazos_perdidos": 900.0, "rechazos_bultos": 2.0,
                         "renta_pesos": 3080.0, "utilidad_pct": 61.6},
        }
    ],
    "totales": {
        "bultos": 10.0, "venta_neta": 14400.0, "costo_mercaderia": 5000.0, "costo_envase": 320.0,
        "costo_mermas": 1500.0, "segunda_bultos": 2.0, "costo_total": 6820.0, "bultos_mermados": 3.0,
        "costo_mermas_cruda": 700.0, "bultos_mermados_cruda": 2.0,
        "costo_mermas_trabajada": 800.0, "bultos_mermados_trabajada": 1.0,
        "devoluciones_bultos": 5.0, "devoluciones_venta": 4500.0,
        "rechazos_perdidos": 900.0, "rechazos_bultos": 2.0,
        "renta_pesos": 3080.0, "utilidad_pct": 61.6, "afuera_bultos": 18.0, "afuera_motivos": 2,
    },
    "afuera_por_motivo": [
        {"motivo": "ajuste_sin_costo", "etiqueta": "Consumió un ajuste (sin costo posible)",
         "bultos": 14.0, "articulos": [{"nombre": "Anco", "bultos": 14.0}]},
        {"motivo": "sin_kilaje", "etiqueta": "Renglón armado sin kilaje cargado",
         "bultos": 4.0, "articulos": [{"nombre": "Kiwi", "bultos": 4.0}]},
    ],
    "fechas_incluidas": [date(2026, 8, 25)],
}


def test_generar_pdf_rentabilidad_real_arma_un_pdf():
    pdf = generar_pdf_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), ["cliente Día"], RESULTADO_REAL)
    assert pdf.startswith(b"%PDF")


def test_generar_excel_rentabilidad_real_lleva_la_cuenta_y_el_afuera():
    excel = generar_excel_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), ["cliente Día"], RESULTADO_REAL)
    hoja = openpyxl.load_workbook(BytesIO(excel)).active

    assert hoja.title == "Rentabilidad Real"
    valores = [str(celda.value) for fila in hoja.iter_rows() for celda in fila if celda.value is not None]
    texto = "\n".join(valores)
    # El subtítulo lleva la regla de la cuenta REAL, para que el archivo
    # se explique solo.
    assert "venta = lo ENVIADO" in texto
    assert "costo FIFO" in texto
    assert "la segunda vale cero" in texto
    # El AFUERA va incluido, con bultos y artículos por motivo — en el
    # papel también es protagonista.
    assert "AFUERA DEL CÁLCULO" in texto
    assert "Consumió un ajuste (sin costo posible)" in texto
    assert "Anco (14)" in texto
    # La cuenta: mermas, segunda y devoluciones con sus columnas.
    assert "Mermas $" in texto
    assert "Segunda bultos" in texto
    assert "Devol. bultos" in texto
    assert "Devoluciones $" in texto
    assert "Banana" in texto
    assert "Total REAL" in texto


def test_generar_excel_rentabilidad_real_abre_la_merma_al_lado_del_total_de_mermas():
    excel = generar_excel_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), ["cliente Día"], RESULTADO_REAL)
    hoja = openpyxl.load_workbook(BytesIO(excel)).active

    encabezados = [c.value for f in hoja.iter_rows() for c in f if c.value == "Mermas $"]
    assert encabezados, "falta la columna Mermas $"
    fila_encabezado = next(f for f in hoja.iter_rows() if f[0].value == "Artículo")
    nombres = [c.value for c in fila_encabezado]
    # El desglose va PEGADO al total de mermas (columnas 7 a 12), para poder
    # sumar o filtrar por "lo trabajado" sin ir a buscarlo al final.
    assert nombres[6:12] == [
        "Mermas $", "Mermas bultos",
        "Mermas crudas $", "Mermas crudas bultos", "Mermas trabajadas $", "Mermas trabajadas bultos",
    ]

    fila_banana = next(f for f in hoja.iter_rows() if f[0].value == "Banana")
    assert fila_banana[6].value == 1500.0   # total de mermas
    assert fila_banana[8].value == 700.0    # crudas $
    assert fila_banana[9].value == 2.0      # crudas bultos
    assert fila_banana[10].value == 800.0   # trabajadas $
    assert fila_banana[11].value == 1.0     # trabajadas bultos
    # Las dos partes cierran contra el total, también en la planilla.
    assert fila_banana[8].value + fila_banana[10].value == fila_banana[6].value

    fila_total = next(f for f in hoja.iter_rows() if f[0].value == "Total REAL")
    assert fila_total[8].value == 700.0 and fila_total[10].value == 800.0
    # El total de bultos mermados también va en el total, no solo el
    # desglose: la columna no puede quedar vacía al lado de dos que suman.
    assert fila_total[7].value == 3.0
    # Y las columnas que estaban después de las mermas siguieron corriéndose
    # enteras: la renta no quedó pisada por el desglose.
    assert fila_total[17].value == 3080.0


def test_generar_pdf_rentabilidad_real_dice_si_lo_tirado_era_crudo_o_trabajado():
    import pypdfium2 as pdfium

    pdf = generar_pdf_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), ["cliente Día"], RESULTADO_REAL)
    documento = pdfium.PdfDocument(pdf)
    texto = "\n".join(pagina.get_textpage().get_text_range() for pagina in documento)

    # Va en su propio párrafo, no adentro de la columna angosta "Mermas".
    assert "Mermas — ¿materia prima o trabajo?" in texto
    assert "cruda $700" in texto
    assert "ya trabajada $800" in texto
    # El párrafo hace wrap, así que se chequea por partes y no por la
    # frase entera de corrido.
    assert "Es el mismo" in texto and "total de $1.500, abierto por lo que se tiró." in texto
    assert "(1 bulto, guías R" in texto  # un solo bulto no lleva la s


def test_exports_reales_sin_datos_no_rompen():
    vacio = {"grupos": [], "totales": {"bultos": 0, "venta_neta": 0, "costo_mercaderia": 0,
                                       "costo_envase": 0, "costo_mermas": 0, "bultos_mermados": 0,
                                       "costo_mermas_cruda": 0, "bultos_mermados_cruda": 0,
                                       "costo_mermas_trabajada": 0, "bultos_mermados_trabajada": 0,
                                       "segunda_bultos": 0,
                                       "costo_total": 0, "renta_pesos": 0, "utilidad_pct": None,
                                       "afuera_bultos": 0, "afuera_motivos": 0},
             "afuera_por_motivo": [], "fechas_incluidas": []}
    assert generar_pdf_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), [], vacio).startswith(b"%PDF")
    hoja = openpyxl.load_workbook(
        BytesIO(generar_excel_rentabilidad_real(date(2026, 8, 18), date(2026, 8, 25), [], vacio))
    ).active
    valores = [str(c.value) for f in hoja.iter_rows() for c in f if c.value is not None]
    assert any("Sin movimientos" in v for v in valores)
